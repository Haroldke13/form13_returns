#!/usr/bin/env python3
"""Backend-friendly workbook merge helpers for pbo_reports-linked exports.

Default CLI usage:
    python3 mergePBOrecords_2.py --related pbo_donations.xlsx --output merged_donations.xlsx

This module keeps the original workbook-to-workbook merge flow, but exposes
importable helpers so the Flask admin backend can generate live exports,
merge them in-process, track progress, and return one cleaned workbook.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEFAULT_REFERENCE = "pbo_reports (2).xlsx"
DEFAULT_REFERENCE_FIELDS = [
    "pbo_name",
    "pbo_name_normalized",
    "pbo_registration_number",
    "reporting_period_start",
    "reporting_period_end",
    "scope",
    "counties",
    "countries_of_operation",
]
DEFAULT_REFERENCE_KEY_ALIASES = ("id", "report_id", "reportid")
DEFAULT_RELATED_KEY_ALIASES = ("report_id", "reportid", "report id", "ReportID", "Report Id", "id")
SHEET_TITLE_LIMIT = 31


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge a report-linked workbook with the PBO reports workbook."
    )
    parser.add_argument(
        "--reference",
        default=DEFAULT_REFERENCE,
        help="Reference workbook path. Defaults to %(default)s.",
    )
    parser.add_argument(
        "--related",
        required=True,
        help="Related workbook path containing a report foreign key such as report_id.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Output workbook path to create.",
    )
    parser.add_argument(
        "--reference-sheet",
        default=None,
        help="Optional reference sheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--related-sheet",
        default=None,
        help="Optional related sheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--reference-key",
        default="id",
        help="Reference workbook primary key column. Defaults to %(default)s.",
    )
    parser.add_argument(
        "--related-key",
        default="report_id",
        help="Related workbook foreign key column. Defaults to %(default)s.",
    )
    parser.add_argument(
        "--reference-fields",
        default=",".join(DEFAULT_REFERENCE_FIELDS),
        help="Comma-separated reference fields to carry into the output.",
    )
    parser.add_argument(
        "--numeric-fields",
        default="",
        help="Optional comma-separated related columns to force as numeric summary fields.",
    )
    parser.add_argument(
        "--label",
        default=None,
        help="Optional label for worksheet prefixes. Defaults to the related file name.",
    )
    return parser.parse_args()


def parse_csv(raw: str) -> list[str]:
    return [part.strip() for part in (raw or "").split(",") if part.strip()]


def safe_get(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def normalize_report_id(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _header_token(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _resolve_header_name(header: list[str], candidates: list[str] | tuple[str, ...]) -> str | None:
    token_to_name = {_header_token(name): name for name in header if str(name or "").strip()}
    for candidate in candidates:
        match = token_to_name.get(_header_token(candidate))
        if match:
            return match
    return None


def _normalize_header_row(row: tuple[Any, ...]) -> list[str]:
    headers: list[str] = []
    used_tokens: dict[str, int] = {}
    for index, value in enumerate(row, start=1):
        header = str(value or "").strip() or f"column_{index}"
        token = _header_token(header) or f"column{index}"
        duplicate_index = used_tokens.get(token, 0)
        used_tokens[token] = duplicate_index + 1
        if duplicate_index:
            header = f"{header}_{duplicate_index + 1}"
        headers.append(header)
    return headers


def load_rows(path: str, sheet_name: str | None = None) -> tuple[list[tuple[Any, ...]], str]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        with file_path.open(newline="", encoding="utf-8-sig") as handle:
            rows = [tuple(row) for row in csv.reader(handle)]
        return rows, file_path.name
    if suffix == ".xls":
        try:
            import pandas as pandas_module
        except Exception as exc:
            raise ValueError("Reading .xls files requires pandas to be installed.") from exc
        dataframe = pandas_module.read_excel(path, sheet_name=sheet_name or 0)
        headers = tuple(str(value or "").strip() for value in dataframe.columns.tolist())
        data_rows = [tuple(row) for row in dataframe.fillna("").itertuples(index=False, name=None)]
        return [headers] + data_rows, str(sheet_name or "Sheet1")

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    selected_sheet = worksheet.title
    workbook.close()
    return rows, selected_sheet


def infer_numeric_column(values: list[Any]) -> bool:
    seen = False
    for value in values:
        if value in (None, "") or isinstance(value, bool):
            continue
        if isinstance(value, (int, float)):
            seen = True
            continue
        try:
            float(str(value).strip())
        except (TypeError, ValueError):
            return False
        seen = True
    return seen


def to_float(value: Any) -> float:
    if value in (None, "") or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return 0.0


def ordered_unique(values: list[Any]) -> str:
    seen: dict[str, None] = {}
    for value in values:
        if value in (None, ""):
            continue
        seen[str(value)] = None
    return ", ".join(seen.keys())


def should_currency_format(name: str) -> bool:
    lowered = name.lower()
    keywords = (
        "amount",
        "total",
        "fee",
        "income",
        "receipt",
        "balance",
        "spending",
        "penalty",
        "grant",
        "donation",
        "value",
    )
    return any(keyword in lowered for keyword in keywords)


def is_date_like(value: Any) -> bool:
    return isinstance(value, (date, datetime))


def autosize_and_format(worksheet) -> None:
    widths: dict[int, int] = {}
    headers = [cell.value for cell in worksheet[1]]
    for col_index, header in enumerate(headers, start=1):
        widths[col_index] = max(len(str(header or "")), 12)
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=col_index)
            if cell.value is not None:
                widths[col_index] = min(max(widths[col_index], len(str(cell.value)) + 2), 60)
                if should_currency_format(str(header or "")) and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
                elif is_date_like(cell.value):
                    cell.number_format = "yyyy-mm-dd"
    for col_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(col_index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)


def write_row_dicts_workbook(
    output_path: str,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_name(sheet_name, set())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    autosize_and_format(worksheet)
    workbook.save(output_path)


def _combine_prefixed_rows(
    reference_label: str,
    reference_row: dict[str, Any],
    reference_columns: list[str],
    related_label: str,
    related_row: dict[str, Any] | None,
    related_columns: list[str],
) -> dict[str, Any]:
    combined = {
        f"{reference_label}__{column}": reference_row.get(column)
        for column in reference_columns
    }
    combined.update(
        {
            f"{related_label}__{column}": (related_row.get(column) if related_row else None)
            for column in related_columns
        }
    )
    return combined


def merge_report_driven_sources(
    reference_rows: list[dict[str, Any]],
    source_specs: list[dict[str, Any]],
    output_path: str,
    reference_label: str = "pbo_reports",
    reference_key: str = "id",
    progress_callback=None,
) -> dict[str, Any]:
    if not reference_rows:
        raise ValueError("Reference dataset is empty.")
    if not source_specs:
        raise ValueError("At least one related source is required.")

    reference_columns = list(reference_rows[0].keys())
    reference_fields = [field for field in DEFAULT_REFERENCE_FIELDS if field in reference_columns]
    if not reference_fields:
        reference_fields = [column for column in reference_columns if column != reference_key]
    reference_map: dict[int, dict[str, Any]] = {}
    for row in reference_rows:
        report_id = normalize_report_id(row.get(reference_key))
        if report_id is not None:
            reference_map[report_id] = row

    results = []
    merged_rows: list[dict[str, Any]] = []
    include_source_column = len(source_specs) > 1
    output_headers = (["source_table"] if include_source_column else []) + ["report_id"] + reference_fields
    seen_headers = set(output_headers)

    for index, spec in enumerate(source_specs, start=1):
        related_label = str(spec.get("label") or f"related_{index}").strip() or f"related_{index}"
        related_key = str(spec.get("related_key") or "report_id").strip() or "report_id"
        related_rows = list(spec.get("rows") or [])
        related_columns = list(
            spec.get("columns")
            or (list(related_rows[0].keys()) if related_rows else [related_key])
        )

        if progress_callback:
            progress_callback(
                {
                    "stage": "merging_database",
                    "detail": f"Joining {related_label} to {reference_label}",
                    "current_source": related_label,
                    "sources_total": len(source_specs),
                    "sources_completed": index - 1,
                }
            )

        related_non_key_columns = [column for column in related_columns if column != related_key]
        matched_count = 0
        unmatched_count = 0
        matched_report_ids: set[int] = set()

        for row in related_rows:
            report_id = normalize_report_id(row.get(related_key))
            if report_id is None or report_id not in reference_map:
                unmatched_count += 1
                continue
            matched_report_ids.add(report_id)
            merged_row = {"report_id": report_id}
            if include_source_column:
                merged_row["source_table"] = related_label
            for field in reference_fields:
                merged_row[field] = reference_map[report_id].get(field)
            for column in related_non_key_columns:
                merged_row[column] = row.get(column)
                if column not in seen_headers:
                    output_headers.append(column)
                    seen_headers.add(column)
            merged_rows.append(merged_row)
            matched_count += 1

        matched_reports = len(matched_report_ids)
        reports_without_match = max(len(reference_map) - matched_reports, 0)
        results.append(
            {
                "label": related_label,
                "related_rows": len(related_rows),
                "joined_rows": matched_count,
                "matched_reports": matched_reports,
                "reports_without_match": reports_without_match,
                "orphan_rows": unmatched_count,
            }
        )

        if progress_callback:
            progress_callback(
                {
                    "stage": "merging_database",
                    "detail": f"Merged {matched_count} cleaned rows from {related_label}",
                    "current_source": related_label,
                    "sources_total": len(source_specs),
                    "sources_completed": index,
                }
            )

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    write_row_dicts_workbook(
        output_path=str(output_file),
        sheet_name="Matched Rows",
        headers=output_headers,
        rows=merged_rows,
    )
    return {
        "reference_label": reference_label,
        "reference_rows": len(reference_rows),
        "output_path": str(output_file),
        "sources": results,
        "rows_written": len(merged_rows),
    }


def build_reference_map(
    reference_rows: list[tuple[Any, ...]],
    reference_key: str,
    reference_fields: list[str],
) -> tuple[dict[int, dict[str, Any]], list[str]]:
    if not reference_rows:
        raise ValueError("Reference workbook is empty.")

    header = _normalize_header_row(reference_rows[0])
    index_by_name = {name: idx for idx, name in enumerate(header)}
    resolved_reference_key = _resolve_header_name(
        header,
        [reference_key, *DEFAULT_REFERENCE_KEY_ALIASES],
    )
    if resolved_reference_key is None:
        raise ValueError(f"Reference key column '{reference_key}' was not found.")

    available_fields: list[str] = []
    field_name_map: dict[str, str] = {}
    for field in reference_fields:
        resolved_name = _resolve_header_name(header, [field])
        if resolved_name:
            available_fields.append(field)
            field_name_map[field] = resolved_name

    report_map: dict[int, dict[str, Any]] = {}
    for row in reference_rows[1:]:
        report_id = normalize_report_id(safe_get(row, index_by_name[resolved_reference_key]))
        if report_id is None:
            continue
        report_map[report_id] = {
            field: safe_get(row, index_by_name[field_name_map[field]])
            for field in available_fields
        }
    return report_map, available_fields


def merge_related_rows(
    related_rows: list[tuple[Any, ...]],
    report_map: dict[int, dict[str, Any]],
    related_key: str,
    forced_numeric_fields: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str], list[str], list[str], str]:
    if not related_rows:
        raise ValueError("Related workbook is empty.")

    related_header = _normalize_header_row(related_rows[0])
    related_index = {name: idx for idx, name in enumerate(related_header)}
    resolved_related_key = _resolve_header_name(
        related_header,
        [related_key, *DEFAULT_RELATED_KEY_ALIASES],
    )
    if resolved_related_key is None:
        raise ValueError(f"Related key column '{related_key}' was not found.")

    related_non_key_columns = [name for name in related_header if name != resolved_related_key]
    matched: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []

    for row in related_rows[1:]:
        report_id = normalize_report_id(safe_get(row, related_index[resolved_related_key]))
        related_values = {
            column: safe_get(row, related_index[column]) for column in related_non_key_columns
        }
        if report_id is None or report_id not in report_map:
            unmatched.append({resolved_related_key: report_id, **related_values})
            continue
        matched.append(
            {
                "report_id": report_id,
                "reference": report_map[report_id],
                "related": related_values,
            }
        )

    summary_candidate_columns = [
        name for name in related_non_key_columns if name not in {"id", resolved_related_key}
    ]
    if forced_numeric_fields:
        numeric_columns = [
            column for column in forced_numeric_fields if column in summary_candidate_columns
        ]
    else:
        numeric_columns = [
            column
            for column in summary_candidate_columns
            if infer_numeric_column([row["related"].get(column) for row in matched])
        ]
    text_columns = [
        column for column in summary_candidate_columns if column not in numeric_columns
    ]
    return matched, unmatched, related_non_key_columns, numeric_columns, text_columns, resolved_related_key


def _safe_sheet_name(name: str, used_names: set[str]) -> str:
    sanitized = re.sub(r"[\\/*?:\[\]]+", " ", str(name or "").strip()).strip()
    sanitized = sanitized or "Sheet"
    sanitized = sanitized[:SHEET_TITLE_LIMIT].rstrip() or "Sheet"
    candidate = sanitized
    duplicate_index = 2
    while candidate in used_names:
        suffix = f"_{duplicate_index}"
        candidate = f"{sanitized[: max(0, SHEET_TITLE_LIMIT - len(suffix))].rstrip()}{suffix}" or f"Sheet{suffix}"
        duplicate_index += 1
    used_names.add(candidate)
    return candidate


def _sheet_title(label: str, suffix: str, used_names: set[str]) -> str:
    suffix_text = f" {suffix}".strip()
    max_label_len = max(1, SHEET_TITLE_LIMIT - len(suffix_text))
    base = re.sub(r"[\\/*?:\[\]]+", " ", str(label or "").strip()).strip() or "Source"
    candidate = f"{base[:max_label_len].rstrip()}{suffix_text}"
    return _safe_sheet_name(candidate.strip(), used_names)


def _append_source_sheets(
    workbook: Workbook,
    used_names: set[str],
    related_key: str,
    source_label: str,
    reference_fields: list[str],
    related_columns: list[str],
    matched: list[dict[str, Any]],
    unmatched: list[dict[str, Any]],
    numeric_columns: list[str],
    text_columns: list[str],
    use_generic_titles: bool = False,
) -> None:
    if use_generic_titles:
        ws_detail = workbook.create_sheet(_safe_sheet_name("Matched Rows", used_names))
    else:
        ws_detail = workbook.create_sheet(_sheet_title(source_label, "Matched", used_names))
    detail_header = [related_key] + reference_fields + related_columns
    ws_detail.append(detail_header)
    for row in matched:
        ws_detail.append(
            [row["report_id"]]
            + [row["reference"].get(field) for field in reference_fields]
            + [row["related"].get(column) for column in related_columns]
        )

    report_summary: dict[int, dict[str, Any]] = {}
    pbo_summary: dict[str, dict[str, Any]] = {}

    for row in matched:
        report_id = row["report_id"]
        pbo_name = row["reference"].get("pbo_name") or "UNKNOWN PBO"

        report_bucket = report_summary.setdefault(
            report_id,
            {
                "reference": row["reference"],
                "row_count": 0,
                "numeric": defaultdict(float),
                "text": defaultdict(list),
            },
        )
        report_bucket["row_count"] += 1
        for column in numeric_columns:
            report_bucket["numeric"][column] += to_float(row["related"].get(column))
        for column in text_columns:
            report_bucket["text"][column].append(row["related"].get(column))

        pbo_bucket = pbo_summary.setdefault(
            pbo_name,
            {
                "report_ids": [],
                "row_count": 0,
                "numeric": defaultdict(float),
                "text": defaultdict(list),
            },
        )
        if report_id not in pbo_bucket["report_ids"]:
            pbo_bucket["report_ids"].append(report_id)
        pbo_bucket["row_count"] += 1
        for column in numeric_columns:
            pbo_bucket["numeric"][column] += to_float(row["related"].get(column))
        for column in text_columns:
            pbo_bucket["text"][column].append(row["related"].get(column))

    if use_generic_titles:
        ws_report = workbook.create_sheet(_safe_sheet_name("Report Summary", used_names))
    else:
        ws_report = workbook.create_sheet(_sheet_title(source_label, "Report Summary", used_names))
    report_header = [related_key] + reference_fields + ["matched_row_count"]
    report_header += [f"total_{column}" for column in numeric_columns]
    report_header += text_columns
    ws_report.append(report_header)
    for report_id in sorted(report_summary):
        bucket = report_summary[report_id]
        ws_report.append(
            [report_id]
            + [bucket["reference"].get(field) for field in reference_fields]
            + [bucket["row_count"]]
            + [bucket["numeric"].get(column, 0.0) for column in numeric_columns]
            + [ordered_unique(bucket["text"].get(column, [])) for column in text_columns]
        )

    if use_generic_titles:
        ws_pbo = workbook.create_sheet(_safe_sheet_name("PBO Summary", used_names))
    else:
        ws_pbo = workbook.create_sheet(_sheet_title(source_label, "PBO Summary", used_names))
    pbo_header = ["pbo_name", "report_ids", "report_count", "matched_row_count"]
    pbo_header += [f"total_{column}" for column in numeric_columns]
    pbo_header += text_columns
    ws_pbo.append(pbo_header)
    for pbo_name in sorted(pbo_summary):
        bucket = pbo_summary[pbo_name]
        ws_pbo.append(
            [
                pbo_name,
                ", ".join(str(report_id) for report_id in bucket["report_ids"]),
                len(bucket["report_ids"]),
                bucket["row_count"],
            ]
            + [bucket["numeric"].get(column, 0.0) for column in numeric_columns]
            + [ordered_unique(bucket["text"].get(column, [])) for column in text_columns]
        )

    if unmatched:
        if use_generic_titles:
            ws_unmatched = workbook.create_sheet(_safe_sheet_name("Unmatched Rows", used_names))
        else:
            ws_unmatched = workbook.create_sheet(_sheet_title(source_label, "Unmatched", used_names))
        unmatched_header = [related_key] + related_columns
        ws_unmatched.append(unmatched_header)
        for row in unmatched:
            ws_unmatched.append([row.get(related_key)] + [row.get(column) for column in related_columns])


def merge_related_workbooks(
    reference_path: str,
    related_specs: list[dict[str, Any]],
    output_path: str,
    reference_sheet: str | None = None,
    reference_key: str = "id",
    reference_fields: list[str] | None = None,
    progress_callback=None,
) -> dict[str, Any]:
    if not related_specs:
        raise ValueError("At least one related workbook is required.")

    selected_reference_fields = list(reference_fields or DEFAULT_REFERENCE_FIELDS)
    if progress_callback:
        progress_callback(
            {
                "stage": "loading_reference",
                "detail": f"Loading reference workbook {Path(reference_path).name}",
                "sources_total": len(related_specs),
                "sources_completed": 0,
            }
        )

    reference_rows, reference_sheet_name = load_rows(reference_path, reference_sheet)
    report_map, available_reference_fields = build_reference_map(
        reference_rows=reference_rows,
        reference_key=reference_key,
        reference_fields=selected_reference_fields,
    )

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Merge Overview"
    overview.append(
        [
            "source_label",
            "source_file",
            "source_sheet",
            "matched_rows",
            "unmatched_rows",
            "numeric_columns",
            "text_columns",
        ]
    )
    used_names = {overview.title}
    source_results = []

    for index, spec in enumerate(related_specs, start=1):
        related_path = str(spec.get("path") or "").strip()
        if not related_path:
            raise ValueError("Each related spec requires a path.")
        source_label = str(spec.get("label") or Path(related_path).stem or f"Source {index}")
        related_sheet = spec.get("sheet")
        forced_numeric_fields = list(spec.get("numeric_fields") or [])
        related_key = str(spec.get("related_key") or "report_id").strip() or "report_id"

        if progress_callback:
            progress_callback(
                {
                    "stage": "loading_related",
                    "detail": f"Loading {source_label}",
                    "current_source": source_label,
                    "sources_total": len(related_specs),
                    "sources_completed": index - 1,
                }
            )

        related_rows, related_sheet_name = load_rows(related_path, related_sheet)
        matched, unmatched, related_columns, numeric_columns, text_columns, resolved_related_key = merge_related_rows(
            related_rows=related_rows,
            report_map=report_map,
            related_key=related_key,
            forced_numeric_fields=forced_numeric_fields,
        )

        _append_source_sheets(
            workbook=workbook,
            used_names=used_names,
            related_key=resolved_related_key,
            source_label=source_label,
            reference_fields=available_reference_fields,
            related_columns=related_columns,
            matched=matched,
            unmatched=unmatched,
            numeric_columns=numeric_columns,
            text_columns=text_columns,
        )
        overview.append(
            [
                source_label,
                Path(related_path).name,
                related_sheet_name,
                len(matched),
                len(unmatched),
                ", ".join(numeric_columns),
                ", ".join(text_columns),
            ]
        )
        source_results.append(
            {
                "label": source_label,
                "path": related_path,
                "sheet": related_sheet_name,
                "matched_rows": len(matched),
                "unmatched_rows": len(unmatched),
                "numeric_columns": numeric_columns,
                "text_columns": text_columns,
            }
        )
        if progress_callback:
            progress_callback(
                {
                    "stage": "merging",
                    "detail": f"Merged {source_label}",
                    "current_source": source_label,
                    "sources_total": len(related_specs),
                    "sources_completed": index,
                }
            )

    for worksheet in workbook.worksheets:
        autosize_and_format(worksheet)

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)

    return {
        "reference_path": str(reference_path),
        "reference_sheet": reference_sheet_name,
        "output_path": str(output_file),
        "reference_fields": available_reference_fields,
        "sources": source_results,
    }


def main() -> None:
    args = parse_args()
    result = merge_related_workbooks(
        reference_path=args.reference,
        related_specs=[
            {
                "path": args.related,
                "sheet": args.related_sheet,
                "label": args.label or Path(args.related).stem,
                "related_key": args.related_key,
                "numeric_fields": parse_csv(args.numeric_fields),
            }
        ],
        output_path=args.output,
        reference_sheet=args.reference_sheet,
        reference_key=args.reference_key,
        reference_fields=parse_csv(args.reference_fields),
    )

    source_result = result["sources"][0]
    print(f"Reference workbook: {Path(result['reference_path']).name} [{result['reference_sheet']}]")
    print(f"Related workbook: {Path(source_result['path']).name} [{source_result['sheet']}]")
    print(f"Output workbook: {result['output_path']}")
    print(f"Matched rows: {source_result['matched_rows']}")
    print(f"Unmatched rows: {source_result['unmatched_rows']}")
    print(f"Numeric summary columns: {', '.join(source_result['numeric_columns']) or 'None'}")
    print(f"Text summary columns: {', '.join(source_result['text_columns']) or 'None'}")


if __name__ == "__main__":
    main()
