#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class ContactExportRow:
    organization_name: str
    organization_contact_number: str
    organization_email: str
    contact_person_name: str
    contact_person_email: str
    contact_person_phone_number: str


def normalize_space(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_key(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", normalize_space(value).upper()).strip()


def first_nonempty(values: Iterable[object]) -> str:
    for value in values:
        cleaned = normalize_space(value)
        if cleaned:
            return cleaned
    return ""


def load_contact_rows(database_path: Path) -> list[ContactExportRow]:
    connection = sqlite3.connect(database_path)
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()
    cursor.execute(
        """
        SELECT
            pbo_name,
            telephone,
            cell_phone,
            email,
            contact_name,
            contact_telephone,
            contact_email,
            reporting_period_end,
            id
        FROM pbo_reports
        ORDER BY COALESCE(reporting_period_end, '') DESC, id DESC
        """
    )

    grouped_rows: dict[str, list[dict[str, object]]] = {}
    for record in cursor.fetchall():
        pbo_name = normalize_space(record["pbo_name"])
        if not pbo_name:
            continue
        grouped_rows.setdefault(normalize_key(pbo_name), []).append(dict(record))

    connection.close()

    export_rows: list[ContactExportRow] = []
    for grouped in grouped_rows.values():
        export_rows.append(
            ContactExportRow(
                organization_name=first_nonempty(row.get("pbo_name") for row in grouped),
                # The database currently stores no org-level phone/email values, so
                # these columns fall back to the contact person details when needed.
                organization_contact_number=first_nonempty(
                    value
                    for row in grouped
                    for value in (
                        row.get("telephone"),
                        row.get("cell_phone"),
                        row.get("contact_telephone"),
                    )
                ),
                organization_email=first_nonempty(
                    value
                    for row in grouped
                    for value in (
                        row.get("email"),
                        row.get("contact_email"),
                    )
                ),
                contact_person_name=first_nonempty(row.get("contact_name") for row in grouped),
                contact_person_email=first_nonempty(row.get("contact_email") for row in grouped),
                contact_person_phone_number=first_nonempty(
                    row.get("contact_telephone") for row in grouped
                ),
            )
        )

    export_rows.sort(key=lambda row: normalize_key(row.organization_name))
    return export_rows


def write_workbook(rows: list[ContactExportRow], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contacts"

    headers = [
        "Organization Name",
        "Organization Contact Number",
        "Organization Email",
        "Contact Person Name",
        "Contact Person Email",
        "Contact Person Phone Number",
    ]
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        sheet.append(
            [
                row.organization_name,
                row.organization_contact_number,
                row.organization_email,
                row.contact_person_name,
                row.contact_person_email,
                row.contact_person_phone_number,
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = max(len(normalize_space(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(max_length + 2, 18),
            48,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export one consolidated contact row per organization from pbo_reports."
    )
    parser.add_argument(
        "--database",
        type=Path,
        default=Path("returnsform14_org_backup.sqlite"),
        help="Path to the SQLite database.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("contacts.xlsx"),
        help="Path to the output workbook.",
    )
    args = parser.parse_args()

    rows = load_contact_rows(args.database)
    write_workbook(rows, args.output)

    missing_all_contact_fields = sum(
        1
        for row in rows
        if not any(
            [
                row.organization_contact_number,
                row.organization_email,
                row.contact_person_name,
                row.contact_person_email,
                row.contact_person_phone_number,
            ]
        )
    )
    print(f"Exported {len(rows)} organizations to {args.output}")
    print(f"Organizations with no contact data at all: {missing_all_contact_fields}")


if __name__ == "__main__":
    main()
