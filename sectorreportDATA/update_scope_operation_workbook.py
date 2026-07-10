from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from scope_operation_utils import format_scope_labels


WORKBOOK_PATH = BASE_DIR / "NGO-Registered-Entities-2026 February_split_by_date.xlsx"
TARGET_HEADER = "SCOPE OF OPERATION"
OBJECTIVES_HEADER = "objectives"


def main():
    workbook = load_workbook(WORKBOOK_PATH)
    updated_counts: dict[str, int] = {}

    for sheet in workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]
        if OBJECTIVES_HEADER not in headers:
            updated_counts[sheet.title] = 0
            continue

        objectives_col = headers.index(OBJECTIVES_HEADER) + 1
        target_col = objectives_col + 1
        if sheet.cell(row=1, column=target_col).value != TARGET_HEADER:
            sheet.insert_cols(target_col)
            sheet.cell(row=1, column=target_col).value = TARGET_HEADER

        updated = 0
        for row_idx in range(2, sheet.max_row + 1):
            objective_value = sheet.cell(row=row_idx, column=objectives_col).value or ""
            sheet.cell(row=row_idx, column=target_col).value = format_scope_labels(str(objective_value))
            updated += 1

        updated_counts[sheet.title] = updated

    workbook.save(WORKBOOK_PATH)
    print(WORKBOOK_PATH)
    for sheet_name, count in updated_counts.items():
        print(f"{sheet_name}: {count}")


if __name__ == "__main__":
    main()
