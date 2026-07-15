
import base64
import click
import sys
import csv
from contextlib import contextmanager
from decimal import Decimal, InvalidOperation
from difflib import get_close_matches
import hashlib
import html
import io
import ipaddress
import json
import math
import random
import smtplib
import shutil
import subprocess
import zipfile
from collections import Counter, defaultdict
from functools import wraps
from flask.cli import with_appcontext
from flask.config import Config

from dotenv import load_dotenv

from flask import Flask, Response, render_template, request, redirect, session, url_for, flash, jsonify, send_file, abort, make_response

from sqlalchemy import MetaData, Table, String, Text, bindparam, cast, create_engine, func, inspect, literal_column, or_, select, text
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import IntegrityError, OperationalError
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from datetime import date, datetime, timezone
from email.message import EmailMessage
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from models import (
    AdminSetting,
    ALL_MODELS,
    db,
    PBOReport,
    Asset,
    IncomeGeneratingActivity,
    Donation,
    Grant,
    Payment,
    BankAccount,
    AuditorEntry,
    StaffBiodata,
    VolunteerBiodata,
    VolunteerPrivilege,
    TrainingRecord,
    TaxWaiverItem,
    Official,
    ProjectImplementation,
    ProjectCarriedOut,
    CollaborationNetworking,
    PostalCodeCache,
    UserActivityLog,
    FieldChangeLog,
    ImportBatch,
    ImportRowError,
    UploadedFile,
    FieldHelpCache,
    FieldHelpDecisionModel,
    FieldHelpMemory,
    FieldHelpRule,
    FieldHelpIntentSample,
    FieldHelpInteraction,
    DataAnalysisTrainingQuestion,
    DataAnalysisBotInteraction,
    REPORT_WORKFLOW_STATUS_VALUES as MODEL_REPORT_WORKFLOW_STATUS_VALUES,
    REPORT_REVIEW_STATUS_VALUES as MODEL_REPORT_REVIEW_STATUS_VALUES,
    IMPORT_BATCH_STATUS_VALUES as MODEL_IMPORT_BATCH_STATUS_VALUES,
    UPLOADED_FILE_STATUS_VALUES as MODEL_UPLOADED_FILE_STATUS_VALUES,
    RETURN_DATE_PLACEHOLDER,
    sqlite_table_has_real_id_primary_key,
)
from mergePBOrecords_2 import (
    DEFAULT_REFERENCE_FIELDS as MERGE_REFERENCE_FIELDS,
    merge_report_driven_sources,
    write_row_dicts_workbook,
)
import os
import re
import secrets
import string
import threading
import time
import uuid
from pathlib import Path
import requests
from textwrap import shorten
from urllib.parse import parse_qs, quote_plus, urlencode, urlparse, urlunparse
from models import User
from datetime import date, timedelta
from zoneinfo import ZoneInfo
# Optional analytics imports for EDA/profiling are loaded lazily so worker boot stays light.
pd = None
ProfileReport = None
TfidfVectorizer = None
RandomForestClassifier = None
DecisionTreeClassifier = None
train_test_split = None
Pipeline = None
cosine_similarity = None

try:
    import fcntl
except ImportError:  # pragma: no cover - non-POSIX hosts
    fcntl = None



load_dotenv()

def normalize_flask_trusted_hosts(value):
    if value and "*" in value:
        return None
    return value


class PBORAConfig(Config):
    def __setitem__(self, key, value):
        if key == "TRUSTED_HOSTS":
            value = normalize_flask_trusted_hosts(value)
        super().__setitem__(key, value)


class PBORAFlask(Flask):
    config_class = PBORAConfig

    @property
    def trusted_hosts(self):
        return getattr(self, "_pbora_trusted_hosts", None)

    @trusted_hosts.setter
    def trusted_hosts(self, value):
        self._pbora_trusted_hosts = normalize_flask_trusted_hosts(value)


app = PBORAFlask(__name__)
BACKUP_THREAD_STARTED = False
BACKUP_THREAD_LOCK = threading.Lock()
app.config['MAX_CONTENT_LENGTH'] = max(int(os.getenv('MAX_REQUEST_MB', '260') or 260), 50) * 1024 * 1024
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': int(os.getenv('SQLALCHEMY_POOL_RECYCLE', '3600')),
    'pool_timeout': int(os.getenv('SQLALCHEMY_POOL_TIMEOUT', '30')),
}
app.config['RESET_TOKEN_MAX_AGE_SECONDS'] = int(os.getenv('RESET_TOKEN_MAX_AGE_SECONDS', '3600'))
app.config['SUPERADMIN_RESET_FORWARD_EMAIL'] = os.getenv('SUPERADMIN_RESET_FORWARD_EMAIL', 'joelharold@ymail.com')

def env_flag(name, default=False):
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


DATA_ANALYSIS_RUNTIME_ASSET_PATHS = [
    Path(__file__).resolve().parent / 'form14_data_analysis_training_dataset.json',
    Path(__file__).resolve().parent / 'form14_data_analysis_filename_questions.json',
    Path(__file__).resolve().parent / 'form14_data_analysis_live_questions.json',
    Path(__file__).resolve().parent / 'form14_data_analysis_keyword_schema.json',
    Path(__file__).resolve().parent / 'form14_data_analysis_live_keyword_schema.json',
]


def data_analysis_runtime_assets_available():
    return all(path.exists() for path in DATA_ANALYSIS_RUNTIME_ASSET_PATHS)


DATA_ANALYSIS_RUNTIME_ASSETS_PRESENT = data_analysis_runtime_assets_available()
CRUD_ONLY_MODE = env_flag('CRUD_ONLY_MODE', True)
OCR_FEATURES_ENABLED = env_flag('OCR_FEATURES_ENABLED', not CRUD_ONLY_MODE)
VISUAL_ANALYSIS_ENABLED = env_flag('VISUAL_ANALYSIS_ENABLED', not CRUD_ONLY_MODE)
SECTOR_ANALYTICS_ENABLED = env_flag('SECTOR_ANALYTICS_ENABLED', not CRUD_ONLY_MODE)
TF_RISK_ENABLED = env_flag('TF_RISK_ENABLED', not CRUD_ONLY_MODE)
DATA_ANALYSIS_BOT_ENABLED = (
    env_flag('DATA_ANALYSIS_BOT_ENABLED', False)
    and not CRUD_ONLY_MODE
    and DATA_ANALYSIS_RUNTIME_ASSETS_PRESENT
)
SHOW_ADMIN_ANALYSIS_FLOATING_BOT = (
    env_flag('SHOW_ADMIN_ANALYSIS_FLOATING_BOT', False)
    and DATA_ANALYSIS_BOT_ENABLED
)
DATA_ANALYSIS_BACKGROUND_ONLY = True
print(
    f"[startup] data-entry mode {'enabled' if not DATA_ANALYSIS_BOT_ENABLED else 'disabled'} "
    f"(data analysis runtime assets {'present' if DATA_ANALYSIS_RUNTIME_ASSETS_PRESENT else 'missing'}; "
    f"crud_only_mode={'on' if CRUD_ONLY_MODE else 'off'})"
)


def get_pandas():
    global pd
    if pd is None:
        import pandas as pandas_module
        pd = pandas_module
    return pd


def get_profile_report_class():
    global ProfileReport
    if ProfileReport is not None:
        return ProfileReport
    try:
        from ydata_profiling import ProfileReport as profile_report_class
    except Exception:
        try:
            from pandas_profiling import ProfileReport as profile_report_class
        except Exception:
            profile_report_class = None
    ProfileReport = profile_report_class
    return ProfileReport


def get_sklearn_components():
    global TfidfVectorizer, RandomForestClassifier, DecisionTreeClassifier, train_test_split, Pipeline, cosine_similarity
    if all(component is not None for component in (
        TfidfVectorizer,
        RandomForestClassifier,
        DecisionTreeClassifier,
        train_test_split,
        Pipeline,
        cosine_similarity,
    )):
        return {
            'available': True,
            'TfidfVectorizer': TfidfVectorizer,
            'RandomForestClassifier': RandomForestClassifier,
            'DecisionTreeClassifier': DecisionTreeClassifier,
            'train_test_split': train_test_split,
            'Pipeline': Pipeline,
            'cosine_similarity': cosine_similarity,
        }
    try:
        from sklearn.ensemble import RandomForestClassifier as rf_cls
        from sklearn.tree import DecisionTreeClassifier as dt_cls
        from sklearn.feature_extraction.text import TfidfVectorizer as tfidf_cls
        from sklearn.model_selection import train_test_split as split_fn
        from sklearn.pipeline import Pipeline as pipe_cls
        from sklearn.metrics.pairwise import cosine_similarity as cosine_fn
    except Exception:
        return {'available': False}

    TfidfVectorizer = tfidf_cls
    RandomForestClassifier = rf_cls
    DecisionTreeClassifier = dt_cls
    train_test_split = split_fn
    Pipeline = pipe_cls
    cosine_similarity = cosine_fn
    return {
        'available': True,
        'TfidfVectorizer': TfidfVectorizer,
        'RandomForestClassifier': RandomForestClassifier,
        'DecisionTreeClassifier': DecisionTreeClassifier,
        'train_test_split': train_test_split,
        'Pipeline': Pipeline,
        'cosine_similarity': cosine_similarity,
    }


def render_plotly_figure_svg(figure, width=900, height=420):
    try:
        svg_bytes = figure.to_image(format='svg', width=width, height=height)
    except Exception:
        return None
    return base64.b64encode(svg_bytes).decode('utf-8')


def render_matplotlib_figure_svg(fig):
    output = io.StringIO()
    fig.savefig(output, format='svg', bbox_inches='tight', transparent=True)
    return base64.b64encode(output.getvalue().encode('utf-8')).decode('utf-8')


def build_my_files_activity_chart(activity_counts):
    if not activity_counts:
        return None

    try:
        from plotly import graph_objects as go
    except Exception:
        return None

    labels = [item[0] for item in activity_counts]
    values = [item[1] for item in activity_counts]

    figure = go.Figure()
    figure.add_trace(go.Scatter(
        x=labels,
        y=values,
        mode='lines+markers',
        line=dict(color='#0f766e', width=3, shape='spline', smoothing=1.1),
        marker=dict(size=8, color='#14b8a6', line=dict(color='#ccfbf1', width=2)),
        fill='tozeroy',
        fillcolor='rgba(20, 184, 166, 0.16)',
        hovertemplate='%{x}<br>%{y} touches<extra></extra>',
    ))
    figure.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(255,255,255,0)',
        margin=dict(l=28, r=18, t=18, b=28),
        xaxis=dict(
            title='',
            tickfont=dict(color='#475569'),
            showgrid=False,
            zeroline=False,
        ),
        yaxis=dict(
            title='',
            tickfont=dict(color='#475569'),
            gridcolor='rgba(148, 163, 184, 0.18)',
            zeroline=False,
        ),
        showlegend=False,
    )
    return render_plotly_figure_svg(figure)


def build_my_files_workflow_chart(workflow_counts):
    if not workflow_counts:
        return None

    try:
        from plotly import graph_objects as go
    except Exception:
        return None

    labels = list(workflow_counts.keys())
    values = list(workflow_counts.values())
    colors = ['#2563eb', '#0f766e', '#f59e0b', '#ef4444', '#8b5cf6', '#14b8a6', '#94a3b8']

    figure = go.Figure(go.Pie(
        labels=labels,
        values=values,
        hole=0.62,
        sort=False,
        marker=dict(colors=colors[:len(labels)]),
        textinfo='label+percent',
        hovertemplate='%{label}: %{value} records<extra></extra>',
    ))
    figure.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=12, r=12, t=12, b=12),
        showlegend=False,
    )
    return render_plotly_figure_svg(figure, width=520, height=420)


def build_my_files_bar_chart(data_points, title, color):
    if not data_points:
        return None

    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import seaborn as sns
    except Exception:
        return None

    labels = [item[0] for item in data_points]
    values = [item[1] for item in data_points]

    sns.set_theme(style='whitegrid')
    fig, ax = plt.subplots(figsize=(8.4, 4.4))
    sns.barplot(x=values, y=labels, ax=ax, color=color)
    ax.set_title(title, loc='left', fontsize=14, fontweight='bold', color='#0f172a')
    ax.set_xlabel('')
    ax.set_ylabel('')
    ax.tick_params(axis='x', colors='#475569')
    ax.tick_params(axis='y', colors='#0f172a')
    ax.grid(axis='x', color='#e2e8f0', linewidth=1)
    ax.grid(axis='y', visible=False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#cbd5e1')

    for index, value in enumerate(values):
        ax.text(value + 0.05, index, str(value), va='center', ha='left', color='#334155', fontsize=10)

    fig.patch.set_alpha(0)
    ax.set_facecolor((1, 1, 1, 0))
    svg_data = render_matplotlib_figure_svg(fig)
    plt.close(fig)
    return svg_data


def utc_now():
    return datetime.now(timezone.utc)


def env_csv(name):
    value = os.getenv(name, "")
    return [item.strip() for item in value.split(",") if item.strip()]


def env_auto_value(value):
    return str(value or "").strip().lower() in {"", "auto", "detect", "dynamic", "<server-ip>", "<server_ip>"}


def unique_text_items(items):
    seen = set()
    result = []
    for item in items:
        normalized = str(item or "").strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        result.append(normalized)
    return result


def command_stdout(args):
    try:
        completed = subprocess.run(args, check=False, text=True, capture_output=True)
    except OSError:
        return ""
    if completed.returncode != 0:
        return ""
    return completed.stdout.strip()


def runtime_ipv4_hosts():
    candidates = []
    route_output = command_stdout(["ip", "-4", "route", "get", "1.1.1.1"])
    route_match = re.search(r"\bsrc\s+(\d+\.\d+\.\d+\.\d+)", route_output)
    if route_match:
        candidates.append(route_match.group(1))

    hostname_ips = command_stdout(["hostname", "-I"])
    candidates.extend(hostname_ips.split())

    valid_hosts = []
    for candidate in unique_text_items(candidates):
        try:
            address = ipaddress.ip_address(candidate)
        except ValueError:
            continue
        if address.version == 4 and not address.is_loopback:
            valid_hosts.append(candidate)
    return valid_hosts


def normalize_trusted_host_candidate(host):
    host = str(host or "").strip()
    if not host:
        return ""
    if "://" in host:
        parsed = urlparse(host)
        host = parsed.hostname or ""
    elif ":" in host and host.count(":") == 1:
        host = host.split(":", 1)[0]
    return host.strip()


def resolve_trusted_hosts(raw_hosts, app_host_ip=None, runtime_hosts=None):
    hosts = []
    auto_requested = False

    for raw_host in raw_hosts or []:
        host = normalize_trusted_host_candidate(raw_host)
        if env_auto_value(host):
            auto_requested = True
            continue
        hosts.append(host)

    app_host = normalize_trusted_host_candidate(app_host_ip)
    if app_host:
        if env_auto_value(app_host):
            auto_requested = True
        else:
            hosts.append(app_host)

    if auto_requested:
        for host in runtime_hosts if runtime_hosts is not None else runtime_ipv4_hosts():
            normalized_host = normalize_trusted_host_candidate(host)
            if not normalized_host:
                continue
            try:
                address = ipaddress.ip_address(normalized_host)
            except ValueError:
                continue
            if address.version == 4 and not address.is_loopback:
                hosts.append(normalized_host)

    return unique_text_items(hosts)


def env_first(*names):
    for name in names:
        value = os.getenv(name)
        if value is None:
            continue
        normalized = value.strip()
        if normalized:
            return normalized
    return None


@app.context_processor
def inject_feature_flags():
    return {
        'crud_only_mode': CRUD_ONLY_MODE,
        'ocr_features_enabled': OCR_FEATURES_ENABLED,
        'data_analysis_bot_enabled': DATA_ANALYSIS_BOT_ENABLED,
        'visual_analysis_enabled': VISUAL_ANALYSIS_ENABLED,
        'sector_analytics_enabled': SECTOR_ANALYTICS_ENABLED,
        'tf_risk_enabled': TF_RISK_ENABLED,
        'show_admin_analysis_floating_bot': SHOW_ADMIN_ANALYSIS_FLOATING_BOT,
    }


def data_analysis_bot_disabled_message():
    return 'Data Analysis Bot is temporarily disabled for this deployment phase (data-entry mode).'


def data_analysis_bot_disabled_response():
    message = data_analysis_bot_disabled_message()
    wants_json = (
        request.is_json
        or request.path.startswith('/admin/data-analysis-bot')
        or request.path.startswith('/admin/ai/')
        or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        or 'application/json' in (request.headers.get('Accept') or '')
    )
    if wants_json:
        return jsonify({'error': message, 'disabled': True}), 503
    flash(message, 'warning')
    return redirect(url_for('admin_view'))


def require_data_analysis_bot(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if DATA_ANALYSIS_BOT_ENABLED:
            return view_func(*args, **kwargs)
        abort(404)

    return wrapped


def require_feature_enabled(flag_name):
    def decorator(view_func):
        @wraps(view_func)
        def wrapped(*args, **kwargs):
            if globals().get(flag_name, False):
                return view_func(*args, **kwargs)
            abort(404)

        return wrapped

    return decorator


require_ocr_features = require_feature_enabled('OCR_FEATURES_ENABLED')
require_sector_analytics = require_feature_enabled('SECTOR_ANALYTICS_ENABLED')
require_tf_risk = require_feature_enabled('TF_RISK_ENABLED')


def get_openai_api_key():
    return (os.getenv('OPENAI_API_KEY') or '').strip()


def seed_admin_emails():
    return {item.lower() for item in env_csv("ADMIN_SEEDED_EMAILS")}


def get_reset_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'])


def generate_reset_token(email):
    serializer = get_reset_serializer()
    return serializer.dumps(email, salt='password-reset')


def verify_reset_token(token):
    serializer = get_reset_serializer()
    max_age = app.config.get('RESET_TOKEN_MAX_AGE_SECONDS', 3600)
    return serializer.loads(token, salt='password-reset', max_age=max_age)


def running_schema_command():
    args = " ".join(sys.argv).lower()
    return (
        "reset-db" in args
        or "init-db" in args
        or "sync-db-schema" in args
        or env_flag("SKIP_SCHEMA_CHECK", False)
    )


def schema_work_disabled():
    return env_flag("SKIP_SCHEMA_CHECK", False)


def database_url_dialect(database_url):
    normalized = str(database_url or "").strip().lower()
    if normalized.startswith("sqlite:///"):
        return "sqlite"
    if normalized.startswith("postgres://") or normalized.startswith("postgresql://") or normalized.startswith("postgresql+"):
        return "postgresql"
    if normalized.startswith("mysql://") or normalized.startswith("mysql+"):
        return "mysql"
    return ""


def database_url_is_sqlite(database_url):
    return database_url_dialect(database_url) == "sqlite"


def database_url_is_postgresql(database_url):
    return database_url_dialect(database_url) == "postgresql"


def database_url_is_mysql(database_url):
    return database_url_dialect(database_url) == "mysql"


def build_database_url_from_parts(scheme, *, host=None, port=None, database_name=None, username=None, password=None, query_params=None):
    if not scheme:
        return None
    if not database_name:
        return None

    netloc = ""
    if username:
        netloc = quote_plus(username)
        if password is not None:
            netloc = f"{netloc}:{quote_plus(password)}"
        if host:
            netloc = f"{netloc}@{host}"
    elif host:
        netloc = host

    if host and port:
        netloc = f"{netloc}:{port}"

    query_string = urlencode(query_params or {}, doseq=True)
    return urlunparse(
        (
            scheme,
            netloc,
            f"/{str(database_name).lstrip('/')}",
            "",
            query_string,
            "",
        )
    )


def build_database_url_from_env_parts():
    dialect = (env_first("DATABASE_DIALECT", "DB_DIALECT", "DB_BACKEND") or "").lower()
    if not dialect:
        return None

    if dialect in {"sqlite", "sqlite3"}:
        sqlite_path = env_first("SQLITE_DATABASE_PATH", "SQLITE_PATH", "DB_PATH")
        if not sqlite_path:
            return None
        return f"sqlite:///{sqlite_path}"

    if dialect in {"postgres", "postgresql"}:
        query_params = {}
        sslmode = env_first("DB_SSL_MODE", "POSTGRES_SSL_MODE")
        if sslmode:
            query_params["sslmode"] = sslmode
        return build_database_url_from_parts(
            "postgresql",
            host=env_first("DB_HOST", "POSTGRES_HOST"),
            port=env_first("DB_PORT", "POSTGRES_PORT",) or "5432",
            database_name=env_first("DB_NAME", "POSTGRES_DB", "POSTGRES_DATABASE"),
            username=env_first("DB_USER", "POSTGRES_USER"),
            password=env_first("DB_PASSWORD", "POSTGRES_PASSWORD"),
            query_params=query_params,
        )

    if dialect == "mysql":
        mysql_driver = env_first("MYSQL_DRIVER", "DB_DRIVER") or "pymysql"
        query_params = {
            "charset": env_first("MYSQL_CHARSET", "DB_CHARSET") or "utf8mb4",
        }
        return build_database_url_from_parts(
            f"mysql+{mysql_driver}",
            host=env_first("MYSQL_HOST", "DB_HOST") or "localhost",
            port=env_first("MYSQL_PORT", "DB_PORT") or "3306",
            database_name=env_first("MYSQL_DATABASE", "MYSQL_DB", "DB_NAME"),
            username=env_first("MYSQL_USER", "DB_USER"),
            password=env_first("MYSQL_PASSWORD", "DB_PASSWORD"),
            query_params=query_params,
        )

    return None


def redact_database_url(database_url):
    raw_value = str(database_url or "").strip()
    if not raw_value:
        return raw_value
    if database_url_is_sqlite(raw_value):
        return raw_value

    parsed = urlparse(raw_value)
    if parsed.hostname is None:
        return raw_value

    username = parsed.username or ""
    auth = quote_plus(username) if username else ""
    if parsed.password is not None:
        auth = f"{auth}:***" if auth else "***"
    netloc = auth
    if netloc:
        netloc = f"{netloc}@{parsed.hostname}"
    else:
        netloc = parsed.hostname
    if parsed.port:
        netloc = f"{netloc}:{parsed.port}"

    return urlunparse((parsed.scheme, netloc, parsed.path, parsed.params, parsed.query, parsed.fragment))


def default_postgresql_sslmode(database_url):
    configured_sslmode = env_first("DB_SSL_MODE", "POSTGRES_SSL_MODE")
    if configured_sslmode:
        return configured_sslmode

    parsed = urlparse(str(database_url or ""))
    host = parsed.hostname
    if not host:
        return "require"
    if host in {"localhost", "127.0.0.1", "::1"}:
        return "disable"

    try:
        host_ip = ipaddress.ip_address(host)
    except ValueError:
        return "require"

    if host_ip.is_private or host_ip.is_loopback or host_ip.is_link_local:
        return "disable"

    return "require"


def normalize_database_url(raw_url):
    if not raw_url:
        return 'sqlite:///form14.sqlite'

    database_url = raw_url.strip()
    if database_url.startswith('sqlite:///') and not database_url.startswith('sqlite:////'):
        raw_path = database_url.replace('sqlite:///', '', 1)
        if raw_path and raw_path != ':memory:':
            absolute_path = Path(app.root_path, raw_path).resolve()
            database_url = f'sqlite:///{absolute_path}'

    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)

    if database_url.startswith('mysql://'):
        database_url = database_url.replace('mysql://', 'mysql+pymysql://', 1)

    if database_url_is_postgresql(database_url):
        parsed = urlparse(database_url)
        query_params = parse_qs(parsed.query, keep_blank_values=True)
        if 'sslmode' not in {key.lower() for key in query_params}:
            query_params['sslmode'] = [default_postgresql_sslmode(database_url)]
            database_url = urlunparse((
                parsed.scheme,
                parsed.netloc,
                parsed.path,
                parsed.params,
                urlencode(query_params, doseq=True),
                parsed.fragment,
            ))

    if database_url_is_mysql(database_url):
        parsed = urlparse(database_url)
        query_params = parse_qs(parsed.query, keep_blank_values=True)
        if 'charset' not in {key.lower() for key in query_params}:
            query_params['charset'] = ['utf8mb4']
            database_url = urlunparse((
                parsed.scheme,
                parsed.netloc,
                parsed.path,
                parsed.params,
                urlencode(query_params, doseq=True),
                parsed.fragment,
            ))

    return database_url


def is_production_environment():
    app_env = (os.getenv('APP_ENV') or '').strip().lower()
    render_hostname = os.getenv('RENDER_EXTERNAL_HOSTNAME')
    return (
        app_env == 'production'
        or env_flag('RENDER', False)
        or env_flag('FLASK_ENV_PRODUCTION', False)
        or bool(render_hostname)
    )


def resolve_primary_database_url():
    configured_url = env_first('INTERNAL_DATABASE_URL')
    return normalize_database_url(configured_url or build_database_url_from_env_parts())


def resolve_backup_database_url():
    return resolve_primary_database_url()


def resolve_sqlite_bootstrap_source_url():
    configured_url = env_first('SQLITE_BOOTSTRAP_SOURCE_URL', 'SQLITE_IMPORT_SOURCE_URL')
    if configured_url:
        return normalize_database_url(configured_url)

    for candidate in (
        'db_bootstrap/returnsform14_org_backup.sqlite',
        'returnsform14_org_backup.sqlite',
    ):
        candidate_path = resolve_directory_path(candidate, base_path=app.root_path)
        if candidate_path is not None and candidate_path.exists():
            return normalize_database_url(f'sqlite:///{candidate_path}')

    return normalize_database_url('sqlite:///returnsform14_org_backup.sqlite')


def resolve_directory_path(path_value, base_path=None):
    raw_value = str(path_value or '').strip()
    if not raw_value:
        return None
    candidate = Path(raw_value).expanduser()
    if not candidate.is_absolute():
        anchor = Path(base_path).expanduser() if base_path else Path(app.root_path)
        if not anchor.is_absolute():
            anchor = (Path(app.root_path) / anchor).resolve()
        candidate = anchor / candidate
    return candidate.resolve()


def resolve_storage_root_directory():
    configured_root = resolve_directory_path(
        os.getenv('PERSISTENT_STORAGE_ROOT'),
        base_path=app.root_path,
    )
    if configured_root is not None:
        return configured_root
    return Path(app.instance_path).resolve()


def resolve_storage_directory(env_name, default_name, storage_root=None):
    root = Path(storage_root or resolve_storage_root_directory()).resolve()
    configured_path = resolve_directory_path(os.getenv(env_name), base_path=root)
    if configured_path is not None:
        return configured_path
    return (root / default_name).resolve()


def configured_storage_directory(config_key, default_name):
    storage_root = resolve_directory_path(
        app.config.get('PERSISTENT_STORAGE_ROOT'),
        base_path=app.root_path,
    ) or Path(app.instance_path).resolve()
    configured_path = resolve_directory_path(app.config.get(config_key), base_path=storage_root)
    target = configured_path or (storage_root / default_name).resolve()
    target.mkdir(parents=True, exist_ok=True)
    app.config[config_key] = str(target)
    return target


def _load_database_backup_module():
    import database_backup

    return database_backup


def run_backup_once(*args, **kwargs):
    return _load_database_backup_module().run_backup_once(*args, **kwargs)


def build_drive_service(*args, **kwargs):
    return _load_database_backup_module().build_drive_service(*args, **kwargs)


def build_sqlalchemy_engine_options(database_url):
    engine_options = dict(app.config.get('SQLALCHEMY_ENGINE_OPTIONS') or {})

    pool_size = os.getenv('SQLALCHEMY_POOL_SIZE')
    max_overflow = os.getenv('SQLALCHEMY_MAX_OVERFLOW')
    connect_timeout = os.getenv('SQLALCHEMY_CONNECT_TIMEOUT')

    if pool_size:
        engine_options['pool_size'] = int(pool_size)
    if max_overflow:
        engine_options['max_overflow'] = int(max_overflow)

    connect_args = dict(engine_options.get('connect_args') or {})
    if database_url_is_sqlite(database_url):
        connect_args.setdefault('check_same_thread', False)
    elif connect_timeout and (database_url_is_mysql(database_url) or database_url_is_postgresql(database_url)):
        connect_args.setdefault('connect_timeout', int(connect_timeout))

    if connect_args:
        engine_options['connect_args'] = connect_args

    return engine_options


SQLITE_TO_MYSQL_BOOTSTRAP_SETTING_KEY = 'sqlite_mysql_bootstrap_import_completed'
DATABASE_DUMP_BOOTSTRAP_SETTING_KEY = 'database_dump_bootstrap_import_completed'
DATABASE_DUMP_BOOTSTRAP_CANDIDATES = (
    'db_bootstrap/returnsform14_org_backup.mysql.sql',
    'db_bootstrap/returnsform14_org_backup.dump',
    'db_bootstrap/returnsform14_org_backup.sql',
    'db_bootstrap/returnsform14_org_backup.sqlite',
    'db_bootstrap/returnsform14_org_backup.db',
)


# 🔹 DATABASE + SECURITY CONFIGURATION
database_url = resolve_primary_database_url()
raw_allowed_hosts = (
    env_csv('ALLOWED_HOSTS')
    + env_csv('ADDITIONAL_ALLOWED_HOSTS')
    + env_csv('NGROK_ALLOWED_HOSTS')
)
allowed_hosts = resolve_trusted_hosts(raw_allowed_hosts, app_host_ip=os.getenv('PBORA_APP_HOST_IP'))
render_hostname = os.getenv('RENDER_EXTERNAL_HOSTNAME')
is_production = is_production_environment()
development_hosts = ['127.0.0.1', 'localhost']
storage_root = resolve_storage_root_directory()
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = build_sqlalchemy_engine_options(database_url)

if render_hostname and render_hostname not in allowed_hosts:
    allowed_hosts.append(render_hostname)

for host in development_hosts:
    if host not in allowed_hosts:
        allowed_hosts.append(host)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'change-me-in-env')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['REMEMBER_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = os.getenv('SESSION_COOKIE_SAMESITE', 'Lax')
app.config['SESSION_COOKIE_SECURE'] = env_flag('SESSION_COOKIE_SECURE', is_production)
app.config['REMEMBER_COOKIE_SECURE'] = env_flag('SESSION_COOKIE_SECURE', is_production)
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
app.config['REMEMBER_COOKIE_REFRESH_EACH_REQUEST'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(
    days=int(os.getenv('PERMANENT_SESSION_LIFETIME_DAYS', '365'))
)
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(
    days=int(os.getenv('REMEMBER_COOKIE_DURATION_DAYS', '365'))
)
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_CONTENT_LENGTH', str(16 * 1024 * 1024)))
app.config['PERSISTENT_STORAGE_ROOT'] = str(storage_root)
app.config['PREFERRED_URL_SCHEME'] = os.getenv(
    'PREFERRED_URL_SCHEME',
    'https' if app.config['SESSION_COOKIE_SECURE'] else 'http',
)
app.config['UPLOAD_FOLDER'] = str(
    resolve_storage_directory('UPLOAD_FOLDER', 'uploads', storage_root=storage_root)
)
app.config['REPORT_MERGE_EXPORT_DIR'] = str(
    resolve_storage_directory('REPORT_MERGE_EXPORT_DIR', 'report_merge_exports', storage_root=storage_root)
)
app.config['REPORT_MERGE_TEMP_DIR'] = str(
    resolve_storage_directory('REPORT_MERGE_TEMP_DIR', 'report_merge_temp', storage_root=storage_root)
)
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', '587'))
app.config['MAIL_USE_TLS'] = env_flag('MAIL_USE_TLS', True)
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_DEFAULT_SENDER', app.config['MAIL_USERNAME'])
app.config['GOOGLE_DRIVE_ENABLED'] = env_flag('GOOGLE_DRIVE_ENABLED', False)
app.config['DB_BACKUP_INTERVAL_SECONDS'] = int(os.getenv('DB_BACKUP_INTERVAL_SECONDS', '3600'))
app.config['APP_TIMEZONE'] = os.getenv('APP_TIMEZONE', 'Africa/Nairobi')

ALLOWED_UPLOAD_EXTENSIONS = {'csv', 'xlsx', 'xls', 'pdf', 'doc', 'docx', 'png', 'jpg', 'jpeg'}
REPORT_STATUS_VALUES = set(MODEL_REPORT_WORKFLOW_STATUS_VALUES)
REVIEW_STATUS_VALUES = set(MODEL_REPORT_REVIEW_STATUS_VALUES)
IMPORT_BATCH_STATUS_VALUES = set(MODEL_IMPORT_BATCH_STATUS_VALUES)
UPLOADED_FILE_STATUS_VALUES = set(MODEL_UPLOADED_FILE_STATUS_VALUES)
TRACKED_REPORT_FIELDS = [
    'pbo_name',
    'pbo_registration_number',
    'reporting_period_start',
    'reporting_period_end',
    'return_date',
    'audited',
    'late_returns',
    'outstanding_penalty',
    'date_received',
    'received_by',
    'designate_by_pco',
    'workflow_status',
    'review_status',
    'review_notes',
    'return_reason',
    'duplicate_flag',
    'data_source',
]
BACKUP_RUNTIME_STATUS = {
    'running': False,
    'stage': 'idle',
    'detail': None,
    'started_at': None,
    'last_backup_at': None,
    'last_backup_archive': None,
    'last_backup_error': None,
    'backup_dir': None,
    'files_total': 0,
    'files_completed': 0,
    'tables_total': 0,
    'tables_completed': 0,
    'current_file': None,
    'current_table': None,
    'drive_upload_succeeded': None,
    'local_backup_succeeded': None,
    'latest_drive_upload': None,
    'drive_refresh_command': None,
    'drive_auth_recommendation': None,
    'drive_credential_type': None,
    'drive_credential_source': None,
}
BACKUP_RUNTIME_LOCK = threading.Lock()
NOON_BACKUP_WORKER_STARTED = False
BACKUP_ROLLBACK_UPLOAD_DIR = Path('/tmp/form14_rollback_uploads')
BACKUP_VERSION_CACHE = {
    'updated_at': 0.0,
    'payload': None,
}
REPORT_MERGE_RUNTIME_STATUS = {
    'running': False,
    'stage': 'idle',
    'detail': None,
    'started_at': None,
    'finished_at': None,
    'current_source': None,
    'sources_total': 0,
    'sources_completed': 0,
    'progress_percent': 0,
    'selected_sources': [],
    'download_ready': False,
    'download_token': None,
    'download_name': None,
    'download_url': None,
    'last_error': None,
    'reference_name': None,
}
REPORT_MERGE_RUNTIME_LOCK = threading.Lock()
REPORT_MERGE_OUTPUTS: dict[str, dict[str, str]] = {}
DATA_INTERPRETATION_JOB_LOCK = threading.Lock()
if allowed_hosts:
    app.config['TRUSTED_HOSTS'] = allowed_hosts
    app.trusted_hosts = allowed_hosts

NOMINATIM_URL = 'https://nominatim.openstreetmap.org/search'
NOMINATIM_USER_AGENT = os.getenv(
    'NOMINATIM_USER_AGENT',
    'form14-app/1.0 (admin@example.com)'
)
NOMINATIM_HEADERS = {'User-Agent': NOMINATIM_USER_AGENT}
NOMINATIM_MIN_INTERVAL_SECONDS = 1.0
NOMINATIM_LAST_REQUEST_AT = 0.0
NOMINATIM_RATE_LOCK = threading.Lock()
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1)


# 🔹 INITIALIZE DB
db.init_app(app)
if VISUAL_ANALYSIS_ENABLED:
    from analysis_view import analysis_bp
    from analysis2_view import analysis2_bp

    app.register_blueprint(analysis_bp)
    app.register_blueprint(analysis2_bp)

# --- CLI COMMAND TO UPDATE ALL RISK SCORES ---
@app.cli.command('update-all-risk-scores')
@with_appcontext
def update_all_risk_scores():
    """Batch update all PBOReport risk scores using compute_tf_risk."""
    from models import PBOReport, db
    updated = 0
    reports = PBOReport.query.all()
    for report in reports:
        report.update_risk_score(compute_tf_risk)
        updated += 1
    db.session.commit()
    click.echo(f"Updated risk scores for {updated} reports.")


@app.cli.command('init-db')
@with_appcontext
def init_db():
    """Create missing tables from models.py and seed users."""
    initialize_database(reset=False, seed_users=env_flag('SEED_USERS_ON_STARTUP', True))
    click.echo("Database initialization complete.")


@app.cli.command('sync-db-schema')
@with_appcontext
def sync_db_schema():
    """Apply non-destructive schema sync so the live DB matches models.py."""
    initialize_database(
        reset=False,
        seed_users=False,
        apply_schema_changes=True,
        run_postgresql_audit=False,
        sync_default_admin=False,
    )
    click.echo("Database schema sync complete.")


@app.cli.command('import-sqlite-to-mysql')
@click.option('--source-url', default=None, help='SQLite source URL. Defaults to SQLITE_BOOTSTRAP_SOURCE_URL or sqlite:///returnsform14_org_backup.sqlite')
@click.option('--force', is_flag=True, help='Import even if MySQL already has rows or the completion marker exists.')
@with_appcontext
def import_sqlite_to_mysql_command(source_url=None, force=False):
    """One-time bootstrap import from SQLite into the current MySQL database."""
    result = import_sqlite_into_current_mysql_database(source_url=source_url, force=force)
    click.echo(json.dumps(result, indent=2, sort_keys=True))

def report_database_state():
    if running_schema_command():
        return
    database_uri = app.config['SQLALCHEMY_DATABASE_URI']
    if database_url_is_sqlite(database_uri):
        relative_path = database_uri.replace('sqlite:///', '', 1)
        db_path = relative_path if os.path.isabs(relative_path) else os.path.join(app.root_path, relative_path)
        db_dir = os.path.dirname(db_path)
        if db_dir:
            os.makedirs(db_dir, exist_ok=True)
        print(f"✅ Database configured: {db_path}")
        return

    try:
        print(f"✅ Database ready: {redact_database_url(database_uri)}")
    except Exception as exc:
        # Handle transient DB host resolution/network errors gracefully in dev mode
        if 'could not translate host name' in str(exc) or 'Could not connect' in str(exc) or 'OperationalError' in type(exc).__name__:
            print(f"⚠️  Temporary DB connection issue: {exc}")
            print("⚠️  Continuing startup in limited mode. Some DB-dependent features may not work until DB is available.")
            return
        raise


def admin_setting_value(key):
    setting = AdminSetting.query.filter_by(key=key).first()
    return setting.value if setting is not None else None


def target_database_first_nonempty_app_table():
    existing_table_names = set(inspect(db.engine).get_table_names())
    for table in db.metadata.sorted_tables:
        if table.name not in existing_table_names:
            continue
        row = db.session.execute(select(literal_column('1')).select_from(table).limit(1)).first()
        if row is not None:
            return table.name
    return None


def target_database_has_existing_app_tables():
    existing_table_names = set(inspect(db.engine).get_table_names())
    return any(table.name in existing_table_names for table in db.metadata.sorted_tables)


def source_database_first_nonempty_app_table(source_connection):
    source_inspector = inspect(source_connection)
    source_table_names = set(source_inspector.get_table_names())

    for table in db.metadata.sorted_tables:
        if table.name not in source_table_names:
            continue
        source_table = Table(table.name, MetaData(), autoload_with=source_connection)
        row = source_connection.execute(
            select(literal_column('1')).select_from(source_table).limit(1)
        ).first()
        if row is not None:
            return table.name
    return None


def set_admin_setting_value(key, value):
    setting = AdminSetting.query.filter_by(key=key).first()
    if setting is None:
        setting = AdminSetting(key=key)
        db.session.add(setting)
    setting.value = value
    return setting


def admin_setting_value_if_available(key):
    if AdminSetting.__tablename__ not in set(inspect(db.engine).get_table_names()):
        return None
    return admin_setting_value(key)


def resolve_database_dump_bootstrap_source_path(source_path=None):
    configured_path = source_path or env_first(
        'DATABASE_DUMP_BOOTSTRAP_PATH',
        'DATABASE_BOOTSTRAP_SOURCE_PATH',
        'DB_BOOTSTRAP_SOURCE_PATH',
    )
    if configured_path:
        return resolve_directory_path(configured_path, base_path=app.root_path)

    for candidate in DATABASE_DUMP_BOOTSTRAP_CANDIDATES:
        candidate_path = resolve_directory_path(candidate, base_path=app.root_path)
        if candidate_path is not None and candidate_path.exists():
            return candidate_path
    return None


def database_dump_bootstrap_mode(payload_path, target_dialect=None):
    dialect_name = target_dialect or db.engine.dialect.name
    file_name = payload_path.name.lower()
    suffix = payload_path.suffix.lower()
    if dialect_name == 'mysql' and suffix == '.sql' and not file_name.endswith('.postgres.sql'):
        return 'mysql_sql'
    if dialect_name == 'postgresql' and suffix == '.dump':
        return 'postgresql_dump'
    if dialect_name == 'postgresql' and suffix == '.sql' and not file_name.endswith('.mysql.sql'):
        return 'postgresql_dump'
    if dialect_name == 'sqlite' and suffix in {'.sqlite', '.db'}:
        return 'sqlite_file'
    return None


def drop_application_tables_if_present():
    if not target_database_has_existing_app_tables():
        return False
    db.session.remove()
    db.drop_all()
    try:
        db.engine.dispose()
    except Exception:
        pass
    return True


def record_database_dump_bootstrap(result):
    if not result or not result.get('performed'):
        return
    payload = json.dumps({
        'source_path': result.get('source_path'),
        'bootstrap_mode': result.get('bootstrap_mode'),
        'target_dialect': result.get('target_dialect'),
        'restored_at': utc_now().isoformat(timespec='seconds'),
    })
    set_admin_setting_value(DATABASE_DUMP_BOOTSTRAP_SETTING_KEY, payload)
    db.session.commit()


def _source_select_for_import(source_connection, source_table, target_table):
    source_columns = {column.name for column in source_table.columns}
    target_columns = [column.name for column in target_table.columns if column.name in source_columns]
    select_columns = []
    use_sqlite_rowid_as_id = (
        source_connection.dialect.name == 'sqlite'
        and 'id' in target_columns
        and not sqlite_table_has_real_id_primary_key(source_connection, source_table.name)
    )
    source_rowid_label = '__source_sqlite_rowid'
    if use_sqlite_rowid_as_id:
        select_columns.append(literal_column('rowid').label(source_rowid_label))
    select_columns.extend(source_table.c[column_name] for column_name in target_columns)
    return (
        select(*select_columns).select_from(source_table),
        target_columns,
        use_sqlite_rowid_as_id,
        source_rowid_label,
    )


def import_database_dump_into_current_database(source_path=None, *, force=False):
    if not force and not env_flag('AUTO_IMPORT_DATABASE_DUMP_ON_STARTUP', False):
        return {'performed': False, 'reason': 'auto_import_disabled'}

    marker_value = admin_setting_value_if_available(DATABASE_DUMP_BOOTSTRAP_SETTING_KEY)
    if marker_value and not force:
        return {'performed': False, 'reason': 'marker_exists', 'marker_value': marker_value}

    target_has_rows = target_database_first_nonempty_app_table()
    if target_has_rows and not force:
        return {'performed': False, 'reason': 'target_already_has_rows', 'table': target_has_rows}

    resolved_source_path = resolve_database_dump_bootstrap_source_path(source_path)
    if resolved_source_path is None:
        return {'performed': False, 'reason': 'source_path_not_configured'}
    if not resolved_source_path.exists():
        return {
            'performed': False,
            'reason': 'source_dump_missing',
            'source_path': str(resolved_source_path),
        }

    bootstrap_mode = database_dump_bootstrap_mode(resolved_source_path)
    if bootstrap_mode is None:
        return {
            'performed': False,
            'reason': 'unsupported_dump_for_target_database',
            'source_path': str(resolved_source_path),
            'target_dialect': db.engine.dialect.name,
        }

    dropped_existing_tables = drop_application_tables_if_present()

    db.session.remove()
    try:
        db.engine.dispose()
    except Exception:
        pass

    database_url = app.config['SQLALCHEMY_DATABASE_URI']
    if bootstrap_mode == 'mysql_sql':
        _restore_mysql_from_payload(database_url, resolved_source_path)
    elif bootstrap_mode == 'postgresql_dump':
        _restore_postgres_from_payload(database_url, resolved_source_path)
    else:
        _restore_sqlite_from_payload(database_url, resolved_source_path)

    try:
        db.engine.dispose()
    except Exception:
        pass

    print(
        '✅ Imported deployment bootstrap dump into current database: '
        f'{resolved_source_path.name} -> {db.engine.dialect.name}.'
    )
    return {
        'performed': True,
        'source_path': str(resolved_source_path),
        'bootstrap_mode': bootstrap_mode,
        'target_dialect': db.engine.dialect.name,
        'dropped_existing_tables': dropped_existing_tables,
    }


def import_sqlite_into_current_mysql_database(source_url=None, *, force=False, batch_size=500):
    if db.engine.dialect.name != 'mysql':
        return {'performed': False, 'reason': 'target_database_is_not_mysql'}

    if not force and not env_flag('AUTO_IMPORT_SQLITE_TO_MYSQL_ON_STARTUP', True):
        return {'performed': False, 'reason': 'auto_import_disabled'}

    marker_value = admin_setting_value(SQLITE_TO_MYSQL_BOOTSTRAP_SETTING_KEY)
    if marker_value and not force:
        return {'performed': False, 'reason': 'marker_exists', 'marker_value': marker_value}

    target_has_rows = target_database_first_nonempty_app_table()
    if target_has_rows and not force:
        return {'performed': False, 'reason': 'target_already_has_rows', 'table': target_has_rows}

    resolved_source_url = normalize_database_url(source_url or resolve_sqlite_bootstrap_source_url())
    if not database_url_is_sqlite(resolved_source_url):
        return {'performed': False, 'reason': 'source_database_is_not_sqlite', 'source_url': resolved_source_url}

    source_path = _resolve_sqlite_path_from_url(resolved_source_url)
    if source_path is None or not source_path.exists():
        return {
            'performed': False,
            'reason': 'source_sqlite_missing',
            'source_url': resolved_source_url,
            'source_path': str(source_path) if source_path is not None else None,
        }

    source_engine = create_engine(resolved_source_url)
    imported_rows = {}

    try:
        with source_engine.connect() as source_connection:
            source_has_rows = source_database_first_nonempty_app_table(source_connection)
            if source_has_rows is None:
                return {
                    'performed': False,
                    'reason': 'source_sqlite_empty',
                    'source_url': resolved_source_url,
                    'source_path': str(source_path),
                }

            source_table_names = set(inspect(source_connection).get_table_names())
            db.session.execute(text('SET FOREIGN_KEY_CHECKS=0'))
            try:
                for target_table in db.metadata.sorted_tables:
                    if target_table.name not in source_table_names:
                        continue

                    source_table = Table(target_table.name, MetaData(), autoload_with=source_connection)
                    statement, target_columns, use_sqlite_rowid_as_id, source_rowid_label = _source_select_for_import(
                        source_connection,
                        source_table,
                        target_table,
                    )
                    if not target_columns:
                        continue

                    inserted_count = 0
                    batch = []
                    for row in source_connection.execute(statement).mappings():
                        payload = {column_name: row.get(column_name) for column_name in target_columns}
                        if use_sqlite_rowid_as_id and row.get(source_rowid_label) is not None:
                            payload['id'] = int(row.get(source_rowid_label))
                        batch.append(payload)
                        if len(batch) >= batch_size:
                            db.session.execute(target_table.insert(), batch)
                            inserted_count += len(batch)
                            batch = []

                    if batch:
                        db.session.execute(target_table.insert(), batch)
                        inserted_count += len(batch)

                    imported_rows[target_table.name] = inserted_count

                marker_payload = json.dumps({
                    'source_url': resolved_source_url,
                    'source_path': str(source_path),
                    'imported_at': utc_now().isoformat(timespec='seconds'),
                    'imported_tables': {key: value for key, value in imported_rows.items() if value > 0},
                })
                set_admin_setting_value(SQLITE_TO_MYSQL_BOOTSTRAP_SETTING_KEY, marker_payload)
                db.session.commit()
            except Exception:
                db.session.rollback()
                raise
            finally:
                try:
                    db.session.execute(text('SET FOREIGN_KEY_CHECKS=1'))
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                    raise
    finally:
        source_engine.dispose()

    imported_table_count = sum(1 for row_count in imported_rows.values() if row_count > 0)
    imported_row_total = sum(imported_rows.values())
    print(
        '✅ Imported SQLite records into MySQL bootstrap target: '
        f'{imported_row_total} row(s) across {imported_table_count} table(s).'
    )
    return {
        'performed': True,
        'source_url': resolved_source_url,
        'source_path': str(source_path),
        'source_first_nonempty_table': source_has_rows,
        'imported_rows': imported_rows,
        'imported_row_total': imported_row_total,
        'imported_table_count': imported_table_count,
    }


def auto_import_sqlite_into_empty_mysql_database():
    if db.engine.dialect.name != 'mysql':
        return {'performed': False, 'reason': 'target_database_is_not_mysql'}
    return import_sqlite_into_current_mysql_database(force=False)


def seed_configured_users():
    try:
        from create_users import seed_users_from_runtime
        seed_users_from_runtime(app_instance=app, initialize_database_fn=initialize_database)
    except Exception as exc:
        print(f"⚠️ User seeding skipped: {exc}")


def clear_application_tables(preserve_tables=None):
    preserve_tables = set(preserve_tables or [])
    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())

    if db.engine.dialect.name == 'sqlite':
        db.session.execute(text('PRAGMA foreign_keys=OFF'))

    for table in reversed(db.metadata.sorted_tables):
        if table.name in preserve_tables or table.name not in existing_tables:
            continue
        db.session.execute(table.delete())

    if db.engine.dialect.name == 'sqlite':
        db.session.execute(text('PRAGMA foreign_keys=ON'))

    db.session.commit()


def sql_literal(value, dialect_name):
    if value is None:
        return 'NULL'
    if isinstance(value, bool):
        if dialect_name in {'sqlite', 'mysql'}:
            return '1' if value else '0'
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (datetime, date)):
        return f"'{value.isoformat()}'"
    escaped = str(value).replace("'", "''")
    return f"'{escaped}'"


def dialect_quote_identifier(identifier):
    identifier = str(identifier).replace('\x00', '')
    dialect_name = getattr(db.engine.dialect, 'name', '')
    if dialect_name == 'postgresql':
        return '"' + identifier.replace('\"', '\"\"') + '"'
    if dialect_name in {'mysql', 'mariadb'}:
        return '`' + identifier.replace('`', '``') + '`'
    return db.engine.dialect.identifier_preparer.quote(identifier)


def build_add_column_statement(table_name, column, dialect_name):
    column_type = column.type.compile(dialect=db.engine.dialect)
    quoted_table = dialect_quote_identifier(table_name)
    quoted_column = dialect_quote_identifier(column.name)
    statement = f'ALTER TABLE {quoted_table} ADD COLUMN {quoted_column} {column_type}'

    scalar_default = None
    if column.default is not None:
        default_arg = getattr(column.default, 'arg', None)
        if not callable(default_arg):
            scalar_default = default_arg

    if scalar_default is not None:
        statement = f'{statement} DEFAULT {sql_literal(scalar_default, dialect_name)}'

    if not column.nullable and not column.primary_key:
        if scalar_default is None:
            return None
        statement = f'{statement} NOT NULL'

    return statement


def build_modify_column_statement(table_name, column, dialect_name):
    if dialect_name != 'mysql':
        return None

    quoted_table = dialect_quote_identifier(table_name)
    quoted_column = dialect_quote_identifier(column.name)
    column_type = column.type.compile(dialect=db.engine.dialect)
    statement = f'ALTER TABLE {quoted_table} MODIFY COLUMN {quoted_column} {column_type}'

    scalar_default = None
    if column.default is not None:
        default_arg = getattr(column.default, 'arg', None)
        if not callable(default_arg):
            scalar_default = default_arg

    if not column.nullable and not column.primary_key:
        statement = f'{statement} NOT NULL'
    elif column.nullable:
        statement = f'{statement} NULL'

    if scalar_default is not None and not isinstance(column.type, Text):
        statement = f'{statement} DEFAULT {sql_literal(scalar_default, dialect_name)}'

    return statement


def ensure_model_columns(model):
    """Add model columns that are missing from an existing table."""
    inspector = inspect(db.engine)
    table_name = model.__table__.name
    if table_name not in inspector.get_table_names():
        return [], []

    existing_columns = {column['name'] for column in inspector.get_columns(table_name)}
    dialect_name = db.engine.dialect.name
    added_columns = []
    skipped_columns = []

    for column in model.__table__.columns:
        if column.name in existing_columns or column.primary_key:
            continue

        statement = build_add_column_statement(table_name, column, dialect_name)
        if statement is None:
            skipped_columns.append(column.name)
            continue

        db.session.execute(text(statement))
        added_columns.append(column.name)

    if added_columns:
        db.session.commit()
        print(f"✅ Added missing columns to {table_name}: {', '.join(sorted(added_columns))}")

    if skipped_columns:
        print(
            f"⚠️ Skipped non-null columns without scalar defaults on {table_name}: "
            f"{', '.join(sorted(skipped_columns))}"
        )

    return added_columns, skipped_columns


def compiled_column_type_name(column_type):
    if column_type is None:
        return None
    compile_method = getattr(column_type, 'compile', None)
    if callable(compile_method):
        try:
            return str(compile_method(dialect=db.engine.dialect)).upper()
        except Exception:
            pass
    return str(column_type).upper()


def ensure_model_string_capacities(model, column_names=None):
    """Widen legacy VARCHAR columns or promote them to TEXT when the model allows it."""
    inspector = inspect(db.engine)
    table_name = model.__table__.name
    if table_name not in inspector.get_table_names():
        return []

    dialect_name = db.engine.dialect.name
    existing_columns = {column['name']: column for column in inspector.get_columns(table_name)}
    widened_columns = []
    quoted_table = dialect_quote_identifier(table_name)

    if column_names is None:
        column_names = [
            column.name
            for column in model.__table__.columns
            if isinstance(column.type, String)
            and (
                getattr(column.type, 'length', None) is not None
                or isinstance(column.type, Text)
            )
        ]

    for column_name in column_names:
        model_column = model.__table__.columns.get(column_name)
        existing_column = existing_columns.get(column_name)
        if model_column is None or existing_column is None:
            continue

        desired_length = getattr(model_column.type, 'length', None)
        current_length = getattr(existing_column.get('type'), 'length', None)
        model_uses_text = isinstance(model_column.type, Text)
        desired_compiled_type = compiled_column_type_name(model_column.type)
        current_type = existing_column.get('type')
        current_compiled_type = compiled_column_type_name(current_type)

        quoted_column = dialect_quote_identifier(column_name)

        if model_uses_text:
            if dialect_name == 'mysql':
                if current_compiled_type == desired_compiled_type:
                    continue
                statement = build_modify_column_statement(table_name, model_column, dialect_name)
                if statement:
                    db.session.execute(text(statement))
                    widened_columns.append(f"{column_name}={current_compiled_type or current_length}->{desired_compiled_type}")
                continue

            if current_length is None:
                continue

            if dialect_name == 'postgresql':
                db.session.execute(
                    text(f'ALTER TABLE {quoted_table} ALTER COLUMN {quoted_column} TYPE TEXT')
                )
                widened_columns.append(f"{column_name}={current_length}->TEXT")
                continue

            if dialect_name == 'sqlite':
                # SQLite stores VARCHAR lengths as advisory type metadata and does not enforce them.
                continue

            if dialect_name == 'mysql':
                statement = build_modify_column_statement(table_name, model_column, dialect_name)
                if statement:
                    db.session.execute(text(statement))
                    widened_columns.append(f"{column_name}={current_length}->TEXT")
                continue

            continue

        if desired_length is None or current_length is None or current_length >= desired_length:
            continue

        if dialect_name == 'postgresql':
            db.session.execute(
                text(
                    f'ALTER TABLE {quoted_table} '
                    f'ALTER COLUMN {quoted_column} TYPE VARCHAR({int(desired_length)})'
                )
            )
            widened_columns.append(f"{column_name}={current_length}->{desired_length}")
            continue

        if dialect_name == 'sqlite':
            # SQLite stores VARCHAR lengths as advisory type metadata and does not enforce them.
            continue

        if dialect_name == 'mysql':
            statement = build_modify_column_statement(table_name, model_column, dialect_name)
            if statement:
                db.session.execute(text(statement))
                widened_columns.append(f"{column_name}={current_length}->{desired_length}")
            continue

    if widened_columns:
        db.session.commit()
        print(f"✅ Widened legacy string columns on {table_name}: {', '.join(widened_columns)}")

    return widened_columns


def ensure_all_model_string_capacities(models=None):
    """Apply legacy VARCHAR widening across all registered models."""
    if schema_work_disabled():
        return

    for model in models or ALL_MODELS:
        ensure_model_string_capacities(model)


def ensure_legacy_user_schema():
    """Backfill columns added to the users table after initial deployments."""
    if schema_work_disabled():
        return

    ensure_model_columns(User)
    ensure_model_string_capacities(User)

    for boolean_column in ('must_change_password', 'is_superadmin', 'can_manage_all_records'):
        db.session.execute(
            text(
                f"UPDATE users_for_form14 "
                f"SET {boolean_column} = :false_value "
                f"WHERE {boolean_column} IS NULL"
            ),
            {'false_value': False},
        )
    db.session.execute(
        text(
            "UPDATE users_for_form14 "
            "SET failed_login_attempts = 0 "
            "WHERE failed_login_attempts IS NULL"
        )
    )
    db.session.commit()


def backfill_missing_return_dates():
    """Hard-code the legacy missing return date marker for preexisting rows."""
    if schema_work_disabled():
        return 0

    inspector = inspect(db.engine)
    table_name = PBOReport.__tablename__
    if table_name not in inspector.get_table_names():
        return 0

    existing_columns = {column['name'] for column in inspector.get_columns(table_name)}
    if 'return_date' not in existing_columns:
        return 0

    quoted_table = dialect_quote_identifier(table_name)
    quoted_column = dialect_quote_identifier('return_date')
    dialect_name = db.engine.dialect.name
    if dialect_name in {'mysql', 'mariadb'}:
        missing_predicate = f"{quoted_column} IS NULL OR TRIM(CAST({quoted_column} AS CHAR)) = ''"
    else:
        missing_predicate = f"{quoted_column} IS NULL OR TRIM(CAST({quoted_column} AS TEXT)) = ''"
    result = db.session.execute(
        text(
            f"UPDATE {quoted_table} "
            f"SET {quoted_column} = :return_date "
            f"WHERE {missing_predicate}"
        ),
        {'return_date': RETURN_DATE_PLACEHOLDER.isoformat()},
    )
    db.session.commit()
    return max(result.rowcount or 0, 0)


def ensure_legacy_report_schema():
    """Backfill columns added to pbo_reports after initial deployments."""
    if schema_work_disabled():
        return

    ensure_model_columns(PBOReport)
    backfill_missing_return_dates()
    ensure_model_string_capacities(PBOReport)


POSTGRESQL_STATUS_SPECS = (
    {
        'table_name': 'pbo_reports',
        'column_name': 'workflow_status',
        'constraint_name': 'ck_pbo_reports_workflow_status_allowed',
        'allowed_values': MODEL_REPORT_WORKFLOW_STATUS_VALUES,
        'default_value': 'draft',
        'normalizer': lambda value, default='draft': normalize_report_status(value, default=default),
    },
    {
        'table_name': 'pbo_reports',
        'column_name': 'review_status',
        'constraint_name': 'ck_pbo_reports_review_status_allowed',
        'allowed_values': MODEL_REPORT_REVIEW_STATUS_VALUES,
        'default_value': 'pending',
        'normalizer': lambda value, default='pending': normalize_review_status(value, default=default),
    },
    {
        'table_name': 'import_batches',
        'column_name': 'status',
        'constraint_name': 'ck_import_batches_status_allowed',
        'allowed_values': MODEL_IMPORT_BATCH_STATUS_VALUES,
        'default_value': 'pending',
        'normalizer': lambda value, default='pending': normalize_choice_status(value, IMPORT_BATCH_STATUS_VALUES, default),
    },
    {
        'table_name': 'uploaded_files',
        'column_name': 'status',
        'constraint_name': 'ck_uploaded_files_status_allowed',
        'allowed_values': MODEL_UPLOADED_FILE_STATUS_VALUES,
        'default_value': 'uploaded',
        'normalizer': lambda value, default='uploaded': normalize_choice_status(value, UPLOADED_FILE_STATUS_VALUES, default),
    },
)


def postgres_quote_identifier(identifier):
    return db.engine.dialect.identifier_preparer.quote(identifier)


def postgresql_table_columns(table_name):
    inspector = inspect(db.engine)
    return {column['name']: column for column in inspector.get_columns(table_name)}


def iter_required_foreign_key_specs(models=None):
    for model in models or ALL_MODELS:
        table = getattr(model, '__table__', None)
        if table is None:
            continue
        for column in table.columns:
            if column.nullable or not column.foreign_keys:
                continue
            foreign_key = next(iter(column.foreign_keys), None)
            if foreign_key is None or foreign_key.column is None:
                continue
            yield {
                'table_name': table.name,
                'column_name': column.name,
                'referred_table': foreign_key.column.table.name,
                'referred_column': foreign_key.column.name,
                'constraint_name': f"fk_{table.name}_{column.name}_{foreign_key.column.table.name}",
                'ondelete': getattr(foreign_key, 'ondelete', None),
            }


def normalize_existing_status_values(table_name, column_name, default_value, normalizer):
    quoted_table = postgres_quote_identifier(table_name)
    quoted_column = postgres_quote_identifier(column_name)
    updates = 0

    blank_result = db.session.execute(
        text(
            f"UPDATE {quoted_table} "
            f"SET {quoted_column} = :default_value "
            f"WHERE {quoted_column} IS NULL OR BTRIM(CAST({quoted_column} AS TEXT)) = ''"
        ),
        {'default_value': default_value},
    )
    updates += max(blank_result.rowcount or 0, 0)

    distinct_values = db.session.execute(
        text(f"SELECT DISTINCT {quoted_column} FROM {quoted_table} WHERE {quoted_column} IS NOT NULL")
    ).scalars().all()

    remapped_values = {}
    for raw_value in distinct_values:
        normalized_value = normalizer(raw_value, default=default_value)
        if normalized_value == raw_value:
            continue
        result = db.session.execute(
            text(
                f"UPDATE {quoted_table} "
                f"SET {quoted_column} = :normalized_value "
                f"WHERE {quoted_column} = :raw_value"
            ),
            {
                'normalized_value': normalized_value,
                'raw_value': raw_value,
            },
        )
        rowcount = max(result.rowcount or 0, 0)
        updates += rowcount
        if rowcount > 0:
            remapped_values[str(raw_value)] = normalized_value

    if updates:
        db.session.commit()

    return {
        'updated_rows': updates,
        'remapped_values': remapped_values,
    }


def status_default_matches(existing_default, expected_default):
    if existing_default in (None, ''):
        return False
    normalized_default = str(existing_default).strip().strip("'").lower()
    return expected_default.lower() in normalized_default


def collect_postgresql_status_constraint_audit():
    dialect_name = db.engine.dialect.name
    audit_payload = {
        'available': dialect_name == 'postgresql',
        'dialect': dialect_name,
        'generated_at': datetime.now(ZoneInfo(app.config['APP_TIMEZONE'])).strftime('%Y-%m-%d %H:%M:%S %Z'),
        'rows': [],
        'blocking_rows': [],
        'summary': {
            'total_columns': 0,
            'blocked_columns': 0,
            'blocked_rows': 0,
            'ready_columns': 0,
            'already_hardened_columns': 0,
        },
        'message': None,
    }
    if dialect_name != 'postgresql':
        audit_payload['message'] = 'Available only when running on PostgreSQL.'
        return audit_payload

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())
    rows = []

    for spec in POSTGRESQL_STATUS_SPECS:
        table_name = spec['table_name']
        column_name = spec['column_name']
        if table_name not in existing_tables:
            continue

        existing_columns = postgresql_table_columns(table_name)
        column_info = existing_columns.get(column_name)
        if column_info is None:
            continue

        quoted_table = postgres_quote_identifier(table_name)
        quoted_column = postgres_quote_identifier(column_name)
        allowed_sql = ", ".join(sql_literal(value, 'postgresql') for value in spec['allowed_values'])
        blank_row_count = int(
            db.session.execute(
                text(
                    f"SELECT COUNT(*) "
                    f"FROM {quoted_table} "
                    f"WHERE {quoted_column} IS NULL OR BTRIM(CAST({quoted_column} AS TEXT)) = ''"
                )
            ).scalar()
            or 0
        )
        invalid_row_count = int(
            db.session.execute(
                text(
                    f"SELECT COUNT(*) "
                    f"FROM {quoted_table} "
                    f"WHERE {quoted_column} IS NOT NULL "
                    f"AND BTRIM(CAST({quoted_column} AS TEXT)) <> '' "
                    f"AND LOWER(BTRIM(CAST({quoted_column} AS TEXT))) NOT IN ({allowed_sql})"
                )
            ).scalar()
            or 0
        )
        check_constraints = inspector.get_check_constraints(table_name)
        has_constraint = any((constraint.get('name') or '') == spec['constraint_name'] for constraint in check_constraints)
        has_expected_default = status_default_matches(column_info.get('default'), spec['default_value'])
        is_blocked = blank_row_count > 0 or invalid_row_count > 0
        is_already_hardened = (
            column_info.get('nullable') is False
            and has_constraint
            and has_expected_default
        )

        if is_blocked:
            status_key = 'blocked'
            status_label = 'Blocked'
            note = 'Legacy blank or invalid values must be reviewed first. This audit does not change them.'
        elif is_already_hardened:
            status_key = 'already_hardened'
            status_label = 'Already Hardened'
            note = 'This status column already has the expected constraint, default, and NOT NULL protection.'
        else:
            status_key = 'ready'
            status_label = 'Ready to Harden'
            note = 'No blank or invalid values were found for this status column.'

        rows.append(
            {
                'table_name': table_name,
                'column_name': column_name,
                'default_value': spec['default_value'],
                'nullable': bool(column_info.get('nullable')),
                'has_constraint': has_constraint,
                'has_expected_default': has_expected_default,
                'blank_row_count': blank_row_count,
                'invalid_row_count': invalid_row_count,
                'blocking_row_count': blank_row_count + invalid_row_count,
                'status_key': status_key,
                'status_label': status_label,
                'note': note,
            }
        )

    status_sort_order = {'blocked': 0, 'ready': 1, 'already_hardened': 2}
    rows.sort(key=lambda row: (status_sort_order.get(row['status_key'], 99), row['table_name'], row['column_name']))
    blocking_rows = [row for row in rows if row['status_key'] == 'blocked']
    ready_rows = [row for row in rows if row['status_key'] == 'ready']
    already_hardened_rows = [row for row in rows if row['status_key'] == 'already_hardened']

    audit_payload['rows'] = rows
    audit_payload['blocking_rows'] = blocking_rows
    audit_payload['summary'] = {
        'total_columns': len(rows),
        'blocked_columns': len(blocking_rows),
        'blocked_rows': sum(row['blocking_row_count'] for row in blocking_rows),
        'ready_columns': len(ready_rows),
        'already_hardened_columns': len(already_hardened_rows),
    }
    if blocking_rows:
        audit_payload['message'] = 'Blocking status values were found. Status hardening will be skipped for those columns until you review them.'
    else:
        audit_payload['message'] = 'No blocking status values were found for the inspected status columns.'

    return audit_payload


def ensure_postgresql_status_constraints():
    if schema_work_disabled() or db.engine.dialect.name != 'postgresql':
        return {'applied': [], 'skipped': []}

    status_audit = collect_postgresql_status_constraint_audit()
    applied_constraints = []
    skipped_constraints = []

    for row in status_audit['rows']:
        if row['status_key'] == 'blocked':
            skipped_constraints.append(
                {
                    'table_name': row['table_name'],
                    'column_name': row['column_name'],
                    'blank_row_count': row['blank_row_count'],
                    'invalid_row_count': row['invalid_row_count'],
                }
            )
            continue

        if row['status_key'] == 'already_hardened':
            continue

        spec = next(
            (
                item for item in POSTGRESQL_STATUS_SPECS
                if item['table_name'] == row['table_name'] and item['column_name'] == row['column_name']
            ),
            None,
        )
        if spec is None:
            continue

        quoted_table = postgres_quote_identifier(spec['table_name'])
        quoted_column = postgres_quote_identifier(spec['column_name'])
        quoted_constraint = postgres_quote_identifier(spec['constraint_name'])
        allowed_sql = ", ".join(sql_literal(value, 'postgresql') for value in spec['allowed_values'])

        db.session.execute(
            text(
                f"ALTER TABLE {quoted_table} "
                f"ALTER COLUMN {quoted_column} SET DEFAULT {sql_literal(spec['default_value'], 'postgresql')}"
            )
        )
        db.session.execute(text(f"ALTER TABLE {quoted_table} ALTER COLUMN {quoted_column} SET NOT NULL"))
        # Commit DDL before SQLAlchemy inspector opens a second connection.
        db.session.commit()

        check_constraints = inspect(db.engine).get_check_constraints(spec['table_name'])
        if not any((constraint.get('name') or '') == spec['constraint_name'] for constraint in check_constraints):
            db.session.execute(
                text(
                    f"ALTER TABLE {quoted_table} "
                    f"ADD CONSTRAINT {quoted_constraint} "
                    f"CHECK (LOWER(BTRIM({quoted_column})) IN ({allowed_sql}))"
                )
            )

        db.session.commit()
        applied_constraints.append(
            {
                'table_name': spec['table_name'],
                'column_name': spec['column_name'],
                'constraint_name': spec['constraint_name'],
            }
        )

    if applied_constraints:
        print(
            "✅ Hardened PostgreSQL status columns: "
            + ", ".join(
                f"{item['table_name']}.{item['column_name']}"
                for item in applied_constraints
            )
        )

    if skipped_constraints:
        print(
            "⚠️ Skipped PostgreSQL status hardening to preserve legacy values: "
            + ", ".join(
                (
                    f"{item['table_name']}.{item['column_name']}"
                    f" (blank={item['blank_row_count']}, invalid={item['invalid_row_count']})"
                )
                for item in skipped_constraints
            )
        )

    return {'applied': applied_constraints, 'skipped': skipped_constraints}


def ensure_postgresql_required_foreign_keys():
    if schema_work_disabled() or db.engine.dialect.name != 'postgresql':
        return {'applied': [], 'skipped': []}

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())
    repaired_relationships = []
    skipped_relationships = []

    for spec in iter_required_foreign_key_specs():
        table_name = spec['table_name']
        referred_table = spec['referred_table']
        column_name = spec['column_name']
        if table_name not in existing_tables or referred_table not in existing_tables:
            continue

        existing_columns = postgresql_table_columns(table_name)
        if column_name not in existing_columns:
            continue

        quoted_table = postgres_quote_identifier(table_name)
        quoted_column = postgres_quote_identifier(column_name)
        quoted_referred_table = postgres_quote_identifier(referred_table)
        quoted_referred_column = postgres_quote_identifier(spec['referred_column'])

        null_row_count = int(
            db.session.execute(
                text(f"SELECT COUNT(*) FROM {quoted_table} WHERE {quoted_column} IS NULL")
            ).scalar()
            or 0
        )
        orphan_row_count = int(
            db.session.execute(
                text(
                    f"SELECT COUNT(*) "
                    f"FROM {quoted_table} AS child "
                    f"WHERE child.{quoted_column} IS NOT NULL "
                    f"AND NOT EXISTS ("
                    f"    SELECT 1 FROM {quoted_referred_table} AS parent "
                    f"    WHERE parent.{quoted_referred_column} = child.{quoted_column}"
                    f")"
                )
            ).scalar()
            or 0
        )
        if null_row_count or orphan_row_count:
            skipped_relationships.append(
                {
                    'table_name': table_name,
                    'column_name': column_name,
                    'referred_table': referred_table,
                    'null_row_count': null_row_count,
                    'orphan_row_count': orphan_row_count,
                }
            )
            continue

        matching_foreign_keys = [
            foreign_key
            for foreign_key in inspect(db.engine).get_foreign_keys(table_name)
            if foreign_key.get('constrained_columns') == [column_name]
            and foreign_key.get('referred_table') == referred_table
            and foreign_key.get('referred_columns') == [spec['referred_column']]
        ]
        desired_ondelete = (spec.get('ondelete') or '').upper()
        column_is_nullable = bool(existing_columns[column_name].get('nullable', True))
        has_desired_foreign_key = any(
            (((foreign_key.get('options') or {}).get('ondelete') or '').upper() == desired_ondelete)
            for foreign_key in matching_foreign_keys
        )

        if not column_is_nullable and has_desired_foreign_key:
            continue

        changed = False
        for foreign_key in matching_foreign_keys:
            existing_ondelete = ((foreign_key.get('options') or {}).get('ondelete') or '').upper()
            if existing_ondelete == desired_ondelete:
                continue
            constraint_name = foreign_key.get('name')
            if constraint_name:
                db.session.execute(
                    text(
                        f"ALTER TABLE {quoted_table} "
                        f"DROP CONSTRAINT {postgres_quote_identifier(constraint_name)}"
                    )
                )
                changed = True

        if column_is_nullable:
            db.session.execute(text(f"ALTER TABLE {quoted_table} ALTER COLUMN {quoted_column} SET NOT NULL"))
            changed = True

        if changed:
            # Commit DDL before SQLAlchemy inspector opens a second connection.
            db.session.commit()

        foreign_keys = inspect(db.engine).get_foreign_keys(table_name)
        has_desired_foreign_key = any(
            foreign_key.get('constrained_columns') == [column_name]
            and foreign_key.get('referred_table') == referred_table
            and foreign_key.get('referred_columns') == [spec['referred_column']]
            and (((foreign_key.get('options') or {}).get('ondelete') or '').upper() == desired_ondelete)
            for foreign_key in foreign_keys
        )
        if not has_desired_foreign_key:
            ondelete_sql = f" ON DELETE {desired_ondelete}" if desired_ondelete else ""
            db.session.execute(
                text(
                    f"ALTER TABLE {quoted_table} "
                    f"ADD CONSTRAINT {postgres_quote_identifier(spec['constraint_name'])} "
                    f"FOREIGN KEY ({quoted_column}) "
                    f"REFERENCES {quoted_referred_table} ({quoted_referred_column})"
                    f"{ondelete_sql}"
                )
            )
            changed = True

        if changed:
            db.session.commit()
            repaired_relationships.append(
                {
                    'table_name': table_name,
                    'column_name': column_name,
                    'referred_table': referred_table,
                }
            )

    if repaired_relationships:
        print(
            "✅ Hardened PostgreSQL foreign keys: "
            + ", ".join(
                f"{item['table_name']}.{item['column_name']}->{item['referred_table']}"
                for item in repaired_relationships
            )
        )

    if skipped_relationships:
        print(
            "⚠️ Skipped PostgreSQL FK hardening to preserve legacy rows: "
            + ", ".join(
                (
                    f"{item['table_name']}.{item['column_name']}->{item['referred_table']}"
                    f" (null={item['null_row_count']}, orphan={item['orphan_row_count']})"
                )
                for item in skipped_relationships
            )
        )

    return {'applied': repaired_relationships, 'skipped': skipped_relationships}


def collect_postgresql_required_foreign_key_audit():
    dialect_name = db.engine.dialect.name
    audit_payload = {
        'available': dialect_name == 'postgresql',
        'dialect': dialect_name,
        'generated_at': datetime.now(ZoneInfo(app.config['APP_TIMEZONE'])).strftime('%Y-%m-%d %H:%M:%S %Z'),
        'rows': [],
        'blocking_rows': [],
        'summary': {
            'total_relations': 0,
            'blocked_relations': 0,
            'blocked_rows': 0,
            'ready_relations': 0,
            'already_hardened_relations': 0,
        },
        'message': None,
    }
    if dialect_name != 'postgresql':
        audit_payload['message'] = 'Available only when running on PostgreSQL.'
        return audit_payload

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())
    rows = []

    for spec in iter_required_foreign_key_specs():
        table_name = spec['table_name']
        referred_table = spec['referred_table']
        column_name = spec['column_name']
        if table_name not in existing_tables or referred_table not in existing_tables:
            continue

        existing_columns = postgresql_table_columns(table_name)
        column_info = existing_columns.get(column_name)
        if column_info is None:
            continue

        quoted_table = postgres_quote_identifier(table_name)
        quoted_column = postgres_quote_identifier(column_name)
        quoted_referred_table = postgres_quote_identifier(referred_table)
        quoted_referred_column = postgres_quote_identifier(spec['referred_column'])
        null_row_count = int(
            db.session.execute(
                text(f"SELECT COUNT(*) FROM {quoted_table} WHERE {quoted_column} IS NULL")
            ).scalar()
            or 0
        )
        orphan_row_count = int(
            db.session.execute(
                text(
                    f"SELECT COUNT(*) "
                    f"FROM {quoted_table} AS child "
                    f"WHERE child.{quoted_column} IS NOT NULL "
                    f"AND NOT EXISTS ("
                    f"    SELECT 1 FROM {quoted_referred_table} AS parent "
                    f"    WHERE parent.{quoted_referred_column} = child.{quoted_column}"
                    f")"
                )
            ).scalar()
            or 0
        )
        foreign_keys = inspector.get_foreign_keys(table_name)
        matching_foreign_keys = [
            foreign_key
            for foreign_key in foreign_keys
            if foreign_key.get('constrained_columns') == [column_name]
            and foreign_key.get('referred_table') == referred_table
            and foreign_key.get('referred_columns') == [spec['referred_column']]
        ]
        desired_ondelete = (spec.get('ondelete') or '').upper()
        has_desired_foreign_key = any(
            (((foreign_key.get('options') or {}).get('ondelete') or '').upper() == desired_ondelete)
            for foreign_key in matching_foreign_keys
        )
        is_blocked = null_row_count > 0 or orphan_row_count > 0
        is_already_hardened = (column_info.get('nullable') is False) and has_desired_foreign_key

        if is_blocked:
            status_key = 'blocked'
            status_label = 'Blocked'
            note = 'Legacy rows must be reviewed first. This audit does not change them.'
        elif is_already_hardened:
            status_key = 'already_hardened'
            status_label = 'Already Hardened'
            note = 'This relation already has the required NOT NULL and FK protection.'
        else:
            status_key = 'ready'
            status_label = 'Ready to Harden'
            note = 'No null or orphan rows detected for this relation.'

        rows.append(
            {
                'table_name': table_name,
                'column_name': column_name,
                'referred_table': referred_table,
                'referred_column': spec['referred_column'],
                'nullable': bool(column_info.get('nullable')),
                'matching_foreign_key_count': len(matching_foreign_keys),
                'has_desired_foreign_key': has_desired_foreign_key,
                'desired_ondelete': desired_ondelete or 'NONE',
                'null_row_count': null_row_count,
                'orphan_row_count': orphan_row_count,
                'blocking_row_count': null_row_count + orphan_row_count,
                'status_key': status_key,
                'status_label': status_label,
                'note': note,
            }
        )

    status_sort_order = {'blocked': 0, 'ready': 1, 'already_hardened': 2}
    rows.sort(key=lambda row: (status_sort_order.get(row['status_key'], 99), row['table_name'], row['column_name']))
    blocking_rows = [row for row in rows if row['status_key'] == 'blocked']
    ready_rows = [row for row in rows if row['status_key'] == 'ready']
    already_hardened_rows = [row for row in rows if row['status_key'] == 'already_hardened']

    audit_payload['rows'] = rows
    audit_payload['blocking_rows'] = blocking_rows
    audit_payload['summary'] = {
        'total_relations': len(rows),
        'blocked_relations': len(blocking_rows),
        'blocked_rows': sum(row['blocking_row_count'] for row in blocking_rows),
        'ready_relations': len(ready_rows),
        'already_hardened_relations': len(already_hardened_rows),
    }
    if blocking_rows:
        audit_payload['message'] = 'Blocking rows were found. Strict FK hardening will be skipped for those tables until you review them.'
    else:
        audit_payload['message'] = 'No blocking rows were found for the inspected required foreign keys.'

    return audit_payload


def ensure_postgresql_schema_integrity():
    if schema_work_disabled() or db.engine.dialect.name != 'postgresql':
        return {
            'status_constraints': [],
            'skipped_status_constraints': [],
            'required_foreign_keys': [],
            'skipped_required_foreign_keys': [],
        }

    status_constraints = ensure_postgresql_status_constraints()
    required_foreign_keys = ensure_postgresql_required_foreign_keys()

    return {
        'status_constraints': status_constraints['applied'],
        'skipped_status_constraints': status_constraints['skipped'],
        'required_foreign_keys': required_foreign_keys['applied'],
        'skipped_required_foreign_keys': required_foreign_keys['skipped'],
    }


def audit_postgresql_schema_integrity():
    if schema_work_disabled() or db.engine.dialect.name != 'postgresql':
        return {
            'available': False,
            'status_audit': None,
            'required_foreign_key_audit': None,
        }

    status_audit = collect_postgresql_status_constraint_audit()
    required_foreign_key_audit = collect_postgresql_required_foreign_key_audit()
    blocked_status_columns = status_audit['summary']['blocked_columns']
    blocked_required_relations = required_foreign_key_audit['summary']['blocked_relations']
    if blocked_status_columns or blocked_required_relations:
        print(
            "⚠️ PostgreSQL startup audit only: "
            f"{blocked_status_columns} blocked status column(s), "
            f"{blocked_required_relations} blocked required FK relation(s). "
            "No schema changes were applied."
        )
    else:
        print("✅ PostgreSQL startup audit only: no blocking status or required FK issues detected. No schema changes were applied.")

    return {
        'available': True,
        'status_audit': status_audit,
        'required_foreign_key_audit': required_foreign_key_audit,
    }


def repair_sqlite_rowid_backed_ids():
    """Backfill legacy SQLite tables where the `id` column exists but was stored as NULL."""
    if schema_work_disabled() or db.engine.dialect.name != 'sqlite':
        return {}

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())
    repaired_tables = {}

    for table_name in sorted(db.metadata.tables):
        table = db.metadata.tables[table_name]
        table_name = table.name
        if table_name not in existing_tables or 'id' not in table.c:
            continue

        quoted_table = '"' + table_name.replace('"', '""') + '"'
        null_count = int(db.session.execute(text(f'SELECT COUNT(*) FROM {quoted_table} WHERE id IS NULL')).scalar() or 0)
        if null_count <= 0:
            continue

        conflicting_rowid_count = int(
            db.session.execute(
                text(
                    f'''
                    SELECT COUNT(*)
                    FROM {quoted_table} AS missing_id_rows
                    WHERE missing_id_rows.id IS NULL
                    AND EXISTS (
                        SELECT 1
                        FROM {quoted_table} AS existing_id_rows
                        WHERE existing_id_rows.id = missing_id_rows.rowid
                    )
                    '''
                )
            ).scalar()
            or 0
        )
        if conflicting_rowid_count > 0:
            non_null_count = int(db.session.execute(text(f'SELECT COUNT(*) FROM {quoted_table} WHERE id IS NOT NULL')).scalar() or 0)
            print(
                f"⚠️ Skipped legacy id repair for {table_name}: "
                f"{null_count} NULL ids remain alongside {non_null_count} populated ids, "
                f"with {conflicting_rowid_count} rowid/id collisions."
            )
            continue

        db.session.execute(text(f'UPDATE {quoted_table} SET id = rowid WHERE id IS NULL'))
        repaired_tables[table_name] = null_count

    if repaired_tables:
        db.session.commit()
        repaired_summary = ', '.join(f'{name}={count}' for name, count in sorted(repaired_tables.items()))
        print(f"✅ Repaired legacy SQLite ids from rowid: {repaired_summary}")

    return repaired_tables


def normalize_user_email_key(value):
    if value is None:
        return None
    normalized = str(value).strip().lower()
    return normalized or None


def user_role_priority(role):
    normalized = str(role or '').strip().lower()
    if normalized == 'admin':
        return 2
    if normalized == 'designate':
        return 1
    return 0


def is_non_empty_value(value):
    return value not in (None, '')


def truthy_database_value(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on', 't', 'y'}


def coerce_legacy_int(value):
    if value in (None, ''):
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def sortable_legacy_value(value):
    if value in (None, ''):
        return ''
    return str(value)


def candidate_value(candidate, field_name, default=None):
    if isinstance(candidate, dict):
        return candidate.get(field_name, default)
    return getattr(candidate, field_name, default)


def choose_preferred_user_candidate(candidates, reference_count_by_id=None):
    reference_count_by_id = reference_count_by_id or {}
    candidates = [candidate for candidate in candidates if candidate is not None]
    if not candidates:
        return None

    def sort_key(candidate):
        candidate_id = coerce_legacy_int(candidate_value(candidate, 'id'))
        reference_count = reference_count_by_id.get(candidate_id, 0)
        return (
            int(reference_count or 0),
            1 if truthy_database_value(candidate_value(candidate, 'is_superadmin')) else 0,
            1 if truthy_database_value(candidate_value(candidate, 'can_manage_all_records')) else 0,
            user_role_priority(candidate_value(candidate, 'role')),
            1 if truthy_database_value(candidate_value(candidate, 'is_authorized')) else 0,
            1 if is_non_empty_value(candidate_value(candidate, 'password_hash')) else 0,
            1 if is_non_empty_value(candidate_value(candidate, 'password_changed_at')) else 0,
            sortable_legacy_value(candidate_value(candidate, 'password_changed_at')),
            1 if is_non_empty_value(candidate_value(candidate, 'last_login_at')) else 0,
            sortable_legacy_value(candidate_value(candidate, 'last_login_at')),
            1 if is_non_empty_value(candidate_value(candidate, 'full_name')) else 0,
            1 if is_non_empty_value(candidate_value(candidate, 'department')) else 0,
            1 if is_non_empty_value(candidate_value(candidate, 'phone')) else 0,
            1 if is_non_empty_value(candidate_value(candidate, 'email')) else 0,
            -(candidate_id or 0),
        )

    return max(candidates, key=sort_key)


def find_users_by_email(email):
    normalized = normalize_user_email_key(email)
    if not normalized:
        return []
    return [
        user
        for user in (
            User.query
            .filter(User.email.isnot(None))
            .filter(func.lower(User.email) == normalized)
            .all()
        )
        if user is not None
    ]


def find_preferred_user_by_email(email):
    return choose_preferred_user_candidate(find_users_by_email(email))


def sqlite_user_reference_columns():
    if schema_work_disabled() or db.engine.dialect.name != 'sqlite':
        return []

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())
    reference_columns = []
    for table in db.metadata.tables.values():
        if table.name not in existing_tables:
            continue
        for column in table.c:
            if any(
                foreign_key.column.table.name == 'users_for_form14'
                and foreign_key.column.name == 'id'
                for foreign_key in column.foreign_keys
            ):
                reference_columns.append((table.name, column.name))
    return reference_columns


def build_sqlite_user_reference_count_map(user_ids, reference_columns):
    counts = defaultdict(int)
    normalized_user_ids = [user_id for user_id in {coerce_legacy_int(value) for value in user_ids} if user_id is not None]
    if not normalized_user_ids:
        return counts

    for table_name, column_name in reference_columns:
        quoted_table = '"' + table_name.replace('"', '""') + '"'
        quoted_column = '"' + column_name.replace('"', '""') + '"'
        statement = text(
            f'''
            SELECT {quoted_column} AS user_id, COUNT(*) AS ref_count
            FROM {quoted_table}
            WHERE {quoted_column} IN :user_ids
            GROUP BY {quoted_column}
            '''
        ).bindparams(bindparam('user_ids', expanding=True))
        for row in db.session.execute(statement, {'user_ids': normalized_user_ids}).mappings():
            user_id = coerce_legacy_int(row.get('user_id'))
            if user_id is None:
                continue
            counts[user_id] += int(row.get('ref_count') or 0)

    return counts


def latest_row_for_field(rows, field_name):
    candidates = [
        row for row in rows
        if is_non_empty_value(candidate_value(row, field_name))
    ]
    if not candidates:
        return None
    return max(
        candidates,
        key=lambda row: (
            sortable_legacy_value(candidate_value(row, field_name)),
            -(coerce_legacy_int(candidate_value(row, 'id')) or 0),
        ),
    )


def best_row_with_field(rows, field_name, reference_count_by_id=None):
    candidates = [
        row for row in rows
        if is_non_empty_value(candidate_value(row, field_name))
    ]
    return choose_preferred_user_candidate(candidates, reference_count_by_id=reference_count_by_id)


def merge_duplicate_user_rows(rows, canonical_id, reference_count_by_id=None):
    reference_count_by_id = reference_count_by_id or {}
    canonical_row = choose_preferred_user_candidate(rows, reference_count_by_id=reference_count_by_id)
    if canonical_row is None:
        return {}

    best_email_row = best_row_with_field(rows, 'email', reference_count_by_id=reference_count_by_id)
    best_name_row = best_row_with_field(rows, 'full_name', reference_count_by_id=reference_count_by_id)
    best_phone_row = best_row_with_field(rows, 'phone', reference_count_by_id=reference_count_by_id)
    best_department_row = best_row_with_field(rows, 'department', reference_count_by_id=reference_count_by_id)
    best_authorized_by_row = best_row_with_field(rows, 'authorized_by_id', reference_count_by_id=reference_count_by_id)
    best_report_row = best_row_with_field(rows, 'report_id', reference_count_by_id=reference_count_by_id)
    last_login_row = latest_row_for_field(rows, 'last_login_at')
    last_failed_login_row = latest_row_for_field(rows, 'last_failed_login_at')
    authorized_at_row = latest_row_for_field(rows, 'authorized_at')
    password_row = max(
        [
            row for row in rows
            if is_non_empty_value(candidate_value(row, 'password_hash'))
        ],
        key=lambda row: (
            1 if is_non_empty_value(candidate_value(row, 'password_changed_at')) else 0,
            sortable_legacy_value(candidate_value(row, 'password_changed_at')),
            1 if is_non_empty_value(candidate_value(row, 'last_login_at')) else 0,
            sortable_legacy_value(candidate_value(row, 'last_login_at')),
            -(coerce_legacy_int(candidate_value(row, 'id')) or 0),
        ),
        default=None,
    )

    known_must_change_password_flags = [
        truthy_database_value(candidate_value(row, 'must_change_password'))
        for row in rows
        if candidate_value(row, 'must_change_password') is not None
    ]
    failed_attempt_candidates = [
        coerce_legacy_int(candidate_value(row, 'failed_login_attempts'))
        for row in rows
    ]
    failed_attempt_candidates = [value for value in failed_attempt_candidates if value is not None]
    best_role = max((candidate_value(row, 'role') for row in rows), key=user_role_priority, default='user')

    merged_values = {
        'email': candidate_value(best_email_row or canonical_row, 'email'),
        'full_name': candidate_value(best_name_row or canonical_row, 'full_name'),
        'phone': candidate_value(best_phone_row or canonical_row, 'phone'),
        'department': candidate_value(best_department_row or canonical_row, 'department'),
        'must_change_password': all(known_must_change_password_flags) if known_must_change_password_flags else False,
        'password_hash': candidate_value(password_row or canonical_row, 'password_hash'),
        'password_changed_at': candidate_value(password_row or canonical_row, 'password_changed_at'),
        'role': 'admin' if user_role_priority(best_role) >= user_role_priority('admin') else best_role,
        'is_superadmin': any(truthy_database_value(candidate_value(row, 'is_superadmin')) for row in rows),
        'can_manage_all_records': any(truthy_database_value(candidate_value(row, 'can_manage_all_records')) for row in rows),
        'is_authorized': any(truthy_database_value(candidate_value(row, 'is_authorized')) for row in rows),
        'authorized_at': candidate_value(authorized_at_row or canonical_row, 'authorized_at'),
        'authorized_by_id': coerce_legacy_int(candidate_value(best_authorized_by_row or canonical_row, 'authorized_by_id')),
        'last_login_at': candidate_value(last_login_row or canonical_row, 'last_login_at'),
        'last_login_ip': candidate_value(last_login_row or canonical_row, 'last_login_ip'),
        'failed_login_attempts': max(failed_attempt_candidates) if failed_attempt_candidates else 0,
        'last_failed_login_at': candidate_value(last_failed_login_row or canonical_row, 'last_failed_login_at'),
        'report_id': candidate_value(best_report_row or canonical_row, 'report_id'),
    }

    if merged_values['is_superadmin']:
        merged_values['role'] = 'admin'
        merged_values['can_manage_all_records'] = True
        merged_values['is_authorized'] = True
    if merged_values['authorized_by_id'] == canonical_id:
        merged_values['authorized_by_id'] = None
    return merged_values


def dedupe_sqlite_users_by_email():
    if schema_work_disabled() or db.engine.dialect.name != 'sqlite':
        return {}

    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())
    if 'users_for_form14' not in existing_tables:
        return {}

    user_rows = [
        dict(row)
        for row in db.session.execute(
            text('SELECT rowid AS sqlite_rowid, * FROM users_for_form14 WHERE email IS NOT NULL')
        ).mappings()
    ]
    grouped_rows = defaultdict(list)
    for row in user_rows:
        email_key = normalize_user_email_key(row.get('email'))
        if email_key:
            grouped_rows[email_key].append(row)

    duplicate_groups = {
        email_key: rows
        for email_key, rows in grouped_rows.items()
        if len(rows) > 1
    }
    if not duplicate_groups:
        return {}

    reference_columns = sqlite_user_reference_columns()
    all_duplicate_ids = [
        user_id
        for rows in duplicate_groups.values()
        for user_id in [coerce_legacy_int(row.get('id'))]
        if user_id is not None
    ]
    reference_count_by_id = build_sqlite_user_reference_count_map(all_duplicate_ids, reference_columns)
    user_columns = {
        column['name']
        for column in inspector.get_columns('users_for_form14')
    }
    mergeable_fields = [
        field_name
        for field_name in (
            'email',
            'full_name',
            'phone',
            'department',
            'must_change_password',
            'password_hash',
            'password_changed_at',
            'role',
            'is_superadmin',
            'can_manage_all_records',
            'is_authorized',
            'authorized_at',
            'authorized_by_id',
            'last_login_at',
            'last_login_ip',
            'failed_login_attempts',
            'last_failed_login_at',
            'report_id',
        )
        if field_name in user_columns
    ]

    dedupe_summary = {}
    db.session.execute(text('PRAGMA foreign_keys=OFF'))
    try:
        for email_key, rows in sorted(duplicate_groups.items()):
            canonical_row = choose_preferred_user_candidate(rows, reference_count_by_id=reference_count_by_id)
            canonical_id = coerce_legacy_int(candidate_value(canonical_row, 'id'))
            duplicate_ids = sorted(
                {
                    coerce_legacy_int(row.get('id'))
                    for row in rows
                    if coerce_legacy_int(row.get('id')) is not None and coerce_legacy_int(row.get('id')) != canonical_id
                }
            )
            if canonical_id is None or not duplicate_ids:
                continue

            merged_values = merge_duplicate_user_rows(rows, canonical_id, reference_count_by_id=reference_count_by_id)
            if merged_values.get('authorized_by_id') in duplicate_ids:
                merged_values['authorized_by_id'] = canonical_id

            if mergeable_fields:
                update_clause = ', '.join(f'"{field_name}" = :{field_name}' for field_name in mergeable_fields)
                params = {field_name: merged_values.get(field_name) for field_name in mergeable_fields}
                params['canonical_id'] = canonical_id
                db.session.execute(
                    text(f'UPDATE "users_for_form14" SET {update_clause} WHERE id = :canonical_id'),
                    params,
                )

            for duplicate_id in duplicate_ids:
                for table_name, column_name in reference_columns:
                    quoted_table = '"' + table_name.replace('"', '""') + '"'
                    quoted_column = '"' + column_name.replace('"', '""') + '"'
                    db.session.execute(
                        text(
                            f'UPDATE {quoted_table} '
                            f'SET {quoted_column} = :canonical_id '
                            f'WHERE {quoted_column} = :duplicate_id'
                        ),
                        {
                            'canonical_id': canonical_id,
                            'duplicate_id': duplicate_id,
                        },
                    )
                db.session.execute(
                    text('DELETE FROM "users_for_form14" WHERE id = :duplicate_id'),
                    {'duplicate_id': duplicate_id},
                )

            dedupe_summary[email_key] = {
                'kept_id': canonical_id,
                'removed_ids': duplicate_ids,
            }

        db.session.commit()
    except Exception:
        db.session.rollback()
        raise
    finally:
        db.session.execute(text('PRAGMA foreign_keys=ON'))
        db.session.commit()

    if dedupe_summary:
        removed_total = sum(len(item['removed_ids']) for item in dedupe_summary.values())
        print(
            '✅ Deduplicated legacy SQLite users by email: '
            f'{len(dedupe_summary)} account(s), {removed_total} duplicate row(s) removed.'
        )

    return dedupe_summary


def initialize_database(
    reset=False,
    seed_users=True,
    apply_schema_changes=True,
    run_postgresql_audit=False,
    sync_default_admin=True,
):
    if schema_work_disabled():
        return

    sqlite_mysql_import_result = None
    database_dump_bootstrap_result = None

    if reset and apply_schema_changes:
        db.drop_all()

    if apply_schema_changes:
        database_dump_bootstrap_result = import_database_dump_into_current_database(force=False)
        db.create_all()
        ensure_legacy_user_schema()
        ensure_legacy_report_schema()
        ensure_all_model_string_capacities()
        ensure_postgresql_schema_integrity()
        if database_dump_bootstrap_result and database_dump_bootstrap_result.get('performed'):
            record_database_dump_bootstrap(database_dump_bootstrap_result)
        repair_sqlite_rowid_backed_ids()
        dedupe_sqlite_users_by_email()
        sqlite_mysql_import_result = auto_import_sqlite_into_empty_mysql_database()
    elif run_postgresql_audit and db.engine.dialect.name == 'postgresql':
        audit_postgresql_schema_integrity()

    report_database_state()
    if sync_default_admin:
        ensure_default_admin()

    if seed_users:
        seed_configured_users()

    if apply_schema_changes:
        repair_sqlite_rowid_backed_ids()
        dedupe_sqlite_users_by_email()

    return {
        'database_dump_bootstrap_result': database_dump_bootstrap_result,
        'sqlite_mysql_import_result': sqlite_mysql_import_result,
    }


def startup_database_bootstrap_lock_path():
    storage_root = resolve_storage_root_directory()
    storage_root.mkdir(parents=True, exist_ok=True)
    return storage_root / '.startup_database_bootstrap.lock'


@contextmanager
def startup_database_bootstrap_lock():
    lock_path = startup_database_bootstrap_lock_path()
    with lock_path.open('a+', encoding='utf-8') as lock_handle:
        if fcntl is not None:
            fcntl.flock(lock_handle.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            if fcntl is not None:
                fcntl.flock(lock_handle.fileno(), fcntl.LOCK_UN)


def bootstrap_database_on_startup():
    if running_schema_command() or schema_work_disabled():
        return

    try:
        with startup_database_bootstrap_lock():
            with app.app_context():
                is_postgresql = db.engine.dialect.name == 'postgresql'
                apply_schema_changes = not is_postgresql
                initialize_database(
                    reset=False,
                    seed_users=False,
                    apply_schema_changes=apply_schema_changes,
                    run_postgresql_audit=is_postgresql and not apply_schema_changes,
                    sync_default_admin=False,
                )
    except OperationalError as exc:
        print(f"⚠️ Database bootstrap skipped due to connection error: {exc}")
    except Exception as exc:
        if 'could not translate host name' in str(exc) or 'Could not connect' in str(exc):
            print(f"⚠️ Database bootstrap skipped due to connection error: {exc}")
            return
        raise


def current_database_dialect():
    return db.engine.dialect.name


def uses_legacy_sqlite_identity_for_model(instance):
    table = getattr(instance, '__table__', None)
    if table is None or 'id' not in table.c:
        return False
    if current_database_dialect() != 'sqlite':
        return False
    return not sqlite_table_has_real_id_primary_key(db.session.connection(), table.name)


def ensure_persisted_primary_key(instance, label='record'):
    table = getattr(instance, '__table__', None)
    if table is None or 'id' not in table.c:
        return None

    instance_id = getattr(instance, 'id', None)
    if instance_id in (None, ''):
        raise RuntimeError(f'{label.capitalize()} was flushed without a persistent id.')

    if not uses_legacy_sqlite_identity_for_model(instance):
        return instance_id

    escaped_table_name = table.name.replace('"', '""')
    persisted_row = db.session.execute(
        text(f'SELECT rowid, id FROM "{escaped_table_name}" WHERE id = :record_id'),
        {'record_id': int(instance_id)},
    ).mappings().first()
    if persisted_row is None:
        raise RuntimeError(
            f'{label.capitalize()} flush did not persist a matching row in legacy SQLite mode.'
        )
    return instance_id


def add_and_flush_new_instance(instance, label='record'):
    db.session.add(instance)
    db.session.flush()
    ensure_persisted_primary_key(instance, label=label)
    return instance


def assign_missing_legacy_sqlite_ids(model_class, rows):
    table = getattr(model_class, '__table__', None)
    if table is None or 'id' not in table.c:
        return
    if current_database_dialect() != 'sqlite':
        return
    if sqlite_table_has_real_id_primary_key(db.session.connection(), table.name):
        return

    current_max_id = db.session.execute(select(func.max(table.c.id))).scalar()
    next_id = int(current_max_id or 0) + 1
    used_ids = {
        int(row.id)
        for row in (rows or [])
        if getattr(row, 'id', None) not in (None, '')
    }
    if used_ids:
        next_id = max(next_id, max(used_ids) + 1)

    for row in rows or []:
        if getattr(row, 'id', None) not in (None, ''):
            continue
        while next_id in used_ids:
            next_id += 1
        row.id = next_id
        used_ids.add(next_id)
        next_id += 1


MISSING_DATE_PLACEHOLDER = "99/99/9999"
RETURN_DATE_PLACEHOLDER_DISPLAY = "9/9/9999"


def parse_date(value):
    """Helper function to parse date strings"""
    if not value:
        return None
    normalized = value.strip()
    if normalized == MISSING_DATE_PLACEHOLDER:
        return None
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def normalize_date_raw_value(value):
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def set_reporting_period_fields(report, start_value=None, end_value=None):
    start_raw = normalize_date_raw_value(start_value)
    end_raw = normalize_date_raw_value(end_value)

    parsed_start = parse_date(start_raw)
    parsed_end = parse_date(end_raw)

    # Keep the standard Form 14 behavior: when a valid start date exists,
    # backfill a missing or invalid end date to one year minus one day.
    if parsed_start and (parsed_end is None or parsed_end < parsed_start):
        parsed_end = derive_reporting_period_end(parsed_start)
        end_raw = parsed_end.isoformat() if parsed_end else end_raw

    report.reporting_period_start = parsed_start
    report.reporting_period_end = parsed_end
    report.reporting_period_start_raw = start_raw if start_raw and report.reporting_period_start is None else None
    report.reporting_period_end_raw = end_raw if end_raw and report.reporting_period_end is None else None


def reporting_period_display(date_value, raw_value=None):
    if date_value:
        return format_date(date_value)
    raw = normalize_date_raw_value(raw_value)
    return raw or MISSING_DATE_PLACEHOLDER


def reporting_period_input_value(date_value, raw_value=None):
    raw = normalize_date_raw_value(raw_value)
    if raw:
        return raw
    return date_value.isoformat() if date_value else ""


def parse_return_date(value):
    return parse_date(value) or RETURN_DATE_PLACEHOLDER


def return_date_input_value(value):
    return (value or RETURN_DATE_PLACEHOLDER).isoformat()


def return_date_display(value):
    normalized = value or RETURN_DATE_PLACEHOLDER
    if normalized == RETURN_DATE_PLACEHOLDER:
        return RETURN_DATE_PLACEHOLDER_DISPLAY
    return format_date(normalized)


def report_owner_display(report):
    if report.user:
        return report.user.full_name or report.user.email
    if report.last_modified_by:
        return report.last_modified_by.full_name or report.last_modified_by.email

    uploaded_by_users = [
        uploaded_file.uploaded_by
        for uploaded_file in sorted(
            report.uploaded_files or [],
            key=lambda item: item.created_at or utc_now(),
            reverse=True,
        )
        if getattr(uploaded_file, 'uploaded_by', None)
    ]
    if uploaded_by_users:
        user = uploaded_by_users[0]
        return user.full_name or user.email
    return 'Unassigned'


def user_display_name(user):
    if user is None:
        return None
    return user.full_name or user.email or f'User #{user.id}'


def user_username_display(user):
    if user is None:
        return None
    email = getattr(user, 'email', None)
    login_name = get_login_name(email) if email else None
    if login_name:
        return login_name
    if email:
        return email
    if getattr(user, 'full_name', None):
        return user.full_name
    if getattr(user, 'id', None) is not None:
        return f'User #{user.id}'
    return None


def report_associated_username_display(report):
    if report is None:
        return 'Unassigned'

    candidate_users = []
    if getattr(report, 'user', None):
        candidate_users.append(report.user)
    if getattr(report, 'last_modified_by', None):
        candidate_users.append(report.last_modified_by)
    candidate_users.extend(
        uploaded_file.uploaded_by
        for uploaded_file in sorted(
            report.uploaded_files or [],
            key=lambda item: item.created_at or utc_now(),
            reverse=True,
        )
        if getattr(uploaded_file, 'uploaded_by', None)
    )

    for user in candidate_users:
        username = user_username_display(user)
        if username:
            return username
    return 'Unassigned'


def report_updater_display(report):
    if report is None:
        return 'Unassigned'
    if report.last_modified_by:
        return user_display_name(report.last_modified_by) or 'Unassigned'
    if report.user:
        return user_display_name(report.user) or 'Unassigned'
    return 'Unassigned'


def report_matches_search_query(report, search_query):
    term = (search_query or '').strip().lower()
    if not term:
        return True

    search_values = [
        getattr(report, 'id', None),
        getattr(report, 'pbo_name', None),
        getattr(report, 'pbo_registration_number', None),
        getattr(report, 'form_14', None),
        getattr(report, 'contact_name', None),
        getattr(report, 'contact_email', None),
    ]
    for value in search_values:
        if value is None:
            continue
        if term in str(value).lower():
            return True
    return False


def maybe_assign_report_owner(report, actor):
    if report is None or actor is None or not getattr(actor, 'is_authenticated', False):
        return
    if getattr(actor, 'role', None) == 'admin':
        return
    if report.user_id == actor.id:
        return

    owner = report.user
    if owner is None or getattr(owner, 'role', None) == 'admin':
        report.user_id = actor.id


def format_date(value):
    """Format date as DD/MM/YYYY for display."""
    return value.strftime("%d/%m/%Y") if value else MISSING_DATE_PLACEHOLDER


def to_app_timezone(value):
    if value is None:
        return None
    app_timezone = ZoneInfo(app.config.get('APP_TIMEZONE', 'Africa/Nairobi'))
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(app_timezone)
    return value


def format_datetime(value):
    """Format datetime as DD/MM/YYYY HH:MM for display."""
    localized_value = to_app_timezone(value)
    return localized_value.strftime("%d/%m/%Y %H:%M") if localized_value else ""


def get_app_today():
    timezone = ZoneInfo(app.config.get('APP_TIMEZONE', 'Africa/Nairobi'))
    return utc_now().astimezone(timezone).date()


def find_duplicate_ocr_upload(filename, user, target_date=None):
    normalized_filename = secure_filename(filename or "")
    if not normalized_filename:
        return None

    lookup_date = target_date or get_app_today()
    query = UploadedFile.query.filter(
        UploadedFile.category == 'ocr_source',
        func.upper(UploadedFile.original_filename) == normalized_filename.upper(),
    )
    if not can_manage_all_records(user):
        query = query.filter(UploadedFile.uploaded_by_id == user.id)

    candidates = query.order_by(UploadedFile.created_at.desc(), UploadedFile.id.desc()).all()
    for candidate in candidates:
        candidate_created_at = to_app_timezone(candidate.created_at)
        if candidate_created_at is None or candidate_created_at.date() == lookup_date:
            return candidate
    return None


def get_current_fiscal_year(date_value):
    start_year = date_value.year if date_value.month >= 7 else date_value.year - 1
    return start_year, start_year + 1


def fiscal_start_year(date_value):
    return date_value.year if date_value.month >= 7 else date_value.year - 1


def parse_fiscal_year(value, default_start_year, default_end_year):
    if not value:
        return default_start_year, default_end_year
    parts = value.split('-')
    if len(parts) != 2:
        return default_start_year, default_end_year
    try:
        start_year = int(parts[0])
        end_year = int(parts[1])
    except ValueError:
        return default_start_year, default_end_year
    if end_year < start_year:
        return default_start_year, default_end_year
    return start_year, end_year


def fiscal_year_date_range(requested_fy=None):
    today = get_app_today()
    default_start_year, default_end_year = get_current_fiscal_year(today)
    start_year, end_year = parse_fiscal_year(
        requested_fy,
        default_start_year,
        default_end_year,
    )
    return (
        datetime(start_year, 7, 1).date(),
        datetime(end_year, 6, 30).date(),
        start_year,
        end_year,
    )


def parse_iso_date(value):
    text_value = (value or '').strip()
    if not text_value:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            continue
    return None


def derive_reporting_period_end(start_date):
    if start_date is None:
        return None
    try:
        next_year_same_day = start_date.replace(year=start_date.year + 1)
    except ValueError:
        # Handles Feb 29 -> Feb 28 in non-leap years.
        next_year_same_day = datetime(start_date.year + 1, 3, 1).date()
    return next_year_same_day - timedelta(days=1)


def resolve_reporting_download_window(requested_fy=None, period_start_value=None, period_end_value=None, stream=None):
    requested_stream = (stream or '').strip().lower()
    period_start = parse_iso_date(period_start_value)
    period_end = parse_iso_date(period_end_value)

    if requested_stream == 'year':
        period_start = None
        period_end = None

    if period_start:
        expected_end = derive_reporting_period_end(period_start)
        if period_end is None or period_end < period_start:
            period_end = expected_end
        return (
            period_start,
            period_end,
            {
                'mode': 'period',
                'stream': 'period',
                'start_label': period_start.isoformat(),
                'end_label': period_end.isoformat() if period_end else '',
                'filename_label': f"{period_start.isoformat()}_to_{period_end.isoformat()}" if period_end else period_start.isoformat(),
            },
        )

    range_start, range_end, start_year, end_year = fiscal_year_date_range(requested_fy)
    return (
        range_start,
        range_end,
        {
            'mode': 'fy',
            'stream': 'year',
            'start_label': range_start.isoformat(),
            'end_label': range_end.isoformat(),
            'filename_label': f"{start_year}-{end_year}",
            'start_year': start_year,
            'end_year': end_year,
        },
    )


def reports_download_base_query():
    query = PBOReport.query.order_by(PBOReport.created_at.desc())
    if not can_manage_all_records(current_user):
        query = query.filter(PBOReport.user_id == current_user.id)
    return query


def reports_for_download():
    return reports_download_base_query().all()


def report_matches_fiscal_year(report, range_start, range_end):
    period_start = report.reporting_period_start
    period_end = report.reporting_period_end
    if period_start:
        return range_start <= period_start <= range_end
    anchor_date = period_end
    if anchor_date is None and report.created_at is not None:
        created_value = to_app_timezone(report.created_at)
        anchor_date = created_value.date() if created_value is not None else None
    if anchor_date is None:
        return False
    return range_start <= anchor_date <= range_end


@app.template_filter("format_date")
def jinja_format_date(value):
    return format_date(value)


@app.template_filter("format_datetime")
def jinja_format_datetime(value):
    return format_datetime(value)


def parse_float(value):
    """Helper function to parse float values"""
    if value is None:
        return None
    try:
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            value = re.sub(r"\s+", "", value)
            if "," in value and "." not in value:
                value = value.replace(",", ".")
            else:
                value = value.replace(",", "")
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError, TypeError):
        return None


def float_or_zero(value):
    """Parse a numeric value and return 0.0 when missing/invalid."""
    parsed = parse_float(value)
    return float(parsed) if parsed is not None else 0.0


LEGACY_9999_ZERO_CUTOFF_DATE = datetime(2026, 3, 23).date()


def legacy_9999_output_reference_date(report):
    if report is None:
        return None
    for attr_name in ('created_at', 'submitted_at', 'updated_at'):
        raw_value = getattr(report, attr_name, None)
        if raw_value is None:
            continue
        if hasattr(raw_value, 'date'):
            try:
                return raw_value.date()
            except TypeError:
                pass
        if isinstance(raw_value, datetime):
            return raw_value.date()
    return None


def report_uses_legacy_9999_zero_rule(report):
    reference_date = legacy_9999_output_reference_date(report)
    return bool(reference_date and reference_date < LEGACY_9999_ZERO_CUTOFF_DATE)


def legacy_zero_output_value(report, value, field_name=None):
    if not report_uses_legacy_9999_zero_rule(report):
        return value
    if isinstance(value, bool):
        return value
    if field_name:
        normalized_field = str(field_name).strip().lower().replace(' ', '_').replace('-', '_')
        compact_field = normalized_field.replace('_', '')
        if normalized_field == 'id' or normalized_field.endswith('_id') or compact_field in {'reportid', 'userid'}:
            return value
    if isinstance(value, (int, float, Decimal)):
        if float(value) == 9999.0:
            return 0 if isinstance(value, int) else 0.0
    return value


def legacy_zero_output_structure(report, value, field_name=None):
    if isinstance(value, dict):
        return {
            key: legacy_zero_output_structure(report, nested_value, field_name=key)
            for key, nested_value in value.items()
        }
    if isinstance(value, list):
        return [legacy_zero_output_structure(report, item, field_name=field_name) for item in value]
    if isinstance(value, tuple):
        return tuple(legacy_zero_output_structure(report, item, field_name=field_name) for item in value)
    return legacy_zero_output_value(report, value, field_name=field_name)


def legacy_zero_float(report, value):
    return float_or_zero(legacy_zero_output_value(report, value))


def legacy_zero_int(report, value):
    return parse_int(legacy_zero_output_value(report, value))


def sum_numeric_attr(rows, attr_name):
    """Sum numeric attributes from SQLAlchemy child rows safely."""
    total = 0.0
    for row in rows or []:
        parsed = parse_float(getattr(row, attr_name, None))
        if parsed is not None:
            total += float(parsed)
    return total


def apply_backend_section_b_math(report):
    """
    Compute Section B totals on the server so frontend JS totals are not trusted.
    - income_b2_total = donations + IGA
    - receipts_total = opening cash balance + income_b2_total
    - cash_bank_balance remains user-entered and is never auto-computed here
    """
    donation_total = sum_numeric_attr(getattr(report, 'donations', []), 'amount')
    iga_total = sum_numeric_attr(getattr(report, 'igas', []), 'amount')
    income_total = donation_total + iga_total
    opening_balance = parse_float(getattr(report, 'cash_balance_previous_year', None))
    has_income_rows = bool(getattr(report, 'donations', [])) or bool(getattr(report, 'igas', []))

    if has_income_rows or opening_balance is not None:
        report.income_b2_total = round(income_total, 2)
        report.receipts_total = round(float_or_zero(opening_balance) + income_total, 2)
    else:
        report.income_b2_total = None
        report.receipts_total = None

def build_section_b_preview(
    opening_balance=None,
    donation_amounts=None,
    iga_amounts=None,
    payment_kenya_amounts=None,
    payment_other_amounts=None,
    project_kenya_amounts=None,
    project_other_amounts=None,
):
    donation_values = [float_or_zero(value) for value in (donation_amounts or [])]
    iga_values = [float_or_zero(value) for value in (iga_amounts or [])]
    payment_kenya_values = [float_or_zero(value) for value in (payment_kenya_amounts or [])]
    payment_other_values = [float_or_zero(value) for value in (payment_other_amounts or [])]
    project_kenya_values = [float_or_zero(value) for value in (project_kenya_amounts or [])]
    project_other_values = [float_or_zero(value) for value in (project_other_amounts or [])]

    donor_total = round(sum(donation_values), 2)
    iga_total = round(sum(iga_values), 2)
    payments_kenya_total = round(sum(payment_kenya_values), 2)
    payments_other_total = round(sum(payment_other_values), 2)
    project_kenya_total = round(sum(project_kenya_values), 2)
    project_other_total = round(sum(project_other_values), 2)
    payment_row_totals = [
        round(payment_kenya_values[index] + payment_other_values[index], 2)
        for index in range(max(len(payment_kenya_values), len(payment_other_values)))
    ]
    payments_total = round(sum(payment_row_totals), 2)

    opening_value = parse_float(opening_balance)
    has_income_rows = any(str(value or '').strip() for value in (donation_amounts or [])) or any(
        str(value or '').strip() for value in (iga_amounts or [])
    )

    income_b2_total = None
    receipts_total = None
    if has_income_rows or opening_value is not None:
        income_b2_total = round(donor_total + iga_total, 2)
        receipts_total = round(float_or_zero(opening_value) + income_b2_total, 2)

    return {
        'donor_total': donor_total,
        'iga_total': iga_total,
        'income_b2_total': income_b2_total,
        'receipts_total': receipts_total,
        'payments_kenya_total': payments_kenya_total,
        'payments_other_total': payments_other_total,
        'payments_total': payments_total,
        'payment_row_totals': payment_row_totals,
        'project_kenya_total': project_kenya_total,
        'project_other_total': project_other_total,
    }


def resolve_project_spending_amount(project_row, report=None):
    """Prefer B6 spending_per_county; fallback to Kenya+Other spend when B6 is null."""
    if project_row is None:
        return 0.0
    spending_value = parse_float(
        legacy_zero_output_value(report, getattr(project_row, "spending_per_county", None))
    )
    if spending_value is not None:
        return float(spending_value)
    kenya_amount = parse_float(
        legacy_zero_output_value(report, getattr(project_row, "amount_spent_kenya", None))
    ) or 0.0
    other_amount = parse_float(
        legacy_zero_output_value(report, getattr(project_row, "amount_spent_other", None))
    ) or 0.0
    return float(kenya_amount) + float(other_amount)


def parse_int(value):
    """Helper function to parse integer values"""
    if not value:
        return 0
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return int(value)
    except (ValueError, TypeError):
        return 0


def parse_optional_int(value):
    """Parse integer values for optional fields, returning None when blank/invalid."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace(",", "").strip()
    if value == "":
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


SECTION_C_SUMMARY_CATEGORY_PREFIX = "SECTION_C SUMMARY | "
SECTION_C_DETAIL_CATEGORIES = (
    "BELOW35",
    "ABOVE35",
    "MALE",
    "FEMALE",
    "PWD",
)
SECTION_C_BIODATA_CATEGORY_ALIASES = {
    "BELOW35": "BELOW35",
    "ABOVE35": "ABOVE35",
    "MALE": "MALE",
    "FEMALE": "FEMALE",
    "PWD": "PWD",
    "VOLBELOW35": "BELOW35",
    "VOLABOVE35": "ABOVE35",
}
SECTION_C_STAFF_SUMMARY_DEFINITIONS = (
    (
        f"{SECTION_C_SUMMARY_CATEGORY_PREFIX}STAFF STATIONED IN KENYA | KENYAN | PREVIOUS/CURRENT",
        "staff_kenyan_prev",
        "staff_kenyan_current",
    ),
    (
        f"{SECTION_C_SUMMARY_CATEGORY_PREFIX}STAFF STATIONED IN KENYA | FOREIGN | PREVIOUS/CURRENT",
        "staff_foreign_prev",
        "staff_foreign_current",
    ),
    (
        f"{SECTION_C_SUMMARY_CATEGORY_PREFIX}STAFF STATIONED IN KENYA | KENYAN | CAME IN/LEFT",
        "staff_kenyan_came_in",
        "staff_kenyan_left",
    ),
    (
        f"{SECTION_C_SUMMARY_CATEGORY_PREFIX}STAFF STATIONED IN KENYA | FOREIGN | CAME IN/LEFT",
        "staff_foreign_came_in",
        "staff_foreign_left",
    ),
    (
        f"{SECTION_C_SUMMARY_CATEGORY_PREFIX}STAFF STATIONED IN OTHER COUNTRIES | KENYAN | PREVIOUS/CURRENT",
        "staff_other_kenyan_prev",
        "staff_other_kenyan_current",
    ),
    (
        f"{SECTION_C_SUMMARY_CATEGORY_PREFIX}STAFF STATIONED IN OTHER COUNTRIES | FOREIGN | PREVIOUS/CURRENT",
        "staff_other_foreign_prev",
        "staff_other_foreign_current",
    ),
)
SECTION_C_VOLUNTEER_SUMMARY_DEFINITIONS = (
    (
        f"{SECTION_C_SUMMARY_CATEGORY_PREFIX}VOLUNTEERS / INTERNS | KENYAN | PREVIOUS/CURRENT",
        "volunteers_kenyan_prev",
        "volunteers_kenyan_current",
    ),
    (
        f"{SECTION_C_SUMMARY_CATEGORY_PREFIX}VOLUNTEERS / INTERNS | FOREIGN | PREVIOUS/CURRENT",
        "volunteers_foreign_prev",
        "volunteers_foreign_current",
    ),
)


def is_blank_form_value(value):
    return value is None or (isinstance(value, str) and value.strip() == "")


def to_upper(value):
    """Helper function to convert text to uppercase"""
    if isinstance(value, str):
        normalized = value.strip()
        return normalized.upper() if normalized else None
    return value


def normalize_pbo_name_value(value):
    normalized = to_upper(value)
    if not normalized:
        return None
    normalized = re.sub(r"[^A-Z0-9]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized or None


KENYAN_PHONE_LOCAL_PREFIX = "07"
KENYAN_PHONE_COUNTRY_CODE = "+254"
KENYAN_PHONE_LOCAL_PREFIXES = ("07", "01")


def normalize_phone(value, country_code=None):
    """Normalize phone values while enforcing Kenyan length rules."""
    if value is None:
        return None

    raw = value.strip()
    if not raw:
        return None

    normalized_country = to_upper(country_code)
    digits_only = re.sub(r"\D", "", raw)
    is_kenyan = (
        normalized_country == KENYAN_PHONE_COUNTRY_CODE
        or raw.startswith(KENYAN_PHONE_COUNTRY_CODE)
        or digits_only.startswith("254")
        or any(digits_only.startswith(prefix) for prefix in KENYAN_PHONE_LOCAL_PREFIXES)
    )

    if not is_kenyan:
        return to_upper(raw)

    if normalized_country == KENYAN_PHONE_COUNTRY_CODE or raw.startswith(KENYAN_PHONE_COUNTRY_CODE) or digits_only.startswith("254"):
        local_digits = digits_only
        if local_digits.startswith("254"):
            local_digits = local_digits[3:]
        elif local_digits.startswith("0"):
            local_digits = local_digits[1:]
        if local_digits.startswith(("7", "1")):
            local_digits = local_digits[:9]
            return f"{KENYAN_PHONE_COUNTRY_CODE}{local_digits}" if local_digits else KENYAN_PHONE_COUNTRY_CODE
        return to_upper(raw)

    if any(digits_only.startswith(prefix) for prefix in KENYAN_PHONE_LOCAL_PREFIXES):
        return digits_only[:10]

    return to_upper(raw)


def normalize_login_email(value):
    if not value:
        return None
    return value.strip().lower() or None


def get_login_name(email):
    normalized = normalize_login_email(email)
    if not normalized:
        return None
    return normalized.split("@", 1)[0]


def build_login_email(username, email):
    normalized_email = normalize_login_email(email)
    normalized_username = normalize_login_email(username)

    if not normalized_email:
        return None, "Username or email is required."

    if "@" not in normalized_email:
        if not re.fullmatch(r"[a-z0-9._%+-]+", normalized_email):
            return None, "Username can only contain letters, numbers, dots, underscores, percent, plus, and hyphen."
        return normalized_email, None

    if normalized_username:
        if "@" in normalized_username:
            return None, "Username should not include @ or a domain."
        if not re.fullmatch(r"[a-z0-9._%+-]+", normalized_username):
            return None, "Username can only contain letters, numbers, dots, underscores, percent, plus, and hyphen."
        _, domain = normalized_email.split("@", 1)
        return f"{normalized_username}@{domain}", None

    return normalized_email, None


def resolve_user_login(identifier):
    normalized = normalize_login_email(identifier)
    if not normalized:
        return None, None

    exact_user = find_preferred_user_by_email(normalized)
    if exact_user is not None and getattr(exact_user, 'email', None):
        return exact_user, None

    username_matches_by_email = defaultdict(list)
    for user in User.query.filter(User.email.isnot(None)).all():
        if user is None or not getattr(user, 'email', None):
            continue
        if get_login_name(user.email) == normalized:
            username_matches_by_email[normalize_user_email_key(user.email) or f'user:{user.id}'].append(user)

    username_matches = [
        choose_preferred_user_candidate(matches)
        for matches in username_matches_by_email.values()
        if matches
    ]
    if len(username_matches) == 1:
        return username_matches[0], None
    if len(username_matches) > 1:
        return None, 'ambiguous'
    return None, None


def normalize_department(value):
    if not value:
        return None
    normalized = re.sub(r"\s+", " ", value.strip())
    return normalized.upper() if normalized else None


def validate_password_policy(password):
    return [
        (len(password) >= 8, "Password must be at least 8 characters."),
        (re.search(r'[A-Z]', password), "Password must contain an uppercase letter."),
        (re.search(r'[a-z]', password), "Password must contain a lowercase letter."),
        (re.search(r'\d', password), "Password must contain a digit."),
        (re.search(r'[!@#$%^&*(),.?\":{}|<>_\-+=;\'\[\]\\/~`]', password), "Password must contain a special character."),
    ]


def generate_admin_password(length=14):
    alphabet = string.ascii_letters + string.digits
    core = ''.join(secrets.choice(alphabet) for _ in range(max(8, length - 2)))
    return f'{core}A1'


def get_form_upper(key):
    """Get form value and convert to uppercase"""
    value = request.form.get(key)
    return to_upper(value)


def get_form_list_upper(key):
    """Get form list values and convert each to uppercase"""
    values = request.form.getlist(key)
    return [to_upper(v) for v in values]


def get_form_list_upper_any(*keys):
    """Get form list values from the first populated key and convert each to uppercase."""
    for key in keys:
        values = request.form.getlist(key)
        normalized = [to_upper(v) for v in values if to_upper(v)]
        if normalized:
            return normalized
    return []


def get_form_list_any(*keys):
    """Get form list values from the first key that is present in the payload."""
    for key in keys:
        values = request.form.getlist(key)
        if values:
            return values
    return []


def merge_form_list_upper(preferred_key, fallback_key):
    """Merge two repeated form fields row-by-row, preferring the first key when populated."""
    preferred_values = request.form.getlist(preferred_key)
    fallback_values = request.form.getlist(fallback_key)
    row_count = max(len(preferred_values), len(fallback_values))
    resolved = []

    for index in range(row_count):
        preferred_value = to_upper(preferred_values[index]) if index < len(preferred_values) else None
        fallback_value = to_upper(fallback_values[index]) if index < len(fallback_values) else None
        resolved.append(preferred_value or fallback_value)

    return resolved


def get_form_phone(key, country_code_key=None):
    """Get and normalize a phone value from the form."""
    values = request.form.getlist(key)
    value = next(
        (item for item in reversed(values) if isinstance(item, str) and item.strip()),
        None,
    )
    country_code = request.form.get(country_code_key) if country_code_key else None
    if not country_code:
        country_code = next(
            (item for item in values if isinstance(item, str) and item.strip().startswith("+")),
            None,
        )
    return normalize_phone(value, country_code)


def get_form_phone_joined(key, country_code_key=None):
    """Get one or more phone values from the form and store them as comma-separated text."""
    values = request.form.getlist(key)
    if not key.endswith("[]"):
        values.extend(request.form.getlist(f"{key}[]"))
    custom_values = request.form.getlist(f"{key}_custom")
    custom_values.extend(request.form.getlist(f"{key}_custom[]"))

    country_code = request.form.get(country_code_key) if country_code_key else None
    normalized_values = []
    seen = set()

    for raw_value in values:
        normalized = normalize_phone(raw_value, country_code)
        if not normalized:
            continue
        if normalized in seen:
            continue
        seen.add(normalized)
        normalized_values.append(normalized)

    for raw_value in custom_values:
        normalized = to_upper(raw_value)
        if not normalized:
            continue
        if normalized in seen:
            continue
        seen.add(normalized)
        normalized_values.append(normalized)

    return ", ".join(normalized_values) if normalized_values else None


def get_form_text_joined_upper(key):
    """Get one or more text values from the form and store them as uppercase comma-separated text."""
    values = request.form.getlist(key)
    if not key.endswith("[]"):
        values.extend(request.form.getlist(f"{key}[]"))

    normalized_values = []
    seen = set()

    for raw_value in values:
        normalized = to_upper(raw_value)
        if not normalized:
            continue
        if normalized in seen:
            continue
        seen.add(normalized)
        normalized_values.append(normalized)

    return ", ".join(normalized_values) if normalized_values else None


def get_form_list_phone(key):
    """Get phone values from the form and normalize them."""
    values = request.form.getlist(key)
    return [normalize_phone(v) for v in values]


def get_form_list_item(key, index):
    """Return a list item by index from the submitted form values."""
    values = request.form.getlist(key)
    return values[index] if index < len(values) else None


def is_section_c_summary_category(category):
    normalized = to_upper(category)
    return bool(normalized and normalized.startswith(SECTION_C_SUMMARY_CATEGORY_PREFIX))


def normalize_section_c_biodata_category(category):
    normalized = to_upper(category)
    if not normalized:
        return None
    return SECTION_C_BIODATA_CATEGORY_ALIASES.get(normalized, normalized)


def report_phone_number_tokens(report):
    tokens = set()
    if report is None:
        return tokens

    for attr_name in ("telephone", "cell_phone", "contact_telephone"):
        raw_value = getattr(report, attr_name, None)
        if not raw_value:
            continue
        for part in str(raw_value).split(","):
            digits = re.sub(r"\D", "", part or "")
            if not digits:
                continue
            candidate_values = {digits}
            if digits.startswith("254") and len(digits) > 3:
                candidate_values.add(digits[3:])
            if digits.startswith("0") and len(digits) > 1:
                candidate_values.add(digits[1:])
            for candidate in candidate_values:
                try:
                    tokens.add(int(candidate))
                except (TypeError, ValueError):
                    continue
    return tokens


def clean_biodata_count_value(report, value):
    normalized = legacy_zero_output_value(report, value)
    return parse_optional_int(normalized)


def is_phone_like_biodata_pair(report, prev_value, curr_value):
    tokens = report_phone_number_tokens(report)
    if not tokens:
        return False
    prev_candidate = parse_optional_int(prev_value)
    curr_candidate = parse_optional_int(curr_value)
    return (
        prev_candidate is not None
        and curr_candidate is not None
        and prev_candidate in tokens
        and curr_candidate in tokens
    )


def is_legacy_placeholder_biodata_pair(report, prev_value, curr_value):
    if not report_uses_legacy_9999_zero_rule(report):
        return False
    prev_candidate = parse_optional_int(prev_value)
    curr_candidate = parse_optional_int(curr_value)
    return prev_candidate == 9999 and curr_candidate == 9999


def report_has_positive_staff_summary(report):
    for _, prev_attr, curr_attr in SECTION_C_STAFF_SUMMARY_DEFINITIONS:
        if (clean_biodata_count_value(report, getattr(report, prev_attr, None)) or 0) > 0:
            return True
        if (clean_biodata_count_value(report, getattr(report, curr_attr, None)) or 0) > 0:
            return True
    return False


def report_has_positive_volunteer_summary(report):
    for _, prev_attr, curr_attr in SECTION_C_VOLUNTEER_SUMMARY_DEFINITIONS:
        if (clean_biodata_count_value(report, getattr(report, prev_attr, None)) or 0) > 0:
            return True
        if (clean_biodata_count_value(report, getattr(report, curr_attr, None)) or 0) > 0:
            return True
    return False


def extract_clean_biodata_rows(report, rows, *, summary_has_positive_counts=False):
    cleaned_rows = []
    for row in rows or []:
        category = normalize_section_c_biodata_category(getattr(row, "category", None))
        if not category or is_section_c_summary_category(category):
            continue

        raw_prev = getattr(row, "prev_year", None)
        raw_curr = getattr(row, "curr_year", None)
        if is_legacy_placeholder_biodata_pair(report, raw_prev, raw_curr):
            continue
        if is_phone_like_biodata_pair(report, raw_prev, raw_curr):
            continue

        prev_year = clean_biodata_count_value(report, raw_prev)
        curr_year = clean_biodata_count_value(report, raw_curr)
        if prev_year is None and curr_year is None:
            continue

        cleaned_rows.append({
            "category": category,
            "prev_year": prev_year,
            "curr_year": curr_year,
        })

    if (
        cleaned_rows
        and summary_has_positive_counts
        and not any((row["prev_year"] or 0) > 0 or (row["curr_year"] or 0) > 0 for row in cleaned_rows)
    ):
        return []

    return cleaned_rows


def build_standardized_section_c_detail_rows(report, rows, model_class, *, summary_has_positive_counts=False):
    category_buckets = {}
    materialize_standard_rows = summary_has_positive_counts

    for row in rows or []:
        category = normalize_section_c_biodata_category(getattr(row, "category", None))
        if not category or is_section_c_summary_category(category):
            continue

        raw_prev = getattr(row, "prev_year", None)
        raw_curr = getattr(row, "curr_year", None)
        if is_legacy_placeholder_biodata_pair(report, raw_prev, raw_curr):
            continue
        if is_phone_like_biodata_pair(report, raw_prev, raw_curr):
            continue

        prev_year = clean_biodata_count_value(report, raw_prev)
        curr_year = clean_biodata_count_value(report, raw_curr)
        if prev_year is None and curr_year is None:
            continue

        materialize_standard_rows = True
        entry = category_buckets.get(category)
        if entry is None:
            category_buckets[category] = {
                "row": row,
                "prev_year": prev_year,
                "curr_year": curr_year,
            }
            continue

        if prev_year is not None:
            entry["prev_year"] = (entry["prev_year"] or 0) + prev_year
        if curr_year is not None:
            entry["curr_year"] = (entry["curr_year"] or 0) + curr_year

    standard_rows = {}
    extra_rows = []
    for category, entry in category_buckets.items():
        row = entry["row"] or model_class()
        row.category = category
        row.prev_year = entry["prev_year"]
        row.curr_year = entry["curr_year"]
        if category in SECTION_C_DETAIL_CATEGORIES:
            standard_rows[category] = row
        else:
            extra_rows.append(row)

    result_rows = []
    if materialize_standard_rows:
        for category in SECTION_C_DETAIL_CATEGORIES:
            row = standard_rows.get(category) or model_class(category=category)
            row.category = category
            row.prev_year = row.prev_year if row.prev_year is not None else 0
            row.curr_year = row.curr_year if row.curr_year is not None else 0
            result_rows.append(row)
    else:
        for category in SECTION_C_DETAIL_CATEGORIES:
            row = standard_rows.get(category)
            if row is not None:
                result_rows.append(row)

    result_rows.extend(extra_rows)
    return result_rows


def get_display_staff_biodata_rows(report):
    return extract_clean_biodata_rows(
        report,
        report.staff_biodata,
        summary_has_positive_counts=report_has_positive_staff_summary(report),
    )


def get_display_volunteer_biodata_rows(report):
    return extract_clean_biodata_rows(
        report,
        report.volunteer_biodata,
        summary_has_positive_counts=report_has_positive_volunteer_summary(report),
    )


def build_section_c_summary_rows(report, model_class, definitions):
    rows = []
    for category, prev_attr, curr_attr in definitions:
        prev_year = clean_biodata_count_value(report, getattr(report, prev_attr, None))
        curr_year = clean_biodata_count_value(report, getattr(report, curr_attr, None))
        if prev_year is None and curr_year is None:
            continue
        rows.append(model_class(category=category, prev_year=prev_year, curr_year=curr_year))
    return rows


def sync_report_section_c_auxiliary_tables(report):
    cleaned_staff_rows = build_standardized_section_c_detail_rows(
        report,
        report.staff_biodata,
        StaffBiodata,
        summary_has_positive_counts=report_has_positive_staff_summary(report),
    )
    cleaned_volunteer_rows = build_standardized_section_c_detail_rows(
        report,
        report.volunteer_biodata,
        VolunteerBiodata,
        summary_has_positive_counts=report_has_positive_volunteer_summary(report),
    )

    staff_rows = cleaned_staff_rows + build_section_c_summary_rows(
        report,
        StaffBiodata,
        SECTION_C_STAFF_SUMMARY_DEFINITIONS,
    )
    volunteer_rows = cleaned_volunteer_rows + build_section_c_summary_rows(
        report,
        VolunteerBiodata,
        SECTION_C_VOLUNTEER_SUMMARY_DEFINITIONS,
    )
    assign_missing_legacy_sqlite_ids(StaffBiodata, staff_rows)
    assign_missing_legacy_sqlite_ids(VolunteerBiodata, volunteer_rows)

    report.staff_biodata[:] = staff_rows
    report.volunteer_biodata[:] = volunteer_rows


def repair_section_c_biodata_records(report_ids=None):
    query = PBOReport.query.options(
        selectinload(PBOReport.staff_biodata),
        selectinload(PBOReport.volunteer_biodata),
    )
    if report_ids:
        query = query.filter(PBOReport.id.in_(list(report_ids)))

    repaired_reports = 0
    for report in query.all():
        before_staff = [
            (getattr(row, "category", None), getattr(row, "prev_year", None), getattr(row, "curr_year", None))
            for row in (report.staff_biodata or [])
        ]
        before_volunteer = [
            (getattr(row, "category", None), getattr(row, "prev_year", None), getattr(row, "curr_year", None))
            for row in (report.volunteer_biodata or [])
        ]

        sync_report_section_c_auxiliary_tables(report)

        after_staff = [
            (getattr(row, "category", None), getattr(row, "prev_year", None), getattr(row, "curr_year", None))
            for row in (report.staff_biodata or [])
        ]
        after_volunteer = [
            (getattr(row, "category", None), getattr(row, "prev_year", None), getattr(row, "curr_year", None))
            for row in (report.volunteer_biodata or [])
        ]

        if before_staff != after_staff or before_volunteer != after_volunteer:
            repaired_reports += 1

    if repaired_reports:
        db.session.commit()

    return repaired_reports


def get_form_list_with_other(key, other_key, trigger_value="OTHER"):
    """Resolve list values that may use a companion text field for 'Other'."""
    values = request.form.getlist(key)
    other_values = request.form.getlist(other_key)
    row_count = max(len(values), len(other_values))
    resolved = []

    for index in range(row_count):
        value = to_upper(values[index]) if index < len(values) else None
        other_value = to_upper(other_values[index]) if index < len(other_values) else None

        if value == trigger_value or (not value and other_value):
            resolved.append(other_value)
        else:
            resolved.append(value)

    return resolved


def get_form_list_with_other_raw(key, other_key, trigger_value="OTHER"):
    """Resolve list values for 'Other' without forcing uppercase casing."""
    values = request.form.getlist(key)
    other_values = request.form.getlist(other_key)
    row_count = max(len(values), len(other_values))
    resolved = []
    trigger_upper = str(trigger_value or "").upper()

    for index in range(row_count):
        value = (values[index] if index < len(values) else "") or ""
        other_value = (other_values[index] if index < len(other_values) else "") or ""
        value = value.strip()
        other_value = other_value.strip()

        if value.upper() == trigger_upper or (not value and other_value):
            resolved.append(other_value or None)
        else:
            resolved.append(value or None)

    return resolved


def get_checkbox_value(form_key):
    """Helper to get checkbox value - returns True if checked"""
    return form_key in request.form


def get_inactive_section_flag(*form_keys):
    """Return True when any supported hidden section flag marks the section inactive."""
    for key in form_keys:
        value = request.form.get(key)
        if value is None:
            continue
        if str(value).strip() == "1":
            return True
    return False


def is_integer_input(value):
    if value is None:
        return True
    raw = str(value).strip()
    if raw == "":
        return True
    if "." in raw:
        return False
    normalized = re.sub(r"[,\s]", "", raw)
    return normalized.isdigit()


def validate_integer_fields():
    return True


def normalize_scope(value):
    value = to_upper(value)
    return value if value in {"NATIONAL", "INTERNATIONAL"} else None


def make_excel_sheet_name(name, used_names=None, fallback='Sheet'):
    raw_name = (name or '').strip()
    sanitized = re.sub(r'[\[\]\:\*\?\/\\]', '_', raw_name)
    sanitized = re.sub(r'[\x00-\x1f]', '', sanitized).strip().strip("'")
    sanitized = sanitized or fallback

    base_name = sanitized[:31] or fallback
    candidate = base_name
    suffix = 2

    if used_names is None:
        return candidate

    while candidate in used_names:
        suffix_text = f"_{suffix}"
        trimmed_base = base_name[: max(1, 31 - len(suffix_text))]
        candidate = f"{trimmed_base}{suffix_text}"
        suffix += 1

    used_names.add(candidate)
    return candidate


def display_scope(value):
    return value.title() if isinstance(value, str) and value else "N/A"


def clear_disabled_officials_section(report):
    """Clear officials/signature fields when the section is intentionally inactive."""
    report.officials.clear()
    report.submitter_fullname = None
    report.signature = None
    report.submission_date = None


def reporting_period_start_month(value):
    if value is None:
        return None
    return getattr(value, 'month', None)


def find_duplicate_pbo_report(
    pbo_name,
    reporting_period_start=None,
    exclude_report_id=None,
    require_start_month_match=False,
):
    normalized_name = normalize_pbo_name_value(pbo_name)
    if not normalized_name:
        return None

    query = PBOReport.query.filter(
        or_(
            PBOReport.pbo_name_normalized == normalized_name,
            PBOReport.pbo_name == to_upper(pbo_name),
        )
    )
    if exclude_report_id is not None:
        query = query.filter(PBOReport.id != exclude_report_id)

    if not require_start_month_match:
        return query.order_by(PBOReport.id.asc()).first()

    candidate_reports = query.order_by(PBOReport.id.asc()).all()
    if not candidate_reports:
        return None

    start_month = reporting_period_start_month(reporting_period_start)
    if start_month is None:
        return None

    for candidate in candidate_reports:
        if reporting_period_start_month(candidate.reporting_period_start) == start_month:
            return candidate
    return None


def assign_pbo_name_normalized(report):
    normalized_name = normalize_pbo_name_value(report.pbo_name)
    if not normalized_name:
        report.pbo_name_normalized = None
        return None

    query = PBOReport.query.filter(PBOReport.pbo_name_normalized == normalized_name)
    if report.id is not None:
        query = query.filter(PBOReport.id != report.id)
    existing_report = query.order_by(PBOReport.id.asc()).first()

    if existing_report is not None:
        report.pbo_name_normalized = None
        return normalized_name

    report.pbo_name_normalized = normalized_name
    return normalized_name


def issue_submission_token():
    return uuid.uuid4().hex.upper()


def get_or_create_submission_token():
    return issue_submission_token()


def reset_submission_token():
    return issue_submission_token()


def form_submission_matches_existing_report(existing_report):
    if existing_report is None:
        return False

    submitted_name = get_form_upper('pbo_name')
    submitted_start_raw = normalize_date_raw_value(request.form.get('reporting_period_start'))
    submitted_end_raw = normalize_date_raw_value(request.form.get('reporting_period_end'))

    def _report_date_matches(submitted_raw, stored_date, stored_raw):
        normalized_stored_raw = normalize_date_raw_value(stored_raw)
        if not submitted_raw:
            return stored_date is None and normalized_stored_raw is None

        parsed_submitted = parse_date(submitted_raw)
        if parsed_submitted is not None and stored_date is not None:
            return parsed_submitted == stored_date

        return submitted_raw == normalized_stored_raw

    return (
        submitted_name == to_upper(existing_report.pbo_name)
        and _report_date_matches(
            submitted_start_raw,
            existing_report.reporting_period_start,
            existing_report.reporting_period_start_raw,
        )
        and _report_date_matches(
            submitted_end_raw,
            existing_report.reporting_period_end,
            existing_report.reporting_period_end_raw,
        )
    )


def summarize_stale_submission_token(existing_report, submitted_token):
    submitted_name = get_form_upper('pbo_name') or 'N/A'
    submitted_start_raw = normalize_date_raw_value(request.form.get('reporting_period_start')) or 'N/A'
    submitted_end_raw = normalize_date_raw_value(request.form.get('reporting_period_end')) or 'N/A'
    existing_name = to_upper(existing_report.pbo_name) or 'N/A'
    existing_start = (
        normalize_date_raw_value(existing_report.reporting_period_start_raw)
        or (
            existing_report.reporting_period_start.isoformat()
            if existing_report.reporting_period_start else 'N/A'
        )
    )
    existing_end = (
        normalize_date_raw_value(existing_report.reporting_period_end_raw)
        or (
            existing_report.reporting_period_end.isoformat()
            if existing_report.reporting_period_end else 'N/A'
        )
    )
    token_preview = (submitted_token or '').strip().upper()[:12] or 'N/A'
    return (
        f'Stale Form 14 token detected: token={token_preview}, existing_report_id={existing_report.id}, '
        f'existing_identity={existing_name} [{existing_start} to {existing_end}], '
        f'submitted_identity={submitted_name} [{submitted_start_raw} to {submitted_end_raw}]'
    )


def render_form14_response(submission_token=None, prefill=None):
    response = make_response(
        render_template(
            'form14.html',
            submission_token=submission_token or get_or_create_submission_token(),
            prefill=prefill or empty_form14_prefill(),
        )
    )
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['X-Form14-No-Cache'] = '1'
    return response


def get_request_ip():
    forwarded_for = request.headers.get('X-Forwarded-For', '')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return request.remote_addr


def summarize_report(report):
    if not report:
        return None
    name = report.pbo_name or "UNNAMED PBO"
    period_end = format_date(report.reporting_period_end) if report.reporting_period_end else "NO PERIOD END"
    return f"{name} | Period End: {period_end} | Report #{report.id or 'NEW'}"


def parse_datetime_local(value):
    if not value:
        return None
    value = value.strip()
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def normalize_report_status(value, default='draft'):
    normalized = (value or '').strip().lower()
    return normalized if normalized in REPORT_STATUS_VALUES else default


def normalize_review_status(value, default='pending'):
    normalized = (value or '').strip().lower()
    return normalized if normalized in REVIEW_STATUS_VALUES else default


def normalize_choice_status(value, allowed_values, default):
    normalized = (value or '').strip().lower()
    return normalized if normalized in allowed_values else default


def get_admin_setting(key, default=None):
    setting = AdminSetting.query.filter_by(key=key).first()
    return setting.value if setting and setting.value not in (None, '') else default


def set_admin_setting(key, value, user=None):
    setting = AdminSetting.query.filter_by(key=key).first()
    if setting is None:
        setting = AdminSetting(key=key)
        db.session.add(setting)
    setting.value = value
    setting.updated_by_id = user.id if user else None
    setting.updated_at = utc_now()
    return setting


def get_admin_setting_json(key, default=None):
    raw_value = get_admin_setting(key, None)
    if raw_value in (None, ''):
        return default
    try:
        return json.loads(raw_value)
    except (TypeError, ValueError):
        return default


def set_admin_setting_json(key, value, user=None):
    serialized = json.dumps(value, ensure_ascii=False)
    return set_admin_setting(key, serialized, user=user)


def data_interp_status_setting_key(user_id):
    return f'data_interpretation_job_status_user_{int(user_id)}'


def data_interp_result_setting_key(user_id):
    return f'data_interpretation_job_result_user_{int(user_id)}'


def data_interp_scope_label_for_user(user):
    return 'Analysis scoped to the full sector-report database'


def data_interp_default_job_snapshot(selected_fiscal_year=None):
    return {
        'running': False,
        'status': 'idle',
        'detail': 'No background analysis has run yet.',
        'selected_fiscal_year': selected_fiscal_year,
        'scope_label': None,
        'started_at': None,
        'finished_at': None,
        'progress_percent': 0,
        'last_error': None,
        'analysis_error': None,
        'result_ready': False,
        'analysis_focus_label': None,
    }


def data_interp_present_job_snapshot(snapshot):
    base_snapshot = data_interp_default_job_snapshot()
    if isinstance(snapshot, dict):
        base_snapshot.update(snapshot)

    status = (base_snapshot.get('status') or '').strip().lower()
    if base_snapshot.get('running'):
        status = status or 'running'
    elif not status:
        status = 'idle'

    if status in {'queued', 'running'}:
        status_label = 'Running'
        status_class = 'is-neutral'
    elif status == 'completed':
        if base_snapshot.get('analysis_error'):
            status_label = 'Completed With Notes'
            status_class = 'is-warn'
        else:
            status_label = 'Completed'
            status_class = 'is-good'
    elif status == 'failed':
        status_label = 'Failed'
        status_class = 'is-bad'
    else:
        status_label = 'Idle'
        status_class = 'is-neutral'

    detail = (
        base_snapshot.get('detail')
        or base_snapshot.get('analysis_error')
        or base_snapshot.get('last_error')
        or 'No background analysis has run yet.'
    )

    base_snapshot['status'] = status
    base_snapshot['status_label'] = status_label
    base_snapshot['status_class'] = status_class
    base_snapshot['detail'] = detail
    return base_snapshot


def get_data_interpretation_job_snapshot_for_user(user_id):
    if not user_id:
        return data_interp_present_job_snapshot({})
    stored = get_admin_setting_json(data_interp_status_setting_key(user_id), {}) or {}
    if not isinstance(stored, dict):
        stored = {}
    return data_interp_present_job_snapshot(stored)


def store_data_interpretation_job_snapshot_for_user(user_id, payload, user=None):
    if not user_id:
        return data_interp_present_job_snapshot(payload or {})
    current_payload = get_admin_setting_json(data_interp_status_setting_key(user_id), {}) or {}
    if not isinstance(current_payload, dict):
        current_payload = {}
    current_payload.update(payload or {})
    set_admin_setting_json(data_interp_status_setting_key(user_id), current_payload, user=user)
    db.session.commit()
    return data_interp_present_job_snapshot(current_payload)


def get_data_interpretation_result_payload_for_user(user_id):
    if not user_id:
        return {}
    stored = get_admin_setting_json(data_interp_result_setting_key(user_id), {}) or {}
    if not isinstance(stored, dict):
        stored = {}
    return stored


def get_data_interpretation_result_for_user(user_id, selected_fiscal_year=None):
    stored = get_data_interpretation_result_payload_for_user(user_id)
    if not stored:
        return None
    if selected_fiscal_year and stored.get('selected_fiscal_year') != selected_fiscal_year:
        return None
    result = stored.get('analysis_result')
    return result if isinstance(result, dict) else None


def store_data_interpretation_result_for_user(user_id, selected_fiscal_year, analysis_result, user=None):
    payload = {
        'selected_fiscal_year': selected_fiscal_year,
        'stored_at': utc_now().isoformat(timespec='seconds'),
        'analysis_result': analysis_result,
    }
    if user_id:
        set_admin_setting_json(data_interp_result_setting_key(user_id), payload, user=user)
        db.session.commit()
    return payload


def report_snapshot(report):
    snapshot = {}
    for field in TRACKED_REPORT_FIELDS:
        value = getattr(report, field, None)
        if isinstance(value, datetime):
            snapshot[field] = value.isoformat(sep=' ', timespec='seconds')
        elif hasattr(value, 'isoformat'):
            snapshot[field] = value.isoformat()
        else:
            snapshot[field] = '' if value is None else str(value)
    return snapshot


def record_report_field_changes(report, before_snapshot, action, user=None):
    after_snapshot = report_snapshot(report)
    actor = user if user is not None else (current_user if current_user.is_authenticated else None)
    for field_name, old_value in before_snapshot.items():
        new_value = after_snapshot.get(field_name, '')
        if old_value != new_value:
            db.session.add(FieldChangeLog(
                report_id=report.id,
                user_id=actor.id if actor else None,
                action=action,
                field_name=field_name,
                old_value=old_value,
                new_value=new_value,
            ))


def allowed_upload(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_UPLOAD_EXTENSIONS


def ensure_upload_folder():
    target = configured_storage_directory('UPLOAD_FOLDER', 'uploads')
    return str(target)


def save_uploaded_file(file_storage, category, report=None, batch=None, user=None, status='uploaded', original_filename_override=None):
    upload_dir = Path(ensure_upload_folder())
    original_name = secure_filename(original_filename_override or file_storage.filename or 'upload.bin')
    extension = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else 'bin'
    stored_name = f"{utc_now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}.{extension}"
    storage_path = upload_dir / stored_name
    file_storage.save(str(storage_path))

    digest = hashlib.sha256()
    with open(storage_path, 'rb') as uploaded_stream:
        for chunk in iter(lambda: uploaded_stream.read(8192), b''):
            digest.update(chunk)

    uploaded_file = UploadedFile(
        report_id=report.id if report else None,
        batch_id=batch.id if batch else None,
        uploaded_by_id=user.id if user else None,
        category=category,
        original_filename=original_name,
        stored_filename=stored_name,
        storage_path=str(storage_path),
        mime_type=getattr(file_storage, 'mimetype', None),
        file_size=os.path.getsize(storage_path),
        sha256_hash=digest.hexdigest(),
        status=status,
    )
    db.session.add(uploaded_file)
    return uploaded_file


def create_import_error(batch, row_number, message, field_name=None, row_payload=None):
    db.session.add(ImportRowError(
        batch_id=batch.id,
        row_number=row_number,
        field_name=field_name,
        error_message=message,
        row_payload=json.dumps(row_payload, default=str) if row_payload is not None else None,
    ))
    batch.error_rows += 1


def load_tabular_import_rows(file_path, filename):
    ext = filename.rsplit('.', 1)[1].lower()
    if ext == 'csv':
        with open(file_path, newline='', encoding='utf-8-sig') as handle:
            return list(csv.DictReader(handle))
    if ext in {'xlsx', 'xls'}:
        try:
            pandas_module = get_pandas()
        except Exception:
            raise ValueError('Excel import requires pandas/openpyxl to be installed.')
        dataframe = pandas_module.read_excel(file_path)
        return dataframe.fillna('').to_dict(orient='records')
    raise ValueError('Unsupported import format.')


def allowed_image_upload(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'pdf'}


def build_batch_ocr_export_rows(saved_uploads, user):
    from data import build_report_from_ocr_upload

    rows = []
    for uploaded_file in saved_uploads:
        try:
            report, file_results, warnings = build_report_from_ocr_upload(
                user,
                [uploaded_file.storage_path],
                [uploaded_file.original_filename],
            )
            report.user_id = user.id
            report.last_modified_by_id = user.id
            report.data_source = 'ocr_batch_upload'
            assign_pbo_name_normalized(report)
            add_and_flush_new_instance(report, label='OCR report')
            report.duplicate_flag = find_duplicate_pbo_report(
                report.pbo_name,
                reporting_period_start=report.reporting_period_start,
                exclude_report_id=report.id,
                require_start_month_match=True,
            ) is not None
            report.form_14 = uploaded_file.original_filename
            report.update_risk_score(compute_tf_risk)

            result = file_results[0] if file_results else {"status": "ocr_processed", "text": "", "error": None}
            uploaded_file.report_id = report.id
            uploaded_file.status = result['status']
            uploaded_file.extracted_text = result['text']
            uploaded_file.error_message = result['error']

            log_user_activity(
                'ocr_batch_report_imported',
                report=report,
                summary=f'{user.email} created report {report.id} from batch OCR file {uploaded_file.original_filename}.',
            )

            rows.append({
                'report_id': report.id,
                'source_file': uploaded_file.original_filename,
                'status': uploaded_file.status,
                'error': uploaded_file.error_message or '',
                'warnings': ' | '.join(warnings),
                'pbo_name': report.pbo_name or '',
                'registration_number': report.registration_number or '',
                'pin_number': report.pin_number or '',
                'date_of_registration': report.date_of_registration or '',
                'postal_address': report.postal_address or '',
                'physical_address': report.physical_address or '',
                'telephone': report.telephone or '',
                'cell_phone': report.cell_phone or '',
                'email': report.email or '',
                'website': report.website or '',
                'contact_name': report.contact_name or '',
                'contact_position': report.contact_position or '',
                'contact_telephone': report.contact_telephone or '',
                'contact_email': report.contact_email or '',
                'contact_nationality': report.contact_nationality or '',
                'counties': report.counties or '',
                'cash_balance_previous_year': report.cash_balance_previous_year or '',
                'income_b2_total': report.income_b2_total or '',
                'receipts_total': report.receipts_total or '',
                'cash_bank_balance': report.cash_bank_balance or '',
                'audited': report.audited or '',
                'extracted_text': uploaded_file.extracted_text or '',
            })
        except Exception as exc:
            uploaded_file.status = 'ocr_failed'
            uploaded_file.extracted_text = ''
            uploaded_file.error_message = str(exc)
            rows.append({
                'report_id': '',
                'source_file': uploaded_file.original_filename,
                'status': 'ocr_failed',
                'error': str(exc),
                'warnings': '',
                'pbo_name': '',
                'registration_number': '',
                'pin_number': '',
                'date_of_registration': '',
                'postal_address': '',
                'physical_address': '',
                'telephone': '',
                'cell_phone': '',
                'email': '',
                'website': '',
                'contact_name': '',
                'contact_position': '',
                'contact_telephone': '',
                'contact_email': '',
                'contact_nationality': '',
                'counties': '',
                'cash_balance_previous_year': '',
                'income_b2_total': '',
                'receipts_total': '',
                'cash_bank_balance': '',
                'audited': '',
                'extracted_text': '',
            })

    return rows


def mail_configured():
    return bool(
        app.config.get('MAIL_SERVER')
        and app.config.get('MAIL_USERNAME')
        and app.config.get('MAIL_PASSWORD')
        and app.config.get('MAIL_DEFAULT_SENDER')
    )


def send_email_message(subject, recipients, body):
    if not recipients or not mail_configured():
        return False

    message = EmailMessage()
    message['Subject'] = subject
    message['From'] = app.config['MAIL_DEFAULT_SENDER']
    message['To'] = ', '.join(recipients)
    message.set_content(body)

    with smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT']) as smtp:
        if app.config['MAIL_USE_TLS']:
            smtp.starttls()
        smtp.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        smtp.send_message(message)
    return True


def empty_form14_prefill():
    return {
        'return_date': return_date_input_value(None),
        'income_b2_total': None,
        'receipts_total': None,
        'cash_bank_balance': None,
    }


def send_email_async(subject, recipients, body):
    if not recipients or not mail_configured():
        return

    def _deliver():
        try:
            send_email_message(subject, recipients, body)
        except Exception as exc:
            app.logger.exception('Email delivery failed: %s', exc)

    threading.Thread(target=_deliver, name='async-mailer', daemon=True).start()


def notify_user_login(user):
    if not user or user.role == 'admin' or not user.email:
        return
    send_email_async(
        subject='Form 14 login notification',
        recipients=[user.email],
        body=(
            f'Hello,\n\n'
            f'Your Form 14 account ({user.email}) logged in successfully on '
            f'{utc_now().strftime("%Y-%m-%d %H:%M:%S")} UTC.\n\n'
            f'If this was not you, please contact the administrator immediately.'
        ),
    )


def perform_application_backup():
    if app.config.get('GOOGLE_DRIVE_ENABLED'):
        os.environ.pop('GOOGLE_DRIVE_DISABLE_UPLOAD', None)
    else:
        os.environ['GOOGLE_DRIVE_DISABLE_UPLOAD'] = '1'
    result = run_backup_once(
        database_url=resolve_backup_database_url(),
        progress_callback=update_backup_runtime_status,
    )
    guidance = result.get('drive_upload_guidance') or {}
    credential_details = result.get('drive_credential_details') or {}
    with BACKUP_RUNTIME_LOCK:
        BACKUP_RUNTIME_STATUS['last_backup_at'] = utc_now().isoformat(timespec='seconds')
        BACKUP_RUNTIME_STATUS['last_backup_archive'] = str(result['archive_path']) if result else None
        BACKUP_RUNTIME_STATUS['last_backup_error'] = result.get('drive_upload_error')
        BACKUP_RUNTIME_STATUS['local_backup_succeeded'] = result.get('local_backup_succeeded')
        BACKUP_RUNTIME_STATUS['drive_upload_succeeded'] = result.get('drive_upload_succeeded')
        BACKUP_RUNTIME_STATUS['drive_refresh_command'] = guidance.get('refresh_command')
        BACKUP_RUNTIME_STATUS['drive_auth_recommendation'] = guidance.get('recommendation')
        BACKUP_RUNTIME_STATUS['drive_credential_type'] = credential_details.get('type')
        BACKUP_RUNTIME_STATUS['drive_credential_source'] = credential_details.get('source')
    return result


def update_backup_runtime_status(progress_payload=None, **kwargs):
    payload = dict(progress_payload or {})
    payload.update(kwargs)
    with BACKUP_RUNTIME_LOCK:
        BACKUP_RUNTIME_STATUS.update(payload)


def backup_status_snapshot():
    with BACKUP_RUNTIME_LOCK:
        return dict(BACKUP_RUNTIME_STATUS)


def update_report_merge_runtime_status(progress_payload=None, **kwargs):
    payload = dict(progress_payload or {})
    payload.update(kwargs)
    with REPORT_MERGE_RUNTIME_LOCK:
        REPORT_MERGE_RUNTIME_STATUS.update(payload)


def report_merge_status_snapshot():
    with REPORT_MERGE_RUNTIME_LOCK:
        return dict(REPORT_MERGE_RUNTIME_STATUS)


def report_merge_export_directory():
    return configured_storage_directory('REPORT_MERGE_EXPORT_DIR', 'report_merge_exports')


def report_merge_temp_directory():
    return configured_storage_directory('REPORT_MERGE_TEMP_DIR', 'report_merge_temp')


def report_merge_download_name(file_name):
    name = Path(file_name or '').name
    if re.match(r'^[0-9a-f]{32}_', name):
        name = name[33:]
    if re.match(r'^.+__\d{8}__\d{6}\.xlsx$', name, flags=re.IGNORECASE):
        return name
    match = re.match(r'^(.*)_(\d{8})_(\d{6})\.xlsx$', name, flags=re.IGNORECASE)
    if match:
        prefix, file_date, file_time = match.groups()
        return f"{prefix}__{file_date}__{file_time}.xlsx"
    return name


def report_merge_output_prefix(selected_source_key=None, uploaded_filename=None):
    selected_key = re.sub(r'[^a-z0-9_]+', '_', str(selected_source_key or '').strip().lower()).strip('_')
    if selected_key:
        return selected_key
    uploaded_stem = re.sub(
        r'[^a-z0-9_]+',
        '_',
        Path(str(uploaded_filename or '')).stem.strip().lower(),
    ).strip('_')
    if uploaded_stem:
        return uploaded_stem
    return 'merged_reports'


def list_report_merge_history(limit=40):
    root = report_merge_export_directory()
    rows = []
    if not root.exists():
        return rows
    for file_path in root.glob('*.xlsx'):
        if not file_path.is_file():
            continue
        stat_result = file_path.stat()
        modified_at = datetime.fromtimestamp(stat_result.st_mtime, timezone.utc)
        download_name = report_merge_download_name(file_path.name)
        rows.append({
            'id': file_path.name,
            'stored_name': file_path.name,
            'download_name': download_name,
            'display_name': download_name.rsplit('.', 1)[0],
            'modified_at': modified_at.isoformat(),
            'modified_at_display': modified_at.strftime('%Y-%m-%d %H:%M:%S UTC'),
            'size_bytes': int(stat_result.st_size or 0),
            'size_display': _format_file_size(stat_result.st_size or 0),
        })
    rows.sort(key=lambda item: item.get('modified_at') or '', reverse=True)
    return rows[:limit]


def _drop_report_merge_output_cache_for_path(target_path):
    target_resolved = Path(target_path).resolve()
    for token, item in list(REPORT_MERGE_OUTPUTS.items()):
        item_path = item.get('path')
        if not item_path:
            continue
        try:
            cached_resolved = Path(item_path).resolve()
        except Exception:
            continue
        if cached_resolved == target_resolved:
            REPORT_MERGE_OUTPUTS.pop(token, None)


def delete_report_merge_history_file(file_name):
    root = report_merge_export_directory()
    selected = (file_name or '').strip()
    if not selected:
        return False
    target = (root / selected).resolve()
    if not _path_within_base(target, root):
        return False
    if not target.exists() or not target.is_file():
        return False
    target.unlink()
    _drop_report_merge_output_cache_for_path(target)
    return True


def delete_all_report_merge_history_files():
    root = report_merge_export_directory()
    deleted_count = 0
    if not root.exists():
        return deleted_count
    for file_path in root.glob('*.xlsx'):
        if not file_path.is_file():
            continue
        resolved_path = file_path.resolve()
        if not _path_within_base(resolved_path, root):
            continue
        file_path.unlink()
        _drop_report_merge_output_cache_for_path(resolved_path)
        deleted_count += 1
    return deleted_count


def latest_report_merge_reference_label():
    return f"pbo_reports_live_{utc_now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def report_merge_reference_columns():
    return [column.name for column in PBOReport.__table__.columns]


def serialize_merge_value(value):
    return value


def _build_merge_rows_from_table(model, order_by=None):
    table = model.__table__
    columns = [column.name for column in table.columns]
    sqlite_rowid_label = None
    use_sqlite_rowid_as_id = False
    select_columns = []
    if db.engine.dialect.name == 'sqlite':
        sqlite_rowid_label = '__sqlite_rowid'
        use_sqlite_rowid_as_id = not sqlite_table_has_real_id_primary_key(db.session.connection(), table.name)
        select_columns.append(literal_column('rowid').label(sqlite_rowid_label))
    select_columns.extend(table.c[column_name] for column_name in columns)

    statement = select(*select_columns).select_from(table)
    order_clauses = []
    for column_name, direction in (order_by or []):
        if str(column_name).lower() == 'rowid':
            if db.engine.dialect.name != 'sqlite':
                continue
            sort_column = literal_column('rowid')
        else:
            sort_column = table.c.get(column_name)
            if sort_column is None:
                continue
        direction_upper = str(direction or 'ASC').upper()
        order_clauses.append(sort_column.desc() if direction_upper == 'DESC' else sort_column.asc())
    if order_clauses:
        statement = statement.order_by(*order_clauses)

    result = db.session.execute(statement).mappings().all()
    rows = []
    for mapping in result:
        row = {column: serialize_merge_value(mapping.get(column)) for column in columns}
        sqlite_rowid = mapping.get(sqlite_rowid_label) if sqlite_rowid_label else None
        if sqlite_rowid_label and use_sqlite_rowid_as_id and sqlite_rowid is not None:
            row['id'] = int(sqlite_rowid)
        elif sqlite_rowid_label and row.get('id') in (None, '') and sqlite_rowid is not None:
            row['id'] = int(sqlite_rowid)
        rows.append(row)
    return columns, rows


def build_report_merge_reference_rows():
    _, rows = _build_merge_rows_from_table(
        PBOReport,
        order_by=[('created_at', 'DESC'), ('id', 'DESC'), ('rowid', 'DESC')],
    )
    return rows


def get_report_merge_related_models():
    return [
        Asset,
        IncomeGeneratingActivity,
        Donation,
        Grant,
        Payment,
        BankAccount,
        AuditorEntry,
        StaffBiodata,
        VolunteerBiodata,
        VolunteerPrivilege,
        TrainingRecord,
        TaxWaiverItem,
        Official,
        ProjectImplementation,
        ProjectCarriedOut,
        CollaborationNetworking,
    ]


def get_report_merge_source_definitions():
    definitions = []
    for model in get_report_merge_related_models():
        table_name = getattr(model, '__tablename__', '').strip()
        if not table_name:
            continue
        definitions.append({
            'key': table_name,
            'label': table_name,
            'model': model,
            'description': f'Database table {table_name} joined to pbo_reports through report_id.',
        })
    definitions.sort(key=lambda item: item['label'])
    return definitions


def get_report_merge_source_definition_map():
    return {row['key']: row for row in get_report_merge_source_definitions()}


def build_report_merge_source_export(source_key, output_path=None):
    definition = get_report_merge_source_definition_map().get(source_key)
    if definition is None:
        raise RuntimeError('Choose a valid live database merge source.')
    model = definition['model']
    order_by = [('report_id', 'ASC'), ('id', 'ASC'), ('rowid', 'ASC')] if hasattr(model, 'report_id') else [('id', 'ASC'), ('rowid', 'ASC')]
    columns, rows = _build_merge_rows_from_table(model, order_by=order_by)
    if output_path:
        write_row_dicts_workbook(
            output_path=str(output_path),
            sheet_name=definition['label'],
            headers=columns,
            rows=rows,
        )
    return {
        'definition': definition,
        'rows_count': len(rows),
        'path': str(output_path) if output_path else None,
        'rows': rows,
        'columns': columns,
    }


def perform_report_merge_export(selected_source_key=None, uploaded_file_id=None):
    selected_key = (selected_source_key or '').strip()
    if not selected_key and not uploaded_file_id:
        raise RuntimeError('Choose a live database source or upload a merge workbook first.')

    output_dir = report_merge_export_directory()
    reference_name = latest_report_merge_reference_label()

    update_report_merge_runtime_status(
        stage='preparing',
        detail='Preparing latest live pbo_reports database rows',
        reference_name=reference_name,
        progress_percent=5,
    )
    reference_rows = build_report_merge_reference_rows()

    source_specs = []
    selected_sources = []
    if selected_key:
        definition = get_report_merge_source_definition_map().get(selected_key)
        if definition is None:
            raise RuntimeError('Choose a valid live database source.')
        update_report_merge_runtime_status(
            stage='preparing',
            detail=f"Loading database table {definition['label']}",
            current_source=definition['label'],
            progress_percent=12,
        )
        source_export = build_report_merge_source_export(selected_key)
        source_specs.append({
            'label': definition['label'],
            'rows': source_export['rows'],
            'columns': source_export['columns'],
            'related_key': 'report_id',
        })
        selected_sources.append(definition['label'])

    uploaded_merge_file = None
    if uploaded_file_id:
        uploaded_merge_file = db.session.get(UploadedFile, uploaded_file_id)
        if uploaded_merge_file is None:
            raise RuntimeError('Uploaded merge file could not be found.')
        update_report_merge_runtime_status(
            stage='preparing',
            detail=f"Loading uploaded workbook {uploaded_merge_file.original_filename}",
            current_source=uploaded_merge_file.original_filename,
            progress_percent=15,
        )
        uploaded_rows = load_tabular_import_rows(
            uploaded_merge_file.storage_path,
            uploaded_merge_file.original_filename,
        )
        uploaded_columns = list(uploaded_rows[0].keys()) if uploaded_rows else ['report_id']
        source_specs.append({
            'label': Path(uploaded_merge_file.original_filename or '').stem or 'uploaded_merge_file',
            'rows': uploaded_rows,
            'columns': uploaded_columns,
            'related_key': 'report_id',
        })
        selected_sources.append(f"Upload: {uploaded_merge_file.original_filename}")

    if not source_specs:
        raise RuntimeError('Choose a live database source or upload a merge workbook first.')

    update_report_merge_runtime_status(
        stage='merging',
        detail='Joining database sources to pbo_reports',
        current_source=None,
        sources_total=len(source_specs),
        sources_completed=0,
        selected_sources=selected_sources,
        progress_percent=18,
    )

    def progress_callback(payload):
        total = int(payload.get('sources_total') or len(source_specs) or 1)
        completed = int(payload.get('sources_completed') or 0)
        progress_percent = 18 + int((completed / total) * 72)
        update_report_merge_runtime_status(
            stage=payload.get('stage') or 'merging_database',
            detail=payload.get('detail') or 'Joining database sources to pbo_reports',
            current_source=payload.get('current_source'),
            sources_total=total,
            sources_completed=completed,
            selected_sources=selected_sources,
            progress_percent=min(progress_percent, 95),
        )

    output_token = uuid.uuid4().hex
    timestamp = utc_now().strftime('%Y%m%d__%H%M%S')
    output_prefix = report_merge_output_prefix(
        selected_source_key=selected_key,
        uploaded_filename=uploaded_merge_file.original_filename if uploaded_merge_file else None,
    )
    output_name = f"{output_prefix}__{timestamp}.xlsx"
    stored_name = f"{output_token}_{output_name}"
    output_path = output_dir / stored_name
    merge_result = merge_report_driven_sources(
        reference_rows=reference_rows,
        source_specs=source_specs,
        output_path=str(output_path),
        reference_label='pbo_reports',
        reference_key='id',
        progress_callback=progress_callback,
    )
    REPORT_MERGE_OUTPUTS[output_token] = {
        'path': str(output_path),
        'download_name': output_name,
    }
    update_report_merge_runtime_status(
        running=False,
        stage='complete',
        detail='Database merge complete. Download the cleaned workbook below.',
        finished_at=utc_now().isoformat(timespec='seconds'),
        current_source=None,
        sources_total=len(source_specs),
        sources_completed=len(source_specs),
        selected_sources=selected_sources,
        progress_percent=100,
        download_ready=True,
        download_token=output_token,
        download_name=output_name,
        download_url=f"/admin/export/reports/files/{stored_name}",
        last_error=None,
        reference_name=reference_name,
    )
    return merge_result


def _run_report_merge_export_job(triggered_by_user_id=None, selected_source_key=None, uploaded_file_id=None):
    with app.app_context():
        try:
            perform_report_merge_export(
                selected_source_key=selected_source_key,
                uploaded_file_id=uploaded_file_id,
            )
        except Exception as exc:
            db.session.rollback()
            update_report_merge_runtime_status(
                running=False,
                stage='failed',
                detail='Report merge failed',
                finished_at=utc_now().isoformat(timespec='seconds'),
                current_source=None,
                progress_percent=0,
                download_ready=False,
                download_token=None,
                download_name=None,
                download_url=None,
                last_error=str(exc),
            )
            app.logger.exception('Report merge job failed: %s', exc)
        finally:
            db.session.remove()


def start_report_merge_export_job(triggered_by_user_id=None, selected_source_key=None, uploaded_file_id=None):
    with REPORT_MERGE_RUNTIME_LOCK:
        if REPORT_MERGE_RUNTIME_STATUS.get('running'):
            return False

    total_sources = (1 if (selected_source_key or '').strip() else 0) + (1 if uploaded_file_id else 0)
    update_report_merge_runtime_status(
        running=True,
        stage='queued',
        detail='Queueing report merge job',
        started_at=utc_now().isoformat(timespec='seconds'),
        finished_at=None,
        current_source=None,
        sources_total=total_sources,
        sources_completed=0,
        selected_sources=[],
        progress_percent=1,
        download_ready=False,
        download_token=None,
        download_name=None,
        download_url=None,
        last_error=None,
        reference_name=latest_report_merge_reference_label(),
    )
    worker = threading.Thread(
        target=_run_report_merge_export_job,
        kwargs={
            'triggered_by_user_id': triggered_by_user_id,
            'selected_source_key': selected_source_key,
            'uploaded_file_id': uploaded_file_id,
        },
        name='report-merge-export-worker',
        daemon=True,
    )
    worker.start()
    return True


def backup_root_directory():
    return Path(os.getenv('DB_BACKUP_DIR', './backups')).resolve()


def _path_within_base(path_value, base_path):
    try:
        path_value.resolve().relative_to(base_path.resolve())
        return True
    except Exception:
        return False


def _format_file_size(num_bytes):
    try:
        size = float(num_bytes or 0)
    except Exception:
        size = 0.0
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    idx = 0
    while size >= 1024 and idx < len(units) - 1:
        size /= 1024.0
        idx += 1
    return f"{size:.1f} {units[idx]}"


def list_local_backup_versions(limit=200):
    root = backup_root_directory()
    if not root.exists():
        return []
    allowed_suffixes = {'.zip', '.dump', '.sql', '.sqlite', '.db'}
    rows = []
    for file_path in root.rglob('*'):
        if not file_path.is_file():
            continue
        if file_path.suffix.lower() not in allowed_suffixes:
            continue
        stat_result = file_path.stat()
        rel_path = str(file_path.resolve().relative_to(root.resolve()))
        modified_at = datetime.fromtimestamp(stat_result.st_mtime, timezone.utc)
        rows.append({
            'id': rel_path,
            'source': 'local',
            'name': file_path.name,
            'relative_path': rel_path,
            'absolute_path': str(file_path.resolve()),
            'modified_at': modified_at.isoformat(),
            'modified_at_display': modified_at.strftime('%Y-%m-%d %H:%M:%S UTC'),
            'size_bytes': int(stat_result.st_size or 0),
            'size_display': _format_file_size(stat_result.st_size or 0),
            'suffix': file_path.suffix.lower(),
            'label': f"{file_path.name} ({modified_at.strftime('%Y-%m-%d %H:%M')}, {_format_file_size(stat_result.st_size or 0)})",
        })
    rows.sort(key=lambda item: item.get('modified_at') or '', reverse=True)
    return rows[:limit]


def list_google_backup_versions(limit=120):
    folder_id = (os.getenv('GOOGLE_DRIVE_FOLDER_ID') or '').strip()
    if not folder_id:
        return {'items': [], 'error': 'GOOGLE_DRIVE_FOLDER_ID is not configured.'}
    try:
        service = build_drive_service()
        query = f"'{folder_id}' in parents and trashed = false"
        response = service.files().list(
            q=query,
            pageSize=limit,
            orderBy='modifiedTime desc',
            fields='files(id,name,modifiedTime,size,mimeType)',
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        files = response.get('files') or []
        allowed_suffixes = ('.zip', '.dump', '.sql', '.sqlite', '.db')
        rows = []
        for item in files:
            file_name = str(item.get('name') or '').strip()
            if not file_name:
                continue
            suffix = Path(file_name).suffix.lower()
            if suffix not in allowed_suffixes:
                continue
            size_bytes = int(item.get('size') or 0)
            modified_iso = item.get('modifiedTime') or ''
            modified_display = modified_iso
            try:
                modified_display = datetime.fromisoformat(modified_iso.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S UTC')
            except Exception:
                pass
            rows.append({
                'id': str(item.get('id') or '').strip(),
                'source': 'google_drive',
                'name': file_name,
                'modified_at': modified_iso,
                'modified_at_display': modified_display,
                'size_bytes': size_bytes,
                'size_display': _format_file_size(size_bytes),
                'suffix': suffix,
                'label': f"{file_name} ({modified_display}, {_format_file_size(size_bytes)})",
            })
        return {'items': rows, 'error': None}
    except Exception as exc:
        return {'items': [], 'error': str(exc)}


def get_backup_rollback_catalog(force=False):
    now_ts = time.time()
    cache_ttl = max(int(os.getenv('BACKUP_VERSION_CACHE_TTL_SECONDS', '90') or 90), 15)
    cached_payload = BACKUP_VERSION_CACHE.get('payload')
    cached_at = float(BACKUP_VERSION_CACHE.get('updated_at') or 0.0)
    if (not force) and cached_payload and (now_ts - cached_at) < cache_ttl:
        return cached_payload

    local_versions = list_local_backup_versions(limit=220)
    google_result = list_google_backup_versions(limit=140)
    payload = {
        'local_versions': local_versions,
        'google_versions': google_result.get('items') or [],
        'google_error': google_result.get('error'),
        'generated_at': utc_now().isoformat(timespec='seconds'),
    }
    BACKUP_VERSION_CACHE['payload'] = payload
    BACKUP_VERSION_CACHE['updated_at'] = now_ts
    return payload


def _resolve_local_backup_selection(relative_path):
    root = backup_root_directory()
    if not root.exists():
        raise RuntimeError('Backup root directory does not exist.')
    selected = (relative_path or '').strip()
    if not selected:
        raise RuntimeError('Select a backup file first.')
    resolved = (root / selected).resolve()
    if not _path_within_base(resolved, root):
        raise RuntimeError('Invalid backup path selection.')
    if not resolved.exists() or not resolved.is_file():
        raise RuntimeError('Selected backup file no longer exists.')
    return resolved


def _save_uploaded_rollback_file(file_storage):
    if file_storage is None or not getattr(file_storage, 'filename', ''):
        raise RuntimeError('Choose a rollback file to upload.')
    safe_name = secure_filename(file_storage.filename or '')
    if not safe_name:
        raise RuntimeError('Invalid uploaded rollback file name.')
    BACKUP_ROLLBACK_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(safe_name).suffix.lower()
    if suffix not in {'.zip', '.dump', '.sql', '.sqlite', '.db'}:
        raise RuntimeError('Unsupported rollback file type.')
    target = BACKUP_ROLLBACK_UPLOAD_DIR / f"upload_{uuid.uuid4().hex}_{safe_name}"
    file_storage.save(str(target))
    return target


def _download_google_backup_file(file_id):
    selected_id = (file_id or '').strip()
    if not selected_id:
        raise RuntimeError('Select a Google backup file first.')
    BACKUP_ROLLBACK_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    try:
        from googleapiclient.http import MediaIoBaseDownload  # type: ignore
    except Exception as exc:
        raise RuntimeError(f'Google download client is not available: {exc}') from exc

    service = build_drive_service()
    meta = service.files().get(fileId=selected_id, fields='id,name').execute()
    file_name = secure_filename(str(meta.get('name') or 'google_backup_file'))
    if not file_name:
        file_name = f'google_backup_{selected_id}'
    request_handle = service.files().get_media(fileId=selected_id)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request_handle)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    target = BACKUP_ROLLBACK_UPLOAD_DIR / f"gdrive_{uuid.uuid4().hex}_{file_name}"
    target.write_bytes(buffer.getvalue())
    return target


def _resolve_restore_payload_file(source_file_path):
    if source_file_path.suffix.lower() != '.zip':
        return source_file_path
    extraction_dir = BACKUP_ROLLBACK_UPLOAD_DIR / f"extract_{uuid.uuid4().hex}"
    extraction_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(source_file_path, 'r') as archive:
        archive.extractall(extraction_dir)
    candidates = []
    for file_path in extraction_dir.rglob('*'):
        if file_path.is_file() and file_path.suffix.lower() in {'.dump', '.sql', '.sqlite', '.db'}:
            candidates.append(file_path)
    if not candidates:
        raise RuntimeError('ZIP file has no supported database payload (.dump/.sql/.sqlite/.db).')
    preferred_ext = {'.dump': 0, '.sql': 1, '.sqlite': 2, '.db': 3}
    candidates.sort(
        key=lambda item: (
            preferred_ext.get(item.suffix.lower(), 99),
            -item.stat().st_size,
        )
    )
    return candidates[0]


def _resolve_sqlite_path_from_url(database_url):
    if not database_url_is_sqlite(database_url):
        return None
    raw_path = database_url.replace('sqlite:///', '', 1)
    db_path = Path(raw_path)
    return db_path if db_path.is_absolute() else (Path.cwd() / db_path)


def _resolve_mysql_connection_from_url(database_url):
    if not database_url_is_mysql(database_url):
        return None
    parsed = urlparse(database_url)
    database_name = (parsed.path or "").lstrip("/")
    if not database_name:
        raise RuntimeError('MySQL database URL is missing the database name.')
    return {
        'host': parsed.hostname or 'localhost',
        'port': parsed.port or 3306,
        'username': parsed.username or '',
        'password': parsed.password or '',
        'database': database_name,
    }


def _mysql_cli_env(password):
    command_env = os.environ.copy()
    if password:
        command_env['MYSQL_PWD'] = password
    return command_env


def _create_mysql_pre_rollback_snapshot(database_url, snapshot_dir, timestamp):
    snapshot_path = snapshot_dir / f'pre_rollback_{timestamp}.sql'
    config = _resolve_mysql_connection_from_url(database_url)
    cmd = [
        'mysqldump',
        '--host', config['host'],
        '--port', str(config['port']),
        '--user', config['username'],
        '--single-transaction',
        '--skip-lock-tables',
        config['database'],
    ]
    with snapshot_path.open('wb') as snapshot_file:
        process = subprocess.run(
            cmd,
            stdout=snapshot_file,
            stderr=subprocess.PIPE,
            env=_mysql_cli_env(config['password']),
        )
    if process.returncode != 0:
        error_message = (process.stderr or b'').decode('utf-8', errors='replace').strip()
        raise RuntimeError(f'Pre-rollback mysqldump failed: {error_message or "Unknown MySQL dump error"}')
    return snapshot_path


def _create_pre_rollback_snapshot(database_url):
    root = backup_root_directory()
    timestamp = utc_now().strftime('%Y%m%d_%H%M%S')
    snapshot_dir = root / 'rollback_snapshots'
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    if database_url_is_postgresql(database_url):
        snapshot_path = snapshot_dir / f'pre_rollback_{timestamp}.dump'
        cmd = ['pg_dump', '--format=custom', '--file', str(snapshot_path), database_url]
        process = subprocess.run(cmd, capture_output=True, text=True)
        if process.returncode != 0:
            raise RuntimeError(f'Pre-rollback pg_dump failed: {(process.stderr or process.stdout or "").strip()}')
        return snapshot_path
    if database_url_is_mysql(database_url):
        return _create_mysql_pre_rollback_snapshot(database_url, snapshot_dir, timestamp)
    sqlite_path = _resolve_sqlite_path_from_url(database_url)
    if sqlite_path is None or not sqlite_path.exists():
        raise RuntimeError('Active SQLite database file was not found for snapshot.')
    snapshot_path = snapshot_dir / f'pre_rollback_{timestamp}{sqlite_path.suffix or ".sqlite"}'
    shutil.copy2(sqlite_path, snapshot_path)
    return snapshot_path


def _restore_postgres_from_payload(database_url, payload_file):
    suffix = payload_file.suffix.lower()
    if suffix == '.dump':
        cmd = [
            'pg_restore',
            '--clean',
            '--if-exists',
            '--no-owner',
            '--no-privileges',
            '--single-transaction',
            '--dbname',
            database_url,
            str(payload_file),
        ]
    elif suffix == '.sql':
        cmd = ['psql', database_url, '-v', 'ON_ERROR_STOP=1', '-f', str(payload_file)]
    else:
        raise RuntimeError('Postgres rollback supports .dump or .sql payloads only.')
    process = subprocess.run(cmd, capture_output=True, text=True)
    if process.returncode != 0:
        message = (process.stderr or process.stdout or 'Unknown restore error').strip()
        raise RuntimeError(f'Rollback restore failed: {message}')


def _restore_sqlite_from_payload(database_url, payload_file):
    suffix = payload_file.suffix.lower()
    if suffix not in {'.sqlite', '.db'}:
        raise RuntimeError('SQLite rollback supports .sqlite or .db payloads only.')
    target_path = _resolve_sqlite_path_from_url(database_url)
    if target_path is None:
        raise RuntimeError('Current database URL is not SQLite.')
    target_path.parent.mkdir(parents=True, exist_ok=True)
    if not payload_file.exists():
        raise RuntimeError('Rollback payload file no longer exists.')
    shutil.copy2(payload_file, target_path)


def _restore_mysql_from_payload(database_url, payload_file):
    suffix = payload_file.suffix.lower()
    if suffix != '.sql':
        raise RuntimeError('MySQL rollback currently supports .sql payloads only.')
    config = _resolve_mysql_connection_from_url(database_url)
    cmd = [
        'mysql',
        '--host', config['host'],
        '--port', str(config['port']),
        '--user', config['username'],
        config['database'],
    ]
    with payload_file.open('rb') as payload_stream:
        process = subprocess.run(
            cmd,
            stdin=payload_stream,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=_mysql_cli_env(config['password']),
        )
    if process.returncode != 0:
        error_message = (process.stderr or process.stdout or b'').decode('utf-8', errors='replace').strip()
        raise RuntimeError(f'Rollback restore failed: {error_message or "Unknown MySQL restore error"}')


def perform_database_rollback(source_mode, source_value, uploaded_file=None):
    database_url = resolve_backup_database_url()
    selected_mode = (source_mode or '').strip().lower()
    if selected_mode not in {'local', 'google', 'upload'}:
        raise RuntimeError('Invalid rollback source selected.')

    if selected_mode == 'local':
        source_file = _resolve_local_backup_selection(source_value)
    elif selected_mode == 'google':
        source_file = _download_google_backup_file(source_value)
    else:
        source_file = _save_uploaded_rollback_file(uploaded_file)

    payload_file = _resolve_restore_payload_file(source_file)

    db.session.remove()
    try:
        db.engine.dispose()
    except Exception:
        pass

    snapshot_path = _create_pre_rollback_snapshot(database_url)
    if database_url_is_postgresql(database_url):
        _restore_postgres_from_payload(database_url, payload_file)
    elif database_url_is_mysql(database_url):
        _restore_mysql_from_payload(database_url, payload_file)
    else:
        _restore_sqlite_from_payload(database_url, payload_file)

    try:
        db.engine.dispose()
    except Exception:
        pass

    return {
        'database_url': redact_database_url(database_url),
        'source_mode': selected_mode,
        'selected_file': str(source_file),
        'payload_file': str(payload_file),
        'pre_snapshot': str(snapshot_path),
    }


def _run_backup_job(triggered_by='system'):
    try:
        update_backup_runtime_status(
            running=True,
            stage='starting',
            detail=f'Starting backup requested by {triggered_by}',
            started_at=utc_now().isoformat(timespec='seconds'),
            last_backup_error=None,
            local_backup_succeeded=None,
            drive_upload_succeeded=None,
            files_total=0,
            files_completed=0,
            tables_total=0,
            tables_completed=0,
            current_file=None,
            current_table=None,
            latest_drive_upload=None,
            drive_refresh_command=None,
            drive_auth_recommendation=None,
            drive_credential_type=None,
            drive_credential_source=None,
        )
        result = perform_application_backup()
        final_stage = 'complete' if result.get('drive_upload_succeeded') or result.get('drive_upload_error') is None else 'upload_failed'
        guidance = result.get('drive_upload_guidance') or {}
        credential_details = result.get('drive_credential_details') or {}
        update_backup_runtime_status(
            running=False,
            stage=final_stage,
            detail='Backup finished' if final_stage == 'complete' else 'Backup finished locally but Google Drive upload failed',
            backup_dir=str(result['backup_dir']) if result and result.get('backup_dir') else None,
            last_backup_archive=str(result['archive_path']) if result and result.get('archive_path') else None,
            last_backup_at=utc_now().isoformat(timespec='seconds'),
            last_backup_error=result.get('drive_upload_error'),
            local_backup_succeeded=result.get('local_backup_succeeded'),
            drive_upload_succeeded=result.get('drive_upload_succeeded'),
            drive_refresh_command=guidance.get('refresh_command'),
            drive_auth_recommendation=guidance.get('recommendation'),
            drive_credential_type=credential_details.get('type'),
            drive_credential_source=credential_details.get('source'),
        )
    except Exception as exc:
        update_backup_runtime_status(
            running=False,
            stage='failed',
            detail='Backup failed',
            last_backup_error=str(exc),
        )
        app.logger.exception('Backup job failed: %s', exc)


def start_application_backup(triggered_by='system'):
    with BACKUP_RUNTIME_LOCK:
        if BACKUP_RUNTIME_STATUS.get('running'):
            return False
    worker = threading.Thread(
        target=_run_backup_job,
        kwargs={'triggered_by': triggered_by},
        name=f'backup-worker-{triggered_by}',
        daemon=True,
    )
    worker.start()
    return True


def _noon_backup_worker():
    last_run_date = None
    timezone = ZoneInfo(app.config.get('APP_TIMEZONE', 'Africa/Nairobi'))
    while True:
        try:
            now = datetime.now(timezone)
            if now.hour == 12 and last_run_date != now.date():
                if start_application_backup(triggered_by='scheduler'):
                    last_run_date = now.date()
        except Exception as exc:
            update_backup_runtime_status(last_backup_error=str(exc))
            app.logger.exception('Scheduled noon backup failed: %s', exc)
        time.sleep(60)


def start_noon_backup_worker():
    global NOON_BACKUP_WORKER_STARTED
    if NOON_BACKUP_WORKER_STARTED:
        return
    worker = threading.Thread(target=_noon_backup_worker, name='noon-backup-worker', daemon=True)
    worker.start()
    NOON_BACKUP_WORKER_STARTED = True


def log_user_activity(action, report=None, summary=None, user=None):
    actor = user if user is not None else (current_user if current_user.is_authenticated else None)
    entry = UserActivityLog(
        user_id=actor.id if actor else None,
        report_id=report.id if report and report.id else None,
        action=action,
        route=request.path if request else None,
        summary=summary or summarize_report(report),
        ip_address=get_request_ip() if request else None,
        user_agent=request.headers.get('User-Agent', '')[:512] if request else None,
    )
    db.session.add(entry)

    if report is not None:
        now = utc_now()
        report.last_activity_at = now
        if action in {'report_created', 'report_updated', 'admin_report_updated', 'admin_designate_updated'}:
            report.last_modified_by_id = actor.id if actor else None
        if action in {'report_opened', 'report_viewed'}:
            report.last_viewed_at = now
        if action == 'report_created' and report.submitted_at is None:
            report.submitted_at = now


def build_deleted_file_activity_summary(uploaded_file):
    if uploaded_file is None:
        return None
    payload = {
        'file_id': uploaded_file.id,
        'file_name': uploaded_file.original_filename,
        'stored_filename': uploaded_file.stored_filename,
        'category': uploaded_file.category,
        'uploaded_by_email': uploaded_file.uploaded_by.email if getattr(uploaded_file, 'uploaded_by', None) and uploaded_file.uploaded_by.email else None,
        'uploaded_by_name': uploaded_file.uploaded_by.full_name if getattr(uploaded_file, 'uploaded_by', None) and uploaded_file.uploaded_by.full_name else None,
    }
    return json.dumps(payload, sort_keys=True)


def parse_deleted_file_activity_summary(summary):
    if not summary:
        return {}
    try:
        payload = json.loads(summary)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def describe_uploaded_file_impacts(uploaded_file):
    impacts = []
    category = (uploaded_file.category or '').strip().lower()

    if uploaded_file.report is not None:
        report_label = f"Report #{uploaded_file.report.id}"
        impacts.append(f"{report_label} uploaded-file history")
        impacts.append("My Files / Open File access")
        if category == 'ocr_source':
            impacts.append(f"{report_label} OCR source trace")
        elif category == 'supporting_document':
            impacts.append(f"{report_label} supporting-document attachment")
        else:
            impacts.append(f"{report_label} file attachment")

    if uploaded_file.batch is not None:
        impacts.append(f"Import batch #{uploaded_file.batch.id} upload history")

    if uploaded_file.extracted_text:
        impacts.append("Stored OCR text snapshot")

    if not impacts:
        impacts.append("Uploaded file archive only")

    deduped_impacts = []
    for item in impacts:
        if item not in deduped_impacts:
            deduped_impacts.append(item)
    return deduped_impacts


def get_report_edit_field_help_metadata():
    return {
        'required_fields': [
            {
                'name': 'reporting_period_start',
                'label': 'Start Date',
                'state': 'enabled',
                'required': True,
                'purpose': 'Defines the opening date of the reporting period for the return.',
            },
            {
                'name': 'pbo_name',
                'label': 'PBO Name',
                'state': 'enabled',
                'required': True,
                'purpose': 'Identifies the organization being reported and is central to record matching.',
            },
        ],
        'sector_report_focus_fields': [
            {
                'name': 'scope',
                'label': 'Scope',
                'state': 'enabled',
                'used_in_sector_report': True,
                'purpose': 'Shows whether the PBO is national or international, which is used in grouped reporting.',
            },
            {
                'name': 'county',
                'label': 'Counties Of Operation',
                'state': 'enabled',
                'used_in_sector_report': True,
                'purpose': 'Lists where the PBO operated during the reporting period for county-level reporting.',
            },
            {
                'name': 'sector[]',
                'label': 'Project Sector',
                'state': 'enabled',
                'used_in_sector_report': True,
                'purpose': 'Captures the sector for each implementation row, which is a core sector-report dimension.',
            },
            {
                'name': 'project_county[]',
                'label': 'Project County',
                'state': 'locked_for_edit',
                'used_in_sector_report': True,
                'purpose': 'Feeds county-level project analysis even when the edit page locks the visible control.',
            },
            {
                'name': 'project_beneficiaries_no[]',
                'label': 'Project Beneficiaries',
                'state': 'locked_for_edit',
                'used_in_sector_report': True,
                'purpose': 'Supports beneficiary totals and sector impact summaries.',
            },
            {
                'name': 'project_spending_per_county[]',
                'label': 'Project Spending Per County',
                'state': 'locked_for_edit',
                'used_in_sector_report': True,
                'purpose': 'Provides county-level spending values for the sector report output.',
            },
            {
                'name': 'project_duration_years[]',
                'label': 'Project Duration',
                'state': 'locked_for_edit',
                'used_in_sector_report': True,
                'purpose': 'Helps describe the duration of implemented projects in analysis outputs.',
            },
            {
                'name': 'project_completion_status[]',
                'label': 'Project Completion Status',
                'state': 'locked_for_edit',
                'used_in_sector_report': True,
                'purpose': 'Used to summarize whether projects are complete, ongoing, or planned.',
            },
            {
                'name': 'payment_description[]',
                'label': 'Payment Categories',
                'state': 'enabled',
                'used_in_sector_report': True,
                'purpose': 'Supports spending analysis by expenditure type.',
            },
            {
                'name': 'payment_kenya[]',
                'label': 'Kenya Payments',
                'state': 'enabled',
                'used_in_sector_report': True,
                'purpose': 'Provides the Kenya-based payment values used in financial summaries.',
            },
            {
                'name': 'payment_other[]',
                'label': 'Other Country Payments',
                'state': 'enabled',
                'used_in_sector_report': True,
                'purpose': 'Adds non-Kenya spending values to payment reporting.',
            },
            {
                'name': 'assets_item[]',
                'label': 'Assets',
                'state': 'enabled',
                'used_in_sector_report': True,
                'purpose': 'Contributes to asset-based rollups and organization profile summaries.',
            },
            {
                'name': 'donor_amount[]',
                'label': 'Donations',
                'state': 'enabled',
                'used_in_sector_report': True,
                'purpose': 'Feeds donation totals and funding-source reporting.',
            },
            {
                'name': 'grant_amount[]',
                'label': 'Grants',
                'state': 'disabled',
                'used_in_sector_report': False,
                'purpose': 'This grants-awarded section is locked on edit and is not part of the main sector-report extraction focus.',
            },
            {
                'name': 'staff_*',
                'label': 'Staff And Volunteer Totals',
                'state': 'enabled',
                'used_in_sector_report': True,
                'purpose': 'Supports workforce and volunteer summary outputs tied to the report.',
            },
        ],
        'always_disabled_fields': [
            {
                'name': 'pbo_registration_number',
                'label': 'PBO Registration Number',
                'state': 'disabled',
                'required': False,
                'reason': 'reference_only',
                'used_in_sector_report': False,
                'purpose': 'Shown for reference only on edit and not intended for sector-report data entry on this screen.',
            },
            {
                'name': 'pbo_registration_date',
                'label': 'PBO Registration Date',
                'state': 'disabled',
                'required': False,
                'reason': 'reference_only',
                'used_in_sector_report': False,
                'purpose': 'Displayed as a protected reference value instead of an editable sector-report field here.',
            },
            {
                'name': 'kra_pin',
                'label': 'KRA PIN',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'Hidden and disabled in the edit form because it is not part of the sector-report extraction focus.',
            },
            {
                'name': 'pin_number',
                'label': 'PIN Number',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'Hidden and disabled in the edit form because it is not part of the sector-report extraction focus.',
            },
            {
                'name': 'postal_address',
                'label': 'Postal Address',
                'state': 'disabled',
                'required': False,
                'reason': 'reference_only',
                'used_in_sector_report': False,
                'purpose': 'Kept as a protected reference field and not used in sector, county, spending, or implementation rollups.',
            },
            {
                'name': 'physical_address',
                'label': 'Physical Address',
                'state': 'disabled',
                'required': False,
                'reason': 'reference_only',
                'used_in_sector_report': False,
                'purpose': 'Shown for background reference and not used in sector-report aggregation.',
            },
            {
                'name': 'telephone',
                'label': 'Telephone',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'Contact details stay locked because they are not part of sector or project reporting outputs.',
            },
            {
                'name': 'cell_phone',
                'label': 'Cell Phone',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'Contact details stay locked because they are not part of sector or project reporting outputs.',
            },
            {
                'name': 'email',
                'label': 'PBO Email',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'This contact field does not feed the sector report dataset.',
            },
            {
                'name': 'website',
                'label': 'Website',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'Website links are not used in sector, county, spending, or implementation summaries.',
            },
            {
                'name': 'social_media',
                'label': 'Social Media',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'Social links are outside the sector-report extraction scope.',
            },
            {
                'name': 'contact_position',
                'label': 'Contact Position',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'This contact-management detail is not used in sector reporting.',
            },
            {
                'name': 'contact_nationality',
                'label': 'Contact Nationality',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'This contact demographic field is not part of sector-report outputs.',
            },
            {
                'name': 'contact_gender',
                'label': 'Contact Gender',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'This contact demographic field is not part of sector-report outputs.',
            },
            {
                'name': 'gov_tax_waiver_amount',
                'label': 'Government Tax Waiver Amount',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'This government-specific value is not one of the main sector-report collection dimensions.',
            },
            {
                'name': 'gov_other_specify',
                'label': 'Government Other Specify',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'This government-specific text does not feed sector, county, or project implementation reporting.',
            },
            {
                'name': 'gov_other_amount',
                'label': 'Government Other Amount',
                'state': 'disabled',
                'required': False,
                'reason': 'out_of_scope_for_sector_report',
                'used_in_sector_report': False,
                'purpose': 'This government-specific amount does not feed the main sector report output.',
            },
        ],
        'conditional_fields': [
            {
                'group': 'Audit fields',
                'state': 'conditional',
                'trigger': 'Enabled only when Audited is Yes.',
                'used_in_sector_report': False,
                'purpose': 'Collects audit firm and auditor details only for audited organizations, but these are not core sector-report fields.',
            },
            {
                'group': 'Government tax waiver section',
                'state': 'conditional',
                'trigger': 'Disabled when the government sections are marked inactive.',
                'used_in_sector_report': False,
                'purpose': 'Avoids collecting government waiver fields when that section does not apply and keeps non-sector data out of the main reporting path.',
            },
            {
                'group': 'Officials section',
                'state': 'conditional',
                'trigger': 'Some official-role cells are read-only or disabled by design.',
                'used_in_sector_report': False,
                'purpose': 'Keeps governance role structure controlled while still allowing review, but this section is outside the main sector-report extraction focus.',
            },
            {
                'group': 'Project implementation detail fields',
                'state': 'conditional',
                'trigger': 'Some visible controls are locked on edit even though their stored values still support sector reports.',
                'used_in_sector_report': True,
                'purpose': 'Project sector, county, spending, beneficiaries, duration, and completion data still matter for sector reporting even when edit mode locks the visible widgets.',
            },
        ],
        'field_aliases': {
            'sector report': ['scope', 'counties of operation', 'project sector', 'project county', 'project spending per county', 'project beneficiaries', 'payment categories', 'kenya payments', 'other country payments'],
            'project implementation': ['project sector', 'project county', 'project beneficiaries', 'project spending per county', 'project duration', 'project completion status'],
            'disabled fields': ['pbo registration number', 'pbo registration date', 'kra pin', 'pin number', 'postal address', 'physical address', 'telephone', 'cell phone', 'email', 'website', 'social media', 'contact position', 'contact nationality', 'contact gender', 'government tax waiver amount'],
        },
        'notes': [
            'On the current edit page, only Start Date and PBO Name should be treated as required fields.',
            'For sector reporting, the main collection focus is on sectors, counties, project implementation rows, spending, beneficiaries, scope, and related financial summary fields.',
            'Many disabled fields are disabled because they are reference-only, out of scope for the sector report, or intentionally locked in edit mode.',
            'Some project fields may look disabled on the edit page but still correspond to stored project data that feeds sector reporting.',
        ],
    }


def build_report_edit_field_help_context():
    metadata = get_report_edit_field_help_metadata()
    required_items = list(metadata.get('required_fields') or [])
    sector_focus_items = list(metadata.get('sector_report_focus_fields') or [])
    disabled_items = list(metadata.get('always_disabled_fields') or [])

    extra_allowed_fields = [
        {'name': 'contact_name', 'label': 'Contact Name', 'purpose': 'Primary contact used for accountability checks.'},
        {'name': 'contact_telephone', 'label': 'Contact Telephone', 'purpose': 'Supports contact verification for filing follow-up.'},
        {'name': 'contact_email', 'label': 'Contact Email', 'purpose': 'Supports communication and filing follow-up.'},
        {'name': 'countries_of_operation', 'label': 'Countries Of Operation', 'purpose': 'Used to classify cross-border operations.'},
        {'name': 'assets_item[]', 'label': 'Asset Rows', 'purpose': 'Supports high-level asset analysis for reporting.'},
        {'name': 'donor_name[]', 'label': 'Donor Name', 'purpose': 'Supports funding-source analysis in reports.'},
        {'name': 'iga_description[]', 'label': 'Income Generating Activities', 'purpose': 'Supports sustainability insights for the sector.'},
        {'name': 'audited', 'label': 'Audited Flag', 'purpose': 'Supports audited-accounts reporting and controls.'},
        {'name': 'staff_kenyan_current', 'label': 'Staff Totals', 'purpose': 'Supports employment analysis in the sector report.'},
        {'name': 'volunteers_kenyan_current', 'label': 'Volunteer Totals', 'purpose': 'Supports volunteer and intern analysis in the sector report.'},
        {'name': 'collaboration', 'label': 'Collaborations And Networking', 'purpose': 'Feeds collaboration and networking analysis.'},
    ]

    heading_map = {
        'reporting_period_start': '2.2.1 Submission of Annual Reports',
        'pbo_name': '2.2.1 Submission of Annual Reports',
        'scope': '2.3.1 Sectors of Operation',
        'county': '2.3.4 Counties of Operation',
        'countries_of_operation': '2.3.4 Counties of Operation',
        'sector[]': '2.3.1 Sectors of Operation',
        'project_county[]': '2.3.6 Project Implementation per Counties',
        'project_beneficiaries_no[]': '2.3.6 Project Implementation per Counties',
        'project_spending_per_county[]': '2.3.3 Utilisation of Funds on Projects',
        'project_duration_years[]': '2.3.6 Project Implementation per Counties',
        'project_completion_status[]': '2.3.6 Project Implementation per Counties',
        'payment_description[]': '2.2.3 Funds Received',
        'payment_kenya[]': '2.2.3 Funds Received',
        'payment_other[]': '2.2.3 Funds Received',
        'assets_item[]': '2.2.3 Funds Received',
        'donor_amount[]': '2.2.3 Funds Received',
        'staff_*': '2.3.10 Employment',
        'staff_kenyan_current': '2.3.10 Employment',
        'volunteers_kenyan_current': '2.3.10 Employment',
        'collaboration': '2.3.11 Collaborations and Networking',
        'audited': '2.2.2 Audited Accounts',
    }

    disallowed_names = {
        (item.get('name') or '').strip().lower()
        for item in disabled_items
        if (item.get('name') or '').strip()
    }
    for item in sector_focus_items:
        state = (item.get('state') or '').strip().lower()
        name = (item.get('name') or '').strip().lower()
        if state == 'disabled' and name:
            disallowed_names.add(name)

    allowed_candidates = required_items + sector_focus_items + extra_allowed_fields
    allowed_rules = []
    seen_allowed = set()
    for item in allowed_candidates:
        name = (item.get('name') or item.get('label') or '').strip()
        label = (item.get('label') or item.get('group') or name).strip()
        if not label:
            continue
        key = ((name or label).lower(), label.lower())
        if key in seen_allowed:
            continue
        seen_allowed.add(key)
        if name.lower() in disallowed_names:
            continue
        allowed_rules.append({
            'name': name or label,
            'label': label,
            'heading': heading_map.get((name or '').lower()),
            'required': bool(item.get('required')),
            'used_in_sector_report': bool(item.get('used_in_sector_report', True)),
            'state': item.get('state') or 'enabled',
            'purpose': item.get('purpose') or '',
        })

    disallowed_rules = []
    seen_disallowed = set()
    for item in disabled_items:
        name = (item.get('name') or item.get('label') or '').strip()
        label = (item.get('label') or name).strip()
        if not label:
            continue
        key = ((name or label).lower(), label.lower())
        if key in seen_disallowed:
            continue
        seen_disallowed.add(key)
        disallowed_rules.append({
            'name': name or label,
            'label': label,
            'heading': heading_map.get((name or '').lower()),
            'required': False,
            'used_in_sector_report': False,
            'state': 'disabled',
            'purpose': item.get('purpose') or '',
        })

    enabled_labels = [item['label'] for item in required_items]
    allowed_labels = [item['label'] for item in allowed_rules]
    disabled_labels = [item['label'] for item in disallowed_rules]
    sector_focus_labels = [item['label'] for item in sector_focus_items]
    sector_report_headings = sorted({
        heading for heading in heading_map.values() if heading
    })

    metadata['sector_report_headings'] = sector_report_headings
    metadata['all_allowed_fields'] = allowed_rules
    metadata['all_disallowed_fields'] = disallowed_rules

    return {
        'metadata': metadata,
        'enabled_summary': enabled_labels,
        'allowed_summary': allowed_labels,
        'disabled_summary': disabled_labels,
        'disallowed_summary': disabled_labels,
        'sector_focus_summary': sector_focus_labels,
        'allowed_rules': allowed_rules,
        'disallowed_rules': disallowed_rules,
        'heading_summary': sector_report_headings,
    }


FIELD_HELP_SAMPLE_SIZE = 50
FIELD_HELP_TRAIN_SIZE = 30
FIELD_HELP_TEST_SIZE = 20
FIELD_HELP_INTENT_TRAIN_SIZE = 30
FIELD_HELP_INTENT_TEST_SIZE = 20
FIELD_HELP_INTENT_MODEL_KEY_SUFFIX = '__intent'
FIELD_HELP_INTENT_DATASET_FILE = Path(__file__).resolve().parent / 'field_help_intent_dataset.json'

FIELD_HELP_DISALLOWED_ACTION_TERMS = {
    'edit', 'change', 'modify', 'override', 'enable', 'unlock', 'force',
    'fill', 'input', 'type', 'write', 'update', 'submit', 'manually',
}
FIELD_HELP_FAQ_TERMS = {'why', 'how', 'what', 'which', 'explain', 'help'}
FIELD_HELP_GENERAL_TERMS = {'sector', 'report', 'county', 'project', 'employment', 'collaboration', 'network'}
FIELD_HELP_SECURITY_TERMS = {'script', 'injection', 'sql', 'xss', 'or 1=1', '<script', '--'}


def load_field_help_intent_dataset():
    try:
        payload = json.loads(FIELD_HELP_INTENT_DATASET_FILE.read_text(encoding='utf-8'))
    except Exception:
        payload = []
    valid_rows = []
    for row in payload if isinstance(payload, list) else []:
        text_value = (row.get('input') or '').strip()
        label_value = (row.get('label') or '').strip().upper()
        if not text_value or not label_value:
            continue
        valid_rows.append({'input': text_value, 'label': label_value})
    return valid_rows


def ensure_field_help_intent_samples(page_key):
    normalized_key = normalize_field_help_page_key(page_key)
    existing = (
        FieldHelpIntentSample.query
        .filter_by(page_key=normalized_key)
        .order_by(FieldHelpIntentSample.id.asc())
        .all()
    )
    if existing:
        existing_labels = {row.label for row in existing}
        if len(existing) >= FIELD_HELP_SAMPLE_SIZE and len(existing_labels) >= 4:
            return existing
        for row in existing:
            db.session.delete(row)
        db.session.commit()

    dataset = load_field_help_intent_dataset()
    if not dataset:
        dataset = [{'input': 'Enter email user@example.com', 'label': 'VALID_INPUT'}]
    rng = random.Random(14014)
    rng.shuffle(dataset)

    train_cutoff = min(FIELD_HELP_INTENT_TRAIN_SIZE, len(dataset))
    for index, row in enumerate(dataset):
        dataset_split = 'train' if index < train_cutoff else 'test'
        db.session.add(FieldHelpIntentSample(
            page_key=normalized_key,
            input_text=row['input'],
            label=row['label'],
            dataset_split=dataset_split,
        ))
    db.session.commit()
    return (
        FieldHelpIntentSample.query
        .filter_by(page_key=normalized_key)
        .order_by(FieldHelpIntentSample.id.asc())
        .all()
    )


def ensure_field_help_intent_model(page_key):
    normalized_key = normalize_field_help_page_key(page_key)
    model_key = f"{normalized_key}{FIELD_HELP_INTENT_MODEL_KEY_SUFFIX}"
    row = FieldHelpDecisionModel.query.filter_by(page_key=model_key).first()
    if row:
        try:
            payload = json.loads(row.model_json or '{}')
            if payload.get('labels') and len(payload.get('labels') or []) >= 4:
                return payload, row
        except Exception:
            pass

    samples = ensure_field_help_intent_samples(normalized_key)
    train_set = []
    test_set = []
    for sample in samples:
        item = {
            'question': sample.input_text,
            'category': sample.label,
        }
        if sample.dataset_split == 'test':
            test_set.append(item)
        else:
            train_set.append(item)

    if not train_set:
        train_set = [{'question': 'enter email user@example.com', 'category': 'VALID_INPUT'}]
    if not test_set:
        test_set = train_set[:1]

    model_payload = train_field_help_decision_model(train_set)
    correct = 0
    for sample in test_set:
        predicted_label, _ = predict_field_help_category(sample['question'], model_payload)
        if predicted_label == sample['category']:
            correct += 1
    accuracy = correct / len(test_set) if test_set else 1.0

    if row is None:
        row = FieldHelpDecisionModel(
            page_key=model_key,
            model_json=json.dumps(model_payload, ensure_ascii=True),
            train_size=len(train_set),
            test_size=len(test_set),
            accuracy=accuracy,
        )
        db.session.add(row)
    else:
        row.model_json = json.dumps(model_payload, ensure_ascii=True)
        row.train_size = len(train_set)
        row.test_size = len(test_set)
        row.accuracy = accuracy
    db.session.commit()
    return model_payload, row


def classify_field_help_route(question, predicted_label):
    lowered = (question or '').lower()
    tokens = set(tokenize_field_help_text(lowered))
    if any(term in lowered for term in FIELD_HELP_SECURITY_TERMS):
        return 'FIELD_VALIDATION'
    if tokens.intersection(FIELD_HELP_FAQ_TERMS):
        return 'FAQ'
    if predicted_label in {'VALID_INPUT', 'INVALID_INPUT', 'DISALLOWED_ACTION', 'EDGE_CASE', 'SECURITY_THREAT'}:
        return 'FIELD_VALIDATION'
    if tokens.intersection(FIELD_HELP_GENERAL_TERMS):
        return 'GENERAL_QUERY'
    return 'UNCERTAIN'


def detect_disallowed_field_action(question, field_help_context):
    lowered = (question or '').lower()
    token_set = set(tokenize_field_help_text(lowered))
    if not token_set.intersection(FIELD_HELP_DISALLOWED_ACTION_TERMS):
        return None
    for rule in field_help_context.get('disallowed_rules') or []:
        label = (rule.get('label') or '').strip().lower()
        name = (rule.get('name') or '').strip().lower()
        if not label and not name:
            continue
        if label and label in lowered:
            return rule
        if name and name in lowered:
            return rule
    return None


def detect_validation_issues(question):
    lowered = (question or '').lower()
    issues = []

    if '<script' in lowered or 'or 1=1' in lowered or '--' in lowered:
        issues.append('SECURITY_THREAT')

    email_match = re.search(r'email[^a-z0-9]*([a-z0-9._%+\-]+@[a-z0-9.\-]+)', lowered)
    if email_match:
        email_value = email_match.group(1)
        if not re.match(r'^[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}$', email_value):
            issues.append('INVALID_EMAIL')

    phone_match = re.search(r'phone[^0-9+]*([+0-9a-z]+)', lowered)
    if phone_match:
        phone_value = phone_match.group(1)
        if re.search(r'[a-z]', phone_value):
            issues.append('INVALID_PHONE')
        digits = re.sub(r'[^0-9]', '', phone_value)
        if len(digits) < 9:
            issues.append('INVALID_PHONE')

    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', lowered)
    if date_match:
        try:
            parsed = datetime.strptime(date_match.group(1), '%Y-%m-%d').date()
            if parsed > date.today():
                issues.append('FUTURE_DATE')
        except ValueError:
            issues.append('INVALID_DATE')

    kra_match = re.search(r'\b([a-z]\d{9}[a-z])\b', lowered, flags=re.IGNORECASE)
    if 'kra' in lowered and kra_match is None:
        issues.append('INVALID_KRA_PIN')

    return sorted(set(issues))


def find_field_rule_match(question, field_help_context, allowed=True):
    lowered = (question or '').lower()
    candidates = field_help_context.get('allowed_rules' if allowed else 'disallowed_rules') or []
    for rule in candidates:
        label = (rule.get('label') or '').lower()
        name = (rule.get('name') or '').lower()
        if (label and label in lowered) or (name and name in lowered):
            return rule
    return None


def search_field_help_knowledge(question, field_help_context, limit=2):
    question_tokens = set(tokenize_field_help_text(question))
    if not question_tokens:
        return []
    ranked = []
    for record in build_field_help_search_records(field_help_context):
        text_tokens = set(tokenize_field_help_text(f"{record.get('label', '')} {record.get('text', '')}"))
        overlap = len(question_tokens.intersection(text_tokens))
        if overlap <= 0:
            continue
        ranked.append((overlap, record))
    ranked.sort(key=lambda item: item[0], reverse=True)
    return [row for _, row in ranked[:limit]]


def store_field_help_interaction(page_key, question, answer, intent=None, route=None, response_source=None, confidence=None, blocked=False):
    normalized_key = normalize_field_help_page_key(page_key)
    db.session.add(FieldHelpInteraction(
        page_key=normalized_key,
        question=(question or '').strip(),
        normalized_question=(question or '').strip().lower(),
        intent=(intent or '')[:80] or None,
        route=(route or '')[:80] or None,
        response_source=(response_source or '')[:80] or None,
        answer=(answer or '').strip(),
        confidence=confidence,
        blocked=bool(blocked),
    ))
    db.session.commit()


def normalize_field_help_page_key(page_key):
    normalized = re.sub(r'[^a-z0-9_.-]+', '_', str(page_key or 'global').strip().lower())
    normalized = normalized.strip('_')
    return (normalized or 'global')[:120]


def tokenize_field_help_text(text):
    return [token for token in re.findall(r'[a-z0-9]+', (text or '').lower()) if len(token) > 1]


def build_field_help_search_records(field_help_context):
    metadata = field_help_context.get('metadata') or {}
    records = []

    def append_items(items, category, label_key='label'):
        for item in items or []:
            label = (item.get(label_key) or '').strip()
            if not label:
                continue
            purpose = (item.get('purpose') or '').strip()
            trigger = (item.get('trigger') or '').strip()
            reason = (item.get('reason') or '').strip()
            details = " ".join(part for part in [purpose, trigger, reason] if part)
            records.append({
                'category': category,
                'label': label,
                'text': details or f"{label} field guidance",
            })

    append_items(metadata.get('required_fields'), 'required')
    append_items(field_help_context.get('allowed_rules'), 'allowed')
    append_items(field_help_context.get('disallowed_rules'), 'disabled')
    append_items(metadata.get('sector_report_focus_fields'), 'sector_focus')
    append_items(metadata.get('conditional_fields'), 'conditional', label_key='group')

    for label in field_help_context.get('heading_summary') or []:
        records.append({
            'category': 'sector_heading',
            'label': label,
            'text': f"Sector report heading from Annual NGO Sector Report 2019/2020: {label}",
        })

    for index, note in enumerate(metadata.get('notes') or []):
        note_text = (note or '').strip()
        if not note_text:
            continue
        records.append({
            'category': 'note',
            'label': f'Note {index + 1}',
            'text': note_text,
        })

    return records


def persist_field_help_context(page_key, field_help_context):
    normalized_key = normalize_field_help_page_key(page_key)
    records = build_field_help_search_records(field_help_context)
    context_json = json.dumps(field_help_context, ensure_ascii=True)
    records_json = json.dumps(records, ensure_ascii=True)
    row = FieldHelpCache.query.filter_by(page_key=normalized_key).first()
    if row is None:
        row = FieldHelpCache(
            page_key=normalized_key,
            context_json=context_json,
            search_index_json=records_json,
        )
        db.session.add(row)
    else:
        if row.context_json != context_json or row.search_index_json != records_json:
            row.context_json = context_json
            row.search_index_json = records_json
    db.session.commit()

    sync_field_help_rules(normalized_key, field_help_context)


def sync_field_help_rules(page_key, field_help_context):
    """Seed missing field-help rules from code defaults without overriding admin edits."""
    normalized_key = normalize_field_help_page_key(page_key)
    allowed_rules = field_help_context.get('allowed_rules') or []
    disallowed_rules = field_help_context.get('disallowed_rules') or []
    required_labels = set(field_help_context.get('enabled_summary') or [])

    desired_rows = {}

    def ingest(rule, is_allowed, is_disabled):
        field_name = (rule.get('name') or rule.get('label') or '').strip()[:160]
        field_label = (rule.get('label') or field_name).strip()[:255]
        if not field_name or not field_label:
            return
        key = field_name.lower()
        desired_rows[key] = {
            'field_name': field_name,
            'field_label': field_label,
            'heading': (rule.get('heading') or '')[:255] or None,
            'is_allowed': bool(is_allowed),
            'is_disabled': bool(is_disabled),
            'is_required': field_label in required_labels,
            'used_in_sector_report': bool(rule.get('used_in_sector_report', is_allowed)),
            'notes': (rule.get('purpose') or '')[:4000] or None,
        }

    for rule in allowed_rules:
        ingest(rule, True, False)
    for rule in disallowed_rules:
        ingest(rule, False, True)

    existing_keys = {
        row.field_name.lower()
        for row in FieldHelpRule.query.filter_by(page_key=normalized_key).all()
    }

    changed = False
    for key, payload in desired_rows.items():
        if key in existing_keys:
            continue
        db.session.add(FieldHelpRule(page_key=normalized_key, **payload))
        changed = True

    if changed:
        db.session.commit()


def load_persisted_field_help_context(page_key):
    normalized_key = normalize_field_help_page_key(page_key)
    row = FieldHelpCache.query.filter_by(page_key=normalized_key).first()
    if row is None:
        return None
    try:
        return json.loads(row.context_json or '{}')
    except Exception:
        return None


def apply_db_field_help_rules(page_key, field_help_context):
    normalized_key = normalize_field_help_page_key(page_key)
    candidate_keys = [normalized_key]
    if normalized_key != 'report_edit':
        candidate_keys.append('report_edit')
    if normalized_key != 'global':
        candidate_keys.append('global')

    rows = (
        FieldHelpRule.query
        .filter(FieldHelpRule.page_key.in_(candidate_keys))
        .order_by(FieldHelpRule.updated_at.desc(), FieldHelpRule.id.desc())
        .all()
    )
    if not rows:
        return field_help_context

    priority = {key: idx for idx, key in enumerate(candidate_keys)}
    rows.sort(key=lambda item: (priority.get(item.page_key, 99), item.id))

    selected = {}
    for row in rows:
        key = (row.field_name or '').strip().lower()
        if not key or key in selected:
            continue
        selected[key] = row

    allowed_rules = []
    disallowed_rules = []
    required_labels = []
    heading_summary = set()

    for row in selected.values():
        rule_payload = {
            'name': row.field_name,
            'label': row.field_label,
            'heading': row.heading,
            'required': bool(row.is_required),
            'used_in_sector_report': bool(row.used_in_sector_report),
            'state': 'disabled' if row.is_disabled else 'enabled',
            'purpose': row.notes or '',
        }
        if row.is_required:
            required_labels.append(row.field_label)
        if row.heading:
            heading_summary.add(row.heading)
        if row.is_allowed and not row.is_disabled:
            allowed_rules.append(rule_payload)
        if row.is_disabled:
            disallowed_rules.append(rule_payload)

    if not allowed_rules and not disallowed_rules:
        return field_help_context

    context = dict(field_help_context or {})
    metadata = dict(context.get('metadata') or {})

    context['allowed_rules'] = allowed_rules
    context['disallowed_rules'] = disallowed_rules
    context['allowed_summary'] = [rule['label'] for rule in allowed_rules]
    context['disallowed_summary'] = [rule['label'] for rule in disallowed_rules]
    context['disabled_summary'] = context['disallowed_summary']
    context['enabled_summary'] = required_labels or context.get('enabled_summary') or []
    context['heading_summary'] = sorted(heading_summary)

    metadata['all_allowed_fields'] = allowed_rules
    metadata['all_disallowed_fields'] = disallowed_rules
    metadata['sector_report_headings'] = context['heading_summary']
    context['metadata'] = metadata
    return context


def build_field_help_samples(search_records, page_key):
    templates = {
        'required': [
            "is {label} required",
            "what is required about {label}",
            "why do i need {label}",
        ],
        'disabled': [
            "why is {label} disabled",
            "can i edit {label}",
            "what is the rule for {label}",
        ],
        'allowed': [
            "is {label} allowed",
            "can i fill {label}",
            "how is {label} used",
        ],
        'sector_focus': [
            "does {label} affect sector report",
            "how is {label} used in sector reporting",
            "is {label} part of sector report",
        ],
        'sector_heading': [
            "what does {label} cover",
            "which fields map to {label}",
            "explain heading {label}",
        ],
        'conditional': [
            "when does {label} apply",
            "why is {label} conditional",
            "what triggers {label}",
        ],
        'note': [
            "explain {label}",
            "help me understand {label}",
            "what should i know about {label}",
        ],
    }
    samples = []
    for record in search_records:
        category = record.get('category') or 'note'
        label = (record.get('label') or 'field').lower()
        for template in templates.get(category, templates['note']):
            samples.append({
                'question': template.format(label=label),
                'category': category,
            })
        samples.append({
            'question': f"{label} {record.get('text', '')}".strip().lower(),
            'category': category,
        })

    if not samples:
        samples = [{'question': 'what fields are required', 'category': 'required'}]

    rng = random.Random(1000 + sum(ord(ch) for ch in normalize_field_help_page_key(page_key)))
    rng.shuffle(samples)
    while len(samples) < FIELD_HELP_SAMPLE_SIZE:
        samples.append(rng.choice(samples))
    return samples[:FIELD_HELP_SAMPLE_SIZE]


def train_field_help_decision_model(samples):
    label_counts = Counter()
    token_counts = {}
    for sample in samples:
        label = sample.get('category') or 'note'
        label_counts[label] += 1
        counter = token_counts.setdefault(label, Counter())
        for token in tokenize_field_help_text(sample.get('question')):
            counter[token] += 1

    vocab = set()
    token_totals = {}
    for label, counter in token_counts.items():
        vocab.update(counter.keys())
        token_totals[label] = sum(counter.values())

    return {
        'labels': sorted(label_counts.keys()),
        'label_counts': dict(label_counts),
        'token_counts': {label: dict(counter) for label, counter in token_counts.items()},
        'token_totals': token_totals,
        'vocab_size': max(len(vocab), 1),
    }


def predict_field_help_category(question, model_payload):
    labels = model_payload.get('labels') or []
    if not labels:
        return 'note', 0.0

    label_counts = model_payload.get('label_counts') or {}
    token_counts = model_payload.get('token_counts') or {}
    token_totals = model_payload.get('token_totals') or {}
    vocab_size = int(model_payload.get('vocab_size') or 1)
    total_docs = max(sum(label_counts.values()), 1)
    question_tokens = tokenize_field_help_text(question)
    if not question_tokens:
        return labels[0], 0.0

    best_label = labels[0]
    best_score = -1e9
    for label in labels:
        label_doc_count = label_counts.get(label, 0)
        prior = (label_doc_count + 1) / (total_docs + len(labels))
        score = math.log(prior)
        label_token_counts = token_counts.get(label, {})
        token_total = token_totals.get(label, 0)
        denom = token_total + vocab_size + 1
        for token in question_tokens:
            numerator = label_token_counts.get(token, 0) + 1
            score += math.log(numerator / denom)
        if score > best_score:
            best_label = label
            best_score = score

    return best_label, best_score


def ensure_field_help_decision_model(page_key, field_help_context):
    normalized_key = normalize_field_help_page_key(page_key)
    row = FieldHelpDecisionModel.query.filter_by(page_key=normalized_key).first()
    if row:
        try:
            payload = json.loads(row.model_json or '{}')
            if payload.get('labels'):
                return payload, row
        except Exception:
            pass

    records = build_field_help_search_records(field_help_context)
    samples = build_field_help_samples(records, normalized_key)
    train_set = samples[:FIELD_HELP_TRAIN_SIZE]
    test_set = samples[FIELD_HELP_TRAIN_SIZE:FIELD_HELP_TRAIN_SIZE + FIELD_HELP_TEST_SIZE]
    model_payload = train_field_help_decision_model(train_set)

    correct = 0
    for sample in test_set:
        predicted_label, _ = predict_field_help_category(sample.get('question'), model_payload)
        if predicted_label == sample.get('category'):
            correct += 1
    accuracy = (correct / len(test_set)) if test_set else 1.0

    if row is None:
        row = FieldHelpDecisionModel(
            page_key=normalized_key,
            model_json=json.dumps(model_payload, ensure_ascii=True),
            train_size=len(train_set),
            test_size=len(test_set),
            accuracy=accuracy,
        )
        db.session.add(row)
    else:
        row.model_json = json.dumps(model_payload, ensure_ascii=True)
        row.train_size = len(train_set)
        row.test_size = len(test_set)
        row.accuracy = accuracy
    db.session.commit()
    return model_payload, row


def lookup_field_help_memory(page_key, question):
    normalized_key = normalize_field_help_page_key(page_key)
    question_tokens = set(tokenize_field_help_text(question))
    if not question_tokens:
        return None

    entries = (
        FieldHelpMemory.query
        .filter(FieldHelpMemory.page_key.in_([normalized_key, 'report_edit', 'global']))
        .order_by(FieldHelpMemory.created_at.desc())
        .limit(250)
        .all()
    )
    best_entry = None
    best_score = 0.0
    for entry in entries:
        entry_tokens = set(tokenize_field_help_text(entry.question))
        if not entry_tokens:
            continue
        overlap = len(question_tokens.intersection(entry_tokens))
        union_size = len(question_tokens.union(entry_tokens)) or 1
        score = overlap / union_size
        if score > best_score:
            best_score = score
            best_entry = entry
    if best_entry and best_score >= 0.35:
        return {'answer': best_entry.answer, 'score': best_score}
    return None


def store_field_help_memory(page_key, question, answer, source, score=None):
    normalized_key = normalize_field_help_page_key(page_key)
    db.session.add(FieldHelpMemory(
        page_key=normalized_key,
        question=(question or '').strip(),
        answer=(answer or '').strip(),
        source=(source or 'local').strip()[:50],
        score=score,
    ))
    db.session.commit()


def compose_local_field_help_answer(question, field_help_context, page_key):
    normalized_question = (question or '').strip()
    lowered = normalized_question.lower()

    blocked_rule = detect_disallowed_field_action(normalized_question, field_help_context)
    if blocked_rule:
        answer = (
            f"`{blocked_rule.get('label')}` is blocked on this form because it is outside sector-report collection. "
            "Please use allowed fields only."
        )
        return answer, 'blocked-rule'

    intent_payload, intent_model_row = ensure_field_help_intent_model(page_key)
    predicted_intent, intent_score = predict_field_help_category(normalized_question, intent_payload)
    route = classify_field_help_route(normalized_question, predicted_intent)
    confidence = 1 / (1 + math.exp(-max(min(intent_score, 8), -8)))

    required_fields = field_help_context.get('enabled_summary') or []
    allowed_fields = field_help_context.get('allowed_summary') or []
    disallowed_fields = field_help_context.get('disallowed_summary') or []
    heading_summary = field_help_context.get('heading_summary') or []

    if 'all allowed' in lowered or 'allowed fields' in lowered or ('allowed' in lowered and 'field' in lowered):
        answer = (
            f"Allowed fields ({len(allowed_fields)}): {', '.join(allowed_fields)}. "
            "These are the fields used for sector-report analytics."
        )
        return answer, 'rules-allowed-list'

    if 'all disabled' in lowered or 'all disallowed' in lowered or 'disabled fields' in lowered:
        answer = (
            f"Disallowed/disabled fields ({len(disallowed_fields)}): {', '.join(disallowed_fields)}. "
            "They are intentionally out of scope for the sector report."
        )
        return answer, 'rules-disallowed-list'

    if route == 'FAQ':
        memory_match = lookup_field_help_memory(page_key, normalized_question)
        if memory_match:
            return memory_match['answer'], 'faq-cache'
        field_match = find_field_rule_match(normalized_question, field_help_context, allowed=True) or find_field_rule_match(normalized_question, field_help_context, allowed=False)
        if field_match:
            heading = field_match.get('heading')
            heading_text = f" Linked sector heading: {heading}." if heading else ''
            answer = f"{field_match.get('label')} guidance: {field_match.get('purpose') or 'Use this field according to form rules.'}{heading_text}"
            return answer, 'faq-rules'
        records = search_field_help_knowledge(normalized_question, field_help_context, limit=2)
        if records:
            answer = " ".join(f"{row.get('label')}: {row.get('text')}" for row in records)
            return answer, 'faq-knowledge'

    if route == 'FIELD_VALIDATION':
        issues = detect_validation_issues(normalized_question)
        disallowed_match = find_field_rule_match(normalized_question, field_help_context, allowed=False)
        if predicted_intent in {'DISALLOWED_ACTION', 'SECURITY_THREAT'} or disallowed_match:
            blocked_label = disallowed_match.get('label') if disallowed_match else 'this action'
            answer = (
                f"{blocked_label} is disallowed on this form. "
                "The bot can only assist with allowed sector-report fields."
            )
            return answer, 'validation-blocked'
        if issues:
            issues_text = ", ".join(issue.replace('_', ' ').lower() for issue in issues)
            answer = f"Validation warning: {issues_text}. Please correct the value and try again."
            return answer, 'validation-issues'
        allowed_match = find_field_rule_match(normalized_question, field_help_context, allowed=True)
        if allowed_match:
            heading = allowed_match.get('heading')
            heading_text = f" It maps to sector heading `{heading}`." if heading else ''
            answer = (
                f"`{allowed_match.get('label')}` is allowed. "
                f"{allowed_match.get('purpose') or 'You can provide this value for sector analytics.'}{heading_text}"
            )
            return answer, 'validation-rules'
        answer = (
            f"Use allowed fields ({len(allowed_fields)}) and avoid disabled fields ({len(disallowed_fields)}). "
            f"Required now: {', '.join(required_fields)}."
        )
        return answer, 'validation-general'

    if route == 'GENERAL_QUERY':
        records = search_field_help_knowledge(normalized_question, field_help_context, limit=3)
        if records:
            answer = " ".join(f"{row.get('label')}: {row.get('text')}" for row in records)
            return answer, 'knowledge-search'
        if heading_summary:
            answer = (
                "Sector-report headings available for mapping are "
                f"{', '.join(heading_summary)}."
            )
            return answer, 'knowledge-headings'

    memory_match = lookup_field_help_memory(page_key, normalized_question)
    if memory_match:
        return memory_match['answer'], 'fallback-cache'

    answer = (
        f"I am using local rules + cache + intent routing. Required: {', '.join(required_fields)}. "
        f"Allowed fields: {len(allowed_fields)}. Disabled fields: {len(disallowed_fields)}. "
        f"Intent model accuracy: {intent_model_row.accuracy:.0%} on {intent_model_row.test_size} test prompts."
    )
    return answer, 'fallback-local'


def get_field_help_context_for_page(page_key):
    normalized_key = normalize_field_help_page_key(page_key)
    cached = load_persisted_field_help_context(normalized_key)
    if cached and cached.get('allowed_summary') and cached.get('disallowed_summary'):
        sync_field_help_rules(normalized_key, cached)
        cached = apply_db_field_help_rules(normalized_key, cached)
        ensure_field_help_intent_samples(normalized_key)
        ensure_field_help_intent_model(normalized_key)
        return cached

    context = build_report_edit_field_help_context()
    metadata = context.get('metadata') or {}
    notes = list(metadata.get('notes') or [])
    if normalized_key not in {'report_edit', 'report_field_help'}:
        notes.append('This page is using cached field-help knowledge for local guidance.')
    metadata['notes'] = notes
    context['metadata'] = metadata
    persist_field_help_context(normalized_key, context)
    context = apply_db_field_help_rules(normalized_key, context)
    ensure_field_help_decision_model(normalized_key, context)
    ensure_field_help_intent_samples(normalized_key)
    ensure_field_help_intent_model(normalized_key)
    return context


def answer_field_help_question(question, field_help_context, page_key, prefer_openai=False):
    normalized_key = normalize_field_help_page_key(page_key)
    field_help_context = apply_db_field_help_rules(normalized_key, field_help_context)
    persist_field_help_context(normalized_key, field_help_context)
    ensure_field_help_decision_model(normalized_key, field_help_context)
    intent_payload, _ = ensure_field_help_intent_model(normalized_key)
    predicted_intent, intent_score = predict_field_help_category(question, intent_payload)
    route = classify_field_help_route(question, predicted_intent)

    answer, source = compose_local_field_help_answer(question, field_help_context, normalized_key)
    if prefer_openai and source.startswith('fallback') and get_openai_api_key():
        try:
            answer = call_openai_field_help_chat(question, field_help_context)
            source = 'openai-fallback'
        except Exception:
            app.logger.exception('OpenAI field-help request failed; using local responder')
    store_field_help_memory(normalized_key, question, answer, source)
    confidence = 1 / (1 + math.exp(-max(min(intent_score, 8), -8)))
    store_field_help_interaction(
        page_key=normalized_key,
        question=question,
        answer=answer,
        intent=predicted_intent,
        route=route,
        response_source=source,
        confidence=confidence,
        blocked=source in {'blocked-rule', 'validation-blocked'},
    )
    return answer, source


def call_openai_field_help_chat(question, field_help_context):
    api_key = get_openai_api_key()
    if not api_key:
        raise RuntimeError('OPENAI_API_KEY is not configured on the server.')

    system_prompt = (
        "You are a concise assistant for a Form 14 edit screen that supports sector reporting. "
        "Answer only about enabled fields, disabled fields, conditional field behavior, required fields, "
        "how fields are used for data collection, and whether a field matters for sector reporting. "
        "Do not invent fields that are not in the provided metadata. "
        "State clearly that only Start Date and PBO Name are required on this edit page. "
        "Treat sector reporting as focused on sectors, counties, project implementation rows, spending, beneficiaries, scope, and related financial summary fields. "
        "When a field is disabled, explain whether it is reference-only, out of scope for sector reporting, or still used indirectly through stored project data. "
        "Be brief, practical, and user-facing."
    )
    payload = {
        'model': 'gpt-5-mini',
        'input': [
            {
                'role': 'system',
                'content': [
                    {'type': 'input_text', 'text': system_prompt},
                ],
            },
            {
                'role': 'user',
                'content': [
                    {
                        'type': 'input_text',
                        'text': (
                            "Field metadata:\n"
                            f"{json.dumps(field_help_context['metadata'], indent=2)}\n\n"
                            f"User question: {question}"
                        ),
                    },
                ],
            },
        ],
        'max_output_tokens': 350,
    }
    response = requests.post(
        'https://api.openai.com/v1/responses',
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
        },
        json=payload,
        timeout=30,
    )
    response.raise_for_status()
    body = response.json()
    output_text = body.get('output_text')
    if output_text:
        return output_text.strip()

    parts = []
    for item in body.get('output', []):
        for content in item.get('content', []):
            text_value = content.get('text')
            if text_value:
                parts.append(text_value)
    answer = "\n".join(part.strip() for part in parts if part.strip()).strip()
    return answer or 'No response was returned from the assistant.'


DATA_ANALYSIS_DATASET_PATH = Path(__file__).resolve().parent / 'form14_data_analysis_training_dataset.json'
DATA_ANALYSIS_SUPPLEMENTAL_DATASET_PATHS = [
    Path(__file__).resolve().parent / 'form14_data_analysis_filename_questions.json',
    Path(__file__).resolve().parent / 'form14_data_analysis_live_questions.json',
]
DATA_ANALYSIS_KEYWORD_SCHEMA_PATHS = [
    Path(__file__).resolve().parent / 'form14_data_analysis_keyword_schema.json',
    Path(__file__).resolve().parent / 'form14_data_analysis_live_keyword_schema.json',
]
DATA_ANALYSIS_PAGE_KEY = 'report_edit'
DATA_ANALYSIS_RESULT_DIR = Path('/tmp/form14_analysis_bot_results')
DATA_ANALYSIS_MODEL_CACHE = {}
DATA_ANALYSIS_DATASET_CACHE = {
    'cache_key': None,
    'payload': None,
}
DATA_ANALYSIS_TRAIN_JOB = {
    'running': False,
    'started_at': None,
    'finished_at': None,
    'last_error': '',
    'last_result': {},
}
DATA_ANALYSIS_TRAIN_LOCK = threading.Lock()
DATA_ANALYSIS_BACKGROUND_BOOT_STARTED = False
DATA_ANALYSIS_BACKGROUND_BOOT_LOCK = threading.Lock()
DATA_ANALYSIS_INTERNAL_SYNC_CACHE = {
    'url': None,
    'ran_at': 0.0,
    'result': None,
}
DATA_ANALYSIS_EXTERNAL_CACHE_DIR = Path(__file__).resolve().parent / 'backups' / 'data_analysis_json_cache'
DATA_ANALYSIS_MAX_QUESTIONS_DEFAULT = max(int(os.getenv('DATA_ANALYSIS_MAX_QUESTIONS', '120000') or 120000), 10000)
DATA_ANALYSIS_EXTERNAL_JSON_TARGETS = [
    {
        'form_key': 'training_dataset_url',
        'file_key': 'training_dataset_file',
        'label': 'Training Dataset JSON',
        'path': DATA_ANALYSIS_DATASET_PATH,
        'setting_key': 'data_analysis_external_training_dataset_url',
    },
    {
        'form_key': 'filename_questions_url',
        'file_key': 'filename_questions_file',
        'label': 'Filename Questions JSON',
        'path': DATA_ANALYSIS_SUPPLEMENTAL_DATASET_PATHS[0],
        'setting_key': 'data_analysis_external_filename_questions_url',
    },
    {
        'form_key': 'live_questions_url',
        'file_key': 'live_questions_file',
        'label': 'Live Questions JSON',
        'path': DATA_ANALYSIS_SUPPLEMENTAL_DATASET_PATHS[1],
        'setting_key': 'data_analysis_external_live_questions_url',
    },
]


def normalize_data_analysis_name(value):
    text = (value or '').strip().lower()
    text = re.sub(r'[^a-z0-9]+', '_', text)
    return re.sub(r'_+', '_', text).strip('_')


def title_from_normalized(value):
    text = (value or '').strip().replace('_', ' ')
    return re.sub(r'\s+', ' ', text).strip().title()


def build_data_analysis_aliases(name, columns):
    aliases = {name, name.replace('_', ' ')}
    if name.startswith('pbo_'):
        aliases.add(name[4:])
        aliases.add(name[4:].replace('_', ' '))
    for column in columns:
        normalized_column = normalize_data_analysis_name(column)
        if not normalized_column:
            continue
        for alias in expand_data_analysis_column_aliases(normalized_column):
            aliases.add(alias)
    return sorted(alias for alias in aliases if alias)


def data_analysis_column_semantic_tags(column_name):
    normalized = normalize_data_analysis_name(column_name)
    tags = set()
    if any(token in normalized for token in ('county', 'counties', 'country', 'location', 'region', 'scope')):
        tags.add('location')
    if any(token in normalized for token in ('name', 'organisation', 'organization', 'pbo', 'ngo')):
        tags.add('entity')
    if any(token in normalized for token in ('amount', 'value', 'total', 'spending', 'cost', 'budget', 'fee')):
        tags.add('finance')
    if any(token in normalized for token in ('date', 'period', 'year', 'month', 'created', 'updated')):
        tags.add('time')
    if any(token in normalized for token in ('status', 'state', 'flag', 'workflow', 'review')):
        tags.add('status')
    if any(token in normalized for token in ('count', 'number', 'qty', 'quantity')):
        tags.add('count')
    if any(token in normalized for token in ('email', 'phone', 'telephone', 'contact', 'address')):
        tags.add('contact')
    if any(token in normalized for token in ('risk', 'score', 'compliance', 'validation', 'duplicate')):
        tags.add('risk')
    if not tags:
        tags.add('general')
    return sorted(tags)


def expand_data_analysis_column_aliases(column_name):
    normalized = normalize_data_analysis_name(column_name)
    if not normalized:
        return []
    aliases = {
        normalized,
        normalized.replace('_', ' '),
    }
    if 'county' in normalized or 'counties' in normalized:
        aliases.update({'county', 'counties', 'location', 'geography', 'region'})
    if any(token in normalized for token in ('organisation', 'organization', 'pbo', 'ngo', 'name')):
        aliases.update({'organization', 'organizations', 'organisation', 'organisations', 'entity', 'entities'})
    if any(token in normalized for token in ('amount', 'value', 'spending', 'total', 'cost', 'budget')):
        aliases.update({'amount', 'total', 'value', 'financial metric', 'money'})
    if any(token in normalized for token in ('count', 'number', 'qty', 'quantity')):
        aliases.update({'count', 'counts', 'number', 'totals'})
    if any(token in normalized for token in ('date', 'period', 'year', 'month', 'created', 'updated')):
        aliases.update({'date', 'time', 'period', 'timeline'})
    if any(token in normalized for token in ('status', 'state', 'flag', 'workflow', 'review')):
        aliases.update({'status', 'state', 'workflow'})
    if any(token in normalized for token in ('email', 'phone', 'telephone', 'contact')):
        aliases.update({'contact', 'communication', 'email', 'phone'})
    if any(token in normalized for token in ('risk', 'score', 'compliance', 'validation', 'duplicate')):
        aliases.update({'risk', 'score', 'compliance', 'quality'})
    return sorted(alias for alias in aliases if alias)


def infer_data_analysis_intent_for_column(column_name, variant_index):
    tags = set(data_analysis_column_semantic_tags(column_name))
    if 'location' in tags:
        return ['distribution', 'aggregation', 'ranking', 'compare_groups'][variant_index % 4]
    if 'finance' in tags:
        return ['aggregation', 'ranking', 'trend_analysis', 'anomaly_detection'][variant_index % 4]
    if 'time' in tags:
        return ['trend_analysis', 'distribution', 'validation_check', 'aggregation'][variant_index % 4]
    if 'status' in tags or 'risk' in tags:
        return ['validation_check', 'distribution', 'anomaly_detection', 'compare_groups'][variant_index % 4]
    if 'count' in tags:
        return ['aggregation', 'ranking', 'compare_groups', 'trend_analysis'][variant_index % 4]
    if 'entity' in tags:
        return ['distribution', 'ranking', 'aggregation', 'validation_check'][variant_index % 4]
    return ['aggregation', 'distribution', 'validation_check', 'compare_groups'][variant_index % 4]


def build_data_analysis_column_query_variants(source_name, column_name, companion_columns, max_queries=1100):
    source_pretty = title_from_normalized(source_name)
    column_pretty = title_from_normalized(column_name)
    companion = title_from_normalized(companion_columns[0]) if companion_columns else 'Related Fields'
    companion_two = title_from_normalized(companion_columns[1]) if len(companion_columns) > 1 else companion
    scope_contexts = [
        'across all records',
        'for the current reporting periods',
        'for trend-ready analysis',
        'for executive dashboard use',
        'for validation and quality checks',
    ]
    row_views = [
        'at row level',
        'using row-level interactions',
        'with row and column cross-checks',
        'using row-level filters for non-empty values',
    ]
    analysis_lenses = [
        'with a count summary',
        'with mean or average where numeric',
        'with grouped rankings',
        'with anomaly checks',
        'with completeness checks',
        'with top-vs-bottom comparisons',
    ]
    prefixes = [
        'Show', 'List', 'Give me', 'Break down', 'Analyze', 'Summarize', 'Explain', 'Compare', 'Rank',
        'Tell me', 'Compute', 'Evaluate', 'Review', 'Profile',
    ]
    suffixes = [
        'for decision-making.',
        'for policy reporting.',
        'for a regulator-friendly dashboard.',
        'for executive summary use.',
        'with clear statistical interpretation.',
        'and highlight key anomalies.',
        'and include confidence checks.',
        'and make the output presentation-ready.',
    ]
    rows = []
    idx = 0
    for prefix in prefixes:
        for context in scope_contexts:
            for row_view in row_views:
                for lens in analysis_lenses:
                    for suffix in suffixes:
                        intent = infer_data_analysis_intent_for_column(column_name, idx)
                        question = (
                            f"{prefix} {column_pretty} in {source_pretty} {context}, "
                            f"{row_view}, correlating with {companion} and {companion_two}, {lens}, {suffix}"
                        )
                        rows.append({
                            'question': re.sub(r'\s+', ' ', question).strip(),
                            'intent': intent,
                            'fields_hint': [normalize_data_analysis_name(column_name)] + [normalize_data_analysis_name(col) for col in companion_columns[:3]],
                            'answer_style': 'brief_analytic',
                        })
                        idx += 1
                        if len(rows) >= max_queries:
                            return rows
    return rows[:max_queries]


def build_data_analysis_questions_for_source(source_name, normalized_columns, start_id=1):
    columns = [normalize_data_analysis_name(col) for col in (normalized_columns or []) if normalize_data_analysis_name(col)]
    top_columns = columns[:6]
    metric_column = next(
        (col for col in columns if any(token in col for token in ('amount', 'total', 'count', 'value', 'number', 'spending', 'beneficiaries', 'score'))),
        columns[0] if columns else 'id',
    )
    group_column = next(
        (col for col in columns if any(token in col for token in ('name', 'sector', 'county', 'category', 'scope', 'role', 'country', 'partner', 'training', 'currency'))),
        columns[0] if columns else 'id',
    )
    rows = []
    next_id = start_id
    base_prompts = [
        (f"Show summary statistics for {title_from_normalized(source_name)} using key columns.", 'aggregation'),
        (f"List top {group_column.replace('_', ' ')} by {metric_column.replace('_', ' ')} in {title_from_normalized(source_name)}.", 'ranking'),
        (f"Show distribution patterns for {group_column.replace('_', ' ')} in {title_from_normalized(source_name)}.", 'distribution'),
        (f"Detect anomalies in {title_from_normalized(source_name)} using {metric_column.replace('_', ' ')}.", 'anomaly_detection'),
        (f"Validate missing and inconsistent values in {title_from_normalized(source_name)}.", 'validation_check'),
        (f"Run trend analysis on {title_from_normalized(source_name)} over available periods.", 'trend_analysis'),
    ]
    for question_text, intent in base_prompts:
        rows.append({
            'id': next_id,
            'question': question_text,
            'intent': intent,
            'target_domain': source_name,
            'fields_hint': top_columns,
            'answer_style': 'brief_analytic',
        })
        next_id += 1

    for column_name in columns:
        companions = [col for col in columns if col != column_name]
        variants = build_data_analysis_column_query_variants(
            source_name=source_name,
            column_name=column_name,
            companion_columns=companions,
            max_queries=1100,
        )
        for variant in variants:
            rows.append({
                'id': next_id,
                'question': variant['question'],
                'intent': variant['intent'],
                'target_domain': source_name,
                'fields_hint': variant['fields_hint'],
                'answer_style': variant['answer_style'],
            })
            next_id += 1
    return rows


def build_global_county_source_row():
    columns = [
        'county',
        'counties',
        'project_county',
        'project_row_counties',
        'counties_of_operation',
        'county_count',
    ]
    normalized_columns = [normalize_data_analysis_name(column) for column in columns]
    return {
        'filename': 'all_database_county_scan',
        'normalized_name': GLOBAL_COUNTY_SCAN_DOMAIN,
        'sheet_name': 'database_virtual',
        'aliases': [
            'all county sources',
            'all counties',
            'cross table county scan',
            'county',
            'counties',
            'county distribution',
            'organizations by county',
            'organizations per county',
        ],
        'columns': columns,
        'normalized_columns': normalized_columns,
    }


def build_global_county_questions(start_id=1):
    prompts = [
        ('How many organizations are in Nairobi City County across all tables with county fields?', 'aggregation'),
        ('Count organizations per county using all database tables with county or counties columns.', 'distribution'),
        ('Show county distribution across the whole database, not one table.', 'distribution'),
        ('List top counties by organization count across all county/counties fields.', 'ranking'),
        ('Compare Nairobi County and Mombasa County using whole-database county data.', 'compare_groups'),
        ('Show every table that contains county or counties and matched row counts.', 'validation_check'),
    ]
    county_names = sorted(format_county_display(name) for name in COUNTY_NAME_HINTS)
    location_templates = [
        ("How many organizations are in {county}?", 'aggregation'),
        ("Count organizations located in {county}.", 'aggregation'),
        ("Number of NGOs operating in {county}.", 'aggregation'),
        ("Total PBO entities in {county}.", 'aggregation'),
        ("How many organization names match {county} county records?", 'aggregation'),
        ("Give organization count for {county} across all tables.", 'aggregation'),
        ("How many submissions mention {county} in county columns?", 'aggregation'),
        ("Find entities linked to {county} using whole database search.", 'aggregation'),
        ("Show county distribution and highlight {county}.", 'distribution'),
        ("Rank counties and show where {county} appears.", 'ranking'),
        ("Compare {county} with Nairobi City County on organization count.", 'compare_groups'),
        ("Validate county-related rows for {county} in all sources.", 'validation_check'),
    ]
    generated = []
    for county in county_names:
        for template, intent in location_templates:
            generated.append((template.format(county=county), intent))
            if len(generated) >= 300:
                break
        if len(generated) >= 300:
            break
    prompts.extend(generated)
    rows = []
    next_id = start_id
    for question_text, intent in prompts:
        rows.append({
            'id': next_id,
            'question': question_text,
            'intent': intent,
            'target_domain': GLOBAL_COUNTY_SCAN_DOMAIN,
            'fields_hint': ['county', 'counties', 'project_county', 'counties_of_operation'],
            'answer_style': 'brief_analytic',
        })
        next_id += 1
    return rows


def sync_data_analysis_live_keyword_schema_from_internal_db():
    internal_url = normalize_database_url(os.getenv('INTERNAL_DATABASE_URL'))
    if not internal_url:
        return {'updated': False, 'reason': 'INTERNAL_DATABASE_URL not set.'}

    try:
        engine = create_engine(internal_url)
        inspector = inspect(engine)
        available_tables = set(inspector.get_table_names())
        target_tables = sorted(
            table_name
            for table_name in available_tables
            if table_name and table_name != 'alembic_version'
        )
        sources = []
        question_rows = []
        next_id = 1
        for table_name in target_tables:
            if table_name not in available_tables:
                continue
            columns = [str(column.get('name') or '').strip() for column in inspector.get_columns(table_name)]
            columns = [column for column in columns if column]
            normalized_columns = [normalize_data_analysis_name(column) for column in columns if normalize_data_analysis_name(column)]
            source_name = normalize_data_analysis_name(table_name)
            sources.append({
                'filename': table_name,
                'normalized_name': source_name,
                'sheet_name': 'database',
                'aliases': build_data_analysis_aliases(source_name, columns),
                'columns': columns,
                'normalized_columns': normalized_columns,
            })
            generated = build_data_analysis_questions_for_source(source_name, normalized_columns, start_id=next_id)
            question_rows.extend(generated)
            next_id += len(generated)

        # Add a synthetic global county/cross-table source so routing can use all matching tables.
        sources.append(build_global_county_source_row())
        global_county_questions = build_global_county_questions(start_id=next_id)
        question_rows.extend(global_county_questions)
        next_id += len(global_county_questions)
    except Exception as exc:
        return {'updated': False, 'reason': str(exc)}
    finally:
        try:
            engine.dispose()
        except Exception:
            pass

    if not sources:
        return {
            'updated': False,
            'reason': 'No target tables were found in INTERNAL_DATABASE_URL.',
            'sources': 0,
            'questions': 0,
        }

    keyword_schema_payload = {
        'version': 1,
        'description': 'Live keyword schema introspected from INTERNAL_DATABASE_URL.',
        'sources': sources,
        'global_keywords': sorted({
            keyword
            for source in sources
            for keyword in (source.get('normalized_columns') or [])
            if keyword
        }),
    }
    DATA_ANALYSIS_KEYWORD_SCHEMA_PATHS[1].write_text(
        json.dumps(keyword_schema_payload, indent=2, ensure_ascii=True) + '\n',
        encoding='utf-8',
    )

    supplemental_payload = {
        'dataset_name': 'form14_live_database_keyword_questions',
        'purpose': 'Supplemental prompts generated from INTERNAL_DATABASE_URL schema introspection.',
        'questions': question_rows,
    }
    DATA_ANALYSIS_SUPPLEMENTAL_DATASET_PATHS[1].write_text(
        json.dumps(supplemental_payload, indent=2, ensure_ascii=True) + '\n',
        encoding='utf-8',
    )
    return {'updated': True, 'sources': len(sources), 'questions': len(question_rows)}


def auto_sync_data_analysis_from_internal_db(force=False):
    internal_url = normalize_database_url(os.getenv('INTERNAL_DATABASE_URL'))
    if not internal_url:
        return {'checked': False, 'updated': False, 'reason': 'INTERNAL_DATABASE_URL not set.'}

    cache_ttl = max(int(os.getenv('DATA_ANALYSIS_INTERNAL_SYNC_TTL_SECONDS', '300') or 300), 30)
    now_ts = time.time()
    last_url = DATA_ANALYSIS_INTERNAL_SYNC_CACHE.get('url')
    last_ran_at = float(DATA_ANALYSIS_INTERNAL_SYNC_CACHE.get('ran_at') or 0.0)
    cached_result = DATA_ANALYSIS_INTERNAL_SYNC_CACHE.get('result')
    cache_is_fresh = (
        (not force)
        and last_url == internal_url
        and cached_result is not None
        and (now_ts - last_ran_at) < cache_ttl
    )
    if cache_is_fresh:
        return {
            **cached_result,
            'checked': True,
            'cached': True,
            'ttl_seconds': cache_ttl,
        }

    result = sync_data_analysis_live_keyword_schema_from_internal_db()
    DATA_ANALYSIS_INTERNAL_SYNC_CACHE['url'] = internal_url
    DATA_ANALYSIS_INTERNAL_SYNC_CACHE['ran_at'] = now_ts
    DATA_ANALYSIS_INTERNAL_SYNC_CACHE['result'] = result
    return {
        **result,
        'checked': True,
        'cached': False,
        'ttl_seconds': cache_ttl,
    }


def _normalize_external_json_url(raw_url):
    url = (raw_url or '').strip()
    if not url:
        return ''
    parsed = urlparse(url)
    if parsed.scheme not in {'http', 'https'} or not parsed.netloc:
        return ''

    host = (parsed.netloc or '').lower()
    if 'drive.google.com' in host:
        query_params = parse_qs(parsed.query or '')
        file_id = (query_params.get('id') or [''])[0].strip()
        if not file_id:
            path_parts = [part for part in (parsed.path or '').split('/') if part]
            if 'd' in path_parts:
                try:
                    file_id = path_parts[path_parts.index('d') + 1]
                except Exception:
                    file_id = ''
        if file_id:
            return f"https://drive.google.com/uc?export=download&id={file_id}"
    return url


def _download_external_questions_json(raw_url):
    normalized_url = _normalize_external_json_url(raw_url)
    if not normalized_url:
        raise ValueError('Invalid URL. Use a full http(s) link.')

    response = requests.get(
        normalized_url,
        timeout=180,
        allow_redirects=True,
        headers={'User-Agent': 'form14-data-analysis-sync/1.0'},
    )
    response.raise_for_status()
    content_text = (response.text or '').lstrip('\ufeff').strip()
    if not content_text:
        raise ValueError('Downloaded file is empty.')

    try:
        payload = json.loads(content_text)
    except Exception as exc:
        raise ValueError(f'Link did not return valid JSON: {exc}') from exc

    if not isinstance(payload, dict):
        raise ValueError('JSON root must be an object.')
    questions = payload.get('questions')
    if not isinstance(questions, list):
        raise ValueError("JSON must include a top-level 'questions' list.")

    normalized_text = content_text + '\n' if not content_text.endswith('\n') else content_text
    return {
        'text': normalized_text,
        'question_count': len(questions),
        'download_url': normalized_url,
        'bytes': len(normalized_text.encode('utf-8')),
    }


def _parse_uploaded_questions_json(uploaded_file):
    if uploaded_file is None or not getattr(uploaded_file, 'filename', None):
        return None
    file_name = (uploaded_file.filename or '').strip()
    if not file_name:
        return None

    raw_bytes = uploaded_file.read()
    if not raw_bytes:
        raise ValueError(f'Uploaded file `{file_name}` is empty.')
    if len(raw_bytes) > (220 * 1024 * 1024):
        raise ValueError(f'Uploaded file `{file_name}` is too large. Keep it below 220 MB.')

    try:
        content_text = raw_bytes.decode('utf-8')
    except UnicodeDecodeError as exc:
        raise ValueError(f'Uploaded file `{file_name}` must be UTF-8 JSON.') from exc

    content_text = content_text.lstrip('\ufeff').strip()
    if not content_text:
        raise ValueError(f'Uploaded file `{file_name}` is empty.')

    try:
        payload = json.loads(content_text)
    except Exception as exc:
        raise ValueError(f'Uploaded file `{file_name}` is not valid JSON: {exc}') from exc

    if not isinstance(payload, dict):
        raise ValueError(f'Uploaded file `{file_name}` must have a JSON object root.')
    questions = payload.get('questions')
    if not isinstance(questions, list):
        raise ValueError(f"Uploaded file `{file_name}` must include a top-level 'questions' list.")

    normalized_text = content_text + '\n' if not content_text.endswith('\n') else content_text
    return {
        'text': normalized_text,
        'question_count': len(questions),
        'bytes': len(normalized_text.encode('utf-8')),
        'filename': file_name,
    }


def sync_data_analysis_json_from_sources(link_values, upload_files=None):
    timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    DATA_ANALYSIS_EXTERNAL_CACHE_DIR.mkdir(parents=True, exist_ok=True)

    result_rows = []
    for target in DATA_ANALYSIS_EXTERNAL_JSON_TARGETS:
        form_key = target['form_key']
        file_key = target.get('file_key') or f"{form_key}_file"
        label = target['label']
        target_path = Path(target['path'])
        source_url = (link_values.get(form_key) or '').strip()
        uploaded_file = (upload_files or {}).get(file_key) if upload_files is not None else None
        uploaded_result = _parse_uploaded_questions_json(uploaded_file)
        if uploaded_result:
            sync_result = dict(uploaded_result)
            sync_result['source_type'] = 'upload'
            sync_result['source_ref'] = uploaded_result.get('filename') or 'uploaded_file'
        else:
            if not source_url:
                raise ValueError(f'Missing link/upload for {label}.')
            download_result = _download_external_questions_json(source_url)
            sync_result = dict(download_result)
            sync_result['source_type'] = 'link'
            sync_result['source_ref'] = download_result.get('download_url')

        base_name = target_path.stem
        backup_before_path = None
        if target_path.exists():
            backup_before_path = DATA_ANALYSIS_EXTERNAL_CACHE_DIR / f"{base_name}_before_sync_{timestamp}.json"
            shutil.copy2(target_path, backup_before_path)

        source_suffix = 'upload' if sync_result.get('source_type') == 'upload' else 'link'
        cached_download_path = DATA_ANALYSIS_EXTERNAL_CACHE_DIR / f"{base_name}_from_{source_suffix}_{timestamp}.json"
        cached_download_path.write_text(sync_result['text'], encoding='utf-8')

        temp_path = target_path.with_suffix(target_path.suffix + '.tmp')
        temp_path.write_text(sync_result['text'], encoding='utf-8')
        temp_path.replace(target_path)

        result_rows.append({
            'form_key': form_key,
            'file_key': file_key,
            'label': label,
            'target': str(target_path),
            'cached_copy': str(cached_download_path),
            'previous_backup': str(backup_before_path) if backup_before_path else None,
            'question_count': int(sync_result['question_count']),
            'size_mb': round(float(sync_result['bytes']) / (1024 * 1024), 2),
            'source_type': sync_result.get('source_type'),
            'source_ref': sync_result.get('source_ref'),
        })

    DATA_ANALYSIS_DATASET_CACHE['cache_key'] = None
    DATA_ANALYSIS_DATASET_CACHE['payload'] = None
    return {
        'updated': True,
        'files': result_rows,
        'cached_at': datetime.now(timezone.utc).isoformat(),
    }


def load_data_analysis_dataset(dataset_path=DATA_ANALYSIS_DATASET_PATH):
    max_questions = max(int(os.getenv('DATA_ANALYSIS_MAX_QUESTIONS', str(DATA_ANALYSIS_MAX_QUESTIONS_DEFAULT)) or DATA_ANALYSIS_MAX_QUESTIONS_DEFAULT), 10000)
    mtimes = []
    tracked_paths = [Path(dataset_path)] + list(DATA_ANALYSIS_SUPPLEMENTAL_DATASET_PATHS) + list(DATA_ANALYSIS_KEYWORD_SCHEMA_PATHS)
    for tracked_path in tracked_paths:
        try:
            mtimes.append((str(tracked_path), int(tracked_path.stat().st_mtime)))
        except Exception:
            mtimes.append((str(tracked_path), 0))
    cache_key = (tuple(mtimes), max_questions)
    if DATA_ANALYSIS_DATASET_CACHE.get('cache_key') == cache_key and isinstance(DATA_ANALYSIS_DATASET_CACHE.get('payload'), dict):
        return DATA_ANALYSIS_DATASET_CACHE['payload']

    default_payload = {'questions': [], 'enabled_fields': [], 'intents': []}
    try:
        payload = json.loads(Path(dataset_path).read_text(encoding='utf-8'))
    except Exception:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}

    questions = payload.get('questions') if isinstance(payload.get('questions'), list) else []
    enabled_fields = payload.get('enabled_fields') if isinstance(payload.get('enabled_fields'), list) else []
    intents = payload.get('intents') if isinstance(payload.get('intents'), list) else []

    # Merge supplemental training prompts generated from filename/column keyword catalogs.
    merged_questions = list(questions)
    dedupe_keys = set()
    for row in merged_questions:
        if not isinstance(row, dict):
            continue
        dedupe_keys.add(((row.get('question') or '').strip().lower(), (row.get('intent') or '').strip().lower()))

    for extra_path in DATA_ANALYSIS_SUPPLEMENTAL_DATASET_PATHS:
        try:
            extra_payload = json.loads(Path(extra_path).read_text(encoding='utf-8'))
        except Exception:
            continue
        if not isinstance(extra_payload, dict):
            continue
        for row in (extra_payload.get('questions') or []):
            if not isinstance(row, dict):
                continue
            question = (row.get('question') or '').strip()
            intent = (row.get('intent') or '').strip().lower()
            if not question or not intent:
                continue
            dedupe_key = (question.lower(), intent)
            if dedupe_key in dedupe_keys:
                continue
            merged_questions.append(row)
            dedupe_keys.add(dedupe_key)

    if len(merged_questions) > max_questions:
        grouped_rows = defaultdict(list)
        for row in merged_questions:
            if not isinstance(row, dict):
                continue
            grouped_rows[(
                (row.get('intent') or '').strip().lower() or 'unknown',
                (row.get('target_domain') or row.get('domain') or '').strip().lower() or 'unknown',
            )].append(row)
        bucket_lists = sorted(grouped_rows.values(), key=len, reverse=True)
        bucket_indexes = [0] * len(bucket_lists)
        trimmed_questions = []
        while len(trimmed_questions) < max_questions:
            added = False
            for bucket_idx, bucket in enumerate(bucket_lists):
                index = bucket_indexes[bucket_idx]
                if index < len(bucket):
                    trimmed_questions.append(bucket[index])
                    bucket_indexes[bucket_idx] += 1
                    added = True
                    if len(trimmed_questions) >= max_questions:
                        break
            if not added:
                break
        merged_questions = trimmed_questions

    keyword_schema = load_data_analysis_keyword_schema()

    payload.update(default_payload)
    payload['questions'] = merged_questions
    payload['enabled_fields'] = enabled_fields
    payload['intents'] = intents
    payload['keyword_schema'] = keyword_schema
    payload['question_limit'] = max_questions
    DATA_ANALYSIS_DATASET_CACHE['cache_key'] = cache_key
    DATA_ANALYSIS_DATASET_CACHE['payload'] = payload
    return payload


def upsert_data_analysis_training_dataset(page_key=DATA_ANALYSIS_PAGE_KEY, replace=False, auto_sync_internal=True):
    normalized_key = normalize_field_help_page_key(page_key)
    sync_result = auto_sync_data_analysis_from_internal_db(force=replace) if auto_sync_internal else {
        'checked': False,
        'updated': False,
        'reason': 'Internal DB auto sync disabled.',
    }
    dataset = load_data_analysis_dataset()
    questions = dataset.get('questions') or []
    if not questions:
        return {'inserted': 0, 'train': 0, 'test': 0, 'internal_sync': sync_result}

    if replace:
        DataAnalysisTrainingQuestion.query.filter_by(page_key=normalized_key).delete()
        db.session.commit()

    existing = DataAnalysisTrainingQuestion.query.filter_by(page_key=normalized_key).count()
    if existing and not replace:
        train_count = DataAnalysisTrainingQuestion.query.filter_by(page_key=normalized_key, dataset_split='train').count()
        test_count = DataAnalysisTrainingQuestion.query.filter_by(page_key=normalized_key, dataset_split='test').count()
        return {'inserted': 0, 'train': train_count, 'test': test_count, 'internal_sync': sync_result}

    if len(questions) <= 2:
        train_size = len(questions)
        test_size = 0
    else:
        test_size = max(60, int(len(questions) * 0.15))
        test_size = min(test_size, max(len(questions) - 1, 1))
        train_size = max(len(questions) - test_size, 1)
    for idx, row in enumerate(questions):
        question = (row.get('question') or '').strip()
        intent = (row.get('intent') or '').strip().lower()
        if not question or not intent:
            continue
        if idx < train_size:
            split = 'train'
        elif idx < train_size + test_size:
            split = 'test'
        else:
            split = 'train'
        db.session.add(DataAnalysisTrainingQuestion(
            page_key=normalized_key,
            question=question,
            intent=intent,
            target_domain=(row.get('target_domain') or '').strip()[:120] or None,
            fields_hint_json=json.dumps(row.get('fields_hint') or [], ensure_ascii=True),
            answer_style=(row.get('answer_style') or '').strip()[:80] or None,
            dataset_split=split,
        ))
    db.session.commit()

    train_count = DataAnalysisTrainingQuestion.query.filter_by(page_key=normalized_key, dataset_split='train').count()
    test_count = DataAnalysisTrainingQuestion.query.filter_by(page_key=normalized_key, dataset_split='test').count()
    return {'inserted': train_count + test_count, 'train': train_count, 'test': test_count, 'internal_sync': sync_result}


def get_data_analysis_training_rows(page_key=DATA_ANALYSIS_PAGE_KEY, bootstrap=True):
    normalized_key = normalize_field_help_page_key(page_key)
    rows = (
        DataAnalysisTrainingQuestion.query
        .filter_by(page_key=normalized_key)
        .order_by(DataAnalysisTrainingQuestion.id.asc())
        .all()
    )
    if not rows and bootstrap:
        upsert_data_analysis_training_dataset(normalized_key, replace=False)
        rows = (
            DataAnalysisTrainingQuestion.query
            .filter_by(page_key=normalized_key)
            .order_by(DataAnalysisTrainingQuestion.id.asc())
            .all()
        )
    return rows


def train_data_analysis_intent_model(
    page_key=DATA_ANALYSIS_PAGE_KEY,
    force=False,
    model_family='rf',
    allow_bootstrap=True,
    allow_inline_training=True,
):
    normalized_key = normalize_field_help_page_key(page_key)
    cache_key = f"{normalized_key}:{model_family}"
    if not force and cache_key in DATA_ANALYSIS_MODEL_CACHE:
        return DATA_ANALYSIS_MODEL_CACHE[cache_key]
    if not force and not allow_inline_training:
        return {
            'available': False,
            'reason': 'Background training is required before the data analysis model can be used.',
        }

    rows = get_data_analysis_training_rows(normalized_key, bootstrap=allow_bootstrap)
    samples = [(row.question, row.intent, row.dataset_split) for row in rows if row.question and row.intent]
    max_model_samples = max(int(os.getenv('DATA_ANALYSIS_MODEL_MAX_SAMPLES', '40000') or 40000), 5000)
    if len(samples) > max_model_samples:
        grouped_samples = defaultdict(list)
        for sample in samples:
            grouped_samples[(sample[1], sample[2] or 'train')].append(sample)
        bucket_lists = sorted(grouped_samples.values(), key=len, reverse=True)
        bucket_indexes = [0] * len(bucket_lists)
        trimmed_samples = []
        while len(trimmed_samples) < max_model_samples:
            added = False
            for bucket_index, bucket in enumerate(bucket_lists):
                index = bucket_indexes[bucket_index]
                if index < len(bucket):
                    trimmed_samples.append(bucket[index])
                    bucket_indexes[bucket_index] += 1
                    added = True
                    if len(trimmed_samples) >= max_model_samples:
                        break
            if not added:
                break
        samples = trimmed_samples
    if not samples:
        state = {'available': False, 'reason': 'No training rows found.'}
        DATA_ANALYSIS_MODEL_CACHE[cache_key] = state
        return state

    sklearn_components = get_sklearn_components()
    if not sklearn_components.get('available'):
        state = {
            'available': False,
            'reason': 'scikit-learn is not installed.',
            'samples': len(samples),
            'labels': sorted({intent for _, intent, _ in samples}),
        }
        DATA_ANALYSIS_MODEL_CACHE[cache_key] = state
        return state

    train_rows = [(q, i) for q, i, split in samples if split == 'train']
    test_rows = [(q, i) for q, i, split in samples if split == 'test']
    if not train_rows:
        train_rows = [(q, i) for q, i, _ in samples]
    if not test_rows:
        split_fn = sklearn_components['train_test_split']
        q_list = [q for q, _, _ in samples]
        i_list = [i for _, i, _ in samples]
        x_train, x_test, y_train, y_test = split_fn(q_list, i_list, test_size=0.2, random_state=42, stratify=i_list if len(set(i_list)) > 1 else None)
        train_rows = list(zip(x_train, y_train))
        test_rows = list(zip(x_test, y_test))

    x_train = [q for q, _ in train_rows]
    y_train = [i for _, i in train_rows]
    x_test = [q for q, _ in test_rows]
    y_test = [i for _, i in test_rows]

    if model_family == 'dt':
        clf = sklearn_components['DecisionTreeClassifier'](max_depth=20, random_state=42)
    else:
        clf = sklearn_components['RandomForestClassifier'](
            n_estimators=320,
            max_depth=20,
            min_samples_leaf=1,
            random_state=42,
        )

    pipeline = sklearn_components['Pipeline']([
        ('tfidf', sklearn_components['TfidfVectorizer'](ngram_range=(1, 2), min_df=1, max_features=6000)),
        ('clf', clf),
    ])
    pipeline.fit(x_train, y_train)
    accuracy = pipeline.score(x_test, y_test) if x_test else 1.0

    tfidf = pipeline.named_steps['tfidf']
    question_matrix = tfidf.transform([q for q, _, _ in samples])
    state = {
        'available': True,
        'model_family': model_family,
        'pipeline': pipeline,
        'accuracy': float(accuracy),
        'labels': sorted(set(y_train + y_test)),
        'samples': samples,
        'question_matrix': question_matrix,
        'questions': [q for q, _, _ in samples],
        'intents': [i for _, i, _ in samples],
        'updated_at': utc_now().isoformat(),
    }
    DATA_ANALYSIS_MODEL_CACHE[cache_key] = state
    return state


def get_data_analysis_train_job_snapshot():
    with DATA_ANALYSIS_TRAIN_LOCK:
        return {
            'running': bool(DATA_ANALYSIS_TRAIN_JOB.get('running')),
            'started_at': DATA_ANALYSIS_TRAIN_JOB.get('started_at'),
            'finished_at': DATA_ANALYSIS_TRAIN_JOB.get('finished_at'),
            'last_error': DATA_ANALYSIS_TRAIN_JOB.get('last_error') or '',
            'last_result': dict(DATA_ANALYSIS_TRAIN_JOB.get('last_result') or {}),
        }


def ensure_data_analysis_background_train_job(replace_rows=False):
    global DATA_ANALYSIS_BACKGROUND_BOOT_STARTED
    if not DATA_ANALYSIS_BOT_ENABLED:
        return False

    with DATA_ANALYSIS_BACKGROUND_BOOT_LOCK:
        if DATA_ANALYSIS_TRAIN_JOB.get('running') or DATA_ANALYSIS_BACKGROUND_BOOT_STARTED:
            return False
        DATA_ANALYSIS_BACKGROUND_BOOT_STARTED = True

    with DATA_ANALYSIS_TRAIN_LOCK:
        DATA_ANALYSIS_TRAIN_JOB['running'] = True
        DATA_ANALYSIS_TRAIN_JOB['started_at'] = utc_now().isoformat()
        DATA_ANALYSIS_TRAIN_JOB['finished_at'] = None
        DATA_ANALYSIS_TRAIN_JOB['last_error'] = ''
        DATA_ANALYSIS_TRAIN_JOB['last_result'] = {}

    worker = threading.Thread(
        target=_background_data_analysis_train_job,
        args=(replace_rows,),
        daemon=True,
        name='data-analysis-train-worker-auto',
    )
    worker.start()
    return True


def _background_data_analysis_train_job(replace_rows=False):
    with app.app_context():
        try:
            DATA_ANALYSIS_MODEL_CACHE.clear()
            loader_result = upsert_data_analysis_training_dataset(
                page_key=DATA_ANALYSIS_PAGE_KEY,
                replace=replace_rows,
            )
            model_state = train_data_analysis_intent_model(
                page_key=DATA_ANALYSIS_PAGE_KEY,
                force=True,
                model_family='rf',
                allow_bootstrap=False,
            )
            set_admin_setting('data_analysis_model_available', '1' if model_state.get('available') else '0')
            set_admin_setting('data_analysis_model_accuracy', str(float(model_state.get('accuracy', 0.0) or 0.0)))
            set_admin_setting('data_analysis_model_labels', ','.join(model_state.get('labels', []) or []))
            set_admin_setting('data_analysis_model_trained_at', utc_now().isoformat())
            db.session.commit()
            with DATA_ANALYSIS_TRAIN_LOCK:
                DATA_ANALYSIS_TRAIN_JOB['last_result'] = {
                    'available': bool(model_state.get('available')),
                    'accuracy': float(model_state.get('accuracy', 0.0) or 0.0),
                    'labels': len(model_state.get('labels', []) or []),
                    'train': int(loader_result.get('train', 0) or 0),
                    'test': int(loader_result.get('test', 0) or 0),
                }
                DATA_ANALYSIS_TRAIN_JOB['last_error'] = ''
        except Exception as exc:
            db.session.rollback()
            app.logger.exception('Background data analysis training failed')
            with DATA_ANALYSIS_TRAIN_LOCK:
                DATA_ANALYSIS_TRAIN_JOB['last_error'] = str(exc)
                DATA_ANALYSIS_TRAIN_JOB['last_result'] = {}
        finally:
            with DATA_ANALYSIS_TRAIN_LOCK:
                DATA_ANALYSIS_TRAIN_JOB['running'] = False
                DATA_ANALYSIS_TRAIN_JOB['finished_at'] = utc_now().isoformat()


def predict_data_analysis_intent(question, page_key=DATA_ANALYSIS_PAGE_KEY):
    question = (question or '').strip()
    if not question:
        return {'intent': 'aggregation', 'confidence': 0.0, 'model_available': False}

    lowered = question.lower()
    heuristic_map = {
        'duplicates': ['duplicate', 'same name', 'repeat', 'duplication'],
        'completeness_check': ['quality', 'missing', 'complete', 'completeness', 'coverage'],
        'trend_analysis': ['trend', 'over time', 'yearly', 'monthly', 'historical'],
        'ranking': ['rank', 'top', 'highest', 'leading', 'largest', 'smallest'],
        'compare_groups': ['compare', 'versus', ' vs ', 'difference between'],
        'distribution': ['distribution', 'spread', 'breakdown', 'share of'],
        'anomaly_detection': ['anomaly', 'outlier', 'unusual', 'spike'],
        'validation_check': ['validate', 'invalid', 'rule', 'consistency check', 'quality check'],
    }
    heuristic_scores = defaultdict(float)
    for intent_name, terms in heuristic_map.items():
        for term in terms:
            if term in lowered:
                heuristic_scores[intent_name] += 0.34 if ' ' in term else 0.22

    if DATA_ANALYSIS_BACKGROUND_ONLY:
        ensure_data_analysis_background_train_job(replace_rows=False)
    model_state = train_data_analysis_intent_model(
        page_key=page_key,
        force=False,
        model_family='rf',
        allow_bootstrap=False,
        allow_inline_training=not DATA_ANALYSIS_BACKGROUND_ONLY,
    )
    domain_context = infer_data_analysis_domain(question)
    domain = (domain_context.get('domain') or '').strip()

    # Domain priors improve intent routing when users mention source files or columns.
    if domain in {'pbo_project_implementations', 'pbo_projects_carried_out'}:
        heuristic_scores['distribution'] += 0.18
        heuristic_scores['compare_groups'] += 0.14
    elif domain in {'pbo_donations', 'pbo_payments'}:
        heuristic_scores['ranking'] += 0.16
        heuristic_scores['aggregation'] += 0.14
    elif domain in {'pbo_reports', 'returnsform14_org_backup_manifest'}:
        heuristic_scores['validation_check'] += 0.24

    if model_state.get('available'):
        pipeline = model_state['pipeline']
        intent = pipeline.predict([question])[0]
        confidence = 0.0
        if hasattr(pipeline.named_steps['clf'], 'predict_proba'):
            probabilities = pipeline.predict_proba([question])[0]
            labels = list(getattr(pipeline.named_steps['clf'], 'classes_', []))
            score_map = {str(label): float(prob) for label, prob in zip(labels, probabilities)}
            for label, bonus in heuristic_scores.items():
                score_map[label] = score_map.get(label, 0.0) + min(0.28, bonus)
            if score_map:
                ranked = sorted(score_map.items(), key=lambda item: item[1], reverse=True)
                intent = ranked[0][0]
                confidence = min(1.0, ranked[0][1])
        return {
            'intent': intent,
            'confidence': confidence,
            'model_available': True,
            'accuracy': model_state.get('accuracy', 0.0),
        }

    if heuristic_scores:
        ranked = sorted(heuristic_scores.items(), key=lambda item: item[1], reverse=True)
        intent = ranked[0][0]
        confidence = min(0.9, 0.5 + ranked[0][1])
    else:
        intent = 'aggregation'
        confidence = 0.45
    return {'intent': intent, 'confidence': confidence, 'model_available': False, 'accuracy': 0.0}


def retrieve_similar_data_analysis_questions(question, page_key=DATA_ANALYSIS_PAGE_KEY, top_n=4):
    if DATA_ANALYSIS_BACKGROUND_ONLY:
        ensure_data_analysis_background_train_job(replace_rows=False)
    model_state = train_data_analysis_intent_model(
        page_key=page_key,
        force=False,
        model_family='rf',
        allow_bootstrap=False,
        allow_inline_training=not DATA_ANALYSIS_BACKGROUND_ONLY,
    )
    if not model_state.get('available'):
        return []
    try:
        tfidf = model_state['pipeline'].named_steps['tfidf']
        vector = tfidf.transform([question])
        similarities = cosine_similarity(vector, model_state['question_matrix'])[0]
        ranked_indices = sorted(range(len(similarities)), key=lambda idx: similarities[idx], reverse=True)
        results = []
        for idx in ranked_indices[:top_n]:
            results.append({
                'question': model_state['questions'][idx],
                'intent': model_state['intents'][idx],
                'score': float(similarities[idx]),
            })
        return results
    except Exception:
        return []


DATA_ANALYSIS_DOMAIN_HINTS = {
    'projects': 'pbo_project_implementations',
    'project': 'pbo_project_implementations',
    'implementation': 'pbo_project_implementations',
    'payments': 'pbo_payments',
    'payment': 'pbo_payments',
    'donations': 'pbo_donations',
    'donors': 'pbo_donations',
    'donor': 'pbo_donations',
    'assets': 'pbo_assets',
    'asset': 'pbo_assets',
    'officials': 'pbo_officials',
    'official': 'pbo_officials',
    'governance': 'pbo_officials',
    'banks': 'pbo_bank_accounts',
    'bank': 'pbo_bank_accounts',
    'account': 'pbo_bank_accounts',
    'auditors': 'pbo_auditors',
    'auditor': 'pbo_auditors',
    'training': 'pbo_training_records',
    'trainings': 'pbo_training_records',
    'volunteer': 'pbo_volunteer_privileges',
    'volunteers': 'pbo_volunteer_privileges',
    'intern': 'pbo_volunteer_privileges',
    'collaboration': 'pbo_collaboration_networking',
    'networking': 'pbo_collaboration_networking',
    'partner': 'pbo_collaboration_networking',
    'county': 'all_county_sources',
    'counties': 'all_county_sources',
    'reports': 'pbo_reports',
    'report': 'pbo_reports',
}

GLOBAL_COUNTY_SCAN_DOMAIN = 'all_county_sources'
COUNTY_MATCH_TERMS = ('county', 'counties')
ORGANIZATION_TERMS = {'organization', 'organizations', 'organisation', 'organisations', 'ngo', 'ngos', 'pbo', 'pbos'}
COUNT_TERMS = {'count', 'counts', 'number', 'numbers', 'many', 'total', 'totals'}
MEAN_TERMS = {'mean', 'average', 'avg'}
SUM_TERMS = {'sum', 'total', 'totals'}
DISTINCT_TERMS = {'distinct', 'unique', 'different'}
NON_MEANABLE_COLUMN_TERMS = {
    'account_number',
    'phone',
    'telephone',
    'mobile',
    'email',
    'id',
    'token',
    'pin',
    'registration_no',
    'registration_number',
}
ANALYSIS_MATCH_STOPWORDS = {
    'the', 'a', 'an', 'for', 'to', 'of', 'in', 'on', 'at', 'by', 'from', 'with', 'and',
    'show', 'give', 'me', 'list', 'tell', 'about', 'please', 'across', 'all', 'table',
    'tables', 'data', 'database', 'what', 'which', 'is', 'are', 'there', 'per',
}
COUNTY_NAME_HINTS = {
    'baringo', 'bomet', 'bungoma', 'busia', 'elgeyo marakwet', 'embu', 'garissa', 'homa bay',
    'isiolo', 'kajiado', 'kakamega', 'kericho', 'kiambu', 'kilifi', 'kirinyaga', 'kisii',
    'kisumu', 'kitui', 'kwale', 'laikipia', 'lamu', 'machakos', 'makueni', 'mandera', 'marsabit',
    'meru', 'migori', 'mombasa', 'muranga', 'nairobi', 'nairobi city', 'nakuru', 'nandi', 'narok',
    'nyamira', 'nyandarua', 'nyeri', 'samburu', 'siaya', 'taita taveta', 'tana river', 'tharaka nithi',
    'trans nzoia', 'turkana', 'uasin gishu', 'vihiga', 'wajir', 'west pokot',
}


def normalize_analysis_keyword(value):
    text = (value or '').strip().lower()
    text = re.sub(r'[^a-z0-9]+', '_', text)
    return re.sub(r'_+', '_', text).strip('_')


def tokenize_analysis_text(value):
    text = (value or '').strip().lower()
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return [token for token in text.split() if token]


def is_county_query(question):
    lowered = (question or '').strip().lower()
    if any(term in lowered for term in COUNTY_MATCH_TERMS):
        return True
    normalized = normalize_county_value(lowered)
    if not normalized:
        return False
    return any(
        hint == normalized or hint in normalized or normalized in hint
        for hint in COUNTY_NAME_HINTS
    )


def is_organization_count_query(question):
    tokens = set(tokenize_analysis_text(question))
    if not tokens:
        return False
    has_org = bool(tokens & ORGANIZATION_TERMS)
    has_count = bool(tokens & COUNT_TERMS) or 'how many' in (question or '').strip().lower()
    return has_org and has_count


def detect_data_analysis_operation(question):
    lowered = (question or '').strip().lower()
    tokens = set(tokenize_analysis_text(lowered))
    if not lowered:
        return None
    if (
        {'bank', 'banks'} & tokens
        and (tokens & ORGANIZATION_TERMS)
        and (tokens & COUNT_TERMS or 'how many' in lowered)
        and (is_county_query(question) or 'per county' in lowered or 'by county' in lowered)
    ):
        return 'bank_org_count_by_county'
    if any(token in tokens for token in MEAN_TERMS) or 'average of' in lowered:
        return 'mean'
    if any(token in tokens for token in SUM_TERMS) and 'summary' not in lowered:
        return 'sum'
    if (
        'each column' in lowered
        or 'every column' in lowered
        or 'all columns' in lowered
        or 'column profile' in lowered
        or 'column stats' in lowered
        or 'row items for each' in lowered
    ):
        return 'column_profile'
    if tokens & COUNT_TERMS or 'how many' in lowered:
        return 'count'
    return None


def is_non_meanable_column(column_name):
    normalized = normalize_data_analysis_name(column_name)
    return any(term in normalized for term in NON_MEANABLE_COLUMN_TERMS)


def _valid_text_series(series):
    return (
        series.astype(str)
        .str.strip()
        .replace({'': None, 'none': None, 'None': None, 'nan': None, 'NaN': None, 'NULL': None, 'null': None})
    )


def _frame_candidates_for_question(frames, domain_context=None):
    prioritized = []
    seen = set()
    domain = (domain_context or {}).get('domain')
    domain_df = resolve_domain_frame(domain, frames)
    if domain_df is not None and not domain_df.empty:
        frame_name = DATA_ANALYSIS_DOMAIN_FRAME_MAP.get(domain)
        if frame_name:
            prioritized.append((frame_name, domain_df))
            seen.add(frame_name)

    for frame_name, df in frames.items():
        if frame_name in seen:
            continue
        if df is None or df.empty:
            continue
        prioritized.append((frame_name, df))
    return prioritized


def match_best_analysis_column(frames, question, domain_context=None, numeric_required=False):
    pandas_module = get_pandas()
    lowered = (question or '').strip().lower()
    normalized_question = normalize_data_analysis_name(lowered)
    question_tokens = {
        token for token in tokenize_analysis_text(lowered)
        if len(token) > 2 and token not in ANALYSIS_MATCH_STOPWORDS
    }

    best = None
    for frame_name, df in _frame_candidates_for_question(frames, domain_context=domain_context):
        for column in df.columns:
            normalized_column = normalize_data_analysis_name(column)
            if not normalized_column:
                continue
            column_tokens = {token for token in normalized_column.split('_') if token and token not in ANALYSIS_MATCH_STOPWORDS}
            overlap = question_tokens & column_tokens
            score = float(len(overlap) * 2)
            if normalized_column in normalized_question:
                score += 2.8
            elif normalized_column.replace('_', ' ') in lowered:
                score += 2.2
            for alias in expand_data_analysis_column_aliases(normalized_column):
                alias_norm = normalize_data_analysis_name(alias)
                if not alias_norm:
                    continue
                if alias_norm in normalized_question:
                    score += 0.9
                alias_tokens = {token for token in alias_norm.split('_') if token}
                score += 0.25 * len(question_tokens & alias_tokens)

            if score <= 0:
                continue
            if numeric_required:
                numeric_series = pandas_module.to_numeric(df[column], errors='coerce')
                if int(numeric_series.notna().sum()) == 0:
                    continue
            candidate = {
                'frame_name': frame_name,
                'column': column,
                'score': score,
            }
            if best is None or candidate['score'] > best['score']:
                best = candidate
    return best


def normalize_county_value(value):
    text = str(value or '').strip().lower()
    text = text.replace('&', ' and ')
    text = re.sub(r'\bcity county\b', ' county', text)
    text = re.sub(r'\bcounty\b', ' ', text)
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def format_county_display(value):
    normalized = normalize_county_value(value)
    if not normalized:
        return 'County'
    if normalized in {'nairobi', 'nairobi city'}:
        return 'Nairobi City County'
    label = ' '.join(word.capitalize() for word in normalized.split())
    return label if label.endswith(' County') else f"{label} County"


def extract_county_name_from_question(question, frames):
    lowered = (question or '').strip().lower()
    if not lowered:
        return None

    known_map = {}
    for frame_name in ('counties_df', 'projects_df'):
        df = frames.get(frame_name)
        if df is None or df.empty:
            continue
        for column in df.columns:
            normalized_col = normalize_analysis_keyword(column)
            if 'county' not in normalized_col and 'counties' not in normalized_col:
                continue
            for raw_value in df[column].dropna().tolist():
                for part in split_multi_values(raw_value):
                    county_value = str(part or '').strip()
                    if not county_value:
                        continue
                    key = normalize_county_value(county_value)
                    if key:
                        known_map.setdefault(key, county_value)

    match_patterns = re.findall(r'([a-z][a-z\s\-]{2,80})\s+county\b', lowered)
    match_patterns.extend(re.findall(r'\bcounty\s+([a-z][a-z\s\-]{2,80})\b', lowered))
    match_patterns.extend(re.findall(r'\bcounty\s+of\s+([a-z][a-z\s\-]{2,80})\b', lowered))
    match_patterns.extend(re.findall(r'\b(?:in|within|across|for|from|inside)\s+([a-z][a-z\s\-]{2,80})\b', lowered))
    candidates = []
    leading_stopwords = {
        'how', 'many', 'organization', 'organizations', 'organisation', 'organisations',
        'in', 'within', 'for', 'from', 'across', 'show', 'count', 'number', 'of', 'the',
        'is', 'are', 'which', 'what', 'where', 'by', 'all', 'data', 'has', 'have', 'highest',
        'lowest', 'please', 'with', 'that', 'appear', 'appears', 'and',
    }
    for raw_candidate in match_patterns:
        tokens = [token for token in re.split(r'\s+', raw_candidate.strip()) if token]
        while tokens and tokens[0] in leading_stopwords:
            tokens.pop(0)
        while tokens and tokens[-1] in leading_stopwords:
            tokens.pop()
        candidate = ' '.join(tokens)
        cleaned = normalize_county_value(candidate)
        if cleaned:
            candidates.append(cleaned)

    if not candidates:
        question_key = normalize_county_value(lowered)
        for known_key in known_map.keys():
            if known_key and known_key in question_key:
                return format_county_display(known_map[known_key])
        return None

    for candidate in candidates:
        if candidate in known_map:
            return format_county_display(known_map[candidate])

    if known_map:
        closest = get_close_matches(candidates[0], list(known_map.keys()), n=1, cutoff=0.65)
        if closest:
            return format_county_display(known_map.get(closest[0]))

    if candidates[0] in COUNTY_NAME_HINTS:
        return format_county_display(candidates[0])
    return None


def _value_has_county_match(raw_value, target_norm):
    for part in split_multi_values(raw_value):
        normalized = normalize_county_value(part)
        if not normalized:
            continue
        if normalized == target_norm or normalized in target_norm or target_norm in normalized:
            return True
    normalized_raw = normalize_county_value(raw_value)
    if not normalized_raw:
        return False
    return normalized_raw == target_norm or normalized_raw in target_norm or target_norm in normalized_raw


def collect_county_metrics_across_frames(frames, target_county=None):
    target_norm = normalize_county_value(target_county) if target_county else ''
    metrics = []
    report_ids = set()

    for frame_name, df in frames.items():
        if df is None or df.empty:
            continue
        has_report_id = 'report_id' in df.columns
        for column in df.columns:
            normalized_col = normalize_analysis_keyword(column)
            if 'county' not in normalized_col and 'counties' not in normalized_col:
                continue

            series = df[column].fillna('').astype(str)
            if target_norm:
                mask = series.apply(lambda value: _value_has_county_match(value, target_norm))
            else:
                mask = series.str.strip().ne('')

            matched_count = int(mask.sum())
            if matched_count == 0:
                continue

            matched_reports = 0
            if has_report_id:
                ids = set(df.loc[mask, 'report_id'].dropna().tolist())
                matched_reports = len(ids)
                report_ids.update(ids)

            metrics.append({
                'source_frame': frame_name,
                'column': column,
                'matched_rows': matched_count,
                'matched_reports': matched_reports,
            })

    return metrics, report_ids


def collect_county_table_scan_from_database(target_county=None):
    target_norm = normalize_county_value(target_county) if target_county else ''
    inspector = inspect(db.engine)
    table_rows = []

    for table_name in inspector.get_table_names():
        county_columns = []
        for column in inspector.get_columns(table_name):
            column_name = str(column.get('name') or '').strip()
            normalized = normalize_analysis_keyword(column_name)
            if not normalized:
                continue
            if 'county' in normalized or 'counties' in normalized:
                county_columns.append(column_name)

        if not county_columns:
            continue

        table_obj = Table(table_name, MetaData(), autoload_with=db.engine)
        for column_name in county_columns:
            column_obj = table_obj.c.get(column_name)
            if column_obj is None:
                continue

            base_query = select(func.count()).select_from(table_obj)
            if target_norm:
                filter_value = f"%{target_norm.upper()}%"
                base_query = base_query.where(
                    func.upper(cast(column_obj, String)).like(filter_value)
                )
            else:
                base_query = base_query.where(column_obj.isnot(None))

            matched_rows = int(db.session.execute(base_query).scalar() or 0)
            if matched_rows <= 0:
                continue

            table_rows.append({
                'table_name': table_name,
                'column': column_name,
                'matched_rows': matched_rows,
            })

    return table_rows


def collect_organization_name_matches_from_database(target_county):
    target_norm = normalize_county_value(target_county)
    if not target_norm:
        return {'unique_org_count': 0, 'table_rows': [], 'organization_names': []}

    inspector = inspect(db.engine)
    org_name_preferences = [
        'pbo_name',
        'organization_name',
        'organisation_name',
        'org_name',
        'ngo_name',
        'name',
    ]
    unique_names = set()
    table_rows = []

    for table_name in inspector.get_table_names():
        columns = [str(col.get('name') or '').strip() for col in inspector.get_columns(table_name)]
        normalized_map = {normalize_analysis_keyword(col): col for col in columns if col}
        county_columns = [
            col for col in columns
            if 'county' in normalize_analysis_keyword(col) or 'counties' in normalize_analysis_keyword(col)
        ]
        if not county_columns:
            continue

        org_col = None
        for pref in org_name_preferences:
            mapped = normalized_map.get(pref)
            if mapped:
                org_col = mapped
                break
        if not org_col:
            continue

        table_obj = Table(table_name, MetaData(), autoload_with=db.engine)
        county_col = table_obj.c.get(county_columns[0])
        org_name_col = table_obj.c.get(org_col)
        if county_col is None or org_name_col is None:
            continue

        county_filter = func.upper(cast(county_col, String)).like(f"%{target_norm.upper()}%")
        non_empty_org = func.trim(cast(org_name_col, String)) != ''
        base_filter = county_filter & org_name_col.isnot(None) & non_empty_org

        row_count_query = select(func.count()).select_from(table_obj).where(base_filter)
        matched_rows = int(db.session.execute(row_count_query).scalar() or 0)
        if matched_rows <= 0:
            continue

        names_query = (
            select(func.upper(func.trim(cast(org_name_col, String))))
            .select_from(table_obj)
            .where(base_filter)
            .distinct()
        )
        names = [row[0] for row in db.session.execute(names_query).fetchall() if row and row[0]]
        for name in names:
            unique_names.add(str(name).strip())

        table_rows.append({
            'table_name': table_name,
            'organization_column': org_col,
            'county_column': county_columns[0],
            'matched_rows': matched_rows,
            'matched_organizations': len(names),
        })

    return {
        'unique_org_count': len(unique_names),
        'table_rows': table_rows,
        'organization_names': sorted(unique_names),
    }


def load_data_analysis_keyword_schema():
    merged_sources = []
    merged_keywords = set()
    seen_sources = set()
    for schema_path in DATA_ANALYSIS_KEYWORD_SCHEMA_PATHS:
        try:
            payload = json.loads(schema_path.read_text(encoding='utf-8'))
        except Exception:
            continue
        if not isinstance(payload, dict):
            continue
        sources = payload.get('sources') if isinstance(payload.get('sources'), list) else []
        for row in sources:
            if not isinstance(row, dict):
                continue
            source_name = (row.get('normalized_name') or '').strip()
            if not source_name:
                continue
            if source_name in seen_sources:
                continue
            seen_sources.add(source_name)
            merged_sources.append(row)
        for keyword in (payload.get('global_keywords') or []):
            normalized_keyword = normalize_analysis_keyword(keyword)
            if normalized_keyword:
                merged_keywords.add(normalized_keyword)
    return {'sources': merged_sources, 'global_keywords': sorted(merged_keywords)}


def infer_data_analysis_domain(question):
    lowered = (question or '').strip().lower()
    normalized_question = normalize_analysis_keyword(lowered)
    question_tokens = set(tokenize_analysis_text(lowered))
    if not lowered:
        return {'domain': None, 'source': None, 'score': 0.0, 'matched_terms': [], 'matched_columns': []}

    schema = load_data_analysis_keyword_schema()
    scored = []
    for source in schema.get('sources', []):
        source_name = (source.get('normalized_name') or '').strip()
        if not source_name:
            continue
        source_score = 0.0
        matched_terms = []
        matched_columns = []

        for alias in (source.get('aliases') or []):
            alias_text = (alias or '').strip().lower()
            if not alias_text:
                continue
            alias_norm = normalize_analysis_keyword(alias_text)
            if alias_text in lowered:
                source_score += 3.0 if len(alias_text.split()) > 1 else 1.8
                matched_terms.append(alias_text)
            elif alias_norm and alias_norm in normalized_question:
                source_score += 1.2
                matched_terms.append(alias_text)
            elif alias_norm in question_tokens:
                source_score += 1.0
                matched_terms.append(alias_text)

        for column in (source.get('normalized_columns') or []):
            column_name = normalize_analysis_keyword(column)
            if not column_name:
                continue
            if column_name in normalized_question:
                source_score += 1.1
                matched_columns.append(column_name)
            elif column_name.replace('_', ' ') in lowered:
                source_score += 0.9
                matched_columns.append(column_name)

        if source_score > 0:
            scored.append({
                'domain': source_name,
                'source': source.get('filename'),
                'score': source_score,
                'matched_terms': sorted(set(matched_terms)),
                'matched_columns': sorted(set(matched_columns)),
            })

    # Add direct domain hints so plain-language prompts like "projects" route correctly.
    for token in question_tokens:
        hinted_domain = DATA_ANALYSIS_DOMAIN_HINTS.get(token)
        if not hinted_domain:
            continue
        existing = next((item for item in scored if item['domain'] == hinted_domain), None)
        if existing:
            existing['score'] += 2.6
            existing['matched_terms'] = sorted(set(existing['matched_terms'] + [token]))
        else:
            scored.append({
                'domain': hinted_domain,
                'source': None,
                'score': 2.6,
                'matched_terms': [token],
                'matched_columns': [],
            })

    # Fuzzy column matching fallback for misspelled/unseen keywords.
    if not scored and question_tokens:
        for source in schema.get('sources', []):
            source_name = (source.get('normalized_name') or '').strip()
            if not source_name:
                continue
            keyword_pool = set()
            for alias in (source.get('aliases') or []):
                normalized_alias = normalize_analysis_keyword(alias)
                if normalized_alias:
                    keyword_pool.add(normalized_alias)
            for column in (source.get('normalized_columns') or []):
                normalized_column = normalize_analysis_keyword(column)
                if normalized_column:
                    keyword_pool.add(normalized_column)
            if not keyword_pool:
                continue

            score = 0.0
            matched_terms = []
            for token in question_tokens:
                if len(token) < 4:
                    continue
                close = get_close_matches(token, list(keyword_pool), n=1, cutoff=0.82)
                if not close:
                    continue
                score += 0.85
                matched_terms.append(close[0].replace('_', ' '))

            if score > 0:
                scored.append({
                    'domain': source_name,
                    'source': source.get('filename'),
                    'score': score,
                    'matched_terms': sorted(set(matched_terms)),
                    'matched_columns': [],
                })

    if not scored:
        return {'domain': None, 'source': None, 'score': 0.0, 'matched_terms': [], 'matched_columns': []}

    ranked = sorted(scored, key=lambda item: item['score'], reverse=True)
    return ranked[0]


def split_multi_values(raw_value):
    if raw_value is None:
        return []
    text_value = str(raw_value).replace('\n', ',').replace(';', ',')
    return [part.strip() for part in text_value.split(',') if part.strip()]


def collect_data_analysis_frames():
    pandas_module = get_pandas()
    reports = (
        PBOReport.query
        .options(
            selectinload(PBOReport.assets),
            selectinload(PBOReport.auditors),
            selectinload(PBOReport.bank_accounts),
            selectinload(PBOReport.collaborations),
            selectinload(PBOReport.payments),
            selectinload(PBOReport.donations),
            selectinload(PBOReport.officials),
            selectinload(PBOReport.project_implementations),
            selectinload(PBOReport.projects_carried_out),
            selectinload(PBOReport.training_records),
            selectinload(PBOReport.volunteer_privileges),
        )
        .order_by(PBOReport.id.asc())
        .all()
    )

    report_rows = []
    asset_rows = []
    auditor_rows = []
    bank_account_rows = []
    collaboration_rows = []
    payment_rows = []
    donation_rows = []
    official_rows = []
    project_rows = []
    projects_carried_out_rows = []
    county_rows = []
    training_rows = []
    volunteer_privilege_rows = []

    for report in reports:
        start_date = report.reporting_period_start
        year_value = start_date.year if start_date else None
        contact_phones = split_multi_values(report.contact_telephone)
        contact_emails = split_multi_values(report.contact_email)
        counties = split_multi_values(report.counties)
        payments_total = 0.0
        kenya_total = 0.0
        other_total = 0.0
        for item in (report.payments or []):
            kenya_value = legacy_zero_float(report, item.kenya_amount)
            other_value = legacy_zero_float(report, item.other_amount)
            kenya_total += kenya_value
            other_total += other_value
            payments_total += kenya_value + other_value
            payment_rows.append(legacy_zero_output_structure(report, {
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'description': item.description,
                'kenya_amount': kenya_value,
                'other_amount': other_value,
                'total_amount': kenya_value + other_value,
            }))

        donations_total = 0.0
        for item in (report.donations or []):
            amount_value = legacy_zero_float(report, item.amount)
            donations_total += amount_value
            donation_rows.append(legacy_zero_output_structure(report, {
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'donor_name': item.name,
                'country': item.country,
                'category': item.category,
                'amount': amount_value,
            }))

        project_beneficiaries = 0
        project_spending = 0.0
        for item in (report.project_implementations or []):
            beneficiaries = legacy_zero_int(report, item.beneficiaries_no)
            spending = resolve_project_spending_amount(item, report=report)
            project_beneficiaries += beneficiaries
            project_spending += spending
            project_rows.append(legacy_zero_output_structure(report, {
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'sector': item.sector,
                'county': item.county,
                'beneficiaries': beneficiaries,
                'spending': spending,
                'duration_years': legacy_zero_float(report, item.duration_years),
                'completion_status': item.completion_status,
            }))

        trainings_total = 0
        for item in (report.training_records or []):
            total_participants = legacy_zero_int(report, item.kenyan_count) + legacy_zero_int(report, item.international_count)
            trainings_total += total_participants
            training_rows.append(legacy_zero_output_structure(report, {
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'training_type': item.training_type,
                'participants': total_participants,
            }))

        for item in (report.assets or []):
            asset_rows.append(legacy_zero_output_structure(report, {
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'item': item.item,
                'number': legacy_zero_int(report, item.number),
                'value': legacy_zero_float(report, item.value),
            }))

        for item in (report.auditors or []):
            auditor_rows.append({
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'firm': item.firm,
                'auditor_name': item.auditor_name,
                'practicing_no': item.practicing_no,
            })

        for item in (report.bank_accounts or []):
            bank_account_rows.append({
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'bank_name': item.bank_name,
                'branch': item.branch,
                'account_number': item.account_number,
                'currency': item.currency,
            })

        for item in (report.collaborations or []):
            collaboration_rows.append({
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'partner_type': item.partner_type,
                'info_exchange': item.info_exchange,
                'tech_support_to_partner': item.tech_support_to_partner,
                'tech_support_from_partner': item.tech_support_from_partner,
                'funding_to_partner': item.funding_to_partner,
                'funding_from_partner': item.funding_from_partner,
                'equipment_to_partner': item.equipment_to_partner,
                'equipment_from_partner': item.equipment_from_partner,
            })

        for item in (report.officials or []):
            official_rows.append({
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'role': item.role,
                'name': item.name,
                'nationality': item.nationality,
                'gender': item.gender,
                'email': item.email,
                'residence': item.residence,
                'phone': item.phone,
                'kra_pin': item.kra_pin,
                'professional_qualification': item.professional_qualification,
            })

        for item in (report.projects_carried_out or []):
            projects_carried_out_rows.append(legacy_zero_output_structure(report, {
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'sector': item.sector,
                'carried_forward_kenya': legacy_zero_float(report, item.carried_forward_kenya),
                'carried_forward_other': legacy_zero_float(report, item.carried_forward_other),
                'started_kenya': legacy_zero_float(report, item.started_kenya),
                'started_other': legacy_zero_float(report, item.started_other),
                'completed_kenya': legacy_zero_float(report, item.completed_kenya),
                'completed_other': legacy_zero_float(report, item.completed_other),
            }))

        for item in (report.volunteer_privileges or []):
            volunteer_privilege_rows.append({
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'category': item.category,
                'kenyan_volunteer': int(bool(item.kenyan_volunteer)),
                'kenyan_intern': int(bool(item.kenyan_intern)),
                'international_volunteer': int(bool(item.international_volunteer)),
                'international_intern': int(bool(item.international_intern)),
            })

        for county in counties:
            county_rows.append({
                'report_id': report.id,
                'year': year_value,
                'scope': report.scope,
                'county': county,
            })

        report_rows.append(legacy_zero_output_structure(report, {
            'report_id': report.id,
            'year': year_value,
            'scope': report.scope or 'UNSPECIFIED',
            'pbo_name': report.pbo_name or '',
            'reporting_period_start': start_date.isoformat() if start_date else '',
            'created_at': report.created_at.isoformat() if report.created_at else '',
            'updated_at': report.updated_at.isoformat() if report.updated_at else '',
            'workflow_status': (report.workflow_status or '').strip().lower(),
            'review_status': (report.review_status or '').strip().lower(),
            'risk_score': float(report.risk_score) if report.risk_score is not None else None,
            'county_count': len(counties),
            'contact_phone_count': len(contact_phones),
            'contact_email_count': len(contact_emails),
            'payments_total': payments_total,
            'payments_kenya': kenya_total,
            'payments_other': other_total,
            'donations_total': donations_total,
            'projects_count': len(report.project_implementations or []),
            'project_beneficiaries_total': project_beneficiaries,
            'project_spending_total': project_spending,
            'staff_kenyan_current': legacy_zero_int(report, report.staff_kenyan_current),
            'staff_foreign_current': legacy_zero_int(report, report.staff_foreign_current),
            'staff_kenyan_left': legacy_zero_int(report, report.staff_kenyan_left),
            'staff_foreign_left': legacy_zero_int(report, report.staff_foreign_left),
            'volunteers_kenyan_current': legacy_zero_int(report, report.volunteers_kenyan_current),
            'volunteers_foreign_current': legacy_zero_int(report, report.volunteers_foreign_current),
            'trainings_total': trainings_total,
            'has_contact_name': 1 if (report.contact_name or '').strip() else 0,
            'has_scope': 1 if (report.scope or '').strip() else 0,
            'has_start_date': 1 if start_date else 0,
        }))

    return {
        'reports_df': pandas_module.DataFrame(report_rows),
        'assets_df': pandas_module.DataFrame(asset_rows),
        'auditors_df': pandas_module.DataFrame(auditor_rows),
        'bank_accounts_df': pandas_module.DataFrame(bank_account_rows),
        'collaboration_df': pandas_module.DataFrame(collaboration_rows),
        'payments_df': pandas_module.DataFrame(payment_rows),
        'donations_df': pandas_module.DataFrame(donation_rows),
        'officials_df': pandas_module.DataFrame(official_rows),
        'projects_df': pandas_module.DataFrame(project_rows),
        'projects_carried_out_df': pandas_module.DataFrame(projects_carried_out_rows),
        'counties_df': pandas_module.DataFrame(county_rows),
        'trainings_df': pandas_module.DataFrame(training_rows),
        'volunteer_privileges_df': pandas_module.DataFrame(volunteer_privilege_rows),
    }


def get_plotly_express():
    try:
        import plotly.express as px
        return px
    except Exception:
        return None


def _analysis_result(title, summary_lines=None, figures=None, tables=None, module='aggregation'):
    return {
        'title': title,
        'summary_lines': summary_lines or [],
        'figures': figures or [],
        'tables': tables or [],
        'module': module,
    }


DATA_ANALYSIS_DOMAIN_FRAME_MAP = {
    'sector_report_data': 'reports_df',
    'pbo_reports': 'reports_df',
    'pbo_assets': 'assets_df',
    'pbo_auditors': 'auditors_df',
    'pbo_bank_accounts': 'bank_accounts_df',
    'pbo_collaboration_networking': 'collaboration_df',
    'pbo_donations': 'donations_df',
    'pbo_officials': 'officials_df',
    'pbo_payments': 'payments_df',
    'pbo_project_implementations': 'projects_df',
    'pbo_projects_carried_out': 'projects_carried_out_df',
    'pbo_training_records': 'trainings_df',
    'pbo_volunteer_privileges': 'volunteer_privileges_df',
    'returnsform14_org_backup_manifest': 'reports_df',
}


def resolve_domain_frame(domain, frames):
    if not domain:
        return None
    frame_name = DATA_ANALYSIS_DOMAIN_FRAME_MAP.get(domain)
    if not frame_name:
        return None
    return frames.get(frame_name)


def pick_group_column(df):
    preferred = [
        'sector', 'county', 'scope', 'category', 'role', 'partner_type',
        'bank_name', 'currency', 'training_type', 'name', 'donor_name',
        'description', 'item', 'country',
    ]
    for column in preferred:
        if column in df.columns:
            return column
    for column in df.columns:
        if df[column].dtype == object:
            return column
    return None


def pick_metric_column(df):
    preferred = [
        'amount', 'total_amount', 'spending', 'spending_per_county',
        'project_spending_total', 'beneficiaries', 'participants', 'value',
        'payments_total', 'donations_total', 'number', 'size',
    ]
    numeric_columns = [column for column in df.columns if str(df[column].dtype) != 'object']
    for column in preferred:
        if column in numeric_columns:
            return column
    return numeric_columns[0] if numeric_columns else None


def domain_no_data_result(domain, title, module_name, reason=None):
    domain_label = (domain or 'requested domain').replace('_', ' ')
    note = reason or f"No data available for {domain_label} in the current dataset."
    return _analysis_result(
        f"{title} ({domain_label.title()})",
        [note],
        module=module_name,
    )


def module_bank_org_count_by_county(frames, question):
    pandas_module = get_pandas()
    px = get_plotly_express()
    banks_df = frames.get('bank_accounts_df')
    reports_df = frames.get('reports_df')
    counties_df = frames.get('counties_df')
    if banks_df is None or banks_df.empty:
        return _analysis_result(
            'Bank Organizations by County',
            ['No bank account data is available in the current dataset.'],
            module='column_operations',
        )

    joined = banks_df.copy()
    if reports_df is not None and not reports_df.empty and {'report_id', 'pbo_name'}.issubset(set(reports_df.columns)):
        joined = joined.merge(
            reports_df[['report_id', 'pbo_name']],
            on='report_id',
            how='left',
        )
    if counties_df is not None and not counties_df.empty and {'report_id', 'county'}.issubset(set(counties_df.columns)):
        joined = joined.merge(
            counties_df[['report_id', 'county']],
            on='report_id',
            how='left',
        )
    if 'bank_name' not in joined.columns:
        return _analysis_result(
            'Bank Organizations by County',
            ['Bank name is missing from the available bank account rows.'],
            module='column_operations',
        )

    joined['bank_name'] = _valid_text_series(joined['bank_name'])
    if 'county' in joined.columns:
        joined['county'] = _valid_text_series(joined['county'])
    else:
        joined['county'] = 'UNSPECIFIED COUNTY'

    if 'pbo_name' in joined.columns:
        joined['pbo_name'] = _valid_text_series(joined['pbo_name'])
    else:
        joined['pbo_name'] = None

    target_county = extract_county_name_from_question(question, frames)
    if target_county:
        target_norm = normalize_county_value(target_county)
        joined = joined[joined['county'].fillna('').astype(str).apply(lambda value: _value_has_county_match(value, target_norm))]

    joined = joined[joined['bank_name'].notna()]
    if joined.empty:
        label = format_county_display(target_county) if target_county else 'the selected county filter'
        return _analysis_result(
            'Bank Organizations by County',
            [f'No rows matched {label}.'],
            module='column_operations',
        )

    grouped = (
        joined.groupby(['bank_name', 'county'], as_index=False)
        .agg(
            organization_count=('pbo_name', 'nunique'),
            account_rows=('bank_name', 'count'),
        )
        .sort_values(['organization_count', 'account_rows'], ascending=False)
    )
    grouped['organization_count'] = grouped['organization_count'].fillna(0).astype(int)
    grouped['account_rows'] = grouped['account_rows'].fillna(0).astype(int)

    summary = [
        f"Bank-organization county matrix built from {len(joined)} bank-account rows.",
        f"Unique banks matched: {grouped['bank_name'].nunique()} | county rows: {grouped['county'].nunique()}.",
    ]
    if target_county:
        summary.insert(0, f"Organizations by bank in {format_county_display(target_county)}.")

    figures = []
    if px:
        figures.append(
            px.bar(
                grouped.head(20),
                x='bank_name',
                y='organization_count',
                color='county',
                barmode='group',
                title='Organizations per Bank by County',
                hover_data=['account_rows'],
            )
        )

    preview_cols = ['bank_name', 'county', 'organization_count', 'account_rows']
    preview = grouped[preview_cols].reset_index(drop=True)
    return _analysis_result(
        'Bank Organizations by County',
        summary,
        figures=figures,
        tables=[preview],
        module='column_operations',
    )


def module_column_profile(frames, domain_context=None):
    pandas_module = get_pandas()
    profile_rows = []
    frame_count = 0

    for frame_name, df in _frame_candidates_for_question(frames, domain_context=domain_context):
        if df is None or df.empty:
            continue
        frame_count += 1
        for column in df.columns:
            series = df[column]
            non_null = int(series.notna().sum())
            distinct = int(series.nunique(dropna=True))
            dtype = str(series.dtype)
            row = {
                'source': frame_name,
                'column': column,
                'dtype': dtype,
                'rows': int(len(df)),
                'non_null_rows': non_null,
                'distinct_values': distinct,
                'mean': None,
            }
            if not is_non_meanable_column(column):
                numeric_series = pandas_module.to_numeric(series, errors='coerce')
                if int(numeric_series.notna().sum()) > 0:
                    row['mean'] = round(float(numeric_series.mean()), 4)
            profile_rows.append(row)

    if not profile_rows:
        return _analysis_result(
            'Column Profile',
            ['No rows are available to profile in the current analysis dataset.'],
            module='column_operations',
        )

    profile_df = pandas_module.DataFrame(profile_rows).sort_values(
        ['source', 'non_null_rows', 'distinct_values'],
        ascending=[True, False, False],
    )
    summary = [
        f"Profiled {len(profile_df)} columns across {frame_count} data sources.",
        "Mean values are computed only for numeric columns and skipped for identifiers (for example account numbers, phones, emails, IDs).",
    ]
    return _analysis_result(
        'Column Profile',
        summary,
        tables=[profile_df.reset_index(drop=True)],
        module='column_operations',
    )


def module_column_operations(frames, question, domain_context=None):
    pandas_module = get_pandas()
    px = get_plotly_express()
    operation = detect_data_analysis_operation(question)
    if not operation:
        return None

    if operation == 'bank_org_count_by_county':
        return module_bank_org_count_by_county(frames, question)

    if operation == 'column_profile':
        return module_column_profile(frames, domain_context=domain_context)

    if is_county_query(question):
        return None

    requires_numeric = operation in {'mean', 'sum'}
    candidate = match_best_analysis_column(
        frames,
        question,
        domain_context=domain_context,
        numeric_required=requires_numeric,
    )
    if not candidate:
        if requires_numeric:
            return _analysis_result(
                'Column Operations',
                ['No numeric column matched this question. Try naming a numeric field such as amount, value, spending, beneficiaries, or total.'],
                module='column_operations',
            )
        return None

    frame_name = candidate['frame_name']
    column = candidate['column']
    df = frames.get(frame_name)
    if df is None or df.empty or column not in df.columns:
        return None

    if operation == 'count':
        series = df[column]
        use_distinct = bool(set(tokenize_analysis_text(question)) & DISTINCT_TERMS) or 'how many different' in (question or '').lower()
        if str(series.dtype) == 'object':
            cleaned = _valid_text_series(series)
            value = int(cleaned.nunique(dropna=True)) if use_distinct else int(cleaned.notna().sum())
        else:
            value = int(series.nunique(dropna=True)) if use_distinct else int(series.notna().sum())
        summary = [
            f"Count for {column.replace('_', ' ')} in {frame_name.replace('_df', '')}: {value}.",
            f"Mode: {'distinct values' if use_distinct else 'non-empty rows'} from {len(df)} total rows.",
        ]
        preview_columns = [col for col in ['report_id', 'year', 'scope', 'pbo_name', 'county', column] if col in df.columns]
        if not preview_columns:
            preview_columns = [column]
        table = df[preview_columns].head(50).reset_index(drop=True)
        return _analysis_result(
            'Column Count',
            summary,
            tables=[table],
            module='column_operations',
        )

    if operation in {'mean', 'sum'}:
        if is_non_meanable_column(column):
            return _analysis_result(
                'Column Operations',
                [f"{column.replace('_', ' ').title()} is treated as an identifier field, so {operation} is not computed to avoid misleading statistics."],
                module='column_operations',
            )

        numeric_series = pandas_module.to_numeric(df[column], errors='coerce')
        clean_df = df.copy()
        clean_df['_metric_value'] = numeric_series
        clean_df = clean_df[clean_df['_metric_value'].notna()]
        if clean_df.empty:
            return _analysis_result(
                'Column Operations',
                [f"No numeric values were found in {column.replace('_', ' ')} for this operation."],
                module='column_operations',
            )

        metric_value = float(clean_df['_metric_value'].mean()) if operation == 'mean' else float(clean_df['_metric_value'].sum())
        operation_label = 'Mean' if operation == 'mean' else 'Sum'
        summary = [
            f"{operation_label} of {column.replace('_', ' ')} in {frame_name.replace('_df', '')}: {metric_value:,.4f}.",
            f"Computed from {len(clean_df)} numeric rows (ignored non-numeric/blank values).",
        ]

        tables = []
        figures = []
        grouping_hints = ['county', 'scope', 'sector', 'bank_name', 'category', 'currency', 'training_type']
        group_col = next((hint for hint in grouping_hints if hint in clean_df.columns), None)
        if group_col and ('by ' in (question or '').lower() or 'per ' in (question or '').lower()):
            grouped = (
                clean_df.groupby(group_col, as_index=False)['_metric_value']
                .agg('mean' if operation == 'mean' else 'sum')
                .sort_values('_metric_value', ascending=False)
                .head(30)
            )
            grouped.rename(columns={'_metric_value': operation}, inplace=True)
            tables.append(grouped.reset_index(drop=True))
            if px:
                figures.append(
                    px.bar(
                        grouped,
                        x=group_col,
                        y=operation,
                        title=f"{operation_label} of {column.replace('_', ' ').title()} by {group_col.replace('_', ' ').title()}",
                    )
                )

        preview_cols = [col for col in ['report_id', 'year', 'scope', 'pbo_name', 'county', column] if col in clean_df.columns]
        if not preview_cols:
            preview_cols = [column]
        tables.append(clean_df[preview_cols].head(50).reset_index(drop=True))

        return _analysis_result(
            f'Column {operation_label}',
            summary,
            figures=figures,
            tables=tables,
            module='column_operations',
        )

    return None


def module_county_cross_database(frames, question, domain_context=None):
    pandas_module = get_pandas()
    px = get_plotly_express()
    lowered_question = (question or '').strip().lower()
    target_county = extract_county_name_from_question(question, frames)
    display_county = format_county_display(target_county) if target_county else None
    org_count_query = is_organization_count_query(question)
    frame_metrics, report_ids = collect_county_metrics_across_frames(frames, target_county=target_county)
    table_scan_rows = collect_county_table_scan_from_database(target_county=target_county)
    org_table_matches = (
        collect_organization_name_matches_from_database(target_county)
        if (org_count_query and target_county)
        else {'unique_org_count': 0, 'table_rows': [], 'organization_names': []}
    )

    if not frame_metrics and not table_scan_rows and not org_table_matches.get('table_rows'):
        if display_county:
            return _analysis_result(
                'Cross-Database County Scan',
                [f"No data available for {display_county} across county/counties columns."],
                module='county_cross_database',
            )
        return _analysis_result(
            'Cross-Database County Scan',
            ['No county/counties columns with data were found in the current database.'],
            module='county_cross_database',
        )

    frame_df = pandas_module.DataFrame(frame_metrics)
    table_df = pandas_module.DataFrame(table_scan_rows)
    county_frame = frames.get('counties_df')

    if not display_county and county_frame is not None and not county_frame.empty and {'county', 'report_id'}.issubset(set(county_frame.columns)):
        county_distribution = (
            county_frame[county_frame['county'].fillna('').astype(str).str.strip() != '']
            .groupby('county', as_index=False)
            .agg(
                matched_rows=('report_id', 'count'),
                organization_count=('report_id', 'nunique'),
            )
            .sort_values('organization_count', ascending=False)
        )
        if not county_distribution.empty and (
            org_count_query
            or 'distribution' in lowered_question
            or 'by county' in lowered_question
            or 'top' in lowered_question
            or 'highest' in lowered_question
            or 'rank' in lowered_question
            or 'most' in lowered_question
        ):
            top_row = county_distribution.iloc[0]
            summary = [
                f"Top county by organizations: {format_county_display(top_row['county'])} ({int(top_row['organization_count'])}).",
                f"Computed county distribution across {int(county_distribution['organization_count'].sum())} organization-county matches.",
            ]
            figures = []
            if px:
                figures.append(
                    px.bar(
                        county_distribution.head(20),
                        x='county',
                        y='organization_count',
                        title='Organizations by County (Cross-Database Scan)',
                        hover_data=['matched_rows'],
                    )
                )
            return _analysis_result(
                'Cross-Database County Distribution',
                summary,
                figures=figures,
                tables=[county_distribution.reset_index(drop=True)],
                module='county_cross_database',
            )

    summary = []
    if display_county and org_count_query and org_table_matches.get('table_rows'):
        summary.append(
            f"Organizations in {display_county}: {int(org_table_matches.get('unique_org_count') or 0)}."
        )
        summary.append(
            f"Matched rows in tables containing both organization-name and county columns: "
            f"{sum(int(row.get('matched_rows', 0)) for row in org_table_matches.get('table_rows', []))}."
        )
    elif display_county:
        summary.append(f"Organizations in {display_county}: {len(report_ids)}.")
    else:
        summary.append("Cross-table county scan completed for all county/counties columns.")

    if not frame_df.empty:
        summary.append(
            f"Frame scan matched {int(frame_df['matched_rows'].sum())} rows across "
            f"{frame_df['source_frame'].nunique()} in-app datasets."
        )
    if not table_df.empty:
        summary.append(
            f"Database scan matched {int(table_df['matched_rows'].sum())} rows across "
            f"{table_df['table_name'].nunique()} tables."
        )

    tables = []
    figures = []
    if not frame_df.empty:
        frame_view = frame_df.sort_values('matched_rows', ascending=False).reset_index(drop=True)
        tables.append(frame_view)
        if px:
            bar_df = frame_view.head(12).copy()
            bar_df['source_column'] = bar_df['source_frame'] + '::' + bar_df['column']
            figures.append(
                px.bar(
                    bar_df,
                    x='source_column',
                    y='matched_rows',
                    title='County Matches Across In-App Datasets',
                    hover_data=['matched_reports'],
                )
            )
    if not table_df.empty:
        tables.append(table_df.sort_values('matched_rows', ascending=False).reset_index(drop=True))
    if org_table_matches.get('table_rows'):
        tables.append(
            pandas_module.DataFrame(org_table_matches['table_rows'])
            .sort_values('matched_rows', ascending=False)
            .reset_index(drop=True)
        )

    return _analysis_result(
        'Cross-Database County Scan',
        summary,
        figures=figures,
        tables=tables,
        module='county_cross_database',
    )


def module_aggregation(frames, question=None, domain_context=None):
    domain = (domain_context or {}).get('domain')
    domain_df = resolve_domain_frame(domain, frames)
    if domain and (domain_df is None or domain_df.empty):
        return domain_no_data_result(domain, 'Aggregation Summary', 'aggregation')
    if domain_df is not None and not domain_df.empty:
        group_col = pick_group_column(domain_df)
        metric_col = pick_metric_column(domain_df)
        title_suffix = domain.replace('_', ' ').title() if domain else 'Domain'
        summary = [f"Rows analyzed for {title_suffix}: {len(domain_df)}."]
        if metric_col:
            total_metric = float(domain_df[metric_col].fillna(0).sum())
            summary.append(f"Total {metric_col.replace('_', ' ')}: {total_metric:,.2f}.")
        if group_col:
            unique_groups = int(domain_df[group_col].astype(str).str.strip().replace({'': None}).notna().sum())
            summary.append(f"Populated {group_col.replace('_', ' ')} rows: {unique_groups}.")
        columns = [col for col in [group_col, metric_col, 'report_id', 'year', 'scope'] if col and col in domain_df.columns]
        if not columns:
            columns = list(domain_df.columns[:6])
        return _analysis_result(
            f'Aggregation Summary ({title_suffix})',
            summary,
            tables=[domain_df[columns].head(50)],
            module='aggregation',
        )

    df = frames['reports_df']
    if df.empty:
        return _analysis_result('Dataset Summary', ['No report data is available yet.'], module='aggregation')
    total_reports = int(len(df))
    total_payments = float(df['payments_total'].sum())
    total_donations = float(df['donations_total'].sum())
    total_projects = int(df['projects_count'].sum())
    summary = [
        f"Total submissions analyzed: {total_reports}.",
        f"Total payments logged: KES {total_payments:,.2f}.",
        f"Total donations logged: KES {total_donations:,.2f}.",
        f"Total project rows captured: {total_projects}.",
    ]
    table = df[['report_id', 'year', 'scope', 'payments_total', 'donations_total', 'projects_count']].head(30)
    return _analysis_result('Aggregation Summary', summary, tables=[table], module='aggregation')


def module_distribution(frames, question, domain_context=None):
    px = get_plotly_express()
    domain = (domain_context or {}).get('domain')
    domain_df = resolve_domain_frame(domain, frames)
    if domain and (domain_df is None or domain_df.empty):
        return domain_no_data_result(domain, 'Distribution', 'distribution')
    if domain_df is not None and not domain_df.empty:
        group_col = pick_group_column(domain_df)
        if group_col:
            grouped = domain_df.groupby(group_col, as_index=False).size().sort_values('size', ascending=False).head(25)
            title = f"Distribution ({domain.replace('_', ' ').title()})"
            fig = px.bar(grouped, x=group_col, y='size', title=title) if px else None
            return _analysis_result(
                title,
                [f"Distribution by {group_col.replace('_', ' ')} using {domain.replace('_', ' ')} domain keywords."],
                figures=[fig] if fig else [],
                tables=[grouped],
                module='distribution',
            )

    lowered = (question or '').lower()
    if 'county' in lowered:
        df = frames['counties_df']
        if df.empty:
            return _analysis_result('County Distribution', ['No county records available.'], module='distribution')
        grouped = df.groupby('county', as_index=False).size().sort_values('size', ascending=False).head(20)
        fig = px.bar(grouped, x='county', y='size', title='Top Counties in Enabled Field Data') if px else None
        return _analysis_result('County Distribution', ['Distribution by county extracted from enabled county fields.'], figures=[fig] if fig else [], tables=[grouped], module='distribution')
    if 'sector' in lowered:
        df = frames['projects_df']
        if df.empty:
            return _analysis_result('Sector Distribution', ['No project sector rows available.'], module='distribution')
        grouped = df.groupby('sector', as_index=False).size().sort_values('size', ascending=False).head(20)
        fig = px.bar(grouped, x='sector', y='size', title='Project Sector Distribution') if px else None
        return _analysis_result('Sector Distribution', ['Distribution by project sector from enabled project rows.'], figures=[fig] if fig else [], tables=[grouped], module='distribution')
    df = frames['reports_df']
    grouped = df.groupby('scope', as_index=False).size().sort_values('size', ascending=False)
    fig = px.pie(grouped, names='scope', values='size', title='Scope Distribution') if px and not grouped.empty else None
    return _analysis_result('Scope Distribution', ['Distribution by scope from enabled scope field values.'], figures=[fig] if fig else [], tables=[grouped], module='distribution')


def module_trend_analysis(frames, domain_context=None):
    px = get_plotly_express()
    domain = (domain_context or {}).get('domain')
    domain_df = resolve_domain_frame(domain, frames)
    if domain and (domain_df is None or domain_df.empty):
        return domain_no_data_result(domain, 'Trend Analysis', 'trend_analysis')
    if domain and (
        'year' not in domain_df.columns
        or domain_df['year'].dropna().shape[0] == 0
    ):
        return domain_no_data_result(
            domain,
            'Trend Analysis',
            'trend_analysis',
            reason=f"No year data available for {domain.replace('_', ' ')} trend analysis.",
        )
    if domain_df is not None and not domain_df.empty and 'year' in domain_df.columns and domain_df['year'].dropna().shape[0] > 0:
        metric_col = pick_metric_column(domain_df)
        if metric_col:
            grouped = (
                domain_df[domain_df['year'].notna()]
                .groupby('year', as_index=False)[metric_col]
                .sum()
                .sort_values('year')
            )
            title = f"Trend Analysis ({domain.replace('_', ' ').title()})"
            fig = px.line(grouped, x='year', y=metric_col, title=title) if px else None
            return _analysis_result(
                title,
                [f"Yearly trend of {metric_col.replace('_', ' ')} from {domain.replace('_', ' ')}."],
                figures=[fig] if fig else [],
                tables=[grouped],
                module='trend_analysis',
            )

    df = frames['reports_df']
    if df.empty or df['year'].dropna().empty:
        return _analysis_result('Trend Analysis', ['No year-based trend data is available yet.'], module='trend_analysis')
    grouped = (
        df[df['year'].notna()]
        .groupby('year', as_index=False)[['payments_total', 'donations_total', 'project_spending_total']]
        .sum()
        .sort_values('year')
    )
    fig = px.line(grouped, x='year', y=['payments_total', 'donations_total', 'project_spending_total'], title='Yearly Trend: Payments, Donations, Project Spending') if px else None
    return _analysis_result('Trend Analysis', ['Yearly trend over enabled financial/project fields.'], figures=[fig] if fig else [], tables=[grouped], module='trend_analysis')


def module_ranking(frames, domain_context=None):
    px = get_plotly_express()
    domain = (domain_context or {}).get('domain')
    domain_df = resolve_domain_frame(domain, frames)
    if domain and (domain_df is None or domain_df.empty):
        return domain_no_data_result(domain, 'Ranking', 'ranking')
    if domain_df is not None and not domain_df.empty:
        group_col = pick_group_column(domain_df)
        metric_col = pick_metric_column(domain_df)
        if group_col and metric_col:
            grouped = (
                domain_df.groupby(group_col, as_index=False)[metric_col]
                .sum()
                .sort_values(metric_col, ascending=False)
                .head(20)
            )
            title = f"Ranking ({domain.replace('_', ' ').title()})"
            fig = px.bar(grouped, x=group_col, y=metric_col, title=title) if px else None
            return _analysis_result(
                title,
                [f"Top {group_col.replace('_', ' ')} ranked by {metric_col.replace('_', ' ')}."],
                figures=[fig] if fig else [],
                tables=[grouped],
                module='ranking',
            )

    donations = frames['donations_df']
    if donations.empty:
        return _analysis_result('Ranking', ['No donation data is available for ranking.'], module='ranking')
    grouped = donations.groupby('donor_name', as_index=False)['amount'].sum().sort_values('amount', ascending=False).head(20)
    fig = px.bar(grouped, x='donor_name', y='amount', title='Top Donors by Total Donation Amount') if px else None
    return _analysis_result('Ranking', ['Top donor ranking from enabled donation fields.'], figures=[fig] if fig else [], tables=[grouped], module='ranking')


def module_compare_groups(frames, domain_context=None):
    px = get_plotly_express()
    domain = (domain_context or {}).get('domain')
    domain_df = resolve_domain_frame(domain, frames)
    if domain and (domain_df is None or domain_df.empty):
        return domain_no_data_result(domain, 'Compare Groups', 'compare_groups')
    if domain_df is not None and not domain_df.empty:
        group_col = pick_group_column(domain_df)
        metric_col = pick_metric_column(domain_df)
        if group_col and metric_col:
            grouped = (
                domain_df.groupby(group_col, as_index=False)[metric_col]
                .sum()
                .sort_values(metric_col, ascending=False)
                .head(20)
            )
            title = f"Compare Groups ({domain.replace('_', ' ').title()})"
            fig = px.bar(grouped, x=group_col, y=metric_col, barmode='group', title=title) if px else None
            return _analysis_result(
                title,
                [f"Group comparison by {group_col.replace('_', ' ')} and {metric_col.replace('_', ' ')}."],
                figures=[fig] if fig else [],
                tables=[grouped],
                module='compare_groups',
            )

    df = frames['reports_df']
    if df.empty:
        return _analysis_result('Compare Groups', ['No report data available for group comparison.'], module='compare_groups')
    grouped = df.groupby('scope', as_index=False)[['payments_total', 'donations_total', 'project_beneficiaries_total']].sum()
    fig = px.bar(grouped, x='scope', y=['payments_total', 'donations_total'], barmode='group', title='Scope Comparison: Payments vs Donations') if px else None
    return _analysis_result('Compare Groups', ['Comparison between scopes using enabled financial fields.'], figures=[fig] if fig else [], tables=[grouped], module='compare_groups')


def module_completeness_check(frames):
    df = frames['reports_df']
    if df.empty:
        return _analysis_result('Completeness Check', ['No report data available for completeness checks.'], module='completeness_check')
    columns = ['pbo_name', 'scope', 'reporting_period_start', 'contact_phone_count', 'contact_email_count']
    results = []
    total_rows = len(df) or 1
    for column in columns:
        if column not in df.columns:
            continue
        if column in {'contact_phone_count', 'contact_email_count'}:
            filled = int((df[column] > 0).sum())
        else:
            filled = int(df[column].astype(str).str.strip().replace({'': None, 'NAN': None, 'NONE': None}).notna().sum())
        results.append({
            'field': column,
            'filled_rows': filled,
            'completeness_percent': round((filled / total_rows) * 100, 2),
        })
    pandas_module = get_pandas()
    table = pandas_module.DataFrame(results).sort_values('completeness_percent', ascending=False)
    summary = [f"Completeness computed across {total_rows} submissions using enabled fields."]
    return _analysis_result('Completeness Check', summary, tables=[table], module='completeness_check')


def module_anomaly_detection(frames, domain_context=None):
    domain = (domain_context or {}).get('domain')
    domain_df = resolve_domain_frame(domain, frames)
    if domain and (domain_df is None or domain_df.empty):
        return domain_no_data_result(domain, 'Anomaly Detection', 'anomaly_detection')
    if domain_df is not None and not domain_df.empty:
        metric_col = pick_metric_column(domain_df)
        if metric_col:
            metric = domain_df[metric_col].fillna(0)
            q1 = metric.quantile(0.25)
            q3 = metric.quantile(0.75)
            iqr = q3 - q1
            lower = q1 - 1.5 * iqr
            upper = q3 + 1.5 * iqr
            columns = [col for col in ['report_id', 'year', 'scope', metric_col] if col in domain_df.columns]
            anomalies = domain_df[(metric < lower) | (metric > upper)][columns]
            summary = [
                f"Detected {len(anomalies)} outlier rows in {domain.replace('_', ' ')} using {metric_col.replace('_', ' ')}.",
                f"Lower bound: {lower:,.2f}; Upper bound: {upper:,.2f}.",
            ]
            return _analysis_result('Anomaly Detection', summary, tables=[anomalies.head(50)], module='anomaly_detection')

    df = frames['reports_df']
    if df.empty:
        return _analysis_result('Anomaly Detection', ['No report data available for anomaly checks.'], module='anomaly_detection')
    metric = df['payments_total'].fillna(0)
    q1 = metric.quantile(0.25)
    q3 = metric.quantile(0.75)
    iqr = q3 - q1
    lower = q1 - 1.5 * iqr
    upper = q3 + 1.5 * iqr
    anomalies = df[(metric < lower) | (metric > upper)][['report_id', 'scope', 'payments_total', 'donations_total']]
    summary = [
        f"Detected {len(anomalies)} payment outlier rows using IQR thresholds.",
        f"Lower bound: {lower:,.2f}; Upper bound: {upper:,.2f}.",
    ]
    return _analysis_result('Anomaly Detection', summary, tables=[anomalies.head(50)], module='anomaly_detection')


def module_validation_check(frames, domain_context=None):
    domain = (domain_context or {}).get('domain')
    domain_df = resolve_domain_frame(domain, frames)
    if domain and (domain_df is None or domain_df.empty):
        return domain_no_data_result(domain, 'Validation & Duplicates', 'validation_check')
    if domain_df is not None and not domain_df.empty:
        null_counts = []
        total_rows = len(domain_df) or 1
        for column in domain_df.columns:
            missing = int(domain_df[column].isna().sum())
            if domain_df[column].dtype == object:
                missing += int((domain_df[column].astype(str).str.strip() == '').sum())
            null_counts.append({
                'column': column,
                'missing_rows': missing,
                'missing_percent': round((missing / total_rows) * 100, 2),
            })
        pandas_module = get_pandas()
        validation_df = pandas_module.DataFrame(null_counts).sort_values('missing_percent', ascending=False)
        summary = [f"Validation profile for {domain.replace('_', ' ')} across {total_rows} rows."]
        return _analysis_result('Validation & Duplicates', summary, tables=[validation_df.head(40)], module='validation_check')

    df = frames['reports_df']
    if df.empty:
        return _analysis_result('Validation & Duplicates', ['No report data available for validation checks.'], module='validation_check')

    duplicate_names = (
        df[df['pbo_name'].astype(str).str.strip() != '']
        .groupby('pbo_name', as_index=False)
        .size()
        .query('size > 1')
        .sort_values('size', ascending=False)
    )
    validation_rows = []
    for _, row in df.iterrows():
        issues = []
        if not str(row.get('pbo_name') or '').strip():
            issues.append('missing_pbo_name')
        if not str(row.get('scope') or '').strip():
            issues.append('missing_scope')
        if not str(row.get('reporting_period_start') or '').strip():
            issues.append('missing_start_date')
        if int(row.get('contact_email_count') or 0) == 0:
            issues.append('missing_contact_email')
        if issues:
            validation_rows.append({
                'report_id': row.get('report_id'),
                'issues': ", ".join(issues),
            })
    pandas_module = get_pandas()
    validation_df = pandas_module.DataFrame(validation_rows)
    summary = [
        f"Duplicate PBO names found: {len(duplicate_names)}.",
        f"Validation issue rows found: {len(validation_df)}.",
    ]
    return _analysis_result('Validation & Duplicates', summary, tables=[duplicate_names.head(30), validation_df.head(50)], module='validation_check')


def build_sector_factor_matrix(frames):
    reports_df = frames.get('reports_df')
    payments_df = frames.get('payments_df')
    donations_df = frames.get('donations_df')
    projects_df = frames.get('projects_df')
    counties_df = frames.get('counties_df')
    trainings_df = frames.get('trainings_df')
    officials_df = frames.get('officials_df')

    rows = []

    def add_row(category, metric, status, value, note):
        rows.append({
            'category': category,
            'metric': metric,
            'status': status,
            'value': value,
            'note': note,
        })

    def add_no_data(category, metric, note='No data available in current dataset.'):
        add_row(category, metric, 'no_data', 'No data available', note)

    if reports_df is None or reports_df.empty:
        add_no_data('Quantitative / Organizational Profile', 'Number of registered NGOs')
        add_no_data('Quantitative / Financial Analysis', 'Total funding received')
        add_no_data('Quantitative / Program & Output', 'Number of beneficiaries reached')
        add_no_data('Quantitative / Employment & Capacity', 'Number of staff (Kenyan vs international)')
        add_no_data('Quantitative / Risk & Compliance', 'Compliance status')
        add_no_data('Quantitative / Trend Analysis', 'Registration and funding trends')
        add_no_data('Qualitative / Sector Trends & Insights', 'Emerging focus areas')
        add_no_data('Qualitative / Governance & Accountability', 'Governance quality indicators')
        add_no_data('Qualitative / Operational Challenges', 'Operational challenge signals')
        add_no_data('Qualitative / Impact Assessment', 'Impact narrative indicators')
        add_no_data('Qualitative / Risk Analysis', 'Risk narrative indicators')
        add_no_data('Qualitative / Stakeholder Perception', 'Donor/public confidence proxy')
        return rows

    total_reports = int(len(reports_df))
    add_row('Quantitative / Organizational Profile', 'Number of registered NGOs', 'available', total_reports, 'Count of report submissions in dataset.')

    status_series = reports_df['workflow_status'].fillna('').astype(str).str.strip().str.lower() if 'workflow_status' in reports_df.columns else None
    if status_series is not None and status_series.astype(bool).any():
        active_count = int(status_series.isin({'submitted', 'validated', 'in_review', 'approved'}).sum())
        inactive_count = int(total_reports - active_count)
        add_row('Quantitative / Organizational Profile', 'Active vs inactive NGOs', 'available', f"{active_count} active / {inactive_count} inactive", 'Active proxy uses workflow status.')
    else:
        add_no_data('Quantitative / Organizational Profile', 'Active vs inactive NGOs')

    if counties_df is not None and not counties_df.empty and 'county' in counties_df.columns:
        county_grouped = counties_df.groupby('county', as_index=False).size().sort_values('size', ascending=False).head(5)
        county_preview = ', '.join(f"{str(row['county'])} ({int(row['size'])})" for _, row in county_grouped.iterrows())
        add_row('Quantitative / Organizational Profile', 'Geographic distribution (counties/regions)', 'available', county_preview or 'Available', 'Top county distribution by records.')
    else:
        add_no_data('Quantitative / Organizational Profile', 'Geographic distribution (counties/regions)')

    if projects_df is not None and not projects_df.empty and 'sector' in projects_df.columns:
        sector_grouped = projects_df.groupby('sector', as_index=False).size().sort_values('size', ascending=False).head(5)
        sector_preview = ', '.join(f"{str(row['sector'])} ({int(row['size'])})" for _, row in sector_grouped.iterrows())
        add_row('Quantitative / Organizational Profile', 'Sector classification', 'available', sector_preview or 'Available', 'Top sectors from project implementation records.')
    else:
        add_no_data('Quantitative / Organizational Profile', 'Sector classification')

    total_donations = float(donations_df['amount'].fillna(0).sum()) if donations_df is not None and not donations_df.empty and 'amount' in donations_df.columns else None
    total_payments = float(payments_df['total_amount'].fillna(0).sum()) if payments_df is not None and not payments_df.empty and 'total_amount' in payments_df.columns else None
    if total_donations is not None or total_payments is not None:
        add_row(
            'Quantitative / Financial Analysis',
            'Total funding received',
            'available',
            f"Donations KES {float(total_donations or 0):,.2f}; Payments KES {float(total_payments or 0):,.2f}",
            'Funding totals from donations/payments tables.',
        )
    else:
        add_no_data('Quantitative / Financial Analysis', 'Total funding received')

    if donations_df is not None and not donations_df.empty and {'category', 'amount'}.issubset(set(donations_df.columns)):
        category_series = donations_df['category'].fillna('').astype(str).str.lower()
        donor_total = float(donations_df['amount'].fillna(0).sum())
        government_total = float(donations_df[category_series.str.contains('gov')]['amount'].fillna(0).sum())
        private_total = float(donations_df[category_series.str.contains('private|corporate')]['amount'].fillna(0).sum())
        donor_pct = (donor_total / donor_total * 100.0) if donor_total > 0 else 0.0
        add_row(
            'Quantitative / Financial Analysis',
            'Funding sources (donor/government/private)',
            'available',
            f"Donor {donor_pct:.1f}% | Government KES {government_total:,.2f} | Private KES {private_total:,.2f}",
            'Source split inferred from donation categories.',
        )
    else:
        add_no_data('Quantitative / Financial Analysis', 'Funding sources (donor/government/private)')

    if payments_df is not None and not payments_df.empty and {'description', 'total_amount'}.issubset(set(payments_df.columns)):
        desc_series = payments_df['description'].fillna('').astype(str).str.lower()
        program_mask = desc_series.str.contains('project|program|implementation|beneficiar|training|grant')
        admin_mask = desc_series.str.contains('admin|office|rent|salary|allowance|utility|transport|audit')
        program_total = float(payments_df[program_mask]['total_amount'].fillna(0).sum())
        admin_total = float(payments_df[admin_mask]['total_amount'].fillna(0).sum())
        if program_total > 0 or admin_total > 0:
            add_row('Quantitative / Financial Analysis', 'Program vs administrative costs', 'available', f"Program KES {program_total:,.2f} / Administrative KES {admin_total:,.2f}", 'Derived using payment-description keyword groups.')
        else:
            add_no_data('Quantitative / Financial Analysis', 'Program vs administrative costs', 'Payment descriptions do not contain enough classification signals.')
    else:
        add_no_data('Quantitative / Financial Analysis', 'Program vs administrative costs')

    if {'payments_total', 'donations_total'}.issubset(set(reports_df.columns)):
        total_budget_series = reports_df['payments_total'].fillna(0) + reports_df['donations_total'].fillna(0)
        add_row('Quantitative / Financial Analysis', 'Average budget per NGO', 'available', f"KES {float(total_budget_series.mean()):,.2f}", 'Average of donations + payments per report row.')
    else:
        add_no_data('Quantitative / Financial Analysis', 'Average budget per NGO')

    if {'year', 'donations_total'}.issubset(set(reports_df.columns)):
        yearly = reports_df[reports_df['year'].notna()].groupby('year', as_index=False)['donations_total'].sum().sort_values('year')
        if len(yearly) >= 2:
            prev_value = float(yearly.iloc[-2]['donations_total'])
            curr_value = float(yearly.iloc[-1]['donations_total'])
            if prev_value > 0:
                growth_pct = ((curr_value - prev_value) / prev_value) * 100.0
                add_row('Quantitative / Financial Analysis', 'Year-on-year funding growth/decline', 'available', f"{growth_pct:.2f}%", f"{int(yearly.iloc[-2]['year'])} to {int(yearly.iloc[-1]['year'])} change.")
            else:
                add_no_data('Quantitative / Financial Analysis', 'Year-on-year funding growth/decline', 'Previous-year funding is zero, growth rate undefined.')
        else:
            add_no_data('Quantitative / Financial Analysis', 'Year-on-year funding growth/decline')
    else:
        add_no_data('Quantitative / Financial Analysis', 'Year-on-year funding growth/decline')

    if projects_df is not None and not projects_df.empty and 'beneficiaries' in projects_df.columns:
        total_beneficiaries = int(projects_df['beneficiaries'].fillna(0).sum())
        add_row('Quantitative / Program & Output', 'Number of beneficiaries reached', 'available', total_beneficiaries, 'Sum of beneficiaries from project implementations.')
    else:
        add_no_data('Quantitative / Program & Output', 'Number of beneficiaries reached')

    if projects_df is not None and not projects_df.empty:
        add_row('Quantitative / Program & Output', 'Number of projects implemented', 'available', int(len(projects_df)), 'Count of project implementation rows.')
    else:
        add_no_data('Quantitative / Program & Output', 'Number of projects implemented')

    if projects_df is not None and not projects_df.empty and 'completion_status' in projects_df.columns:
        completion_series = projects_df['completion_status'].fillna('').astype(str).str.lower()
        completed_count = int(completion_series.str.contains('complete').sum())
        completion_rate = (completed_count / len(projects_df) * 100.0) if len(projects_df) > 0 else 0.0
        add_row('Quantitative / Program & Output', 'Project completion rates (%)', 'available', f"{completion_rate:.2f}%", 'Completed status inferred from completion_status text.')
    else:
        add_no_data('Quantitative / Program & Output', 'Project completion rates (%)')

    if projects_df is not None and not projects_df.empty and {'spending', 'beneficiaries'}.issubset(set(projects_df.columns)):
        total_spending = float(projects_df['spending'].fillna(0).sum())
        total_beneficiaries = float(projects_df['beneficiaries'].fillna(0).sum())
        if total_beneficiaries > 0:
            add_row('Quantitative / Program & Output', 'Cost per beneficiary', 'available', f"KES {total_spending / total_beneficiaries:,.2f}", 'Project spending divided by beneficiaries.')
        else:
            add_no_data('Quantitative / Program & Output', 'Cost per beneficiary', 'Beneficiary total is zero.')
    else:
        add_no_data('Quantitative / Program & Output', 'Cost per beneficiary')

    add_no_data('Quantitative / Program & Output', 'Coverage (urban vs rural)', 'No explicit urban/rural field available in current dataset.')

    if {'staff_kenyan_current', 'staff_foreign_current'}.issubset(set(reports_df.columns)):
        kenyan_staff = int(reports_df['staff_kenyan_current'].fillna(0).sum())
        foreign_staff = int(reports_df['staff_foreign_current'].fillna(0).sum())
        add_row('Quantitative / Employment & Capacity', 'Number of staff (Kenyan vs international)', 'available', f"{kenyan_staff} Kenyan / {foreign_staff} International", 'Aggregated from report staffing fields.')
    else:
        add_no_data('Quantitative / Employment & Capacity', 'Number of staff (Kenyan vs international)')

    if {'volunteers_kenyan_current', 'volunteers_foreign_current'}.issubset(set(reports_df.columns)):
        total_volunteers = int(reports_df['volunteers_kenyan_current'].fillna(0).sum() + reports_df['volunteers_foreign_current'].fillna(0).sum())
        add_row('Quantitative / Employment & Capacity', 'Volunteers count', 'available', total_volunteers, 'Current-year volunteers and interns totals.')
    else:
        add_no_data('Quantitative / Employment & Capacity', 'Volunteers count')

    if trainings_df is not None and not trainings_df.empty and 'participants' in trainings_df.columns:
        add_row('Quantitative / Employment & Capacity', 'Staff training numbers', 'available', int(trainings_df['participants'].fillna(0).sum()), 'Total participants from training records.')
    else:
        add_no_data('Quantitative / Employment & Capacity', 'Staff training numbers')

    if {'staff_kenyan_left', 'staff_foreign_left', 'staff_kenyan_current', 'staff_foreign_current'}.issubset(set(reports_df.columns)):
        left_total = float(reports_df['staff_kenyan_left'].fillna(0).sum() + reports_df['staff_foreign_left'].fillna(0).sum())
        current_total = float(reports_df['staff_kenyan_current'].fillna(0).sum() + reports_df['staff_foreign_current'].fillna(0).sum())
        if current_total > 0:
            turnover_rate = (left_total / current_total) * 100.0
            add_row('Quantitative / Employment & Capacity', 'Staff turnover rate', 'available', f"{turnover_rate:.2f}%", 'Proxy using left vs current staff totals.')
        else:
            add_no_data('Quantitative / Employment & Capacity', 'Staff turnover rate', 'Current staff total is zero.')
    else:
        add_no_data('Quantitative / Employment & Capacity', 'Staff turnover rate')

    if status_series is not None and status_series.astype(bool).any():
        compliant_count = int(status_series.isin({'approved', 'validated'}).sum())
        non_compliant_count = int(total_reports - compliant_count)
        add_row('Quantitative / Risk & Compliance', 'Number of NGOs compliant vs non-compliant', 'available', f"{compliant_count} compliant / {non_compliant_count} non-compliant", 'Compliance proxy based on workflow status.')
    else:
        add_no_data('Quantitative / Risk & Compliance', 'Number of NGOs compliant vs non-compliant')

    if status_series is not None and status_series.astype(bool).any():
        filed_count = int(status_series.isin({'submitted', 'validated', 'in_review', 'approved', 'returned'}).sum())
        filing_rate = (filed_count / total_reports) * 100.0 if total_reports > 0 else 0.0
        add_row('Quantitative / Risk & Compliance', 'Filing rates (annual returns submitted %)', 'available', f"{filing_rate:.2f}%", 'Filed proxy derived from workflow statuses.')
    else:
        add_no_data('Quantitative / Risk & Compliance', 'Filing rates (annual returns submitted %)')

    add_no_data('Quantitative / Risk & Compliance', 'Suspicious transactions reported', 'No suspicious-transactions field is present in current dataset.')

    if 'risk_score' in reports_df.columns and reports_df['risk_score'].notna().any():
        risk_avg = float(reports_df['risk_score'].dropna().mean())
        risk_max = float(reports_df['risk_score'].dropna().max())
        add_row('Quantitative / Risk & Compliance', 'Risk scores (Inherent/Residual proxy)', 'available', f"avg={risk_avg:.2f}, max={risk_max:.2f}", 'Risk score available from report-level field.')
    else:
        add_no_data('Quantitative / Risk & Compliance', 'Risk scores (Inherent/Residual proxy)')

    add_no_data('Quantitative / Risk & Compliance', 'Cash vs digital transactions ratio', 'No transaction-mode split field available in current dataset.')

    if 'year' in reports_df.columns and reports_df['year'].notna().any():
        yearly_counts = reports_df[reports_df['year'].notna()].groupby('year', as_index=False).size().sort_values('year')
        if len(yearly_counts) >= 2:
            start_count = int(yearly_counts.iloc[0]['size'])
            end_count = int(yearly_counts.iloc[-1]['size'])
            add_row('Quantitative / Trend Analysis', 'Growth in NGO registrations over time', 'available', f"{start_count} -> {end_count}", 'Report counts by reporting year.')
        else:
            add_no_data('Quantitative / Trend Analysis', 'Growth in NGO registrations over time')
    else:
        add_no_data('Quantitative / Trend Analysis', 'Growth in NGO registrations over time')

    if {'year', 'donations_total'}.issubset(set(reports_df.columns)) and reports_df['year'].notna().any():
        funding_trend = reports_df[reports_df['year'].notna()].groupby('year', as_index=False)['donations_total'].sum().sort_values('year').tail(5)
        preview = ', '.join(f"{int(row['year'])}: {float(row['donations_total']):,.2f}" for _, row in funding_trend.iterrows())
        add_row('Quantitative / Trend Analysis', 'Funding trends over 3-5 years', 'available', preview or 'Available', 'Donation totals by year.')
    else:
        add_no_data('Quantitative / Trend Analysis', 'Funding trends over 3-5 years')

    if projects_df is not None and not projects_df.empty and {'year', 'sector'}.issubset(set(projects_df.columns)):
        sector_year = projects_df[projects_df['year'].notna()].groupby(['year', 'sector'], as_index=False).size()
        if not sector_year.empty:
            latest_year = int(sector_year['year'].max())
            latest_slice = sector_year[sector_year['year'] == latest_year].sort_values('size', ascending=False).head(3)
            preview = ', '.join(f"{str(row['sector'])} ({int(row['size'])})" for _, row in latest_slice.iterrows())
            add_row('Quantitative / Trend Analysis', 'Sector growth trends', 'available', preview or 'Available', f"Top sectors in {latest_year}.")
        else:
            add_no_data('Quantitative / Trend Analysis', 'Sector growth trends')
    else:
        add_no_data('Quantitative / Trend Analysis', 'Sector growth trends')

    if 'risk_score' in reports_df.columns and reports_df['risk_score'].notna().any() and counties_df is not None and not counties_df.empty:
        risk_df = reports_df[['report_id', 'risk_score']].dropna()
        county_risk = counties_df.merge(risk_df, on='report_id', how='inner')
        if not county_risk.empty:
            grouped = county_risk.groupby('county', as_index=False)['risk_score'].mean().sort_values('risk_score', ascending=False).head(5)
            preview = ', '.join(f"{str(row['county'])} ({float(row['risk_score']):.2f})" for _, row in grouped.iterrows())
            add_row('Quantitative / Trend Analysis', 'Geographic risk trends', 'available', preview or 'Available', 'Average risk score by county.')
        else:
            add_no_data('Quantitative / Trend Analysis', 'Geographic risk trends')
    else:
        add_no_data('Quantitative / Trend Analysis', 'Geographic risk trends')

    if projects_df is not None and not projects_df.empty and 'sector' in projects_df.columns:
        dominant_sector = projects_df['sector'].fillna('').astype(str).str.strip()
        dominant_sector = dominant_sector[dominant_sector != '']
        if not dominant_sector.empty:
            top_sector = dominant_sector.value_counts().index[0]
            add_row('Qualitative / Sector Trends & Insights', 'Emerging focus areas', 'available', top_sector, 'Dominant sector based on project records.')
        else:
            add_no_data('Qualitative / Sector Trends & Insights', 'Emerging focus areas')
    else:
        add_no_data('Qualitative / Sector Trends & Insights', 'Emerging focus areas')
    add_no_data('Qualitative / Sector Trends & Insights', 'Shift in donor priorities', 'No longitudinal donor-priority labels available.')
    add_no_data('Qualitative / Sector Trends & Insights', 'Impact of global events', 'No event-tag field available for causal attribution.')

    if officials_df is not None and not officials_df.empty:
        role_count = int(officials_df['role'].fillna('').astype(str).str.strip().ne('').sum()) if 'role' in officials_df.columns else int(len(officials_df))
        add_row('Qualitative / Governance & Accountability', 'Quality of governance structures', 'available', f"{role_count} governance-role entries", 'Derived from officials records.')
        add_row('Qualitative / Governance & Accountability', 'Board effectiveness', 'available', 'Proxy available', 'Use meeting/election fields in pbo_reports for deeper analysis.')
        add_row('Qualitative / Governance & Accountability', 'Transparency practices', 'available', 'Proxy available', 'Officials contact details and signatures can be audited.')
        add_row('Qualitative / Governance & Accountability', 'Internal controls strength', 'available', 'Proxy available', 'Audit and governance fields available for checks.')
    else:
        add_no_data('Qualitative / Governance & Accountability', 'Quality of governance structures')
        add_no_data('Qualitative / Governance & Accountability', 'Board effectiveness')
        add_no_data('Qualitative / Governance & Accountability', 'Transparency practices')
        add_no_data('Qualitative / Governance & Accountability', 'Internal controls strength')

    if reports_df['donations_total'].fillna(0).sum() > 0 and reports_df['payments_total'].fillna(0).sum() > 0:
        add_row('Qualitative / Operational Challenges', 'Funding instability', 'available', 'Signal detected', 'Funding inflow and expenditure variability can be profiled.')
    else:
        add_no_data('Qualitative / Operational Challenges', 'Funding instability')
    add_no_data('Qualitative / Operational Challenges', 'Regulatory burdens', 'No direct burden index field available.')
    if trainings_df is not None and not trainings_df.empty:
        add_row('Qualitative / Operational Challenges', 'Capacity gaps', 'available', 'Training records present', 'Training entries can be used as capacity proxy.')
    else:
        add_no_data('Qualitative / Operational Challenges', 'Capacity gaps')
    add_no_data('Qualitative / Operational Challenges', 'Technology adoption issues', 'No explicit digital-adoption indicator in current dataset.')

    if projects_df is not None and not projects_df.empty and {'beneficiaries', 'spending'}.issubset(set(projects_df.columns)):
        beneficiaries_total = int(projects_df['beneficiaries'].fillna(0).sum())
        spending_total = float(projects_df['spending'].fillna(0).sum())
        add_row('Qualitative / Impact Assessment', 'Community impact stories (proxy)', 'available', f"{beneficiaries_total} beneficiaries, KES {spending_total:,.2f} spending", 'Narrative proxy from project outputs.')
        add_row('Qualitative / Impact Assessment', 'Long-term sustainability of interventions', 'available', 'Proxy available', 'Project duration and completion fields can support sustainability narratives.')
    else:
        add_no_data('Qualitative / Impact Assessment', 'Community impact stories (proxy)')
        add_no_data('Qualitative / Impact Assessment', 'Long-term sustainability of interventions')
    add_no_data('Qualitative / Impact Assessment', 'Case studies of successful programs', 'No case-study narrative field in current dataset.')

    if 'risk_score' in reports_df.columns and reports_df['risk_score'].notna().any():
        add_row('Qualitative / Risk Analysis', 'Terror financing and AML/CFT risks', 'available', 'Risk score available', 'Use risk_score and transaction tables for follow-up risk narratives.')
    else:
        add_no_data('Qualitative / Risk Analysis', 'Terror financing and AML/CFT risks')
    add_no_data('Qualitative / Risk Analysis', 'Fraud and misuse of funds', 'No explicit fraud-case variable available.')
    if counties_df is not None and not counties_df.empty:
        add_row('Qualitative / Risk Analysis', 'Geographic risk exposure', 'available', 'County coverage available', 'Combine counties with risk score for geographic exposure narrative.')
    else:
        add_no_data('Qualitative / Risk Analysis', 'Geographic risk exposure')
    if frames.get('collaboration_df') is not None and not frames.get('collaboration_df').empty:
        add_row('Qualitative / Risk Analysis', 'Partner risks', 'available', 'Partner linkage records available', 'Collaboration/networking fields can inform partner-risk checks.')
    else:
        add_no_data('Qualitative / Risk Analysis', 'Partner risks')

    add_no_data('Qualitative / Stakeholder Perception', 'Donor confidence levels', 'No survey/sentiment field available.')
    add_no_data('Qualitative / Stakeholder Perception', 'Public trust in NGOs', 'No public-trust index field available.')
    add_no_data('Qualitative / Stakeholder Perception', 'Government-NGO relationships', 'No perception/satisfaction field available.')

    return rows


def merge_sector_factor_matrix_into_result(result_payload, frames, question=None):
    pandas_module = get_pandas()
    factor_rows = build_sector_factor_matrix(frames)
    factor_df = pandas_module.DataFrame(factor_rows)
    if factor_df.empty:
        return {'total': 0, 'available': 0, 'no_data': 0, 'no_data_matches': [], 'available_matches': []}

    available_count = int((factor_df['status'] == 'available').sum())
    no_data_count = int((factor_df['status'] == 'no_data').sum())
    total_count = int(len(factor_df))

    no_data_matches = []
    available_matches = []
    question_text = (question or '').strip().lower()
    question_tokens = set(tokenize_analysis_text(question_text))
    if question_text and question_tokens:
        scored_rows = []
        for row in factor_rows:
            metric = str(row.get('metric') or '').strip()
            category = str(row.get('category') or '').strip()
            status = str(row.get('status') or '').strip().lower()
            metric_tokens = {token for token in tokenize_analysis_text(metric) if len(token) > 3}
            category_tokens = {token for token in tokenize_analysis_text(category) if len(token) > 3}
            overlap_metric = metric_tokens & question_tokens
            overlap_category = category_tokens & question_tokens
            score = len(overlap_metric) * 2 + len(overlap_category)
            if metric and metric.lower() in question_text:
                score += 4
            if score < 2:
                continue
            scored_rows.append((score, metric, status))
        scored_rows.sort(key=lambda item: item[0], reverse=True)
        seen_metrics = set()
        for _, metric, status in scored_rows:
            if metric in seen_metrics:
                continue
            seen_metrics.add(metric)
            if status == 'no_data':
                no_data_matches.append(metric)
            else:
                available_matches.append(metric)
            if len(no_data_matches) + len(available_matches) >= 6:
                break

    result_payload.setdefault('tables', [])
    result_payload['tables'].append(factor_df)
    result_payload.setdefault('summary_lines', [])
    result_payload['summary_lines'].append(
        f"Sector factor check: {available_count} available, {no_data_count} with no data."
    )
    if no_data_matches:
        result_payload['summary_lines'].append(
            "No data available for: " + ", ".join(no_data_matches[:3]) + "."
        )
    result_payload['factor_summary'] = {
        'total': total_count,
        'available': available_count,
        'no_data': no_data_count,
        'no_data_matches': no_data_matches[:5],
        'available_matches': available_matches[:5],
        'explanation': (
            f"Checked {total_count} sector-analysis metrics. "
            f"{available_count} have usable data; {no_data_count} currently have no data."
        ),
    }
    return result_payload['factor_summary']


def run_data_analysis_module(intent, question, frames, domain_context=None):
    intent = (intent or 'aggregation').strip().lower()
    operation_result = module_column_operations(frames, question, domain_context=domain_context)
    if operation_result is not None:
        return operation_result
    if is_county_query(question):
        return module_county_cross_database(frames, question, domain_context=domain_context)
    module_map = {
        'aggregation': lambda _frames: module_aggregation(_frames, question=question, domain_context=domain_context),
        'distribution': lambda _frames: module_distribution(_frames, question, domain_context=domain_context),
        'trend_analysis': lambda _frames: module_trend_analysis(_frames, domain_context=domain_context),
        'ranking': lambda _frames: module_ranking(_frames, domain_context=domain_context),
        'compare_groups': lambda _frames: module_compare_groups(_frames, domain_context=domain_context),
        'completeness_check': module_completeness_check,
        'anomaly_detection': lambda _frames: module_anomaly_detection(_frames, domain_context=domain_context),
        'duplicates': lambda _frames: module_validation_check(_frames, domain_context=domain_context),
        'validation_check': lambda _frames: module_validation_check(_frames, domain_context=domain_context),
    }
    func = module_map.get(intent, lambda _frames: module_aggregation(_frames, question=question, domain_context=domain_context))
    return func(frames)


def cleanup_data_analysis_result_files(max_age_seconds=3600 * 24):
    DATA_ANALYSIS_RESULT_DIR.mkdir(parents=True, exist_ok=True)
    now_ts = time.time()
    for file_path in DATA_ANALYSIS_RESULT_DIR.glob('*.html'):
        try:
            if now_ts - file_path.stat().st_mtime > max_age_seconds:
                file_path.unlink(missing_ok=True)
        except Exception:
            continue


def render_data_analysis_html(question, intent, result_payload, similar_questions=None):
    similar_questions = similar_questions or []
    title = html.escape(result_payload.get('title') or 'Data Analysis Result')
    safe_question = html.escape(question or '')
    summary_html = ''.join(f"<li>{html.escape(line)}</li>" for line in (result_payload.get('summary_lines') or []))

    figure_html_parts = []
    include_plotly = True
    for figure in (result_payload.get('figures') or []):
        if figure is None:
            continue
        try:
            figure_html_parts.append(figure.to_html(full_html=False, include_plotlyjs='cdn' if include_plotly else False))
            include_plotly = False
        except Exception:
            continue

    table_html_parts = []
    for table in (result_payload.get('tables') or []):
        try:
            table_html_parts.append(table.to_html(index=False, classes='table table-sm table-striped'))
        except Exception:
            continue

    similar_html = ''.join(
        f"<li>{html.escape(item.get('question', ''))} <small>({html.escape(str(item.get('intent', '')))}, score={item.get('score', 0):.2f})</small></li>"
        for item in similar_questions
    )
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{title}</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 24px; background: #f8fafc; color: #0f172a; }}
    .wrap {{ max-width: 1200px; margin: 0 auto; }}
    .card {{ background: #fff; border: 1px solid #dbe3f4; border-radius: 12px; padding: 16px; margin-bottom: 16px; }}
    .muted {{ color: #475569; }}
    .table {{ width: 100%; border-collapse: collapse; }}
    .table th, .table td {{ border: 1px solid #e2e8f0; padding: 6px; text-align: left; }}
    .actions {{ display: flex; justify-content: flex-end; margin-bottom: 12px; }}
    .btn {{ border: 1px solid #0f766e; background: #0f766e; color: #fff; border-radius: 8px; padding: 8px 12px; cursor: pointer; font-size: 0.85rem; }}
    .btn:hover {{ background: #0d5e58; border-color: #0d5e58; }}
    @media print {{
      body {{ margin: 0; background: #fff; }}
      .card {{ border: 1px solid #ddd; box-shadow: none; break-inside: avoid; }}
      .no-print {{ display: none !important; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="actions no-print">
      <button class="btn" onclick="window.print()">Download PDF report</button>
    </div>
    <div class="card">
      <h2>{title}</h2>
      <p class="muted"><strong>Intent:</strong> {html.escape(intent or 'aggregation')}</p>
      <p class="muted"><strong>Question:</strong> {safe_question}</p>
      <ul>{summary_html}</ul>
    </div>
    <div class="card">{''.join(figure_html_parts) or '<p class="muted">No visualization generated for this result.</p>'}</div>
    <div class="card">{''.join(table_html_parts) or '<p class="muted">No table output generated for this result.</p>'}</div>
    <div class="card">
      <h4>Similar Training Questions</h4>
      <ul>{similar_html or '<li>None</li>'}</ul>
    </div>
  </div>
</body>
</html>"""


def save_data_analysis_html_result(question, intent, result_payload, similar_questions=None):
    cleanup_data_analysis_result_files()
    DATA_ANALYSIS_RESULT_DIR.mkdir(parents=True, exist_ok=True)
    token = uuid.uuid4().hex
    html_content = render_data_analysis_html(question, intent, result_payload, similar_questions=similar_questions)
    file_path = DATA_ANALYSIS_RESULT_DIR / f'{token}.html'
    file_path.write_text(html_content, encoding='utf-8')
    return token, str(file_path)


def answer_data_analysis_question(question, page_key=DATA_ANALYSIS_PAGE_KEY):
    field_context = get_field_help_context_for_page(page_key)
    lowered = (question or '').lower()
    for label in field_context.get('disallowed_summary', []):
        if label and label.lower() in lowered:
            warning = (
                f"`{label}` is disabled/out of scope for enabled-field analytics. "
                "Ask using enabled fields (scope, counties, payments, projects, donors, staffing, trainings)."
            )
            return {
                'answer': warning,
                'intent': 'validation_check',
                'confidence': 1.0,
                'blocked': True,
                'result_token': None,
                'result_path': None,
                'result_payload': _analysis_result('Validation Warning', [warning], module='validation_check'),
                'similar_questions': [],
            }

    prediction = predict_data_analysis_intent(question, page_key=page_key)
    domain_context = infer_data_analysis_domain(question)
    frames = collect_data_analysis_frames()
    if is_county_query(question):
        question_tokens = set(tokenize_analysis_text(question))
        semantic_terms = {'county', 'counties'}
        if question_tokens & ORGANIZATION_TERMS:
            semantic_terms.add('organizations')
        location_terms = {
            token for token in question_tokens
            if token in {'nairobi', 'mombasa', 'nakuru', 'kiambu', 'uasin', 'gishu', 'kisumu', 'machakos'}
        }
        matched_terms = sorted(set((domain_context.get('matched_terms') or []) + list(semantic_terms) + list(location_terms)))
        matched_columns = ['pbo_name', 'county', 'counties', 'counties_of_operation', 'project_county']
        domain_context = {
            'domain': GLOBAL_COUNTY_SCAN_DOMAIN,
            'source': 'cross_table_county_scan_org' if is_organization_count_query(question) else 'cross_table_county_scan',
            'score': max(float(domain_context.get('score') or 0.0), 9.9),
            'matched_terms': matched_terms,
            'matched_columns': matched_columns,
        }
    result_payload = run_data_analysis_module(prediction['intent'], question, frames, domain_context=domain_context)
    factor_summary = merge_sector_factor_matrix_into_result(result_payload, frames, question=question)
    similar_questions = retrieve_similar_data_analysis_questions(question, page_key=page_key, top_n=5)
    token, path = save_data_analysis_html_result(question, prediction['intent'], result_payload, similar_questions=similar_questions)
    summary_lines = result_payload.get('summary_lines') or []
    answer = summary_lines[0] if summary_lines else f"Analysis completed using module `{result_payload.get('module')}`."
    module_name = (result_payload.get('module') or '').strip().lower()
    if (
        factor_summary.get('no_data_matches')
        and not factor_summary.get('available_matches')
        and module_name not in {'county_cross_database', 'column_operations'}
    ):
        answer = "No data available for " + ", ".join(factor_summary['no_data_matches'][:2]) + " in the current dataset."
    if domain_context.get('domain'):
        answer = f"{answer} (Routed domain: {domain_context.get('domain')})"
    return {
        'answer': answer,
        'intent': prediction['intent'],
        'confidence': prediction.get('confidence', 0.0),
        'blocked': False,
        'domain': domain_context.get('domain'),
        'domain_source': domain_context.get('source'),
        'domain_score': float(domain_context.get('score') or 0.0),
        'domain_terms': domain_context.get('matched_terms', []),
        'domain_columns': domain_context.get('matched_columns', []),
        'result_token': token,
        'result_path': path,
        'result_payload': result_payload,
        'similar_questions': similar_questions,
        'factor_summary': factor_summary,
    }


def ensure_default_admin():
    if running_schema_command():
        return
    admin_email = (os.getenv('ADMIN_USER_EMAIL') or '').strip().lower()
    admin_password = os.getenv('ADMIN_USER_PASSWORD')
    admin_name = (os.getenv('ADMIN_USER_NAME') or '').strip() or None

    if not admin_email or not admin_password:
        print("⚠️ Skipping default admin setup: ADMIN_USER_EMAIL or ADMIN_USER_PASSWORD is missing.")
        return
    try:
        admin_user = find_preferred_user_by_email(admin_email)

        if admin_user is None:
            admin_user = User(
                email=admin_email,
                full_name=admin_name,
                role='admin',
                is_superadmin=True,
                can_manage_all_records=True,
                is_authorized=True,
            )
            admin_user.set_password(admin_password, mark_changed=False)
            db.session.add(admin_user)
            db.session.commit()
            print(f"✅ Default admin user created: {admin_email}")
            return

        updated = False
        if admin_user.role != 'admin':
            admin_user.role = 'admin'
            updated = True
        if admin_name and admin_user.full_name != admin_name:
            admin_user.full_name = admin_name
            updated = True
        if not getattr(admin_user, 'is_superadmin', False):
            admin_user.is_superadmin = True
            updated = True
        if not getattr(admin_user, 'can_manage_all_records', False):
            admin_user.can_manage_all_records = True
            updated = True
        if not admin_user.is_authorized:
            admin_user.is_authorized = True
            updated = True
        if not admin_user.password_hash or not admin_user.check_password(admin_password):
            admin_user.set_password(admin_password, mark_changed=False)
            updated = True

        if updated:
            db.session.commit()
            print(f"✅ Default admin credentials refreshed for: {admin_email}")
    except Exception as exc:
        if 'could not translate host name' in str(exc) or 'Could not connect' in str(exc) or 'OperationalError' in type(exc).__name__:
            print(f"⚠️  Temporary DB connection issue during default admin setup: {exc}")
            print("⚠️  Continuing startup in limited mode. Default admin sync will resume when DB is available.")
            db.session.rollback()
            return
        raise


bootstrap_database_on_startup()


if env_flag('GOOGLE_DRIVE_ENABLED', False):
    start_noon_backup_worker()


def build_sector_report_rows(reports):
    """Return flattened report rows suitable for review, cleaning, and downstream analysis."""
    disabled_template = {
        "chairperson_name": None,
        "chairperson_nationality": None,
        "chairperson_gender": None,
        "chairperson_email": None,
        "chairperson_residence": None,
        "chairperson_phone": None,
        "chairperson_kra_pin": None,
        "chairperson_qualification": None,
        "secretary_name": None,
        "secretary_nationality": None,
        "secretary_gender": None,
        "secretary_email": None,
        "secretary_residence": None,
        "secretary_phone": None,
        "secretary_kra_pin": None,
        "secretary_qualification": None,
        "treasurer_name": None,
        "treasurer_nationality": None,
        "treasurer_gender": None,
        "treasurer_email": None,
        "treasurer_residence": None,
        "treasurer_phone": None,
        "treasurer_kra_pin": None,
        "treasurer_qualification": None,
        "chairperson_signature": None,
        "secretary_signature": None,
        "treasurer_signature": None,
        "submitter_fullname": None,
        "submitter_signature": None,
        "submission_date": None,
    }

    rows = []
    for report in reports:
        project_rows = list(report.project_implementations)
        bank_rows = list(report.bank_accounts)
        projects_carried = list(report.projects_carried_out)
        project_row_countries = []
        project_country_sectors = {"KENYA": [], "OTHER": [], "BOTH": [], "UNSPECIFIED": []}

        for index, project in enumerate(project_rows):
            spent_kenya = legacy_zero_float(report, project.amount_spent_kenya)
            spent_other = legacy_zero_float(report, project.amount_spent_other)
            if spent_kenya > 0 and spent_other > 0:
                project_country = "BOTH"
            elif spent_kenya > 0:
                project_country = "KENYA"
            elif spent_other > 0:
                project_country = "OTHER"
            else:
                project_country = "UNSPECIFIED"

            project_country_sectors[project_country].append(project.sector or "UNNAMED")
            project_row_countries.append(
                f"PROJECT {index + 1}: {(project.sector or 'UNNAMED')} - {project_country}"
            )

        row = {
            "updated_by_name": report_updater_display(report),
            "report_id": report.id,
            "pbo_name": report.pbo_name,
            "scope": report.scope,
            "reporting_period_start": format_date(report.reporting_period_start),
            "reporting_period_end": format_date(report.reporting_period_end),
            "return_date": return_date_display(report.return_date),
            "contact_name": report.contact_name,
            "contact_telephone": report.contact_telephone,
            "countries_of_operation": report.countries_of_operation,
            "counties": report.counties,
            "cash_balance_previous_year": legacy_zero_output_value(report, report.cash_balance_previous_year),
            "cash_bank_balance": legacy_zero_output_value(report, report.cash_bank_balance),
            "audited": report.audited,
            "staff_kenyan_current": legacy_zero_output_value(report, report.staff_kenyan_current),
            "staff_foreign_current": legacy_zero_output_value(report, report.staff_foreign_current),
            "volunteers_kenyan_current": legacy_zero_output_value(report, report.volunteers_kenyan_current),
            "volunteers_foreign_current": legacy_zero_output_value(report, report.volunteers_foreign_current),
            "project_implementation_method": report.project_implementation_method,
            "project_implementation_count": len(project_rows),
            "project_row_names": "; ".join(
                f"PROJECT {index + 1}: {project.sector or 'UNNAMED'}"
                for index, project in enumerate(project_rows)
            ) or None,
            "project_row_country": "; ".join(project_row_countries) or None,
            "project_country_row_count": ", ".join(
                f"{country}: {'; '.join(sectors)}"
                for country, sectors in project_country_sectors.items()
                if sectors
            ) or None,
            "project_row_counties": "; ".join(
                project.county for project in project_rows if project.county
            ) or None,
            "project_row_beneficiary_categories": "; ".join(
                project.vulnerable_group for project in project_rows if project.vulnerable_group
            ) or None,
            "project_row_completion_statuses": "; ".join(
                project.completion_status for project in project_rows if project.completion_status
            ) or None,
            "project_beneficiaries_total": sum(legacy_zero_int(report, project.beneficiaries_no) for project in project_rows),
            "project_spending_per_county_total": sum(resolve_project_spending_amount(project, report=report) for project in project_rows),
            "project_duration_years_total": sum(legacy_zero_float(report, project.duration_years) for project in project_rows),
            "project_amount_spent_kenya_total": sum(legacy_zero_float(report, project.amount_spent_kenya) for project in project_rows),
            "project_amount_spent_other_total": sum(legacy_zero_float(report, project.amount_spent_other) for project in project_rows),
            "projects_carried_out_count": len(projects_carried),
            "projects_carried_out_names": "; ".join(
                f"PROJECT {index + 1}: {project.sector or 'UNNAMED'}"
                for index, project in enumerate(projects_carried)
            ) or None,
            "donations_total": sum(legacy_zero_float(report, donation.amount) for donation in report.donations),
            "payments_total": sum(
                legacy_zero_float(report, payment.kenya_amount) + legacy_zero_float(report, payment.other_amount)
                for payment in report.payments
            ),
            "bank_accounts_count": len(bank_rows),
            "bank_account_names": "; ".join(
                bank.bank_name for bank in bank_rows if bank.bank_name
            ) or None,
            "created_at": format_datetime(report.created_at),
            "updated_at": format_datetime(report.updated_at),
            **disabled_template,
        }
        rows.append(legacy_zero_output_structure(report, row))
    return rows


def get_sector_report_columns():
    allowed_columns = [
        ("updated_by_name", "Updated By"),
        ("report_id", "Report ID"),
        ("pbo_name", "PBO Name"),
        ("scope", "Scope"),
        ("reporting_period_start", "Reporting Start"),
        ("reporting_period_end", "Reporting End"),
        ("return_date", "Return Date"),
        ("contact_name", "Contact Name"),
        ("contact_telephone", "Contact Telephone"),
        ("countries_of_operation", "Countries of Operation"),
        ("counties", "Counties"),
        ("cash_balance_previous_year", "Opening Balance"),
        ("cash_bank_balance", "Closing Balance"),
        ("audited", "Audited"),
        ("staff_kenyan_current", "Kenyan Staff Current"),
        ("staff_foreign_current", "Foreign Staff Current"),
        ("volunteers_kenyan_current", "Kenyan Volunteers Current"),
        ("volunteers_foreign_current", "Foreign Volunteers Current"),
        ("project_implementation_method", "Implementation Method"),
        ("project_implementation_count", "Project Implementation Row Count"),
        ("project_row_names", "Project Row Names"),
        ("project_row_country", "Project Country Per Row"),
        ("project_country_row_count", "Project Country Row Count"),
        ("project_row_counties", "Project Row Counties"),
        ("project_row_beneficiary_categories", "Project Beneficiary Categories"),
        ("project_row_completion_statuses", "Project Completion Statuses"),
        ("project_beneficiaries_total", "Project Beneficiaries Total"),
        ("project_spending_per_county_total", "Project Spending Per County Total"),
        ("project_duration_years_total", "Project Duration Years Total"),
        ("project_amount_spent_kenya_total", "Project Amount Spent Kenya Total"),
        ("project_amount_spent_other_total", "Project Amount Spent Other Total"),
        ("projects_carried_out_count", "Projects Carried Out Row Count"),
        ("projects_carried_out_names", "Projects Carried Out Names"),
        ("donations_total", "Donations Total"),
        ("payments_total", "Payments Total"),
        ("bank_accounts_count", "Bank Accounts Count"),
        ("bank_account_names", "Bank Account Names"),
        ("created_at", "Created At"),
        ("updated_at", "Updated At"),
    ]
    disabled_columns = [
        ("chairperson_name", "Chairperson Name"),
        ("chairperson_nationality", "Chairperson Nationality"),
        ("chairperson_gender", "Chairperson Gender"),
        ("chairperson_email", "Chairperson Email"),
        ("chairperson_residence", "Chairperson Residence"),
        ("chairperson_phone", "Chairperson Phone"),
        ("chairperson_kra_pin", "Chairperson KRA PIN"),
        ("chairperson_qualification", "Chairperson Qualification"),
        ("secretary_name", "Secretary Name"),
        ("secretary_nationality", "Secretary Nationality"),
        ("secretary_gender", "Secretary Gender"),
        ("secretary_email", "Secretary Email"),
        ("secretary_residence", "Secretary Residence"),
        ("secretary_phone", "Secretary Phone"),
        ("secretary_kra_pin", "Secretary KRA PIN"),
        ("secretary_qualification", "Secretary Qualification"),
        ("treasurer_name", "Treasurer Name"),
        ("treasurer_nationality", "Treasurer Nationality"),
        ("treasurer_gender", "Treasurer Gender"),
        ("treasurer_email", "Treasurer Email"),
        ("treasurer_residence", "Treasurer Residence"),
        ("treasurer_phone", "Treasurer Phone"),
        ("treasurer_kra_pin", "Treasurer KRA PIN"),
        ("treasurer_qualification", "Treasurer Qualification"),
        ("chairperson_signature", "Chairperson Signature"),
        ("secretary_signature", "Secretary Signature"),
        ("treasurer_signature", "Treasurer Signature"),
        ("submitter_fullname", "Submitter Full Name"),
        ("submitter_signature", "Submitter Signature"),
        ("submission_date", "Submission Date"),
    ]
    money_columns = {
        "cash_balance_previous_year",
        "cash_bank_balance",
        "donations_total",
        "payments_total",
        "project_spending_per_county_total",
        "project_amount_spent_kenya_total",
        "project_amount_spent_other_total",
    }
    return allowed_columns, disabled_columns, money_columns


def enforce_nominatim_rate_limit():
    """Return seconds to wait if rate limited, otherwise 0."""
    global NOMINATIM_LAST_REQUEST_AT
    with NOMINATIM_RATE_LOCK:
        now = time.monotonic()
        elapsed = now - NOMINATIM_LAST_REQUEST_AT
        if elapsed < NOMINATIM_MIN_INTERVAL_SECONDS:
            return NOMINATIM_MIN_INTERVAL_SECONDS - elapsed
        NOMINATIM_LAST_REQUEST_AT = now
    return 0.0


def get_town_from_postal_code(postal_code):
    cached = PostalCodeCache.query.filter_by(postal_code=postal_code).first()
    if cached and cached.town:
        return cached.town, None, 200

    retry_after = enforce_nominatim_rate_limit()
    if retry_after > 0:
        return None, {
            "message": "Rate limit exceeded. Try again shortly.",
            "retry_after": round(retry_after, 2),
        }, 429

    def fetch_nominatim(params):
        try:
            response = requests.get(
                NOMINATIM_URL,
                params=params,
                headers=NOMINATIM_HEADERS,
                timeout=10,
            )
        except requests.RequestException:
            return None, {"message": "Address service unavailable"}, 503

        if response.status_code == 429:
            return None, {"message": "Address service rate limited"}, 429

        if response.status_code != 200:
            return None, {"message": "Address service unavailable"}, 503

        try:
            data = response.json()
        except ValueError:
            return None, {"message": "Address service unavailable"}, 503

        return data, None, 200

    base_params = {
        'countrycodes': 'ke',
        'format': 'json',
        'addressdetails': 1,
        'limit': 1,
    }

    data, error, status = fetch_nominatim({
        **base_params,
        'postalcode': postal_code,
    })
    if error:
        return None, error, status

    if not data:
        data, error, status = fetch_nominatim({
            **base_params,
            'q': f"{postal_code} Kenya",
        })
        if error:
            return None, error, status

    if not data:
        return None, {"message": "Invalid postal code"}, 400

    address = data[0].get('address', {})
    town = (
        address.get('town')
        or address.get('city')
        or address.get('municipality')
        or address.get('county')
    )

    if not town:
        return None, {"message": "Town not found"}, 404

    if cached:
        cached.town = town
    else:
        cached = PostalCodeCache(postal_code=postal_code, town=town)
        db.session.add(cached)
    db.session.commit()

    return town, None, 200



from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'
login_manager.refresh_view = 'login'
login_manager.needs_refresh_message = 'Please reauthenticate to access this page.'
login_manager.needs_refresh_message_category = 'info'

PUBLIC_ENDPOINTS = {
    'login',
    'signup',
    'forgot_password',
    'reset_password',
    'global_field_help_context',
    'static',
}

AUTH_REDIRECT_ENDPOINTS = {
    'login',
    'signup',
    'forgot_password',
    'reset_password',
}



@login_manager.user_loader
def load_user(user_id):
    try:
        user_id = int(user_id)
    except (TypeError, ValueError):
        return None
    user = db.session.get(User, user_id)
    # Only return user if they are authorized
    if user and user.is_authorized:
        return user
    return None


@app.before_request
def redirect_to_canonical_host():
    if not env_flag('FORCE_CANONICAL_HOST', False):
        return None

    canonical_host = normalize_trusted_host_candidate(os.getenv('CANONICAL_HOSTNAME'))
    if not canonical_host:
        return None

    request_host = normalize_trusted_host_candidate(request.host)
    if not request_host or request_host == canonical_host:
        return None

    target = urlunparse((
        request.scheme,
        canonical_host,
        request.path,
        '',
        request.query_string.decode('utf-8', errors='ignore'),
        '',
    ))
    return redirect(target, code=308)


@app.before_request
def enforce_access_rules():
    endpoint = request.endpoint or ''
    normalized_endpoint = endpoint.lower()

    if request.method == 'OPTIONS':
        return None

    if normalized_endpoint.endswith('.static') or normalized_endpoint in PUBLIC_ENDPOINTS:
        if current_user.is_authenticated and normalized_endpoint in AUTH_REDIRECT_ENDPOINTS:
            return redirect(url_for('home'))
        return None

    if current_user.is_authenticated:
        return None

    if request.path.startswith('/api/') or request.is_json:
        return jsonify({
            'error': 'Authentication required',
            'login_url': url_for('login'),
        }), 401

    return redirect(url_for('login', next=request.url))


@app.before_request
def validate_user_session():
    """Ensure session data matches the current user for proper isolation."""
    endpoint = (request.endpoint or '').lower()
    if endpoint in PUBLIC_ENDPOINTS or endpoint == 'logout' or endpoint.endswith('.static'):
        return

    if current_user.is_authenticated:
        # Verify session user data matches current user
        session_user_id = session.get('user_id')
        session_user_email = session.get('user_email')

        if session_user_id and session_user_email:
            if str(current_user.id) != str(session_user_id) or current_user.email != session_user_email:
                # Session mismatch - logout user to prevent conflicts
                logout_user()
                session.clear()
                flash('Session conflict detected. Please log in again.', 'warning')
                return redirect(url_for('login'))
        else:
            # Rehydrate missing session data for remembered authenticated users.
            session['user_id'] = current_user.id
            session['user_email'] = current_user.email
            session.modified = True


def establish_user_session(user, *, fresh=True):
    session.clear()
    session.permanent = True
    login_user(user, remember=True, fresh=fresh)
    session['user_id'] = user.id
    session['user_email'] = user.email
    session['login_time'] = time.time()
    session.modified = True


@app.after_request
def add_no_store_headers(response):
    endpoint = (request.endpoint or '').lower()
    sensitive_endpoint_prefixes = (
        'login',
        'logout',
        'change_password',
        'forgot_password',
        'reset_password',
        'report_detail',
        'report_edit',
        'my_files',
        'reports_list',
    )
    is_html_response = (response.mimetype or '').startswith('text/html')
    is_sensitive_response = (
        is_html_response
        or current_user.is_authenticated
        or endpoint.startswith(sensitive_endpoint_prefixes)
    )

    if is_sensitive_response:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0, private'
        response.headers['CDN-Cache-Control'] = 'no-store'
        response.headers['Cloudflare-CDN-Cache-Control'] = 'no-store'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        existing_vary = response.headers.get('Vary')
        response.headers['Vary'] = 'Cookie' if not existing_vary else f'{existing_vary}, Cookie'

    if endpoint == 'logout':
        response.headers['Clear-Site-Data'] = '"cache", "cookies", "storage"'

    return response


def enforce_password_change():
    # Password change enforcement disabled to allow users to continue working
    # even after password resets. All user data is preserved in the database.
    return None

# Decorator for admin only
from functools import wraps
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if (
            not current_user.is_authenticated
            or current_user.role != 'admin'
            or not current_user.is_authorized
        ):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def is_superadmin(user):
    return bool(user and getattr(user, 'role', None) == 'admin' and getattr(user, 'is_superadmin', False))


def can_manage_all_records(user):
    return bool(
        user
        and getattr(user, 'role', None) == 'admin'
        and (
            getattr(user, 'is_superadmin', False)
            or getattr(user, 'can_manage_all_records', False)
        )
    )


QUALIFYING_REPORT_WORK_ACTIVITY_ACTIONS = frozenset({
    'admin_designate_updated',
    'admin_report_created',
    'admin_report_updated',
    'ocr_batch_report_imported',
    'ocr_report_imported',
    'report_created',
    'report_file_deleted',
    'report_file_uploaded',
    'report_imported',
    'report_updated',
})
QUALIFYING_REPORT_WORK_FIELD_CHANGE_ACTIONS = frozenset({
    'admin_designate_updated',
    'admin_report_created',
    'admin_report_updated',
    'report_created',
    'report_imported',
    'report_updated',
})


def form_submitted_report_source_filter():
    return func.lower(func.trim(func.coalesce(PBOReport.data_source, 'form'))) == 'form'


def report_submission_timestamp(report):
    if report is None:
        return None
    return getattr(report, 'submitted_at', None) or getattr(report, 'created_at', None)


def build_form_submitted_report_index(user_ids=None, activity_start=None, activity_end=None):
    normalized_user_ids = sorted({
        user_id for user_id in (
            coerce_legacy_int(raw_user_id) for raw_user_id in (user_ids or [])
        )
        if user_id is not None
    })
    submitted_by_user_id = defaultdict(dict)

    query = db.session.query(
        PBOReport.user_id.label('user_id'),
        PBOReport.id.label('report_id'),
        PBOReport.submitted_at.label('submitted_at'),
        PBOReport.created_at.label('created_at'),
    ).filter(
        PBOReport.user_id.isnot(None),
        form_submitted_report_source_filter(),
    )
    if normalized_user_ids:
        query = query.filter(PBOReport.user_id.in_(normalized_user_ids))
    if activity_start and activity_end:
        query = query.filter(
            or_(
                PBOReport.submitted_at.between(activity_start, activity_end),
                PBOReport.created_at.between(activity_start, activity_end),
            )
        )

    for row in query.all():
        normalized_user_id = coerce_legacy_int(row.user_id)
        normalized_report_id = coerce_legacy_int(row.report_id)
        if normalized_user_id is None or normalized_report_id is None:
            continue
        submitted_at = row.submitted_at or row.created_at
        existing_value = submitted_by_user_id[normalized_user_id].get(normalized_report_id)
        if normalized_report_id not in submitted_by_user_id[normalized_user_id]:
            submitted_by_user_id[normalized_user_id][normalized_report_id] = submitted_at
            continue
        if submitted_at is not None and (existing_value is None or submitted_at > existing_value):
            submitted_by_user_id[normalized_user_id][normalized_report_id] = submitted_at

    return submitted_by_user_id


def build_worked_report_last_touched_index(user_ids=None, activity_start=None, activity_end=None, report_ids=None):
    normalized_user_ids = sorted({
        user_id for user_id in (
            coerce_legacy_int(raw_user_id) for raw_user_id in (user_ids or [])
        )
        if user_id is not None
    })
    normalized_report_ids = sorted({
        report_id for report_id in (
            coerce_legacy_int(raw_report_id) for raw_report_id in (report_ids or [])
        )
        if report_id is not None
    })
    last_touched_by_user_id = defaultdict(dict)

    def apply_common_filters(query, user_column, report_column, time_column):
        query = query.filter(user_column.isnot(None), report_column.isnot(None))
        if normalized_user_ids:
            query = query.filter(user_column.in_(normalized_user_ids))
        if normalized_report_ids:
            query = query.filter(report_column.in_(normalized_report_ids))
        if activity_start and activity_end:
            query = query.filter(time_column >= activity_start, time_column <= activity_end)
        return query

    def register_touch(user_id, report_id, touched_at):
        normalized_user_id = coerce_legacy_int(user_id)
        normalized_report_id = coerce_legacy_int(report_id)
        if normalized_user_id is None or normalized_report_id is None:
            return
        existing_value = last_touched_by_user_id[normalized_user_id].get(normalized_report_id)
        if normalized_report_id not in last_touched_by_user_id[normalized_user_id]:
            last_touched_by_user_id[normalized_user_id][normalized_report_id] = touched_at
            return
        if touched_at is not None and (existing_value is None or touched_at > existing_value):
            last_touched_by_user_id[normalized_user_id][normalized_report_id] = touched_at

    owner_query = apply_common_filters(
        db.session.query(
            PBOReport.user_id.label('user_id'),
            PBOReport.id.label('report_id'),
            PBOReport.created_at.label('touched_at'),
        ),
        PBOReport.user_id,
        PBOReport.id,
        PBOReport.created_at,
    )
    for row in owner_query.all():
        register_touch(row.user_id, row.report_id, row.touched_at)

    modifier_query = apply_common_filters(
        db.session.query(
            PBOReport.last_modified_by_id.label('user_id'),
            PBOReport.id.label('report_id'),
            PBOReport.updated_at.label('touched_at'),
        ),
        PBOReport.last_modified_by_id,
        PBOReport.id,
        PBOReport.updated_at,
    )
    for row in modifier_query.all():
        register_touch(row.user_id, row.report_id, row.touched_at)

    activity_query = apply_common_filters(
        db.session.query(
            UserActivityLog.user_id.label('user_id'),
            UserActivityLog.report_id.label('report_id'),
            UserActivityLog.created_at.label('touched_at'),
        ).filter(UserActivityLog.action.in_(tuple(QUALIFYING_REPORT_WORK_ACTIVITY_ACTIONS))),
        UserActivityLog.user_id,
        UserActivityLog.report_id,
        UserActivityLog.created_at,
    )
    for row in activity_query.all():
        register_touch(row.user_id, row.report_id, row.touched_at)

    field_change_query = apply_common_filters(
        db.session.query(
            FieldChangeLog.user_id.label('user_id'),
            FieldChangeLog.report_id.label('report_id'),
            FieldChangeLog.created_at.label('touched_at'),
        ).filter(FieldChangeLog.action.in_(tuple(QUALIFYING_REPORT_WORK_FIELD_CHANGE_ACTIONS))),
        FieldChangeLog.user_id,
        FieldChangeLog.report_id,
        FieldChangeLog.created_at,
    )
    for row in field_change_query.all():
        register_touch(row.user_id, row.report_id, row.touched_at)

    uploaded_file_query = apply_common_filters(
        db.session.query(
            UploadedFile.uploaded_by_id.label('user_id'),
            UploadedFile.report_id.label('report_id'),
            UploadedFile.created_at.label('touched_at'),
        ),
        UploadedFile.uploaded_by_id,
        UploadedFile.report_id,
        UploadedFile.created_at,
    )
    for row in uploaded_file_query.all():
        register_touch(row.user_id, row.report_id, row.touched_at)

    return last_touched_by_user_id


def combine_worked_report_last_touched(last_touched_index, user_ids):
    combined = {}
    normalized_user_ids = sorted({
        user_id for user_id in (
            coerce_legacy_int(raw_user_id) for raw_user_id in (user_ids or [])
        )
        if user_id is not None
    })
    for user_id in normalized_user_ids:
        for report_id, touched_at in (last_touched_index.get(user_id) or {}).items():
            existing_value = combined.get(report_id)
            if report_id not in combined:
                combined[report_id] = touched_at
                continue
            if touched_at is not None and (existing_value is None or touched_at > existing_value):
                combined[report_id] = touched_at
    return combined


def get_user_worked_report_ids(user):
    if not user or not getattr(user, 'is_authenticated', False):
        return set()

    worked_index = build_worked_report_last_touched_index(user_ids=[user.id])
    return set((worked_index.get(user.id) or {}).keys())


def user_has_worked_on_report(user, report):
    if not user or not getattr(user, 'is_authenticated', False) or report is None:
        return False
    if report.user_id == user.id or report.last_modified_by_id == user.id:
        return True

    return (
        db.session.query(UserActivityLog.id)
        .filter(
            UserActivityLog.user_id == user.id,
            UserActivityLog.report_id == report.id,
            UserActivityLog.action.in_(tuple(QUALIFYING_REPORT_WORK_ACTIVITY_ACTIONS)),
        )
        .first()
        is not None
        or db.session.query(FieldChangeLog.id)
        .filter(
            FieldChangeLog.user_id == user.id,
            FieldChangeLog.report_id == report.id,
            FieldChangeLog.action.in_(tuple(QUALIFYING_REPORT_WORK_FIELD_CHANGE_ACTIONS)),
        )
        .first()
        is not None
        or db.session.query(UploadedFile.id)
        .filter(
            UploadedFile.uploaded_by_id == user.id,
            UploadedFile.report_id == report.id,
        )
        .first()
        is not None
    )


def can_edit_report_record(user, report):
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    if report is None:
        return False
    if can_manage_all_records(user):
        return True
    return user_has_worked_on_report(user, report)


def can_view_report_record(user, report):
    return can_edit_report_record(user, report)


def can_view_uploaded_file_record(user, uploaded_file):
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    if uploaded_file is None:
        return False
    if can_manage_all_records(user):
        return True
    if uploaded_file.uploaded_by_id == user.id:
        return True
    if uploaded_file.report is not None:
        return can_view_report_record(user, uploaded_file.report)
    return False


def require_report_access(report, write=False):
    allowed = can_edit_report_record(current_user, report) if write else can_view_report_record(current_user, report)
    if allowed:
        return None
    action = 'edit' if write else 'view'
    log_user_activity(
        f'report_{action}_denied',
        report=report,
        user=current_user if current_user.is_authenticated else None,
        summary=f'{getattr(current_user, "email", "Anonymous")} was denied permission to {action} report #{report.id}',
    )
    db.session.commit()
    flash('You do not have permission to access that record.', 'danger')
    return redirect(url_for('reports_list'))


def delete_report_uploaded_files(report):
    if report is None:
        return

    for uploaded_file in list(report.uploaded_files or []):
        log_user_activity(
            'report_file_deleted',
            report=report,
            summary=build_deleted_file_activity_summary(uploaded_file),
            user=current_user if current_user.is_authenticated else None,
        )
        storage_path = uploaded_file.storage_path
        if storage_path and os.path.exists(storage_path):
            try:
                os.remove(storage_path)
            except OSError:
                app.logger.warning('Could not remove uploaded file from disk: %s', storage_path)


def require_superadmin():
    if is_superadmin(current_user):
        return None
    log_user_activity(
        'superadmin_access_denied',
        user=current_user if current_user.is_authenticated else None,
        summary=f'{getattr(current_user, "email", "Unknown")} attempted a superadmin-only action',
    )
    db.session.commit()
    flash('Only a superadmin can perform that action.', 'danger')
    return redirect(url_for('authorize_accounts'))


def require_admin_editor_access(target_user):
    if not target_user:
        abort(404)
    if getattr(target_user, 'role', None) == 'admin' and target_user.id != current_user.id and not is_superadmin(current_user):
        log_user_activity(
            'admin_user_edit_denied',
            user=current_user if current_user.is_authenticated else None,
            summary=f'{getattr(current_user, "email", "Unknown")} attempted to edit admin account {getattr(target_user, "email", "Unknown")}',
        )
        db.session.commit()
        flash('Only a superadmin can edit another admin account.', 'danger')
        return redirect(url_for('authorize_accounts'))
    return None



# Home page route: display and process the main form14
@app.route('/api/section-b-preview', methods=['POST'])
@login_required
def section_b_preview():
    payload = request.get_json(silent=True) or {}
    preview = build_section_b_preview(
        opening_balance=payload.get('cash_balance_previous_year'),
        donation_amounts=payload.get('donor_amounts') or [],
        iga_amounts=payload.get('iga_amounts') or [],
        payment_kenya_amounts=payload.get('payment_kenya_amounts') or [],
        payment_other_amounts=payload.get('payment_other_amounts') or [],
        project_kenya_amounts=payload.get('project_kenya_amounts') or [],
        project_other_amounts=payload.get('project_other_amounts') or [],
    )
    return jsonify(preview)


@app.route('/', methods=['GET', 'POST'])
@login_required
def home():
        submission_token = get_or_create_submission_token()
        if request.method == 'POST':
            try:
                posted_submission_token = to_upper((request.form.get('submission_token') or '').strip()) or submission_token
                existing_report = PBOReport.query.filter_by(submission_token=posted_submission_token).first()
                if existing_report is not None:
                    if form_submission_matches_existing_report(existing_report):
                        flash('This form was already submitted. Showing the saved report instead.', 'info')
                        return redirect(url_for('report_detail', report_id=existing_report.id, clear_form14_draft=1))
                    log_user_activity(
                        'form14_stale_submission_token_detected',
                        report=existing_report,
                        summary=summarize_stale_submission_token(existing_report, posted_submission_token),
                    )
                    posted_submission_token = reset_submission_token()

                # Create a new report instance
                report = PBOReport()
                report.submission_token = posted_submission_token

                # Scalar fields
                set_reporting_period_fields(
                    report,
                    request.form.get('reporting_period_start'),
                    request.form.get('reporting_period_end'),
                )
                report.return_date = parse_return_date(request.form.get('return_date'))
                report.pbo_name = get_form_upper('pbo_name')
                report.pbo_registration_number = get_form_upper('pbo_registration_number')
                report.pbo_registration_date = parse_date(request.form.get('pbo_registration_date'))

                report.kra_pin = None
                report.postal_address = get_form_upper('postal_address')
                report.physical_address = get_form_upper('physical_address')
                report.telephone = get_form_phone('telephone', 'telephone_country_code')
                report.cell_phone = get_form_phone('cell_phone', 'cellphone_country_code')
                report.email = get_form_upper('email')
                report.website = get_form_upper('website')
                report.social_media = get_form_upper('social_media')

                report.contact_name = get_form_upper('contact_name')
                report.contact_position = get_form_upper('contact_position')
                report.contact_telephone = get_form_phone_joined('contact_telephone', 'contact_country_code')
                report.contact_email = get_form_text_joined_upper('contact_email')
                report.contact_nationality = request.form.get('contact_nationality')
                report.contact_gender = request.form.get('contact_gender')
                assign_pbo_name_normalized(report)

                report.registration_number = get_form_upper('registration_number')
                report.pin_number = None
                report.date_of_registration = parse_date(request.form.get('date_of_registration'))
                report.scope = normalize_scope(request.form.get('scope'))

                countries_of_operation = get_form_list_upper('country_of_operation')
                report.countries_of_operation = ", ".join(countries_of_operation) if countries_of_operation else None
                counties = get_form_list_upper('county')
                report.counties = ", ".join(counties) if counties else None

                report.audited = get_form_upper('audited')
                report.assets_stolen = get_form_upper('assets_stolen')
                report.cash_balance_previous_year = parse_float(request.form.get('cash_balance_previous_year'))
                report.cash_bank_balance = parse_float(request.form.get('cash_bank_balance'))

                # Section C - Staff/volunteer numbers
                report.staff_kenyan_prev = parse_int(get_form_list_item('staff_kenyan[]', 0))
                report.staff_kenyan_current = parse_int(get_form_list_item('staff_kenyan[]', 1))
                report.staff_kenyan_came_in = parse_int(get_form_list_item('staff_kenyan[]', 2))
                report.staff_kenyan_left = parse_int(get_form_list_item('staff_kenyan[]', 3))
                report.staff_foreign_prev = parse_int(get_form_list_item('staff_foreign[]', 0))
                report.staff_foreign_current = parse_int(get_form_list_item('staff_foreign[]', 1))
                report.staff_foreign_came_in = parse_int(get_form_list_item('staff_foreign[]', 2))
                report.staff_foreign_left = parse_int(get_form_list_item('staff_foreign[]', 3))
                report.staff_other_kenyan_prev = parse_int(get_form_list_item('staff_other_kenyan[]', 0))
                report.staff_other_kenyan_current = parse_int(get_form_list_item('staff_other_kenyan[]', 1))
                report.staff_other_foreign_prev = parse_int(get_form_list_item('staff_other_foreign[]', 0))
                report.staff_other_foreign_current = parse_int(get_form_list_item('staff_other_foreign[]', 1))
                report.volunteers_kenyan_prev = parse_int(get_form_list_item('volunteers_kenyan[]', 0))
                report.volunteers_kenyan_current = parse_int(get_form_list_item('volunteers_kenyan[]', 1))
                report.volunteers_foreign_prev = parse_int(get_form_list_item('volunteers_foreign[]', 0))
                report.volunteers_foreign_current = parse_int(get_form_list_item('volunteers_foreign[]', 1))

                # Section D - Project implementation method
                implementation = get_form_list_upper('implementation[]')
                report.project_implementation_method = ", ".join(implementation) if implementation else None

                # Contributions in kind
                report.local_material = get_checkbox_value("local_material")
                report.local_material_amount = parse_float(request.form.get("local_material_amount"))
                report.local_labour = get_checkbox_value("local_labour")
                report.local_labour_amount = parse_float(request.form.get("local_labour_amount"))
                report.local_financial = get_checkbox_value("local_financial")
                report.local_financial_amount = parse_float(request.form.get("local_financial_amount"))
                report.local_other = get_checkbox_value("local_other")
                report.local_other_specify = get_form_upper("local_other_specify")
                report.local_other_amount = parse_float(request.form.get("local_other_amount"))
                government_sections_inactive = get_inactive_section_flag(
                    "government_sections_inactive",
                    "government_section_inactive",
                )
                if government_sections_inactive:
                    report.gov_tax_waiver = None
                    report.gov_tax_waiver_amount = None
                    report.gov_other = None
                    report.gov_other_specify = None
                    report.gov_other_amount = None
                else:
                    report.gov_tax_waiver = get_checkbox_value("gov_tax_waiver")
                    report.gov_tax_waiver_amount = parse_float(request.form.get("gov_tax_waiver_amount"))
                    report.gov_other = get_checkbox_value("gov_other")
                    report.gov_other_specify = get_form_upper("gov_other_specify")
                    report.gov_other_amount = parse_float(request.form.get("gov_other_amount"))

                # Section B - Assets
                assets_items = get_form_list_upper('assets_item[]')
                assets_numbers = request.form.getlist('assets_number[]')
                assets_values = request.form.getlist('assets_value[]')
                for i in range(0, min(len(assets_items), len(assets_numbers), len(assets_values))):
                    if assets_items[i] or assets_numbers[i] or assets_values[i]:
                        report.assets.append(Asset(
                            item=assets_items[i],
                            number=parse_int(assets_numbers[i]),
                            value=parse_float(assets_values[i]),
                        ))

                # Section B - Donations
                donor_names = get_form_list_upper('donor_name[]')
                donor_categories = get_form_list_with_other('donor_category[]', 'donor_category_other[]', trigger_value='OTHER')
                donor_countries = get_form_list_with_other(
                    'donor_country[]',
                    'donor_country_other[]',
                    trigger_value='OTHER_COUNTRY',
                )
                donor_amounts = request.form.getlist('donor_amount[]')
                rows = max(len(donor_names), len(donor_categories), len(donor_countries), len(donor_amounts))
                for i in range(0, rows):
                    donor_name = donor_names[i] if i < len(donor_names) else None
                    donor_category = donor_categories[i] if i < len(donor_categories) else None
                    donor_country = donor_countries[i] if i < len(donor_countries) else None
                    donor_amount = donor_amounts[i] if i < len(donor_amounts) else None
                    if donor_name or donor_category or donor_country or donor_amount:
                        report.donations.append(Donation(
                            name=donor_name,
                            category=donor_category,
                            country=donor_country,
                            amount=parse_float(donor_amount),
                        ))

                # Section B - Grants to other PBOs
                grant_names = get_form_list_upper('grant_name[]')
                grant_regs = get_form_list_upper('grant_registration_no[]')
                grant_countries = get_form_list_upper('grant_country[]')
                grant_amounts = request.form.getlist('grant_amount[]')
                grant_rows = max(len(grant_names), len(grant_regs), len(grant_countries), len(grant_amounts))
                for i in range(0, grant_rows):
                    grant_name = grant_names[i] if i < len(grant_names) else None
                    grant_reg = grant_regs[i] if i < len(grant_regs) else None
                    grant_country = grant_countries[i] if i < len(grant_countries) else None
                    grant_amount = grant_amounts[i] if i < len(grant_amounts) else None
                    if grant_name or grant_reg or grant_country or grant_amount:
                        report.grants.append(Grant(
                            name=grant_name,
                            registration_no=grant_reg,
                            country=grant_country,
                            amount=parse_float(grant_amount),
                        ))

                # Section B - Payments (dynamic rows from submitted form)
                payment_descriptions = get_form_list_with_other_raw(
                    'payment_description[]',
                    'payment_description_other[]',
                    trigger_value='OTHER',
                )
                payment_kenya = get_form_list_any('payment_kenya[]', 'payment_kenya_amount[]')
                payment_other = get_form_list_any('payment_other[]', 'payment_other_amount[]')
                payment_rows = max(len(payment_descriptions), len(payment_kenya), len(payment_other))
                for idx in range(payment_rows):
                    description = payment_descriptions[idx] if idx < len(payment_descriptions) else None
                    kenya_val = payment_kenya[idx] if idx < len(payment_kenya) else None
                    other_val = payment_other[idx] if idx < len(payment_other) else None
                    if description or kenya_val or other_val:
                        report.payments.append(Payment(
                            description=description,
                            kenya_amount=parse_float(kenya_val),
                            other_amount=parse_float(other_val),
                        ))

                # Section B - Banking details
                bank_names = get_form_list_with_other('bank_name[]', 'bank_name_other[]')
                bank_branches = merge_form_list_upper('bank_branch[]', 'bank_branch_select[]')
                bank_accounts = [value.strip() for value in request.form.getlist('bank_account_number[]')]
                bank_currencies = get_form_list_upper('bank_currency[]')
                bank_rows = max(len(bank_names), len(bank_branches), len(bank_accounts), len(bank_currencies))
                for i in range(0, bank_rows):
                    bank_name = bank_names[i] if i < len(bank_names) else ''
                    bank_branch = bank_branches[i] if i < len(bank_branches) else ''
                    bank_account = bank_accounts[i] if i < len(bank_accounts) else ''
                    bank_currency = bank_currencies[i] if i < len(bank_currencies) else ''
                    if bank_name or bank_branch or bank_account or bank_currency:
                        bank_account = bank_account or "00000000"
                        bank_currency = bank_currency or "KES"
                        report.bank_accounts.append(BankAccount(
                            bank_name=bank_name,
                            branch=bank_branch,
                            account_number=bank_account,
                            currency=bank_currency,
                        ))

                # Section B - Auditor info
                audit_firms = get_form_list_upper('audit_firm[]')
                auditor_names = get_form_list_upper('auditor_name[]')
                auditor_practicing = get_form_list_upper('auditor_practicing_no[]')
                for i in range(0, min(len(audit_firms), len(auditor_names), len(auditor_practicing))):
                    if audit_firms[i] or auditor_names[i] or auditor_practicing[i]:
                        report.auditors.append(AuditorEntry(
                            firm=audit_firms[i],
                            auditor_name=auditor_names[i],
                            practicing_no=auditor_practicing[i],
                        ))

                # Section C - Staff biodata
                staff_categories = get_form_list_upper('biodata_item[]')
                staff_prev = request.form.getlist('prev-year5[]')
                staff_curr = request.form.getlist('curr-year5[]')
                for i in range(0, min(len(staff_categories), len(staff_prev), len(staff_curr))):
                    category = normalize_section_c_biodata_category(staff_categories[i])
                    raw_prev = staff_prev[i]
                    raw_curr = staff_curr[i]
                    if not category or is_section_c_summary_category(category):
                        continue
                    if is_blank_form_value(raw_prev) and is_blank_form_value(raw_curr):
                        continue
                    report.staff_biodata.append(StaffBiodata(
                        category=category,
                        prev_year=parse_optional_int(raw_prev),
                        curr_year=parse_optional_int(raw_curr),
                    ))

                # Section C - Volunteer biodata
                vol_categories = get_form_list_upper('volbiodata_item[]')
                vol_prev = request.form.getlist('prev-volntr[]')
                vol_curr = request.form.getlist('curr-volntr[]')
                for i in range(0, min(len(vol_categories), len(vol_prev), len(vol_curr))):
                    category = normalize_section_c_biodata_category(vol_categories[i])
                    raw_prev = vol_prev[i]
                    raw_curr = vol_curr[i]
                    if not category or is_section_c_summary_category(category):
                        continue
                    if is_blank_form_value(raw_prev) and is_blank_form_value(raw_curr):
                        continue
                    report.volunteer_biodata.append(VolunteerBiodata(
                        category=category,
                        prev_year=parse_optional_int(raw_prev),
                        curr_year=parse_optional_int(raw_curr),
                    ))

                # Section C - Volunteer privileges
                priv_categories = get_form_list_with_other('vol_priv_category[]', 'vol_priv_category_other[]', trigger_value='OTHER')
                priv_kenyan_vol = set(get_form_list_upper('vol_priv_kenyan_volunteer[]'))
                priv_kenyan_intern = set(get_form_list_upper('vol_priv_kenyan_intern[]'))
                priv_int_vol = set(get_form_list_upper('vol_priv_international_volunteer[]'))
                priv_int_intern = set(get_form_list_upper('vol_priv_international_intern[]'))
                for category in priv_categories:
                    if category:
                        report.volunteer_privileges.append(VolunteerPrivilege(
                            category=category,
                            kenyan_volunteer=category in priv_kenyan_vol,
                            kenyan_intern=category in priv_kenyan_intern,
                            international_volunteer=category in priv_int_vol,
                            international_intern=category in priv_int_intern,
                        ))

                # Section C - Staff training
                training_types = get_form_list_upper('training_type[]')
                training_kenyan = request.form.getlist('training_kenyan[]')
                training_foreign = request.form.getlist('training_foreign[]')
                for i in range(0, min(len(training_types), len(training_kenyan), len(training_foreign))):
                    if training_types[i] or training_kenyan[i] or training_foreign[i]:
                        report.training_records.append(TrainingRecord(
                            training_type=training_types[i],
                            kenyan_count=parse_int(training_kenyan[i]),
                            international_count=parse_int(training_foreign[i]),
                        ))

                # Tax waiver items
                if government_sections_inactive:
                    report.tax_waiver_items.clear()
                else:
                    tax_descriptions = get_form_list_upper('items_exemption_description[]')
                    tax_quantities = request.form.getlist('items_exemption_quantity[]')
                    tax_types = get_form_list_upper('items_exemption_type[]')
                    tax_amounts = request.form.getlist('items_exemption_amount[]')
                    tax_certificates = get_form_list_upper('items_exemption_certificate[]')
                    tax_rows = max(
                        len(tax_descriptions),
                        len(tax_quantities),
                        len(tax_types),
                        len(tax_amounts),
                        len(tax_certificates),
                    )
                    for i in range(tax_rows):
                        description = tax_descriptions[i] if i < len(tax_descriptions) else ""
                        quantity = tax_quantities[i] if i < len(tax_quantities) else ""
                        exemption_type = tax_types[i] if i < len(tax_types) else ""
                        amount = tax_amounts[i] if i < len(tax_amounts) else ""
                        certificate = tax_certificates[i] if i < len(tax_certificates) else ""
                        if any([description, quantity, exemption_type, amount, certificate]):
                            report.tax_waiver_items.append(TaxWaiverItem(
                                item_description=description,
                                quantity=parse_int(quantity),
                                exemption_type=exemption_type,
                                estimated_tax_waived=parse_float(amount),
                                certificate_approval_no=certificate,
                            ))

                # Section D - Project implementation details
                project_sectors = get_form_list_with_other('project_sector[]', 'project_sector_other[]', trigger_value='OTHER')
                project_counties = get_form_list_upper('project_county[]')
                project_vulnerable_groups = get_form_list_with_other(
                    'project_vulnerable_group[]',
                    'project_vulnerable_group_other[]',
                    trigger_value='OTHERS',
                )
                project_beneficiaries = request.form.getlist('project_beneficiaries_no[]')
                project_spending = request.form.getlist('project_spending_per_county[]')
                project_duration = request.form.getlist('project_duration_years[]')
                project_completion = get_form_list_upper('project_completion_status[]')
                project_spent_kenya = request.form.getlist('project_amount_spent_kenya[]')
                project_spent_other = request.form.getlist('project_amount_spent_other[]')
                project_rows = max(
                    len(project_sectors),
                    len(project_counties),
                    len(project_vulnerable_groups),
                    len(project_beneficiaries),
                    len(project_spending),
                    len(project_duration),
                    len(project_completion),
                    len(project_spent_kenya),
                    len(project_spent_other),
                )
                for i in range(0, project_rows):
                    sector = project_sectors[i] if i < len(project_sectors) else None
                    county = project_counties[i] if i < len(project_counties) else None
                    vulnerable_group = project_vulnerable_groups[i] if i < len(project_vulnerable_groups) else None
                    beneficiaries = project_beneficiaries[i] if i < len(project_beneficiaries) else ""
                    spending = project_spending[i] if i < len(project_spending) else ""
                    duration = project_duration[i] if i < len(project_duration) else ""
                    completion = project_completion[i] if i < len(project_completion) else None
                    spent_kenya = project_spent_kenya[i] if i < len(project_spent_kenya) else ""
                    spent_other = project_spent_other[i] if i < len(project_spent_other) else ""
                    if any([
                        sector,
                        county,
                        vulnerable_group,
                        beneficiaries,
                        spending,
                        duration,
                        completion,
                        spent_kenya,
                        spent_other,
                    ]):
                        report.project_implementations.append(ProjectImplementation(
                            sector=sector,
                            county=county,
                            vulnerable_group=vulnerable_group,
                            beneficiaries_no=parse_optional_int(beneficiaries),
                            spending_per_county=parse_float(spending),
                            duration_years=parse_float(duration),
                            completion_status=completion,
                            amount_spent_kenya=parse_float(spent_kenya),
                            amount_spent_other=parse_float(spent_other),
                        ))

                # Section D1 - Projects carried out
                carried_sectors = get_form_list_with_other('projects_sector[]', 'projects_sector_other[]', trigger_value='OTHER')
                carried_forward_kenya = set(get_form_list_upper('projects_carried_forward_kenya[]'))
                carried_forward_other = set(get_form_list_upper('projects_carried_forward_other[]'))
                carried_started_kenya = set(get_form_list_upper('projects_started_kenya[]'))
                carried_started_other = set(get_form_list_upper('projects_started_other[]'))
                carried_completed_kenya = set(get_form_list_upper('projects_completed_kenya[]'))
                carried_completed_other = set(get_form_list_upper('projects_completed_other[]'))
                for sector in carried_sectors:
                    if not sector:
                        continue
                    report.projects_carried_out.append(ProjectCarriedOut(
                        sector=sector,
                        carried_forward_kenya=sector if sector in carried_forward_kenya else None,
                        carried_forward_other=sector if sector in carried_forward_other else None,
                        started_kenya=sector if sector in carried_started_kenya else None,
                        started_other=sector if sector in carried_started_other else None,
                        completed_kenya=sector if sector in carried_completed_kenya else None,
                        completed_other=sector if sector in carried_completed_other else None,
                    ))

                # Section D3 - Collaboration & Networking
                collab_partner_types = get_form_list_with_other(
                    'collab_partner_type[]',
                    'collab_partner_type_other[]',
                    trigger_value='OTHER',
                )
                collab_info_exchange = set(get_form_list_upper('collab_info_exchange[]'))
                collab_tech_support_to = set(get_form_list_upper('collab_tech_support_to[]'))
                collab_tech_support_from = set(get_form_list_upper('collab_tech_support_from[]'))
                collab_funding_to = set(get_form_list_upper('collab_funding_to[]'))
                collab_funding_from = set(get_form_list_upper('collab_funding_from[]'))
                collab_equipment_to = set(get_form_list_upper('collab_equipment_to[]'))
                collab_equipment_from = set(get_form_list_upper('collab_equipment_from[]'))
                for partner_type in collab_partner_types:
                    if not partner_type:
                        continue
                    report.collaborations.append(CollaborationNetworking(
                        partner_type=partner_type,
                        info_exchange=partner_type if partner_type in collab_info_exchange else None,
                        tech_support_to_partner=partner_type if partner_type in collab_tech_support_to else None,
                        tech_support_from_partner=partner_type if partner_type in collab_tech_support_from else None,
                        funding_to_partner=partner_type if partner_type in collab_funding_to else None,
                        funding_from_partner=partner_type if partner_type in collab_funding_from else None,
                        equipment_to_partner=partner_type if partner_type in collab_equipment_to else None,
                        equipment_from_partner=partner_type if partner_type in collab_equipment_from else None,
                    ))

                # Section E - Officials
                if get_inactive_section_flag("officials_section_inactive"):
                    clear_disabled_officials_section(report)
                else:
                    official_roles = get_form_list_upper('official_role[]')
                    official_names = get_form_list_upper('official_name[]')
                    official_nationalities = get_form_list_upper('official_nationality[]')
                    official_genders = get_form_list_upper('official_gender[]')
                    official_emails = get_form_list_upper('official_email[]')
                    official_residences = get_form_list_upper('official_residence[]')
                    official_phones = get_form_list_upper('official_phone[]')
                    official_kra_pins = get_form_list_upper('official_kra_pin[]')
                    official_qualifications = get_form_list_upper('official_professional_qualification[]')
                    official_signatures = request.form.getlist('official_signature[]')
                    official_rows = min(
                        len(official_roles),
                        len(official_names),
                        len(official_nationalities),
                        len(official_genders),
                        len(official_emails),
                        len(official_residences),
                        len(official_phones),
                    )
                    for i in range(0, official_rows):
                        if any([
                            official_roles[i],
                            official_names[i],
                            official_nationalities[i],
                            official_genders[i],
                            official_emails[i],
                            official_residences[i],
                            official_phones[i],
                        ]):
                            report.officials.append(Official(
                                role=official_roles[i] or "OFFICIAL",
                                name=official_names[i],
                                nationality=official_nationalities[i],
                                gender=official_genders[i],
                                email=official_emails[i],
                                residence=official_residences[i],
                                phone=official_phones[i],
                                kra_pin=official_kra_pins[i] if i < len(official_kra_pins) else None,
                                professional_qualification=official_qualifications[i] if i < len(official_qualifications) else None,
                                signature=official_signatures[i] if i < len(official_signatures) else None,
                            ))

                report.number_of_directors = parse_int(request.form.get('number_of_directors'))
                report.number_of_registered_members = parse_int(request.form.get('number_of_registered_members'))
                report.number_of_board_meetings = parse_int(request.form.get('number_of_board_meetings'))

                ef = get_form_list_upper_any('election_frequency[]', 'election_frequency')
                report.election_frequency = ", ".join(ef) if ef else None
                report.election_frequency_other = get_form_upper('election_frequency_other') if 'OTHER' in ef else None
                report.date_last_agm = parse_date(request.form.get('date_last_agm'))
                report.date_last_election = parse_date(request.form.get('date_last_election'))
                report.date_last_board_meeting = parse_date(request.form.get('date_last_board_meeting'))

                report.membership_number_of_directors = parse_optional_int(request.form.get("membership_number_of_directors"))
                report.membership_number_of_registered_members = parse_optional_int(request.form.get("membership_number_of_registered_members"))
                report.membership_number_of_board_meetings = parse_optional_int(request.form.get("membership_number_of_board_meetings"))
                report.membership_date_last_agm = parse_date(request.form.get("membership_date_last_agm"))
                report.membership_date_last_election = parse_date(request.form.get("membership_date_last_election"))
                report.non_membership_number_of_directors = parse_optional_int(request.form.get("non_membership_number_of_directors"))
                report.non_membership_number_of_board_meetings = parse_optional_int(request.form.get("non_membership_number_of_board_meetings"))
                report.non_membership_date_last_board_meeting = parse_date(request.form.get("non_membership_date_last_board_meeting"))
                report.non_membership_date_last_election = parse_date(request.form.get("non_membership_date_last_election"))
                if get_inactive_section_flag("officials_section_inactive"):
                    clear_disabled_officials_section(report)
                else:
                    report.submitter_fullname = get_form_upper("submitter_fullname")
                    report.signature = request.form.get("signature")
                    report.submission_date = parse_date(request.form.get("submission_date"))

                # Server-side totals (authoritative)
                apply_backend_section_b_math(report)
                sync_report_section_c_auxiliary_tables(report)

                # Compute and store risk score
                report.workflow_status = 'submitted'
                report.review_status = 'pending'
                report.data_source = 'form'
                report.duplicate_flag = find_duplicate_pbo_report(
                    report.pbo_name,
                    reporting_period_start=report.reporting_period_start,
                    require_start_month_match=True,
                ) is not None
                report.update_risk_score(compute_tf_risk)
                if current_user.is_authenticated:
                    report.user_id = current_user.id
                    report.last_modified_by_id = current_user.id

                add_and_flush_new_instance(report, label='report')
                record_report_field_changes(report, {field: '' for field in TRACKED_REPORT_FIELDS}, 'report_created', current_user if current_user.is_authenticated else None)
                log_user_activity('report_created', report=report)
                db.session.commit()
                flash('Form submitted successfully!', 'success')
                return redirect(url_for('report_detail', report_id=report.id, clear_form14_draft=1))
            except IntegrityError:
                db.session.rollback()
                existing_report = PBOReport.query.filter_by(submission_token=posted_submission_token).first()
                if existing_report is not None and form_submission_matches_existing_report(existing_report):
                    flash('This form was already submitted. Showing the saved report instead.', 'info')
                    return redirect(url_for('report_detail', report_id=existing_report.id, clear_form14_draft=1))
                flash('A duplicate submission was blocked. Please try again.', 'warning')
                return render_form14_response(submission_token=reset_submission_token(), prefill=empty_form14_prefill())
            except Exception as e:
                db.session.rollback()
                flash(f'Error submitting form: {str(e)}', 'error')
                return render_form14_response(submission_token=reset_submission_token(), prefill=empty_form14_prefill())
        return render_form14_response(submission_token=submission_token, prefill=empty_form14_prefill())


@app.route('/api/form14/submission-token', methods=['GET'])
@login_required
def form14_submission_token():
    response = jsonify({"submission_token": issue_submission_token()})
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response



# Signup route
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        flash('Sign up is currently disabled.', 'warning')
        return redirect(url_for('signup'))

    return render_template('signup.html', signup_disabled=True)
import re
# Login route
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        raw_email = request.form.get('email') or ''
        email = normalize_login_email(raw_email)
        password = request.form.get('password') or ''
        raw_department = request.form.get('department') or ''
        department_input = normalize_department(raw_department)
        if not email:
            flash('Email is required.', 'danger')
            return render_template('login.html', email=raw_email, department=raw_department)

        user, login_resolution = resolve_user_login(raw_email)
        if login_resolution == 'ambiguous':
            flash('That username matches multiple accounts. Please log in with your full email address.', 'danger')
            return render_template('login.html', email=raw_email, department=raw_department)

        stored_department = normalize_department(user.department or '') if user else ''
        credentials_match = user and user.check_password(password)
        department_match = bool(department_input) and stored_department == department_input

        if credentials_match and (
            stored_department == ''
            or not department_input
            or department_match
        ):
            if not stored_department and department_input:
                user.department = department_input

            if not user.is_authorized:
                user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
                user.last_failed_login_at = utc_now()
                log_user_activity('login_blocked', user=user, summary=f'{user.email} attempted login before authorization')
                db.session.commit()
                flash('Your account is pending admin authorization.', 'warning')
                return render_template('login.html', email=raw_email, department=raw_department)

            if current_user.is_authenticated:
                logout_user()
            establish_user_session(user)
            user.last_login_at = utc_now()
            user.last_login_ip = get_request_ip()
            prior_failed_attempts = user.failed_login_attempts or 0
            user.failed_login_attempts = 0
            log_user_activity('login', user=user, summary=f'{user.email} logged in')
            db.session.commit()
            notify_user_login(user)
            # Password change enforcement removed - users can continue working
            if prior_failed_attempts >= 10:
                flash('Too many failed attempts. Recent activity detected.', 'info')
            flash('Logged in successfully.', 'success')
            return redirect(url_for('home'))

        if user:
            user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
            user.last_failed_login_at = utc_now()
            log_user_activity('login_failed', user=user, summary=f'Failed login for {user.email}')
            db.session.commit()

        if user and credentials_match and stored_department and not department_match:
            flash('Email and password are valid, but department does not match.', 'danger')
        else:
            flash('Invalid credentials.', 'danger')
    return render_template('login.html', email=request.form.get('email') if request.method == 'POST' else None,
                           department=request.form.get('department') if request.method == 'POST' else None)


@app.route('/password/forgot', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        raw_email = request.form.get('email') or ''
        user, _ = resolve_user_login(raw_email)

        if user and user.email:
            token = generate_reset_token(user.email)
            reset_url = url_for('reset_password', token=token, _external=True)
            recipient = user.email
            if user.is_superadmin:
                recipient = app.config.get('SUPERADMIN_RESET_FORWARD_EMAIL')
            send_email_async(
                subject='Form 14 password reset',
                recipients=[recipient],
                body=(
                    f'Hello,\n\n'
                    f'A password reset was requested for your Form 14 account.\n\n'
                    f'Reset link (valid for {app.config.get("RESET_TOKEN_MAX_AGE_SECONDS", 3600)//60} minutes):\n'
                    f'{reset_url}\n\n'
                    f'If you did not request this reset, you can ignore this email.'
                ),
            )
        flash('If the account exists, a reset link has been emailed.', 'success')
        return redirect(url_for('login'))

    return render_template('forgot_password.html')


@app.route('/password/reset/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        email = verify_reset_token(token)
    except (SignatureExpired, BadSignature):
        flash('The reset link is invalid or has expired.', 'danger')
        return redirect(url_for('forgot_password'))

    user = User.query.filter_by(email=email).first()
    if user is None:
        flash('The reset link is invalid or has expired.', 'danger')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        password = request.form.get('password') or ''
        confirm = request.form.get('confirm_password') or ''
        if password != confirm:
            flash('Passwords do not match.', 'danger')
            return redirect(url_for('reset_password', token=token))
        for valid, msg in validate_password_policy(password):
            if not valid:
                flash(msg, 'danger')
                return redirect(url_for('reset_password', token=token))

        user.set_password(password)
        db.session.commit()

        recipient = user.email
        if user.is_superadmin:
            recipient = app.config.get('SUPERADMIN_RESET_FORWARD_EMAIL')
        send_email_async(
            subject='Form 14 password reset successful',
            recipients=[recipient],
            body=(
                f'Hello,\n\n'
                f'The password for your Form 14 account ({user.email}) was just reset.\n\n'
                f'If you did not perform this action, contact the administrator immediately.'
            ),
        )

        if current_user.is_authenticated:
            logout_user()
            session.clear()

        flash('Password reset successful. Please log in again.', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password.html', token=token, email=email)


@app.route('/profile', methods=['GET', 'POST'])
@app.route('/password/change', methods=['GET', 'POST'])
@login_required
def change_password():
    flash('Profile and password change are currently disabled.', 'warning')
    return redirect(url_for('home'))

# Logout route
@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    log_user_activity('logout', user=current_user, summary=f'{current_user.email} logged out')
    db.session.commit()
    logout_user()
    session.clear()
    flash('Logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/reports/ocr-upload', methods=['GET', 'POST'])
@require_ocr_features
@login_required
def report_ocr_upload():
    recent_reports_query = (
        PBOReport.query
        .filter(func.upper(PBOReport.data_source) == 'OCR_UPLOAD')
        .order_by(PBOReport.created_at.desc())
    )
    recent_uploads_query = (
        UploadedFile.query
        .filter(UploadedFile.category == 'ocr_source')
        .order_by(UploadedFile.created_at.desc())
    )

    if not can_manage_all_records(current_user):
        recent_reports_query = recent_reports_query.filter_by(user_id=current_user.id)
        recent_uploads_query = recent_uploads_query.filter_by(uploaded_by_id=current_user.id)

    if request.method == 'POST':
        uploads = [item for item in request.files.getlist('images') if item and item.filename]
        if not uploads:
            flash('Choose at least one image before uploading to the database.', 'warning')
            return redirect(url_for('report_ocr_upload'))
        if len(uploads) > 10:
            flash('You can upload a maximum of 10 images per OCR import.', 'danger')
            return redirect(url_for('report_ocr_upload'))

        invalid_files = [upload.filename for upload in uploads if not allowed_image_upload(upload.filename)]
        if invalid_files:
            flash(f"Unsupported image type: {', '.join(invalid_files)}", 'danger')
            return redirect(url_for('report_ocr_upload'))

        ocr_upload_date = get_app_today()
        for upload in uploads:
            duplicate_file = find_duplicate_ocr_upload(upload.filename, current_user, target_date=ocr_upload_date)
            if duplicate_file is not None:
                flash(
                    f'A same-day OCR upload named "{duplicate_file.original_filename}" already exists. Opening the existing file instead.',
                    'warning',
                )
                return redirect(url_for('my_file_open', file_id=duplicate_file.id))

        saved_uploads = []
        saved_upload_ids = []
        try:
            for upload in uploads:
                saved_uploads.append(
                    save_uploaded_file(upload, category='ocr_source', user=current_user, status='processing')
                )
            db.session.commit()
            saved_upload_ids = [item.id for item in saved_uploads]

            from data import build_report_from_ocr_upload

            report, file_results, warnings = build_report_from_ocr_upload(
                current_user,
                [item.storage_path for item in saved_uploads],
                [item.original_filename for item in saved_uploads],
            )
            report.user_id = current_user.id
            report.last_modified_by_id = current_user.id
            report.data_source = 'ocr_upload'
            assign_pbo_name_normalized(report)
            add_and_flush_new_instance(report, label='OCR report')
            report.duplicate_flag = find_duplicate_pbo_report(
                report.pbo_name,
                reporting_period_start=report.reporting_period_start,
                exclude_report_id=report.id,
                require_start_month_match=True,
            ) is not None
            report.form_14 = ', '.join(item.original_filename for item in saved_uploads)
            report.update_risk_score(compute_tf_risk)

            linked_uploads = (
                UploadedFile.query
                .filter(UploadedFile.id.in_(saved_upload_ids))
                .order_by(UploadedFile.id.asc())
                .all()
            )
            for uploaded_file, result in zip(linked_uploads, file_results):
                uploaded_file.report_id = report.id
                uploaded_file.status = result['status']
                uploaded_file.extracted_text = result['text']
                uploaded_file.error_message = result['error']

            log_user_activity(
                'ocr_report_imported',
                report=report,
                summary=f'{current_user.email} created report {report.id} from {len(saved_uploads)} uploaded image(s).',
            )
            db.session.commit()

            processed_count = sum(1 for result in file_results if result['status'] == 'ocr_processed')
            failed_count = len(file_results) - processed_count
            flash(
                f'OCR import created report #{report.id}. {processed_count} image(s) processed successfully.',
                'success',
            )
            if failed_count:
                flash(f'{failed_count} image(s) could not be read cleanly. Review the extracted fields carefully.', 'warning')
            for warning in warnings:
                flash(warning, 'warning')
            return redirect(url_for('report_edit', report_id=report.id))
        except Exception as exc:
            db.session.rollback()
            if saved_upload_ids:
                for uploaded_file in UploadedFile.query.filter(UploadedFile.id.in_(saved_upload_ids)).all():
                    uploaded_file.status = 'ocr_failed'
                    uploaded_file.error_message = str(exc)
                db.session.commit()
            flash(f'OCR import failed: {exc}', 'danger')
            return redirect(url_for('report_ocr_upload'))

    recent_reports = recent_reports_query.limit(10).all()
    recent_uploads = recent_uploads_query.limit(20).all()
    return render_template(
        'ocr_upload.html',
        recent_reports=recent_reports,
        recent_uploads=recent_uploads,
    )


@app.route('/reports/ocr-batch-export', methods=['GET', 'POST'])
@require_ocr_features
@login_required
def report_ocr_batch_export():
    recent_uploads_query = (
        UploadedFile.query
        .filter(UploadedFile.category == 'ocr_batch_source')
        .order_by(UploadedFile.created_at.desc())
    )

    if not can_manage_all_records(current_user):
        recent_uploads_query = recent_uploads_query.filter_by(uploaded_by_id=current_user.id)

    if request.method == 'POST':
        uploads = [item for item in request.files.getlist('documents') if item and item.filename]
        if not uploads:
            flash('Choose at least one document for batch OCR export.', 'warning')
            return redirect(url_for('report_ocr_batch_export'))
        if len(uploads) > 50:
            flash('You can upload a maximum of 50 files per batch OCR export.', 'danger')
            return redirect(url_for('report_ocr_batch_export'))

        invalid_files = [upload.filename for upload in uploads if not allowed_image_upload(upload.filename)]
        if invalid_files:
            flash(f"Unsupported file type: {', '.join(invalid_files)}", 'danger')
            return redirect(url_for('report_ocr_batch_export'))

        saved_uploads = []
        try:
            for upload in uploads:
                saved_uploads.append(
                    save_uploaded_file(upload, category='ocr_batch_source', user=current_user, status='processing')
                )
            db.session.commit()

            rows = build_batch_ocr_export_rows(saved_uploads, current_user)
            db.session.commit()

            pandas_module = get_pandas()
            dataframe = pandas_module.DataFrame(rows)
            output = io.BytesIO()
            with pandas_module.ExcelWriter(output, engine='openpyxl') as writer:
                dataframe.to_excel(writer, index=False, sheet_name='OCR Batch Export')
            output.seek(0)

            processed_count = sum(1 for row in rows if row['status'] == 'ocr_processed')
            failed_count = len(rows) - processed_count
            log_user_activity(
                'ocr_batch_exported',
                summary=f'{current_user.email} batch-processed {len(rows)} OCR file(s) into Excel.',
            )
            flash(
                f'Batch OCR export prepared. {processed_count} file(s) processed successfully, {failed_count} failed.',
                'success' if processed_count else 'warning',
            )
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'ocr_batch_export_{utc_now().strftime("%Y%m%d%H%M%S")}.xlsx',
            )
        except Exception as exc:
            db.session.rollback()
            for uploaded_file in saved_uploads:
                try:
                    uploaded_file.status = 'ocr_failed'
                    uploaded_file.error_message = str(exc)
                    db.session.add(uploaded_file)
                except Exception:
                    pass
            db.session.commit()
            flash(f'Batch OCR export failed: {exc}', 'danger')
            return redirect(url_for('report_ocr_batch_export'))

    recent_uploads = recent_uploads_query.limit(25).all()
    return render_template(
        'ocr_batch_export.html',
        recent_uploads=recent_uploads,
    )


@app.route('/admin/users')
@admin_required
@login_required
def authorize_accounts():
    users = User.query.order_by(User.is_authorized.asc(), User.email.asc()).all()
    return render_template(
        'authorize_accounts.html',
        users=users,
        current_user_is_superadmin=is_superadmin(current_user),
        seed_admin_emails=seed_admin_emails(),
    )


@app.route('/admin/users/create', methods=['POST'])
@admin_required
@login_required
def create_user_account():
    denied = require_superadmin()
    if denied:
        return denied

    raw_identifier = request.form.get('login_identifier') or ''
    raw_full_name = request.form.get('full_name') or ''
    raw_phone = request.form.get('phone') or ''
    raw_department = request.form.get('department') or ''
    password = request.form.get('password') or ''
    confirm_password = request.form.get('confirm_password') or ''
    generate_password_requested = request.form.get('generate_password') == '1'
    authorize_now = request.form.get('is_authorized') == '1'
    must_change_password = request.form.get('must_change_password', '1') == '1'

    email, email_error = build_login_email('', raw_identifier)
    if email_error:
        flash(email_error, 'danger')
        return redirect(url_for('authorize_accounts'))

    if User.query.filter_by(email=email).first():
        flash('An account with that username/email already exists.', 'danger')
        return redirect(url_for('authorize_accounts'))

    if generate_password_requested and (password or confirm_password):
        flash('Use either a manual password or Generate Password, not both.', 'danger')
        return redirect(url_for('authorize_accounts'))

    if not generate_password_requested and not password:
        flash('Enter a password or choose Generate Password for the new user.', 'danger')
        return redirect(url_for('authorize_accounts'))

    generated_password = None
    if generate_password_requested:
        generated_password = generate_admin_password()
        password_to_store = generated_password
    else:
        if password != confirm_password:
            flash('Password and confirm password must match.', 'danger')
            return redirect(url_for('authorize_accounts'))
        for valid, msg in validate_password_policy(password):
            if not valid:
                flash(msg, 'danger')
                return redirect(url_for('authorize_accounts'))
        password_to_store = password

    new_user = User(
        email=email,
        full_name=re.sub(r'\s+', ' ', raw_full_name.strip()) if raw_full_name.strip() else None,
        phone=normalize_phone(raw_phone) if raw_phone.strip() else None,
        department=normalize_department(raw_department),
        role='user',
        is_superadmin=False,
        can_manage_all_records=False,
        is_authorized=authorize_now,
        authorized_at=utc_now() if authorize_now else None,
        authorized_by_id=current_user.id if authorize_now else None,
        must_change_password=must_change_password,
    )
    new_user.set_password(password_to_store, mark_changed=not must_change_password)
    if must_change_password:
        new_user.password_changed_at = None

    add_and_flush_new_instance(new_user, label='user account')
    log_user_activity(
        'user_account_created',
        user=current_user,
        summary=(
            f'{current_user.email} created account {new_user.email}'
            + (' and authorized it immediately' if authorize_now else '')
        ),
    )
    db.session.commit()

    flash(f'User account for {new_user.email} created successfully.', 'success')
    if generated_password:
        flash(f'Generated password for {new_user.email}: {generated_password}', 'warning')
    return redirect(url_for('authorize_accounts'))


@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
@login_required

def edit_user_account(user_id):
    user = db.session.get(User, user_id) or abort(404)
    denied = require_admin_editor_access(user)
    if denied:
        return denied

    current_user_is_superadmin = is_superadmin(current_user)

    if request.method == 'POST':
        raw_username = request.form.get('username') or ''
        raw_email = request.form.get('email') or ''
        raw_full_name = request.form.get('full_name') or ''
        raw_phone = request.form.get('phone') or ''
        raw_department = request.form.get('department') or ''
        raw_role = request.form.get('role') or ''
        password = request.form.get('password') or ''
        confirm_password = request.form.get('confirm_password') or ''
        generate_password_requested = request.form.get('generate_password') == '1'
        raw_is_authorized = request.form.get('is_authorized')
        raw_must_change_password = request.form.get('must_change_password')
        raw_can_manage_all_records = request.form.get('can_manage_all_records')

        updated_email, email_error = build_login_email(raw_username, raw_email)
        department = normalize_department(raw_department)
        updated_full_name = re.sub(r'\s+', ' ', raw_full_name.strip()) if raw_full_name.strip() else None
        updated_phone = normalize_phone(raw_phone) if raw_phone.strip() else None

        if email_error:
            flash(email_error, 'danger')
            return render_template('admin_edit_user.html', user=user, current_user_is_superadmin=current_user_is_superadmin)

        existing_user = User.query.filter_by(email=updated_email).first()
        if existing_user and existing_user.id != user.id:
            flash('Another account already uses that email/username combination.', 'danger')
            return render_template('admin_edit_user.html', user=user, current_user_is_superadmin=current_user_is_superadmin)

        if generate_password_requested and (password or confirm_password):
            flash('Use either a manual password or Generate Password, not both.', 'danger')
            return render_template('admin_edit_user.html', user=user, current_user_is_superadmin=current_user_is_superadmin)

        generated_password = None
        if generate_password_requested:
            generated_password = generate_admin_password()
            user.set_password(generated_password)
        elif password or confirm_password:
            if password != confirm_password:
                flash('Password and confirm password must match.', 'danger')
                return render_template('admin_edit_user.html', user=user, current_user_is_superadmin=current_user_is_superadmin)
            for valid, msg in validate_password_policy(password):
                if not valid:
                    flash(msg, 'danger')
                    return render_template('admin_edit_user.html', user=user, current_user_is_superadmin=current_user_is_superadmin)
            user.set_password(password)

        old_email = user.email
        old_full_name = user.full_name
        old_phone = user.phone
        old_department = user.department
        old_role = user.role
        old_is_authorized = user.is_authorized
        old_must_change_password = user.must_change_password
        old_can_manage_all_records = user.can_manage_all_records

        user.email = updated_email
        user.full_name = updated_full_name
        user.phone = updated_phone
        user.department = department

        # This screen is intentionally limited to account details only.
        # Role, authorization, password-policy flags, and broad access are managed
        # from the dedicated admin actions on the user-management page.

        db.session.commit()

        if current_user.id == user.id:
            session['user_email'] = user.email
            session.modified = True

        changed_parts = []
        if old_email != user.email:
            changed_parts.append(f'email/login changed from {old_email or "none"} to {user.email}')
        if old_full_name != user.full_name:
            changed_parts.append(f'full name changed from {old_full_name or "none"} to {user.full_name or "none"}')
        if old_phone != user.phone:
            changed_parts.append(f'phone changed from {old_phone or "none"} to {user.phone or "none"}')
        if old_department != user.department:
            changed_parts.append(f'department changed from {old_department or "none"} to {user.department or "none"}')
        if old_role != user.role:
            changed_parts.append(f'role changed from {old_role or "none"} to {user.role or "none"}')
        if old_is_authorized != user.is_authorized:
            changed_parts.append(f'authorization changed from {old_is_authorized} to {user.is_authorized}')
        if old_must_change_password != user.must_change_password:
            changed_parts.append(f'must_change_password changed from {old_must_change_password} to {user.must_change_password}')
        if old_can_manage_all_records != user.can_manage_all_records:
            changed_parts.append(
                f'can_manage_all_records changed from {old_can_manage_all_records} to {user.can_manage_all_records}'
            )
        if password:
            changed_parts.append('password updated')
        if generated_password:
            changed_parts.append('password auto-generated')

        log_user_activity(
            'user_account_updated',
            user=current_user,
            summary=(
                f'{current_user.email} updated account settings for {user.email}: '
                + (', '.join(changed_parts) if changed_parts else 'no field changes submitted')
            ),
        )
        db.session.commit()

        flash(f'User account for {user.email} updated successfully.', 'success')
        if generated_password:
            flash(f'Generated password for {user.email}: {generated_password}', 'warning')
        return redirect(url_for('authorize_accounts'))

    return render_template('admin_edit_user.html', user=user, current_user_is_superadmin=current_user_is_superadmin)


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
@login_required
def delete_user_account(user_id):
    denied = require_superadmin()
    if denied:
        return denied

    user = db.session.get(User, user_id) or abort(404)
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'warning')
        return redirect(url_for('authorize_accounts'))
    if user.is_superadmin:
        flash('Superadmin accounts cannot be deleted.', 'warning')
        return redirect(url_for('authorize_accounts'))

    target_email = user.email or f'User #{user.id}'

    db.session.execute(
        text("UPDATE users_for_form14 SET authorized_by_id = NULL WHERE authorized_by_id = :user_id"),
        {'user_id': user.id},
    )
    db.session.execute(
        text(
            "UPDATE pbo_reports "
            "SET user_id = NULL, "
            "    last_modified_by_id = CASE WHEN last_modified_by_id = :user_id THEN NULL ELSE last_modified_by_id END, "
            "    reviewed_by_id = CASE WHEN reviewed_by_id = :user_id THEN NULL ELSE reviewed_by_id END "
            "WHERE user_id = :user_id OR last_modified_by_id = :user_id OR reviewed_by_id = :user_id"
        ),
        {'user_id': user.id},
    )
    db.session.execute(
        text("UPDATE admin_settings SET updated_by_id = NULL WHERE updated_by_id = :user_id"),
        {'user_id': user.id},
    )
    db.session.execute(
        text("UPDATE import_batches SET created_by_id = NULL WHERE created_by_id = :user_id"),
        {'user_id': user.id},
    )
    db.session.execute(
        text("UPDATE uploaded_files SET uploaded_by_id = NULL WHERE uploaded_by_id = :user_id"),
        {'user_id': user.id},
    )
    db.session.execute(
        text("UPDATE field_change_logs SET user_id = NULL WHERE user_id = :user_id"),
        {'user_id': user.id},
    )
    db.session.execute(
        text("UPDATE user_activity_logs SET user_id = NULL WHERE user_id = :user_id"),
        {'user_id': user.id},
    )
    db.session.execute(
        text("UPDATE data_analysis_bot_interactions SET user_id = NULL WHERE user_id = :user_id"),
        {'user_id': user.id},
    )
    db.session.execute(
        text("DELETE FROM users_for_form14 WHERE id = :user_id"),
        {'user_id': user.id},
    )

    log_user_activity(
        'user_account_deleted',
        user=current_user,
        summary=(
            f'{current_user.email} deleted account {target_email}. '
            'Linked reports remain in the system as unassigned records.'
        ),
    )
    db.session.commit()

    flash(f'User account {target_email} deleted successfully.', 'success')
    return redirect(url_for('authorize_accounts'))


@app.route('/admin/users/<int:user_id>/authorize', methods=['POST'])
@admin_required
@login_required

def authorize_account(user_id):
    denied = require_superadmin()
    if denied:
        return denied
    user = db.session.get(User, user_id) or abort(404)
    user.is_authorized = True
    user.authorized_at = utc_now()
    user.authorized_by_id = current_user.id
    log_user_activity('user_authorized', user=current_user, summary=f'{current_user.email} authorized {user.email}')
    db.session.commit()
    flash(f'{user.email} authorized successfully.', 'success')
    return redirect(url_for('authorize_accounts'))


@app.route('/admin/users/<int:user_id>/revoke', methods=['POST'])
@admin_required
@login_required

def revoke_account(user_id):
    denied = require_superadmin()
    if denied:
        return denied
    user = db.session.get(User, user_id) or abort(404)
    if user.is_superadmin:
        flash('Superadmin access cannot be revoked.', 'warning')
        return redirect(url_for('authorize_accounts'))
    if user.id == current_user.id:
        flash('You cannot revoke your own admin access.', 'warning')
        return redirect(url_for('authorize_accounts'))
    user.is_authorized = False
    log_user_activity('user_authorization_revoked', user=current_user, summary=f'{current_user.email} revoked authorization for {user.email}')
    db.session.commit()
    flash(f'{user.email} authorization revoked.', 'info')
    return redirect(url_for('authorize_accounts'))


@app.route('/admin/users/<int:user_id>/grant-record-access', methods=['POST'])
@admin_required
@login_required

def grant_record_access(user_id):
    denied = require_superadmin()
    if denied:
        return denied
    user = db.session.get(User, user_id) or abort(404)
    user.can_manage_all_records = True
    log_user_activity('record_access_granted', user=current_user, summary=f'{current_user.email} granted broad record access to {user.email}')
    db.session.commit()
    flash(f'{user.email} can now manage other users records.', 'success')
    return redirect(url_for('authorize_accounts'))


@app.route('/admin/users/<int:user_id>/revoke-record-access', methods=['POST'])
@admin_required
@login_required

def revoke_record_access(user_id):
    denied = require_superadmin()
    if denied:
        return denied
    user = db.session.get(User, user_id) or abort(404)
    if user.id == current_user.id:
        flash('You cannot revoke your own record-management privilege.', 'warning')
        return redirect(url_for('authorize_accounts'))
    user.can_manage_all_records = False
    log_user_activity('record_access_revoked', user=current_user, summary=f'{current_user.email} revoked broad record access from {user.email}')
    db.session.commit()
    flash(f'{user.email} can no longer manage other users records.', 'info')
    return redirect(url_for('authorize_accounts'))


@app.route('/admin/users/<int:user_id>/grant-admin', methods=['POST'])
@admin_required
@login_required

def grant_admin_role(user_id):
    denied = require_superadmin()
    if denied:
        return denied
    user = db.session.get(User, user_id) or abort(404)
    if user.email.lower() not in seed_admin_emails():
        flash('Admin role changes are limited to the seeded admin accounts.', 'warning')
        return redirect(url_for('authorize_accounts'))
    user.role = 'admin'
    user.is_authorized = True
    user.is_superadmin = False
    user.can_manage_all_records = False
    log_user_activity('admin_role_granted', user=current_user, summary=f'{current_user.email} granted admin role to {user.email}')
    db.session.commit()
    flash(f'{user.email} is now an admin.', 'success')
    return redirect(url_for('authorize_accounts'))


@app.route('/admin/users/<int:user_id>/revoke-admin', methods=['POST'])
@admin_required
@login_required

def revoke_admin_role(user_id):
    denied = require_superadmin()
    if denied:
        return denied
    user = db.session.get(User, user_id) or abort(404)
    if user.is_superadmin:
        flash('Superadmin access cannot be revoked.', 'warning')
        return redirect(url_for('authorize_accounts'))
    if user.email.lower() not in seed_admin_emails():
        flash('Admin role changes are limited to the seeded admin accounts.', 'warning')
        return redirect(url_for('authorize_accounts'))
    if user.id == current_user.id:
        flash('You cannot revoke your own admin role.', 'warning')
        return redirect(url_for('authorize_accounts'))
    user.role = 'user'
    user.is_superadmin = False
    user.can_manage_all_records = False
    log_user_activity('admin_role_revoked', user=current_user, summary=f'{current_user.email} revoked admin role from {user.email}')
    db.session.commit()
    flash(f'{user.email} is now a regular user.', 'info')
    return redirect(url_for('authorize_accounts'))

@app.route('/admin/users/change_admin', methods=['POST'])
@admin_required
@login_required

def change_admin():
    denied = require_superadmin()
    if denied:
        return jsonify({'success': False, 'message': 'Only a superadmin can perform that action.'}), 403

    data = request.get_json(silent=True) or {}
    user_id = data.get('user_id')
    make_admin = data.get('admin')

    if user_id is None or make_admin is None:
        return jsonify({'success': False, 'message': 'Missing user_id or admin value.'}), 400

    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'success': False, 'message': 'User not found.'}), 404

    if user.id == current_user.id and not make_admin:
        return jsonify({'success': False, 'message': 'You cannot revoke your own admin role.'}), 400
    if user.is_superadmin and not make_admin:
        return jsonify({'success': False, 'message': 'Superadmin access cannot be revoked.'}), 400

    make_admin = bool(make_admin)
    user.role = 'admin' if make_admin else 'user'
    user.is_authorized = True if make_admin else user.is_authorized
    user.is_superadmin = False if not make_admin else user.is_superadmin
    user.can_manage_all_records = False if not make_admin else user.can_manage_all_records

    action = 'granted' if make_admin else 'revoked'
    log_user_activity(
        f'admin_role_{"granted" if make_admin else "revoked"}',
        user=current_user,
        summary=f'{current_user.email} {action} admin role {"to" if make_admin else "from"} {user.email}',
    )
    db.session.commit()
    return jsonify({'success': True, 'message': f'Admin access {action} successfully for {user.email}.'})

# Create route (admin only)
@app.route('/admin/create', methods=['GET', 'POST'])
@admin_required
@login_required

def admin_create():
    if request.method == 'POST':
        # Example: create new report
        # Add form handling here
        flash('Created successfully.', 'success')
        return redirect(url_for('admin_view'))
    return render_template('admin_create.html')

# Admin view route
def build_admin_user_activity_rows(users, activity_date=None):
    activity_start = None
    activity_end = None
    if activity_date:
        activity_start = datetime.combine(activity_date, datetime.min.time())
        activity_end = datetime.combine(activity_date, datetime.max.time())

    grouped_users = defaultdict(list)
    for user in users:
        if user is None:
            continue
        group_key = normalize_user_email_key(getattr(user, 'email', None)) or f'user-id:{getattr(user, "id", "unknown")}'
        grouped_users[group_key].append(user)

    submitted_report_index = build_form_submitted_report_index(
        user_ids=[
            getattr(user, 'id', None)
            for user in users
            if user is not None and getattr(user, 'id', None) is not None
        ],
        activity_start=activity_start,
        activity_end=activity_end,
    )
    report_cache = {}
    user_activity_rows = []
    for group_key in sorted(grouped_users, key=str):
        user_group = [user for user in grouped_users[group_key] if user is not None]
        if not user_group:
            continue

        canonical_user = choose_preferred_user_candidate(user_group)
        if canonical_user is None:
            continue

        submitted_reports = []
        submitted_report_ids = set()
        activity_report_last_touched = {}
        group_activities = []
        for user in user_group:
            for report in (getattr(user, 'reports', None) or []):
                if report is None or getattr(report, 'id', None) is None:
                    continue
                if report.id in submitted_report_ids:
                    continue
                if getattr(report, 'user_id', None) != getattr(user, 'id', None):
                    continue
                if report_submission_timestamp(report) is None:
                    continue
                report_source = str(getattr(report, 'data_source', None) or 'form').strip().lower()
                if report_source != 'form':
                    continue
                if activity_start and activity_end:
                    submitted_at = report_submission_timestamp(report)
                    if submitted_at is None or not (activity_start <= submitted_at <= activity_end):
                        continue
                submitted_reports.append(report)
                submitted_report_ids.add(report.id)
            for entry in (getattr(user, 'activity_logs', None) or []):
                if (
                    entry is None
                    or getattr(entry, 'created_at', None) is None
                    or (
                        activity_start
                        and not (activity_start <= entry.created_at <= activity_end)
                    )
                ):
                    continue
                group_activities.append(entry)
                activity_report_id = getattr(entry, 'report_id', None)
                if activity_report_id is not None:
                    previous_touched_at = activity_report_last_touched.get(activity_report_id)
                    if previous_touched_at is None or entry.created_at > previous_touched_at:
                        activity_report_last_touched[activity_report_id] = entry.created_at

        group_user_ids = [
            getattr(user, 'id', None)
            for user in user_group
            if getattr(user, 'id', None) is not None
        ]
        group_submitted_last_touched = combine_worked_report_last_touched(
            submitted_report_index,
            group_user_ids,
        )
        all_activities = sorted(
            group_activities,
            key=lambda entry: entry.created_at or datetime.min,
            reverse=True,
        )
        if activity_start and not all_activities and not group_submitted_last_touched:
            continue

        recent_activities = all_activities[:15]
        worked_report_items = sorted(
            group_submitted_last_touched.items(),
            key=lambda item: (item[1] or datetime.min, item[0]),
            reverse=True,
        )
        recent_report_last_touched = dict(group_submitted_last_touched)
        for report_id, touched_at in activity_report_last_touched.items():
            previous_touched_at = recent_report_last_touched.get(report_id)
            if previous_touched_at is None or touched_at > previous_touched_at:
                recent_report_last_touched[report_id] = touched_at
        recent_report_items = sorted(
            recent_report_last_touched.items(),
            key=lambda item: (item[1] or datetime.min, item[0]),
            reverse=True,
        )
        touched_report_ids = [report_id for report_id, _ in recent_report_items[:15]]

        owned_by_id = {report.id: report for report in submitted_reports if getattr(report, 'id', None)}
        touched_reports = []
        for report_id in touched_report_ids:
            report = owned_by_id.get(report_id) or report_cache.get(report_id)
            if report is None:
                report = db.session.get(PBOReport, report_id)
                report_cache[report_id] = report
            if report:
                touched_reports.append(report)

        last_worked_report = None
        last_worked_at = worked_report_items[0][1] if worked_report_items else None
        if worked_report_items:
            last_worked_report_id = worked_report_items[0][0]
            last_worked_report = owned_by_id.get(last_worked_report_id) or report_cache.get(last_worked_report_id)
            if last_worked_report is None:
                last_worked_report = db.session.get(PBOReport, last_worked_report_id)
                report_cache[last_worked_report_id] = last_worked_report
            if last_worked_report is None:
                last_worked_at = None

        display_role = max(
            (getattr(user, 'role', None) for user in user_group),
            key=user_role_priority,
            default=getattr(canonical_user, 'role', None),
        )
        display_last_login_at = max(
            (
                getattr(user, 'last_login_at', None)
                for user in user_group
                if getattr(user, 'last_login_at', None) is not None
            ),
            default=None,
        )
        display_failed_attempts = max(
            (getattr(user, 'failed_login_attempts', None) or 0 for user in user_group),
            default=0,
        )

        user_activity_rows.append({
            'user': canonical_user,
            'owned_reports': submitted_reports,
            'touched_reports': touched_reports,
            'worked_file_count': len(group_submitted_last_touched),
            'recent_activities': recent_activities,
            'last_seen_at': recent_activities[0].created_at if recent_activities else None,
            'last_worked_report': last_worked_report,
            'last_worked_at': last_worked_at,
            'display_email': getattr(canonical_user, 'email', None) or (f'User #{canonical_user.id}' if getattr(canonical_user, 'id', None) is not None else 'Unknown user'),
            'display_role': display_role or getattr(canonical_user, 'role', None) or 'user',
            'display_is_authorized': any(getattr(user, 'is_authorized', False) for user in user_group),
            'display_last_login_at': display_last_login_at,
            'display_failed_attempts': display_failed_attempts,
            'display_full_name': next(
                (
                    getattr(user, 'full_name', None)
                    for user in [canonical_user] + user_group
                    if getattr(user, 'full_name', None)
                ),
                None,
            ),
            'group_email': normalize_user_email_key(getattr(canonical_user, 'email', None)),
            'group_user_ids': sorted(
                {
                    user.id
                    for user in user_group
                    if getattr(user, 'id', None) is not None
                }
            ),
            'group_size': len(user_group),
        })

    return user_activity_rows


def build_admin_user_activity_file_rows(search_query, activity_date=None):
    term = (search_query or '').strip()
    if not term:
        return []

    activity_start = None
    activity_end = None
    if activity_date:
        activity_start = datetime.combine(activity_date, datetime.min.time())
        activity_end = datetime.combine(activity_date, datetime.max.time())

    pattern = f"%{term}%"
    query = (
        PBOReport.query
        .options(
            selectinload(PBOReport.user),
            selectinload(PBOReport.last_modified_by),
        )
        .filter(
            or_(
                cast(PBOReport.id, String).ilike(pattern),
                PBOReport.pbo_name.ilike(pattern),
                PBOReport.pbo_registration_number.ilike(pattern),
                PBOReport.form_14.ilike(pattern),
            )
        )
    )
    if activity_start and activity_end:
        query = query.filter(
            or_(
                PBOReport.updated_at.between(activity_start, activity_end),
                PBOReport.last_activity_at.between(activity_start, activity_end),
                PBOReport.created_at.between(activity_start, activity_end),
            )
        )

    matched_reports = (
        query
        .order_by(PBOReport.updated_at.desc(), PBOReport.id.desc())
        .limit(100)
        .all()
    )
    return [
        {
            'report': report,
            'updated_by': report_updater_display(report),
            'updated_at': report.updated_at or report.last_activity_at or report.created_at,
        }
        for report in matched_reports
    ]


@app.route('/admin', methods=['GET', 'POST'])
@admin_required
@login_required

def admin_view():
    current_user_is_superadmin = is_superadmin(current_user)
    deadlines = {
        'edit_deadline': get_admin_setting('edit_deadline', ''),
        'form_submission_deadline': get_admin_setting('form_submission_deadline', ''),
    }
    if request.method == 'POST' and 'edit_deadline' in request.form:
        edit_deadline = request.form.get('edit_deadline')
        form_submission_deadline = request.form.get('form_submission_deadline')
        set_admin_setting('edit_deadline', edit_deadline, current_user)
        set_admin_setting('form_submission_deadline', form_submission_deadline, current_user)
        deadlines['edit_deadline'] = edit_deadline
        deadlines['form_submission_deadline'] = form_submission_deadline
        log_user_activity('admin_deadlines_updated', summary='Admin updated annual return deadlines')
        db.session.commit()
        flash('Deadlines updated successfully!', 'success')

    audit_action = request.args.get('action', '').strip()
    audit_user_id = request.args.get('user_id', type=int)
    audit_report_id = request.args.get('report_id', type=int)
    audit_from = parse_date(request.args.get('from_date'))
    audit_to = parse_date(request.args.get('to_date'))
    activity_date = parse_date(request.args.get('activity_date'))
    activity_search = request.args.get('activity_search', '').strip()
    requested_fk_audit = request.args.get('run_fk_audit') == '1'
    run_fk_audit = requested_fk_audit and current_user_is_superadmin
    if requested_fk_audit and not current_user_is_superadmin:
        flash('PostgreSQL hardening audits are available to superadmins only.', 'warning')

    users = (
        User.query
        .options(selectinload(User.reports), selectinload(User.activity_logs))
        .order_by(User.email.asc())
        .all()
    )
    user_activity_rows = build_admin_user_activity_rows(users, activity_date=activity_date)
    user_activity_file_rows = build_admin_user_activity_file_rows(activity_search, activity_date=activity_date)

    audit_query = UserActivityLog.query.options(
        selectinload(UserActivityLog.user),
        selectinload(UserActivityLog.report),
    )
    if audit_action:
        audit_query = audit_query.filter(UserActivityLog.action == audit_action)
    if audit_user_id:
        audit_query = audit_query.filter(UserActivityLog.user_id == audit_user_id)
    if audit_report_id:
        audit_query = audit_query.filter(UserActivityLog.report_id == audit_report_id)
    if audit_from:
        audit_query = audit_query.filter(UserActivityLog.created_at >= datetime.combine(audit_from, datetime.min.time()))
    if audit_to:
        audit_query = audit_query.filter(UserActivityLog.created_at <= datetime.combine(audit_to, datetime.max.time()))

    audit_logs = audit_query.order_by(UserActivityLog.created_at.desc()).limit(200).all()
    workflow_reports = (
        PBOReport.query
        .options(
            selectinload(PBOReport.user),
            selectinload(PBOReport.last_modified_by),
            selectinload(PBOReport.reviewed_by),
            selectinload(PBOReport.uploaded_files).selectinload(UploadedFile.uploaded_by),
            selectinload(PBOReport.field_changes),
        )
        .filter(PBOReport.workflow_status.in_(['submitted', 'validated', 'in_review', 'returned']))
        .order_by(PBOReport.updated_at.desc(), PBOReport.id.desc())
        .limit(100)
        .all()
    )
    for report in workflow_reports:
        report.owner_display = report_owner_display(report)
    import_batches = (
        ImportBatch.query
        .options(selectinload(ImportBatch.row_errors), selectinload(ImportBatch.uploaded_files))
        .order_by(ImportBatch.created_at.desc())
        .limit(20)
        .all()
    )
    recent_report_files = (
        PBOReport.query
        .options(selectinload(PBOReport.user))
        .order_by(PBOReport.updated_at.desc(), PBOReport.id.desc())
        .limit(30)
        .all()
    )
    deleted_file_logs = (
        UserActivityLog.query
        .options(
            selectinload(UserActivityLog.user),
            selectinload(UserActivityLog.report),
        )
        .filter(UserActivityLog.action == 'report_file_deleted')
        .order_by(UserActivityLog.created_at.desc())
        .limit(100)
        .all()
    )
    deleted_file_rows = []
    for log in deleted_file_logs:
        payload = parse_deleted_file_activity_summary(log.summary)
        uploaded_by_parts = [
            payload.get('uploaded_by_name') or '',
            payload.get('uploaded_by_email') or '',
        ]
        uploaded_by_display = " / ".join([part for part in uploaded_by_parts if part]) or 'Unknown'
        deleted_file_rows.append({
            'index': payload.get('file_id') or 'N/A',
            'file_name': payload.get('file_name') or 'Unknown file',
            'category': payload.get('category') or 'uncategorized',
            'uploaded_by': uploaded_by_display,
            'deleted_by': log.user,
            'report': log.report,
            'deleted_at': log.created_at,
        })
    dashboard_metrics = {
        'pending_users': User.query.filter_by(is_authorized=False).count(),
        'reports_pending_review': PBOReport.query.filter(PBOReport.workflow_status.in_(['submitted', 'validated', 'in_review'])).count(),
        'reports_returned': PBOReport.query.filter_by(workflow_status='returned').count(),
        'reports_approved': PBOReport.query.filter_by(workflow_status='approved').count(),
        'duplicate_reports': PBOReport.query.filter_by(duplicate_flag=True).count(),
        'imports_with_errors': ImportBatch.query.filter(ImportBatch.error_rows > 0).count(),
    }
    audit_actions = [row[0] for row in db.session.query(UserActivityLog.action).distinct().order_by(UserActivityLog.action.asc()).all()]
    rollback_catalog = {
        'local_versions': [],
        'google_versions': [],
        'google_error': None,
        'generated_at': None,
    }
    if current_user_is_superadmin:
        rollback_catalog = get_backup_rollback_catalog(force=request.args.get('refresh_backups') == '1')
    postgres_fk_audit = collect_postgresql_required_foreign_key_audit() if run_fk_audit else None
    postgres_status_audit = collect_postgresql_status_constraint_audit() if run_fk_audit else None

    return render_template(
        'admin_view.html',
        deadlines=deadlines,
        merged_report_files=list_report_merge_history(),
        report_merge_sources=get_report_merge_source_definitions(),
        report_merge_status=report_merge_status_snapshot(),
        report_merge_reference_label=latest_report_merge_reference_label(),
        user_activity_rows=user_activity_rows,
        user_activity_file_rows=user_activity_file_rows,
        activity_search_query=activity_search,
        selected_activity_date=request.args.get('activity_date', ''),
        user_activity_filters={
            'activity_date': request.args.get('activity_date', ''),
            'activity_search': activity_search,
        },
        audit_logs=audit_logs,
        audit_actions=audit_actions,
        audit_filters={
            'action': audit_action,
            'user_id': audit_user_id,
            'report_id': audit_report_id,
            'from_date': request.args.get('from_date', ''),
            'to_date': request.args.get('to_date', ''),
        },
        workflow_reports=workflow_reports,
        dashboard_metrics=dashboard_metrics,
        import_batches=import_batches,
        recent_report_files=recent_report_files,
        deleted_file_rows=deleted_file_rows,
        users=users,
        backup_runtime_status=backup_status_snapshot(),
        rollback_catalog=rollback_catalog,
        postgres_fk_audit=postgres_fk_audit,
        postgres_status_audit=postgres_status_audit,
        run_fk_audit=run_fk_audit,
        current_user_is_superadmin=current_user_is_superadmin,
    )


@app.route('/admin/user-activity-panel')
@admin_required
@login_required
def admin_user_activity_panel():
    activity_date = parse_date(request.args.get('activity_date'))
    activity_search = request.args.get('activity_search', '').strip()
    users = (
        User.query
        .options(selectinload(User.reports), selectinload(User.activity_logs))
        .order_by(User.email.asc())
        .all()
    )
    user_activity_rows = build_admin_user_activity_rows(users, activity_date=activity_date)
    user_activity_file_rows = build_admin_user_activity_file_rows(activity_search, activity_date=activity_date)
    return render_template(
        '_admin_user_activity_table.html',
        user_activity_rows=user_activity_rows,
        user_activity_file_rows=user_activity_file_rows,
        selected_activity_date=request.args.get('activity_date', ''),
        activity_search_query=activity_search,
    )


@app.route('/admin/user-activity/export')
@admin_required
@login_required
def admin_user_activity_export():
    activity_date = parse_date(request.args.get('activity_date'))
    users = (
        User.query
        .options(selectinload(User.reports), selectinload(User.activity_logs))
        .order_by(User.email.asc())
        .all()
    )
    user_activity_rows = build_admin_user_activity_rows(users, activity_date=activity_date)
    export_rows = []
    for row in user_activity_rows:
        user = row['user']
        last_worked_report = row.get('last_worked_report')
        export_rows.append({
            'Activity Date': activity_date.isoformat() if activity_date else 'All Records',
            'User Name': row.get('display_email') or user.email or f'User #{user.id}',
            'Full Name': row.get('display_full_name') or user.full_name or '',
            'Department': user.department or '',
            'No of Files Worked On': int(row.get('worked_file_count') or 0),
            'Last Login': format_datetime(row.get('display_last_login_at')) if row.get('display_last_login_at') else '',
            'Name of Last File Worked On': (
                last_worked_report.pbo_name
                if last_worked_report and last_worked_report.pbo_name
                else (f"Report #{last_worked_report.id}" if last_worked_report else '')
            ),
            'Last File Update Time': format_datetime(row.get('last_worked_at')) if row.get('last_worked_at') else '',
        })

    pandas_module = get_pandas()
    dataframe = pandas_module.DataFrame(export_rows)
    buffer = io.BytesIO()
    with pandas_module.ExcelWriter(buffer, engine='openpyxl') as writer:
        dataframe.to_excel(writer, index=False, sheet_name='User Activity')
        worksheet = writer.sheets['User Activity']
        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 40)

    buffer.seek(0)
    timestamp = utc_now().strftime('%Y%m%d_%H%M%S')
    filename_prefix = (
        f'user_activity_{activity_date.strftime("%Y%m%d")}'
        if activity_date else
        'user_activity_all'
    )
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{filename_prefix}_{timestamp}.xlsx',
    )


@app.route('/admin/users/<int:user_id>/worked-files')
@admin_required
@login_required
def admin_user_worked_files(user_id):
    target_user = User.query.get_or_404(user_id)
    search_query = request.args.get('q', '').strip()
    aggregation_email = normalize_user_email_key(request.args.get('email'))
    matched_users = [target_user]
    if aggregation_email:
        grouped_matches = find_users_by_email(aggregation_email)
        if grouped_matches:
            matched_users = grouped_matches

    canonical_user = choose_preferred_user_candidate(matched_users) or target_user
    matched_user_ids = sorted({
        user.id for user in matched_users
        if user is not None and getattr(user, 'id', None) is not None
    })

    touched_report_map = combine_worked_report_last_touched(
        build_form_submitted_report_index(user_ids=matched_user_ids),
        matched_user_ids,
    )
    touched_rows = sorted(
        touched_report_map.items(),
        key=lambda item: (item[1] or datetime.min, item[0]),
        reverse=True,
    )

    report_ids = [report_id for report_id, _ in touched_rows if report_id]
    reports_by_id = {}
    if report_ids:
        reports_by_id = {
            report.id: report
            for report in PBOReport.query.filter(PBOReport.id.in_(report_ids)).all()
        }

    worked_reports = []
    for report_id, last_touched_at in touched_rows:
        report = reports_by_id.get(report_id)
        if not report:
            continue
        if search_query and not report_matches_search_query(report, search_query):
            continue
        worked_reports.append({
            'report': report,
            'last_touched_at': last_touched_at,
        })

    return render_template(
        'admin_user_worked_files.html',
        target_user=canonical_user,
        worked_reports=worked_reports,
        total_worked_file_count=len(touched_report_map),
        filtered_worked_file_count=len(worked_reports),
        search_query=search_query,
        aggregation_email=aggregation_email,
        aggregation_user_count=len(matched_user_ids),
    )


@app.route('/admin/field-help-rules', methods=['GET', 'POST'])
@admin_required
def admin_field_help_rules():
    def as_bool(value):
        return str(value or '').strip().lower() in {'1', 'true', 'yes', 'on'}

    selected_page_key = normalize_field_help_page_key(request.values.get('page_key') or 'report_edit')

    if request.method == 'POST':
        action = (request.form.get('action') or 'update').strip().lower()
        selected_page_key = normalize_field_help_page_key(request.form.get('page_key') or selected_page_key)

        if action == 'create':
            field_name = (request.form.get('field_name') or '').strip()
            field_label = (request.form.get('field_label') or '').strip()
            if not field_name or not field_label:
                flash('Field name and field label are required.', 'warning')
                return redirect(url_for('admin_field_help_rules', page_key=selected_page_key))

            rule = FieldHelpRule(
                page_key=selected_page_key,
                field_name=field_name[:160],
                field_label=field_label[:255],
                heading=((request.form.get('heading') or '').strip() or None),
                is_allowed=as_bool(request.form.get('is_allowed')),
                is_disabled=as_bool(request.form.get('is_disabled')),
                is_required=as_bool(request.form.get('is_required')),
                used_in_sector_report=as_bool(request.form.get('used_in_sector_report')),
                notes=((request.form.get('notes') or '').strip() or None),
            )
            if rule.is_disabled:
                rule.is_allowed = False
            db.session.add(rule)
            db.session.commit()
            flash('Field-help rule created.', 'success')
        elif action == 'delete':
            rule_id = request.form.get('rule_id', type=int)
            rule = db.session.get(FieldHelpRule, rule_id) if rule_id else None
            if not rule:
                flash('Rule not found.', 'warning')
            else:
                selected_page_key = rule.page_key
                db.session.delete(rule)
                db.session.commit()
                flash('Field-help rule deleted.', 'info')
        else:
            rule_id = request.form.get('rule_id', type=int)
            rule = db.session.get(FieldHelpRule, rule_id) if rule_id else None
            if not rule:
                flash('Rule not found.', 'warning')
            else:
                rule.page_key = selected_page_key
                rule.field_name = (request.form.get('field_name') or rule.field_name or '').strip()[:160]
                rule.field_label = (request.form.get('field_label') or rule.field_label or '').strip()[:255]
                rule.heading = ((request.form.get('heading') or '').strip() or None)
                rule.notes = ((request.form.get('notes') or '').strip() or None)
                rule.is_allowed = as_bool(request.form.get('is_allowed'))
                rule.is_disabled = as_bool(request.form.get('is_disabled'))
                rule.is_required = as_bool(request.form.get('is_required'))
                rule.used_in_sector_report = as_bool(request.form.get('used_in_sector_report'))
                if rule.is_disabled:
                    rule.is_allowed = False
                db.session.commit()
                flash('Field-help rule updated.', 'success')

        try:
            refreshed_context = get_field_help_context_for_page(selected_page_key)
            ensure_field_help_decision_model(selected_page_key, refreshed_context)
        except Exception:
            app.logger.exception('Failed to refresh field-help model after admin rule update')

        return redirect(url_for('admin_field_help_rules', page_key=selected_page_key))

    page_key_rows = (
        db.session.query(FieldHelpRule.page_key)
        .distinct()
        .order_by(FieldHelpRule.page_key.asc())
        .all()
    )
    known_page_keys = sorted({row[0] for row in page_key_rows if row and row[0]} | {'report_edit', 'global'})
    rules = (
        FieldHelpRule.query
        .filter_by(page_key=selected_page_key)
        .order_by(FieldHelpRule.is_disabled.desc(), FieldHelpRule.field_label.asc())
        .all()
    )
    metrics = {
        'total': len(rules),
        'allowed': sum(1 for rule in rules if rule.is_allowed and not rule.is_disabled),
        'disabled': sum(1 for rule in rules if rule.is_disabled),
        'required': sum(1 for rule in rules if rule.is_required),
    }
    return render_template(
        'admin_field_help_rules.html',
        rules=rules,
        metrics=metrics,
        page_keys=known_page_keys,
        selected_page_key=selected_page_key,
    )


@app.route('/admin/data-analysis-bot')
@admin_required
@require_data_analysis_bot
def admin_data_analysis_bot():
    cache_key = f"{normalize_field_help_page_key(DATA_ANALYSIS_PAGE_KEY)}:rf"
    cached_model = DATA_ANALYSIS_MODEL_CACHE.get(cache_key) or {}
    recent_rows = (
        DataAnalysisBotInteraction.query
        .options(selectinload(DataAnalysisBotInteraction.user))
        .order_by(DataAnalysisBotInteraction.created_at.desc())
        .limit(50)
        .all()
    )
    persisted_labels = (get_admin_setting('data_analysis_model_labels', '') or '').strip()
    training_row_count = DataAnalysisTrainingQuestion.query.filter_by(
        page_key=normalize_field_help_page_key(DATA_ANALYSIS_PAGE_KEY)
    ).count()
    metrics = {
        'model_available': bool(cached_model.get('available')) or (get_admin_setting('data_analysis_model_available', '0') == '1'),
        'accuracy': float(cached_model.get('accuracy', get_admin_setting('data_analysis_model_accuracy', 0.0)) or 0.0),
        'labels': cached_model.get('labels', []) or ([label for label in persisted_labels.split(',') if label] if persisted_labels else []),
        'sample_count': len(cached_model.get('samples', [])) if cached_model.get('available') else training_row_count,
    }
    external_links = {
        target['form_key']: (get_admin_setting(target['setting_key'], '') or '')
        for target in DATA_ANALYSIS_EXTERNAL_JSON_TARGETS
    }
    external_sync_meta = {
        'last_at': get_admin_setting('data_analysis_external_sync_last_at', ''),
    }
    train_job = get_data_analysis_train_job_snapshot()
    return render_template(
        'admin_data_analysis_bot.html',
        metrics=metrics,
        recent_rows=recent_rows,
        external_links=external_links,
        external_sync_meta=external_sync_meta,
        external_json_targets=DATA_ANALYSIS_EXTERNAL_JSON_TARGETS,
        train_job=train_job,
    )


@app.route('/admin/data-analysis-bot/train', methods=['POST'])
@admin_required
@require_data_analysis_bot
def admin_data_analysis_bot_train():
    replace_rows = (request.form.get('replace') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
    with DATA_ANALYSIS_TRAIN_LOCK:
        if DATA_ANALYSIS_TRAIN_JOB.get('running'):
            flash('Training is already running in background. Refresh this page in a bit for status.', 'warning')
            return redirect(url_for('admin_data_analysis_bot'))
        DATA_ANALYSIS_TRAIN_JOB['running'] = True
        DATA_ANALYSIS_TRAIN_JOB['started_at'] = utc_now().isoformat()
        DATA_ANALYSIS_TRAIN_JOB['finished_at'] = None
        DATA_ANALYSIS_TRAIN_JOB['last_error'] = ''
        DATA_ANALYSIS_TRAIN_JOB['last_result'] = {}

    worker = threading.Thread(
        target=_background_data_analysis_train_job,
        args=(replace_rows,),
        daemon=True,
        name='data-analysis-train-worker',
    )
    worker.start()
    flash('Reload + Train started in background. You can keep using the page and refresh status shortly.', 'info')
    return redirect(url_for('admin_data_analysis_bot'))


@app.route('/admin/data-analysis-bot/external-sync', methods=['POST'])
@admin_required
@require_data_analysis_bot
def admin_data_analysis_bot_external_sync():
    link_values = {
        target['form_key']: (request.form.get(target['form_key']) or '').strip()
        for target in DATA_ANALYSIS_EXTERNAL_JSON_TARGETS
    }
    upload_values = {
        target.get('file_key') or f"{target['form_key']}_file": request.files.get(target.get('file_key') or f"{target['form_key']}_file")
        for target in DATA_ANALYSIS_EXTERNAL_JSON_TARGETS
    }
    missing_labels = [
        target['label']
        for target in DATA_ANALYSIS_EXTERNAL_JSON_TARGETS
        if not link_values.get(target['form_key']) and not ((upload_values.get(target.get('file_key') or f"{target['form_key']}_file") or None) and (upload_values.get(target.get('file_key') or f"{target['form_key']}_file").filename or '').strip())
    ]
    if missing_labels:
        flash(
            "Provide all three datasets (link or upload per file) before syncing: " + ", ".join(missing_labels) + ".",
            'warning',
        )
        return redirect(url_for('admin_data_analysis_bot') + '#external-json-sync')

    try:
        sync_result = sync_data_analysis_json_from_sources(link_values, upload_files=upload_values)
    except Exception as exc:
        flash(f'External JSON sync failed: {exc}', 'danger')
        return redirect(url_for('admin_data_analysis_bot') + '#external-json-sync')

    for target in DATA_ANALYSIS_EXTERNAL_JSON_TARGETS:
        form_key = target['form_key']
        link_value = (link_values.get(form_key) or '').strip()
        if link_value:
            set_admin_setting(target['setting_key'], link_value, current_user)
    set_admin_setting('data_analysis_external_sync_last_at', sync_result.get('cached_at', ''), current_user)
    db.session.commit()

    file_summary = "; ".join(
        f"{row['label']}: {row['question_count']} questions ({row['size_mb']} MB)"
        for row in (sync_result.get('files') or [])
    )
    flash(f'External links synced and cached locally. {file_summary}', 'success')

    run_train = (request.form.get('train_after_sync') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
    if run_train:
        replace_rows = (request.form.get('replace_after_sync') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
        with DATA_ANALYSIS_TRAIN_LOCK:
            if DATA_ANALYSIS_TRAIN_JOB.get('running'):
                flash('Sync complete. Training is already running in background.', 'warning')
            else:
                DATA_ANALYSIS_TRAIN_JOB['running'] = True
                DATA_ANALYSIS_TRAIN_JOB['started_at'] = utc_now().isoformat()
                DATA_ANALYSIS_TRAIN_JOB['finished_at'] = None
                DATA_ANALYSIS_TRAIN_JOB['last_error'] = ''
                DATA_ANALYSIS_TRAIN_JOB['last_result'] = {}
                worker = threading.Thread(
                    target=_background_data_analysis_train_job,
                    args=(replace_rows,),
                    daemon=True,
                    name='data-analysis-train-worker-sync',
                )
                worker.start()
                flash('Sync complete. Reload + train started in background.', 'info')
    return redirect(url_for('admin_data_analysis_bot') + '#external-json-sync')


@app.route('/admin/data-analysis-bot/train-status')
@admin_required
@require_data_analysis_bot
def admin_data_analysis_bot_train_status():
    snapshot = get_data_analysis_train_job_snapshot()
    return jsonify(snapshot)


@app.route('/admin/data-analysis-bot/context')
@admin_required
@require_data_analysis_bot
def admin_data_analysis_bot_context():
    dataset = load_data_analysis_dataset()
    field_context = get_field_help_context_for_page(DATA_ANALYSIS_PAGE_KEY)
    keyword_schema = dataset.get('keyword_schema') or {}
    return jsonify({
        'enabled_fields': dataset.get('enabled_fields', []),
        'allowed_fields': field_context.get('allowed_summary', []),
        'disabled_fields': field_context.get('disallowed_summary', []),
        'intents': dataset.get('intents', []),
        'question_count': len(dataset.get('questions', [])),
        'keyword_sources': keyword_schema.get('sources', []),
        'keyword_count': len(keyword_schema.get('global_keywords', [])),
    })


def _admin_data_analysis_ask(question):
    result = answer_data_analysis_question(question, page_key=DATA_ANALYSIS_PAGE_KEY)
    result_url = None
    if result.get('result_token'):
        result_url = url_for('admin_data_analysis_result', token=result['result_token'])

    db.session.add(DataAnalysisBotInteraction(
        page_key=DATA_ANALYSIS_PAGE_KEY,
        question=(question or '').strip(),
        intent=result.get('intent'),
        confidence=float(result.get('confidence') or 0.0),
        answer=(result.get('answer') or '').strip(),
        result_token=result.get('result_token'),
        result_path=result.get('result_path'),
        user_id=current_user.id if current_user.is_authenticated else None,
    ))
    db.session.commit()

    return {
        'answer': result.get('answer'),
        'intent': result.get('intent'),
        'confidence': float(result.get('confidence') or 0.0),
        'domain': result.get('domain'),
        'domain_source': result.get('domain_source'),
        'domain_score': float(result.get('domain_score') or 0.0),
        'domain_terms': result.get('domain_terms', []),
        'domain_columns': result.get('domain_columns', []),
        'blocked': bool(result.get('blocked')),
        'module': (result.get('result_payload') or {}).get('module'),
        'title': (result.get('result_payload') or {}).get('title'),
        'summary_lines': (result.get('result_payload') or {}).get('summary_lines', []),
        'result_url': result_url,
        'similar_questions': result.get('similar_questions', []),
        'factor_summary': result.get('factor_summary', {}),
    }


@app.route('/admin/data-analysis-bot/ask', methods=['POST'])
@admin_required
@require_data_analysis_bot
def admin_data_analysis_bot_ask():
    payload = request.get_json(silent=True) or {}
    question = (payload.get('question') or '').strip()
    if not question:
        return jsonify({'error': 'Ask a data analysis question first.'}), 400
    return jsonify(_admin_data_analysis_ask(question))


@app.route('/admin/ai/analyze-question', methods=['POST'])
@admin_required
@require_data_analysis_bot
def admin_ai_analyze_question():
    payload = request.get_json(silent=True) or {}
    question = (payload.get('question') or '').strip()
    if not question:
        return jsonify({'error': 'Ask a data analysis question first.'}), 400
    return jsonify(_admin_data_analysis_ask(question))


@app.route('/admin/data-analysis-bot/result/<token>')
@admin_required
@require_data_analysis_bot
def admin_data_analysis_result(token):
    safe_token = re.sub(r'[^a-z0-9]+', '', (token or '').lower())
    if not safe_token:
        abort(404)
    file_path = DATA_ANALYSIS_RESULT_DIR / f'{safe_token}.html'
    if not file_path.exists():
        abort(404)
    return send_file(file_path, mimetype='text/html')


@app.route('/admin/backup', methods=['POST'])
@admin_required

def admin_backup_database():
    if start_application_backup(triggered_by='admin'):
        flash('Database backup started. Progress will update below.', 'success')
    else:
        flash('A database backup is already running. Watch the progress panel below.', 'warning')
    return redirect(url_for('admin_view'))


@app.route('/admin/backup/status')
@admin_required
def admin_backup_status():
    return jsonify(backup_status_snapshot())


@app.route('/admin/export/reports/merge', methods=['POST'])
@admin_required
def admin_start_report_merge():
    selected_source = (request.form.get('merge_related_source') or '').strip()
    merge_upload = request.files.get('merge_upload_file')
    anchor = url_for('admin_view') + '#report-merge-tools'

    if report_merge_status_snapshot().get('running'):
        flash('A report merge is already running. Watch the export panel below.', 'warning')
        return redirect(anchor)

    if selected_source and selected_source not in get_report_merge_source_definition_map():
        flash('Choose a valid live database merge source.', 'warning')
        return redirect(anchor)

    has_upload = bool(merge_upload and merge_upload.filename)
    if not selected_source and not has_upload:
        flash('Choose a live database source or upload a merge workbook first.', 'warning')
        return redirect(anchor)

    uploaded_file_id = None
    uploaded_name = None
    if has_upload:
        ext = merge_upload.filename.rsplit('.', 1)[1].lower() if '.' in merge_upload.filename else ''
        if ext not in {'csv', 'xlsx', 'xls'}:
            flash('Uploaded merge files must be CSV or Excel workbooks.', 'danger')
            return redirect(anchor)
        uploaded_merge_file = save_uploaded_file(
            merge_upload,
            category='report_merge_upload',
            user=current_user,
            status='uploaded',
        )
        db.session.flush()
        ensure_persisted_primary_key(uploaded_merge_file, label='uploaded merge file')
        uploaded_file_id = uploaded_merge_file.id
        uploaded_name = uploaded_merge_file.original_filename

    selected_labels = []
    if selected_source:
        selected_labels.append(get_report_merge_source_definition_map()[selected_source]['label'])
    if uploaded_name:
        selected_labels.append(f'Upload: {uploaded_name}')
    log_user_activity(
        'report_merge_requested',
        user=current_user,
        summary=f"{current_user.email} requested report merge using {', '.join(selected_labels) or 'no sources'}",
    )
    db.session.commit()

    started = start_report_merge_export_job(
        triggered_by_user_id=current_user.id,
        selected_source_key=selected_source,
        uploaded_file_id=uploaded_file_id,
    )
    if started:
        flash('Report merge started. Progress will update below.', 'success')
    else:
        flash('A report merge is already running. Watch the export panel below.', 'warning')
    return redirect(anchor)


@app.route('/admin/export/reports/status')
@admin_required
def admin_report_merge_status():
    return jsonify(report_merge_status_snapshot())


@app.route('/admin/export/reports/files/<path:file_name>')
@admin_required
def admin_report_merge_file_download(file_name):
    root = report_merge_export_directory()
    selected = (file_name or '').strip()
    if not selected:
        abort(404)
    target = (root / selected).resolve()
    if not _path_within_base(target, root):
        abort(404)
    if not target.exists() or not target.is_file():
        abort(404)
    return send_file(
        target,
        as_attachment=True,
        download_name=report_merge_download_name(target.name),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/admin/export/reports/files/<path:file_name>/delete', methods=['POST'])
@admin_required
def admin_report_merge_file_delete(file_name):
    deleted = delete_report_merge_history_file(file_name)
    flash(
        'Merged workbook deleted.' if deleted else 'Merged workbook could not be found.',
        'success' if deleted else 'warning',
    )
    return redirect(url_for('admin_view') + '#report-merge-history')


@app.route('/admin/export/reports/files/delete-all', methods=['POST'])
@admin_required
def admin_report_merge_file_delete_all():
    deleted_count = delete_all_report_merge_history_files()
    if deleted_count:
        flash(f'Deleted {deleted_count} merged workbook(s).', 'success')
    else:
        flash('No merged workbooks were available to delete.', 'info')
    return redirect(url_for('admin_view') + '#report-merge-history')


@app.route('/admin/export/reports/download/<token>')
@admin_required
def admin_report_merge_download(token):
    item = REPORT_MERGE_OUTPUTS.get((token or '').strip())
    if not item:
        abort(404)
    file_path = item.get('path')
    if not file_path or not os.path.exists(file_path):
        abort(404)
    return send_file(
        file_path,
        as_attachment=True,
        download_name=item.get('download_name') or os.path.basename(file_path),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/admin/backup/rollback', methods=['POST'])
@admin_required
def admin_backup_rollback():
    denied = require_superadmin()
    if denied:
        return denied

    confirm_text = (request.form.get('confirm_rollback') or '').strip().upper()
    if confirm_text != 'ROLLBACK':
        flash('Rollback cancelled. Type ROLLBACK in the confirmation box to proceed.', 'warning')
        return redirect(url_for('admin_view') + '#rollback-tools')

    source_mode = (request.form.get('rollback_source_mode') or '').strip().lower()
    source_value = ''
    if source_mode == 'local':
        source_value = (request.form.get('local_backup_id') or '').strip()
    elif source_mode == 'google':
        source_value = (request.form.get('google_backup_id') or '').strip()

    try:
        rollback_result = perform_database_rollback(
            source_mode=source_mode,
            source_value=source_value,
            uploaded_file=request.files.get('rollback_upload_file'),
        )
        BACKUP_VERSION_CACHE['updated_at'] = 0.0
        BACKUP_VERSION_CACHE['payload'] = None
        try:
            log_user_activity(
                'database_rollback',
                summary=(
                    f"{current_user.email} rolled back database from {rollback_result.get('source_mode')} source; "
                    f"payload={rollback_result.get('payload_file')} ; pre_snapshot={rollback_result.get('pre_snapshot')}"
                ),
                user=current_user,
            )
            db.session.commit()
        except Exception:
            db.session.rollback()
        flash(
            'Database rollback completed successfully. '
            f"Pre-rollback snapshot saved at {rollback_result.get('pre_snapshot')}.",
            'success',
        )
    except Exception as exc:
        db.session.rollback()
        flash(f'Rollback failed: {exc}', 'danger')

    return redirect(url_for('admin_view') + '#rollback-tools')


@app.route('/admin/uploaded-files')
@admin_required
def admin_uploaded_files():
    denied = require_superadmin()
    if denied:
        return denied

    uploaded_files = (
        UploadedFile.query
        .options(
            selectinload(UploadedFile.report),
            selectinload(UploadedFile.batch),
            selectinload(UploadedFile.uploaded_by),
        )
        .order_by(UploadedFile.created_at.desc(), UploadedFile.id.desc())
        .all()
    )

    file_rows = [
        {
            'file': uploaded_file,
            'impacts': describe_uploaded_file_impacts(uploaded_file),
        }
        for uploaded_file in uploaded_files
    ]

    return render_template(
        'admin_uploaded_files.html',
        file_rows=file_rows,
        generated_at=format_datetime(utc_now()),
    )


@app.route('/admin/uploaded-files/<int:file_id>/delete', methods=['POST'])
@admin_required
def admin_delete_uploaded_file(file_id):
    denied = require_superadmin()
    if denied:
        return denied

    uploaded_file = (
        UploadedFile.query
        .options(
            selectinload(UploadedFile.report),
            selectinload(UploadedFile.batch),
            selectinload(UploadedFile.uploaded_by),
        )
        .filter(UploadedFile.id == file_id)
        .first()
        or abort(404)
    )

    linked_report = uploaded_file.report
    linked_batch = uploaded_file.batch
    log_user_activity(
        'report_file_deleted',
        report=linked_report,
        summary=build_deleted_file_activity_summary(uploaded_file),
        user=current_user if current_user.is_authenticated else None,
    )

    storage_path = uploaded_file.storage_path
    if storage_path and os.path.exists(storage_path):
        try:
            os.remove(storage_path)
        except OSError:
            app.logger.warning('Could not remove uploaded file from disk: %s', storage_path)

    db.session.delete(uploaded_file)
    log_user_activity(
        'superadmin_uploaded_file_deleted',
        report=linked_report,
        user=current_user,
        summary=(
            f'{current_user.email} deleted uploaded file {file_id} '
            f'({linked_report.pbo_name if linked_report and linked_report.pbo_name else uploaded_file.original_filename})'
            f'{f" from import batch #{linked_batch.id}" if linked_batch else ""}'
        ),
    )
    db.session.commit()
    flash(f'Uploaded file {uploaded_file.original_filename} deleted successfully.', 'success')
    return redirect(url_for('admin_uploaded_files'))


@app.route('/admin/audit/export.csv')
@admin_required
def export_audit_logs():
    audit_action = request.args.get('action', '').strip()
    audit_user_id = request.args.get('user_id', type=int)
    audit_report_id = request.args.get('report_id', type=int)
    audit_from = parse_date(request.args.get('from_date'))
    audit_to = parse_date(request.args.get('to_date'))

    query = UserActivityLog.query.options(selectinload(UserActivityLog.user), selectinload(UserActivityLog.report))
    if audit_action:
        query = query.filter(UserActivityLog.action == audit_action)
    if audit_user_id:
        query = query.filter(UserActivityLog.user_id == audit_user_id)
    if audit_report_id:
        query = query.filter(UserActivityLog.report_id == audit_report_id)
    if audit_from:
        query = query.filter(UserActivityLog.created_at >= datetime.combine(audit_from, datetime.min.time()))
    if audit_to:
        query = query.filter(UserActivityLog.created_at <= datetime.combine(audit_to, datetime.max.time()))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['created_at', 'action', 'user', 'report_id', 'route', 'summary', 'ip_address', 'user_agent'])
    for log in query.order_by(UserActivityLog.created_at.desc()).all():
        writer.writerow([
            log.created_at.isoformat(sep=' ', timespec='seconds') if log.created_at else '',
            log.action,
            log.user.email if log.user else '',
            log.report_id or '',
            log.route or '',
            log.summary or '',
            log.ip_address or '',
            log.user_agent or '',
        ])
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=audit_logs.csv'},
    )


@app.route('/admin/report/<int:report_id>/workflow', methods=['POST'])
@admin_required
def update_report_workflow(report_id):
    report = db.session.get(PBOReport, report_id) or abort(404)
    denied = require_report_access(report, write=True)
    if denied:
        return denied
    before = report_snapshot(report)
    action = request.form.get('workflow_action', '').strip().lower()
    notes = (request.form.get('review_notes') or '').strip()
    return_reason = (request.form.get('return_reason') or '').strip()

    if action == 'validate':
        report.workflow_status = 'validated'
        report.review_status = 'pending'
    elif action == 'start_review':
        report.workflow_status = 'in_review'
        report.review_status = 'reviewed'
    elif action == 'approve':
        report.workflow_status = 'approved'
        report.review_status = 'approved'
    elif action == 'return':
        report.workflow_status = 'returned'
        report.review_status = 'returned'
        report.return_reason = return_reason or notes or report.return_reason
    elif action == 'submit':
        report.workflow_status = 'submitted'
        report.review_status = 'pending'
    else:
        flash('Invalid workflow action.', 'danger')
        return redirect(url_for('admin_view'))

    report.review_notes = notes or report.review_notes
    report.reviewed_by_id = current_user.id
    report.reviewed_at = utc_now()
    record_report_field_changes(report, before, f'workflow_{action}', current_user)
    log_user_activity(f'workflow_{action}', report=report, summary=f'{current_user.email} set workflow to {report.workflow_status}')
    db.session.commit()
    flash('Workflow updated successfully.', 'success')
    return redirect(request.referrer or url_for('admin_view'))


@app.route('/admin/review/approve-all', methods=['POST'])
@admin_required
def approve_all_review_queue():
    review_notes = (request.form.get('review_notes') or '').strip()
    reports = (
        PBOReport.query
        .filter(PBOReport.workflow_status.in_(['submitted', 'validated', 'in_review', 'returned']))
        .order_by(PBOReport.updated_at.desc(), PBOReport.id.desc())
        .all()
    )

    if not reports:
        flash('No reports are currently in the review queue.', 'info')
        return redirect(url_for('admin_view'))

    approved_count = 0
    for report in reports:
        denied = require_report_access(report, write=True)
        if denied:
            continue
        before = report_snapshot(report)
        report.workflow_status = 'approved'
        report.review_status = 'approved'
        report.reviewed_by_id = current_user.id
        report.reviewed_at = utc_now()
        if review_notes:
            report.review_notes = review_notes
        record_report_field_changes(report, before, 'workflow_approve_all', current_user)
        log_user_activity(
            'workflow_approve_all',
            report=report,
            summary=f'{current_user.email} approved report #{report.id} from the review queue',
        )
        approved_count += 1

    db.session.commit()
    flash(f'Approved {approved_count} report(s) from the review queue.', 'success')
    return redirect(url_for('admin_view'))


@app.route('/admin/report/<int:report_id>/upload', methods=['POST'])
@admin_required
def upload_report_file(report_id):
    report = db.session.get(PBOReport, report_id) or abort(404)
    denied = require_report_access(report, write=True)
    if denied:
        return denied
    upload = request.files.get('file')
    category = (request.form.get('category') or 'supporting_document').strip()
    if upload is None or not upload.filename:
        flash('Choose a file to upload.', 'warning')
        return redirect(request.referrer or url_for('admin_view'))
    if not allowed_upload(upload.filename):
        flash('Unsupported file type.', 'danger')
        return redirect(request.referrer or url_for('admin_view'))

    save_uploaded_file(upload, category=category, report=report, user=current_user)
    log_user_activity('report_file_uploaded', report=report, summary=f'{current_user.email} uploaded {upload.filename}')
    db.session.commit()
    flash('File uploaded successfully.', 'success')
    return redirect(request.referrer or url_for('admin_view'))


@app.route('/admin/import/reports', methods=['POST'])
@admin_required
def import_reports():
    upload = request.files.get('import_file')
    if upload is None or not upload.filename:
        flash('Choose a CSV or Excel file to import.', 'warning')
        return redirect(url_for('admin_view'))
    if not allowed_upload(upload.filename):
        flash('Unsupported import file type.', 'danger')
        return redirect(url_for('admin_view'))

    batch = ImportBatch(
        source_type='file_import',
        original_filename=secure_filename(upload.filename),
        status='processing',
        created_by_id=current_user.id,
        notes=(request.form.get('notes') or '').strip() or None,
    )
    add_and_flush_new_instance(batch, label='import batch')
    uploaded_file = save_uploaded_file(upload, category='import_batch', batch=batch, user=current_user, status='processed')
    batch.stored_filename = uploaded_file.stored_filename

    try:
        rows = load_tabular_import_rows(uploaded_file.storage_path, uploaded_file.original_filename)
        batch.total_rows = len(rows)
        if env_flag('REPLACE_REPORTS_ON_IMPORT', False):
            clear_application_tables(
                preserve_tables={
                    'users_for_form14',
                    'admin_settings',
                    'postal_code_cache',
                    'import_batches',
                    'import_row_errors',
                    'uploaded_files',
                }
            )
        for index, raw_row in enumerate(rows, start=2):
            row = {str(key).strip(): raw_row.get(key) for key in raw_row.keys()}
            pbo_name = to_upper(row.get('pbo_name') or row.get('PBO Name'))
            registration_number = to_upper(row.get('pbo_registration_number') or row.get('registration_number'))
            if not pbo_name:
                create_import_error(batch, index, 'PBO name is required.', 'pbo_name', row)
                continue

            report = None
            if registration_number:
                report = PBOReport.query.filter_by(pbo_registration_number=registration_number).first()
            if report is None:
                report = find_duplicate_pbo_report(pbo_name)

            created = False
            if report is None:
                report = PBOReport(
                    pbo_name=pbo_name,
                    pbo_registration_number=registration_number,
                    workflow_status='validated',
                    review_status='pending',
                    data_source='import',
                    import_batch_id=batch.id,
                    user_id=current_user.id,
                )
                add_and_flush_new_instance(report, label='imported report')
                created = True

            before = report_snapshot(report)
            report.pbo_name = pbo_name
            assign_pbo_name_normalized(report)
            report.pbo_registration_number = registration_number or report.pbo_registration_number
            set_reporting_period_fields(
                report,
                row.get('reporting_period_start') or row.get('Reporting Period Start'),
                row.get('reporting_period_end') or row.get('Reporting Period End'),
            )
            report.return_date = parse_return_date(row.get('return_date') or row.get('Return Date'))
            report.audited = to_upper(row.get('audited') or row.get('Audited'))
            report.late_returns = to_upper(row.get('late_returns') or row.get('Late Returns'))
            report.outstanding_penalty = row.get('outstanding_penalty') or row.get('Outstanding Penalty') or report.outstanding_penalty
            report.email = to_upper(row.get('email') or row.get('Email'))
            report.telephone = to_upper(row.get('telephone') or row.get('Telephone'))
            report.workflow_status = normalize_report_status(row.get('workflow_status'), default='validated')
            report.review_status = normalize_review_status(row.get('review_status'), default='pending')
            report.data_source = 'import'
            report.import_batch_id = batch.id
            report.duplicate_flag = find_duplicate_pbo_report(
                report.pbo_name,
                reporting_period_start=report.reporting_period_start,
                exclude_report_id=report.id,
                require_start_month_match=True,
            ) is not None
            record_report_field_changes(report, before, 'report_imported', current_user)
            log_user_activity('report_imported', report=report, summary=f'Import batch #{batch.id} {"created" if created else "updated"} report #{report.id}')
            batch.success_rows += 1

        batch.status = 'completed_with_errors' if batch.error_rows else 'completed'
        batch.completed_at = utc_now()
        db.session.commit()
        flash(f'Import finished. Success rows: {batch.success_rows}, errors: {batch.error_rows}.', 'success')
    except Exception as exc:
        batch.status = 'failed'
        batch.completed_at = utc_now()
        create_import_error(batch, 0, str(exc), 'file', {'filename': uploaded_file.original_filename})
        db.session.commit()
        flash(f'Import failed: {exc}', 'danger')

    return redirect(url_for('admin_view'))


# Route to update designate_by_pco field from dropdown
@app.route('/update_designate_by_pco/<int:report_id>', methods=['POST'])
@admin_required
def update_designate_by_pco(report_id):
    designate = request.form.get('designate_by_pco')
    report = db.session.get(PBOReport, report_id) or abort(404)
    denied = require_report_access(report, write=True)
    if denied:
        return denied
    report.designate_by_pco = designate
    log_user_activity('admin_designate_updated', report=report, summary=f'Designate updated to {designate or "N/A"}')
    db.session.commit()
    flash('Designate (By PCO) updated.', 'success')
    # Stay on the same page and preserve pagination/sort
    next_url = request.referrer or url_for('admin_view')
    return redirect(next_url)

def compute_tf_risk(report):
    score = 0
    summary = []
    # Financial compliance
    if getattr(report, 'late_returns', None) == "Yes":
        score += 2
        summary.append("Late returns: +2")
    if getattr(report, 'outstanding_penalty', None):
        try:
            if float(report.outstanding_penalty) > 0:
                score += 3
                summary.append("Outstanding penalty: +3")
        except Exception:
            pass
    if getattr(report, 'audited', None) == "No":
        score += 2
        summary.append("Not audited: +2")

    # Governance
    if getattr(report, 'number_of_directors', None) is not None and report.number_of_directors < 3:
        score += 2
        summary.append("<3 directors: +2")
    if not getattr(report, 'date_last_agm', None) or not getattr(report, 'date_last_board_meeting', None):
        score += 2
        summary.append("Missing AGM/board meeting: +2")

    # Operational risk
    pim = getattr(report, 'project_implementation_method', '') or ''
    if "cash" in pim.lower():
        score += 1
        summary.append("Cash-only project: +1")
    # Collaboration/networking high-risk regions
    collab = getattr(report, 'collaborations', [])
    high_risk_zones = ["somalia", "syria", "yemen"]
    found_zone = False
    for c in collab:
        for zone in high_risk_zones:
            if zone in (getattr(c, 'info_exchange', '') or '').lower():
                score += 3
                summary.append(f"Collab with {zone.title()}: +3")
                found_zone = True
                break
        if found_zone:
            break

    # Staff risk
    if getattr(report, 'volunteers_foreign_current', None) and report.volunteers_foreign_current > 5:
        score += 1
        summary.append(">5 foreign volunteers: +1")

    # Donations/grants: high influx
    total_donations = sum([d.amount or 0 for d in getattr(report, 'donations', [])])
    total_grants = sum([g.amount or 0 for g in getattr(report, 'grants', [])])
    if total_donations > 1_000_000:
        score += 2
        summary.append("High donations influx: +2")
    if total_grants > 1_000_000:
        score += 2
        summary.append("High grants influx: +2")

    return score, summary

# Admin TF risk view (DB pagination, eager loading, uses stored risk_score)
@app.route('/admin/tf_risk')
@require_tf_risk
@admin_required
def admin_tf_risk():
    page = request.args.get('page', 1, type=int)
    per_page_options = [10, 25, 50, 100]
    requested_per_page = request.args.get('per_page', type=int)
    if requested_per_page in per_page_options:
        per_page = requested_per_page
        session['tf_risk_per_page'] = per_page
    else:
        per_page = session.get('tf_risk_per_page', 25)
        if per_page not in per_page_options:
            per_page = 25
    search = request.args.get('search', '', type=str).strip()
    sort_by = request.args.get('sort_by', 'risk_score', type=str)
    sort_dir = request.args.get('sort_dir', 'desc', type=str)
    high_risk_threshold = 6

    query = PBOReport.eager_query().options(
        selectinload(PBOReport.user),
        selectinload(PBOReport.last_modified_by),
    )
    # Search by PBO name or summary
    if search:
        query = query.filter(PBOReport.pbo_name.ilike(f"%{search}%"))
    # Sorting
    if sort_by == 'pbo_name':
        order = PBOReport.pbo_name.asc() if sort_dir == 'asc' else PBOReport.pbo_name.desc()
    else:
        order = PBOReport.risk_score.asc() if sort_dir == 'asc' else PBOReport.risk_score.desc().nullslast()
    query = query.order_by(order)

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    reports = pagination.items
    risk_data = []
    for report in reports:
        # If risk_score is missing, compute and update it
        if report.risk_score is None:
            report.update_risk_score(compute_tf_risk)
            db.session.commit()
        score = report.risk_score
        _, summary = compute_tf_risk(report)
        risk_data.append({
            'report': report,
            'updater_name': report_updater_display(report),
            'score': score,
            'summary': summary,
            'is_high_risk': score >= high_risk_threshold
        })
    return render_template(
        'tf_risk_scores.html',
        risk_data=risk_data,
        high_risk_threshold=high_risk_threshold,
        page=page,
        total_pages=pagination.pages,
        total=pagination.total,
        per_page=per_page,
        per_page_options=per_page_options,
        search=search,
        sort_by=sort_by,
        sort_dir=sort_dir,
        max=max,
        min=min
    )



# Route to update any editable report field from the admin table
@app.route('/update_report_field/<int:report_id>', methods=['POST'])
@admin_required
def update_report_field(report_id):
    report = db.session.get(PBOReport, report_id) or abort(404)
    denied = require_report_access(report, write=True)
    if denied:
        return denied
    before = report_snapshot(report)
    editable_fields = [
        'pbo_name', 'filing_period', 'filling_fee', 'late_returns', 'penalty_paid',
        'outstanding_penalty', 'reporting_period_end', 'return_date', 'form_14', 'audited', 'requests',
        'date_received', 'received_by', 'filed_for_action_by_registry', 'date_filed_by_registry_for_action',
        'designate_by_pco', 'date_assigned', 'review_acknowledgement', 'date_acknowledged_notice_sent',
        'end_of_notice_period', 'notice_countdown'
    ]
    updated = False
    for field in editable_fields:
        if field in request.form:
            value = request.form[field]
            if field == 'return_date':
                setattr(report, field, parse_return_date(value))
            else:
                setattr(report, field, parse_date(value) if field == 'reporting_period_end' else value)
            updated = True
    if updated:
        report.pbo_name = to_upper(report.pbo_name)
        assign_pbo_name_normalized(report)
        report.duplicate_flag = find_duplicate_pbo_report(
            report.pbo_name,
            reporting_period_start=report.reporting_period_start,
            exclude_report_id=report.id,
            require_start_month_match=True,
        ) is not None
        record_report_field_changes(report, before, 'admin_report_updated', current_user)
        log_user_activity('admin_report_updated', report=report, summary='Admin updated report fields from admin view')
        db.session.commit()
        flash('Report updated.', 'success')
    return redirect(request.referrer or url_for('admin_view'))

# Route to add a new report row from the admin table
@app.route('/add_report_row', methods=['POST'])
@admin_required
def add_report_row():
    try:
        report = PBOReport(
            pbo_name=get_form_upper('pbo_name'),
            pbo_registration_number=get_form_upper('pbo_registration_number'),
            filing_period=request.form.get('filing_period'),
            filling_fee=request.form.get('filling_fee'),
            late_returns=request.form.get('late_returns'),
            penalty_paid=request.form.get('penalty_paid'),
            outstanding_penalty=request.form.get('outstanding_penalty'),
            form_14=request.form.get('form_14'),
            audited=request.form.get('audited'),
            requests=request.form.get('requests'),
            date_received=request.form.get('date_received'),
            received_by=request.form.get('received_by'),
            filed_for_action_by_registry=request.form.get('filed_for_action_by_registry'),
            date_filed_by_registry_for_action=request.form.get('date_filed_by_registry_for_action'),
            designate_by_pco=request.form.get('designate_by_pco'),
            date_assigned=request.form.get('date_assigned'),
            review_acknowledgement=request.form.get('review_acknowledgement'),
            date_acknowledged_notice_sent=request.form.get('date_acknowledged_notice_sent'),
            end_of_notice_period=request.form.get('end_of_notice_period'),
            notice_countdown=request.form.get('notice_countdown'),
            workflow_status='validated',
            review_status='pending',
            data_source='admin',
            user_id=current_user.id,
            return_date=parse_return_date(request.form.get('return_date')),
        )
        set_reporting_period_fields(
            report,
            request.form.get('reporting_period_start'),
            request.form.get('reporting_period_end'),
        )
        assign_pbo_name_normalized(report)
        report.duplicate_flag = find_duplicate_pbo_report(
            report.pbo_name,
            reporting_period_start=report.reporting_period_start,
            require_start_month_match=True,
        ) is not None
        add_and_flush_new_instance(report, label='admin report')
        report.update_risk_score(compute_tf_risk)
        record_report_field_changes(report, {field: '' for field in TRACKED_REPORT_FIELDS}, 'admin_report_created', current_user)
        log_user_activity('admin_report_created', report=report, summary='Admin created a report row from admin panel')
        db.session.commit()
        flash('Report row created successfully.', 'success')
        return redirect(url_for('report_detail', report_id=report.id))
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating report row: {str(e)}', 'error')
        return redirect(url_for('admin_view'))


@app.route('/validate-postal', methods=['POST'])
def validate_postal():
    payload = request.get_json(silent=True) or {}
    postal_code = (payload.get('postal_code') or '').strip()

    if not postal_code or not postal_code.isdigit() or len(postal_code) != 5:
        return jsonify({"error": "Invalid Kenyan postal code format"}), 400

    town, error, status = get_town_from_postal_code(postal_code)
    if error:
        response = {"error": error.get("message", "Unknown error")}
        if "retry_after" in error:
            response["retry_after"] = error["retry_after"]
        return jsonify(response), status

    return jsonify({
        "postal_code": postal_code,
        "town": town,
    })


@app.route('/api/check-pbo-name', methods=['GET'])
def check_pbo_name():
    name = request.args.get('name', '')
    exclude_report_id = request.args.get('exclude_report_id', type=int)
    duplicate_report = find_duplicate_pbo_report(
        name,
        reporting_period_start=parse_date(request.args.get('reporting_period_start')),
        exclude_report_id=exclude_report_id,
        require_start_month_match=True,
    )
    return jsonify({
        "duplicate": duplicate_report is not None,
        "report_id": duplicate_report.id if duplicate_report else None,
        "message": "",
    })


@app.route('/report/<int:report_id>')
@login_required
def report_detail(report_id):
    report = db.session.get(PBOReport, report_id) or abort(404)
    denied = require_report_access(report, write=False)
    if denied:
        return denied

    neighbor_reports = []
    if report.pbo_name_normalized and report.reporting_period_start:
        peers = (
            PBOReport.query
            .filter(PBOReport.pbo_name_normalized == report.pbo_name_normalized)
            .filter(PBOReport.reporting_period_start.isnot(None))
            .order_by(PBOReport.reporting_period_start.asc(), PBOReport.id.asc())
            .all()
        )
        current_start = report.reporting_period_start
        prev_report = None
        next_report = None
        for peer in peers:
            if peer.id == report.id:
                continue
            start = peer.reporting_period_start
            if start and start < current_start:
                prev_report = peer
            elif start and start > current_start and next_report is None:
                next_report = peer
                break

        def make_label(record):
            base = (record.pbo_name or "Report").strip().replace(" ", "_")
            year = record.reporting_period_start.year if record.reporting_period_start else "period"
            return f"{base}_{year}"

        for rel in (prev_report, next_report):
            if rel:
                neighbor_reports.append({
                    'id': rel.id,
                    'label': make_label(rel),
                    'url': url_for('report_detail', report_id=rel.id)
                })

    counties_list = [c.strip() for c in (report.counties or "").split(",") if c.strip()]
    countries_of_operation_list = [c.strip() for c in (report.countries_of_operation or "").split(",") if c.strip()]
    election_frequency_list = [e.strip() for e in (report.election_frequency or "").split(",") if e.strip()]
    staff_biodata_rows = get_display_staff_biodata_rows(report)
    volunteer_biodata_rows = get_display_volunteer_biodata_rows(report)

    report.auditor_info = {}
    report.js_totals = {}
    return_filing_date_display = return_date_display(report.return_date)

    context = {
        'report': report,
        'associated_username': report_associated_username_display(report),
        'reporting_period_start_display': reporting_period_display(report.reporting_period_start, report.reporting_period_start_raw),
        'reporting_period_end_display': reporting_period_display(report.reporting_period_end, report.reporting_period_end_raw),
        'return_date_display': return_filing_date_display,
        'return_filing_date_display': return_filing_date_display,
        'can_edit_report': can_edit_report_record(current_user, report),
        'countries_of_operation_list': countries_of_operation_list,
        'counties_list': counties_list,
        'election_frequency_list': election_frequency_list,
        'assets_list': report.assets,
        'donations_list': report.donations,
        'payments_list': report.payments,
        'grants_list': report.grants,
        'banking_list': report.bank_accounts,
        'auditors_list': report.auditors,
        'staff_biodata_list': staff_biodata_rows,
        'volunteer_biodata_list': volunteer_biodata_rows,
        'volunteer_privileges_list': report.volunteer_privileges,
        'training_records_list': report.training_records,
        'tax_waiver_list': report.tax_waiver_items,
        'officials_list': report.officials,
        'project_implementations_list': report.project_implementations,
        'projects_carried_out_list': report.projects_carried_out,
        'collaboration_list': report.collaborations,
        'neighbor_reports': neighbor_reports,
    }

    response = make_response(render_template('report_detail.html', **context))
    # Route-level no-cache marker to reduce stale report-detail renders in browser tabs.
    response.headers['X-Report-Detail-No-Cache'] = '1'
    response.headers['Surrogate-Control'] = 'no-store'
    return response


@app.route('/report/<int:report_id>/delete', methods=['POST'])
@login_required
def report_delete(report_id):
    report = db.session.get(PBOReport, report_id) or abort(404)
    denied = require_report_access(report, write=True)
    if denied:
        return denied

    report_name = report.pbo_name or f'Report #{report.id}'
    delete_report_uploaded_files(report)
    log_user_activity(
        'report_deleted',
        report=report,
        user=current_user if current_user.is_authenticated else None,
        summary=f'{getattr(current_user, "email", "Unknown")} deleted report #{report.id}',
    )
    db.session.delete(report)
    db.session.commit()
    flash(f'{report_name} was deleted successfully.', 'success')
    return redirect(url_for('my_files'))


@app.route('/report/<int:report_id>/edit', methods=['GET', 'POST'])
@login_required
def report_edit(report_id):
    report = db.session.get(PBOReport, report_id) or abort(404)
    denied = require_report_access(report, write=True)
    if denied:
        return denied
    if request.method == 'GET':
        log_user_activity('report_opened', report=report)
        db.session.commit()

    if request.method == 'POST':
        try:
            before = report_snapshot(report)
            maybe_assign_report_owner(report, current_user if current_user.is_authenticated else None)
            # Scalar updates (unchanged)
            set_reporting_period_fields(
                report,
                request.form.get('reporting_period_start'),
                request.form.get('reporting_period_end'),
            )
            report.return_date = parse_return_date(request.form.get('return_date'))
            report.pbo_name = get_form_upper('pbo_name')
            assign_pbo_name_normalized(report)
            report.pbo_registration_number = get_form_upper('pbo_registration_number')
            report.pbo_registration_date = parse_date(request.form.get('pbo_registration_date'))

            report.kra_pin = None
            report.postal_address = get_form_upper('postal_address')
            report.physical_address = get_form_upper('physical_address')
            report.telephone = get_form_phone('telephone', 'telephone_country_code')
            report.cell_phone = get_form_phone('cell_phone', 'cellphone_country_code')
            report.email = get_form_upper('email')
            report.website = get_form_upper('website')
            report.social_media = get_form_upper('social_media')

            report.contact_name = get_form_upper('contact_name')
            report.contact_position = get_form_upper('contact_position')
            report.contact_telephone = get_form_phone_joined('contact_telephone', 'contact_country_code')
            report.contact_email = get_form_text_joined_upper('contact_email')
            report.contact_nationality = request.form.get('contact_nationality')
            report.contact_gender = request.form.get('contact_gender')

            report.registration_number = get_form_upper('registration_number')
            report.pin_number = None
            report.date_of_registration = parse_date(request.form.get('date_of_registration'))
            report.scope = normalize_scope(request.form.get('scope'))

            countries_of_operation = get_form_list_upper('country_of_operation')
            report.countries_of_operation = ", ".join(countries_of_operation) if countries_of_operation else None
            counties = get_form_list_upper('county')
            report.counties = ", ".join(counties) if counties else None

            report.audited = get_form_upper('audited')
            report.assets_stolen = get_form_upper('assets_stolen')
            report.cash_balance_previous_year = parse_float(request.form.get('cash_balance_previous_year'))
            report.cash_bank_balance = parse_float(request.form.get('cash_bank_balance'))

            report.staff_kenyan_prev = parse_int(get_form_list_item('staff_kenyan[]', 0))
            report.staff_kenyan_current = parse_int(get_form_list_item('staff_kenyan[]', 1))
            report.staff_kenyan_came_in = parse_int(get_form_list_item('staff_kenyan[]', 2))
            report.staff_kenyan_left = parse_int(get_form_list_item('staff_kenyan[]', 3))
            report.staff_foreign_prev = parse_int(get_form_list_item('staff_foreign[]', 0))
            report.staff_foreign_current = parse_int(get_form_list_item('staff_foreign[]', 1))
            report.staff_foreign_came_in = parse_int(get_form_list_item('staff_foreign[]', 2))
            report.staff_foreign_left = parse_int(get_form_list_item('staff_foreign[]', 3))
            report.staff_other_kenyan_prev = parse_int(get_form_list_item('staff_other_kenyan[]', 0))
            report.staff_other_kenyan_current = parse_int(get_form_list_item('staff_other_kenyan[]', 1))
            report.staff_other_foreign_prev = parse_int(get_form_list_item('staff_other_foreign[]', 0))
            report.staff_other_foreign_current = parse_int(get_form_list_item('staff_other_foreign[]', 1))
            report.volunteers_kenyan_prev = parse_int(get_form_list_item('volunteers_kenyan[]', 0))
            report.volunteers_kenyan_current = parse_int(get_form_list_item('volunteers_kenyan[]', 1))
            report.volunteers_foreign_prev = parse_int(get_form_list_item('volunteers_foreign[]', 0))
            report.volunteers_foreign_current = parse_int(get_form_list_item('volunteers_foreign[]', 1))

            report.local_material = get_checkbox_value("local_material")
            report.local_material_amount = parse_float(request.form.get("local_material_amount"))
            report.local_labour = get_checkbox_value("local_labour")
            report.local_labour_amount = parse_float(request.form.get("local_labour_amount"))
            report.local_financial = get_checkbox_value("local_financial")
            report.local_financial_amount = parse_float(request.form.get("local_financial_amount"))
            report.local_other = get_checkbox_value("local_other")
            report.local_other_specify = get_form_upper("local_other_specify")
            report.local_other_amount = parse_float(request.form.get("local_other_amount"))

            government_sections_inactive = get_inactive_section_flag(
                "government_sections_inactive",
                "government_section_inactive",
            )
            if government_sections_inactive:
                report.gov_tax_waiver = None
                report.gov_tax_waiver_amount = None
                report.gov_other = None
                report.gov_other_specify = None
                report.gov_other_amount = None
            else:
                report.gov_tax_waiver = get_checkbox_value("gov_tax_waiver")
                report.gov_tax_waiver_amount = parse_float(request.form.get("gov_tax_waiver_amount"))
                report.gov_other = get_checkbox_value("gov_other")
                report.gov_other_specify = get_form_upper("gov_other_specify")
                report.gov_other_amount = parse_float(request.form.get("gov_other_amount"))

            def section_posted(*keys):
                return any(key in request.form for key in keys)

            # Assets
            if section_posted('assets_item[]', 'assets_number[]', 'assets_value[]'):
                report.assets.clear()
                items = request.form.getlist('assets_item[]')
                nums = request.form.getlist('assets_number[]')
                vals = request.form.getlist('assets_value[]')
                for i in range(min(len(items), len(nums), len(vals))):
                    if items[i] or nums[i] or vals[i]:
                        report.assets.append(Asset(
                            item=items[i],
                            number=parse_int(nums[i]),
                            value=parse_float(vals[i]),
                        ))

            # Donations
            if section_posted('donor_name[]', 'donor_category[]', 'donor_country[]', 'donor_amount[]'):
                report.donations.clear()
                dn = request.form.getlist('donor_name[]')
                dc = get_form_list_with_other('donor_category[]', 'donor_category_other[]', trigger_value='OTHER')
                dco = get_form_list_with_other(
                    'donor_country[]',
                    'donor_country_other[]',
                    trigger_value='OTHER_COUNTRY',
                )
                da = request.form.getlist('donor_amount[]')
                donor_rows = max(len(dn), len(dc), len(dco), len(da))
                for i in range(donor_rows):
                    donor_name = dn[i] if i < len(dn) else None
                    donor_category = dc[i] if i < len(dc) else None
                    donor_country = dco[i] if i < len(dco) else None
                    donor_amount = da[i] if i < len(da) else None
                    if donor_name or donor_category or donor_country or donor_amount:
                        report.donations.append(Donation(
                            name=donor_name,
                            category=donor_category,
                            country=donor_country,
                            amount=parse_float(donor_amount),
                        ))

            # Grants
            if section_posted('grant_name[]', 'grant_registration_no[]', 'grant_country[]', 'grant_amount[]'):
                report.grants.clear()
                gname = request.form.getlist('grant_name[]')
                greg = request.form.getlist('grant_registration_no[]')
                gcountry = request.form.getlist('grant_country[]')
                gamt = request.form.getlist('grant_amount[]')
                grant_rows = max(len(gname), len(greg), len(gcountry), len(gamt))
                for i in range(grant_rows):
                    grant_name = gname[i] if i < len(gname) else None
                    grant_reg = greg[i] if i < len(greg) else None
                    grant_country = gcountry[i] if i < len(gcountry) else None
                    grant_amount = gamt[i] if i < len(gamt) else None
                    if grant_name or grant_reg or grant_country or grant_amount:
                        report.grants.append(Grant(
                            name=grant_name,
                            registration_no=grant_reg,
                            country=grant_country,
                            amount=parse_float(grant_amount),
                        ))

            # Payments (support both field-name variants used across templates)
            if section_posted(
                'payment_description[]',
                'payment_kenya[]',
                'payment_other[]',
                'payment_kenya_amount[]',
                'payment_other_amount[]',
            ):
                report.payments.clear()
                payment_descs = get_form_list_with_other_raw(
                    'payment_description[]',
                    'payment_description_other[]',
                    trigger_value='OTHER',
                )
                payment_kenya = get_form_list_any('payment_kenya_amount[]', 'payment_kenya[]')
                payment_other = get_form_list_any('payment_other_amount[]', 'payment_other[]')
                payment_rows = max(len(payment_descs), len(payment_kenya), len(payment_other))
                for i in range(payment_rows):
                    description = payment_descs[i] if i < len(payment_descs) else None
                    kenya_amount = payment_kenya[i] if i < len(payment_kenya) else None
                    other_amount = payment_other[i] if i < len(payment_other) else None
                    if description or kenya_amount or other_amount:
                        report.payments.append(Payment(
                            description=description,
                            kenya_amount=parse_float(kenya_amount),
                            other_amount=parse_float(other_amount),
                        ))

            # Banking
            if section_posted('bank_name[]', 'bank_branch[]', 'bank_branch_select[]', 'bank_account_number[]', 'bank_currency[]'):
                report.bank_accounts.clear()
                bank_names = get_form_list_with_other('bank_name[]', 'bank_name_other[]')
                bank_branches = merge_form_list_upper('bank_branch[]', 'bank_branch_select[]')
                bank_accounts = [value.strip() for value in request.form.getlist('bank_account_number[]')]
                bank_currencies = [value.strip().upper() for value in request.form.getlist('bank_currency[]')]
                bank_rows = max(len(bank_names), len(bank_branches), len(bank_accounts), len(bank_currencies))
                for i in range(bank_rows):
                    bank_name = bank_names[i] if i < len(bank_names) else ''
                    bank_branch = bank_branches[i] if i < len(bank_branches) else ''
                    bank_account = bank_accounts[i] if i < len(bank_accounts) else ''
                    bank_currency = bank_currencies[i] if i < len(bank_currencies) else ''
                    if bank_name or bank_branch or bank_account or bank_currency:
                        bank_account = bank_account or "00000000"
                        bank_currency = bank_currency or "KES"
                        report.bank_accounts.append(BankAccount(
                            bank_name=bank_name,
                            branch=bank_branch,
                            account_number=bank_account,
                            currency=bank_currency,
                        ))

            # Auditors
            if section_posted('audit_firm[]', 'auditor_name[]', 'auditor_practicing_no[]'):
                report.auditors.clear()
                audit_firms = request.form.getlist('audit_firm[]')
                auditor_names = request.form.getlist('auditor_name[]')
                auditor_practicing = request.form.getlist('auditor_practicing_no[]')
                for i in range(min(len(audit_firms), len(auditor_names), len(auditor_practicing))):
                    if audit_firms[i] or auditor_names[i] or auditor_practicing[i]:
                        report.auditors.append(AuditorEntry(
                            firm=audit_firms[i],
                            auditor_name=auditor_names[i],
                            practicing_no=auditor_practicing[i],
                        ))

            # Section C - Staff biodata
            if section_posted('biodata_item[]', 'prev-year5[]', 'curr-year5[]'):
                report.staff_biodata.clear()
                staff_categories = get_form_list_upper('biodata_item[]')
                staff_prev = request.form.getlist('prev-year5[]')
                staff_curr = request.form.getlist('curr-year5[]')
                for i in range(0, min(len(staff_categories), len(staff_prev), len(staff_curr))):
                    category = normalize_section_c_biodata_category(staff_categories[i])
                    raw_prev = staff_prev[i]
                    raw_curr = staff_curr[i]
                    if not category or is_section_c_summary_category(category):
                        continue
                    if is_blank_form_value(raw_prev) and is_blank_form_value(raw_curr):
                        continue
                    report.staff_biodata.append(StaffBiodata(
                        category=category,
                        prev_year=parse_optional_int(raw_prev),
                        curr_year=parse_optional_int(raw_curr),
                    ))

            # Section C - Volunteer biodata
            if section_posted('volbiodata_item[]', 'prev-volntr[]', 'curr-volntr[]'):
                report.volunteer_biodata.clear()
                vol_categories = get_form_list_upper('volbiodata_item[]')
                vol_prev = request.form.getlist('prev-volntr[]')
                vol_curr = request.form.getlist('curr-volntr[]')
                for i in range(0, min(len(vol_categories), len(vol_prev), len(vol_curr))):
                    category = normalize_section_c_biodata_category(vol_categories[i])
                    raw_prev = vol_prev[i]
                    raw_curr = vol_curr[i]
                    if not category or is_section_c_summary_category(category):
                        continue
                    if is_blank_form_value(raw_prev) and is_blank_form_value(raw_curr):
                        continue
                    report.volunteer_biodata.append(VolunteerBiodata(
                        category=category,
                        prev_year=parse_optional_int(raw_prev),
                        curr_year=parse_optional_int(raw_curr),
                    ))

            # Section C - Volunteer privileges
            if section_posted(
                'vol_priv_category[]',
                'vol_priv_kenyan_volunteer[]',
                'vol_priv_kenyan_intern[]',
                'vol_priv_international_volunteer[]',
                'vol_priv_international_intern[]',
            ):
                report.volunteer_privileges.clear()
                priv_categories = get_form_list_upper('vol_priv_category[]')
                priv_kenyan_vol = set(get_form_list_upper('vol_priv_kenyan_volunteer[]'))
                priv_kenyan_intern = set(get_form_list_upper('vol_priv_kenyan_intern[]'))
                priv_int_vol = set(get_form_list_upper('vol_priv_international_volunteer[]'))
                priv_int_intern = set(get_form_list_upper('vol_priv_international_intern[]'))
                for category in priv_categories:
                    if category:
                        report.volunteer_privileges.append(VolunteerPrivilege(
                            category=category,
                            kenyan_volunteer=category in priv_kenyan_vol,
                            kenyan_intern=category in priv_kenyan_intern,
                            international_volunteer=category in priv_int_vol,
                            international_intern=category in priv_int_intern,
                        ))

            # Section C - Staff training
            if section_posted('training_type[]', 'training_kenyan[]', 'training_foreign[]'):
                report.training_records.clear()
                training_types = get_form_list_upper('training_type[]')
                training_kenyan = request.form.getlist('training_kenyan[]')
                training_foreign = request.form.getlist('training_foreign[]')
                for i in range(0, min(len(training_types), len(training_kenyan), len(training_foreign))):
                    if training_types[i] or training_kenyan[i] or training_foreign[i]:
                        report.training_records.append(TrainingRecord(
                            training_type=training_types[i],
                            kenyan_count=parse_int(training_kenyan[i]),
                            international_count=parse_int(training_foreign[i]),
                        ))

            # Tax waiver items
            if (not government_sections_inactive) and section_posted(
                'items_exemption_description[]',
                'items_exemption_quantity[]',
                'items_exemption_type[]',
                'items_exemption_amount[]',
                'items_exemption_certificate[]',
            ):
                report.tax_waiver_items.clear()
                tax_descriptions = get_form_list_upper('items_exemption_description[]')
                tax_quantities = request.form.getlist('items_exemption_quantity[]')
                tax_types = get_form_list_upper('items_exemption_type[]')
                tax_amounts = request.form.getlist('items_exemption_amount[]')
                tax_certificates = get_form_list_upper('items_exemption_certificate[]')
                tax_rows = max(
                    len(tax_descriptions),
                    len(tax_quantities),
                    len(tax_types),
                    len(tax_amounts),
                    len(tax_certificates),
                )
                for i in range(tax_rows):
                    description = tax_descriptions[i] if i < len(tax_descriptions) else ""
                    quantity = tax_quantities[i] if i < len(tax_quantities) else ""
                    exemption_type = tax_types[i] if i < len(tax_types) else ""
                    amount = tax_amounts[i] if i < len(tax_amounts) else ""
                    certificate = tax_certificates[i] if i < len(tax_certificates) else ""
                    if any([description, quantity, exemption_type, amount, certificate]):
                        report.tax_waiver_items.append(TaxWaiverItem(
                            item_description=description,
                            quantity=parse_int(quantity),
                            exemption_type=exemption_type,
                            estimated_tax_waived=parse_float(amount),
                            certificate_approval_no=certificate,
                        ))

            # Section D - Project implementation details
            if section_posted(
                'project_sector[]',
                'project_county[]',
                'project_vulnerable_group[]',
                'project_beneficiaries_no[]',
                'project_spending_per_county[]',
                'project_duration_years[]',
                'project_completion_status[]',
                'project_amount_spent_kenya[]',
                'project_amount_spent_other[]',
            ):
                report.project_implementations.clear()
                project_sectors = get_form_list_upper('project_sector[]')
                project_counties = get_form_list_upper('project_county[]')
                project_vulnerable_groups = get_form_list_upper('project_vulnerable_group[]')
                project_beneficiaries = request.form.getlist('project_beneficiaries_no[]')
                project_spending = request.form.getlist('project_spending_per_county[]')
                project_duration = request.form.getlist('project_duration_years[]')
                project_completion = get_form_list_upper('project_completion_status[]')
                project_spent_kenya = request.form.getlist('project_amount_spent_kenya[]')
                project_spent_other = request.form.getlist('project_amount_spent_other[]')
                project_rows = max(
                    len(project_sectors),
                    len(project_counties),
                    len(project_vulnerable_groups),
                    len(project_beneficiaries),
                    len(project_spending),
                    len(project_duration),
                    len(project_completion),
                    len(project_spent_kenya),
                    len(project_spent_other),
                )
                for i in range(0, project_rows):
                    sector = project_sectors[i] if i < len(project_sectors) else None
                    county = project_counties[i] if i < len(project_counties) else None
                    vulnerable_group = project_vulnerable_groups[i] if i < len(project_vulnerable_groups) else None
                    beneficiaries = project_beneficiaries[i] if i < len(project_beneficiaries) else ""
                    spending = project_spending[i] if i < len(project_spending) else ""
                    duration = project_duration[i] if i < len(project_duration) else ""
                    completion = project_completion[i] if i < len(project_completion) else None
                    spent_kenya = project_spent_kenya[i] if i < len(project_spent_kenya) else ""
                    spent_other = project_spent_other[i] if i < len(project_spent_other) else ""
                    if any([
                        sector,
                        county,
                        vulnerable_group,
                        beneficiaries,
                        spending,
                        duration,
                        completion,
                        spent_kenya,
                        spent_other,
                    ]):
                        report.project_implementations.append(ProjectImplementation(
                            sector=sector,
                            county=county,
                            vulnerable_group=vulnerable_group,
                            beneficiaries_no=parse_optional_int(beneficiaries),
                            spending_per_county=parse_float(spending),
                            duration_years=parse_float(duration),
                            completion_status=completion,
                            amount_spent_kenya=parse_float(spent_kenya),
                            amount_spent_other=parse_float(spent_other),
                        ))

            # Section D1 - Projects carried out
            if section_posted(
                'projects_sector[]',
                'projects_carried_forward_kenya[]',
                'projects_carried_forward_other[]',
                'projects_started_kenya[]',
                'projects_started_other[]',
                'projects_completed_kenya[]',
                'projects_completed_other[]',
            ):
                report.projects_carried_out.clear()
                carried_sectors = get_form_list_upper('projects_sector[]')
                carried_forward_kenya = get_form_list_upper('projects_carried_forward_kenya[]')
                carried_forward_other = get_form_list_upper('projects_carried_forward_other[]')
                carried_started_kenya = get_form_list_upper('projects_started_kenya[]')
                carried_started_other = get_form_list_upper('projects_started_other[]')
                carried_completed_kenya = get_form_list_upper('projects_completed_kenya[]')
                carried_completed_other = get_form_list_upper('projects_completed_other[]')
                carried_rows = min(
                    len(carried_sectors),
                    len(carried_forward_kenya),
                    len(carried_forward_other),
                    len(carried_started_kenya),
                    len(carried_started_other),
                    len(carried_completed_kenya),
                    len(carried_completed_other),
                )
                for i in range(0, carried_rows):
                    if any([
                        carried_sectors[i],
                        carried_forward_kenya[i],
                        carried_forward_other[i],
                        carried_started_kenya[i],
                        carried_started_other[i],
                        carried_completed_kenya[i],
                        carried_completed_other[i],
                    ]):
                        report.projects_carried_out.append(ProjectCarriedOut(
                            sector=carried_sectors[i],
                            carried_forward_kenya=carried_forward_kenya[i] or None,
                            carried_forward_other=carried_forward_other[i] or None,
                            started_kenya=carried_started_kenya[i] or None,
                            started_other=carried_started_other[i] or None,
                            completed_kenya=carried_completed_kenya[i] or None,
                            completed_other=carried_completed_other[i] or None,
                        ))

            # Section D3 - Collaboration & Networking
            if section_posted(
                'collab_partner_type[]',
                'collab_info_exchange[]',
                'collab_tech_support_to[]',
                'collab_tech_support_from[]',
                'collab_funding_to[]',
                'collab_funding_from[]',
                'collab_equipment_to[]',
                'collab_equipment_from[]',
            ):
                report.collaborations.clear()
                collab_partner_types = get_form_list_upper('collab_partner_type[]')
                collab_info_exchange = get_form_list_upper('collab_info_exchange[]')
                collab_tech_support_to = get_form_list_upper('collab_tech_support_to[]')
                collab_tech_support_from = get_form_list_upper('collab_tech_support_from[]')
                collab_funding_to = get_form_list_upper('collab_funding_to[]')
                collab_funding_from = get_form_list_upper('collab_funding_from[]')
                collab_equipment_to = get_form_list_upper('collab_equipment_to[]')
                collab_equipment_from = get_form_list_upper('collab_equipment_from[]')
                collab_rows = min(
                    len(collab_partner_types),
                    len(collab_info_exchange),
                    len(collab_tech_support_to),
                    len(collab_tech_support_from),
                    len(collab_funding_to),
                    len(collab_funding_from),
                    len(collab_equipment_to),
                    len(collab_equipment_from),
                )
                for i in range(0, collab_rows):
                    if any([
                        collab_partner_types[i],
                        collab_info_exchange[i],
                        collab_tech_support_to[i],
                        collab_tech_support_from[i],
                        collab_funding_to[i],
                        collab_funding_from[i],
                        collab_equipment_to[i],
                        collab_equipment_from[i],
                    ]):
                        report.collaborations.append(CollaborationNetworking(
                            partner_type=collab_partner_types[i],
                            info_exchange=collab_info_exchange[i] or None,
                            tech_support_to_partner=collab_tech_support_to[i] or None,
                            tech_support_from_partner=collab_tech_support_from[i] or None,
                            funding_to_partner=collab_funding_to[i] or None,
                            funding_from_partner=collab_funding_from[i] or None,
                            equipment_to_partner=collab_equipment_to[i] or None,
                            equipment_from_partner=collab_equipment_from[i] or None,
                        ))

            # Section E - Officials
            if (not get_inactive_section_flag("officials_section_inactive")) and section_posted(
                'official_role[]',
                'official_name[]',
                'official_nationality[]',
                'official_gender[]',
                'official_email[]',
                'official_residence[]',
                'official_phone[]',
                'official_kra_pin[]',
                'official_professional_qualification[]',
                'official_signature[]',
            ):
                report.officials.clear()
                official_roles = get_form_list_upper('official_role[]')
                official_names = get_form_list_upper('official_name[]')
                official_nationalities = get_form_list_upper('official_nationality[]')
                official_genders = get_form_list_upper('official_gender[]')
                official_emails = get_form_list_upper('official_email[]')
                official_residences = get_form_list_upper('official_residence[]')
                official_phones = get_form_list_phone('official_phone[]')
                official_kra_pins = get_form_list_upper('official_kra_pin[]')
                official_qualifications = get_form_list_upper('official_professional_qualification[]')
                official_signatures = request.form.getlist('official_signature[]')
                official_rows = min(
                    len(official_roles),
                    len(official_names),
                    len(official_nationalities),
                    len(official_genders),
                    len(official_emails),
                    len(official_residences),
                    len(official_phones),
                )
                for i in range(0, official_rows):
                    if any([
                        official_roles[i],
                        official_names[i],
                        official_nationalities[i],
                        official_genders[i],
                        official_emails[i],
                        official_residences[i],
                        official_phones[i],
                    ]):
                        report.officials.append(Official(
                            role=official_roles[i] or "OFFICIAL",
                            name=official_names[i],
                            nationality=official_nationalities[i],
                            gender=official_genders[i],
                            email=official_emails[i],
                            residence=official_residences[i],
                            phone=official_phones[i],
                            kra_pin=official_kra_pins[i] if i < len(official_kra_pins) else None,
                            professional_qualification=official_qualifications[i] if i < len(official_qualifications) else None,
                            signature=official_signatures[i] if i < len(official_signatures) else None,
                        ))

            report.number_of_directors = parse_int(request.form.get('number_of_directors'))
            report.number_of_registered_members = parse_int(request.form.get('number_of_registered_members'))
            report.number_of_board_meetings = parse_int(request.form.get('number_of_board_meetings'))

            ef = get_form_list_upper_any('election_frequency[]', 'election_frequency')
            report.election_frequency = ", ".join(ef) if ef else None
            report.election_frequency_other = get_form_upper('election_frequency_other') if 'OTHER' in ef else None
            report.date_last_agm = parse_date(request.form.get('date_last_agm'))
            report.date_last_election = parse_date(request.form.get('date_last_election'))
            report.date_last_board_meeting = parse_date(request.form.get('date_last_board_meeting'))

            report.membership_number_of_directors = parse_optional_int(request.form.get("membership_number_of_directors"))
            report.membership_number_of_registered_members = parse_optional_int(request.form.get("membership_number_of_registered_members"))
            report.membership_number_of_board_meetings = parse_optional_int(request.form.get("membership_number_of_board_meetings"))
            report.membership_date_last_agm = parse_date(request.form.get("membership_date_last_agm"))
            report.membership_date_last_election = parse_date(request.form.get("membership_date_last_election"))
            report.non_membership_number_of_directors = parse_optional_int(request.form.get("non_membership_number_of_directors"))
            report.non_membership_number_of_board_meetings = parse_optional_int(request.form.get("non_membership_number_of_board_meetings"))
            report.non_membership_date_last_board_meeting = parse_date(request.form.get("non_membership_date_last_board_meeting"))
            report.non_membership_date_last_election = parse_date(request.form.get("non_membership_date_last_election"))
            if get_inactive_section_flag("officials_section_inactive"):
                clear_disabled_officials_section(report)
            else:
                report.submitter_fullname = get_form_upper("submitter_fullname")
                report.signature = request.form.get("signature")
                report.submission_date = parse_date(request.form.get("submission_date"))

            # Server-side totals (authoritative)
            apply_backend_section_b_math(report)
            sync_report_section_c_auxiliary_tables(report)

            if current_user.is_authenticated:
                report.last_modified_by_id = current_user.id
            report.duplicate_flag = find_duplicate_pbo_report(
                report.pbo_name,
                reporting_period_start=report.reporting_period_start,
                exclude_report_id=report.id,
                require_start_month_match=True,
            ) is not None
            if report.workflow_status in {'returned', 'draft'}:
                report.workflow_status = 'submitted'
                report.review_status = 'pending'
            record_report_field_changes(report, before, 'report_updated', current_user if current_user.is_authenticated else None)
            log_user_activity('report_updated', report=report)
            db.session.commit()
            flash('Report updated successfully.', 'success')
            return redirect(url_for('report_detail', report_id=report.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating report: {str(e)}', 'error')
            return redirect(url_for('report_edit', report_id=report.id))

    return_filing_date_display = return_date_display(report.return_date)
    context = {
        'report': report,
        'associated_username': report_associated_username_display(report),
        'reporting_period_start_display': reporting_period_display(report.reporting_period_start, report.reporting_period_start_raw),
        'reporting_period_end_display': reporting_period_display(report.reporting_period_end, report.reporting_period_end_raw),
        'return_date_display': return_filing_date_display,
        'return_filing_date_display': return_filing_date_display,
        'countries_of_operation_list': [c.strip() for c in (report.countries_of_operation or "").split(",") if c.strip()],
        'counties_list': [c.strip() for c in (report.counties or "").split(",") if c.strip()],
        'assets_list': report.assets,
        'donations_list': report.donations,
        'officials_list': [{'name': o.name, 'position': o.role} for o in report.officials],
        'election_frequency_list': [e.strip() for e in (report.election_frequency or "").split(",") if e.strip()],
        'projects_list': [],
    }
    def fmt_date(val):
        return val.isoformat() if val else ""

    staff_biodata_rows = get_display_staff_biodata_rows(report)
    volunteer_biodata_rows = get_display_volunteer_biodata_rows(report)

    prefill = {
        'reporting_period_start': report.reporting_period_start.isoformat() if report.reporting_period_start else "",
        'reporting_period_end': report.reporting_period_end.isoformat() if report.reporting_period_end else "",
        'return_date': return_date_input_value(report.return_date),
        'pbo_name': report.pbo_name or "",
        'pbo_registration_number': report.pbo_registration_number or "",
        'pbo_registration_date': report.pbo_registration_date.isoformat() if report.pbo_registration_date else "",
        'kra_pin': report.kra_pin or "",
        'postal_address': report.postal_address or "",
        'physical_address': report.physical_address or "",
        'telephone': report.telephone or "",
        'cell_phone': report.cell_phone or "",
        'email': report.email or "",
        'website': report.website or "",
        'social_media': report.social_media or "",
        'contact_name': report.contact_name or "",
        'contact_position': report.contact_position or "",
        'contact_telephone': report.contact_telephone or "",
        'contact_email': report.contact_email or "",
        'contact_nationality': report.contact_nationality or "",
        'contact_gender': report.contact_gender or "",
        'registration_number': report.registration_number or "",
        'pin_number': report.pin_number or "",
        'date_of_registration': report.date_of_registration.isoformat() if report.date_of_registration else "",
        'scope': report.scope or "",
        'countries_of_operation': [c.strip() for c in (report.countries_of_operation or "").split(",") if c.strip()],
        'counties': [c.strip() for c in (report.counties or "").split(",") if c.strip()],
        'audited': report.audited or "",
        'assets_stolen': report.assets_stolen or "",
        'cash_balance_previous_year': report.cash_balance_previous_year,
        'income_b2_total': report.income_b2_total,
        'receipts_total': report.receipts_total,
        'cash_bank_balance': report.cash_bank_balance,
        'staff_kenyan': [
            report.staff_kenyan_prev,
            report.staff_kenyan_current,
            report.staff_kenyan_came_in,
            report.staff_kenyan_left,
        ],
        'staff_foreign': [
            report.staff_foreign_prev,
            report.staff_foreign_current,
            report.staff_foreign_came_in,
            report.staff_foreign_left,
        ],
        'staff_other_kenyan': [report.staff_other_kenyan_prev, report.staff_other_kenyan_current],
        'staff_other_foreign': [report.staff_other_foreign_prev, report.staff_other_foreign_current],
        'volunteers_kenyan': [report.volunteers_kenyan_prev, report.volunteers_kenyan_current],
        'volunteers_foreign': [report.volunteers_foreign_prev, report.volunteers_foreign_current],
        'implementation': [i.strip() for i in (report.project_implementation_method or "").split(",") if i.strip()],
        'local_material': bool(report.local_material),
        'local_material_amount': report.local_material_amount,
        'local_labour': bool(report.local_labour),
        'local_labour_amount': report.local_labour_amount,
        'local_financial': bool(report.local_financial),
        'local_financial_amount': report.local_financial_amount,
        'local_other': bool(report.local_other),
        'local_other_specify': report.local_other_specify or "",
        'local_other_amount': report.local_other_amount,
        'gov_tax_waiver': bool(report.gov_tax_waiver),
        'gov_tax_waiver_amount': report.gov_tax_waiver_amount,
        'gov_other': bool(report.gov_other),
        'gov_other_specify': report.gov_other_specify or "",
        'gov_other_amount': report.gov_other_amount,
        'election_frequency': [e.strip() for e in (report.election_frequency or "").split(",") if e.strip()],
        'election_frequency_value': next(
            (
                e.strip().upper()
                for e in (report.election_frequency or "").split(",")
                if e.strip()
            ),
            "",
        ),
        'election_frequency_other': report.election_frequency_other or "",
        'membership_number_of_directors': (
            report.membership_number_of_directors
            if report.membership_number_of_directors is not None
            else report.number_of_directors
        ),
        'membership_number_of_registered_members': (
            report.membership_number_of_registered_members
            if report.membership_number_of_registered_members is not None
            else report.number_of_registered_members
        ),
        'membership_number_of_board_meetings': (
            report.membership_number_of_board_meetings
            if report.membership_number_of_board_meetings is not None
            else report.number_of_board_meetings
        ),
        'membership_date_last_agm': fmt_date(report.membership_date_last_agm or report.date_last_agm),
        'membership_date_last_election': fmt_date(report.membership_date_last_election or report.date_last_election),
        'non_membership_number_of_directors': (
            report.non_membership_number_of_directors
            if report.non_membership_number_of_directors is not None
            else report.number_of_directors
        ),
        'non_membership_number_of_board_meetings': (
            report.non_membership_number_of_board_meetings
            if report.non_membership_number_of_board_meetings is not None
            else report.number_of_board_meetings
        ),
        'non_membership_date_last_board_meeting': fmt_date(
            report.non_membership_date_last_board_meeting or report.date_last_board_meeting
        ),
        'non_membership_date_last_election': fmt_date(
            report.non_membership_date_last_election or report.date_last_election
        ),
        'submitter_fullname': report.submitter_fullname or "",
        'signature': report.signature or "",
        'submission_date': fmt_date(report.submission_date),
        'assets': [{'item': a.item, 'number': a.number, 'value': a.value} for a in report.assets],
        'donations': [{'name': d.name, 'category': d.category, 'country': d.country, 'amount': d.amount} for d in report.donations],
        'grants': [{'name': g.name, 'registration_no': g.registration_no, 'country': g.country, 'amount': g.amount} for g in report.grants],
        'payments': [{'description': p.description, 'kenya_amount': p.kenya_amount, 'other_amount': p.other_amount} for p in report.payments],
        'bank_accounts': [{'bank_name': b.bank_name, 'branch': b.branch, 'account_number': b.account_number, 'currency': b.currency} for b in report.bank_accounts],
        'auditors': [{'firm': a.firm, 'auditor_name': a.auditor_name, 'practicing_no': a.practicing_no} for a in report.auditors],
        'staff_biodata': staff_biodata_rows,
        'volunteer_biodata': volunteer_biodata_rows,
        'volunteer_privileges': [
            {
                'category': v.category,
                'kenyan_volunteer': v.kenyan_volunteer,
                'kenyan_intern': v.kenyan_intern,
                'international_volunteer': v.international_volunteer,
                'international_intern': v.international_intern,
            } for v in report.volunteer_privileges
        ],
        'training_records': [{'training_type': t.training_type, 'kenyan_count': t.kenyan_count, 'international_count': t.international_count} for t in report.training_records],
        'tax_waiver_items': [
            {
                'item_description': t.item_description,
                'quantity': t.quantity,
                'exemption_type': t.exemption_type,
                'estimated_tax_waived': t.estimated_tax_waived,
                'certificate_approval_no': t.certificate_approval_no,
            } for t in report.tax_waiver_items
        ],
        'project_implementations': [
            {
                'sector': p.sector or "",
                'county': p.county or "",
                'vulnerable_group': p.vulnerable_group or "",
                'beneficiaries_no': legacy_zero_output_value(report, p.beneficiaries_no),
                'spending_per_county': p.spending_per_county,
                'duration_years': p.duration_years,
                'completion_status': p.completion_status or "",
                'amount_spent_kenya': p.amount_spent_kenya,
                'amount_spent_other': p.amount_spent_other,
            } for p in report.project_implementations
        ],
        'projects_carried_out': [
            {
                'sector': p.sector,
                'carried_forward_kenya': p.carried_forward_kenya,
                'carried_forward_other': p.carried_forward_other,
                'started_kenya': p.started_kenya,
                'started_other': p.started_other,
                'completed_kenya': p.completed_kenya,
                'completed_other': p.completed_other,
            } for p in report.projects_carried_out
        ],
        'collaboration': [
            {
                'partner_type': c.partner_type,
                'info_exchange': c.info_exchange,
                'tech_support_to_partner': c.tech_support_to_partner,
                'tech_support_from_partner': c.tech_support_from_partner,
                'funding_to_partner': c.funding_to_partner,
                'funding_from_partner': c.funding_from_partner,
                'equipment_to_partner': c.equipment_to_partner,
                'equipment_from_partner': c.equipment_from_partner,
            } for c in report.collaborations
        ],
        'officials': [
            {
                'role': o.role,
                'name': o.name,
                'nationality': o.nationality,
                'gender': o.gender,
                'email': o.email,
                'residence': o.residence,
                'phone': o.phone,
                'kra_pin': o.kra_pin,
                'professional_qualification': o.professional_qualification,
                'signature': o.signature,
            } for o in report.officials
        ],
    }
    prefill = legacy_zero_output_structure(report, prefill)
    field_help_context = build_report_edit_field_help_context()
    persist_field_help_context('report_edit', field_help_context)
    ensure_field_help_decision_model('report_edit', field_help_context)
    ensure_field_help_intent_samples('report_edit')
    ensure_field_help_intent_model('report_edit')
    context['prefill'] = prefill
    return render_template('report_edit.html', **context)


@app.route('/report/<int:report_id>/field-help', methods=['POST'])
@login_required
def report_field_help(report_id):
    report = db.session.get(PBOReport, report_id) or abort(404)
    denied = require_report_access(report, write=False)
    if denied:
        return jsonify({'error': 'You do not have permission to access that report.'}), 403

    payload = request.get_json(silent=True) or {}
    question = (payload.get('question') or '').strip()
    if not question:
        return jsonify({'error': 'Ask a question about enabled or disabled fields.'}), 400

    field_help_context = build_report_edit_field_help_context()
    answer, source = answer_field_help_question(
        question=question,
        field_help_context=field_help_context,
        page_key='report_edit',
        prefer_openai=False,
    )

    return jsonify({
        'answer': answer,
        'required_fields': field_help_context['enabled_summary'],
        'disabled_fields': field_help_context['disallowed_summary'],
        'allowed_fields': field_help_context['allowed_summary'],
        'disallowed_fields': field_help_context['disallowed_summary'],
        'headings': field_help_context.get('heading_summary', []),
        'field_heading_relations': [
            {'field': rule.get('label'), 'heading': rule.get('heading')}
            for rule in field_help_context.get('allowed_rules', [])
            if rule.get('heading')
        ],
        'source': source,
    })


@app.route('/field-help/context', methods=['GET'])
def global_field_help_context():
    page_key = normalize_field_help_page_key(request.args.get('page_key') or request.endpoint or 'global')
    field_help_context = get_field_help_context_for_page(page_key)
    return jsonify({
        'required_fields': field_help_context.get('enabled_summary', []),
        'disabled_fields': field_help_context.get('disallowed_summary', []),
        'allowed_fields': field_help_context.get('allowed_summary', []),
        'disallowed_fields': field_help_context.get('disallowed_summary', []),
        'sector_fields': field_help_context.get('sector_focus_summary', []),
        'headings': field_help_context.get('heading_summary', []),
        'field_heading_relations': [
            {'field': rule.get('label'), 'heading': rule.get('heading')}
            for rule in field_help_context.get('allowed_rules', [])
            if rule.get('heading')
        ],
        'notes': (field_help_context.get('metadata') or {}).get('notes', []),
    })


@app.route('/field-help/chat', methods=['POST'])
def global_field_help_chat():
    payload = request.get_json(silent=True) or {}
    question = (payload.get('question') or '').strip()
    page_key = normalize_field_help_page_key(payload.get('page_key') or request.endpoint or 'global')
    if not question:
        return jsonify({'error': 'Ask a short question so I can help.'}), 400

    field_help_context = get_field_help_context_for_page(page_key)
    answer, source = answer_field_help_question(
        question=question,
        field_help_context=field_help_context,
        page_key=page_key,
        prefer_openai=False,
    )
    return jsonify({
        'answer': answer,
        'source': source,
        'required_fields': field_help_context.get('enabled_summary', []),
        'disabled_fields': field_help_context.get('disallowed_summary', []),
        'allowed_fields': field_help_context.get('allowed_summary', []),
        'disallowed_fields': field_help_context.get('disallowed_summary', []),
        'sector_fields': field_help_context.get('sector_focus_summary', []),
        'headings': field_help_context.get('heading_summary', []),
        'field_heading_relations': [
            {'field': rule.get('label'), 'heading': rule.get('heading')}
            for rule in field_help_context.get('allowed_rules', [])
            if rule.get('heading')
        ],
    })


@app.route('/field-help/ai', methods=['POST'])
def field_help_ai():
    return global_field_help_chat()


def all_files_datetime_rank(value):
    if value is None:
        return float('-inf')
    if isinstance(value, datetime):
        normalized = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
        return normalized.timestamp()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), tzinfo=timezone.utc).timestamp()
    return float('-inf')


def all_files_user_label(user, user_id=None):
    label = user_display_name(user)
    if label:
        return label
    normalized_user_id = coerce_legacy_int(user_id)
    if normalized_user_id is not None:
        return f'User #{normalized_user_id}'
    return 'Unassigned'


def all_files_register_worker(workers, user=None, user_id=None, role='Contributor', touched_at=None):
    normalized_user_id = coerce_legacy_int(user_id if user_id is not None else getattr(user, 'id', None))
    if normalized_user_id is None:
        return None

    entry = workers.setdefault(normalized_user_id, {
        'user_id': normalized_user_id,
        'user': user,
        'roles': [],
        'last_touched_at': None,
    })
    if entry.get('user') is None and user is not None:
        entry['user'] = user
    if role and role not in entry['roles']:
        entry['roles'].append(role)
    if all_files_datetime_rank(touched_at) > all_files_datetime_rank(entry.get('last_touched_at')):
        entry['last_touched_at'] = touched_at
    return entry


def all_files_register_event(events, user=None, user_id=None, role='Contributor', label='Activity', touched_at=None, detail=None):
    normalized_user_id = coerce_legacy_int(user_id if user_id is not None else getattr(user, 'id', None))
    events.append({
        'user_id': normalized_user_id,
        'user': user,
        'role': role,
        'label': label,
        'detail': detail,
        'touched_at': touched_at,
    })


def all_files_role_summary(entry):
    return ', '.join(entry.get('roles') or []) or 'Contributor'


def all_files_worker_summary(worker_entries):
    summary_items = []
    for entry in worker_entries[:6]:
        summary_items.append(
            f"{all_files_user_label(entry.get('user'), entry.get('user_id'))} ({all_files_role_summary(entry)})"
        )
    if len(worker_entries) > 6:
        summary_items.append(f"+{len(worker_entries) - 6} more")
    return '; '.join(summary_items) or 'No tracked worker'


def build_all_files_rows(reports):
    rows = []
    for report in reports:
        workers = {}
        events = []

        def register(user=None, user_id=None, role='Contributor', label='Activity', touched_at=None, detail=None):
            all_files_register_worker(workers, user=user, user_id=user_id, role=role, touched_at=touched_at)
            all_files_register_event(
                events,
                user=user,
                user_id=user_id,
                role=role,
                label=label,
                touched_at=touched_at,
                detail=detail,
            )

        register(
            user=getattr(report, 'user', None),
            user_id=getattr(report, 'user_id', None),
            role='Owner',
            label='Record owner',
            touched_at=getattr(report, 'created_at', None),
        )
        register(
            user=getattr(report, 'last_modified_by', None),
            user_id=getattr(report, 'last_modified_by_id', None),
            role='Editor',
            label='Report updated',
            touched_at=getattr(report, 'updated_at', None),
        )

        uploaded_files = sorted(
            report.uploaded_files or [],
            key=lambda item: all_files_datetime_rank(getattr(item, 'created_at', None)),
            reverse=True,
        )
        for uploaded_file in uploaded_files:
            register(
                user=getattr(uploaded_file, 'uploaded_by', None),
                user_id=getattr(uploaded_file, 'uploaded_by_id', None),
                role='Uploader',
                label='File uploaded',
                touched_at=getattr(uploaded_file, 'created_at', None),
                detail=getattr(uploaded_file, 'original_filename', None),
            )

        for change in report.field_changes or []:
            if getattr(change, 'action', None) not in QUALIFYING_REPORT_WORK_FIELD_CHANGE_ACTIONS:
                continue
            register(
                user=getattr(change, 'user', None),
                user_id=getattr(change, 'user_id', None),
                role='Editor',
                label='Field changed',
                touched_at=getattr(change, 'created_at', None),
                detail=getattr(change, 'field_name', None),
            )

        for activity in report.activity_logs or []:
            if getattr(activity, 'action', None) not in QUALIFYING_REPORT_WORK_ACTIVITY_ACTIONS:
                continue
            action_label = str(getattr(activity, 'action', '') or 'Activity').replace('_', ' ').title()
            summary = shorten(str(getattr(activity, 'summary', '') or ''), width=90, placeholder='...')
            register(
                user=getattr(activity, 'user', None),
                user_id=getattr(activity, 'user_id', None),
                role='Contributor',
                label=action_label,
                touched_at=getattr(activity, 'created_at', None),
                detail=summary or None,
            )

        worker_entries = sorted(
            workers.values(),
            key=lambda entry: (
                all_files_datetime_rank(entry.get('last_touched_at')),
                entry.get('user_id') or 0,
            ),
            reverse=True,
        )
        latest_event = max(
            events,
            key=lambda item: all_files_datetime_rank(item.get('touched_at')),
            default=None,
        )
        latest_worker_entry = None
        if latest_event and latest_event.get('user_id') in workers:
            latest_worker_entry = workers[latest_event['user_id']]
        elif worker_entries:
            latest_worker_entry = worker_entries[0]

        latest_touched_at = (
            latest_event.get('touched_at') if latest_event else None
        ) or getattr(report, 'updated_at', None) or getattr(report, 'created_at', None)
        latest_activity_parts = []
        if latest_event:
            latest_activity_parts.append(latest_event.get('label') or 'Activity')
            if latest_event.get('detail'):
                latest_activity_parts.append(str(latest_event['detail']))
        latest_activity_summary = ': '.join(latest_activity_parts) if latest_activity_parts else 'No tracked activity'

        owner_name = all_files_user_label(getattr(report, 'user', None), getattr(report, 'user_id', None))
        latest_worker_name = (
            all_files_user_label(latest_worker_entry.get('user'), latest_worker_entry.get('user_id'))
            if latest_worker_entry else 'Unassigned'
        )
        latest_worker_roles = all_files_role_summary(latest_worker_entry or {})

        rows.append({
            'report': report,
            'period_display': (
                f"{reporting_period_display(report.reporting_period_start, report.reporting_period_start_raw)}"
                f" to {reporting_period_display(report.reporting_period_end, report.reporting_period_end_raw)}"
            ),
            'return_filing_date': return_date_display(report.return_date),
            'owner_name': owner_name,
            'worker_count': len(worker_entries),
            'worker_ids': [entry.get('user_id') for entry in worker_entries if entry.get('user_id') is not None],
            'worker_summary': all_files_worker_summary(worker_entries),
            'latest_worker': latest_worker_name,
            'latest_worker_roles': latest_worker_roles,
            'latest_activity_summary': latest_activity_summary,
            'latest_activity_at': format_datetime(latest_touched_at),
            'latest_touched_at_raw': latest_touched_at,
            'file_count': len(uploaded_files),
            'file_names': '; '.join(item.original_filename for item in uploaded_files[:4] if item.original_filename),
            'latest_file_id': uploaded_files[0].id if uploaded_files else None,
            'workflow_status': (report.workflow_status or 'draft').replace('_', ' ').title(),
            'updated_at': format_datetime(report.updated_at),
            'created_at': format_datetime(report.created_at),
        })

    rows.sort(
        key=lambda row: (
            all_files_datetime_rank(row.get('latest_touched_at_raw')),
            getattr(row.get('report'), 'id', 0) or 0,
        ),
        reverse=True,
    )
    return rows


def all_files_row_matches_search(row, search_query):
    term = (search_query or '').strip().lower()
    if not term:
        return True
    report = row.get('report')
    search_values = [
        getattr(report, 'id', None),
        getattr(report, 'pbo_name', None),
        getattr(report, 'pbo_registration_number', None),
        getattr(report, 'contact_name', None),
        getattr(report, 'contact_email', None),
        row.get('period_display'),
        row.get('return_filing_date'),
        row.get('owner_name'),
        row.get('worker_summary'),
        row.get('latest_worker'),
        row.get('latest_activity_summary'),
        row.get('file_names'),
        row.get('workflow_status'),
    ]
    return any(term in str(value).lower() for value in search_values if value not in [None, ''])


@app.route('/my-files')
@login_required
def my_files():
    worked_report_ids = get_user_worked_report_ids(current_user)
    rows = []
    reports = []
    dashboard_cards = []

    if worked_report_ids:
        reports = (
            PBOReport.query
            .options(
                selectinload(PBOReport.uploaded_files),
                selectinload(PBOReport.field_changes),
                selectinload(PBOReport.activity_logs),
                selectinload(PBOReport.last_modified_by),
                selectinload(PBOReport.user),
                selectinload(PBOReport.donations),
                selectinload(PBOReport.payments),
                selectinload(PBOReport.project_implementations),
                selectinload(PBOReport.bank_accounts),
            )
            .filter(PBOReport.id.in_(worked_report_ids))
            .order_by(PBOReport.updated_at.desc(), PBOReport.id.desc())
            .all()
        )

        files_added_by_day = Counter()
        records_touched_by_day = {}

        for report in reports:
            uploaded_by_user = [
                item for item in (report.uploaded_files or [])
                if item.uploaded_by_id == current_user.id
            ]
            uploaded_by_user.sort(key=lambda item: item.created_at or utc_now(), reverse=True)
            changes_by_user = [
                item for item in (report.field_changes or [])
                if item.user_id == current_user.id
            ]
            activity_by_user = [
                item for item in (report.activity_logs or [])
                if item.user_id == current_user.id
            ]

            roles = []
            if report.user_id == current_user.id:
                roles.append('Owner')
            if report.last_modified_by_id == current_user.id or changes_by_user:
                roles.append('Editor')
            if uploaded_by_user:
                roles.append('Uploader')
            if activity_by_user and not roles:
                roles.append('Contributor')

            for item in uploaded_by_user:
                if not item.created_at:
                    continue
                day_key = item.created_at.date()
                files_added_by_day[day_key] += 1
                records_touched_by_day.setdefault(day_key, set()).add(report.id)

            latest_touch_candidates = [
                report.updated_at,
                *(item.created_at for item in uploaded_by_user if item.created_at),
                *(item.created_at for item in changes_by_user if item.created_at),
                *(item.created_at for item in activity_by_user if item.created_at),
            ]
            latest_touch = max((item for item in latest_touch_candidates if item is not None), default=None)

            rows.append({
                'report_id': report.id,
                'pbo_name': report.pbo_name,
                'registration_number': report.pbo_registration_number,
                'scope': report.scope,
                'reporting_period_start': reporting_period_display(report.reporting_period_start, report.reporting_period_start_raw),
                'reporting_period_end': reporting_period_display(report.reporting_period_end, report.reporting_period_end_raw),
                'return_filing_date': return_date_display(report.return_date),
                'contact_name': report.contact_name,
                'contact_telephone': report.contact_telephone,
                'contact_email': report.contact_email,
                'countries_of_operation': report.countries_of_operation,
                'counties': report.counties,
                'audited': report.audited,
                'project_implementation_method': report.project_implementation_method,
                'cash_balance_previous_year': report.cash_balance_previous_year,
                'cash_bank_balance': report.cash_bank_balance,
                'donations_total': sum((item.amount or 0) for item in (report.donations or [])),
                'payments_total': sum(((item.kenya_amount or 0) + (item.other_amount or 0)) for item in (report.payments or [])),
                'project_implementation_count': len(report.project_implementations or []),
                'bank_accounts_count': len(report.bank_accounts or []),
                'workflow_status': report.workflow_status,
                'owner_name': report.user.full_name if report.user and report.user.full_name else (report.user.email if report.user else None),
                'worked_as': ', '.join(dict.fromkeys(roles)) or 'Contributor',
                'file_count': len(uploaded_by_user),
                'file_names': '; '.join(item.original_filename for item in uploaded_by_user if item.original_filename) or None,
                'latest_file_id': uploaded_by_user[0].id if uploaded_by_user else None,
                'latest_file_name': uploaded_by_user[0].original_filename if uploaded_by_user else None,
                'last_touched_at': format_datetime(latest_touch),
                'updated_at': format_datetime(report.updated_at),
                'created_at': format_datetime(report.created_at),
            })

        spark_palette = [
            '#dbeafe',
            '#93c5fd',
            '#bfdbfe',
            '#60a5fa',
            '#a7f3d0',
            '#34d399',
            '#fde68a',
            '#fbbf24',
            '#fbcfe8',
            '#f472b6',
        ]
        today_date = utc_now().date()
        upload_days = sorted(files_added_by_day.keys(), reverse=True)
        series_days = upload_days[:20] if upload_days else [today_date]
        files_added_series = []
        max_files_added = 0
        for target_day in series_days:
            files_added = int(files_added_by_day.get(target_day, 0))
            max_files_added = max(max_files_added, files_added)
            records_touched_value = records_touched_by_day.get(target_day, 0)
            if isinstance(records_touched_value, set):
                records_touched = len(records_touched_value)
            else:
                records_touched = int(records_touched_value or 0)
            files_added_series.append({
                'date_label': target_day.strftime('%d %b %Y'),
                'files_added': files_added,
                'records_touched': records_touched,
                'color': spark_palette[len(files_added_series) % len(spark_palette)],
            })
        for item in files_added_series:
            if max_files_added > 0:
                height_percent = max(12, round((item['files_added'] / max_files_added) * 100))
            else:
                height_percent = 10
            item['height_percent'] = height_percent
            item['tooltip'] = (
                f"{item['date_label']}: {item['files_added']} file"
                f"{'' if item['files_added'] == 1 else 's'} added"
                f" across {item['records_touched']} record"
                f"{'' if item['records_touched'] == 1 else 's'}."
            )

        dashboard_cards = [
            {
                'label': 'Daily Files Added',
                'value': sum(item['files_added'] for item in files_added_series),
                'tooltip': 'Latest upload days from your files, up to 20 dates. Hover a bar to see the exact day and how many records it affected.',
                'accent': 'rose',
                'spark_bars': files_added_series,
                'spark_caption': f"Showing {len(files_added_series)} upload day{'' if len(files_added_series) == 1 else 's'}, newest first",
            },
        ]

    columns = [
        ('report_id', 'Report ID'),
        ('pbo_name', 'PBO Name'),
        ('registration_number', 'Registration Number'),
        ('scope', 'Scope'),
        ('reporting_period_start', 'Reporting Start'),
        ('reporting_period_end', 'Reporting End'),
        ('return_filing_date', 'Return Filing Date'),
        ('contact_name', 'Contact Name'),
        ('contact_telephone', 'Contact Telephone'),
        ('contact_email', 'Contact Email'),
        ('countries_of_operation', 'Countries Of Operation'),
        ('counties', 'Counties'),
        ('audited', 'Audited'),
        ('project_implementation_method', 'Implementation Method'),
        ('cash_balance_previous_year', 'Opening Balance'),
        ('cash_bank_balance', 'Closing Balance'),
        ('donations_total', 'Donations Total'),
        ('payments_total', 'Payments Total'),
        ('project_implementation_count', 'Project Rows'),
        ('bank_accounts_count', 'Bank Accounts'),
        ('workflow_status', 'Workflow Status'),
        ('owner_name', 'Record Owner'),
        ('worked_as', 'Worked As'),
        ('file_count', 'My File Count'),
        ('file_names', 'My File Names'),
        ('latest_file_name', 'Latest File'),
        ('last_touched_at', 'Last Touched'),
        ('updated_at', 'Report Updated At'),
        ('created_at', 'Report Created At'),
    ]

    return render_template(
        'my_files.html',
        rows=rows,
        columns=columns,
        money_columns={
            'cash_balance_previous_year',
            'cash_bank_balance',
            'donations_total',
            'payments_total',
        },
        generated_at=format_datetime(utc_now()),
        total_reports=len(rows),
        dashboard_cards=dashboard_cards,
    )


@app.route('/all-files')
@login_required
def all_files():
    if not can_manage_all_records(current_user):
        abort(403)

    search_query = (request.args.get('q') or '').strip()
    reports = (
        PBOReport.query
        .options(
            selectinload(PBOReport.uploaded_files).selectinload(UploadedFile.uploaded_by),
            selectinload(PBOReport.field_changes).selectinload(FieldChangeLog.user),
            selectinload(PBOReport.activity_logs).selectinload(UserActivityLog.user),
            selectinload(PBOReport.last_modified_by),
            selectinload(PBOReport.user),
        )
        .order_by(PBOReport.updated_at.desc(), PBOReport.id.desc())
        .all()
    )
    rows = build_all_files_rows(reports)
    filtered_rows = [
        row for row in rows
        if all_files_row_matches_search(row, search_query)
    ]
    distinct_worker_ids = {
        user_id
        for row in rows
        for user_id in (row.get('worker_ids') or [])
        if user_id is not None
    }
    total_file_count = sum(row.get('file_count') or 0 for row in rows)

    return render_template(
        'all_files.html',
        rows=filtered_rows,
        search_query=search_query,
        generated_at=format_datetime(utc_now()),
        total_reports=len(rows),
        filtered_reports=len(filtered_rows),
        distinct_worker_count=len(distinct_worker_ids),
        total_file_count=total_file_count,
    )


@app.route('/my-files/file/<int:file_id>')
@login_required
def my_file_open(file_id):
    uploaded_file = (
        UploadedFile.query
        .options(selectinload(UploadedFile.report))
        .filter(UploadedFile.id == file_id)
        .first()
        or abort(404)
    )
    if not can_view_uploaded_file_record(current_user, uploaded_file):
        flash('You do not have permission to view that file.', 'danger')
        return redirect(url_for('my_files'))
    if not uploaded_file.storage_path or not os.path.exists(uploaded_file.storage_path):
        flash('The requested file could not be found on disk.', 'warning')
        return redirect(url_for('my_files'))

    return send_file(
        uploaded_file.storage_path,
        as_attachment=False,
        download_name=uploaded_file.original_filename,
        mimetype=uploaded_file.mime_type or 'application/octet-stream',
    )


DATA_INTERPRETATION_SOURCE_LABELS = {
    'reports': 'Form 14 reports',
    'registrations': 'Registration dates',
    'projects': 'Project implementations',
    'project_summary': 'Projects carried out',
    'donations': 'Donations',
    'grants': 'Grants',
    'igas': 'Income generating activities',
    'payments': 'Payments',
    'counties': 'Counties of operation',
    'countries': 'Countries of operation',
    'collaborations': 'Collaborations',
    'staff': 'Staff and volunteers',
    'governance': 'Governance and officials',
    'assets': 'Assets and balances',
    'crosswalks': 'Lookup crosswalks',
}


def data_interp_split_values(raw_value):
    if raw_value in [None, '']:
        return []
    text = str(raw_value).replace('\n', ',').replace(';', ',')
    return [
        re.sub(r'\s+', ' ', chunk).strip()
        for chunk in text.split(',')
        if re.sub(r'\s+', ' ', chunk).strip()
    ]


def data_interp_register_counter(counter, raw_value):
    if raw_value in [None, '']:
        return
    normalized = re.sub(r'\s+', ' ', str(raw_value).strip()).upper()
    if normalized:
        counter[normalized] += 1


def data_interp_reporting_label(report):
    start_date = getattr(report, 'reporting_period_start', None)
    end_date = getattr(report, 'reporting_period_end', None)
    if start_date and end_date:
        return f"{start_date.year}/{str(end_date.year)[-2:]}"
    if start_date:
        return str(start_date.year)
    if end_date:
        return str(end_date.year)
    return None


def data_interp_status_class(status):
    return {
        'Direct': 'is-direct',
        'Partial': 'is-partial',
    }.get(status, 'is-partial')


def data_interp_qa_status_class(status):
    return {
        'Reconciled': 'is-good',
        'Ready': 'is-good',
        'Needs Review': 'is-warn',
        'Mismatch Found': 'is-bad',
        'No Comparison Yet': 'is-neutral',
    }.get(status, 'is-neutral')


def data_interp_build_qa_summary(qa_checks):
    mismatch_count = sum(1 for check in qa_checks if check.get('status') == 'Mismatch Found')
    review_count = sum(1 for check in qa_checks if check.get('status') in {'Needs Review', 'No Comparison Yet'})
    if mismatch_count:
        status = 'Mismatch Found'
        message = (
            f"{mismatch_count} QA check{'s' if mismatch_count != 1 else ''} found reconciliation problems "
            "in this run. Review the flagged items before publishing."
        )
    elif review_count:
        status = 'Needs Review'
        message = (
            f"No blocking numeric mismatches were found, but {review_count} QA check"
            f"{'s' if review_count != 1 else ''} still need cleanup or richer mappings."
        )
    elif qa_checks:
        status = 'Reconciled'
        message = 'All available QA checks reconciled for this analysis run.'
    else:
        status = 'No Comparison Yet'
        message = 'Run the analysis to generate QA status checks.'
    return {
        'status': status,
        'label': status,
        'message': message,
        'status_class': data_interp_qa_status_class(status),
    }


DATA_INTERP_DONOR_CATEGORY_OPTIONS = [
    'Research / Academic Institution',
    'Agency of Kenya Government',
    'National Government',
    'County Government',
    'United Nations Agency',
    'Individual Donors in Kenya / Foreign',
    'Embassy/High Commission',
    'Foundation/Trust',
    'Headquarter of this PBO',
    "Directors' Contributions",
    'Membership Subscription',
    'Returns from investments (e.g., dividends & interest)',
    'Foreign Government Agency (e.g. SIDA, USAID, NORAD, UK AID, Foreign Embassies / High Commissions)',
    'Non-Profit Organizations (PBOs / FBOs)',
    'Members Subscription',
    'Corporate Donors',
    'Foundations',
    'Affiliate / Parent of the PBO',
    'Religious Institutions',
    'International NGOs',
    'Community-Based Organizations (CBOs)',
    'Faith-Based Organizations (FBOs)',
    'Development Partners',
    'Philanthropic Individuals',
    'Multilateral Agencies',
    'Private Sector Companies',
    'Charitable Trusts',
    'Professional Associations',
    'Educational Institutions',
    'NGOs Self Generated Income (eg Consultancy services, Farming & Business Income)',
    'Returns From Investments(eg dividends and interest)',
    'Other (Specified)',
    'Unspecified',
]

DATA_INTERP_COLLAB_PARTNER_OPTIONS = [
    'PBOs',
    'CBOs',
    'FBOs',
    'Research Institutions',
    'Academic Institutions',
    'Health Institutions',
    'National Government',
    'Government Agencies',
    'County Government',
    'Media',
    'Donor Agencies',
    'Corporate',
    'Other (Specified)',
    'Unspecified',
]

DATA_INTERP_COLLAB_NATURE_FIELDS = [
    ('Information Exchange', 'info_exchange'),
    ('Technical Support to Partner', 'tech_support_to_partner'),
    ('Technical Support from Partner', 'tech_support_from_partner'),
    ('Funding to Partner', 'funding_to_partner'),
    ('Funding from Partner', 'funding_from_partner'),
    ('Equipment to Partner', 'equipment_to_partner'),
    ('Equipment from Partner', 'equipment_from_partner'),
]


def build_data_interpretation_topic_plan():
    return [
        {
            'code': '2.0',
            'title': 'Sector Growth',
            'status': 'Partial',
            'objective': 'Combine growth in registrations, filing activity, and funding into one headline trend for the sector.',
            'use_data': [
                'Count newly registered NGOs from `date_of_registration` / `pbo_registration_date` by fiscal year.',
                'Count filed annual returns from reporting-period and workflow fields.',
                'Aggregate funding from donations, grants, IGA income, and report-level receipt totals by year.',
            ],
            'visuals': ['Dual-axis line chart', 'Growth KPI cards', 'Driver waterfall'],
            'deliverable': 'A yearly growth dashboard combining volume and value.',
            'gap': 'Use the current schema to track registrations, filings, and funding together; lifecycle status fields can be added later if you decide to capture them.',
            'source_keys': ['reports', 'registrations', 'donations', 'grants', 'igas'],
        },
        {
            'code': '2.1',
            'title': 'Changes in the Register',
            'status': 'Partial',
            'objective': 'Track register movement that is already visible in the current schema through registrations, filing activity, and workflow progression.',
            'use_data': [
                'Count organisations with registration dates inside the selected fiscal year.',
                'Track workflow progression using `workflow_status`, `submitted_at`, and `submission_date`.',
                'Compare filing activity and newly registered counts across fiscal years to show visible register movement.',
            ],
            'visuals': ['Register activity bars', 'Annual movement chart', 'Status summary table'],
            'deliverable': 'Register activity summary based on the current filing database.',
            'gap': 'This page measures new registrations and active filing movement from Form 14 records already stored in the app.',
            'source_keys': ['registrations', 'reports'],
        },
        {
            'code': '2.2',
            'title': 'Annual Reports',
            'status': 'Partial',
            'objective': 'Track annual return submissions, audited-account compliance, and filing coverage over time.',
            'use_data': [
                'Count reports by fiscal year using reporting-period dates.',
                'Use `workflow_status`, `submitted_at`, and `submission_date` to define filed reports.',
                'Use `audited` to compute audited-account submission counts and rates.',
            ],
            'visuals': ['Submission trend line', 'Audited vs unaudited stacked bars', 'Compliance KPI tiles'],
            'deliverable': 'Annual filing and compliance trend view.',
            'gap': 'Historical replication quality depends on how many older Form 14 years have already been loaded into the database.',
            'source_keys': ['reports', 'governance'],
        },
        {
            'code': '2.3',
            'title': 'NGOs Funding Trend',
            'status': 'Partial',
            'objective': 'Show how reported funding changes year by year across the filing population.',
            'use_data': [
                'Aggregate donation amounts by fiscal year, organisation, and each donor category captured in the donor dropdown, including specified Other entries.',
                'Use the donor-table `category` field to compare how each donation source changes across reporting years.',
                'Use report-level `receipts_total` / `income_b2_total` as a validation check against row-level funding.',
            ],
            'visuals': ['Funding trend line', 'Year-on-year variance bars', 'Top growth contributors table'],
            'deliverable': 'Funding trend series at annual and organisation level.',
            'gap': 'Exact reproduction improves after a donor-type map and full historical backfill of filings.',
            'source_keys': ['reports', 'donations', 'grants', 'igas', 'crosswalks'],
        },
        {
            'code': '3.0',
            'title': 'Sector Contribution to National Development',
            'status': 'Direct',
            'objective': 'Create the chapter-level contribution summary by combining projects, counties, employment, collaborations, and spending.',
            'use_data': [
                'Roll up metrics from project, county, employment, and collaboration tables into one summary layer.',
                'Use project spending, beneficiaries, staffing, and county spread as the headline contribution pillars.',
                'Publish a balanced scorecard that links the sub-sections in Chapter Three.',
            ],
            'visuals': ['Contribution scorecard', 'Summary dashboard', 'Small multiples'],
            'deliverable': 'One summary view that explains how the sector contributes nationally.',
            'gap': 'This depends mostly on completing the supporting Chapter Three sub-analyses.',
            'source_keys': ['reports', 'projects', 'counties', 'staff', 'collaborations'],
        },
        {
            'code': '3.1',
            'title': 'Sectors of Operation for Newly Registered NGOs',
            'status': 'Partial',
            'objective': 'Identify which sectors newly registered NGOs prefer to operate in during the reporting year.',
            'use_data': [
                'Filter reports where registration date falls inside the target fiscal year.',
                'Use the organisation’s available sector evidence from project rows or project summary rows.',
                'Standardize sector spellings before ranking the preferred sectors.',
            ],
            'visuals': ['Ranked sector bars', 'Pareto chart', 'Treemap'],
            'deliverable': 'Preferred sector distribution for the newly registered cohort.',
            'gap': 'Exact replication is stronger if you store a dedicated registration-time sector field instead of inferring from project rows.',
            'source_keys': ['registrations', 'projects', 'project_summary', 'crosswalks'],
        },
        {
            'code': '3.2',
            'title': 'Utilisation of Funds on Programmes/Projects',
            'status': 'Direct',
            'objective': 'Measure how much was spent on programme and project implementation across sectors and geographies.',
            'use_data': [
                'Sum `spending_per_county` with fallback to `amount_spent_kenya + amount_spent_other` per project row.',
                'Group by sector to build sector utilisation tables and by geography to split Kenya vs other countries.',
                'Use beneficiaries and completion status as supporting context for interpretation.',
            ],
            'visuals': ['Sector spending bars', 'Kenya vs other countries stacked bars', 'Bubble chart for spend vs beneficiaries'],
            'deliverable': 'Project utilisation table and chart pack by sector and geography.',
            'gap': 'Only needs sector standardization for clean ranking.',
            'source_keys': ['projects', 'countries', 'crosswalks'],
        },
        {
            'code': '3.3',
            'title': 'Counties of Operation',
            'status': 'Direct',
            'objective': 'Map where NGOs operate using both declared counties of operation and project implementation footprints.',
            'use_data': [
                'Explode the comma-separated `counties` field to one county per row.',
                'Union county declarations with project-level county mentions.',
                'Count organisations, projects, and spending per county.',
            ],
            'visuals': ['County choropleth', 'Ranked county bars', 'County heatmap'],
            'deliverable': 'National county coverage view across all filed returns.',
            'gap': 'Use a county name cleaner so aliases such as Nairobi vs Nairobi City do not split the same county.',
            'source_keys': ['reports', 'counties', 'projects', 'crosswalks'],
        },
        {
            'code': '3.3.1',
            'title': 'County of Operations for Newly Registered NGOs',
            'status': 'Direct',
            'objective': 'Show the county footprint preferred by newly registered NGOs in the target fiscal year.',
            'use_data': [
                'Filter to organisations registered within the fiscal-year window.',
                'Explode their `counties` selections to one county per row.',
                'Count organisations per county and rank the results.',
            ],
            'visuals': ['Top counties bar chart', 'County map', 'Monthly registration-by-county matrix'],
            'deliverable': 'County distribution for the newly registered cohort.',
            'gap': 'This is already well-supported if registration dates and county selections are complete.',
            'source_keys': ['registrations', 'counties', 'reports'],
        },
        {
            'code': '3.3.2',
            'title': 'Project Implementation per County',
            'status': 'Direct',
            'objective': 'Count and compare project implementation intensity across counties.',
            'use_data': [
                'Use project-level county values as the primary implementation geography.',
                'Fallback to report-level counties where project county is missing and the organisation has a clear footprint.',
                'Rank counties by project rows, spend, and beneficiaries.',
            ],
            'visuals': ['County project count bars', 'Spend vs project count scatter', 'County dashboard map'],
            'deliverable': 'County project footprint and implementation intensity summary.',
            'gap': 'A project-county cleaner is useful because some rows store county names inconsistently.',
            'source_keys': ['projects', 'counties', 'crosswalks'],
        },
        {
            'code': '3.4',
            'title': 'Leading NGOs in Utilisation of Funds on Projects',
            'status': 'Direct',
            'objective': 'Rank organisations by declared project spend and separate international and national players where needed.',
            'use_data': [
                'Aggregate project spending totals by report and organisation name.',
                'Use `scope` to split international and national NGOs.',
                'Sort descending and publish top-N tables with contribution shares.',
            ],
            'visuals': ['Top 10 horizontal bars', 'Scope split leaderboard', 'Cumulative share chart'],
            'deliverable': 'Leading-spenders tables by organisation and scope.',
            'gap': 'Organisation name normalization helps prevent duplicate spellings from splitting totals.',
            'source_keys': ['reports', 'projects', 'crosswalks'],
        },
        {
            'code': '3.5',
            'title': 'Leading Sector Players',
            'status': 'Direct',
            'objective': 'Identify the leading NGOs inside each sector based on project spend.',
            'use_data': [
                'Standardize sector labels first.',
                'Group project spending by sector and organisation.',
                'Select the top five or top ten organisations inside each sector.',
            ],
            'visuals': ['Sector leaderboard tables', 'Facet bars by sector', 'Treemap of sector leaders'],
            'deliverable': 'Sector-by-sector league tables for top players.',
            'gap': 'The main requirement is a stable sector normalization map.',
            'source_keys': ['projects', 'reports', 'crosswalks'],
        },
        {
            'code': '3.6',
            'title': 'Bottom-up Economic Transformation Agenda (BETA)',
            'status': 'Partial',
            'objective': 'Reframe project spending and activity into BETA pillars such as health, agriculture, MSMEs, housing, and digital/youth outcomes.',
            'use_data': [
                'Create a sector-to-BETA lookup table approved by the analysis team.',
                'Map project sectors into BETA pillars and aggregate spending, beneficiaries, and county spread.',
                'Profile top organisations contributing to each pillar.',
            ],
            'visuals': ['BETA pillar bars', 'Top NGOs per pillar', 'County-by-pillar heatmap'],
            'deliverable': 'BETA-aligned contribution dashboard with pillar drill-downs.',
            'gap': 'This needs an agreed crosswalk before the figures can be treated as official.',
            'source_keys': ['projects', 'counties', 'crosswalks'],
        },
        {
            'code': '3.7',
            'title': 'Employment Creation',
            'status': 'Direct',
            'objective': 'Measure salaried employment, expatriate presence, and volunteer/intern engagement across the sector.',
            'use_data': [
                'Use staff current-year counts in Kenya and other countries.',
                'Use volunteer current-year counts to measure volunteerism and internship engagement.',
                'Use came-in and left fields to estimate staff turnover where available.',
            ],
            'visuals': ['Employment trend lines', 'Salaried vs volunteer split', 'Kenyan vs expatriate stacked bars'],
            'deliverable': 'Employment and volunteerism summary with local/foreign splits.',
            'gap': 'Long-term trendlines improve as more historical fiscal years are loaded.',
            'source_keys': ['reports', 'staff'],
        },
        {
            'code': '3.8',
            'title': 'Collaborations and Networking',
            'status': 'Direct',
            'objective': 'Measure both the types of institutions NGOs work with and the nature of those collaborations.',
            'use_data': [
                'Count collaboration rows by `partner_type`.',
                'Convert saved collaboration markers into binary values and sum them, treating both Yes/No values and echoed partner-type values as checked flags.',
                'Build two layers: type of collaborator and nature of collaboration.',
            ],
            'visuals': ['Partner-type bar chart', 'Nature-of-collaboration doughnut', 'Matrix heatmap'],
            'deliverable': 'Collaboration profile by partner class and engagement type.',
            'gap': 'Very little extra work is needed beyond cleaning partner type labels.',
            'source_keys': ['collaborations', 'crosswalks'],
        },
        {
            'code': '4.1',
            'title': 'Funds Received',
            'status': 'Partial',
            'objective': 'Analyse where funding comes from, who provides it, and how source patterns change over time.',
            'use_data': [
                'Aggregate donation, grant, and IGA rows by fiscal year and organisation.',
                'Map donor categories into official donor-type buckets.',
                'Map donor countries into continents for geographical distribution.',
            ],
            'visuals': ['Donor-type bars', 'Continent split doughnut', 'Comparative year-on-year funding bars'],
            'deliverable': 'Funds-received chapter with donor-type, continent, and leading-NGO views.',
            'gap': 'Needs donor-type and country-to-continent crosswalks for exact alignment with the official report.',
            'source_keys': ['donations', 'grants', 'igas', 'countries', 'crosswalks'],
        },
        {
            'code': '4.2',
            'title': 'Utilisation of Funds',
            'status': 'Partial',
            'objective': 'Break expenditure into project cost, staffing, administration, and other line items with Kenya vs other-country splits.',
            'use_data': [
                'Use payments as the primary expenditure ledger.',
                'Map payment descriptions into standard expenditure lines such as project cost, local staff, international staff, and administration.',
                'Use assets and project spending as validation layers for tangible-asset and project-cost interpretation.',
            ],
            'visuals': ['Expenditure breakdown doughnut', 'Kenya vs other countries stacked bars', 'Utilisation ratio cards'],
            'deliverable': 'Expenditure structure analysis and utilisation breakdown.',
            'gap': 'A standardized `expense_line_map` is needed because payment descriptions are not yet fully normalized.',
            'source_keys': ['payments', 'projects', 'assets', 'crosswalks'],
        },
        {
            'code': '4.3',
            'title': 'Sector Sustainability',
            'status': 'Partial',
            'objective': 'Estimate sustainability using governance, local funding mix, liquidity, and organisational continuity signals.',
            'use_data': [
                'Combine audited status, local-funding share, self-generated income, opening vs closing balances, and governance indicators.',
                'Use collaboration strength and staffing continuity as supporting resilience signals.',
                'Score each organisation against an agreed sustainability framework and summarize the sector pattern.',
            ],
            'visuals': ['Sustainability radar', 'Composite score distribution', 'Dimension scorecard'],
            'deliverable': 'A sustainability index with interpretable dimensions.',
            'gap': 'This requires an agreed scoring model and benchmark thresholds before figures become policy-grade.',
            'source_keys': ['reports', 'donations', 'igas', 'assets', 'governance', 'staff', 'collaborations', 'crosswalks'],
        },
    ]


def get_plotly_graph_objects():
    try:
        from plotly import graph_objects as go
        return go
    except Exception:
        return None


def data_interp_plot_html(figure):
    if figure is None:
        return None
    return figure.to_html(
        full_html=False,
        include_plotlyjs=False,
        config={
            'displayModeBar': False,
            'responsive': True,
            'scrollZoom': False,
        },
    )


def data_interp_clean_text(raw_value):
    if raw_value in [None, '']:
        return ''
    return re.sub(r'\s+', ' ', str(raw_value).strip())


def data_interp_option_key(raw_value):
    text = data_interp_clean_text(raw_value)
    text = text.replace('’', "'")
    return re.sub(r'\s+', ' ', text).strip().upper()


def data_interp_format_money(value):
    return f"KES {float(value or 0):,.2f}"


def data_interp_percent(value, total):
    if not total:
        return 0.0
    return round((float(value or 0) / float(total)) * 100, 2)


def data_interp_normalize_org(report):
    display_name = data_interp_clean_text(getattr(report, 'pbo_name', None)) or f"Report {report.id}"
    return normalize_pbo_name_value(display_name) or f"REPORT_{report.id}", display_name.upper()


def data_interp_raw_donor_category(raw_value):
    raw_key = data_interp_option_key(raw_value)
    if not raw_key:
        return 'Unspecified'
    alias_map = {
        "HEADQUARTER OF THIS NGO": 'Headquarter of this PBO',
        "DIRECTORS' CONTRIBUTION": "Directors' Contributions",
        "DIRECTORS CONTRIBUTION": "Directors' Contributions",
        "DIRECTORS CONTRIBUTIONS": "Directors' Contributions",
        "NON-PROFIT ORGANIZATIONS (NGOS / FBOS)": 'Non-Profit Organizations (PBOs / FBOs)',
        "NON-PROFIT ORGANIZATIONS (PBOS / FBOS)": 'Non-Profit Organizations (PBOs / FBOs)',
        "FOREIGN GOVERNMENT AGENCY (E.G. SIDA, USAID, NORAD, UK AID, FOREIGN EMBASSIES / HIGH COMMISSIONS)": 'Foreign Government Agency (e.g. SIDA, USAID, NORAD, UK AID, Foreign Embassies / High Commissions)',
        "RETURNS FROM INVESTMENTS (E.G., DIVIDENDS & INTEREST)": 'Returns from investments (e.g., dividends & interest)',
        "RETURNS FROM INVESTMENTS(EG DIVIDENDS AND INTEREST)": 'Returns From Investments(eg dividends and interest)',
        "RETURNS FROM INVESTMENTS(EG DIVIDENDS & INTEREST)": 'Returns From Investments(eg dividends and interest)',
        "OTHER": 'Other (Specified)',
    }
    option_lookup = {data_interp_option_key(option): option for option in DATA_INTERP_DONOR_CATEGORY_OPTIONS}
    if raw_key in option_lookup:
        return option_lookup[raw_key]
    if raw_key in alias_map:
        return alias_map[raw_key]
    return 'Other (Specified)'


def data_interp_display_country(raw_value):
    return data_interp_clean_text(raw_value) or 'Unspecified Country'


def data_interp_normalize_county(raw_value):
    text = data_interp_clean_text(raw_value).upper()
    if not text:
        return None
    text = re.sub(r'[^A-Z0-9\s\'-]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    aliases = {
        'NAIROBI CITY': 'Nairobi',
        'NAIROBI': 'Nairobi',
        'MURANGA': "Murang'a",
        "MURANG'A": "Murang'a",
        'ELGEYO MARAKWET': 'Elgeyo Marakwet',
        'ELGEYO-MARAKWET': 'Elgeyo Marakwet',
        'TAITA TAVETA': 'Taita Taveta',
        'TAITA-TAVETA': 'Taita Taveta',
        'THARAKA NITHI': 'Tharaka Nithi',
        'THARAKA-NITHI': 'Tharaka Nithi',
        'TRANS NZOIA': 'Trans Nzoia',
        'TRANS-NZOIA': 'Trans Nzoia',
        'UASIN GISHU': 'Uasin Gishu',
        'UASIN-GISHU': 'Uasin Gishu',
        'WEST POKOT': 'West Pokot',
        'HOMA BAY': 'Homa Bay',
        'TANA RIVER': 'Tana River',
        'NYANDARUA': 'Nyandarua',
    }
    return aliases.get(text, text.title())


def data_interp_normalize_sector(raw_value):
    text = data_interp_clean_text(raw_value).upper()
    if not text:
        return 'Unspecified'
    if any(token in text for token in ('HIV', 'AIDS', 'HEALTH', 'REPRODUCTIVE')):
        return 'Health'
    if 'CHILD' in text:
        return 'Children'
    if 'EDUCATION' in text:
        return 'Education'
    if 'RELIEF' in text or 'DISASTER' in text:
        return 'Relief / Disaster Management'
    if 'WATER' in text or 'SANITATION' in text or 'WASH' in text:
        return 'Water and Sanitation'
    if 'AGRIC' in text or 'FOOD' in text or 'NUTRITION' in text:
        return 'Agriculture'
    if 'ENVIRONMENT' in text or 'CLIMATE' in text or 'CONSERVATION' in text:
        return 'Environment'
    if 'YOUTH' in text:
        return 'Youth'
    if 'GENDER' in text or 'WOMEN' in text:
        return 'Gender'
    if 'GOVERNANCE' in text or 'HUMAN RIGHTS' in text or 'DEMOCR' in text:
        return 'Governance'
    if 'REFUGEE' in text:
        return 'Refugees'
    if 'DISABILITY' in text:
        return 'Disability'
    if 'PEACE' in text:
        return 'Peace Building'
    if 'MICRO' in text or 'INFORMAL SECTOR' in text or 'ECONOMIC' in text or 'MSME' in text:
        return 'Microfinance / MSMEs'
    if 'ICT' in text or 'DIGITAL' in text or 'TECH' in text or 'INFORMATION' in text:
        return 'ICT / Digital'
    if 'HOUSING' in text or 'SETTLEMENT' in text or 'CONSTRUCTION' in text:
        return 'Housing and Settlement'
    if 'WELFARE' in text:
        return 'Welfare'
    if 'OLD AGE' in text or 'ELDER' in text:
        return 'Old Age'
    return text.title()


def data_interp_normalize_partner_type(raw_value):
    raw_key = data_interp_option_key(raw_value)
    if not raw_key:
        return 'Unspecified'

    option_lookup = {data_interp_option_key(option): option for option in DATA_INTERP_COLLAB_PARTNER_OPTIONS}
    if raw_key in option_lookup:
        return option_lookup[raw_key]

    alias_map = {
        'NGOS': 'PBOs',
        'NGO': 'PBOs',
        'PBOS': 'PBOs',
        'NON-GOVERNMENTAL ORGANIZATIONS': 'PBOs',
        'NON PROFIT ORGANIZATIONS': 'PBOs',
        'CBOS': 'CBOs',
        'COMMUNITY BASED ORGANIZATIONS': 'CBOs',
        'COMMUNITY-BASED ORGANIZATIONS': 'CBOs',
        'FBOS': 'FBOs',
        'FAITH BASED ORGANIZATIONS': 'FBOs',
        'FAITH-BASED ORGANIZATIONS': 'FBOs',
        'RELIGIOUS INSTITUTIONS': 'FBOs',
        'RESEARCH / ACADEMIC INSTITUTIONS': 'Research Institutions',
        'UNIVERSITIES AND COLLEGES': 'Academic Institutions',
        'HOSPITALS AND HEALTH CENTERS': 'Health Institutions',
        'AGENCY OF KENYA GOVERNMENT': 'Government Agencies',
        'GOVERNMENT': 'Government Agencies',
        'GOVERNMENT AGENCY': 'Government Agencies',
        'GOVERNMENT AGENCIES AND DEPARTMENTS': 'Government Agencies',
        'COUNTY GOVERNMENT OFFICES': 'County Government',
        'MEDIA ORGANIZATIONS': 'Media',
        'MEDIA ORGANIZATIONS AND OUTLETS': 'Media',
        'DONOR AGENCY': 'Donor Agencies',
        'PRIVATE SECTOR': 'Corporate',
        'PRIVATE SECTOR COMPANIES': 'Corporate',
        'BUSINESS ORGANIZATIONS': 'Corporate',
        'OTHER': 'Other (Specified)',
    }
    if raw_key in alias_map:
        return alias_map[raw_key]

    if 'PBO' in raw_key or 'NGO' in raw_key:
        return 'PBOs'
    if 'CBO' in raw_key or 'COMMUNITY' in raw_key:
        return 'CBOs'
    if 'FBO' in raw_key or 'FAITH' in raw_key or 'RELIG' in raw_key:
        return 'FBOs'
    if 'RESEARCH' in raw_key:
        return 'Research Institutions'
    if 'ACADEMIC' in raw_key or 'UNIVERSIT' in raw_key or 'COLLEGE' in raw_key:
        return 'Academic Institutions'
    if 'HEALTH' in raw_key or 'HOSPITAL' in raw_key or 'MEDIC' in raw_key:
        return 'Health Institutions'
    if 'COUNTY' in raw_key:
        return 'County Government'
    if 'NATIONAL' in raw_key and 'GOV' in raw_key:
        return 'National Government'
    if 'GOV' in raw_key:
        return 'Government Agencies'
    if 'MEDIA' in raw_key:
        return 'Media'
    if 'DONOR' in raw_key or 'BILATERAL' in raw_key or 'MULTILATERAL' in raw_key:
        return 'Donor Agencies'
    if 'CORPORATE' in raw_key or 'PRIVATE' in raw_key or 'BUSINESS' in raw_key or 'SECTOR' in raw_key:
        return 'Corporate'
    return 'Other (Specified)'


def data_interp_normalize_donor_type(raw_value):
    text = data_interp_clean_text(raw_value).upper()
    if not text or text == 'MISSING':
        return 'Unspecified'
    if 'UNITED NATIONS' in text:
        return 'United Nations Agencies'
    if 'FOREIGN GOVERNMENT' in text or 'EMBASSY' in text or 'HIGH COMMISSION' in text:
        return 'Foreign Government Agencies'
    if 'AGENCY OF KENYA GOVERNMENT' in text or 'KENYA GOVERNMENT' in text or 'NATIONAL GOVERNMENT' in text:
        return 'Kenya Government Agencies'
    if 'DIRECTOR' in text or 'MEMBER' in text:
        return 'Members and Directors'
    if 'INDIVIDUAL' in text:
        return 'Individual Donors'
    if 'SELF GENERATED' in text or 'RETURNS FROM INVESTMENTS' in text or 'BUSINESS INCOME' in text:
        return 'Self-generated Income'
    if 'CORPORATE' in text or 'PRIVATE SECTOR' in text:
        return 'Corporates'
    if 'RESEARCH' in text or 'ACADEMIC' in text:
        return 'Research / Academic Institutions'
    if 'HEADQUARTER' in text or 'AFFILIATE NGO' in text:
        return 'Affiliate NGOs'
    if 'AFFILIATE FBO' in text:
        return 'Affiliate FBOs'
    if 'FBO' in text or 'RELIGIOUS' in text or 'FAITH-BASED' in text:
        return 'Faith-Based Organisations'
    if any(token in text for token in ('NGO', 'CBO', 'FOUNDATION', 'TRUST', 'NON-PROFIT', 'CHARITABLE')):
        return 'NGOs, CBOs, Foundations and Trusts'
    return text.title()


def data_interp_display_country(raw_value):
    text = data_interp_clean_text(raw_value)
    if not text or text.upper() == 'MISSING':
        return 'Unspecified'
    upper = text.upper()
    aliases = {
        'USA': 'United States',
        'U.S.A.': 'United States',
        'U.S.A': 'United States',
        'U.S.': 'United States',
        'US': 'United States',
        'UNITED STATES OF AMERICA': 'United States',
        'UK': 'United Kingdom',
        'U.K.': 'United Kingdom',
        'UNITED KINGDOM': 'United Kingdom',
        'UAE': 'United Arab Emirates',
        'DRC': 'DR Congo',
        'DEMOCRATIC REPUBLIC OF CONGO': 'DR Congo',
    }
    if upper in aliases:
        return aliases[upper]
    if text == upper and len(text) > 4:
        return text.title()
    return text


def data_interp_country_to_continent(raw_value):
    text = data_interp_clean_text(raw_value).upper()
    if not text or text == 'MISSING':
        return 'Unspecified'

    africa = {
        'KENYA', 'UGANDA', 'TANZANIA', 'SOMALIA', 'ETHIOPIA', 'RWANDA', 'BURUNDI', 'SOUTH SUDAN',
        'SUDAN', 'ERITREA', 'NIGERIA', 'SOUTH AFRICA', 'GHANA', 'ZAMBIA', 'ZIMBABWE', 'BOTSWANA',
        'SENEGAL', 'SOMALILAND', 'CONGO', 'DRC', 'MOZAMBIQUE',
    }
    europe = {'UNITED KINGDOM', 'UK', 'NETHERLANDS', 'GERMANY', 'SWEDEN', 'SWITZERLAND', 'ITALY', 'IRELAND', 'DENMARK', 'NORWAY', 'FRANCE', 'BELGIUM', 'SPAIN'}
    north_america = {'UNITED STATES OF AMERICA', 'USA', 'UNITED STATES', 'CANADA'}
    asia = {'JAPAN', 'CHINA', 'INDIA', 'QATAR', 'UAE', 'UNITED ARAB EMIRATES', 'SAUDI ARABIA', 'KOREA'}
    oceania = {'AUSTRALIA', 'NEW ZEALAND'}
    south_america = {'BRAZIL', 'ARGENTINA', 'COLOMBIA', 'PERU', 'CHILE', 'ECUADOR'}

    if text in africa:
        return 'Africa'
    if text in europe:
        return 'Europe'
    if text in north_america:
        return 'North America'
    if text in asia:
        return 'Asia'
    if text in oceania:
        return 'Oceania'
    if text in south_america:
        return 'South America'
    return 'Other / Unspecified'


def data_interp_expense_line(raw_value):
    text = data_interp_clean_text(raw_value).upper()
    if not text:
        return 'Other Expenditure'
    if any(token in text for token in ('SALARY', 'PAYROLL', 'WAGE', 'ALLOWANCE', 'STAFF', 'PERSONNEL')):
        if any(token in text for token in ('EXPAT', 'FOREIGN', 'INTERNATIONAL')):
            return 'International Staff'
        return 'Local Staff'
    if any(token in text for token in ('ADMIN', 'RENT', 'OFFICE', 'UTILITY', 'AUDIT', 'MANAGEMENT', 'TRAVEL', 'COMMUNICATION', 'LEGAL', 'GOVERNANCE')):
        return 'Administration Costs'
    if any(token in text for token in ('ASSET', 'EQUIPMENT', 'FURNITURE', 'VEHICLE', 'COMPUTER', 'BUILDING', 'CONSTRUCTION')):
        return 'Purchase of Tangible Assets'
    if any(token in text for token in ('PROJECT', 'PROGRAM', 'PROGRAMME', 'IMPLEMENT', 'BENEFICIAR', 'FIELD', 'TRAINING', 'RELIEF', 'ACTIVITY')):
        return 'Project Cost'
    return 'Other Expenditure'


def data_interp_beta_pillar(sector_label):
    sector = data_interp_normalize_sector(sector_label)
    if sector in {'Health', 'Water and Sanitation', 'Disability'}:
        return 'Universal Health Coverage'
    if sector in {'Agriculture', 'Environment'}:
        return 'Agriculture and Food Systems'
    if sector in {'Microfinance / MSMEs', 'Welfare'}:
        return 'MSMEs and Local Enterprise'
    if sector in {'Housing and Settlement'}:
        return 'Affordable Housing'
    if sector in {'ICT / Digital', 'Education', 'Youth'}:
        return 'Digital, Skills and Youth Economy'
    return 'Social and Community Development'


def data_interp_is_yes(raw_value):
    return str(raw_value or '').strip().lower() in {'1', 'true', 'yes', 'y'}


def data_interp_collaboration_flag_checked(raw_value):
    text = data_interp_clean_text(raw_value)
    if not text:
        return False
    return text.upper() not in {'0', 'FALSE', 'NO', 'N', 'NONE', 'NULL', 'MISSING'}


def data_interp_is_filed(report):
    workflow_status = str(getattr(report, 'workflow_status', '') or '').strip().lower()
    return bool(
        getattr(report, 'submitted_at', None)
        or getattr(report, 'submission_date', None)
        or workflow_status in {'submitted', 'approved', 'pending', 'returned', 'under review', 'reviewed'}
    )


def data_interp_is_audited(report):
    return data_interp_is_yes(getattr(report, 'audited', None))


def data_interp_report_funding_total(report):
    donation_total = sum(legacy_zero_float(report, getattr(row, 'amount', None)) for row in getattr(report, 'donations', []))
    grant_total = sum(legacy_zero_float(report, getattr(row, 'amount', None)) for row in getattr(report, 'grants', []))
    iga_total = sum(legacy_zero_float(report, getattr(row, 'amount', None)) for row in getattr(report, 'igas', []))
    row_total = donation_total + grant_total + iga_total
    if row_total > 0:
        return row_total
    income_total = legacy_zero_float(report, getattr(report, 'income_b2_total', None))
    if income_total > 0:
        return income_total
    receipts_total = legacy_zero_float(report, getattr(report, 'receipts_total', None))
    opening_balance = legacy_zero_float(report, getattr(report, 'cash_balance_previous_year', None))
    return max(0.0, receipts_total - opening_balance)


def data_interp_report_local_funding_total(report):
    local_total = 0.0
    for donation in getattr(report, 'donations', []):
        if data_interp_clean_text(getattr(donation, 'country', None)).upper() == 'KENYA':
            local_total += legacy_zero_float(report, getattr(donation, 'amount', None))
    for grant in getattr(report, 'grants', []):
        if data_interp_clean_text(getattr(grant, 'country', None)).upper() == 'KENYA':
            local_total += legacy_zero_float(report, getattr(grant, 'amount', None))
    local_total += sum(legacy_zero_float(report, getattr(row, 'amount', None)) for row in getattr(report, 'igas', []))
    return local_total


def data_interp_get_fiscal_year_options():
    today = utc_now().date()
    default_start_year, default_end_year = get_current_fiscal_year(today)
    period_bounds_query = db.session.query(
        func.min(func.coalesce(PBOReport.reporting_period_start, PBOReport.reporting_period_end)),
        func.max(func.coalesce(PBOReport.reporting_period_end, PBOReport.reporting_period_start)),
    )
    scope_label = data_interp_scope_label_for_user(current_user)

    min_period_date, max_period_date = period_bounds_query.first() or (None, None)
    if min_period_date and max_period_date:
        start_year = fiscal_start_year(min_period_date)
        end_year = fiscal_start_year(max_period_date)
    else:
        start_year = default_start_year
        end_year = default_start_year

    option_start_year = min(start_year, default_start_year)
    option_end_year = max(end_year, default_start_year)
    options = [{'value': 'all', 'label': 'All Years'}]
    for year in range(option_start_year, option_end_year + 1):
        options.append({'value': f"{year}-{year + 1}", 'label': f"FY {year}-{year + 1}"})

    selected_value = (request.form.get('fy') or request.args.get('fy') or f"{default_start_year}-{default_end_year}").strip()
    valid_values = {item['value'] for item in options}
    if selected_value not in valid_values:
        selected_value = f"{default_start_year}-{default_end_year}"

    return options, selected_value, scope_label


def data_interp_load_accessible_reports_for_user(user):
    query = PBOReport.query.options(
        selectinload(PBOReport.assets),
        selectinload(PBOReport.igas),
        selectinload(PBOReport.donations),
        selectinload(PBOReport.grants),
        selectinload(PBOReport.payments),
        selectinload(PBOReport.officials),
        selectinload(PBOReport.project_implementations),
        selectinload(PBOReport.projects_carried_out),
        selectinload(PBOReport.collaborations),
    )
    return query.order_by(PBOReport.created_at.desc()).all()


def data_interp_load_accessible_reports():
    return data_interp_load_accessible_reports_for_user(current_user)


def data_interp_sorted_year_labels(labels):
    def sort_key(label):
        if not label:
            return (9999, label)
        base = str(label).split('/')[0]
        try:
            return (int(base), label)
        except ValueError:
            return (9999, label)

    return sorted(labels, key=sort_key)


def data_interp_build_bar_chart(
    go,
    labels,
    values,
    title,
    color='#103a7d',
    horizontal=False,
    money_axis=False,
    height=420,
    x_tickangle=None,
    show_text=True,
):
    if not labels or not values:
        return None
    text_values = [f"{value:,.0f}" for value in values] if show_text else None
    text_position = 'outside' if show_text else 'none'
    if horizontal:
        figure = go.Figure(
            data=[
                go.Bar(
                    y=labels,
                    x=values,
                    orientation='h',
                    marker={'color': color},
                    text=text_values,
                    textposition=text_position,
                    hovertemplate='<b>%{y}</b><br>%{x:,.2f}<extra></extra>' if money_axis else '<b>%{y}</b><br>%{x:,.0f}<extra></extra>',
                )
            ]
        )
    else:
        figure = go.Figure(
            data=[
                go.Bar(
                    x=labels,
                    y=values,
                    marker={'color': color},
                    text=text_values,
                    textposition=text_position,
                    hovertemplate='<b>%{x}</b><br>%{y:,.2f}<extra></extra>' if money_axis else '<b>%{x}</b><br>%{y:,.0f}<extra></extra>',
                )
            ]
        )
    figure.update_layout(
        title={'text': title, 'x': 0.03, 'font': {'size': 16, 'color': '#13253d'}},
        autosize=True,
        height=height,
        margin={'l': 50, 'r': 24, 't': 58, 'b': 52},
        paper_bgcolor='white',
        plot_bgcolor='white',
        showlegend=False,
    )
    figure.update_xaxes(showgrid=not horizontal, gridcolor='#dfe7f1', automargin=True)
    figure.update_yaxes(showgrid=horizontal, gridcolor='#dfe7f1', automargin=True)
    if x_tickangle is not None:
        figure.update_xaxes(tickangle=x_tickangle)
    return data_interp_plot_html(figure)


def data_interp_build_donut_chart(go, labels, values, title, colors=None):
    if not labels or not values:
        return None
    figure = go.Figure(
        data=[
            go.Pie(
                labels=labels,
                values=values,
                hole=0.55,
                marker={'colors': colors or ['#103a7d', '#3a6ab3', '#6c96d8', '#9abde8', '#d0e3f8']},
                textinfo='percent',
                textposition='inside',
                insidetextorientation='auto',
                hovertemplate='<b>%{label}</b><br>%{value:,.2f}<extra></extra>',
            )
        ]
    )
    figure.update_layout(
        title={'text': title, 'x': 0.03, 'font': {'size': 16, 'color': '#13253d'}},
        autosize=True,
        height=420,
        margin={'l': 24, 'r': 24, 't': 58, 'b': 34},
        paper_bgcolor='white',
        showlegend=True,
        legend={'orientation': 'h', 'y': -0.14, 'x': 0},
    )
    return data_interp_plot_html(figure)


def data_interp_build_heatmap(go, x_labels, y_labels, z_values, title, height=420):
    if not x_labels or not y_labels or not z_values:
        return None
    figure = go.Figure(
        data=[
            go.Heatmap(
                x=x_labels,
                y=y_labels,
                z=z_values,
                colorscale=[
                    [0.0, '#edf3fb'],
                    [0.25, '#cfe0f5'],
                    [0.5, '#8fb2e1'],
                    [0.75, '#3a6ab3'],
                    [1.0, '#103a7d'],
                ],
                colorbar={'title': 'Count'},
                hovertemplate='<b>%{y}</b><br>%{x}: %{z:,}<extra></extra>',
            )
        ]
    )
    figure.update_layout(
        title={'text': title, 'x': 0.03, 'font': {'size': 16, 'color': '#13253d'}},
        autosize=True,
        height=height,
        margin={'l': 70, 'r': 24, 't': 58, 'b': 70},
        paper_bgcolor='white',
        plot_bgcolor='white',
    )
    figure.update_xaxes(showgrid=False, tickangle=-20, automargin=True)
    figure.update_yaxes(showgrid=False, automargin=True)
    return data_interp_plot_html(figure)


def data_interp_build_line_chart(go, labels, series, title):
    if not labels or not series:
        return None
    figure = go.Figure()
    for item in series:
        figure.add_trace(
            go.Scatter(
                x=labels,
                y=item['values'],
                mode='lines+markers',
                name=item['label'],
                line={'width': 3, 'color': item.get('color', '#103a7d')},
                marker={'size': 8},
                hovertemplate=f"<b>{item['label']}</b><br>%{{x}}<br>%{{y:,.2f}}<extra></extra>" if item.get('money') else f"<b>{item['label']}</b><br>%{{x}}<br>%{{y:,.0f}}<extra></extra>",
            )
        )
    figure.update_layout(
        title={'text': title, 'x': 0.03, 'font': {'size': 16, 'color': '#13253d'}},
        autosize=True,
        height=420,
        margin={'l': 44, 'r': 24, 't': 58, 'b': 48},
        paper_bgcolor='white',
        plot_bgcolor='white',
        legend={'orientation': 'h', 'y': -0.18, 'x': 0},
    )
    figure.update_xaxes(showgrid=False, automargin=True)
    figure.update_yaxes(showgrid=True, gridcolor='#dfe7f1', automargin=True)
    return data_interp_plot_html(figure)


def data_interp_build_dropdown_line_chart(go, labels, series_map, title):
    if not labels or not series_map:
        return None

    ordered_items = list(series_map.items())
    figure = go.Figure()
    total_values = [
        sum(values[index] for _, values in ordered_items)
        for index in range(len(labels))
    ]
    figure.add_trace(
        go.Scatter(
            x=labels,
            y=total_values,
            mode='lines+markers',
            name='All Donor Categories',
            visible=True,
            line={'width': 3, 'color': '#103a7d'},
            marker={'size': 8},
            hovertemplate='<b>Total</b><br>%{x}<br>%{y:,.2f}<extra></extra>',
        )
    )

    palette = ['#0e8b7d', '#d18a00', '#5f5ce1', '#0b8793', '#d64550', '#35517c']
    for index, (label, values) in enumerate(ordered_items):
        figure.add_trace(
            go.Scatter(
                x=labels,
                y=values,
                mode='lines+markers',
                name=label,
                visible=False,
                line={'width': 3, 'color': palette[index % len(palette)]},
                marker={'size': 8},
                hovertemplate=f"<b>{label}</b><br>%{{x}}<br>%{{y:,.2f}}<extra></extra>",
            )
        )

    buttons = [
        {
            'label': 'All Donor Categories',
            'method': 'update',
            'args': [
                {'visible': [True] + [False] * len(ordered_items)},
                {'title': {'text': title, 'x': 0.03, 'font': {'size': 16, 'color': '#13253d'}}},
            ],
        }
    ]
    for index, (label, _) in enumerate(ordered_items):
        visible = [False] * (len(ordered_items) + 1)
        visible[index + 1] = True
        buttons.append(
            {
                'label': label,
                'method': 'update',
                'args': [
                    {'visible': visible},
                    {'title': {'text': f"{title}: {label}", 'x': 0.03, 'font': {'size': 16, 'color': '#13253d'}}},
                ],
            }
        )

    figure.update_layout(
        title={'text': title, 'x': 0.03, 'font': {'size': 16, 'color': '#13253d'}},
        autosize=True,
        height=460,
        margin={'l': 44, 'r': 24, 't': 92, 'b': 48},
        paper_bgcolor='white',
        plot_bgcolor='white',
        updatemenus=[
            {
                'buttons': buttons,
                'direction': 'down',
                'showactive': True,
                'x': 0.0,
                'y': 1.18,
                'xanchor': 'left',
                'yanchor': 'top',
                'pad': {'r': 10, 't': 6},
            }
        ],
    )
    figure.update_xaxes(showgrid=False, automargin=True)
    figure.update_yaxes(showgrid=True, gridcolor='#dfe7f1', automargin=True)
    return data_interp_plot_html(figure)


def build_data_interpretation_context(scope_label, fiscal_year_options, selected_fiscal_year, analysis_result=None, analysis_requested=False):
    topic_plan = build_data_interpretation_topic_plan()
    chapter_counter = Counter()
    source_counter = Counter()
    readiness_counter = Counter()

    for topic in topic_plan:
        status = topic.get('status') or 'Partial'
        readiness_counter[status] += 1
        if topic['code'].startswith('2'):
            chapter_counter['Chapter 2'] += 1
        elif topic['code'].startswith('3'):
            chapter_counter['Chapter 3'] += 1
        else:
            chapter_counter['Chapter 4'] += 1
        topic['status_class'] = data_interp_status_class(status)
        topic['source_labels'] = [
            DATA_INTERPRETATION_SOURCE_LABELS.get(source_key, source_key.replace('_', ' ').title())
            for source_key in topic.get('source_keys', [])
        ]
        live_result = (analysis_result or {}).get('topic_results', {}).get(topic['code'])
        if live_result:
            topic['live_result'] = live_result
        for source_key in topic.get('source_keys', []):
            source_counter[source_key] += 1

    quality_checks = [
        'Use fiscal-year filters consistently; do not mix registration year and reporting year in the same table.',
        'Normalize organisation names before ranking the top NGOs to avoid split totals.',
        'Standardize sector spelling, county names, and donor categories before charting.',
        'Reconcile row-level totals against `receipts_total`, `payments_total`, and project spend summaries.',
        'Flag missing donor country, county, or sector values so incomplete records do not distort percentages.',
    ]

    workflow_steps = [
        {
            'step': '1. Scope the analysis window',
            'detail': 'Pick the fiscal year first, then limit the report population to all accessible returns or the authorised analyst scope.',
        },
        {
            'step': '2. Build chapter-ready fact tables',
            'detail': 'Use report, project, donation, grant, payment, county, staff, collaboration, and governance tables as separate analysis marts.',
        },
        {
            'step': '3. Standardize the taxonomies',
            'detail': 'Normalize sectors, counties, donor categories, countries, expenditure lines, and BETA mappings before any ranking or trend analysis.',
        },
        {
            'step': '4. Produce the 18 outputs',
            'detail': 'Aggregate each topic at the correct grain: fiscal year, organisation, sector, county, donor type, or BETA pillar.',
        },
        {
            'step': '5. QA before publication',
            'detail': 'Reconcile row-level totals against report totals, validate rankings, and compare chapter summaries with the downloadable Excel extracts.',
        },
    ]

    crosswalks = [
        {
            'name': 'sector_to_beta_pillar.csv',
            'purpose': 'Maps project sectors into approved BETA pillars for Chapter 3.6.',
        },
        {
            'name': 'donor_type_map.csv',
            'purpose': 'Maps raw donation categories into official donor-source groups for Chapter 4.1.',
        },
        {
            'name': 'country_to_continent.csv',
            'purpose': 'Maps donor countries to continents for the geographical distribution of funds.',
        },
        {
            'name': 'expense_line_map.csv',
            'purpose': 'Maps payment descriptions into standard expenditure lines for Chapter 4.2.',
        },
        {
            'name': 'sustainability_weights.json',
            'purpose': 'Stores weights and thresholds used to calculate the Chapter 4.3 sustainability index.',
        },
    ]

    tool_stack = [
        {
            'name': 'SQLAlchemy + SQL views',
            'purpose': 'Use for scoped extraction, joins, fiscal-year filters, and reusable reporting marts.',
        },
        {
            'name': 'Pandas',
            'purpose': 'Use for cleaning comma-separated fields, reshaping data, grouping, ranking, and reconciling totals.',
        },
        {
            'name': 'OpenPyXL + Excel',
            'purpose': 'Use for regulator-ready workbooks, downloadable tables, and quick validation by non-technical users.',
        },
        {
            'name': 'Plotly',
            'purpose': 'Use for interactive charts, in-app storytelling, and drill-down reporting without preloading every dataset on page load.',
        },
        {
            'name': 'GeoJSON / Kenya County Boundaries',
            'purpose': 'Use when you want county choropleths and geographic drill-downs that align with county rankings.',
        },
        {
            'name': 'Power BI / Tableau / Metabase',
            'purpose': 'Use for executive dashboards, scheduled refreshes, and external presentation packs built on the same exports.',
        },
    ]

    data_downloads = [
        {
            'label': 'Sector Report Data',
            'endpoint': 'sector_report_data',
            'description': 'Wide report-level table for profiling, QA, and top-level joins.',
        },
        {
            'label': 'Registration Period Export',
            'endpoint': 'download_registration_period_data',
            'description': 'Registration dates and reporting windows for growth and newly registered cohorts.',
        },
        {
            'label': 'Projects Export',
            'endpoint': 'download_projects_data',
            'description': 'Project sectors, counties, beneficiaries, and spending for Chapter Three.',
        },
        {
            'label': 'Donations Export',
            'endpoint': 'download_donations_data',
            'description': 'Donor category, donor country, and amount for funding analysis.',
        },
        {
            'label': 'Payments Export',
            'endpoint': 'download_payments_data',
            'description': 'Spend-line detail for expenditure and utilisation analysis.',
        },
        {
            'label': 'County Export',
            'endpoint': 'download_county_data',
            'description': 'County drill-down workbook for operational spread and county rankings.',
        },
        {
            'label': 'Staff & Volunteer Export',
            'endpoint': 'download_staff_volunteer_summary',
            'description': 'Employment and volunteer data for Chapter 3.7.',
        },
        {
            'label': 'Collaboration Export',
            'endpoint': 'download_collaboration_data',
            'description': 'Collaboration types and networking nature for Chapter 3.8.',
        },
        {
            'label': 'Governance Export',
            'endpoint': 'download_governance_data',
            'description': 'Governance and submission fields that support sustainability analysis.',
        },
    ]

    coverage_notes = [
        'Chapter Three is the strongest area in the current schema because projects, counties, employment, and collaborations are already structured.',
        'Chapter Four becomes highly reliable after donor-type, continent, and expenditure-line mappings are added.',
        'Chapter Two now focuses on the registration and filing signals already captured inside Form 14 data.',
    ]

    source_chart = {
        'labels': [
            DATA_INTERPRETATION_SOURCE_LABELS.get(source_key, source_key.replace('_', ' ').title())
            for source_key, _ in source_counter.most_common(7)
        ],
        'values': [count for _, count in source_counter.most_common(7)],
    }

    context = {
        'generated_at': format_datetime(utc_now()),
        'reference_report_title': 'Annual NGO Sector Report Year 2022 / 2023',
        'reference_report_publication': 'Published 15 June 2024',
        'scope_label': scope_label,
        'fiscal_year_options': fiscal_year_options,
        'selected_fiscal_year': selected_fiscal_year,
        'analysis_requested': analysis_requested,
        'coverage_notes': coverage_notes,
        'topic_plan': topic_plan,
        'workflow_steps': workflow_steps,
        'quality_checks': quality_checks,
        'crosswalks': crosswalks,
        'tool_stack': tool_stack,
        'data_downloads': data_downloads,
        'charts': {
            'readiness': {
                'labels': ['Direct', 'Partial'],
                'values': [
                    readiness_counter.get('Direct', 0),
                    readiness_counter.get('Partial', 0),
                ],
            },
            'chapters': {
                'labels': ['Chapter 2', 'Chapter 3', 'Chapter 4'],
                'values': [
                    chapter_counter.get('Chapter 2', 0),
                    chapter_counter.get('Chapter 3', 0),
                    chapter_counter.get('Chapter 4', 0),
                ],
            },
            'sources': source_chart,
        },
        'analysis_run': bool(analysis_result and analysis_result.get('analysis_run')),
        'analysis_error': (analysis_result or {}).get('analysis_error'),
        'analysis_generated_at': (analysis_result or {}).get('analysis_generated_at'),
        'analysis_focus_label': (analysis_result or {}).get('analysis_focus_label'),
        'analysis_summary_lines': (analysis_result or {}).get('analysis_summary_lines', []),
        'analysis_cards': (analysis_result or {}).get('analysis_cards', []),
        'qa_summary': (analysis_result or {}).get('qa_summary', data_interp_build_qa_summary([])),
        'qa_checks': (analysis_result or {}).get('qa_checks', []),
        'analysis_sections': (analysis_result or {}).get('analysis_sections', []),
        'top_sector_labels': (analysis_result or {}).get('top_sector_labels', []),
        'top_donor_labels': (analysis_result or {}).get('top_donor_labels', []),
    }
    return context


def build_data_interpretation_analysis(reports, selected_fiscal_year, scope_label):
    go = get_plotly_graph_objects()
    analysis_generated_at = format_datetime(utc_now())
    qa_tolerance = 1.0
    if not reports:
        return {
            'analysis_run': False,
            'analysis_error': 'No accessible reports are available for analysis yet.',
            'analysis_generated_at': analysis_generated_at,
        }

    if selected_fiscal_year == 'all':
        selected_reports = list(reports)
        focus_label = 'All Accessible Years'
        range_start = None
        range_end = None
    else:
        range_start, range_end, start_year, end_year = fiscal_year_date_range(selected_fiscal_year)
        selected_reports = [report for report in reports if report_matches_fiscal_year(report, range_start, range_end)]
        focus_label = f"FY {start_year}-{end_year}"

    if not selected_reports:
        return {
            'analysis_run': False,
            'analysis_error': f'No reports matched {focus_label}.',
            'analysis_generated_at': analysis_generated_at,
        }

    yearly_metrics = defaultdict(lambda: {
        'new_registrations': 0,
        'reports': 0,
        'filed_reports': 0,
        'audited_reports': 0,
        'funding_total': 0.0,
        'project_spend': 0.0,
        'employment_total': 0,
    })
    yearly_donor_category_totals = defaultdict(lambda: defaultdict(float))
    org_year_funding = defaultdict(lambda: defaultdict(float))
    org_display_all_years = {}

    for report in reports:
        reporting_label = data_interp_reporting_label(report)
        if reporting_label:
            yearly_metrics[reporting_label]['reports'] += 1
            if data_interp_is_filed(report):
                yearly_metrics[reporting_label]['filed_reports'] += 1
            if data_interp_is_audited(report):
                yearly_metrics[reporting_label]['audited_reports'] += 1
            yearly_metrics[reporting_label]['funding_total'] += data_interp_report_funding_total(report)
            yearly_metrics[reporting_label]['project_spend'] += sum(
                resolve_project_spending_amount(project, report=report) for project in getattr(report, 'project_implementations', [])
            )
            yearly_metrics[reporting_label]['employment_total'] += (
                legacy_zero_int(report, getattr(report, 'staff_kenyan_current', None))
                + legacy_zero_int(report, getattr(report, 'staff_foreign_current', None))
                + legacy_zero_int(report, getattr(report, 'staff_other_kenyan_current', None))
                + legacy_zero_int(report, getattr(report, 'staff_other_foreign_current', None))
                + legacy_zero_int(report, getattr(report, 'volunteers_kenyan_current', None))
                + legacy_zero_int(report, getattr(report, 'volunteers_foreign_current', None))
            )
            org_key, org_name = data_interp_normalize_org(report)
            org_display_all_years[org_key] = org_name
            org_year_funding[org_key][reporting_label] += data_interp_report_funding_total(report)
            for donation in getattr(report, 'donations', []):
                yearly_donor_category_totals[reporting_label][
                    data_interp_raw_donor_category(getattr(donation, 'category', None))
                ] += float_or_zero(getattr(donation, 'amount', None))

        reg_date = getattr(report, 'date_of_registration', None) or getattr(report, 'pbo_registration_date', None)
        if reg_date:
            reg_label = f"{fiscal_start_year(reg_date)}/{str(fiscal_start_year(reg_date) + 1)[-2:]}"
            yearly_metrics[reg_label]['new_registrations'] += 1

    sector_spend = defaultdict(float)
    sector_project_counts = Counter()
    county_operations_orgs = defaultdict(set)
    county_project_orgs = defaultdict(set)
    county_project_counts = Counter()
    county_spend = defaultdict(float)
    donor_type_amounts = defaultdict(float)
    continent_amounts = defaultdict(float)
    donor_category_amounts = defaultdict(float)
    donor_category_row_counts = Counter()
    donor_category_orgs = defaultdict(set)
    funding_country_amounts = defaultdict(float)
    funding_country_row_counts = Counter()
    funding_country_orgs = defaultdict(set)
    org_project_spend = defaultdict(float)
    org_scope = {}
    org_display = {}
    org_beneficiaries = defaultdict(int)
    sector_org_spend = defaultdict(lambda: defaultdict(float))
    beta_spend = defaultdict(float)
    beta_org_spend = defaultdict(lambda: defaultdict(float))
    partner_type_counts = Counter()
    collaboration_nature_counts = Counter()
    collaboration_partner_nature_counts = defaultdict(Counter)
    expense_line_amounts = defaultdict(float)
    expense_line_kenya = defaultdict(float)
    expense_line_other = defaultdict(float)
    sustainability_rows = []
    org_name_variants = defaultdict(set)
    funding_reconciliation = {
        'comparable': 0,
        'matched': 0,
        'mismatched': 0,
        'missing_support': 0,
        'examples': [],
    }
    project_reconciliation = {
        'comparable': 0,
        'matched': 0,
        'mismatched': 0,
        'fallback_only': 0,
        'examples': [],
    }
    completeness_counts = Counter()

    newly_registered_count = 0
    filed_reports_count = 0
    audited_reports_count = 0
    funding_total = 0.0
    project_spend_total = 0.0
    project_spend_kenya = 0.0
    project_spend_other = 0.0
    payments_total = 0.0
    staff_kenyan_total = 0
    staff_foreign_total = 0
    staff_other_kenyan_total = 0
    staff_other_foreign_total = 0
    volunteers_kenyan_total = 0
    volunteers_foreign_total = 0

    for report in selected_reports:
        org_key, org_name = data_interp_normalize_org(report)
        org_display[org_key] = org_name
        org_scope[org_key] = display_scope(getattr(report, 'scope', None))
        org_name_variants[org_key].add(org_name)

        reg_date = getattr(report, 'date_of_registration', None) or getattr(report, 'pbo_registration_date', None)
        if reg_date and (
            selected_fiscal_year == 'all'
            or (range_start and range_end and range_start <= reg_date <= range_end)
        ):
            newly_registered_count += 1

        if data_interp_is_filed(report):
            filed_reports_count += 1
        if data_interp_is_audited(report):
            audited_reports_count += 1

        report_funding_total = data_interp_report_funding_total(report)
        funding_total += report_funding_total

        declared_counties = []
        for county in data_interp_split_values(getattr(report, 'counties', None)):
            normalized_county = data_interp_normalize_county(county)
            if normalized_county:
                declared_counties.append(normalized_county)
                county_operations_orgs[normalized_county].add(org_key)

        report_donation_total = 0.0
        for donation in getattr(report, 'donations', []):
            amount = legacy_zero_float(report, getattr(donation, 'amount', None))
            report_donation_total += amount
            raw_donor_category = data_interp_raw_donor_category(getattr(donation, 'category', None))
            normalized_donor_type = data_interp_normalize_donor_type(getattr(donation, 'category', None))
            donor_category_amounts[raw_donor_category] += amount
            donor_category_row_counts[raw_donor_category] += 1
            donor_category_orgs[raw_donor_category].add(org_key)
            donor_type_amounts[normalized_donor_type] += amount
            donor_country = data_interp_display_country(getattr(donation, 'country', None))
            funding_country_amounts[donor_country] += amount
            funding_country_row_counts[donor_country] += 1
            funding_country_orgs[donor_country].add(org_key)
            continent_amounts[data_interp_country_to_continent(getattr(donation, 'country', None))] += amount
            if amount > 0 and not data_interp_clean_text(getattr(donation, 'country', None)):
                completeness_counts['donor_country'] += 1
            if amount > 0 and normalized_donor_type == 'Unspecified':
                completeness_counts['donor_type'] += 1

        report_grant_total = 0.0
        for grant in getattr(report, 'grants', []):
            amount = legacy_zero_float(report, getattr(grant, 'amount', None))
            report_grant_total += amount
            donor_type_amounts['Grant Funding (Unclassified)'] += amount
            continent_amounts[data_interp_country_to_continent(getattr(grant, 'country', None))] += amount
            if amount > 0 and not data_interp_clean_text(getattr(grant, 'country', None)):
                completeness_counts['donor_country'] += 1

        iga_total = sum(legacy_zero_float(report, getattr(row, 'amount', None)) for row in getattr(report, 'igas', []))
        if iga_total:
            donor_type_amounts['Self-generated Income'] += iga_total
            continent_amounts['Kenya / Internal'] += iga_total

        row_funding_total = report_donation_total + report_grant_total + iga_total
        stored_income_total = legacy_zero_float(report, getattr(report, 'income_b2_total', None))
        stored_receipts_total = legacy_zero_float(report, getattr(report, 'receipts_total', None))
        opening_balance = legacy_zero_float(report, getattr(report, 'cash_balance_previous_year', None))
        funding_comparable = False
        funding_mismatch_parts = []
        if row_funding_total > 0 and stored_income_total > 0:
            funding_comparable = True
            if abs(row_funding_total - stored_income_total) > qa_tolerance:
                funding_mismatch_parts.append(
                    f"rows {data_interp_format_money(row_funding_total)} vs income_b2_total {data_interp_format_money(stored_income_total)}"
                )
        if row_funding_total > 0 and stored_receipts_total > 0:
            funding_comparable = True
            expected_receipts_total = row_funding_total + opening_balance
            if abs(expected_receipts_total - stored_receipts_total) > qa_tolerance:
                funding_mismatch_parts.append(
                    f"expected receipts {data_interp_format_money(expected_receipts_total)} vs receipts_total {data_interp_format_money(stored_receipts_total)}"
                )
        if funding_comparable:
            funding_reconciliation['comparable'] += 1
            if funding_mismatch_parts:
                funding_reconciliation['mismatched'] += 1
                if len(funding_reconciliation['examples']) < 3:
                    funding_reconciliation['examples'].append(f"{org_name}: {'; '.join(funding_mismatch_parts)}")
            else:
                funding_reconciliation['matched'] += 1
        elif row_funding_total > 0 or stored_income_total > 0 or stored_receipts_total > 0:
            funding_reconciliation['missing_support'] += 1

        report_project_spend = 0.0
        report_project_beneficiaries = 0
        for project in getattr(report, 'project_implementations', []):
            raw_sector = getattr(project, 'sector', None)
            sector = data_interp_normalize_sector(raw_sector)
            spend = resolve_project_spending_amount(project, report=report)
            beneficiaries = legacy_zero_int(report, getattr(project, 'beneficiaries_no', None))
            sector_spend[sector] += spend
            sector_project_counts[sector] += 1
            org_project_spend[org_key] += spend
            org_beneficiaries[org_key] += beneficiaries
            sector_org_spend[sector][org_key] += spend

            beta_pillar = data_interp_beta_pillar(sector)
            beta_spend[beta_pillar] += spend
            beta_org_spend[beta_pillar][org_key] += spend

            raw_project_counties = data_interp_split_values(getattr(project, 'county', None))
            county_values = [
                data_interp_normalize_county(county)
                for county in raw_project_counties
            ]
            county_values = [county for county in county_values if county]
            if not county_values and len(declared_counties) == 1:
                county_values = list(declared_counties)
            if county_values:
                allocated_spend = spend / len(county_values) if county_values else spend
                for county in county_values:
                    county_project_counts[county] += 1
                    county_spend[county] += allocated_spend
                    county_project_orgs[county].add(org_key)

            if not data_interp_clean_text(raw_sector):
                completeness_counts['project_sector'] += 1
            if not raw_project_counties and not declared_counties:
                completeness_counts['project_county'] += 1

            raw_declared_spend = parse_float(legacy_zero_output_value(report, getattr(project, 'spending_per_county', None)))
            kenya_amount = legacy_zero_float(report, getattr(project, 'amount_spent_kenya', None))
            other_amount = legacy_zero_float(report, getattr(project, 'amount_spent_other', None))
            breakdown_total = kenya_amount + other_amount
            if raw_declared_spend not in (None, 0) and breakdown_total > 0:
                project_reconciliation['comparable'] += 1
                if abs(float(raw_declared_spend) - breakdown_total) > qa_tolerance:
                    project_reconciliation['mismatched'] += 1
                    if len(project_reconciliation['examples']) < 3:
                        project_reconciliation['examples'].append(
                            f"{org_name} / {sector}: spending_per_county {data_interp_format_money(raw_declared_spend)} vs Kenya+Other {data_interp_format_money(breakdown_total)}"
                        )
                else:
                    project_reconciliation['matched'] += 1
            elif spend > 0 or breakdown_total > 0:
                project_reconciliation['fallback_only'] += 1
            if kenya_amount == 0 and other_amount == 0 and spend > 0:
                kenya_amount = spend
            project_spend_kenya += kenya_amount
            project_spend_other += other_amount
            report_project_spend += spend
            report_project_beneficiaries += beneficiaries

        project_spend_total += report_project_spend

        for collaboration in getattr(report, 'collaborations', []):
            partner_type = data_interp_normalize_partner_type(getattr(collaboration, 'partner_type', None))
            partner_type_counts[partner_type] += 1
            for nature_label, field_name in DATA_INTERP_COLLAB_NATURE_FIELDS:
                if data_interp_collaboration_flag_checked(getattr(collaboration, field_name, None)):
                    collaboration_nature_counts[nature_label] += 1
                    collaboration_partner_nature_counts[partner_type][nature_label] += 1

        report_local_funding = data_interp_report_local_funding_total(report)
        report_payments_total = 0.0
        for payment in getattr(report, 'payments', []):
            kenya_amount = legacy_zero_float(report, getattr(payment, 'kenya_amount', None))
            other_amount = legacy_zero_float(report, getattr(payment, 'other_amount', None))
            line = data_interp_expense_line(getattr(payment, 'description', None))
            total_amount = kenya_amount + other_amount
            if total_amount > 0 and line == 'Other Expenditure':
                completeness_counts['expense_line'] += 1
            expense_line_amounts[line] += total_amount
            expense_line_kenya[line] += kenya_amount
            expense_line_other[line] += other_amount
            report_payments_total += total_amount
        payments_total += report_payments_total

        staff_kenyan = legacy_zero_int(report, getattr(report, 'staff_kenyan_current', None))
        staff_foreign = legacy_zero_int(report, getattr(report, 'staff_foreign_current', None))
        staff_other_kenyan = legacy_zero_int(report, getattr(report, 'staff_other_kenyan_current', None))
        staff_other_foreign = legacy_zero_int(report, getattr(report, 'staff_other_foreign_current', None))
        volunteers_kenyan = legacy_zero_int(report, getattr(report, 'volunteers_kenyan_current', None))
        volunteers_foreign = legacy_zero_int(report, getattr(report, 'volunteers_foreign_current', None))

        staff_kenyan_total += staff_kenyan
        staff_foreign_total += staff_foreign
        staff_other_kenyan_total += staff_other_kenyan
        staff_other_foreign_total += staff_other_foreign
        volunteers_kenyan_total += volunteers_kenyan
        volunteers_foreign_total += volunteers_foreign

        governance_completeness = 0
        if getattr(report, 'officials', None):
            governance_completeness += 8
        if legacy_zero_int(report, getattr(report, 'number_of_board_meetings', None)) > 0 or legacy_zero_int(report, getattr(report, 'membership_number_of_board_meetings', None)) > 0:
            governance_completeness += 7

        funding_base = max(report_funding_total, 1.0)
        local_share = min(report_local_funding / funding_base, 1.0)
        closing_balance = legacy_zero_float(report, getattr(report, 'cash_bank_balance', None))
        liquidity_ratio = min(closing_balance / funding_base, 1.0)
        collaboration_score = 10 if getattr(report, 'collaborations', None) else 0
        staffing_score = 10 if (staff_kenyan + staff_foreign + staff_other_kenyan + staff_other_foreign + volunteers_kenyan + volunteers_foreign) > 0 else 0
        audited_score = 25 if data_interp_is_audited(report) else 0
        local_score = local_share * 20
        liquidity_score = liquidity_ratio * 20
        sustainability_score = round(audited_score + governance_completeness + local_score + liquidity_score + collaboration_score + staffing_score, 2)
        if sustainability_score >= 70:
            sustainability_band = 'Strong'
        elif sustainability_score >= 45:
            sustainability_band = 'Emerging'
        else:
            sustainability_band = 'Fragile'

        sustainability_rows.append({
            'org_key': org_key,
            'org_name': org_name,
            'score': sustainability_score,
            'band': sustainability_band,
            'local_share': local_share,
        })

    total_employment = (
        staff_kenyan_total
        + staff_foreign_total
        + staff_other_kenyan_total
        + staff_other_foreign_total
        + volunteers_kenyan_total
        + volunteers_foreign_total
    )
    total_volunteers = volunteers_kenyan_total + volunteers_foreign_total

    yearly_labels = data_interp_sorted_year_labels(yearly_metrics.keys())
    yearly_rows = []
    for label in yearly_labels:
        metrics = yearly_metrics[label]
        yearly_rows.append({
            'Fiscal Year': label,
            'New Registrations': f"{metrics['new_registrations']:,}",
            'Filed Reports': f"{metrics['filed_reports']:,}",
            'Audited Reports': f"{metrics['audited_reports']:,}",
            'Funding Captured': data_interp_format_money(metrics['funding_total']),
            'Project Spend': data_interp_format_money(metrics['project_spend']),
        })

    donor_category_series = {
        label: [yearly_donor_category_totals[year].get(label, 0.0) for year in yearly_labels]
        for label in DATA_INTERP_DONOR_CATEGORY_OPTIONS
    }
    donor_category_series = {
        label: values
        for label, values in donor_category_series.items()
        if any(value > 0 for value in values) or label == 'Other (Specified)'
    }

    yearly_variance_labels = []
    yearly_variance_values = []
    previous_metrics = None
    for label in yearly_labels:
        metrics = yearly_metrics[label]
        if previous_metrics is not None:
            yearly_variance_labels.append(label)
            yearly_variance_values.append(metrics['funding_total'] - previous_metrics['funding_total'])
        previous_metrics = metrics

    if selected_fiscal_year == 'all':
        comparison_year_label = yearly_labels[-1] if yearly_labels else None
    else:
        comparison_year_label = f"{range_start.year}/{str(range_end.year)[-2:]}" if range_start and range_end else None
    previous_year_label = None
    if comparison_year_label in yearly_labels:
        comparison_index = yearly_labels.index(comparison_year_label)
        if comparison_index > 0:
            previous_year_label = yearly_labels[comparison_index - 1]

    growth_contributor_rows = []
    if comparison_year_label:
        for org_key, funding_by_year in org_year_funding.items():
            current_amount = funding_by_year.get(comparison_year_label, 0.0)
            previous_amount = funding_by_year.get(previous_year_label, 0.0) if previous_year_label else 0.0
            variance_amount = current_amount - previous_amount
            if current_amount == 0 and previous_amount == 0:
                continue
            growth_contributor_rows.append({
                'Organisation': org_display_all_years.get(org_key, org_key),
                'Current Year': data_interp_format_money(current_amount),
                'Previous Year': data_interp_format_money(previous_amount),
                'Variance': data_interp_format_money(variance_amount),
                'variance_value': variance_amount,
            })
        growth_contributor_rows = sorted(
            growth_contributor_rows,
            key=lambda row: row['variance_value'],
            reverse=True,
        )
        for row in growth_contributor_rows:
            row.pop('variance_value', None)

    top_sector_labels = [label for label, _ in sorted(sector_spend.items(), key=lambda item: item[1], reverse=True)[:4]]
    top_donor_labels = [label for label, _ in sorted(donor_category_amounts.items(), key=lambda item: item[1], reverse=True)[:4]]

    sorted_counties = sorted(
        {
            county: {
                'organisations': len(county_operations_orgs.get(county, set()) | county_project_orgs.get(county, set())),
                'projects': county_project_counts.get(county, 0),
                'spend': county_spend.get(county, 0.0),
            }
            for county in set(county_operations_orgs.keys()) | set(county_project_counts.keys())
        }.items(),
        key=lambda item: (item[1]['projects'], item[1]['organisations'], item[1]['spend']),
        reverse=True,
    )

    leading_orgs = sorted(org_project_spend.items(), key=lambda item: item[1], reverse=True)
    sector_leaders = []
    for sector, orgs in sector_org_spend.items():
        leader_key, leader_value = max(orgs.items(), key=lambda item: item[1])
        sector_leaders.append((sector, leader_key, leader_value))
    sector_leaders = sorted(sector_leaders, key=lambda item: item[2], reverse=True)

    sustainability_rows_sorted = sorted(sustainability_rows, key=lambda row: row['score'], reverse=True)
    sustainability_band_counts = Counter(row['band'] for row in sustainability_rows)
    qa_checks = []

    if selected_fiscal_year == 'all':
        scope_status = 'Ready'
        scope_summary = (
            f"This run intentionally combines all accessible fiscal years across {len(selected_reports):,} report"
            f"{'s' if len(selected_reports) != 1 else ''}."
        )
        scope_metrics = [f"Reports analysed across all years: {len(selected_reports):,}"]
    else:
        scope_mismatch_count = sum(
            1 for report in selected_reports if not report_matches_fiscal_year(report, range_start, range_end)
        )
        scope_status = 'Reconciled' if scope_mismatch_count == 0 else 'Mismatch Found'
        scope_summary = (
            f"All {len(selected_reports):,} selected reports fall inside {focus_label}."
            if scope_mismatch_count == 0
            else f"{scope_mismatch_count:,} selected report{'s' if scope_mismatch_count != 1 else ''} fell outside {focus_label}."
        )
        scope_metrics = [
            f"Reports in selected fiscal window: {len(selected_reports):,}",
            f"Accessible reports excluded from this run: {len(reports) - len(selected_reports):,}",
        ]
    qa_checks.append({
        'title': 'Fiscal-Year Scope Filter',
        'status': scope_status,
        'summary': scope_summary,
        'metrics': scope_metrics,
        'examples': [],
    })

    if funding_reconciliation['mismatched'] > 0:
        funding_status = 'Mismatch Found'
        funding_summary = (
            f"{funding_reconciliation['mismatched']:,} of {funding_reconciliation['comparable']:,} comparable report"
            f"{'s' if funding_reconciliation['comparable'] != 1 else ''} did not reconcile against stored funding headers."
        )
    elif funding_reconciliation['comparable'] > 0 and funding_reconciliation['missing_support'] == 0:
        funding_status = 'Reconciled'
        funding_summary = (
            f"All {funding_reconciliation['comparable']:,} comparable report"
            f"{'s' if funding_reconciliation['comparable'] != 1 else ''} reconciled against income and receipt totals."
        )
    elif funding_reconciliation['comparable'] > 0:
        funding_status = 'Needs Review'
        funding_summary = (
            f"{funding_reconciliation['matched']:,} comparable report"
            f"{'s' if funding_reconciliation['matched'] != 1 else ''} reconciled, but some reports still lack enough header detail for a full check."
        )
    else:
        funding_status = 'No Comparison Yet'
        funding_summary = 'No report had both row-level funding detail and stored header totals for a numeric reconciliation check.'
    qa_checks.append({
        'title': 'Funding Reconciliation',
        'status': funding_status,
        'summary': funding_summary,
        'metrics': [
            f"Comparable reports: {funding_reconciliation['comparable']:,}",
            f"Matched reports: {funding_reconciliation['matched']:,}",
            f"Mismatched reports: {funding_reconciliation['mismatched']:,}",
            f"Reports without enough comparison detail: {funding_reconciliation['missing_support']:,}",
        ],
        'examples': funding_reconciliation['examples'],
    })

    if project_reconciliation['mismatched'] > 0:
        project_status = 'Mismatch Found'
        project_summary = (
            f"{project_reconciliation['mismatched']:,} of {project_reconciliation['comparable']:,} comparable project row"
            f"{'s' if project_reconciliation['comparable'] != 1 else ''} did not reconcile between declared spend and Kenya/Other splits."
        )
    elif project_reconciliation['comparable'] > 0 and project_reconciliation['fallback_only'] == 0:
        project_status = 'Reconciled'
        project_summary = (
            f"All {project_reconciliation['comparable']:,} comparable project row"
            f"{'s' if project_reconciliation['comparable'] != 1 else ''} reconciled cleanly."
        )
    elif project_reconciliation['comparable'] > 0:
        project_status = 'Needs Review'
        project_summary = (
            f"Comparable project rows reconciled, but {project_reconciliation['fallback_only']:,} row"
            f"{'s' if project_reconciliation['fallback_only'] != 1 else ''} still rely on one-sided project-spend values."
        )
    else:
        project_status = 'No Comparison Yet'
        project_summary = 'No project row had both `spending_per_county` and Kenya/Other amounts populated for direct reconciliation.'
    qa_checks.append({
        'title': 'Project Spend Row Reconciliation',
        'status': project_status,
        'summary': project_summary,
        'metrics': [
            f"Comparable project rows: {project_reconciliation['comparable']:,}",
            f"Matched project rows: {project_reconciliation['matched']:,}",
            f"Mismatched project rows: {project_reconciliation['mismatched']:,}",
            f"Fallback-only project rows: {project_reconciliation['fallback_only']:,}",
        ],
        'examples': project_reconciliation['examples'],
    })

    completeness_total = sum(completeness_counts.values())
    if completeness_total == 0:
        completeness_status = 'Reconciled'
        completeness_summary = 'No missing donor geography, project taxonomy, or unmapped spend-line issues were detected in this run.'
    else:
        completeness_status = 'Needs Review'
        completeness_summary = (
            f"{completeness_total:,} classification or completeness issue"
            f"{'s' if completeness_total != 1 else ''} still need attention before publication."
        )
    qa_checks.append({
        'title': 'Classification and Completeness',
        'status': completeness_status,
        'summary': completeness_summary,
        'metrics': [
            f"Missing donor country values: {completeness_counts.get('donor_country', 0):,}",
            f"Unspecified donor categories: {completeness_counts.get('donor_type', 0):,}",
            f"Projects without sector values: {completeness_counts.get('project_sector', 0):,}",
            f"Projects without any county geography: {completeness_counts.get('project_county', 0):,}",
            f"Payment rows still mapped to Other Expenditure: {completeness_counts.get('expense_line', 0):,}",
        ],
        'examples': [],
    })

    org_variant_examples = []
    org_variant_count = 0
    for names in org_name_variants.values():
        if len(names) > 1:
            org_variant_count += 1
            if len(org_variant_examples) < 3:
                org_variant_examples.append(' / '.join(sorted(names)))
    if org_variant_count == 0:
        org_status = 'Reconciled'
        org_summary = 'No split organisation spellings were detected in the selected reporting run.'
    else:
        org_status = 'Needs Review'
        org_summary = (
            f"{org_variant_count:,} normalized organisation profile"
            f"{'s' if org_variant_count != 1 else ''} contain multiple display-name spellings that could split rankings."
        )
    qa_checks.append({
        'title': 'Organisation Ranking Integrity',
        'status': org_status,
        'summary': org_summary,
        'metrics': [
            f"Profiles analysed for ranking: {len(org_name_variants):,}",
            f"Profiles with multiple spellings: {org_variant_count:,}",
        ],
        'examples': org_variant_examples,
    })

    for check in qa_checks:
        check['status_label'] = check['status']
        check['status_class'] = data_interp_qa_status_class(check['status'])

    qa_summary = data_interp_build_qa_summary(qa_checks)

    topic_results = {}
    top_sector = top_sector_labels[0] if top_sector_labels else 'No sector data yet'
    top_county = sorted_counties[0][0] if sorted_counties else 'No county data yet'
    top_org_name = org_display.get(leading_orgs[0][0], 'No organisation data yet') if leading_orgs else 'No organisation data yet'
    top_donor_type = top_donor_labels[0] if top_donor_labels else 'No donor categories yet'
    top_beta = max(beta_spend.items(), key=lambda item: item[1])[0] if beta_spend else 'No mapped BETA pillar yet'
    top_collab_type = partner_type_counts.most_common(1)[0][0] if partner_type_counts else 'No collaboration rows yet'
    top_collab_nature = collaboration_nature_counts.most_common(1)[0][0] if collaboration_nature_counts else 'No collaboration nature yet'
    avg_sustainability = round(sum(row['score'] for row in sustainability_rows) / len(sustainability_rows), 2) if sustainability_rows else 0.0

    topic_results['2.0'] = {'headline': 'Current growth signal', 'detail': f"{newly_registered_count:,} newly registered organisations, {filed_reports_count:,} filed returns, and {data_interp_format_money(funding_total)} in captured funding for {focus_label}."}
    topic_results['2.1'] = {'headline': 'Register activity visible now', 'detail': f"The current schema shows {newly_registered_count:,} registrations in window and {filed_reports_count:,} active filers moving through workflow for {focus_label}."}
    topic_results['2.2'] = {'headline': 'Annual reports status', 'detail': f"{filed_reports_count:,} filed returns were analysed and {audited_reports_count:,} of them were marked audited in {focus_label}."}
    topic_results['2.3'] = {'headline': 'Funding trend signal', 'detail': f"Captured funding totals {data_interp_format_money(funding_total)} for {focus_label}, with trendlines built from all accessible fiscal years."}
    topic_results['3.0'] = {'headline': 'National contribution summary', 'detail': f"{data_interp_format_money(project_spend_total)} in project spend, {len(sorted_counties):,} counties active, and {total_employment:,} people captured in the workforce footprint."}
    topic_results['3.1'] = {'headline': 'Preferred sector', 'detail': f"{top_sector} is currently the strongest sector signal among the available project evidence for the selected reporting window."}
    topic_results['3.2'] = {'headline': 'Programme utilisation', 'detail': f"Projects account for {data_interp_format_money(project_spend_total)}, with {data_interp_format_money(project_spend_kenya)} linked to Kenya and {data_interp_format_money(project_spend_other)} linked to other countries."}
    topic_results['3.3'] = {'headline': 'County footprint', 'detail': f"{len(sorted_counties):,} counties have visible operational or project activity in {focus_label}, led by {top_county}."}
    topic_results['3.3.1'] = {'headline': 'Newly registered county preference', 'detail': f"{top_county} is currently the strongest county signal among organisations whose registration dates fall inside {focus_label}."}
    topic_results['3.3.2'] = {'headline': 'Project intensity by county', 'detail': f"{top_county} currently leads the selected data by project activity and spend."}
    topic_results['3.4'] = {'headline': 'Leading NGO by project spend', 'detail': f"{top_org_name} currently leads the selected dataset on declared project spend."}
    topic_results['3.5'] = {'headline': 'Leading sector player', 'detail': f"The strongest current sector-player signal is {top_sector} led by {top_org_name if leading_orgs else 'the top available organisation'}."}
    topic_results['3.6'] = {'headline': 'Top BETA pillar', 'detail': f"{top_beta} is the leading BETA-aligned pillar in the current mapping for {focus_label}."}
    topic_results['3.7'] = {'headline': 'Employment footprint', 'detail': f"{total_employment:,} staff, volunteers, and interns are visible in the selected returns, with volunteers accounting for {data_interp_percent(total_volunteers, total_employment):,.2f}%."}
    topic_results['3.8'] = {'headline': 'Collaboration pattern', 'detail': f"{top_collab_type} is the most common partner class, while {top_collab_nature} is the strongest collaboration nature signal."}
    topic_results['4.1'] = {'headline': 'Funds received', 'detail': f"{top_donor_type} is the leading donor category visible in the current funding mix for {focus_label}."}
    topic_results['4.2'] = {'headline': 'Utilisation of funds', 'detail': f"Payment analysis currently highlights {max(expense_line_amounts.items(), key=lambda item: item[1])[0] if expense_line_amounts else 'no mapped expenditure line yet'} as the largest visible spend line."}
    topic_results['4.3'] = {'headline': 'Sector sustainability', 'detail': f"The average sustainability score is {avg_sustainability:,.2f}, with {sustainability_band_counts.get('Strong', 0):,} organisations currently landing in the Strong band."}

    analysis_cards = [
        {'label': 'Reports Analysed', 'value': f"{len(selected_reports):,}", 'sub': focus_label},
        {'label': 'New Registrations', 'value': f"{newly_registered_count:,}", 'sub': 'Within selected fiscal window'},
        {'label': 'Funding Captured', 'value': data_interp_format_money(funding_total), 'sub': f"{len(top_donor_labels):,} strong donor categories surfaced"},
        {'label': 'Project Spend', 'value': data_interp_format_money(project_spend_total), 'sub': f"{len(top_sector_labels):,} leading sectors visible"},
        {'label': 'Employment Footprint', 'value': f"{total_employment:,}", 'sub': f"{total_volunteers:,} volunteers and interns captured"},
        {'label': 'Average Sustainability', 'value': f"{avg_sustainability:,.2f}", 'sub': f"{sustainability_band_counts.get('Strong', 0):,} strong profiles in current run"},
    ]

    analysis_summary_lines = [
        f"{len(selected_reports):,} reports were analysed for {focus_label} under {scope_label.lower()}.",
        f"Project spending reaches {data_interp_format_money(project_spend_total)} across {len(top_sector_labels) or 0:,} leading normalized sectors, with {top_sector} currently dominant.",
        f"Funding captured totals {data_interp_format_money(funding_total)}, led by {top_donor_type}, while {top_county} stands out in county-level activity.",
        f"The workforce footprint currently covers {total_employment:,} people, and the collaboration profile is led by {top_collab_type}.",
    ]

    analysis_sections = []
    if go is not None:
        growth_charts = [
            {
                'title': 'Growth Trend by Fiscal Year',
                'subtitle': 'New registrations, filed returns, and funding trendlines across all accessible years.',
                'html': data_interp_build_line_chart(
                    go,
                    yearly_labels,
                    [
                        {'label': 'New Registrations', 'values': [yearly_metrics[label]['new_registrations'] for label in yearly_labels], 'color': '#0e8b7d'},
                        {'label': 'Filed Reports', 'values': [yearly_metrics[label]['filed_reports'] for label in yearly_labels], 'color': '#103a7d'},
                    ],
                    'Registrations and Filed Returns',
                ),
            },
            {
                'title': 'Funding Trend by Donor Category',
                'subtitle': 'Uses the donor-table `category` field and surfaces the donor dropdown options, including Other (Specified), in a Plotly selector.',
                'html': data_interp_build_dropdown_line_chart(
                    go,
                    yearly_labels,
                    donor_category_series,
                    'Funding Trend by Donor Category',
                ),
            },
            {
                'title': 'Year-on-Year Funding Variance',
                'subtitle': 'Shows how total captured funding changes from one fiscal year to the next.',
                'html': data_interp_build_bar_chart(
                    go,
                    yearly_variance_labels,
                    yearly_variance_values,
                    'Year-on-Year Funding Variance',
                    color='#d18a00',
                    money_axis=True,
                    show_text=False,
                ),
            },
            {
                'title': 'Audited vs Filed Returns',
                'subtitle': 'Audited-account coverage against filed-return volumes.',
                'html': data_interp_build_bar_chart(
                    go,
                    yearly_labels,
                    [yearly_metrics[label]['audited_reports'] for label in yearly_labels],
                    'Audited Returns by Fiscal Year',
                    color='#3a6ab3',
                ),
            },
        ]
        analysis_sections.append({
            'title': 'Growth and Reporting',
            'subtitle': 'Live Chapter Two reporting generated when you run the analysis.',
            'charts': [chart for chart in growth_charts if chart.get('html')],
            'tables': [
                {
                    'title': 'Fiscal-Year Growth Summary',
                    'columns': ['Fiscal Year', 'New Registrations', 'Filed Reports', 'Audited Reports', 'Funding Captured', 'Project Spend'],
                    'rows': yearly_rows,
                },
                {
                    'title': f"Top Growth Contributors{f' ({comparison_year_label} vs {previous_year_label})' if comparison_year_label and previous_year_label else ''}",
                    'columns': ['Organisation', 'Current Year', 'Previous Year', 'Variance'],
                    'rows': growth_contributor_rows,
                },
            ],
        })

        collaboration_partner_rows = [
            (label, partner_type_counts.get(label, 0))
            for label in DATA_INTERP_COLLAB_PARTNER_OPTIONS
        ]
        collaboration_nature_rows = [
            (label, collaboration_nature_counts.get(label, 0))
            for label, _ in DATA_INTERP_COLLAB_NATURE_FIELDS
        ]
        collaboration_heatmap_rows = [
            label
            for label, row_count in collaboration_partner_rows
            if row_count > 0 or any(collaboration_partner_nature_counts.get(label, Counter()).values())
        ]

        contribution_charts = [
            {
                'title': 'Top Sectors by Project Spend',
                'subtitle': 'Normalized project sectors ranked by declared spend in the selected analysis window.',
                'html': data_interp_build_bar_chart(
                    go,
                    [item[0] for item in sorted(sector_spend.items(), key=lambda item: item[1], reverse=True)[:10]][::-1],
                    [item[1] for item in sorted(sector_spend.items(), key=lambda item: item[1], reverse=True)[:10]][::-1],
                    'Project Spend by Sector',
                    color='#103a7d',
                    horizontal=True,
                    money_axis=True,
                ),
            },
            {
                'title': 'Top Counties by Project Activity',
                'subtitle': 'Counties ranked by project rows and blended project spend signals.',
                'html': data_interp_build_bar_chart(
                    go,
                    [item[0] for item in sorted_counties[:10]][::-1],
                    [item[1]['projects'] for item in sorted_counties[:10]][::-1],
                    'Project Activity by County',
                    color='#0e8b7d',
                    horizontal=True,
                ),
            },
            {
                'title': 'Leading NGOs by Project Spend',
                'subtitle': 'Top organisations in the selected analysis window.',
                'html': data_interp_build_bar_chart(
                    go,
                    [org_display[item[0]] for item in leading_orgs[:10]][::-1],
                    [item[1] for item in leading_orgs[:10]][::-1],
                    'Leading NGOs by Project Spend',
                    color='#d18a00',
                    horizontal=True,
                    money_axis=True,
                ),
            },
            {
                'title': 'Employment Composition',
                'subtitle': 'Current staffing and volunteer footprint visible in the selected returns.',
                'html': data_interp_build_bar_chart(
                    go,
                    ['Kenyan Staff', 'Foreign Staff', 'Other-Country Kenyan Staff', 'Other-Country Foreign Staff', 'Kenyan Volunteers', 'Foreign Volunteers'],
                    [staff_kenyan_total, staff_foreign_total, staff_other_kenyan_total, staff_other_foreign_total, volunteers_kenyan_total, volunteers_foreign_total],
                    'Employment and Volunteer Composition',
                    color='#3a6ab3',
                ),
            },
            {
                'title': 'BETA Pillar Spend',
                'subtitle': 'Default in-app mapping of project sectors into BETA-aligned pillars.',
                'html': data_interp_build_bar_chart(
                    go,
                    [item[0] for item in sorted(beta_spend.items(), key=lambda item: item[1], reverse=True)],
                    [item[1] for item in sorted(beta_spend.items(), key=lambda item: item[1], reverse=True)],
                    'BETA-aligned Spend',
                    color='#5f5ce1',
                    horizontal=True,
                    money_axis=True,
                ),
            },
            {
                'title': 'Collaboration Types',
                'subtitle': 'Uses the Question D3 partner-type dropdown options, with older NGO labels normalized into PBOs.',
                'html': data_interp_build_bar_chart(
                    go,
                    [item[0] for item in collaboration_partner_rows][::-1],
                    [item[1] for item in collaboration_partner_rows][::-1],
                    'Question D3 Partner Types',
                    color='#0b8793',
                    horizontal=True,
                    height=max(460, 28 * max(len(collaboration_partner_rows), 1)),
                    show_text=False,
                ),
                'card_class': 'is-wide',
                'min_width': 1120,
            },
            {
                'title': 'Nature of Collaboration',
                'subtitle': 'Counts each saved collaboration marker, whether it was stored as Yes/No or echoed back as the selected partner type.',
                'html': data_interp_build_donut_chart(
                    go,
                    [item[0] for item in collaboration_nature_rows if item[1] > 0],
                    [item[1] for item in collaboration_nature_rows if item[1] > 0],
                    'Nature of Collaboration',
                    colors=['#103a7d', '#3a6ab3', '#5d95cd', '#0e8b7d', '#4fb3a2', '#d18a00', '#f1b74d'],
                ),
            },
            {
                'title': 'Collaboration Matrix',
                'subtitle': 'Heatmap of D3 partner types against the collaboration-nature checkboxes saved in each report.',
                'html': data_interp_build_heatmap(
                    go,
                    [label for label, _ in DATA_INTERP_COLLAB_NATURE_FIELDS],
                    collaboration_heatmap_rows,
                    [
                        [collaboration_partner_nature_counts.get(partner_type, Counter()).get(label, 0) for label, _ in DATA_INTERP_COLLAB_NATURE_FIELDS]
                        for partner_type in collaboration_heatmap_rows
                    ],
                    'Partner Type by Collaboration Nature',
                    height=max(420, 42 * max(len(collaboration_heatmap_rows), 1)),
                ),
                'card_class': 'is-wide',
                'min_width': 1240,
            },
        ]
        top_county_rows = [
            {
                'County': county,
                'Organisations': f"{metrics['organisations']:,}",
                'Projects': f"{metrics['projects']:,}",
                'Project Spend': data_interp_format_money(metrics['spend']),
            }
            for county, metrics in sorted_counties
        ]
        sector_leader_rows = [
            {
                'Sector': sector,
                'Leading Organisation': org_display.get(org_key, org_key),
                'Project Spend': data_interp_format_money(amount),
            }
            for sector, org_key, amount in sector_leaders
        ]
        collaboration_table_rows = [
            {
                'Partner Type': label,
                'Collaboration Rows': f"{partner_type_counts.get(label, 0):,}",
                'Information Exchange': f"{collaboration_partner_nature_counts.get(label, Counter()).get('Information Exchange', 0):,}",
                'Technical Support To': f"{collaboration_partner_nature_counts.get(label, Counter()).get('Technical Support to Partner', 0):,}",
                'Technical Support From': f"{collaboration_partner_nature_counts.get(label, Counter()).get('Technical Support from Partner', 0):,}",
                'Funding To': f"{collaboration_partner_nature_counts.get(label, Counter()).get('Funding to Partner', 0):,}",
                'Funding From': f"{collaboration_partner_nature_counts.get(label, Counter()).get('Funding from Partner', 0):,}",
                'Equipment To': f"{collaboration_partner_nature_counts.get(label, Counter()).get('Equipment to Partner', 0):,}",
                'Equipment From': f"{collaboration_partner_nature_counts.get(label, Counter()).get('Equipment from Partner', 0):,}",
            }
            for label in DATA_INTERP_COLLAB_PARTNER_OPTIONS
        ]
        analysis_sections.append({
            'title': 'National Development Contribution',
            'subtitle': 'Live Chapter Three outputs from projects, counties, employment, collaborations, and BETA mappings.',
            'charts': [chart for chart in contribution_charts if chart.get('html')],
            'tables': [
                {
                    'title': 'Top County Summary',
                    'columns': ['County', 'Organisations', 'Projects', 'Project Spend'],
                    'rows': top_county_rows,
                },
                {
                    'title': 'Leading Sector Players',
                    'columns': ['Sector', 'Leading Organisation', 'Project Spend'],
                    'rows': sector_leader_rows,
                },
                {
                    'title': 'Collaboration Profile by Partner Type',
                    'subtitle': 'Question D3 partner-type rows with collaboration-nature counts derived from the saved checkbox fields.',
                    'columns': [
                        'Partner Type',
                        'Collaboration Rows',
                        'Information Exchange',
                        'Technical Support To',
                        'Technical Support From',
                        'Funding To',
                        'Funding From',
                        'Equipment To',
                        'Equipment From',
                    ],
                    'rows': collaboration_table_rows,
                },
            ],
        })

        donor_rows = [
            (label, donor_category_amounts.get(label, 0.0))
            for label in DATA_INTERP_DONOR_CATEGORY_OPTIONS
        ]
        donor_chart_rows = sorted(donor_rows, key=lambda item: (item[1], item[0]), reverse=True)
        country_rows = sorted(funding_country_amounts.items(), key=lambda item: item[1], reverse=True)
        finance_charts = [
            {
                'title': 'Funding by Donor Type',
                'subtitle': 'Uses the donor-table `category` column and displays the donor dropdown options as full analysis rows instead of collapsing them into a short summary.',
                'html': data_interp_build_bar_chart(
                    go,
                    [item[0] for item in donor_chart_rows][::-1],
                    [item[1] for item in donor_chart_rows][::-1],
                    'Funding by Donor Category',
                    color='#103a7d',
                    horizontal=True,
                    money_axis=True,
                    height=max(620, 28 * max(len(donor_chart_rows), 1)),
                    show_text=False,
                ),
                'card_class': 'is-wide',
            },
            {
                'title': 'Funding by Country',
                'subtitle': 'Uses the donor-table `country` column and keeps country-by-country rows visible in a wider statistical display.',
                'html': data_interp_build_bar_chart(
                    go,
                    [item[0] for item in country_rows][::-1],
                    [item[1] for item in country_rows][::-1],
                    'Funding by Country',
                    color='#3a6ab3',
                    horizontal=True,
                    money_axis=True,
                    height=max(540, 28 * max(len(country_rows), 1)),
                    show_text=False,
                ),
                'card_class': 'is-wide',
            },
            {
                'title': 'Expenditure Breakdown',
                'subtitle': 'Payments are mapped into standard expenditure lines during the run.',
                'html': data_interp_build_donut_chart(
                    go,
                    [item[0] for item in sorted(expense_line_amounts.items(), key=lambda item: item[1], reverse=True)],
                    [item[1] for item in sorted(expense_line_amounts.items(), key=lambda item: item[1], reverse=True)],
                    'Visible Expenditure Mix',
                    colors=['#103a7d', '#3a6ab3', '#7aa4df', '#d18a00', '#0e8b7d', '#8b5cf6'],
                ),
            },
            {
                'title': 'Sustainability Score Distribution',
                'subtitle': 'Default in-app sustainability weights applied to the selected reporting window.',
                'html': data_interp_build_bar_chart(
                    go,
                    ['Strong', 'Emerging', 'Fragile'],
                    [
                        sustainability_band_counts.get('Strong', 0),
                        sustainability_band_counts.get('Emerging', 0),
                        sustainability_band_counts.get('Fragile', 0),
                    ],
                    'Sustainability Bands',
                    color='#0e8b7d',
                ),
            },
        ]
        donation_total_for_shares = sum(amount for _, amount in donor_rows) or 0.0
        country_total_for_shares = sum(amount for _, amount in country_rows) or 0.0
        funding_source_rows = [
            {
                'Donor Category': label,
                'Donor Rows': f"{donor_category_row_counts.get(label, 0):,}",
                'Reporting Organisations': f"{len(donor_category_orgs.get(label, set())):,}",
                'Amount': data_interp_format_money(amount),
                'Share': f"{data_interp_percent(amount, donation_total_for_shares):,.2f}%",
            }
            for label, amount in donor_rows
        ]
        funding_country_table_rows = [
            {
                'Country': label,
                'Donor Rows': f"{funding_country_row_counts.get(label, 0):,}",
                'Reporting Organisations': f"{len(funding_country_orgs.get(label, set())):,}",
                'Amount': data_interp_format_money(amount),
                'Share': f"{data_interp_percent(amount, country_total_for_shares):,.2f}%",
            }
            for label, amount in country_rows
        ]
        sustainability_table_rows = [
            {
                'Organisation': row['org_name'],
                'Score': f"{row['score']:,.2f}",
                'Band': row['band'],
                'Local Funding Share': f"{row['local_share'] * 100:,.2f}%",
            }
            for row in sustainability_rows_sorted
        ]
        analysis_sections.append({
            'title': 'Funding, Utilisation, and Sustainability',
            'subtitle': 'Live Chapter Four reporting generated when you run the analysis.',
            'charts': [chart for chart in finance_charts if chart.get('html')],
            'tables': [
                {
                    'title': 'Funding by Donor Category',
                    'subtitle': 'One row per donor dropdown option from the donor table category field, with row counts, organisation counts, amounts, and shares.',
                    'columns': ['Donor Category', 'Donor Rows', 'Reporting Organisations', 'Amount', 'Share'],
                    'rows': funding_source_rows,
                },
                {
                    'title': 'Funding by Country',
                    'subtitle': 'One row per donor-table country value so country-by-country funding can be searched, sorted, and paged.',
                    'columns': ['Country', 'Donor Rows', 'Reporting Organisations', 'Amount', 'Share'],
                    'rows': funding_country_table_rows,
                },
                {
                    'title': 'Top Sustainability Profiles',
                    'subtitle': 'Full organisation-level sustainability table for the selected analysis run.',
                    'columns': ['Organisation', 'Score', 'Band', 'Local Funding Share'],
                    'rows': sustainability_table_rows,
                },
            ],
        })

    return {
        'analysis_run': True,
        'analysis_generated_at': analysis_generated_at,
        'analysis_focus_label': focus_label,
        'analysis_cards': analysis_cards,
        'analysis_summary_lines': analysis_summary_lines,
        'qa_summary': qa_summary,
        'qa_checks': qa_checks,
        'analysis_sections': analysis_sections,
        'top_sector_labels': top_sector_labels,
        'top_donor_labels': top_donor_labels,
        'topic_results': topic_results,
    }


def start_data_interpretation_analysis_job(user_id, selected_fiscal_year, scope_label):
    if not user_id:
        return False

    with DATA_INTERPRETATION_JOB_LOCK:
        current_snapshot = get_data_interpretation_job_snapshot_for_user(user_id)
        if current_snapshot.get('running'):
            return False

        user = db.session.get(User, user_id)
        store_data_interpretation_job_snapshot_for_user(
            user_id,
            {
                'running': True,
                'status': 'queued',
                'detail': 'Analysis has been queued and will start in the background.',
                'selected_fiscal_year': selected_fiscal_year,
                'scope_label': scope_label,
                'started_at': utc_now().isoformat(timespec='seconds'),
                'finished_at': None,
                'progress_percent': 2,
                'last_error': None,
                'analysis_error': None,
                'result_ready': False,
                'analysis_focus_label': None,
            },
            user=user,
        )

    worker = threading.Thread(
        target=_run_data_interpretation_analysis_job,
        kwargs={
            'user_id': user_id,
            'selected_fiscal_year': selected_fiscal_year,
            'scope_label': scope_label,
        },
        name=f'data-interpretation-worker-{user_id}',
        daemon=True,
    )
    worker.start()
    return True


def _run_data_interpretation_analysis_job(user_id, selected_fiscal_year, scope_label):
    with app.app_context():
        user = db.session.get(User, user_id)
        actor = user if user and getattr(user, 'role', None) == 'admin' else None

        try:
            store_data_interpretation_job_snapshot_for_user(
                user_id,
                {
                    'running': True,
                    'status': 'running',
                    'detail': 'Loading accessible reports for the selected fiscal year.',
                    'progress_percent': 20,
                    'last_error': None,
                    'analysis_error': None,
                },
                user=actor,
            )

            reports = data_interp_load_accessible_reports_for_user(user)

            store_data_interpretation_job_snapshot_for_user(
                user_id,
                {
                    'running': True,
                    'status': 'running',
                    'detail': 'Building charts, tables, and QA checks in the background.',
                    'progress_percent': 62,
                },
                user=actor,
            )

            analysis_result = build_data_interpretation_analysis(
                reports,
                selected_fiscal_year,
                scope_label=scope_label,
            )

            store_data_interpretation_job_snapshot_for_user(
                user_id,
                {
                    'running': True,
                    'status': 'running',
                    'detail': 'Saving the latest analysis output for this admin session.',
                    'progress_percent': 88,
                },
                user=actor,
            )

            store_data_interpretation_result_for_user(
                user_id,
                selected_fiscal_year,
                analysis_result,
                user=actor,
            )

            completed_message = (
                analysis_result.get('analysis_error')
                or f"Analysis completed for {analysis_result.get('analysis_focus_label') or selected_fiscal_year}."
            )
            store_data_interpretation_job_snapshot_for_user(
                user_id,
                {
                    'running': False,
                    'status': 'completed',
                    'detail': completed_message,
                    'finished_at': utc_now().isoformat(timespec='seconds'),
                    'progress_percent': 100,
                    'analysis_error': analysis_result.get('analysis_error'),
                    'result_ready': True,
                    'analysis_focus_label': analysis_result.get('analysis_focus_label'),
                },
                user=actor,
            )
        except Exception as exc:
            app.logger.exception('Data interpretation background analysis failed')
            store_data_interpretation_job_snapshot_for_user(
                user_id,
                {
                    'running': False,
                    'status': 'failed',
                    'detail': 'The background analysis run failed. Review the error and try again.',
                    'finished_at': utc_now().isoformat(timespec='seconds'),
                    'progress_percent': 100,
                    'last_error': str(exc),
                    'analysis_error': str(exc),
                    'result_ready': False,
                },
                user=actor,
            )
        finally:
            db.session.remove()


@app.route('/data-interpretations', methods=['GET', 'POST'])
@app.route('/datainterpretations', methods=['GET', 'POST'])
@require_sector_analytics
@admin_required
def data_interpretations():
    fiscal_year_options, selected_fiscal_year, scope_label = data_interp_get_fiscal_year_options()
    analysis_requested = request.method == 'POST' and request.form.get('run_analysis') == '1'

    if analysis_requested:
        started = start_data_interpretation_analysis_job(
            current_user.id,
            selected_fiscal_year,
            scope_label,
        )
        if started:
            flash('Analysis started in background. This page will refresh when the run is ready.', 'info')
        else:
            flash('An analysis run is already in progress. Watch the status panel below.', 'warning')
        return redirect(url_for('data_interpretations', fy=selected_fiscal_year))

    analysis_job = get_data_interpretation_job_snapshot_for_user(current_user.id)
    analysis_result = get_data_interpretation_result_for_user(
        current_user.id,
        selected_fiscal_year=selected_fiscal_year,
    )

    context = build_data_interpretation_context(
        scope_label=scope_label,
        fiscal_year_options=fiscal_year_options,
        selected_fiscal_year=selected_fiscal_year,
        analysis_result=analysis_result,
        analysis_requested=bool(analysis_result) or bool(analysis_job.get('running')),
    )
    context['analysis_job'] = analysis_job
    context['analysis_status_endpoint'] = url_for('data_interpretations_status')
    return render_template('datainterpretations.html', **context)


@app.route('/data-interpretations/status')
@app.route('/datainterpretations/status')
@require_sector_analytics
@admin_required
def data_interpretations_status():
    selected_fiscal_year = (request.args.get('fy') or '').strip() or None
    snapshot = get_data_interpretation_job_snapshot_for_user(current_user.id)
    result_payload = get_data_interpretation_result_payload_for_user(current_user.id)
    latest_result_fiscal_year = (result_payload.get('selected_fiscal_year') or '').strip() or None
    return jsonify({
        **snapshot,
        'selected_fiscal_year': snapshot.get('selected_fiscal_year') or latest_result_fiscal_year,
        'latest_result_fiscal_year': latest_result_fiscal_year,
        'result_ready_for_selected_fiscal_year': bool(
            latest_result_fiscal_year
            and latest_result_fiscal_year == selected_fiscal_year
            and get_data_interpretation_result_for_user(current_user.id, selected_fiscal_year=selected_fiscal_year)
        ),
        'view_url': url_for('data_interpretations', fy=(snapshot.get('selected_fiscal_year') or selected_fiscal_year or 'all')),
    })


@app.route('/reports')
@login_required
def reports_list():
    """List all submitted reports with pagination"""
    from sqlalchemy.orm import selectinload

    page = request.args.get('page', 1, type=int)
    per_page_options = [10, 25, 50, 100]
    requested_per_page = request.args.get('per_page', type=int)
    if requested_per_page in per_page_options:
        per_page = requested_per_page
        session['reports_per_page'] = per_page
    else:
        per_page = session.get('reports_per_page', 25)
        if per_page not in per_page_options:
            per_page = 25
    search_query = (request.args.get('q') or '').strip()
    requested_fiscal_year = (request.args.get('fy') or '').strip()
    requested_period_start = (request.args.get('period_start') or '').strip()
    requested_period_end = (request.args.get('period_end') or '').strip()

    today = utc_now().date()
    default_start_year, default_end_year = get_current_fiscal_year(today)
    period_bounds_query = db.session.query(
        func.min(func.coalesce(PBOReport.reporting_period_start, PBOReport.reporting_period_end)),
        func.max(func.coalesce(PBOReport.reporting_period_end, PBOReport.reporting_period_start)),
    )
    if not can_manage_all_records(current_user):
        period_bounds_query = period_bounds_query.filter(PBOReport.user_id == current_user.id)
    min_period_date, max_period_date = period_bounds_query.first() or (None, None)

    if min_period_date and max_period_date:
        start_year = fiscal_start_year(min_period_date)
        end_year = fiscal_start_year(max_period_date)
    else:
        start_year = default_start_year
        end_year = default_start_year

    option_start_year = min(start_year, 2020)
    option_end_year = max(end_year, default_start_year)
    fiscal_year_options = [
        f"{year}-{year + 1}" for year in range(option_start_year, option_end_year + 1)
    ]
    selected_start_year, selected_end_year = parse_fiscal_year(
        requested_fiscal_year,
        default_start_year,
        default_end_year,
    )
    current_fiscal_year = f"{selected_start_year}-{selected_end_year}"
    if current_fiscal_year not in fiscal_year_options:
        current_fiscal_year = f"{default_start_year}-{default_end_year}"

    _, _, download_window_meta = resolve_reporting_download_window(
        requested_fy=current_fiscal_year,
        period_start_value=requested_period_start,
        period_end_value=requested_period_end,
        stream='period',
    )
    if download_window_meta.get('mode') == 'period':
        download_period_start = download_window_meta.get('start_label', '')
        download_period_end = download_window_meta.get('end_label', '')
    else:
        download_period_start = ''
        download_period_end = ''
    
    report_query = PBOReport.query.options(
        selectinload(PBOReport.assets),
        selectinload(PBOReport.donations),
        selectinload(PBOReport.payments),
        selectinload(PBOReport.project_implementations),
        selectinload(PBOReport.officials),
        selectinload(PBOReport.volunteer_privileges),
        selectinload(PBOReport.training_records),
        selectinload(PBOReport.user),
        selectinload(PBOReport.last_modified_by),
    )
    if not can_manage_all_records(current_user):
        report_query = report_query.filter(PBOReport.user_id == current_user.id)
    if search_query:
        pattern = f"%{search_query}%"
        report_query = report_query.filter(
            or_(
                PBOReport.pbo_name.ilike(pattern),
                PBOReport.pbo_registration_number.ilike(pattern),
                PBOReport.contact_name.ilike(pattern),
                PBOReport.contact_email.ilike(pattern),
            )
        )

    pagination = report_query.order_by(PBOReport.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    reports = pagination.items

    reports_vm = []
    for r in reports:
        category_display = display_scope(r.scope)

        # Assets: sum of asset values
        assets_total = sum((a.value or 0) for a in getattr(r, 'assets', []))

        # Donations: sum of donation amounts
        donations_total = sum((d.amount or 0) for d in getattr(r, 'donations', []))

        # Staff: show Kenyan and Foreign current staff
        staff_display = "Kenyan: {0}, Foreign: {1}".format(
            r.staff_kenyan_current if r.staff_kenyan_current is not None else "N/A",
            r.staff_foreign_current if r.staff_foreign_current is not None else "N/A",
        )

        # Projects: count of project implementations
        projects_count = len(getattr(r, 'project_implementations', []))
        payments_total = sum(((p.kenya_amount or 0) + (p.other_amount or 0)) for p in getattr(r, 'payments', []))
        payments_count = len(getattr(r, 'payments', []))
        payments_summary = (
            f"Rows: {payments_count}, Total: {payments_total:,.2f}"
            if payments_count else "N/A"
        )
        key_answers = [
            f"Audited: {r.audited or 'N/A'}",
            f"Assets issue: {r.assets_stolen or 'N/A'}",
            f"Tax waiver: {'YES' if r.gov_tax_waiver else 'NO'}",
            f"Election: {r.election_frequency or r.election_frequency_other or 'N/A'}",
        ]

        officials_names = ", ".join([o.name for o in r.officials if o.name][:3])
        volunteer_priv_categories = [
            v.category for v in r.volunteer_privileges
            if v.category and any([v.kenyan_volunteer, v.kenyan_intern, v.international_volunteer, v.international_intern])
        ]
        volunteer_privileges_summary = ", ".join(volunteer_priv_categories) if volunteer_priv_categories else "N/A"
        training_total = sum(
            (t.kenyan_count or 0) + (t.international_count or 0)
            for t in r.training_records
        )
        training_summary = (
            f"Types: {len(r.training_records)}, Total: {training_total}"
            if r.training_records else "N/A"
        )
        reports_vm.append({
            'report': r,
            'updater_name': report_updater_display(r),
            'reporting_period_start_display': reporting_period_display(r.reporting_period_start, r.reporting_period_start_raw),
            'reporting_period_end_display': reporting_period_display(r.reporting_period_end, r.reporting_period_end_raw),
            'return_filing_date_display': return_date_display(r.return_date),
            'category_display': category_display,
            'assets_total': assets_total,
            'donations_total': donations_total,
            'payments_summary': payments_summary,
            'staff_display': staff_display,
            'projects_count': projects_count,
            'created_at_display': format_datetime(r.created_at),
            'updated_at_display': format_datetime(r.updated_at),
            'submitted_at_display': format_datetime(r.submitted_at),
            'key_answers_summary': key_answers,
            'staff_summary': staff_display or 'N/A',
            'officials_names': officials_names or 'N/A',
            'volunteer_privileges_summary': volunteer_privileges_summary,
            'training_summary': training_summary,
        })

    return render_template(
        'reports_list.html',
        reports_vm=reports_vm,
        pagination=pagination,
        fiscal_year_options=fiscal_year_options,
        current_fiscal_year=current_fiscal_year,
        download_period_start=download_period_start,
        download_period_end=download_period_end,
        per_page=per_page,
        per_page_options=per_page_options,
        search_query=search_query,
    )


@app.route('/sectorreportdata')
@require_sector_analytics
@login_required
def sector_report_data():
    total_reports = db.session.query(func.count(PBOReport.id)).scalar() or 0
    rows = []
    if total_reports:
        from sqlalchemy.orm import selectinload
        reports = (
            PBOReport.query.options(
                selectinload(PBOReport.igas),
                selectinload(PBOReport.donations),
                selectinload(PBOReport.payments),
                selectinload(PBOReport.bank_accounts),
                selectinload(PBOReport.project_implementations),
                selectinload(PBOReport.projects_carried_out),
                selectinload(PBOReport.user),
                selectinload(PBOReport.last_modified_by),
            )
            .order_by(PBOReport.created_at.desc())
            .all()
        )
        rows = build_sector_report_rows(reports)
    allowed_columns, disabled_columns, money_columns = get_sector_report_columns()
    return render_template(
        'sectorreportdata.html',
        rows=rows,
        total_reports=total_reports,
        allowed_columns=allowed_columns,
        disabled_columns=disabled_columns,
        money_columns=money_columns,
        generated_at=format_datetime(utc_now()),
    )


@app.route('/sectorreportdata.json')
@require_sector_analytics
@login_required
def sector_report_data_json():
    reports = (
        reports_download_base_query()
        .options(
            selectinload(PBOReport.user),
            selectinload(PBOReport.last_modified_by),
        )
        .all()
    )
    return jsonify(build_sector_report_rows(reports))


@app.route('/sectorreportdata/download.xlsx')
@require_sector_analytics
@login_required
def sector_report_data_download_xlsx():
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    from sqlalchemy.orm import selectinload

    reports = (
        PBOReport.query.options(
            selectinload(PBOReport.igas),
            selectinload(PBOReport.donations),
            selectinload(PBOReport.payments),
            selectinload(PBOReport.bank_accounts),
            selectinload(PBOReport.project_implementations),
            selectinload(PBOReport.projects_carried_out),
            selectinload(PBOReport.user),
            selectinload(PBOReport.last_modified_by),
        )
        .order_by(PBOReport.created_at.desc())
        .all()
    )
    rows = build_sector_report_rows(reports)
    allowed_columns, disabled_columns, money_columns = get_sector_report_columns()
    ordered_columns = allowed_columns + disabled_columns

    df = pandas_module.DataFrame(
        [
            {label: row.get(key) for key, label in ordered_columns}
            for row in rows
        ],
        columns=[label for _, label in ordered_columns],
    )

    target_dir = os.path.join(app.root_path, 'static', 'extracted_xls')
    os.makedirs(target_dir, exist_ok=True)
    file_path = os.path.join(target_dir, 'sector_report_data.xlsx')

    buffer = BytesIO()
    with pandas_module.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sector Report Data', index=False)
        worksheet = writer.sheets['Sector Report Data']
        money_labels = {label for key, label in ordered_columns if key in money_columns}

        for col_idx, (_, label) in enumerate(ordered_columns, start=1):
            if label in money_labels:
                for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for value_cell in cell:
                        value_cell.number_format = '#,##0.00'

    buffer.seek(0)
    with pandas_module.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sector Report Data', index=False)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='sector_report_data.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

@app.route('/reports/download-2021-newly-reg-counties')
@login_required
def download_2021_newly_reg_counties():
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font

    # ---------------------------------------------------
    # DATE RANGE
    # ---------------------------------------------------
    today = get_app_today()
    default_start_year, default_end_year = get_current_fiscal_year(today)
    requested_fy = request.args.get('fy')
    start_year, end_year = parse_fiscal_year(
        requested_fy,
        default_start_year,
        default_end_year,
    )

    range_start = datetime(start_year, 7, 1).date()
    range_end = datetime(end_year, 6, 30).date()

    target_dir = os.path.join(app.root_path, 'static', 'extracted_xls')
    os.makedirs(target_dir, exist_ok=True)

    filename = f"{range_start.year}-{range_end.year} newly reg counties of ops.xlsx"
    file_path = os.path.join(target_dir, filename)

    # ---------------------------------------------------
    # FETCH DATA
    # ---------------------------------------------------
    records = PBOReport.query.order_by(PBOReport.created_at.desc()).all()

    rows = []
    for report in records:
        reg_date = report.date_of_registration or report.pbo_registration_date
        if not reg_date or not (range_start <= reg_date <= range_end):
            continue

        is_national = 1 if report.scope == 'NATIONAL' else 0
        is_international = 1 if report.scope == 'INTERNATIONAL' else 0

        rows.append({
            'OrgName': report.pbo_name or '',
            'Scope': display_scope(report.scope),
            'RegDate': format_date(reg_date),
            'County': report.counties or '',
            'scope_national': is_national,
            'scope_international': is_international,
        })

    pandas_module = get_pandas()
    df = pandas_module.DataFrame(rows)

    # ---------------------------------------------------
    # COUNTY DISTRIBUTION + CLEAN ORG LOOKUP
    # ---------------------------------------------------
    county_counts = {}
    county_org_lookup = {}

    for _, row in df.iterrows():
        org = row['OrgName']

        counties = [x.strip() for x in str(row['County']).split(',') if x.strip()]

        for county in counties:
            county_counts[county] = county_counts.get(county, 0) + 1
            county_org_lookup.setdefault(county, set()).add(org)   # ✅ USE SET

    county_dist_df = pandas_module.DataFrame(
        sorted(county_counts.items(), key=lambda x: x[1], reverse=True),
        columns=['County', 'NGO_Count']
    )

    county_dist_df['Organizations'] = ''

    # ---------------------------------------------------
    # SCOPE SUMMARY
    # ---------------------------------------------------
    scope_totals_df = pandas_module.DataFrame([
        {'ScopeType': 'National', 'Total': df['scope_national'].sum()},
        {'ScopeType': 'International', 'Total': df['scope_international'].sum()},
    ])

    scope_totals_df['Organizations'] = ''

    # ---------------------------------------------------
    # REGISTRATIONS PER MONTH
    # ---------------------------------------------------
    df['RegDate_dt'] = pandas_module.to_datetime(df['RegDate'], errors='coerce').dt.date
    df['RegMonth'] = pandas_module.to_datetime(df['RegDate'], errors='coerce').dt.to_period('M')

    reg_month_counts = (
        df.groupby('RegMonth')
        .size()
        .reset_index(name='Registrations')
    )

    reg_month_counts['RegMonth'] = reg_month_counts['RegMonth'].astype(str)

    # ---------------------------------------------------
    # WRITE TO EXCEL
    # ---------------------------------------------------
    with pandas_module.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='NGO Data', index=False)
        county_dist_df.to_excel(writer, sheet_name='County Distribution', index=False)
        scope_totals_df.to_excel(writer, sheet_name='Scope Summary', index=False)
        reg_month_counts.to_excel(writer, sheet_name='Registrations Per Month', index=False)

    wb = load_workbook(file_path)

    ws_county = wb['County Distribution']
    ws_scope = wb['Scope Summary']
    ws_month = wb['Registrations Per Month']
    ws_hidden = wb.create_sheet('OrgLists')
    ws_hidden.sheet_state = 'hidden'

    hidden_col = 1

    # ---------------------------------------------------
    # COUNTY DROPDOWNS (FIXED)
    # ---------------------------------------------------
    for row_idx in range(2, ws_county.max_row + 1):

        county = ws_county.cell(row=row_idx, column=1).value
        orgs = sorted(list(county_org_lookup.get(county, [])))

        if not orgs:
            continue

        # write org list vertically
        for r, org in enumerate(orgs, start=1):
            ws_hidden.cell(row=r, column=hidden_col, value=org)

        col_letter = get_column_letter(hidden_col)
        formula = f"'OrgLists'!${col_letter}$1:${col_letter}${len(orgs)}"

        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws_county.add_data_validation(dv)

        cell = ws_county.cell(row=row_idx, column=3)
        dv.add(cell)
        cell.value = orgs[0]  # auto select first

        hidden_col += 1

    # ---------------------------------------------------
    # SCOPE DROPDOWNS
    # ---------------------------------------------------
    for row_idx in range(2, ws_scope.max_row + 1):

        scope = ws_scope.cell(row=row_idx, column=1).value

        if scope == 'National':
            orgs = df.loc[df['scope_national'] == 1, 'OrgName'].unique().tolist()
        else:
            orgs = df.loc[df['scope_international'] == 1, 'OrgName'].unique().tolist()

        orgs = sorted(orgs)

        if not orgs:
            continue

        for r, org in enumerate(orgs, start=1):
            ws_hidden.cell(row=r, column=hidden_col, value=org)

        col_letter = get_column_letter(hidden_col)
        formula = f"'OrgLists'!${col_letter}$1:${col_letter}${len(orgs)}"

        dv = DataValidation(type="list", formula1=formula)
        ws_scope.add_data_validation(dv)

        cell = ws_scope.cell(row=row_idx, column=3)
        dv.add(cell)
        cell.value = orgs[0]

        hidden_col += 1

    # ---------------------------------------------------
    # ADD EXCEL NATIVE CHARTS
    # ---------------------------------------------------

    # County Chart
    county_chart = BarChart()
    county_chart.title = "NGOs by County"
    data = Reference(ws_county, min_col=2, min_row=1, max_row=ws_county.max_row)
    cats = Reference(ws_county, min_col=1, min_row=2, max_row=ws_county.max_row)
    county_chart.add_data(data, titles_from_data=True)
    county_chart.set_categories(cats)
    ws_county.add_chart(county_chart, "E2")

    # Scope Chart
    scope_chart = PieChart()
    scope_chart.title = "Scope Distribution"
    data = Reference(ws_scope, min_col=2, min_row=1, max_row=ws_scope.max_row)
    labels = Reference(ws_scope, min_col=1, min_row=2, max_row=ws_scope.max_row)
    scope_chart.add_data(data, titles_from_data=True)
    scope_chart.set_categories(labels)
    ws_scope.add_chart(scope_chart, "E2")

    # Registrations Per Month Chart
    month_chart = BarChart()
    month_chart.title = "Registrations Per Month"
    data = Reference(ws_month, min_col=2, min_row=1, max_row=ws_month.max_row)
    cats = Reference(ws_month, min_col=1, min_row=2, max_row=ws_month.max_row)
    month_chart.add_data(data, titles_from_data=True)
    month_chart.set_categories(cats)
    ws_month.add_chart(month_chart, "D2")

    wb.save(file_path)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/reports/download-collaboration-data')
@login_required
def download_collaboration_data():
    """Download XLSX of collaboration data with:
       - Yes/No converted to 1/0
       - Totals
       - Correlations
       - Dropdown organization list per category
    """

    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.datavalidation import DataValidation

    today = get_app_today()
    default_start_year, default_end_year = get_current_fiscal_year(today)
    requested_fy = request.args.get('fy')
    start_year, end_year = parse_fiscal_year(
        requested_fy,
        default_start_year,
        default_end_year,
    )

    range_start = datetime(start_year, 7, 1).date()
    range_end = datetime(end_year, 6, 30).date()

    target_dir = os.path.join(app.root_path, 'static', 'extracted_xls')
    os.makedirs(target_dir, exist_ok=True)
    filename = f"{range_start.year}-{range_end.year} collaboration data.xlsx"
    file_path = os.path.join(target_dir, filename)

    reports = reports_for_download()

    rows = []
    for report in reports:
        reg_date = report.date_of_registration or report.pbo_registration_date
        if not reg_date or not (range_start <= reg_date <= range_end):
            continue

        for c in report.collaborations:
            rows.append({
                'OrgName': report.pbo_name or '',
                'Scope': display_scope(report.scope),
                'InstTypeName': c.partner_type or '',
                'SumOfInfoExchange': c.info_exchange,
                'SumOfTechSupportTo': c.tech_support_to_partner,
                'SumOfTechSupportFrom': c.tech_support_from_partner,
                'SumOfFundingTo': c.funding_to_partner,
                'SumOfFundingFrom': c.funding_from_partner,
                'SumOfEquipmentTo': c.equipment_to_partner,
                'SumOfEquipmentFrom': c.equipment_from_partner,
            })

    df = pandas_module.DataFrame(rows)

    numeric_columns = [
        'SumOfInfoExchange',
        'SumOfTechSupportTo',
        'SumOfTechSupportFrom',
        'SumOfFundingTo',
        'SumOfFundingFrom',
        'SumOfEquipmentTo',
        'SumOfEquipmentFrom',
    ]

    # ---- YES/NO → 1/0 ----
    def yes_no_to_numeric(value):
        if isinstance(value, str):
            val = value.strip().lower()
            if val == 'yes':
                return 1
            elif val == 'no':
                return 0
        if value is True:
            return 1
        if value is False:
            return 0
        return pandas_module.to_numeric(value, errors='coerce')

    for col in numeric_columns:
        df[col] = df[col].apply(yes_no_to_numeric).fillna(0)

    # ---- CALCULATIONS ----
    totals_data = []
    org_lookup = {}  # Store org lists per category

    for col in numeric_columns:
        total_value = int(df[col].sum())
        orgs = sorted(df.loc[df[col] == 1, 'OrgName'].dropna().unique())
        org_lookup[col] = orgs

        totals_data.append({
            'Category': col,
            'Total': total_value,
            'Organizations': ''  # Placeholder (dropdown cell)
        })

    totals_df = pandas_module.DataFrame(totals_data)
    correlation_df = df[numeric_columns].corr()
    describe_df = df[numeric_columns].describe()

    # ---- WRITE TO EXCEL ----
    with pandas_module.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Collaboration Data', index=False)

        wb = writer.book
        ws_summary = wb.create_sheet('Summary & Graphs')
        ws_hidden = wb.create_sheet('OrgLists')
        ws_hidden.sheet_state = 'hidden'

        row_cursor = 1

        # ---- TOTALS TABLE ----
        ws_summary.cell(row=row_cursor, column=1, value="TOTALS").font = Font(bold=True)
        row_cursor += 1

        for r in dataframe_to_rows(totals_df, index=False, header=True):
            ws_summary.append(r)

        totals_start_row = 3
        totals_end_row = totals_start_row + len(totals_df) - 1

        # ---- CREATE DROPDOWN LISTS ----
        hidden_col = 1
        for i, (category, orgs) in enumerate(org_lookup.items()):
            if not orgs:
                continue

            # Write org list vertically in hidden sheet
            for row_idx, org in enumerate(orgs, start=1):
                ws_hidden.cell(row=row_idx, column=hidden_col, value=org)

            col_letter = ws_hidden.cell(row=1, column=hidden_col).column_letter
            formula_range = f"'OrgLists'!${col_letter}$1:${col_letter}${len(orgs)}"

            dv = DataValidation(
                type="list",
                formula1=formula_range,
                allow_blank=True
            )

            ws_summary.add_data_validation(dv)

            # Apply validation to Organizations column (column 3)
            target_row = totals_start_row + i
            dv.add(ws_summary.cell(row=target_row, column=3))

            # Auto-select first item in dropdown
            if orgs:
                ws_summary.cell(row=target_row, column=3, value=orgs[0])


            hidden_col += 1

        row_cursor = totals_end_row + 2

        # ---- BASIC STATS ----
        ws_summary.cell(row=row_cursor, column=1, value="BASIC STATISTICS").font = Font(bold=True)
        row_cursor += 1

        for r in dataframe_to_rows(describe_df, index=True, header=True):
            ws_summary.append(r)

        row_cursor += len(describe_df) + 3

        # ---- CORRELATION ----
        ws_summary.cell(row=row_cursor, column=1, value="CORRELATION MATRIX").font = Font(bold=True)
        row_cursor += 1

        for r in dataframe_to_rows(correlation_df, index=True, header=True):
            ws_summary.append(r)

        # ---- BAR CHART ----
        chart = BarChart()
        chart.title = "Total Collaboration Counts (Yes=1, No=0)"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Collaboration Type"

        data = Reference(
            ws_summary,
            min_col=2,
            min_row=2,
            max_row=totals_end_row
        )

        categories = Reference(
            ws_summary,
            min_col=1,
            min_row=3,
            max_row=totals_end_row
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws_summary.add_chart(chart, "J2")

    return send_file(
        file_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/reports/download-projects-data')
@login_required
def download_projects_data():
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    range_start, range_end, window_meta = resolve_reporting_download_window(
        requested_fy=request.args.get('fy'),
        period_start_value=request.args.get('period_start'),
        period_end_value=request.args.get('period_end'),
        stream=request.args.get('stream'),
    )
    filename_label = window_meta.get('filename_label')
    reports = reports_for_download()

    implementation_rows = []
    carried_out_rows = []

    for report in reports:
        if not report_matches_fiscal_year(report, range_start, range_end):
            continue

        for project in report.project_implementations:
            implementation_rows.append(legacy_zero_output_structure(report, {
                'ReportID': report.id,
                'OrgName': report.pbo_name or '',
                'Scope': display_scope(report.scope),
                'ReportingPeriodStart': format_date(report.reporting_period_start),
                'ReportingPeriodEnd': format_date(report.reporting_period_end),
                'Sector': project.sector or '',
                'County': project.county or '',
                'VulnerableGroup': project.vulnerable_group or '',
                'BeneficiariesNo': legacy_zero_output_value(report, project.beneficiaries_no),
                'SpendingPerCounty': resolve_project_spending_amount(project, report=report),
                'DurationYears': legacy_zero_output_value(report, project.duration_years),
                'CompletionStatus': project.completion_status or '',
                'AmountSpentKenya': legacy_zero_output_value(report, project.amount_spent_kenya),
                'AmountSpentOther': legacy_zero_output_value(report, project.amount_spent_other),
            }))

        for project in report.projects_carried_out:
            carried_out_rows.append(legacy_zero_output_structure(report, {
                'ReportID': report.id,
                'OrgName': report.pbo_name or '',
                'Scope': display_scope(report.scope),
                'ReportingPeriodStart': format_date(report.reporting_period_start),
                'ReportingPeriodEnd': format_date(report.reporting_period_end),
                'Sector': project.sector or '',
                'CarriedForwardKenya': legacy_zero_output_value(report, project.carried_forward_kenya),
                'CarriedForwardOther': legacy_zero_output_value(report, project.carried_forward_other),
                'StartedKenya': legacy_zero_output_value(report, project.started_kenya),
                'StartedOther': legacy_zero_output_value(report, project.started_other),
                'CompletedKenya': legacy_zero_output_value(report, project.completed_kenya),
                'CompletedOther': legacy_zero_output_value(report, project.completed_other),
            }))

    implementation_df = pandas_module.DataFrame(implementation_rows)
    carried_out_df = pandas_module.DataFrame(carried_out_rows)

    buffer = BytesIO()
    with pandas_module.ExcelWriter(buffer, engine='openpyxl') as writer:
        implementation_df.to_excel(writer, sheet_name='Implementations', index=False)
        carried_out_df.to_excel(writer, sheet_name='Projects Carried Out', index=False)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'{filename_label} projects_data.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/reports/download-registration-period-data')
@login_required
def download_registration_period_data():
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    range_start, range_end, window_meta = resolve_reporting_download_window(
        requested_fy=request.args.get('fy'),
        period_start_value=request.args.get('period_start'),
        period_end_value=request.args.get('period_end'),
        stream=request.args.get('stream'),
    )
    filename_label = window_meta.get('filename_label')
    reports = reports_for_download()

    rows = []
    for report in reports:
        if not report_matches_fiscal_year(report, range_start, range_end):
            continue
        rows.append({
            'OrgName': report.pbo_name or '',
            'PBORegistrationNumber': report.pbo_registration_number or '',
            'PBORegistrationDate': format_date(report.pbo_registration_date or report.date_of_registration),
            'ReportingPeriodStart': format_date(report.reporting_period_start),
            'ReportingPeriodEnd': format_date(report.reporting_period_end),
            'ReturnCreatedAt': format_datetime(report.created_at),
        })

    df = pandas_module.DataFrame(rows)
    buffer = BytesIO()
    with pandas_module.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Registration Period', index=False)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'{filename_label} registration_period_data.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/reports/download-category-scope-data')
@login_required
def download_category_scope_data():
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    range_start, range_end, window_meta = resolve_reporting_download_window(
        requested_fy=request.args.get('fy'),
        period_start_value=request.args.get('period_start'),
        period_end_value=request.args.get('period_end'),
        stream=request.args.get('stream'),
    )
    filename_label = window_meta.get('filename_label')
    reports = reports_for_download()

    rows = []
    for report in reports:
        if not report_matches_fiscal_year(report, range_start, range_end):
            continue
        rows.append({
            'OrgName': report.pbo_name or '',
            'Category': display_scope(report.scope),
            'Scope': report.scope or '',
            'Counties': report.counties or '',
            'Audited': report.audited or '',
        })

    df = pandas_module.DataFrame(rows)
    buffer = BytesIO()
    with pandas_module.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Category Scope', index=False)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'{filename_label} category_scope_data.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/reports/download-assets-data')
@login_required
def download_assets_data():
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    range_start, range_end, window_meta = resolve_reporting_download_window(
        requested_fy=request.args.get('fy'),
        period_start_value=request.args.get('period_start'),
        period_end_value=request.args.get('period_end'),
        stream=request.args.get('stream'),
    )
    filename_label = window_meta.get('filename_label')
    reports = reports_for_download()

    rows = []
    for report in reports:
        if not report_matches_fiscal_year(report, range_start, range_end):
            continue
        for asset in report.assets:
            rows.append(legacy_zero_output_structure(report, {
                'OrgName': report.pbo_name or '',
                'ReportingPeriodStart': format_date(report.reporting_period_start),
                'ReportingPeriodEnd': format_date(report.reporting_period_end),
                'AssetItem': asset.item or '',
                'AssetNumber': legacy_zero_output_value(report, asset.number),
                'AssetValue': legacy_zero_output_value(report, asset.value),
            }))

    df = pandas_module.DataFrame(rows)
    buffer = BytesIO()
    with pandas_module.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Assets', index=False)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'{filename_label} assets_data.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/reports/download-donations-data')
@login_required
def download_donations_data():
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    range_start, range_end, window_meta = resolve_reporting_download_window(
        requested_fy=request.args.get('fy'),
        period_start_value=request.args.get('period_start'),
        period_end_value=request.args.get('period_end'),
        stream=request.args.get('stream'),
    )
    filename_label = window_meta.get('filename_label')
    reports = reports_for_download()

    rows = []
    for report in reports:
        if not report_matches_fiscal_year(report, range_start, range_end):
            continue
        for donation in report.donations:
            rows.append(legacy_zero_output_structure(report, {
                'OrgName': report.pbo_name or '',
                'ReportingPeriodStart': format_date(report.reporting_period_start),
                'ReportingPeriodEnd': format_date(report.reporting_period_end),
                'DonorName': donation.name or '',
                'DonorCategory': donation.category or '',
                'DonorCountry': donation.country or '',
                'Amount': legacy_zero_output_value(report, donation.amount),
            }))

    df = pandas_module.DataFrame(rows)
    buffer = BytesIO()
    with pandas_module.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Donations', index=False)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'{filename_label} donations_data.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/reports/download-payments-data')
@login_required
def download_payments_data():
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    range_start, range_end, window_meta = resolve_reporting_download_window(
        requested_fy=request.args.get('fy'),
        period_start_value=request.args.get('period_start'),
        period_end_value=request.args.get('period_end'),
        stream=request.args.get('stream'),
    )
    filename_label = window_meta.get('filename_label')
    reports = reports_for_download()

    rows = []
    for report in reports:
        if not report_matches_fiscal_year(report, range_start, range_end):
            continue

        for payment in report.payments:
            kenya_amount = legacy_zero_float(report, payment.kenya_amount)
            other_amount = legacy_zero_float(report, payment.other_amount)
            rows.append(legacy_zero_output_structure(report, {
                'ReportID': report.id,
                'OrgName': report.pbo_name or '',
                'Scope': display_scope(report.scope),
                'ReportingPeriodStart': format_date(report.reporting_period_start),
                'ReportingPeriodEnd': format_date(report.reporting_period_end),
                'Description': payment.description or '',
                'KenyaAmount': kenya_amount,
                'OtherAmount': other_amount,
                'TotalAmount': kenya_amount + other_amount,
            }))

    df = pandas_module.DataFrame(rows)

    buffer = BytesIO()
    with pandas_module.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Payments', index=False)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'{filename_label} payments_data.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/reports/download-governance-data')
@login_required
def download_governance_data():
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    range_start, range_end, window_meta = resolve_reporting_download_window(
        requested_fy=request.args.get('fy'),
        period_start_value=request.args.get('period_start'),
        period_end_value=request.args.get('period_end'),
        stream=request.args.get('stream'),
    )
    filename_label = window_meta.get('filename_label')
    reports = reports_for_download()

    summary_rows = []
    officials_rows = []

    for report in reports:
        if not report_matches_fiscal_year(report, range_start, range_end):
            continue

        summary_rows.append(legacy_zero_output_structure(report, {
            'ReportID': report.id,
            'OrgName': report.pbo_name or '',
            'Scope': display_scope(report.scope),
            'ReportingPeriodStart': format_date(report.reporting_period_start),
            'ReportingPeriodEnd': format_date(report.reporting_period_end),
            'MembershipDirectors': legacy_zero_output_value(report, report.membership_number_of_directors),
            'MembershipRegisteredMembers': legacy_zero_output_value(report, report.membership_number_of_registered_members),
            'MembershipBoardMeetings': legacy_zero_output_value(report, report.membership_number_of_board_meetings),
            'MembershipLastAGM': format_date(report.membership_date_last_agm),
            'MembershipLastElection': format_date(report.membership_date_last_election),
            'NonMembershipDirectors': legacy_zero_output_value(report, report.non_membership_number_of_directors),
            'NonMembershipBoardMeetings': legacy_zero_output_value(report, report.non_membership_number_of_board_meetings),
            'NonMembershipLastBoardMeeting': format_date(report.non_membership_date_last_board_meeting),
            'NonMembershipLastElection': format_date(report.non_membership_date_last_election),
            'SubmitterFullname': report.submitter_fullname or '',
            'SubmissionDate': format_date(report.submission_date),
        }))

        for official in report.officials:
            officials_rows.append({
                'ReportID': report.id,
                'OrgName': report.pbo_name or '',
                'Role': official.role or '',
                'Name': official.name or '',
                'Nationality': official.nationality or '',
                'Gender': official.gender or '',
                'Email': official.email or '',
                'Residence': official.residence or '',
                'Phone': official.phone or '',
                'KRAPin': official.kra_pin or '',
                'ProfessionalQualification': official.professional_qualification or '',
            })

    summary_df = pandas_module.DataFrame(summary_rows)
    officials_df = pandas_module.DataFrame(officials_rows)

    buffer = BytesIO()
    with pandas_module.ExcelWriter(buffer, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Governance Summary', index=False)
        officials_df.to_excel(writer, sheet_name='Officials', index=False)

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'{filename_label} governance_data.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/reports/download-county-data')
@login_required
def download_county_data():
    """Download XLSX with sheet 'funsource' and one sheet per county, with required columns."""
    try:
        pandas_module = get_pandas()
    except Exception:
        return 'Pandas is not available on the server.', 500

    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference

    range_start, range_end, window_meta = resolve_reporting_download_window(
        requested_fy=request.args.get('fy'),
        period_start_value=request.args.get('period_start'),
        period_end_value=request.args.get('period_end'),
        stream=request.args.get('stream'),
    )
    filename_label = window_meta.get('filename_label')
    reports = reports_for_download()

    # Build rows for the main sheet
    funsource_rows = []
    for report in reports:
        if not report_matches_fiscal_year(report, range_start, range_end):
            continue
        # Section A1
        org_name = report.pbo_name or ''
        postal_address = report.postal_address or ''
        physical_address = report.physical_address or ''
        telephone = report.telephone or ''
        cell_phone = report.cell_phone or ''
        email = report.email or ''
        # Section A2
        contact_name = report.contact_name or ''
        contact_position = report.contact_position or ''
        contact_telephone = report.contact_telephone or ''
        contact_email = report.contact_email or ''
        contact_nationality = report.contact_nationality or ''
        contact_gender = report.contact_gender or ''
        # Scope
        scope = display_scope(report.scope)
        # Amounts
        assets_amt = sum(legacy_zero_float(report, a.value) for a in report.assets)
        donations_amt = sum(legacy_zero_float(report, d.amount) for d in report.donations)
        grants_amt = sum(legacy_zero_float(report, g.amount) for g in report.grants)
        payments_amt = sum(legacy_zero_float(report, p.kenya_amount) for p in report.payments) + sum(
            legacy_zero_float(report, p.other_amount) for p in report.payments
        )
        # Return date
        return_date = format_date(report.created_at) if report.created_at else ''

        funsource_rows.append(legacy_zero_output_structure(report, {
            'Scope': scope,
            'OrgName': org_name,
            'PostalAddress': postal_address,
            'PhysicalAddress': physical_address,
            'Name': contact_name,
            'Telephone': telephone,
            'CellPhone': cell_phone,
            'Email': email,
            'ContactPosition': contact_position,
            'ContactTelephone': contact_telephone,
            'ContactEmail': contact_email,
            'ContactNationality': contact_nationality,
            'ContactGender': contact_gender,
            'AssetsAmount': assets_amt,
            'DonationsAmount': donations_amt,
            'GrantsAmount': grants_amt,
            'PaymentsAmount': payments_amt,
            'ReturnDate': return_date,
            'Counties': report.counties or '',
        }))

    df = pandas_module.DataFrame(funsource_rows)

    # Prepare output file
    target_dir = os.path.join(app.root_path, 'static', 'extracted_xls')
    os.makedirs(target_dir, exist_ok=True)
    filename = f"{filename_label} county data.xlsx"
    file_path = os.path.join(target_dir, filename)

    # Write to Excel
    with pandas_module.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Main sheet
        df.to_excel(writer, sheet_name='funsource', index=False)
        used_sheet_names = {'funsource'}

        # County sheets
        # Get all unique counties
        all_counties = set()
        for counties in df['Counties']:
            if counties:
                for c in [x.strip() for x in counties.split(',') if x.strip()]:
                    all_counties.add(c)
        for county in sorted(all_counties):
            county_df = df[df['Counties'].str.contains(county, na=False)]
            # Remove the Counties column for county sheets
            county_df = county_df.drop(columns=['Counties'])
            county_df.to_excel(
                writer,
                sheet_name=make_excel_sheet_name(county, used_names=used_sheet_names, fallback='County'),
                index=False,
            )

    return send_file(
        file_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )





@app.route('/reports/download-staff-volunteer-summary')
@login_required
def download_staff_volunteer_summary():
        """Download XLSX with staff/volunteer summary variables and graphs."""
        try:
            pandas_module = get_pandas()
        except Exception:
            return 'Pandas is not available on the server.', 500

        # Import BarChart and Reference for Excel charting
        from openpyxl.chart import BarChart, Reference

        range_start, range_end, window_meta = resolve_reporting_download_window(
            requested_fy=request.args.get('fy'),
            period_start_value=request.args.get('period_start'),
            period_end_value=request.args.get('period_end'),
            stream=request.args.get('stream'),
        )
        filename_label = window_meta.get('filename_label')
        reports = reports_for_download()

        summary_rows = []
        for report in reports:
            if not report_matches_fiscal_year(report, range_start, range_end):
                continue
            # Scope as string
            scope = display_scope(report.scope)

            summary_rows.append(legacy_zero_output_structure(report, {
                'OrgName': report.pbo_name or '',
                'PostalAddress': report.postal_address or '',
                'PhysicalAddress': report.physical_address or '',
                'Telephone': report.telephone or '',
                'Email': report.email or '',
                'Scope': scope,
                'SumOfPrevLocalKe': legacy_zero_output_value(report, report.staff_kenyan_prev),
                'SumOfCurLocalKe': legacy_zero_output_value(report, report.staff_kenyan_current),
                'SumOfTurnoverLocalKe': (
                    legacy_zero_int(report, report.staff_kenyan_came_in) - legacy_zero_int(report, report.staff_kenyan_left)
                ) if report.staff_kenyan_came_in is not None and report.staff_kenyan_left is not None else None,
                'SumOfPrevForeignKe': legacy_zero_output_value(report, report.staff_foreign_prev),
                'SumOfCurForeignKe': legacy_zero_output_value(report, report.staff_foreign_current),
                'SumOfTurnoverForeignKe': (
                    legacy_zero_int(report, report.staff_foreign_came_in) - legacy_zero_int(report, report.staff_foreign_left)
                ) if report.staff_foreign_came_in is not None and report.staff_foreign_left is not None else None,
                'SumOfPrevLocalOth': legacy_zero_output_value(report, report.staff_other_kenyan_prev),
                'SumOfCurLocalOth': legacy_zero_output_value(report, report.staff_other_kenyan_current),
                'SumOfTurnoverLocalOth': None,  # Not enough info for turnover in "other" (no came_in/left)
                'SumOfPrevForeignOth': legacy_zero_output_value(report, report.staff_other_foreign_prev),
                'SumOfCurForeignOth': legacy_zero_output_value(report, report.staff_other_foreign_current),
                'SumOfTurnoverForeignOth': None,  # Not enough info for turnover in "other"
                'SumOfPrevVolLocal': legacy_zero_output_value(report, report.volunteers_kenyan_prev),
                'SumOfCurVolLocal': legacy_zero_output_value(report, report.volunteers_kenyan_current),
                'SumOfPrevVolForeign': legacy_zero_output_value(report, report.volunteers_foreign_prev),
                'SumOfCurVolForeign': legacy_zero_output_value(report, report.volunteers_foreign_current),
            }))

        df = pandas_module.DataFrame(summary_rows)

        # Prepare output file
        target_dir = os.path.join(app.root_path, 'static', 'extracted_xls')
        os.makedirs(target_dir, exist_ok=True)
        filename = f"{filename_label} staff_volunteer_summary.xlsx"
        file_path = os.path.join(target_dir, filename)


        # Write to Excel and add graphs in their own sheet, and split data per county
        with pandas_module.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write the summary sheet (all data)
            df.to_excel(writer, sheet_name='summary', index=False)
            wb = writer.book
            ws_summary = wb['summary']
            used_sheet_names = {'summary'}


            # Split data per county and write each to its own sheet
            if 'Scope' in df.columns and 'OrgName' in df.columns:
                # Try to find a column with counties, fallback to Scope if not present
                county_col = None
                for col in ['County', 'Counties', 'Scope']:
                    if col in df.columns:
                        county_col = col
                        break
                if county_col is None:
                    county_col = 'Scope'  # fallback

                # Build a set of all counties
                all_counties = set()
                for val in df[county_col]:
                    if isinstance(val, str):
                        for c in [x.strip() for x in val.split(',') if x.strip()]:
                            all_counties.add(c)
                # Write each county's data to its own sheet
                for county in sorted(all_counties):
                    # Filter rows where county is present in the county_col
                    mask = df[county_col].apply(lambda x: county in x if isinstance(x, str) else False)
                    county_df = df[mask]
                    safe_name = make_excel_sheet_name(county, used_names=used_sheet_names, fallback='County')
                    county_df.to_excel(writer, sheet_name=safe_name, index=False)

                # --- Add sheets for mixed scopes ---
                # Find all unique scope combinations (comma-separated, sorted, >1 scope)
                unique_scopes = set()
                for val in df['Scope']:
                    if isinstance(val, str):
                        scopes = tuple(sorted([s.strip() for s in val.split(',') if s.strip()]))
                        if len(scopes) > 1:
                            unique_scopes.add(scopes)
                for scopes in sorted(unique_scopes):
                    # Build a readable, Excel-safe sheet name
                    sheet_name = '_'.join(scopes)
                    safe_sheet_name = make_excel_sheet_name(
                        sheet_name,
                        used_names=used_sheet_names,
                        fallback='MixedScope',
                    )
                    # Filter rows matching exactly this set of scopes
                    def match_scopes(x):
                        if not isinstance(x, str):
                            return False
                        x_scopes = tuple(sorted([s.strip() for s in x.split(',') if s.strip()]))
                        return x_scopes == scopes
                    mixed_df = df[df['Scope'].apply(match_scopes)]
                    if not mixed_df.empty:
                        mixed_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            # Create a new sheet for graphs
            ws_graphs = wb.create_sheet(make_excel_sheet_name('Graphs', used_names=used_sheet_names))

            # Add bar chart for SumOfCurLocalKe and SumOfCurForeignKe by OrgName
            chart = BarChart()
            chart.title = "Current Staff (Kenyan/Foreign) by Org"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "OrgName"
            max_row = ws_summary.max_row
            orgs = Reference(ws_summary, min_col=1, min_row=2, max_row=max_row)
            local = Reference(ws_summary, min_col=7, min_row=1, max_row=max_row)  # SumOfCurLocalKe
            foreign = Reference(ws_summary, min_col=9, min_row=1, max_row=max_row)  # SumOfCurForeignKe
            chart.add_data(local, titles_from_data=True)
            chart.add_data(foreign, titles_from_data=True)
            chart.set_categories(orgs)
            ws_graphs.add_chart(chart, "B2")

            # Add bar chart for volunteers
            chart2 = BarChart()
            chart2.title = "Current Volunteers (Local/Foreign) by Org"
            chart2.y_axis.title = "Count"
            chart2.x_axis.title = "OrgName"
            vol_local = Reference(ws_summary, min_col=17, min_row=1, max_row=max_row)  # SumOfCurVolLocal
            vol_foreign = Reference(ws_summary, min_col=19, min_row=1, max_row=max_row)  # SumOfCurVolForeign
            chart2.add_data(vol_local, titles_from_data=True)
            chart2.add_data(vol_foreign, titles_from_data=True)
            chart2.set_categories(orgs)
            ws_graphs.add_chart(chart2, "B20")

        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )




@app.route('/eda')
@login_required
def eda():
    """Generate a pandas profiling report and EDA summaries for the reports data."""
    # Check if there is data
    records = PBOReport.query.order_by(PBOReport.created_at.desc()).all()
    if not records:
        msg = 'No data available. Submit some reports first.'
        return render_template('eda.html', error_message=msg)

    # Helper to parse JSON safely
    def jloads(s):
        try:
            return json.loads(s) if s else None
        except Exception:
            return None

    # Build a DataFrame if pandas is available
    df = None
    if pd is not None:
        rows = []
        for r in records:
            base = {c.name: getattr(r, c.name) for c in PBOReport.__table__.columns}
            # Derived metrics from JSON/text columns
            counties = [c.strip() for c in (base.get('counties') or "").split(",") if c.strip()]
            assets = r.assets
            donations = r.donations
            projects = r.project_implementations

            # Simple numeric aggregates
            assets_total_value = sum(legacy_zero_float(r, a.value) for a in assets)
            donations_total_amount = sum(legacy_zero_float(r, d.amount) for d in donations)
            staff_total_count = legacy_zero_int(r, r.staff_kenyan_current) + legacy_zero_int(r, r.staff_foreign_current)
            projects_total_budget = sum(resolve_project_spending_amount(p, report=r) for p in projects)

            base.update({
                'counties_count': len(counties),
                'assets_total_value': assets_total_value,
                'donations_total_amount': donations_total_amount,
                'staff_total_count': staff_total_count,
                'projects_total_budget': projects_total_budget,
            })
            rows.append(legacy_zero_output_structure(r, base))
        try:
            pandas_module = get_pandas()
            df = pandas_module.DataFrame(rows)
        except Exception:
            df = None

    profiling_html = None
    summary_sections = []  # list of (title, html)

    profile_report_class = get_profile_report_class()
    if df is not None and profile_report_class is not None:
        try:
            profile = profile_report_class(df, title='PBO Reports Profiling', minimal=True,
                        correlations={"pearson": False, "spearman": False, "kendall": False},
                        interactions=None,
                        samples=None,
                        duplicates=None,
                        missing_diagrams=False)
            profiling_html = profile.to_html()
        except Exception as e:
            summary_sections.append(('Profiling Error', f'<p>{str(e)}</p>'))

    # Fallback EDA summaries if profiling not available
    if df is not None and profiling_html is None:
        try:
            summary_sections.append(('Shape', f'<p>Rows: {df.shape[0]}, Columns: {df.shape[1]}</p>'))
            dtypes_html = df.dtypes.to_frame('dtype').to_html()
            summary_sections.append(('Dtypes', dtypes_html))
            # Describe for numeric and categorical
            try:
                desc_all = df.describe(include='all', datetime_is_numeric=True).transpose().to_html()
            except Exception:
                desc_all = df.describe(include='all').transpose().to_html()
            summary_sections.append(('Describe (all)', desc_all))
            # Missingness
            na_html = df.isna().sum().to_frame('missing').to_html()
            summary_sections.append(('Missing Values', na_html))
            # Key categorical distributions
            for col in ['audited', 'scope']:
                if col in df.columns:
                    vc = df[col].value_counts(dropna=False).to_frame('count').to_html()
                    summary_sections.append((f'Value Counts: {col}', vc))
        except Exception as e:
            summary_sections.append(('EDA Error', f'<p>{str(e)}</p>'))

    # If pandas missing, show guidance
    try:
        get_pandas()
    except Exception:
        missing = 'pandas is not installed. Install with: pip install pandas'
        return render_template('eda.html', error_message=missing)

    if profiling_html:
        return render_template('eda.html', profiling_html=profiling_html)
    else:
        return render_template('eda.html', summary_sections=summary_sections)























import os

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 8000))  # fallback to 10080
    host = os.environ.get("APP_HOST") or os.environ.get("FLASK_RUN_HOST") or "0.0.0.0"
    app.run(host=host, port=port, debug=env_flag("FLASK_DEBUG", False))
