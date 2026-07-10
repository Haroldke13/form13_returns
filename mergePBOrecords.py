#!/usr/bin/env python3
"""Merge a report-linked workbook with the PBO reports workbook.

Default usage:
    python3 mergePBOrecords.py --related pbo_donations.xlsx --output fundsource.xlsx

You can point it at any other workbook that carries a report foreign key:
    python3 mergePBOrecords.py \
        --reference "pbo_reports (2).xlsx" \
        --related pbo_project_implementations.xlsx \
        --output projectIMPLEMENTATIONS_v2.xlsx
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEFAULT_REFERENCE = "pbo_reports (2).xlsx"
DEFAULT_REFERENCE_FIELDS = [
    "pbo_name",
    "pbo_name_normalized",
    "pbo_registration_number",
    "reporting_period_start",
    "reporting_period_end",
    "scope",
    "counties",
    "countries_of_operation",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge a report-linked workbook with the PBO reports workbook."
    )
    parser.add_argument(
        "--reference",
        default=DEFAULT_REFERENCE,
        help="Reference workbook path. Defaults to %(default)s.",
    )
    parser.add_argument(
        "--related",
        required=True,
        help="Related workbook path containing a report foreign key such as report_id.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Output workbook path to create.",
    )
    parser.add_argument(
        "--reference-sheet",
        default=None,
        help="Optional reference sheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--related-sheet",
        default=None,
        help="Optional related sheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--reference-key",
        default="id",
        help="Reference workbook primary key column. Defaults to %(default)s.",
    )
    parser.add_argument(
        "--related-key",
        default="report_id",
        help="Related workbook foreign key column. Defaults to %(default)s.",
    )
    parser.add_argument(
        "--reference-fields",
        default=",".join(DEFAULT_REFERENCE_FIELDS),
        help="Comma-separated reference fields to carry into the output.",
    )
    parser.add_argument(
        "--numeric-fields",
        default="",
        help="Optional comma-separated related columns to force as numeric summary fields.",
    )
    return parser.parse_args()


def parse_csv(raw: str) -> list[str]:
    return [part.strip() for part in (raw or "").split(",") if part.strip()]


def safe_get(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def normalize_report_id(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def load_rows(path: str, sheet_name: str | None = None) -> tuple[list[tuple[Any, ...]], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    selected_sheet = worksheet.title
    workbook.close()
    return rows, selected_sheet


def infer_numeric_column(values: list[Any]) -> bool:
    seen = False
    for value in values:
        if value in (None, "") or isinstance(value, bool):
            continue
        if isinstance(value, (int, float)):
            seen = True
            continue
        try:
            float(str(value).strip())
        except (TypeError, ValueError):
            return False
        seen = True
    return seen


def to_float(value: Any) -> float:
    if value in (None, "") or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return 0.0


def ordered_unique(values: list[Any]) -> str:
    seen: dict[str, None] = {}
    for value in values:
        if value in (None, ""):
            continue
        seen[str(value)] = None
    return ", ".join(seen.keys())


def should_currency_format(name: str) -> bool:
    lowered = name.lower()
    keywords = (
        "amount",
        "total",
        "fee",
        "income",
        "receipt",
        "balance",
        "spending",
        "penalty",
        "grant",
        "donation",
    )
    return any(keyword in lowered for keyword in keywords)


def is_date_like(value: Any) -> bool:
    return isinstance(value, (date, datetime))


def autosize_and_format(worksheet) -> None:
    widths: dict[int, int] = {}
    headers = [cell.value for cell in worksheet[1]]
    for col_index, header in enumerate(headers, start=1):
        widths[col_index] = max(len(str(header)), 12)
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=col_index)
            if cell.value is not None:
                widths[col_index] = min(max(widths[col_index], len(str(cell.value)) + 2), 60)
                if should_currency_format(str(header)):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00"
                elif is_date_like(cell.value):
                    cell.number_format = "yyyy-mm-dd"
    for col_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(col_index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)


def build_reference_map(
    reference_rows: list[tuple[Any, ...]],
    reference_key: str,
    reference_fields: list[str],
) -> tuple[dict[int, dict[str, Any]], list[str]]:
    if not reference_rows:
        raise ValueError("Reference workbook is empty.")
    header = list(reference_rows[0])
    index_by_name = {name: idx for idx, name in enumerate(header)}
    if reference_key not in index_by_name:
        raise ValueError(f"Reference key column '{reference_key}' was not found.")

    available_fields = [field for field in reference_fields if field in index_by_name]
    report_map: dict[int, dict[str, Any]] = {}
    for row in reference_rows[1:]:
        report_id = normalize_report_id(safe_get(row, index_by_name[reference_key]))
        if report_id is None:
            continue
        report_map[report_id] = {
            field: safe_get(row, index_by_name.get(field)) for field in available_fields
        }
    return report_map, available_fields


def merge_related_rows(
    related_rows: list[tuple[Any, ...]],
    report_map: dict[int, dict[str, Any]],
    related_key: str,
    forced_numeric_fields: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str], list[str], list[str]]:
    if not related_rows:
        raise ValueError("Related workbook is empty.")

    related_header = list(related_rows[0])
    related_index = {name: idx for idx, name in enumerate(related_header)}
    if related_key not in related_index:
        raise ValueError(f"Related key column '{related_key}' was not found.")

    related_non_key_columns = [name for name in related_header if name != related_key]
    matched: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []

    for row in related_rows[1:]:
        report_id = normalize_report_id(safe_get(row, related_index[related_key]))
        related_values = {
            column: safe_get(row, related_index[column]) for column in related_non_key_columns
        }
        if report_id is None or report_id not in report_map:
            unmatched.append({related_key: report_id, **related_values})
            continue
        matched.append(
            {
                "report_id": report_id,
                "reference": report_map[report_id],
                "related": related_values,
            }
        )

    summary_candidate_columns = [
        name for name in related_non_key_columns if name not in {"id", related_key}
    ]
    if forced_numeric_fields:
        numeric_columns = [
            column for column in forced_numeric_fields if column in summary_candidate_columns
        ]
    else:
        numeric_columns = [
            column
            for column in summary_candidate_columns
            if infer_numeric_column([row["related"].get(column) for row in matched])
        ]
    text_columns = [
        column for column in summary_candidate_columns if column not in numeric_columns
    ]
    return matched, unmatched, related_non_key_columns, numeric_columns, text_columns


def write_output(
    output_path: str,
    related_key: str,
    reference_fields: list[str],
    related_columns: list[str],
    matched: list[dict[str, Any]],
    unmatched: list[dict[str, Any]],
    numeric_columns: list[str],
    text_columns: list[str],
) -> None:
    workbook = Workbook()

    ws_detail = workbook.active
    ws_detail.title = "Matched Rows"
    detail_header = [related_key] + reference_fields + related_columns
    ws_detail.append(detail_header)
    for row in matched:
        ws_detail.append(
            [row["report_id"]]
            + [row["reference"].get(field) for field in reference_fields]
            + [row["related"].get(column) for column in related_columns]
        )

    report_summary: dict[int, dict[str, Any]] = {}
    pbo_summary: dict[str, dict[str, Any]] = {}

    for row in matched:
        report_id = row["report_id"]
        pbo_name = row["reference"].get("pbo_name") or "UNKNOWN PBO"

        report_bucket = report_summary.setdefault(
            report_id,
            {
                "reference": row["reference"],
                "row_count": 0,
                "numeric": defaultdict(float),
                "text": defaultdict(list),
            },
        )
        report_bucket["row_count"] += 1
        for column in numeric_columns:
            report_bucket["numeric"][column] += to_float(row["related"].get(column))
        for column in text_columns:
            report_bucket["text"][column].append(row["related"].get(column))

        pbo_bucket = pbo_summary.setdefault(
            pbo_name,
            {
                "report_ids": [],
                "row_count": 0,
                "numeric": defaultdict(float),
                "text": defaultdict(list),
            },
        )
        if report_id not in pbo_bucket["report_ids"]:
            pbo_bucket["report_ids"].append(report_id)
        pbo_bucket["row_count"] += 1
        for column in numeric_columns:
            pbo_bucket["numeric"][column] += to_float(row["related"].get(column))
        for column in text_columns:
            pbo_bucket["text"][column].append(row["related"].get(column))

    ws_report = workbook.create_sheet("Report Summary")
    report_header = [related_key] + reference_fields + ["matched_row_count"]
    report_header += [f"total_{column}" for column in numeric_columns]
    report_header += text_columns
    ws_report.append(report_header)
    for report_id in sorted(report_summary):
        bucket = report_summary[report_id]
        ws_report.append(
            [report_id]
            + [bucket["reference"].get(field) for field in reference_fields]
            + [bucket["row_count"]]
            + [bucket["numeric"].get(column, 0.0) for column in numeric_columns]
            + [ordered_unique(bucket["text"].get(column, [])) for column in text_columns]
        )

    ws_pbo = workbook.create_sheet("PBO Summary")
    pbo_header = ["pbo_name", "report_ids", "report_count", "matched_row_count"]
    pbo_header += [f"total_{column}" for column in numeric_columns]
    pbo_header += text_columns
    ws_pbo.append(pbo_header)
    for pbo_name in sorted(pbo_summary):
        bucket = pbo_summary[pbo_name]
        ws_pbo.append(
            [
                pbo_name,
                ", ".join(str(report_id) for report_id in bucket["report_ids"]),
                len(bucket["report_ids"]),
                bucket["row_count"],
            ]
            + [bucket["numeric"].get(column, 0.0) for column in numeric_columns]
            + [ordered_unique(bucket["text"].get(column, [])) for column in text_columns]
        )

    if unmatched:
        ws_unmatched = workbook.create_sheet("Unmatched Rows")
        unmatched_header = [related_key] + related_columns
        ws_unmatched.append(unmatched_header)
        for row in unmatched:
            ws_unmatched.append([row.get(related_key)] + [row.get(column) for column in related_columns])

    for worksheet in workbook.worksheets:
        autosize_and_format(worksheet)

    workbook.save(output_path)


def main() -> None:
    args = parse_args()

    reference_fields = parse_csv(args.reference_fields)
    reference_rows, reference_sheet = load_rows(args.reference, args.reference_sheet)
    related_rows, related_sheet = load_rows(args.related, args.related_sheet)

    report_map, available_reference_fields = build_reference_map(
        reference_rows=reference_rows,
        reference_key=args.reference_key,
        reference_fields=reference_fields,
    )
    matched, unmatched, related_columns, numeric_columns, text_columns = merge_related_rows(
        related_rows=related_rows,
        report_map=report_map,
        related_key=args.related_key,
        forced_numeric_fields=parse_csv(args.numeric_fields),
    )
    write_output(
        output_path=args.output,
        related_key=args.related_key,
        reference_fields=available_reference_fields,
        related_columns=related_columns,
        matched=matched,
        unmatched=unmatched,
        numeric_columns=numeric_columns,
        text_columns=text_columns,
    )

    print(f"Reference workbook: {Path(args.reference).name} [{reference_sheet}]")
    print(f"Related workbook: {Path(args.related).name} [{related_sheet}]")
    print(f"Output workbook: {args.output}")
    print(f"Matched rows: {len(matched)}")
    print(f"Unmatched rows: {len(unmatched)}")
    print(f"Numeric summary columns: {', '.join(numeric_columns) or 'None'}")
    print(f"Text summary columns: {', '.join(text_columns) or 'None'}")


if __name__ == "__main__":
    main()
