import io
import os
import re
import time
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect, text
from sqlalchemy.exc import IntegrityError
from werkzeug.datastructures import FileStorage
from werkzeug.security import generate_password_hash


def create_user(db, User, email, password, role="user", is_authorized=True, is_superadmin=False, can_manage_all_records=False, department=None):
    user = User(
        email=email,
        role=role,
        is_authorized=is_authorized,
        is_superadmin=is_superadmin,
        can_manage_all_records=can_manage_all_records,
        department=department,
    )
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    return user


def login(client, email, password, department=""):
    return client.post("/login", data={"email": email, "password": password, "department": department}, follow_redirects=True)


def create_and_login_user(app, client, models, email, password, **kwargs):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, email, password, **kwargs)

    response = login(client, email, password)
    assert response.status_code == 200
    return response


def test_backup_export_uses_sqlalchemy_selectables_for_table_reads(app, models, monkeypatch, tmp_path):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "backup-reader@example.com", "BackupReader@123")

    import database_backup as backup_module

    captured_queries = []
    real_read_sql = backup_module.pd.read_sql

    def tracking_read_sql(query, connection, *args, **kwargs):
        captured_queries.append(query)
        return real_read_sql(query, connection, *args, **kwargs)

    monkeypatch.setattr(backup_module.pd, "read_sql", tracking_read_sql)

    with app.app_context():
        files = backup_module.export_database_tables(
            app.config["SQLALCHEMY_DATABASE_URI"],
            tmp_path,
        )

    assert files
    assert captured_queries
    assert all(query.__class__.__name__ != "TextClause" for query in captured_queries)


def test_create_mysql_sql_dump_writes_expected_filename_and_command(monkeypatch, tmp_path):
    import database_backup as backup_module

    captured = {}

    def fake_run(cmd, stdout=None, stderr=None, env=None):
        captured["cmd"] = cmd
        captured["env"] = env
        stdout.write(b"-- mysql dump --\n")
        return SimpleNamespace(returncode=0, stderr=b"")

    monkeypatch.setattr(backup_module.subprocess, "run", fake_run)

    dump_path = backup_module.create_mysql_sql_dump(
        "mysql+pymysql://dump_user:dump_password@db.example.com:3307/form14?charset=utf8mb4",
        tmp_path,
    )

    assert dump_path is not None
    assert dump_path.name == "returnsform14_org_backup.mysql.sql"
    assert dump_path.read_text(encoding="utf-8") == "-- mysql dump --\n"
    assert captured["cmd"][0] == "mysqldump"
    assert "--single-transaction" in captured["cmd"]
    assert captured["cmd"][-1] == "form14"
    assert captured["env"]["MYSQL_PWD"] == "dump_password"


def test_database_dump_bootstrap_mode_maps_supported_extensions(app_code):
    assert app_code.database_dump_bootstrap_mode(Path("seed.mysql.sql"), target_dialect="mysql") == "mysql_sql"
    assert app_code.database_dump_bootstrap_mode(Path("seed.dump"), target_dialect="postgresql") == "postgresql_dump"
    assert app_code.database_dump_bootstrap_mode(Path("seed.sql"), target_dialect="postgresql") == "postgresql_dump"
    assert app_code.database_dump_bootstrap_mode(Path("seed.mysql.sql"), target_dialect="postgresql") is None
    assert app_code.database_dump_bootstrap_mode(Path("seed.sqlite"), target_dialect="sqlite") == "sqlite_file"
    assert app_code.database_dump_bootstrap_mode(Path("seed.sql"), target_dialect="sqlite") is None


def test_record_database_dump_bootstrap_persists_admin_setting(app, app_code):
    with app.app_context():
        app_code.record_database_dump_bootstrap(
            {
                "performed": True,
                "source_path": "/tmp/bootstrap.sql",
                "bootstrap_mode": "mysql_sql",
                "target_dialect": "mysql",
            }
        )

        stored_value = app_code.admin_setting_value(app_code.DATABASE_DUMP_BOOTSTRAP_SETTING_KEY)

    assert "bootstrap.sql" in stored_value
    assert "mysql_sql" in stored_value


def test_google_drive_credential_details_uses_google_application_credentials_even_when_service_account_file_exists(monkeypatch, tmp_path):
    import database_backup as backup_module

    service_account_path = tmp_path / "service-account.json"
    authorized_user_path = tmp_path / "application_default_credentials.json"
    service_account_path.write_text('{"type": "service_account"}', encoding="utf-8")
    authorized_user_path.write_text('{"type": "authorized_user"}', encoding="utf-8")

    monkeypatch.setenv("GOOGLE_DRIVE_SERVICE_ACCOUNT_FILE", str(service_account_path))
    monkeypatch.setenv("GOOGLE_APPLICATION_CREDENTIALS", str(authorized_user_path))

    details = backup_module.google_drive_credential_details()

    assert details["type"] == "authorized_user"
    assert details["source"] == "GOOGLE_APPLICATION_CREDENTIALS"
    assert details["path"] == str(authorized_user_path)


def test_google_drive_auth_guidance_returns_reauthorize_command_for_invalid_grant(monkeypatch, tmp_path):
    import database_backup as backup_module

    authorized_user_path = tmp_path / "application_default_credentials.json"
    client_secret_path = tmp_path / "client_secret_test.json"
    authorized_user_path.write_text('{"type": "authorized_user"}', encoding="utf-8")
    client_secret_path.write_text('{"installed": {"client_id": "abc", "project_id": "demo"}}', encoding="utf-8")

    monkeypatch.setenv("GOOGLE_APPLICATION_CREDENTIALS", str(authorized_user_path))
    monkeypatch.delenv("GOOGLE_DRIVE_SERVICE_ACCOUNT_FILE", raising=False)
    monkeypatch.setenv("GOOGLE_DRIVE_OAUTH_CLIENT_SECRET_FILE", str(client_secret_path))

    friendly_error = backup_module.describe_google_drive_error(
        RuntimeError("invalid_grant: Token has been expired or revoked.")
    )
    guidance = backup_module.google_drive_auth_guidance(error_message=friendly_error)

    assert "Google Drive authorization is invalid" in friendly_error
    assert guidance is not None
    assert guidance["refresh_command"].endswith("--reauthorize-google-drive")
    assert "GOOGLE_APPLICATION_CREDENTIALS" in guidance["recommendation"]


def test_admin_user_activity_can_filter_and_export_by_day(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import UserActivityLog

        create_user(db, User, "activity-admin@example.com", "ActivityAdmin@123", role="admin")
        matching_user = create_user(db, User, "match-user@example.com", "MatchUser@123")
        other_user = create_user(db, User, "other-user@example.com", "OtherUser@123")

        matching_report = PBOReport(pbo_name="Matching Report", user_id=matching_user.id)
        other_report = PBOReport(pbo_name="Other Report", user_id=other_user.id)
        db.session.add_all([matching_report, other_report])
        db.session.commit()

        db.session.add_all([
            UserActivityLog(
                user_id=matching_user.id,
                report_id=matching_report.id,
                action="report_updated",
                created_at=datetime(2026, 3, 20, 10, 30, 0),
            ),
            UserActivityLog(
                user_id=other_user.id,
                report_id=other_report.id,
                action="report_updated",
                created_at=datetime(2026, 3, 21, 9, 15, 0),
            ),
        ])
        db.session.commit()

    login(client, "activity-admin@example.com", "ActivityAdmin@123")

    panel_response = client.get("/admin/user-activity-panel?activity_date=2026-03-20")

    assert panel_response.status_code == 200
    assert b"match-user@example.com" in panel_response.data
    assert b"other-user@example.com" not in panel_response.data
    assert b"MATCHING REPORT" in panel_response.data

    export_response = client.get("/admin/user-activity/export?activity_date=2026-03-20")

    assert export_response.status_code == 200
    assert "user_activity_20260320_" in export_response.headers["Content-Disposition"]

    workbook = load_workbook(io.BytesIO(export_response.data))
    worksheet = workbook["User Activity"]
    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[0][0] == "Activity Date"
    assert rows[1][0] == "2026-03-20"
    assert rows[1][1] == "match-user@example.com"
    assert len(rows) == 2


def test_admin_user_activity_file_search_shows_matching_file_and_updater(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        admin = create_user(db, User, "activity-search-admin@example.com", "ActivitySearchAdmin@123", role="admin")
        updater = create_user(db, User, "file-updater@example.com", "FileUpdater@123")
        report = PBOReport(
            pbo_name="Searchable Activity File",
            pbo_registration_number="REG-ACTIVITY-SEARCH-001",
            user_id=updater.id,
            last_modified_by_id=updater.id,
        )
        db.session.add(report)
        db.session.commit()

    login(client, "activity-search-admin@example.com", "ActivitySearchAdmin@123")
    response = client.get("/admin?activity_search=Searchable Activity File")

    assert response.status_code == 200
    assert b"Searchable Activity File" in response.data
    assert b"file-updater@example.com" in response.data
    assert b"Updated By" in response.data


def test_build_admin_user_activity_rows_skips_none_activity_entries(app_code, monkeypatch):
    class SessionStub:
        def get(self, *args, **kwargs):
            return None

    class ActivityStub:
        def __init__(self, created_at, report_id=None):
            self.created_at = created_at
            self.report_id = report_id

    class UserStub:
        def __init__(self):
            self.id = 7
            self.email = "stub-user@example.com"
            self.reports = [None]
            self.activity_logs = [None, ActivityStub(datetime(2026, 3, 20, 8, 15, 0))]

    monkeypatch.setattr(app_code.db, "session", SessionStub())
    monkeypatch.setattr(app_code, "build_form_submitted_report_index", lambda *args, **kwargs: {})

    rows = app_code.build_admin_user_activity_rows([None, UserStub()])

    assert len(rows) == 1
    assert rows[0]["user"].email == "stub-user@example.com"
    assert len(rows[0]["recent_activities"]) == 1
    assert rows[0]["last_seen_at"] == datetime(2026, 3, 20, 8, 15, 0)


def test_build_admin_user_activity_rows_groups_duplicate_emails(app_code, monkeypatch):
    class ReportStub:
        def __init__(self, report_id, name):
            self.id = report_id
            self.pbo_name = name

    class SessionStub:
        def __init__(self, report_map):
            self.report_map = report_map

        def get(self, model, report_id):
            return self.report_map.get(report_id)

    class ActivityStub:
        def __init__(self, created_at, report_id=None):
            self.created_at = created_at
            self.report_id = report_id

    class UserStub:
        def __init__(self, user_id, email, role, reports, activity_logs, last_login_at=None):
            self.id = user_id
            self.email = email
            self.role = role
            self.reports = reports
            self.activity_logs = activity_logs
            self.is_authorized = True
            self.failed_login_attempts = 0
            self.last_login_at = last_login_at
            self.full_name = None
            self.department = None

    alpha_report = ReportStub(201, "Grouped Alpha")
    beta_report = ReportStub(202, "Grouped Beta")
    report_map = {
        201: alpha_report,
        202: beta_report,
    }
    duplicate_user_a = UserStub(
        1,
        "duplicate@example.com",
        "user",
        [alpha_report],
        [ActivityStub(datetime(2026, 3, 20, 8, 15, 0), 201)],
    )
    duplicate_user_b = UserStub(
        2,
        "duplicate@example.com",
        "admin",
        [beta_report],
        [ActivityStub(datetime(2026, 3, 20, 9, 30, 0), 202)],
        last_login_at=datetime(2026, 3, 20, 9, 45, 0),
    )

    monkeypatch.setattr(app_code.db, "session", SessionStub(report_map))
    monkeypatch.setattr(
        app_code,
        "build_form_submitted_report_index",
        lambda *args, **kwargs: {
            1: {201: datetime(2026, 3, 20, 8, 15, 0)},
            2: {
                201: datetime(2026, 3, 20, 9, 0, 0),
                202: datetime(2026, 3, 20, 9, 30, 0),
            },
        },
    )

    alpha_report.user_id = 1
    alpha_report.data_source = "form"
    alpha_report.submitted_at = datetime(2026, 3, 20, 8, 15, 0)
    alpha_report.created_at = datetime(2026, 3, 20, 8, 15, 0)
    beta_report.user_id = 2
    beta_report.data_source = "form"
    beta_report.submitted_at = datetime(2026, 3, 20, 9, 30, 0)
    beta_report.created_at = datetime(2026, 3, 20, 9, 30, 0)

    rows = app_code.build_admin_user_activity_rows([duplicate_user_a, duplicate_user_b])

    assert len(rows) == 1
    assert rows[0]["display_email"] == "duplicate@example.com"
    assert rows[0]["display_role"] == "admin"
    assert rows[0]["worked_file_count"] == 2
    assert rows[0]["group_size"] == 2
    assert {report.id for report in rows[0]["touched_reports"]} == {201, 202}


def test_select_preferred_existing_user_prefers_authorized_admin_account():
    import create_users as create_users_module

    class UserStub:
        def __init__(self, user_id, role, is_authorized, is_superadmin=False, can_manage_all_records=False, last_login_at=None):
            self.id = user_id
            self.role = role
            self.is_authorized = is_authorized
            self.is_superadmin = is_superadmin
            self.can_manage_all_records = can_manage_all_records
            self.password_hash = "hash"
            self.password_changed_at = None
            self.last_login_at = last_login_at
            self.full_name = None
            self.department = None

    stale_user = UserStub(9, "user", False)
    active_admin = UserStub(
        4,
        "admin",
        True,
        can_manage_all_records=True,
        last_login_at=datetime(2026, 3, 20, 10, 0, 0),
    )

    preferred = create_users_module.select_preferred_existing_user([stale_user, active_admin])

    assert preferred is active_admin


def test_global_field_help_context_is_public(client):
    response = client.get("/field-help/context?page_key=login")

    assert response.status_code == 200
    payload = response.get_json()
    assert isinstance(payload, dict)
    assert "required_fields" in payload


def test_analysis_route_is_disabled_in_crud_only_deployment(app, client, models, app_code):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "analysis-user@example.com", "AnalysisUser@123")

    login(client, "analysis-user@example.com", "AnalysisUser@123")
    response = client.get("/analysis")

    assert app_code.VISUAL_ANALYSIS_ENABLED is False
    assert response.status_code == 404


def test_reports_page_hides_analysis_links_in_crud_only_deployment(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "analysis-funding@example.com", "AnalysisFunding@123", role="admin")

    login(client, "analysis-funding@example.com", "AnalysisFunding@123")
    response = client.get("/reports")

    assert response.status_code == 200
    assert b"Analysis Table" not in response.data
    assert b"Data Interpretations" not in response.data
    assert b"Sector Report Data" not in response.data
    assert b"TF Risk" not in response.data


def test_data_interpretations_routes_are_disabled_in_crud_only_deployment(app, client, models, app_code):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "interpret-admin@example.com", "InterpretAdmin@123", role="admin")

    login(client, "interpret-admin@example.com", "InterpretAdmin@123")
    response = client.get("/datainterpretations")
    start_response = client.post("/datainterpretations", data={"fy": "2024-2025", "run_analysis": "1"})

    assert app_code.SECTOR_ANALYTICS_ENABLED is False
    assert response.status_code == 404
    assert start_response.status_code == 404


def test_data_interpretations_status_endpoint_is_disabled_in_crud_only_deployment(app, client, models, app_code):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "interpret-status-admin@example.com", "InterpretStatusAdmin@123", role="admin")

    login(client, "interpret-status-admin@example.com", "InterpretStatusAdmin@123")
    response = client.get("/datainterpretations/status?fy=all")

    assert app_code.SECTOR_ANALYTICS_ENABLED is False
    assert response.status_code == 404


def test_data_interpretations_uses_whole_database_for_admin_runs(app, client, models, app_code):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        admin = create_user(db, User, "interpret-whole-db-admin@example.com", "InterpretWholeDbAdmin@123", role="admin")
        other_user = create_user(db, User, "interpret-whole-db-owner@example.com", "InterpretWholeDbOwner@123")

        db.session.add(PBOReport(
            user_id=admin.id,
            pbo_name="Admin Scoped Report",
            reporting_period_start=datetime(2024, 7, 1).date(),
            reporting_period_end=datetime(2025, 6, 30).date(),
        ))
        db.session.add(PBOReport(
            user_id=other_user.id,
            pbo_name="Other User Report",
            reporting_period_start=datetime(2024, 7, 1).date(),
            reporting_period_end=datetime(2025, 6, 30).date(),
        ))
        db.session.commit()

        admin_user = db.session.get(User, admin.id)
        reports = app_code.data_interp_load_accessible_reports_for_user(admin_user)
        scope_label = app_code.data_interp_scope_label_for_user(admin_user)

    assert sorted(report.pbo_name for report in reports) == ["ADMIN SCOPED REPORT", "OTHER USER REPORT"]
    assert scope_label == "Analysis scoped to the full sector-report database"


def test_data_interpretations_flags_mismatch_checks(app, client, models, app_code):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import Donation, ProjectImplementation

        admin = create_user(db, User, "interpret-qa@example.com", "InterpretQa@123", role="admin")
        admin_id = admin.id
        report = PBOReport(
            user_id=admin.id,
            pbo_name="QA Drift NGO",
            reporting_period_start=datetime(2024, 7, 1).date(),
            reporting_period_end=datetime(2025, 6, 30).date(),
            date_of_registration=datetime(2024, 9, 1).date(),
            counties="Nairobi",
            income_b2_total=90000,
            receipts_total=95000,
            cash_balance_previous_year=5000,
        )
        db.session.add(report)
        db.session.flush()
        db.session.add(Donation(report_id=report.id, name="Donor Drift", amount=100000))
        db.session.add(
            ProjectImplementation(
                report_id=report.id,
                sector="Health",
                county="Nairobi",
                spending_per_county=150000,
                amount_spent_kenya=100000,
                amount_spent_other=30000,
            )
        )
        db.session.commit()

    with app.app_context():
        admin_user = db.session.get(User, admin_id)
        reports = app_code.data_interp_load_accessible_reports_for_user(admin_user)
        scope_label = app_code.data_interp_scope_label_for_user(admin_user)
        analysis_result = app_code.build_data_interpretation_analysis(reports, "2024-2025", scope_label=scope_label)

    qa_checks = {check["title"]: check["status"] for check in analysis_result["qa_checks"]}
    assert analysis_result["qa_summary"]["status"] == "Mismatch Found"
    assert qa_checks["Funding Reconciliation"] == "Mismatch Found"
    assert qa_checks["Project Spend Row Reconciliation"] == "Mismatch Found"


def test_data_interpretation_analysis_zeroes_legacy_9999_placeholders(app, models, app_code):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import Donation, ProjectImplementation

        admin = create_user(db, User, "interpret-legacy-zero@example.com", "InterpretLegacy@123", role="admin")
        report = PBOReport(
            user_id=admin.id,
            pbo_name="Legacy Analysis NGO",
            reporting_period_start=datetime(2025, 7, 1).date(),
            reporting_period_end=datetime(2026, 6, 30).date(),
            created_at=datetime(2026, 3, 22, 12, 0, 0),
            updated_at=datetime(2026, 3, 22, 12, 0, 0),
            staff_kenyan_current=9999,
            volunteers_kenyan_current=9999,
            income_b2_total=9999,
            receipts_total=9999,
            cash_balance_previous_year=9999,
            cash_bank_balance=9999,
        )
        db.session.add(report)
        db.session.flush()
        db.session.add(Donation(report_id=report.id, name="Legacy Donor", category="Foundation/Trust", amount=9999))
        db.session.add(
            ProjectImplementation(
                report_id=report.id,
                sector="Health",
                beneficiaries_no=9999,
                spending_per_county=9999,
                amount_spent_kenya=9999,
                amount_spent_other=9999,
            )
        )
        db.session.commit()

        analysis = app_code.build_data_interpretation_analysis(
            [report],
            "2025-2026",
            scope_label="Analysis scoped to all accessible sector reports",
        )

    cards = {card["label"]: card["value"] for card in analysis["analysis_cards"]}
    assert cards["Funding Captured"] == "KES 0.00"
    assert cards["Project Spend"] == "KES 0.00"
    assert cards["Employment Footprint"] == "0"


def test_data_interpretation_analysis_uses_d3_collaboration_dropdown_and_saved_flags(app, models, app_code):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import CollaborationNetworking

        admin = create_user(db, User, "interpret-collab@example.com", "InterpretCollab@123", role="admin")
        report = PBOReport(
            user_id=admin.id,
            pbo_name="Collaboration Analysis NGO",
            reporting_period_start=datetime(2024, 7, 1).date(),
            reporting_period_end=datetime(2025, 6, 30).date(),
        )
        db.session.add(report)
        db.session.flush()
        db.session.add_all([
            CollaborationNetworking(
                report_id=report.id,
                partner_type="NGOs",
                info_exchange="NGOs",
                funding_from_partner="Yes",
            ),
            CollaborationNetworking(
                report_id=report.id,
                partner_type="Academic Institutions",
                tech_support_to_partner="Academic Institutions",
            ),
        ])
        db.session.commit()

        analysis = app_code.build_data_interpretation_analysis(
            [report],
            "2024-2025",
            scope_label="Analysis scoped to all accessible sector reports",
        )

    assert "Information Exchange" in analysis["topic_results"]["3.8"]["detail"]

    national_section = next(section for section in analysis["analysis_sections"] if section["title"] == "National Development Contribution")
    collaboration_table = next(table for table in national_section["tables"] if table["title"] == "Collaboration Profile by Partner Type")
    pbo_row = next(row for row in collaboration_table["rows"] if row["Partner Type"] == "PBOs")
    academic_row = next(row for row in collaboration_table["rows"] if row["Partner Type"] == "Academic Institutions")

    assert pbo_row["Collaboration Rows"] == "1"
    assert pbo_row["Information Exchange"] == "1"
    assert pbo_row["Funding From"] == "1"
    assert academic_row["Collaboration Rows"] == "1"
    assert academic_row["Technical Support To"] == "1"


def test_legacy_zero_output_structure_preserves_identifier_fields(app_code):
    legacy_report = SimpleNamespace(created_at=datetime(2026, 3, 22, 12, 0, 0))

    row = app_code.legacy_zero_output_structure(
        legacy_report,
        {
            "report_id": 9999,
            "user_id": 9999,
            "cash_bank_balance": 9999,
            "nested": {"ReportID": 9999, "Amount": 9999},
        },
    )

    assert row["report_id"] == 9999
    assert row["user_id"] == 9999
    assert row["cash_bank_balance"] == 0
    assert row["nested"]["ReportID"] == 9999
    assert row["nested"]["Amount"] == 0


def test_reports_nav_hides_disabled_analysis_links_for_admin(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "nav-admin@example.com", "NavAdmin@123", role="admin")

    login(client, "nav-admin@example.com", "NavAdmin@123")
    response = client.get("/reports")

    assert response.status_code == 200
    assert b"Data Interpretations" not in response.data
    assert b"/datainterpretations" not in response.data
    assert b"/sectorreportdata" not in response.data
    assert b"/admin/tf_risk" not in response.data


def test_data_interpretations_route_is_hidden_from_regular_users_in_crud_only_deployment(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "regular-viewer@example.com", "RegularViewer@123")

    login(client, "regular-viewer@example.com", "RegularViewer@123")
    response = client.get("/datainterpretations", follow_redirects=False)

    assert response.status_code == 404


def test_login_blocks_unauthorized_user(app, client, models):
    db = models["db"]
    User = models["User"]
    with app.app_context():
        create_user(db, User, "pending@example.com", "PendingUser@123", is_authorized=False)

    response = login(client, "pending@example.com", "PendingUser@123")

    assert response.status_code == 200
    assert b"pending admin authorization" in response.data


def test_report_child_rows_require_parent_report(app, models):
    db = models["db"]

    with app.app_context():
        from models import Donation

        db.session.add(Donation(name="Orphan Donor", amount=1000))
        with pytest.raises(ValueError, match="cannot be saved without a parent PBOReport"):
            db.session.flush()
        db.session.rollback()


def test_report_child_rows_save_when_attached_to_parent_report(app, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import Donation

        user = create_user(db, User, "child-parent@example.com", "ChildParent@123")
        report = PBOReport(user_id=user.id, pbo_name="Parent Report NGO")
        report.donations.append(Donation(name="Attached Donor", amount=2500))
        db.session.add(report)
        db.session.flush()

        assert report.id is not None
        assert len(report.donations) == 1
        assert report.donations[0].report_id == report.id
        db.session.rollback()


def test_logout_redirects_to_login_and_protected_pages_stay_blocked(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "logout-user@example.com", "LogoutUser@123")

    login_response = login(client, "logout-user@example.com", "LogoutUser@123")
    assert login_response.status_code == 200
    assert b"Logged in successfully." in login_response.data

    protected_response = client.get("/", follow_redirects=False)
    assert protected_response.status_code == 200

    logout_redirect = client.post("/logout", follow_redirects=False)
    assert logout_redirect.status_code == 302
    assert logout_redirect.headers["Location"].endswith("/login")
    assert logout_redirect.headers["Clear-Site-Data"] == '"cache", "cookies", "storage"'

    logout_response = client.get("/login", follow_redirects=True)
    assert logout_response.status_code == 200
    assert b"Login" in logout_response.data

    after_logout_home = client.get("/", follow_redirects=False)
    assert after_logout_home.status_code == 302
    assert "/login" in after_logout_home.headers["Location"]

    after_logout_api = client.get("/api/check-pbo-name?name=Example", follow_redirects=False)
    assert after_logout_api.status_code == 401
    assert after_logout_api.get_json()["error"] == "Authentication required"


def test_authenticated_user_is_redirected_away_from_login(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "redirect-user@example.com", "RedirectUser@123")

    login(client, "redirect-user@example.com", "RedirectUser@123")

    response = client.get("/login", follow_redirects=False)

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/")


def test_duplicate_submit_token_creates_one_report(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "duplicate-submit@example.com", "DuplicateSubmit@123")

    token = "TOKEN-ABC-123"
    payload = {
        "submission_token": token,
        "pbo_name": "Unique NGO Name",
        "pbo_registration_number": "REG-001",
    }

    first = client.post("/", data=payload, follow_redirects=False)
    second = client.post("/", data=payload, follow_redirects=False)

    assert first.status_code == 302
    assert second.status_code == 302

    with app.app_context():
        rows = PBOReport.query.filter_by(submission_token=token).all()
        assert len(rows) == 1


def test_form14_submission_saves_return_date(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "return-date-submit@example.com", "ReturnDateSubmit@123")

    response = client.post(
        "/",
        data={
            "submission_token": "TOKEN-RETURN-DATE-001",
            "pbo_name": "Return Date Test NGO",
            "reporting_period_start": "2026-01-01",
            "reporting_period_end": "2026-12-31",
            "return_date": "2026-07-15",
        },
        follow_redirects=False,
    )

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token="TOKEN-RETURN-DATE-001").one()
        assert report.return_date == datetime(2026, 7, 15).date()


def test_form14_get_renders_return_date_field_with_placeholder_default(app, client, models):
    create_and_login_user(app, client, models, "return-date-form@example.com", "ReturnDateForm@123")

    response = client.get("/")

    assert response.status_code == 200
    assert b'name="return_date"' in response.data
    assert b'value="9999-09-09"' in response.data


def test_reused_stale_submit_token_with_different_report_identity_creates_new_report(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "stale-token@example.com", "StaleToken@123")

    token = "TOKEN-STALE-001"
    first = client.post(
        "/",
        data={
            "submission_token": token,
            "pbo_name": "First NGO Name",
            "reporting_period_start": "2025-01-01",
            "reporting_period_end": "2025-12-31",
        },
        follow_redirects=False,
    )
    second = client.post(
        "/",
        data={
            "submission_token": token,
            "pbo_name": "Second NGO Name",
            "reporting_period_start": "2026-01-01",
            "reporting_period_end": "2026-12-31",
        },
        follow_redirects=False,
    )

    assert first.status_code == 302
    assert second.status_code == 302

    with app.app_context():
        reports = PBOReport.query.order_by(PBOReport.id.asc()).all()
        assert len(reports) == 2
        assert reports[0].submission_token == token
        assert reports[0].pbo_name == "FIRST NGO NAME"
        assert reports[1].submission_token != token
        assert reports[1].pbo_name == "SECOND NGO NAME"


def test_reused_stale_submit_token_logs_admin_audit_event(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "stale-token-audit@example.com", "StaleTokenAudit@123")

    token = "TOKEN-STALE-AUDIT-001"
    client.post(
        "/",
        data={
            "submission_token": token,
            "pbo_name": "Original NGO Name",
            "reporting_period_start": "2025-01-01",
            "reporting_period_end": "2025-12-31",
        },
        follow_redirects=False,
    )
    client.post(
        "/",
        data={
            "submission_token": token,
            "pbo_name": "Replacement NGO Name",
            "reporting_period_start": "2026-01-01",
            "reporting_period_end": "2026-12-31",
        },
        follow_redirects=False,
    )

    with app.app_context():
        from models import UserActivityLog

        reports = PBOReport.query.order_by(PBOReport.id.asc()).all()
        audit_entry = (
            UserActivityLog.query
            .filter_by(action="form14_stale_submission_token_detected")
            .order_by(UserActivityLog.id.desc())
            .first()
        )

        assert len(reports) == 2
        assert audit_entry is not None
        assert audit_entry.report_id == reports[0].id
        assert "existing_report_id=" in (audit_entry.summary or "")
        assert "ORIGINAL NGO NAME" in (audit_entry.summary or "")
        assert "REPLACEMENT NGO NAME" in (audit_entry.summary or "")


def test_form14_home_response_disables_caching(app, client, models):
    create_and_login_user(app, client, models, "form-cache@example.com", "FormCache@123")

    response = client.get("/")

    assert response.status_code == 200
    assert "no-store" in response.headers["Cache-Control"]
    assert "max-age=0" in response.headers["Cache-Control"]
    assert response.headers["Pragma"] == "no-cache"


def test_form_submission_saves_direct_posted_b6_project_implementation_fields(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "b6-direct@example.com", "B6Direct@123")

    token = "TOKEN-B6-DIRECT-001"
    payload = {
        "submission_token": token,
        "pbo_name": "Direct B6 NGO",
        "pbo_registration_number": "REG-B6-001",
        "project_sector[]": ["Other"],
        "project_sector_other[]": ["Climate Action"],
        "project_county[]": ["Nairobi City"],
        "project_vulnerable_group[]": ["OTHERS"],
        "project_vulnerable_group_other[]": ["Urban youth"],
        "project_beneficiaries_no[]": ["125"],
        "project_spending_per_county[]": ["245000.50"],
        "project_duration_years[]": ["2"],
        "project_completion_status[]": ["Ongoing"],
        "project_amount_spent_kenya[]": ["200000"],
        "project_amount_spent_other[]": ["45000.50"],
    }

    response = client.post("/", data=payload, follow_redirects=False)

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert len(report.project_implementations) == 1

        row = report.project_implementations[0]
        assert row.sector == "CLIMATE ACTION"
        assert row.county == "NAIROBI CITY"
        assert row.vulnerable_group == "URBAN YOUTH"
        assert row.beneficiaries_no == 125
        assert row.spending_per_county == 245000.50
        assert row.duration_years == 2
        assert row.completion_status == "ONGOING"
        assert row.amount_spent_kenya == 200000
        assert row.amount_spent_other == 45000.50


def test_form_submission_accepts_decimal_b3_payment_values(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        create_user(db, User, "decimal-home@example.com", "DecimalHome@123")

    login_response = login(client, "decimal-home@example.com", "DecimalHome@123")
    assert login_response.status_code == 200

    token = "TOKEN-B3-DECIMAL-001"
    response = client.post(
        "/",
        data={
            "submission_token": token,
            "pbo_name": "Decimal Payment NGO",
            "pbo_registration_number": "REG-B3-001",
            "payment_description[]": ["Direct Programme Cost"],
            "payment_kenya[]": ["12345.67"],
            "payment_other[]": ["890.45"],
        },
        follow_redirects=False,
    )

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert len(report.payments) == 1
        assert report.payments[0].description == "DIRECT PROGRAMME COST"
        assert report.payments[0].kenya_amount == pytest.approx(12345.67)
        assert report.payments[0].other_amount == pytest.approx(890.45)


def test_form_submission_accepts_comma_separated_amount_values(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "comma-home@example.com", "CommaHome@123")

    token = "TOKEN-COMMA-AMOUNTS-001"
    response = client.post(
        "/",
        data={
            "submission_token": token,
            "pbo_name": "Comma Amount NGO",
            "pbo_registration_number": "REG-COMMA-001",
            "cash_balance_previous_year": "12,345.67",
            "cash_bank_balance": "7,654.32",
            "donor_name[]": ["Example Donor"],
            "donor_category[]": ["Corporate Donors"],
            "donor_country[]": ["Kenya"],
            "donor_amount[]": ["45,000.75"],
            "payment_description[]": ["Direct Programme Cost"],
            "payment_kenya[]": ["1,234.50"],
            "payment_other[]": ["90.25"],
        },
        follow_redirects=False,
    )

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert report.cash_balance_previous_year == pytest.approx(12345.67)
        assert report.cash_bank_balance == pytest.approx(7654.32)
        assert len(report.donations) == 1
        assert report.donations[0].amount == pytest.approx(45000.75)
        assert len(report.payments) == 1
        assert report.payments[0].kenya_amount == pytest.approx(1234.50)
        assert report.payments[0].other_amount == pytest.approx(90.25)


def test_form_submission_accepts_all_and_other_donor_country_values(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "donor-country-home@example.com", "DonorCountryHome@123")

    token = "TOKEN-DONOR-COUNTRY-001"
    response = client.post(
        "/",
        data={
            "submission_token": token,
            "pbo_name": "Donor Country NGO",
            "pbo_registration_number": "REG-DONOR-COUNTRY-001",
            "donor_name[]": ["Global Fund", "Special Donor"],
            "donor_category[]": ["Foundation/Trust", "Corporate Donors"],
            "donor_country[]": ["All Countries", "OTHER_COUNTRY"],
            "donor_country_other[]": ["", "Atlantis"],
            "donor_amount[]": ["1,000.00", "2,000.00"],
        },
        follow_redirects=False,
    )

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert len(report.donations) == 2
        assert report.donations[0].country == "ALL COUNTRIES"
        assert report.donations[1].country == "ATLANTIS"


def test_section_b_preview_accepts_comma_separated_amount_values(app, client, models):
    create_and_login_user(app, client, models, "comma-preview@example.com", "CommaPreview@123")

    response = client.post(
        "/api/section-b-preview",
        json={
            "cash_balance_previous_year": "12,345.67",
            "donor_amounts": ["1,200.50", "300.00"],
            "iga_amounts": [],
            "payment_kenya_amounts": ["2,000.00", "15.50"],
            "payment_other_amounts": ["150.25", ""],
            "project_kenya_amounts": ["3,250.40"],
            "project_other_amounts": ["75.10"],
        },
    )

    assert response.status_code == 200

    preview = response.get_json()
    assert preview["donor_total"] == pytest.approx(1500.50)
    assert preview["income_b2_total"] == pytest.approx(1500.50)
    assert preview["receipts_total"] == pytest.approx(13846.17)
    assert preview["payments_kenya_total"] == pytest.approx(2015.50)
    assert preview["payments_other_total"] == pytest.approx(150.25)
    assert preview["payments_total"] == pytest.approx(2165.75)
    assert preview["payment_row_totals"] == pytest.approx([2150.25, 15.50])
    assert preview["project_kenya_total"] == pytest.approx(3250.40)
    assert preview["project_other_total"] == pytest.approx(75.10)


def test_report_edit_accepts_decimal_b3_payment_values(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        user = create_user(db, User, "decimal-editor@example.com", "DecimalEdit@123")
        report = PBOReport(
            user_id=user.id,
            pbo_name="Editable Decimal NGO",
            pbo_registration_number="REG-EDIT-B3-001",
            workflow_status="draft",
            review_status="pending",
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login_response = login(client, "decimal-editor@example.com", "DecimalEdit@123")
    assert login_response.status_code == 200

    response = client.post(
        f"/report/{report_id}/edit",
        data={
            "pbo_name": "Editable Decimal NGO",
            "pbo_registration_number": "REG-EDIT-B3-001",
            "payment_description[]": ["Direct Programme Cost"],
            "payment_kenya_amount[]": ["5000.25"],
            "payment_other_amount[]": ["125.75"],
        },
        follow_redirects=False,
    )

    assert response.status_code == 302

    with app.app_context():
        report = db.session.get(PBOReport, report_id)
        assert report is not None
        assert len(report.payments) == 1
        assert report.payments[0].description == "DIRECT PROGRAMME COST"
        assert report.payments[0].kenya_amount == pytest.approx(5000.25)
        assert report.payments[0].other_amount == pytest.approx(125.75)


def test_form_submission_saves_countries_of_operation_and_sector_export_helper_includes_it(app, client, models, app_code):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "countries-ops@example.com", "CountriesOps@123")

    token = "TOKEN-COUNTRIES-OPS-001"
    payload = {
        "submission_token": token,
        "pbo_name": "Countries of Operation NGO",
        "pbo_registration_number": "REG-COUNTRIES-001",
        "country_of_operation": ["Kenya", "Uganda", "Tanzania"],
        "county": ["Nairobi City", "Kisumu"],
    }

    response = client.post("/", data=payload, follow_redirects=False)

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert report.countries_of_operation == "KENYA, UGANDA, TANZANIA"
        assert report.counties == "NAIROBI CITY, KISUMU"

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        export_rows = app_code.build_sector_report_rows([report])

    matching_row = next(row for row in export_rows if row["pbo_name"] == "COUNTRIES OF OPERATION NGO")
    assert matching_row["countries_of_operation"] == "KENYA, UGANDA, TANZANIA"
    assert matching_row["counties"] == "NAIROBI CITY, KISUMU"


def test_form_submission_mirrors_section_c_summary_into_child_tables(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "section-c-summary@example.com", "SectionCSummary@123")

    token = "TOKEN-SECTION-C-SUMMARY-001"
    response = client.post(
        "/",
        data={
            "submission_token": token,
            "pbo_name": "Section C Summary NGO",
            "pbo_registration_number": "REG-SECTION-C-001",
            "staff_kenyan[]": ["12", "15", "4", "1"],
            "staff_foreign[]": ["2", "3", "1", "0"],
            "staff_other_kenyan[]": ["5", "6"],
            "staff_other_foreign[]": ["1", "2"],
            "volunteers_kenyan[]": ["8", "9"],
            "volunteers_foreign[]": ["3", "4"],
        },
        follow_redirects=False,
    )

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None

        staff_rows = {
            (row.category, row.prev_year, row.curr_year)
            for row in report.staff_biodata
            if row.category and row.category.startswith("SECTION_C SUMMARY |")
        }
        volunteer_rows = {
            (row.category, row.prev_year, row.curr_year)
            for row in report.volunteer_biodata
            if row.category and row.category.startswith("SECTION_C SUMMARY |")
        }
        demographic_staff_rows = {
            (row.category, row.prev_year, row.curr_year)
            for row in report.staff_biodata
            if row.category and not row.category.startswith("SECTION_C SUMMARY |")
        }
        demographic_volunteer_rows = {
            (row.category, row.prev_year, row.curr_year)
            for row in report.volunteer_biodata
            if row.category and not row.category.startswith("SECTION_C SUMMARY |")
        }

        assert (
            "SECTION_C SUMMARY | STAFF STATIONED IN KENYA | KENYAN | PREVIOUS/CURRENT",
            12,
            15,
        ) in staff_rows
        assert (
            "SECTION_C SUMMARY | STAFF STATIONED IN KENYA | FOREIGN | CAME IN/LEFT",
            1,
            0,
        ) in staff_rows
        assert (
            "SECTION_C SUMMARY | STAFF STATIONED IN OTHER COUNTRIES | FOREIGN | PREVIOUS/CURRENT",
            1,
            2,
        ) in staff_rows
        assert (
            "SECTION_C SUMMARY | VOLUNTEERS / INTERNS | KENYAN | PREVIOUS/CURRENT",
            8,
            9,
        ) in volunteer_rows
        assert (
            "SECTION_C SUMMARY | VOLUNTEERS / INTERNS | FOREIGN | PREVIOUS/CURRENT",
            3,
            4,
        ) in volunteer_rows
        assert demographic_staff_rows == {
            ("BELOW35", 0, 0),
            ("ABOVE35", 0, 0),
            ("MALE", 0, 0),
            ("FEMALE", 0, 0),
            ("PWD", 0, 0),
        }
        assert demographic_volunteer_rows == {
            ("BELOW35", 0, 0),
            ("ABOVE35", 0, 0),
            ("MALE", 0, 0),
            ("FEMALE", 0, 0),
            ("PWD", 0, 0),
        }


def test_report_edit_skips_blank_hidden_biodata_rows_but_keeps_section_c_summary_rows(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        user = create_user(db, User, "section-c-edit@example.com", "SectionCEdit@123")
        report = PBOReport(
            user_id=user.id,
            pbo_name="Section C Edit NGO",
            pbo_registration_number="REG-SECTION-C-EDIT-001",
            workflow_status="draft",
            review_status="pending",
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login(client, "section-c-edit@example.com", "SectionCEdit@123")
    response = client.post(
        f"/report/{report_id}/edit",
        data={
            "pbo_name": "Section C Edit NGO",
            "pbo_registration_number": "REG-SECTION-C-EDIT-001",
            "staff_kenyan[]": ["3", "4", "1", "0"],
            "staff_foreign[]": ["1", "2", "0", "0"],
            "staff_other_kenyan[]": ["0", "0"],
            "staff_other_foreign[]": ["0", "0"],
            "volunteers_kenyan[]": ["2", "3"],
            "volunteers_foreign[]": ["1", "1"],
            "biodata_item[]": ["Below35", "Above35", "Male", "Female", "PWD"],
            "prev-year5[]": ["", "", "", "", ""],
            "curr-year5[]": ["", "", "", "", ""],
            "volbiodata_item[]": ["volBelow35", "volAbove35", "Male", "Female", "PWD"],
            "prev-volntr[]": ["", "", "", "", ""],
            "curr-volntr[]": ["", "", "", "", ""],
        },
        follow_redirects=False,
    )

    assert response.status_code == 302

    with app.app_context():
        report = db.session.get(PBOReport, report_id)
        assert report is not None

        demographic_staff_rows = [
            row for row in report.staff_biodata
            if row.category and not row.category.startswith("SECTION_C SUMMARY |")
        ]
        demographic_volunteer_rows = [
            row for row in report.volunteer_biodata
            if row.category and not row.category.startswith("SECTION_C SUMMARY |")
        ]
        summary_staff_rows = [
            row for row in report.staff_biodata
            if row.category and row.category.startswith("SECTION_C SUMMARY |")
        ]
        summary_volunteer_rows = [
            row for row in report.volunteer_biodata
            if row.category and row.category.startswith("SECTION_C SUMMARY |")
        ]

        assert sorted((row.category, row.prev_year, row.curr_year) for row in demographic_staff_rows) == [
            ("ABOVE35", 0, 0),
            ("BELOW35", 0, 0),
            ("FEMALE", 0, 0),
            ("MALE", 0, 0),
            ("PWD", 0, 0),
        ]
        assert sorted((row.category, row.prev_year, row.curr_year) for row in demographic_volunteer_rows) == [
            ("ABOVE35", 0, 0),
            ("BELOW35", 0, 0),
            ("FEMALE", 0, 0),
            ("MALE", 0, 0),
            ("PWD", 0, 0),
        ]
        assert len(summary_staff_rows) == 6
        assert len(summary_volunteer_rows) == 2


def test_section_c_biodata_normalizes_standard_variable_names(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        user = create_user(db, User, "section-c-normalize@example.com", "SectionCNormalize@123")
        report = PBOReport(
            user_id=user.id,
            pbo_name="Section C Normalize NGO",
            pbo_registration_number="REG-SECTION-C-NORMALIZE-001",
            workflow_status="draft",
            review_status="pending",
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login(client, "section-c-normalize@example.com", "SectionCNormalize@123")
    response = client.post(
        f"/report/{report_id}/edit",
        data={
            "pbo_name": "Section C Normalize NGO",
            "pbo_registration_number": "REG-SECTION-C-NORMALIZE-001",
            "staff_kenyan[]": ["5", "6", "0", "0"],
            "staff_foreign[]": ["1", "1", "0", "0"],
            "staff_other_kenyan[]": ["0", "0"],
            "staff_other_foreign[]": ["0", "0"],
            "volunteers_kenyan[]": ["4", "5"],
            "volunteers_foreign[]": ["1", "2"],
            "biodata_item[]": ["Below35", "Above35", "Male", "Female", "PWD"],
            "prev-year5[]": ["2", "3", "1", "4", ""],
            "curr-year5[]": ["3", "3", "2", "4", ""],
            "volbiodata_item[]": ["volBelow35", "volAbove35", "Male", "Female", "PWD"],
            "prev-volntr[]": ["1", "3", "2", "", ""],
            "curr-volntr[]": ["2", "3", "2", "", ""],
        },
        follow_redirects=False,
    )

    assert response.status_code == 302

    with app.app_context():
        report = db.session.get(PBOReport, report_id)
        assert report is not None

        staff_rows = sorted(
            (row.category, row.prev_year, row.curr_year)
            for row in report.staff_biodata
            if row.category and not row.category.startswith("SECTION_C SUMMARY |")
        )
        volunteer_rows = sorted(
            (row.category, row.prev_year, row.curr_year)
            for row in report.volunteer_biodata
            if row.category and not row.category.startswith("SECTION_C SUMMARY |")
        )

        assert staff_rows == [
            ("ABOVE35", 3, 3),
            ("BELOW35", 2, 3),
            ("FEMALE", 4, 4),
            ("MALE", 1, 2),
            ("PWD", 0, 0),
        ]
        assert volunteer_rows == [
            ("ABOVE35", 3, 3),
            ("BELOW35", 1, 2),
            ("FEMALE", 0, 0),
            ("MALE", 2, 2),
            ("PWD", 0, 0),
        ]


def test_report_detail_shows_section_c_total_columns(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        user = create_user(db, User, "section-c-detail@example.com", "SectionCDetail@123")
        report = PBOReport(
            user_id=user.id,
            pbo_name="Section C Detail NGO",
            pbo_registration_number="REG-SECTION-C-DETAIL-001",
            workflow_status="submitted",
            review_status="pending",
            staff_kenyan_prev=3,
            staff_foreign_prev=1,
            staff_kenyan_current=4,
            staff_foreign_current=2,
            staff_kenyan_came_in=1,
            staff_foreign_came_in=1,
            staff_kenyan_left=0,
            staff_foreign_left=1,
            staff_other_kenyan_prev=2,
            staff_other_foreign_prev=1,
            staff_other_kenyan_current=3,
            staff_other_foreign_current=1,
            volunteers_kenyan_prev=5,
            volunteers_foreign_prev=2,
            volunteers_kenyan_current=6,
            volunteers_foreign_current=3,
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login(client, "section-c-detail@example.com", "SectionCDetail@123")
    response = client.get(f"/report/{report_id}")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert html.count('title="Total column">Total') >= 3
    assert 'title="Total previous year">4<' in html
    assert 'title="Total current year">6<' in html
    assert 'title="Total previous year">7<' in html


def test_sector_report_helper_zeroes_legacy_9999_placeholders_but_keeps_dates(app, client, models, app_code):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import ProjectImplementation

        user = create_user(db, User, "legacy-zero-user@example.com", "LegacyZero@123")
        report = PBOReport(
            user_id=user.id,
            pbo_name="Legacy Placeholder NGO",
            reporting_period_start=datetime(2025, 7, 1).date(),
            reporting_period_end=datetime(2026, 6, 30).date(),
            cash_balance_previous_year=9999,
            cash_bank_balance=9999,
            staff_kenyan_current=9999,
            volunteers_kenyan_current=9999,
            created_at=datetime(2026, 3, 22, 9, 0, 0),
            updated_at=datetime(2026, 3, 22, 9, 0, 0),
        )
        db.session.add(report)
        db.session.flush()
        db.session.add(
            ProjectImplementation(
                report_id=report.id,
                sector="Health",
                beneficiaries_no=9999,
                spending_per_county=9999,
                amount_spent_kenya=9999,
                amount_spent_other=9999,
            )
        )
        db.session.commit()

    with app.app_context():
        report = PBOReport.query.filter_by(pbo_name="LEGACY PLACEHOLDER NGO").first()
        rows = app_code.build_sector_report_rows([report])

    matching_row = next(row for row in rows if row["pbo_name"] == "LEGACY PLACEHOLDER NGO")
    assert matching_row["cash_balance_previous_year"] == 0
    assert matching_row["cash_bank_balance"] == 0
    assert matching_row["staff_kenyan_current"] == 0
    assert matching_row["volunteers_kenyan_current"] == 0
    assert matching_row["project_beneficiaries_total"] == 0
    assert matching_row["project_spending_per_county_total"] == 0
    assert "22/03/2026" in matching_row["created_at"]


def test_projects_download_zeroes_legacy_9999_placeholders_but_keeps_reporting_dates(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import ProjectImplementation

        user = create_user(db, User, "legacy-project-export@example.com", "LegacyProject@123")
        report = PBOReport(
            user_id=user.id,
            pbo_name="Legacy Project Export NGO",
            reporting_period_start=datetime(2025, 7, 1).date(),
            reporting_period_end=datetime(2026, 6, 30).date(),
            created_at=datetime(2026, 3, 22, 8, 0, 0),
            updated_at=datetime(2026, 3, 22, 8, 0, 0),
        )
        db.session.add(report)
        db.session.flush()
        db.session.add(
            ProjectImplementation(
                report_id=report.id,
                sector="Education",
                county="Nairobi",
                beneficiaries_no=9999,
                spending_per_county=9999,
                duration_years=9999,
                amount_spent_kenya=9999,
                amount_spent_other=9999,
            )
        )
        db.session.commit()

    login(client, "legacy-project-export@example.com", "LegacyProject@123")
    response = client.get("/reports/download-projects-data?fy=2025-2026")

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.data))
    sheet = workbook["Implementations"]
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0][3] == "ReportingPeriodStart"
    assert rows[1][3] == "01/07/2025"
    assert rows[1][4] == "30/06/2026"
    assert rows[1][8] == 0
    assert rows[1][9] == 0
    assert rows[1][10] == 0
    assert rows[1][12] == 0
    assert rows[1][13] == 0


def test_form14_country_of_operation_widget_renders_all_and_other_options(app, client, models):
    create_and_login_user(app, client, models, "country-widget@example.com", "CountryWidget@123")

    response = client.get("/")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert 'value="All Countries"' in html
    assert 'value="Other Country (Specify)"' in html
    assert 'name="country_of_operation_other_toggle"' in html
    assert 'id="country-of-operation-custom-input"' in html


def test_form_submission_accepts_all_countries_of_operation_value(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "all-countries@example.com", "AllCountries@123")

    token = "TOKEN-ALL-COUNTRIES-001"
    response = client.post(
        "/",
        data={
            "submission_token": token,
            "pbo_name": "Global NGO",
            "pbo_registration_number": "REG-GLOBAL-001",
            "country_of_operation": ["All Countries"],
        },
        follow_redirects=False,
    )

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert report.countries_of_operation == "ALL COUNTRIES"


def test_report_edit_country_of_operation_widget_renders_all_and_other_options(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        user = create_user(db, User, "edit-country-widget@example.com", "EditCountryWidget@123")
        report = PBOReport(
            user_id=user.id,
            pbo_name="Editable Country Widget NGO",
            pbo_registration_number="REG-EDIT-COUNTRY-001",
            workflow_status="draft",
            review_status="pending",
            countries_of_operation="KENYA, UGANDA",
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login_response = login(client, "edit-country-widget@example.com", "EditCountryWidget@123")
    assert login_response.status_code == 200

    response = client.get(f"/report/{report_id}/edit")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert 'value="All Countries"' in html
    assert 'value="Other Country (Specify)"' in html
    assert 'name="country_of_operation_other_toggle"' in html
    assert 'id="edit-country-of-operation-custom-input"' in html


def test_report_edit_donor_country_widget_renders_and_persists_all_and_other_options(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import Donation

        user = create_user(db, User, "edit-donor-country@example.com", "EditDonorCountry@123")
        report = PBOReport(
            user_id=user.id,
            pbo_name="Editable Donor Country NGO",
            pbo_registration_number="REG-EDIT-DONOR-COUNTRY-001",
            workflow_status="draft",
            review_status="pending",
        )
        db.session.add(report)
        db.session.flush()
        report.donations.append(Donation(
            name="Legacy Donor",
            category="FOUNDATION/TRUST",
            country="ATLANTIS",
            amount=2500,
        ))
        db.session.commit()
        report_id = report.id

    login_response = login(client, "edit-donor-country@example.com", "EditDonorCountry@123")
    assert login_response.status_code == 200

    edit_response = client.get(f"/report/{report_id}/edit")

    assert edit_response.status_code == 200
    html = edit_response.get_data(as_text=True)
    assert 'value="All Countries"' in html
    assert 'value="OTHER_COUNTRY"' in html
    assert 'class="donor-country-other"' in html
    assert 'data-selected="ATLANTIS"' in html

    save_response = client.post(
        f"/report/{report_id}/edit",
        data={
            "pbo_name": "Editable Donor Country NGO",
            "pbo_registration_number": "REG-EDIT-DONOR-COUNTRY-001",
            "donor_name[]": ["Custom Donor", "Global Donor"],
            "donor_category[]": ["Foundation/Trust", "Corporate Donors"],
            "donor_category_other[]": ["", ""],
            "donor_country[]": ["OTHER_COUNTRY", "All Countries"],
            "donor_country_other[]": ["Atlantis", ""],
            "donor_amount[]": ["1200", "3400"],
        },
        follow_redirects=False,
    )

    assert save_response.status_code == 302

    with app.app_context():
        report = db.session.get(PBOReport, report_id)
        assert report is not None
        assert len(report.donations) == 2
        assert report.donations[0].country == "ATLANTIS"
        assert report.donations[1].country == "ALL COUNTRIES"


def test_report_edit_b6_sector_dropdown_includes_microfinance_and_keeps_custom_other_editable(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import ProjectImplementation

        user = create_user(db, User, "edit-b6-sector@example.com", "EditB6Sector@123")
        report = PBOReport(
            user_id=user.id,
            pbo_name="Editable B6 Sector NGO",
            pbo_registration_number="REG-EDIT-B6-SECTOR-001",
            workflow_status="draft",
            review_status="pending",
        )
        db.session.add(report)
        db.session.flush()
        report.project_implementations.append(ProjectImplementation(
            sector="CLIMATE ACTION",
            county="NAIROBI",
            beneficiaries_no=25,
        ))
        db.session.commit()
        report_id = report.id

    login_response = login(client, "edit-b6-sector@example.com", "EditB6Sector@123")
    assert login_response.status_code == 200

    response = client.get(f"/report/{report_id}/edit")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert 'Microfinance' in html
    custom_other_match = re.search(
        r'<input[^>]*name="project_sector_other\[\]"[^>]*value="CLIMATE ACTION"[^>]*>',
        html,
    )
    assert custom_other_match is not None
    assert 'disabled' not in custom_other_match.group(0)


def test_form14_b6_fields_render_simplified_and_hidden_columns_disabled(app, client, models):
    create_and_login_user(app, client, models, "b6-render@example.com", "B6Render@123")

    response = client.get("/")

    assert response.status_code == 200

    html = response.get_data(as_text=True)
    b6_fragment = html.split('id="project-implementation-section"', 1)[1].split('id="banking-details-section"', 1)[0]

    assert 'class="b6-hidden-col"' in b6_fragment
    assert re.search(
        r'name="project_county\[\]"\s+title="Choose the county for project implementation"\s+disabled',
        b6_fragment,
    )
    assert re.search(
        r'name="project_vulnerable_group\[\]"\s+class="vulnerable-group-select"',
        b6_fragment,
    )
    assert 'onchange="handleOthersOption(this)"' not in b6_fragment
    assert 'name="project_completion_status[]"' in b6_fragment
    assert re.search(
        r'name="project_completion_status\[\]"\s+title="Select current completion status of the project"\s+disabled',
        b6_fragment,
    )
    assert 'name="project_beneficiaries_no[]"' in b6_fragment
    assert 'name="project_spending_per_county[]"' in b6_fragment
    assert 'name="project_duration_years[]"' in b6_fragment
    assert 'name="project_vulnerable_group_other[]"' in b6_fragment
    assert 'name="project_amount_spent_kenya[]"' in b6_fragment
    assert 'name="project_amount_spent_other[]"' in b6_fragment


def test_form_submission_saves_direct_posted_official_fields(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "officials-direct@example.com", "OfficialsDirect@123")

    token = "TOKEN-OFFICIALS-DIRECT-001"
    payload = {
        "submission_token": token,
        "pbo_name": "Officials Direct NGO",
        "pbo_registration_number": "REG-OFF-001",
        "official_role[]": ["Chairperson", "Secretary", "Treasurer"],
        "official_name[]": ["Alice Akinyi", "Ben Otieno", "Carol Njeri"],
        "official_nationality[]": ["Kenya", "Uganda", "Tanzania"],
        "official_gender[]": ["Female", "Male", "Female"],
        "official_email[]": ["alice@example.com", "ben@example.com", "carol@example.com"],
        "official_residence[]": ["Nairobi", "Kisumu", "Mombasa"],
        "official_phone[]": ["+254700000001", "+254700000002", "+254700000003"],
        "official_kra_pin[]": ["A001", "B002", "C003"],
        "official_professional_qualification[]": ["Advocate", "Accountant", "Auditor"],
    }

    response = client.post("/", data=payload, follow_redirects=False)

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert len(report.officials) == 3

        first = report.officials[0]
        assert first.role == "CHAIRPERSON"
        assert first.name == "ALICE AKINYI"
        assert first.nationality == "KENYA"
        assert first.gender == "FEMALE"
        assert first.email == "ALICE@EXAMPLE.COM"
        assert first.residence == "NAIROBI"
        assert first.phone == "+254700000001"
        assert first.kra_pin == "A001"
        assert first.professional_qualification == "ADVOCATE"

        third = report.officials[2]
        assert third.role == "TREASURER"
        assert third.name == "CAROL NJERI"
        assert third.phone == "+254700000003"


def test_governance_election_frequency_prefills_on_edit_and_shows_in_detail(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        owner = create_user(db, User, "governance-owner@example.com", "OwnerUser@123")
        report = PBOReport(
            pbo_name="Governance Prefill NGO",
            pbo_registration_number="REG-GOV-001",
            user_id=owner.id,
            election_frequency="ANNUALLY",
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login(client, "governance-owner@example.com", "OwnerUser@123")

    edit_response = client.get(f"/report/{report_id}/edit")
    detail_response = client.get(f"/report/{report_id}")

    assert edit_response.status_code == 200
    assert detail_response.status_code == 200
    assert b'name="election_frequency[]" value="Annually" checked' in edit_response.data
    assert b"ANNUALLY" in detail_response.data


def test_governance_blank_blocks_save_as_null_and_election_frequency_from_create_form(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "governance-create@example.com", "GovernanceCreate@123")

    token = "TOKEN-GOV-NULLS-001"
    payload = {
        "submission_token": token,
        "pbo_name": "Governance Null NGO",
        "pbo_registration_number": "REG-GOV-NULL-001",
        "election_frequency": "Annually",
        "membership_number_of_directors": "",
        "membership_number_of_registered_members": "",
        "membership_number_of_board_meetings": "",
        "membership_date_last_agm": "",
        "membership_date_last_election": "",
        "non_membership_number_of_directors": "",
        "non_membership_number_of_board_meetings": "",
        "non_membership_date_last_board_meeting": "",
        "non_membership_date_last_election": "",
    }

    response = client.post("/", data=payload, follow_redirects=False)

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert report.election_frequency == "ANNUALLY"
        assert report.membership_number_of_directors is None
        assert report.membership_number_of_registered_members is None
        assert report.membership_number_of_board_meetings is None
        assert report.membership_date_last_agm is None
        assert report.membership_date_last_election is None
        assert report.non_membership_number_of_directors is None
        assert report.non_membership_number_of_board_meetings is None
        assert report.non_membership_date_last_board_meeting is None
        assert report.non_membership_date_last_election is None


def test_form_submission_saves_multiple_contact_emails(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "contact-emails@example.com", "ContactEmails@123")

    token = "TOKEN-CONTACT-EMAILS-001"
    payload = {
        "submission_token": token,
        "pbo_name": "Contact Email NGO",
        "pbo_registration_number": "REG-CONTACT-001",
        "contact_email[]": ["first@example.com", "second@example.com", "first@example.com"],
    }

    response = client.post("/", data=payload, follow_redirects=False)

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert report.contact_email == "FIRST@EXAMPLE.COM, SECOND@EXAMPLE.COM"


def test_form_submission_saves_bank_account_rows(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "bank-rows@example.com", "BankRows@123")

    token = "TOKEN-BANK-001"
    payload = {
        "submission_token": token,
        "pbo_name": "Banking NGO",
        "pbo_registration_number": "REG-BANK-001",
        "bank_name[]": ["KCB BANK KENYA"],
        "bank_name_other[]": [""],
        "bank_branch[]": ["Kisumu"],
        "bank_account_number[]": ["00123456789"],
        "bank_currency[]": ["KES"],
    }

    response = client.post("/", data=payload, follow_redirects=False)

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert len(report.bank_accounts) == 1
        bank = report.bank_accounts[0]
        assert bank.bank_name == "KCB BANK KENYA"
        assert bank.branch == "KISUMU"
        assert bank.account_number == "00123456789"
        assert bank.currency == "KES"


def test_form_submission_defaults_missing_bank_account_number_and_currency(app, client, models):
    PBOReport = models["PBOReport"]

    create_and_login_user(app, client, models, "bank-defaults@example.com", "BankDefaults@123")

    token = "TOKEN-BANK-DEFAULTS-001"
    payload = {
        "submission_token": token,
        "pbo_name": "Banking Defaults NGO",
        "pbo_registration_number": "REG-BANK-DEFAULTS-001",
        "bank_name[]": ["EQUITY BANK KENYA"],
        "bank_name_other[]": [""],
        "bank_branch[]": ["Nairobi"],
        "bank_account_number[]": [""],
        "bank_currency[]": [""],
    }

    response = client.post("/", data=payload, follow_redirects=False)

    assert response.status_code == 302

    with app.app_context():
        report = PBOReport.query.filter_by(submission_token=token).first()
        assert report is not None
        assert len(report.bank_accounts) == 1
        bank = report.bank_accounts[0]
        assert bank.bank_name == "EQUITY BANK KENYA"
        assert bank.branch == "NAIROBI"
        assert bank.account_number == "00000000"
        assert bank.currency == "KES"


def test_report_detail_renders_saved_bank_accounts(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]
    BankAccount = models["BankAccount"]

    with app.app_context():
        owner = create_user(db, User, "bank-owner@example.com", "OwnerUser@123")
        report = PBOReport(
            pbo_name="Bank Detail NGO",
            user_id=owner.id,
            workflow_status="draft",
            review_status="pending",
        )
        report.bank_accounts.append(BankAccount(
            bank_name="NCBA BANK KENYA",
            branch="UPPER HILL",
            account_number="9988776655",
            currency="USD",
        ))
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login(client, "bank-owner@example.com", "OwnerUser@123")
    response = client.get(f"/report/{report_id}")

    assert response.status_code == 200
    assert b"Bank Accounts" in response.data
    assert b"NCBA BANK KENYA" in response.data
    assert b"UPPER HILL" in response.data
    assert b"9988776655" in response.data
    assert b"USD" in response.data


def test_report_edit_persists_default_bank_account_number_and_currency(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        owner = create_user(db, User, "bank-edit-owner@example.com", "OwnerUser@123")
        report = PBOReport(
            pbo_name="Bank Edit NGO",
            user_id=owner.id,
            workflow_status="draft",
            review_status="pending",
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login(client, "bank-edit-owner@example.com", "OwnerUser@123")
    response = client.post(
        f"/report/{report_id}/edit",
        data={
            "pbo_name": "Bank Edit NGO",
            "bank_name[]": ["ABSA BANK KENYA"],
            "bank_name_other[]": [""],
            "bank_branch[]": ["Karen"],
            "bank_account_number[]": [""],
            "bank_currency[]": [""],
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Error updating report" not in response.data

    with app.app_context():
        report = db.session.get(PBOReport, report_id)
        assert report is not None
        assert len(report.bank_accounts) == 1
        bank = report.bank_accounts[0]
        assert bank.bank_name == "ABSA BANK KENYA"
        assert bank.branch == "KAREN"
        assert bank.account_number == "00000000"
        assert bank.currency == "KES"

    detail = client.get(f"/report/{report_id}")
    assert detail.status_code == 200
    assert b"ABSA BANK KENYA" in detail.data
    assert b"KAREN" in detail.data
    assert b"00000000" in detail.data
    assert b"KES" in detail.data


def test_report_detail_and_edit_show_associated_username(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        owner = create_user(db, User, "detail-owner@example.com", "OwnerUser@123")
        report = PBOReport(
            pbo_name="Associated Username NGO",
            user_id=owner.id,
            workflow_status="draft",
            review_status="pending",
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login(client, "detail-owner@example.com", "OwnerUser@123")

    detail_response = client.get(f"/report/{report_id}")
    edit_response = client.get(f"/report/{report_id}/edit")

    assert detail_response.status_code == 200
    assert edit_response.status_code == 200
    assert b"Associated Username" in detail_response.data
    assert b"detail-owner" in detail_response.data
    assert b"Associated Username" in edit_response.data
    assert b"detail-owner" in edit_response.data


def test_report_edit_accepts_inactive_section_flags_and_clears_section_data(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        owner = create_user(db, User, "report-owner@example.com", "OwnerUser@123")
        report = PBOReport(
            pbo_name="Section Flag NGO",
            user_id=owner.id,
            workflow_status="draft",
            review_status="pending",
            gov_tax_waiver=True,
            gov_tax_waiver_amount=1500,
            gov_other=True,
            gov_other_specify="OLD GOV SUPPORT",
            gov_other_amount=2500,
            submitter_fullname="OLD SIGNER",
            signature="OLD-SIGNATURE",
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login(client, "report-owner@example.com", "OwnerUser@123")
    response = client.post(
        f"/report/{report_id}/edit",
        data={
            "pbo_name": "Section Flag NGO Updated",
            "government_sections_inactive": "1",
            "officials_section_inactive": "1",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Error updating report" not in response.data

    with app.app_context():
        report = db.session.get(PBOReport, report_id)
        assert report.pbo_name == "SECTION FLAG NGO UPDATED"
        assert report.gov_tax_waiver is None
        assert report.gov_tax_waiver_amount is None
        assert report.gov_other is None
        assert report.gov_other_specify is None
        assert report.gov_other_amount is None
        assert report.submitter_fullname is None
        assert report.signature is None


def test_reports_list_renders_dates_payments_and_key_answers_columns(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]
    Payment = models["Payment"]

    with app.app_context():
        viewer = create_user(db, User, "reportsviewer@example.com", "ViewerUser@123")
        report = PBOReport(
            pbo_name="Columns NGO",
            pbo_registration_number="REG-COLUMNS-001",
            user_id=viewer.id,
            workflow_status="submitted",
            review_status="pending",
            audited="YES",
            assets_stolen="NO",
            gov_tax_waiver=True,
            election_frequency="ANNUALLY",
        )
        db.session.add(report)
        db.session.commit()
        db.session.add(Payment(report_id=report.id, description="Rent", kenya_amount=1000, other_amount=250))
        db.session.commit()

    login(client, "reportsviewer@example.com", "ViewerUser@123")
    response = client.get("/reports")

    assert response.status_code == 200
    assert b"Dates" in response.data
    assert b"Updated By" in response.data
    assert b"Payments" in response.data
    assert b"Key Answers" in response.data
    assert b"Rows: 1, Total: 1,250.00" in response.data
    assert b"Audited: YES" in response.data
    assert b"Assets issue: NO" in response.data
    assert b"Tax waiver: YES" in response.data
    assert b"Election: ANNUALLY" in response.data


def test_reports_list_supports_search_and_per_page(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        viewer = create_user(db, User, "searchviewer@example.com", "ViewerUser@123")
        for index in range(12):
            report = PBOReport(
                pbo_name=f"SEARCH NGO {index}",
                pbo_registration_number=f"REG-SEARCH-{index:03d}",
                contact_name=f"CONTACT {index}",
                contact_email=f"CONTACT{index}@EXAMPLE.COM",
                user_id=viewer.id,
                workflow_status="submitted",
                review_status="pending",
            )
            db.session.add(report)
        db.session.commit()

    login(client, "searchviewer@example.com", "ViewerUser@123")

    filtered = client.get("/reports?q=SEARCH NGO 11&per_page=50")
    assert filtered.status_code == 200
    assert b"SEARCH NGO 11" in filtered.data
    assert b"SEARCH NGO 10" not in filtered.data
    assert b'value="50"' in filtered.data

    expanded = client.get("/reports?per_page=50")
    assert expanded.status_code == 200
    assert b"SEARCH NGO 11" in expanded.data
    assert b"SEARCH NGO 0" in expanded.data


def test_sector_report_data_and_tf_risk_routes_are_disabled_in_crud_only_deployment(app, client, models, app_code):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "table-admin@example.com", "TableAdmin@123", role="admin")

    login(client, "table-admin@example.com", "TableAdmin@123")

    sector_response = client.get("/sectorreportdata")
    risk_response = client.get("/admin/tf_risk")

    assert app_code.SECTOR_ANALYTICS_ENABLED is False
    assert app_code.TF_RISK_ENABLED is False
    assert sector_response.status_code == 404
    assert risk_response.status_code == 404


def test_regular_admin_cannot_edit_other_users_record_without_privilege(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        owner = create_user(db, User, "owner@example.com", "OwnerUser@123")
        admin = create_user(db, User, "admin2@example.com", "AdminTwo@123", role="admin")
        report = PBOReport(
            pbo_name="Owner NGO",
            user_id=owner.id,
            workflow_status="draft",
            review_status="pending",
        )
        db.session.add(report)
        db.session.commit()
        report_id = report.id

    login(client, "admin2@example.com", "AdminTwo@123")
    denied = client.post(
        f"/update_report_field/{report_id}",
        data={"pbo_name": "CHANGED BY ADMIN"},
        follow_redirects=True,
    )

    assert denied.status_code == 200
    assert b"do not have permission" in denied.data

    with app.app_context():
        report = db.session.get(PBOReport, report_id)
        assert report.pbo_name == "OWNER NGO"
        admin = User.query.filter_by(email="admin2@example.com").first()
        admin.can_manage_all_records = True
        db.session.commit()

    client.get("/logout", follow_redirects=True)
    login(client, "admin2@example.com", "AdminTwo@123")
    allowed = client.post(
        f"/update_report_field/{report_id}",
        data={"pbo_name": "CHANGED BY ADMIN"},
        follow_redirects=True,
    )

    assert allowed.status_code == 200
    with app.app_context():
        report = db.session.get(PBOReport, report_id)
        assert report.pbo_name == "CHANGED BY ADMIN"


def test_user_keeps_report_owner_after_admin_edits_same_record(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]
    UploadedFile = models["UploadedFile"]

    with app.app_context():
        user = create_user(db, User, "shared-owner@example.com", "SharedOwner@123")
        admin = create_user(
            db,
            User,
            "shared-admin@example.com",
            "SharedAdmin@123",
            role="admin",
            can_manage_all_records=True,
        )
        report = PBOReport(
            pbo_name="Shared Ownership NGO",
            pbo_registration_number="REG-SHARED-001",
            user_id=admin.id,
            workflow_status="draft",
            review_status="pending",
        )
        db.session.add(report)
        db.session.commit()
        db.session.add(UploadedFile(
            report_id=report.id,
            uploaded_by_id=user.id,
            category="supporting_document",
            original_filename="shared-owner-proof.pdf",
            stored_filename="shared-owner-proof.pdf",
            storage_path="/tmp/shared-owner-proof.pdf",
            status="uploaded",
        ))
        db.session.commit()
        report_id = report.id
        user_id = user.id
        admin_id = admin.id

    login(client, "shared-owner@example.com", "SharedOwner@123")
    user_edit = client.post(
        f"/report/{report_id}/edit",
        data={
            "pbo_name": "Shared Ownership NGO",
            "pbo_registration_number": "REG-SHARED-001",
        },
        follow_redirects=False,
    )

    assert user_edit.status_code == 302

    client.get("/logout", follow_redirects=True)
    login(client, "shared-admin@example.com", "SharedAdmin@123")
    admin_edit = client.post(
        f"/update_report_field/{report_id}",
        data={"pbo_name": "Changed By Admin"},
        follow_redirects=False,
    )

    assert admin_edit.status_code == 302

    with app.app_context():
        report = db.session.get(PBOReport, report_id)
        assert report.user_id == user_id
        assert report.last_modified_by_id == admin_id


def test_admin_user_worked_files_supports_search(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        admin = create_user(db, User, "worked-search-admin@example.com", "WorkedSearchAdmin@123", role="admin")
        target_user = create_user(db, User, "worked-search-user@example.com", "WorkedSearchUser@123")
        matching_report = PBOReport(
            pbo_name="Worked Search Match",
            pbo_registration_number="REG-WORKED-001",
            user_id=target_user.id,
            data_source="form",
            submitted_at=datetime(2026, 3, 21, 9, 0, 0),
        )
        other_report = PBOReport(
            pbo_name="Worked Search Other",
            pbo_registration_number="REG-WORKED-002",
            user_id=target_user.id,
            data_source="form",
            submitted_at=datetime(2026, 3, 21, 10, 0, 0),
        )
        db.session.add_all([matching_report, other_report])
        db.session.commit()
        target_user_id = target_user.id

    login(client, "worked-search-admin@example.com", "WorkedSearchAdmin@123")
    response = client.get(f"/admin/users/{target_user_id}/worked-files?q=Worked Search Match")

    assert response.status_code == 200
    assert b"Worked Search Match" in response.data
    assert b"Worked Search Other" not in response.data


def test_denied_report_access_does_not_count_as_worked_file(app, client, models, app_code):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        owner = create_user(db, User, "protected-owner@example.com", "ProtectedOwner@123")
        admin = create_user(db, User, "limited-admin@example.com", "LimitedAdmin@123", role="admin")
        report = PBOReport(
            pbo_name="Protected Report NGO",
            user_id=owner.id,
            workflow_status="submitted",
            review_status="pending",
        )
        db.session.add(report)
        db.session.commit()
        admin_id = admin.id
        report_id = report.id

    login(client, "limited-admin@example.com", "LimitedAdmin@123")

    first_denied = client.get(f"/report/{report_id}", follow_redirects=True)
    second_denied = client.get(f"/report/{report_id}", follow_redirects=True)
    worked_files = client.get(f"/admin/users/{admin_id}/worked-files")

    assert first_denied.status_code == 200
    assert second_denied.status_code == 200
    assert b"do not have permission" in first_denied.data
    assert b"do not have permission" in second_denied.data
    assert worked_files.status_code == 200
    assert b"Protected Report NGO" not in worked_files.data

    with app.app_context():
        admin = db.session.get(User, admin_id)
        report = db.session.get(PBOReport, report_id)
        assert app_code.user_has_worked_on_report(admin, report) is False


def test_workflow_approval_does_not_add_report_to_admin_worked_files(app, client, models, app_code):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        owner = create_user(db, User, "approval-owner@example.com", "ApprovalOwner@123")
        admin = create_user(
            db,
            User,
            "approval-admin@example.com",
            "ApprovalAdmin@123",
            role="admin",
            can_manage_all_records=True,
        )
        report = PBOReport(
            pbo_name="Approval Queue NGO",
            user_id=owner.id,
            workflow_status="submitted",
            review_status="pending",
        )
        db.session.add(report)
        db.session.commit()
        admin_id = admin.id
        report_id = report.id

    login(client, "approval-admin@example.com", "ApprovalAdmin@123")
    approval = client.post(
        f"/admin/report/{report_id}/workflow",
        data={"workflow_action": "approve"},
        follow_redirects=True,
    )
    worked_files = client.get(f"/admin/users/{admin_id}/worked-files")

    assert approval.status_code == 200
    assert worked_files.status_code == 200
    assert b"Approval Queue NGO" not in worked_files.data

    with app.app_context():
        admin = db.session.get(User, admin_id)
        report = db.session.get(PBOReport, report_id)
        assert report.workflow_status == "approved"
        assert app_code.user_has_worked_on_report(admin, report) is False


def test_admin_edit_does_not_add_report_to_admin_submitted_files(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        owner = create_user(db, User, "edit-owner@example.com", "EditOwner@123")
        admin = create_user(
            db,
            User,
            "edit-admin@example.com",
            "EditAdmin@123",
            role="admin",
            can_manage_all_records=True,
        )
        report = PBOReport(
            pbo_name="Edited By Admin NGO",
            pbo_registration_number="REG-EDIT-001",
            user_id=owner.id,
            data_source="form",
            workflow_status="submitted",
            review_status="pending",
            submitted_at=datetime(2026, 3, 21, 11, 0, 0),
        )
        db.session.add(report)
        db.session.commit()
        admin_id = admin.id
        report_id = report.id

    login(client, "edit-admin@example.com", "EditAdmin@123")
    response = client.post(
        f"/update_report_field/{report_id}",
        data={"pbo_name": "Edited By Admin NGO Updated"},
        follow_redirects=True,
    )
    worked_files = client.get(f"/admin/users/{admin_id}/worked-files")

    assert response.status_code == 200
    assert worked_files.status_code == 200
    assert b"Edited By Admin NGO" not in worked_files.data


def test_backup_kickoff_and_status_endpoint(app, client, models, app_code, monkeypatch):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(
            db,
            User,
            "backupadmin@example.com",
            "BackupAdmin@123",
            role="admin",
            is_superadmin=True,
            can_manage_all_records=True,
        )

    def fake_backup():
        app_code.update_backup_runtime_status(stage="uploading", detail="Uploading test backup", files_total=2, files_completed=1)
        time.sleep(0.05)
        return {
            "backup_dir": "backups/test",
            "archive_path": "backups/test/returnsform14_org_backup.zip",
            "local_backup_succeeded": True,
            "drive_upload_succeeded": True,
            "drive_upload_error": None,
        }

    monkeypatch.setattr(app_code, "perform_application_backup", fake_backup)

    login(client, "backupadmin@example.com", "BackupAdmin@123")
    response = client.post("/admin/backup", follow_redirects=True)
    assert response.status_code == 200

    final_payload = None
    for _ in range(20):
        status = client.get("/admin/backup/status")
        payload = status.get_json()
        if not payload["running"] and payload["last_backup_at"]:
            final_payload = payload
            break
        time.sleep(0.05)

    assert final_payload is not None
    assert final_payload["drive_upload_succeeded"] is True
    assert final_payload["last_backup_archive"].endswith("returnsform14_org_backup.zip")


def test_admin_view_renders_report_merge_tools(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(
            db,
            User,
            "mergepaneladmin@example.com",
            "MergePanelAdmin@123",
            role="admin",
            is_superadmin=True,
            can_manage_all_records=True,
        )

    login(client, "mergepaneladmin@example.com", "MergePanelAdmin@123")
    response = client.get("/admin")

    assert response.status_code == 200
    assert b"Export Reports" in response.data
    assert b"pbo_reports (main database table)" in response.data
    assert b'name="merge_related_source"' in response.data
    assert b'name="merge_upload_file"' in response.data
    assert b"pbo_payments" in response.data
    assert b"Merge Files" in response.data
    assert b"Download Cleaned Data" in response.data
    assert b"PostgreSQL Hardening Audit" in response.data
    assert b"Run Audit" in response.data


def test_save_uploaded_file_uses_configured_persistent_upload_folder(app, models, app_code, tmp_path, monkeypatch):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        user = create_user(db, User, "persistupload@example.com", "PersistUpload@123")
        storage_root = tmp_path / "persistent-storage"
        monkeypatch.setitem(app.config, "PERSISTENT_STORAGE_ROOT", str(storage_root))
        monkeypatch.setitem(app.config, "UPLOAD_FOLDER", "pdf_uploads")

        upload = FileStorage(
            stream=io.BytesIO(b"%PDF-1.4 merged pdf"),
            filename="merged-proof.pdf",
            content_type="application/pdf",
        )

        uploaded_file = app_code.save_uploaded_file(
            upload,
            category="ocr_source",
            user=user,
            status="uploaded",
        )
        db.session.flush()

        stored_path = Path(uploaded_file.storage_path)
        assert stored_path.parent == storage_root / "pdf_uploads"
        assert stored_path.exists()
        assert uploaded_file.original_filename == "merged-proof.pdf"


def test_report_merge_export_uses_configured_persistent_directory(app, models, app_code, tmp_path, monkeypatch):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]
    Payment = models["Payment"]

    with app.app_context():
        admin = create_user(
            db,
            User,
            "persistmerge@example.com",
            "PersistMerge@123",
            role="admin",
            can_manage_all_records=True,
        )
        report = PBOReport(
            pbo_name="Persistent Merge NGO",
            pbo_registration_number="REG-PERSIST-001",
            workflow_status="submitted",
            review_status="pending",
            user_id=admin.id,
        )
        db.session.add(report)
        db.session.commit()
        db.session.add(Payment(
            report_id=report.id,
            description="Travel",
            kenya_amount=1200,
            other_amount=300,
        ))
        db.session.commit()

        storage_root = tmp_path / "persistent-storage"
        monkeypatch.setitem(app.config, "PERSISTENT_STORAGE_ROOT", str(storage_root))
        monkeypatch.setitem(app.config, "REPORT_MERGE_EXPORT_DIR", "merged_exports")
        monkeypatch.setitem(app.config, "REPORT_MERGE_TEMP_DIR", "merged_temp")

        app_code.perform_report_merge_export(selected_source_key="pbo_payments")

        export_dir = storage_root / "merged_exports"
        exported_files = list(export_dir.glob("*.xlsx"))
        assert len(exported_files) == 1
        assert exported_files[0].exists()
        assert "pbo_payments__" in exported_files[0].name


def test_admin_view_runs_read_only_fk_audit_on_non_postgres(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(
            db,
            User,
            "fkauditadmin@example.com",
            "FkAuditAdmin@123",
            role="admin",
            is_superadmin=True,
            can_manage_all_records=True,
        )

    login(client, "fkauditadmin@example.com", "FkAuditAdmin@123")
    response = client.get("/admin?run_fk_audit=1")

    assert response.status_code == 200
    assert b"PostgreSQL Hardening Audit" in response.data
    assert b"Available only when running on PostgreSQL." in response.data
    assert b"does not delete rows" in response.data
    assert b"Status Columns" in response.data


def test_admin_view_hides_hardening_audit_from_non_superadmin(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(
            db,
            User,
            "plainadmin@example.com",
            "PlainAdmin@123",
            role="admin",
            can_manage_all_records=True,
        )

    login(client, "plainadmin@example.com", "PlainAdmin@123")
    response = client.get("/admin?run_fk_audit=1", follow_redirects=True)

    assert response.status_code == 200
    assert b"PostgreSQL Hardening Audit" not in response.data
    assert b"superadmins only" in response.data


def test_initialize_database_uses_audit_only_mode_for_postgresql_startup(app, app_code, monkeypatch):
    with app.app_context():
        mutation_calls = []
        audit_calls = []
        side_effect_calls = []

        monkeypatch.setattr(app_code.db.engine.dialect, "name", "postgresql", raising=False)
        monkeypatch.setattr(app_code, "report_database_state", lambda: None)
        monkeypatch.setattr(app_code, "ensure_default_admin", lambda: side_effect_calls.append("ensure_default_admin"))
        monkeypatch.setattr(app_code, "seed_configured_users", lambda: side_effect_calls.append("seed_configured_users"))
        monkeypatch.setattr(app_code, "repair_sqlite_rowid_backed_ids", lambda: mutation_calls.append("sqlite_repair"))
        monkeypatch.setattr(app_code, "dedupe_sqlite_users_by_email", lambda: mutation_calls.append("sqlite_dedupe"))
        monkeypatch.setattr(app_code.db, "create_all", lambda: mutation_calls.append("create_all"))
        monkeypatch.setattr(app_code, "ensure_legacy_user_schema", lambda: mutation_calls.append("legacy_user"))
        monkeypatch.setattr(app_code, "ensure_legacy_report_schema", lambda: mutation_calls.append("legacy_report"))
        monkeypatch.setattr(app_code, "ensure_all_model_string_capacities", lambda: mutation_calls.append("string_caps"))
        monkeypatch.setattr(app_code, "ensure_postgresql_schema_integrity", lambda: mutation_calls.append("postgres_mutation"))
        monkeypatch.setattr(app_code, "audit_postgresql_schema_integrity", lambda: audit_calls.append("postgres_audit") or {"available": True})

        app_code.initialize_database(
            reset=False,
            seed_users=False,
            apply_schema_changes=False,
            run_postgresql_audit=True,
            sync_default_admin=False,
        )

    assert mutation_calls == []
    assert audit_calls == ["postgres_audit"]
    assert side_effect_calls == []


def test_bootstrap_database_on_startup_keeps_seed_and_admin_sync_off_for_postgresql(app, app_code, monkeypatch):
    startup_calls = []

    monkeypatch.setattr(app_code, "running_schema_command", lambda: False)
    monkeypatch.setattr(app_code, "schema_work_disabled", lambda: False)
    monkeypatch.setattr(app_code.db.engine.dialect, "name", "postgresql", raising=False)

    def capture_initialize_database(**kwargs):
        startup_calls.append(kwargs)

    monkeypatch.setattr(app_code, "initialize_database", capture_initialize_database)

    app_code.bootstrap_database_on_startup()

    assert startup_calls == [
        {
            "reset": False,
            "seed_users": False,
            "apply_schema_changes": False,
            "run_postgresql_audit": True,
            "sync_default_admin": False,
        }
    ]


def test_admin_report_merge_runs_and_downloads_cleaned_workbook(app, client, models):
    db = models["db"]
    User = models["User"]
    PBOReport = models["PBOReport"]
    Payment = models["Payment"]

    with app.app_context():
        admin = create_user(
            db,
            User,
            "mergeadmin@example.com",
            "MergeAdmin@123",
            role="admin",
            can_manage_all_records=True,
        )
        report = PBOReport(
            pbo_name="Merge Payment NGO",
            pbo_registration_number="REG-MERGE-001",
            workflow_status="submitted",
            review_status="pending",
            user_id=admin.id,
        )
        db.session.add(report)
        db.session.commit()
        db.session.add(Payment(
            report_id=report.id,
            description="Rent",
            kenya_amount=7500,
            other_amount=500,
        ))
        db.session.commit()
        report_id = report.id

    workbook_buffer = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Legacy Data"
    worksheet.append(["report_id", "legacy_tag", "legacy_amount"])
    worksheet.append([report_id, "ARCHIVE MATCH", 123])
    workbook.save(workbook_buffer)
    workbook_buffer.seek(0)

    login(client, "mergeadmin@example.com", "MergeAdmin@123")
    response = client.post(
        "/admin/export/reports/merge",
        data={
            "merge_related_source": "pbo_payments",
            "merge_upload_file": (workbook_buffer, "legacy_merge.xlsx"),
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert response.status_code == 200

    final_payload = None
    for _ in range(40):
        status_response = client.get("/admin/export/reports/status")
        payload = status_response.get_json()
        if not payload["running"] and payload["download_ready"]:
            final_payload = payload
            break
        time.sleep(0.05)

    assert final_payload is not None
    assert final_payload["stage"] == "complete"
    assert final_payload["sources_total"] == 2
    assert final_payload["sources_completed"] == 2
    assert "pbo_payments" in " ".join(final_payload["selected_sources"])
    assert "legacy_merge.xlsx" in " ".join(final_payload["selected_sources"])
    assert final_payload["download_url"]
    assert "pbo_payments__" in final_payload["download_url"]

    download_response = client.get(final_payload["download_url"])
    assert download_response.status_code == 200
    assert download_response.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "pbo_payments__" in download_response.headers["Content-Disposition"]

    merged_workbook = load_workbook(io.BytesIO(download_response.data))
    assert merged_workbook.sheetnames == ["Matched Rows"]
    worksheet = merged_workbook["Matched Rows"]
    headers = [cell.value for cell in worksheet[1]]
    assert "source_table" in headers
    assert "report_id" in headers
    assert "pbo_name" in headers
    assert "legacy_tag" in headers
    assert "kenya_amount" in headers

    all_values = []
    for row in worksheet.iter_rows(values_only=True):
        all_values.extend([str(value) for value in row if value is not None])
    merged_text = " ".join(all_values)
    assert "MERGE PAYMENT NGO" in merged_text.upper()
    assert "ARCHIVE MATCH" in merged_text


def test_merge_export_loaders_use_table_reads_not_orm_query(app, models, app_code, monkeypatch, tmp_path):
    db = models["db"]
    PBOReport = models["PBOReport"]
    Payment = models["Payment"]
    User = models["User"]

    with app.app_context():
        user = create_user(db, User, "tablemerge@example.com", "TableMerge@123")
        report = PBOReport(
            pbo_name="Legacy Org",
            pbo_name_normalized="LEGACY ORG",
            reporting_period_start=datetime(2024, 1, 1),
            reporting_period_end=datetime(2024, 12, 31),
            scope="NATIONAL",
            counties="NAIROBI CITY",
            workflow_status="submitted",
            review_status="pending",
            user_id=user.id,
        )
        db.session.add(report)
        db.session.commit()
        db.session.add(
            Payment(
                report_id=report.id,
                description="Legacy Donor",
                kenya_amount=1234,
            )
        )
        db.session.commit()
        report_id = report.id
    monkeypatch.setattr(app_code.PBOReport, "query", None, raising=False)
    monkeypatch.setattr(app_code.Payment, "query", None, raising=False)

    with app.app_context():
        reference_rows = app_code.build_report_merge_reference_rows()
        source_export = app_code.build_report_merge_source_export("pbo_payments")

    inserted_reference = next(row for row in reference_rows if row["id"] == report_id)
    inserted_payment = next(row for row in source_export["rows"] if row["report_id"] == report_id)

    assert inserted_reference["pbo_name"].upper() == "LEGACY ORG"
    assert inserted_payment["report_id"] == report_id

    output_path = tmp_path / "legacy_merge.xlsx"
    app_code.merge_report_driven_sources(
        reference_rows=reference_rows,
        source_specs=[{
            "label": "pbo_payments",
            "rows": source_export["rows"],
            "columns": source_export["columns"],
            "related_key": "report_id",
        }],
        output_path=str(output_path),
        reference_label="pbo_reports",
        reference_key="id",
    )

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    assert workbook.sheetnames == ["Matched Rows"]
    rows = list(workbook["Matched Rows"].iter_rows(values_only=True))
    assert rows[0] == (
        "report_id",
        "pbo_name",
        "pbo_name_normalized",
        "pbo_registration_number",
        "reporting_period_start",
        "reporting_period_end",
        "scope",
        "counties",
        "countries_of_operation",
        "id",
        "description",
        "kenya_amount",
        "other_amount",
    )
    merged_row = next(row for row in rows[1:] if row[0] == report_id)
    assert str(merged_row[1]).upper() == "LEGACY ORG"
    assert str(merged_row[10]).upper() == "LEGACY DONOR"
    assert merged_row[11] == 1234


def test_build_merge_rows_from_table_uses_rowid_for_legacy_sqlite_tables(app, models, app_code, monkeypatch):
    from models import StaffBiodata

    class FakeResult:
        def mappings(self):
            return self

        def all(self):
            return [{
                "__sqlite_rowid": 77,
                "id": 6,
                "report_id": 10,
                "category": "SECTION_C SUMMARY | ROW",
                "prev_year": 1,
                "curr_year": 2,
            }]

    with app.app_context():
        monkeypatch.setattr(app_code, "sqlite_table_has_real_id_primary_key", lambda connection, table_name: False)
        monkeypatch.setattr(app_code.db.session, "connection", lambda: object())
        monkeypatch.setattr(app_code.db.session, "execute", lambda statement: FakeResult())

        columns, rows = app_code._build_merge_rows_from_table(StaffBiodata)

    assert columns == ["id", "report_id", "category", "prev_year", "curr_year"]
    assert rows == [{
        "id": 77,
        "report_id": 10,
        "category": "SECTION_C SUMMARY | ROW",
        "prev_year": 1,
        "curr_year": 2,
    }]


def test_admin_view_lists_previous_merged_files_and_downloads(app, client, models, app_code):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(
            db,
            User,
            "mergehistoryadmin@example.com",
            "MergeHistoryAdmin@123",
            role="admin",
            can_manage_all_records=True,
        )
        export_dir = app_code.report_merge_export_directory()
        stored_name = "1234567890abcdef1234567890abcdef_pbo_projects_carried_out__20260325__231955.xlsx"
        file_path = export_dir / stored_name
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Merge Overview"
        worksheet.append(["source_label", "matched_rows"])
        worksheet.append(["Payments", 1])
        workbook.save(file_path)
        fixed_timestamp = datetime(2026, 3, 25, 23, 19, 55).timestamp()
        os.utime(file_path, (fixed_timestamp, fixed_timestamp))
        export_dir_path = export_dir
        history_rows = app_code.list_report_merge_history(limit=200)
        expected_modified_display = next(
            item["modified_at_display"]
            for item in history_rows
            if item["stored_name"] == stored_name
        )

    login(client, "mergehistoryadmin@example.com", "MergeHistoryAdmin@123")
    response = client.get("/admin")

    assert response.status_code == 200
    assert b"Previous Merged Files" in response.data
    assert b"pbo_projects_carried_out__20260325__231955.xlsx" in response.data
    assert expected_modified_display.encode() in response.data
    assert b"Delete" in response.data
    assert b"Delete All" in response.data

    download_response = client.get(f"/admin/export/reports/files/{stored_name}")
    assert download_response.status_code == 200
    assert download_response.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "pbo_projects_carried_out__20260325__231955.xlsx" in download_response.headers["Content-Disposition"]

    delete_response = client.post(
        f"/admin/export/reports/files/{stored_name}/delete",
        follow_redirects=True,
    )
    assert delete_response.status_code == 200
    assert not (export_dir_path / stored_name).exists()
    missing_download_response = client.get(f"/admin/export/reports/files/{stored_name}")
    assert missing_download_response.status_code == 404

    second_stored_name = "abcdefabcdefabcdefabcdefabcdefab_pbo_payments__20260326__010101.xlsx"
    third_stored_name = "fedcbafedcbafedcbafedcbafedcbafe_pbo_donations__20260326__020202.xlsx"
    with app.app_context():
        for next_name in (second_stored_name, third_stored_name):
            next_path = export_dir_path / next_name
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Merge Overview"
            worksheet.append(["source_label", "matched_rows"])
            worksheet.append(["Test", 1])
            workbook.save(next_path)

    delete_all_response = client.post(
        "/admin/export/reports/files/delete-all",
        follow_redirects=True,
    )
    assert delete_all_response.status_code == 200
    assert not (export_dir_path / second_stored_name).exists()
    assert not (export_dir_path / third_stored_name).exists()
    assert b"No merged files have been saved yet." in delete_all_response.data


def test_ocr_upload_route_is_disabled_in_crud_only_deployment(app, client, models, app_code):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "ocradmin@example.com", "OcrAdmin@123", role="admin")

    login(client, "ocradmin@example.com", "OcrAdmin@123")
    response = client.post(
        "/reports/ocr-upload",
        data={"images": (io.BytesIO(b"fake-image-bytes"), "page1.jpg")},
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    assert app_code.OCR_FEATURES_ENABLED is False
    assert response.status_code == 404


def test_ocr_upload_pdf_route_is_disabled_in_crud_only_deployment(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "ocrpdf@example.com", "OcrPdf@123", role="admin")

    login(client, "ocrpdf@example.com", "OcrPdf@123")
    response = client.post(
        "/reports/ocr-upload",
        data={"images": (io.BytesIO(b"%PDF-1.4 fake"), "scan.pdf")},
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    assert response.status_code == 404


def test_ocr_upload_failure_route_is_disabled_in_crud_only_deployment(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "ocrfail@example.com", "OcrFail@123", role="admin")

    login(client, "ocrfail@example.com", "OcrFail@123")
    response = client.post(
        "/reports/ocr-upload",
        data={"images": (io.BytesIO(b"fake-image-bytes"), "page1.jpg")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )

    assert response.status_code == 404


def test_ocr_upload_duplicate_filename_route_is_disabled_in_crud_only_deployment(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "ocrdup@example.com", "OcrDup@123", role="admin")

    login(client, "ocrdup@example.com", "OcrDup@123")
    response = client.post(
        "/reports/ocr-upload",
        data={"images": (io.BytesIO(b"fake-image-bytes"), "page1.jpg")},
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    assert response.status_code == 404


def test_normalize_phone_accepts_kenyan_01_prefix(app_code):
    assert app_code.normalize_phone("0112345678") == "0112345678"
    assert app_code.normalize_phone("+254112345678", "+254") == "+254112345678"


def test_missing_string_and_integer_defaults(app_code):
    assert app_code.to_upper("") is None
    assert app_code.to_upper("   ") is None
    assert app_code.parse_int("") == 0
    assert app_code.parse_int(None) == 0
    assert app_code.parse_int("abc") == 0


def test_normalize_database_url_resolves_relative_sqlite_path_to_project_root(app_code):
    normalized = app_code.normalize_database_url("sqlite:///form14.db")
    assert normalized == f"sqlite:///{(Path(app_code.app.root_path) / 'form14.db').resolve()}"


def test_resolve_sqlite_bootstrap_source_url_prefers_db_bootstrap_candidate(app_code, monkeypatch, tmp_path):
    monkeypatch.delenv("SQLITE_BOOTSTRAP_SOURCE_URL", raising=False)
    monkeypatch.delenv("SQLITE_IMPORT_SOURCE_URL", raising=False)
    monkeypatch.setattr(app_code.app, "root_path", str(tmp_path))

    bootstrap_path = tmp_path / "db_bootstrap" / "returnsform14_org_backup.sqlite"
    bootstrap_path.parent.mkdir(parents=True, exist_ok=True)
    bootstrap_path.write_bytes(b"sqlite bootstrap placeholder")

    resolved = app_code.resolve_sqlite_bootstrap_source_url()

    assert resolved == f"sqlite:///{bootstrap_path.resolve()}"


def test_normalize_database_url_upgrades_mysql_driver_and_charset(app_code):
    normalized = app_code.normalize_database_url("mysql://demo:secret@localhost/sample_db")
    assert normalized == "mysql+pymysql://demo:secret@localhost/sample_db?charset=utf8mb4"


def test_resolve_primary_database_url_builds_mysql_url_from_env_parts(app_code, monkeypatch):
    monkeypatch.delenv("INTERNAL_DATABASE_URL", raising=False)
    monkeypatch.setenv("DATABASE_DIALECT", "mysql")
    monkeypatch.setenv("MYSQL_HOST", "localhost")
    monkeypatch.setenv("MYSQL_PORT", "3306")
    monkeypatch.setenv("MYSQL_DATABASE", "haroldda_annual_returns")
    monkeypatch.setenv("MYSQL_USER", "haroldda_complianceresearch")
    monkeypatch.setenv("MYSQL_PASSWORD", "ResearchandCompliance_2026")

    resolved = app_code.resolve_primary_database_url()

    assert resolved == (
        "mysql+pymysql://haroldda_complianceresearch:ResearchandCompliance_2026"
        "@localhost:3306/haroldda_annual_returns?charset=utf8mb4"
    )


def test_redact_database_url_masks_network_database_passwords(app_code):
    redacted = app_code.redact_database_url(
        "mysql+pymysql://haroldda_complianceresearch:ResearchandCompliance_2026@localhost:3306/haroldda_annual_returns?charset=utf8mb4"
    )
    assert "ResearchandCompliance_2026" not in redacted
    assert ":***@" in redacted


def test_resolve_user_login_skips_none_query_rows(app_code, monkeypatch):
    matched_user = type("MatchedUser", (), {"email": "bnzomo"})()

    class QueryStub:
        def filter_by(self, **kwargs):
            return self

        def first(self):
            return None

        def filter(self, *args, **kwargs):
            return self

        def all(self):
            return [None, matched_user]

    monkeypatch.setattr(app_code.User, "query", QueryStub(), raising=False)

    user, resolution = app_code.resolve_user_login("bnzomo")

    assert resolution is None
    assert user is matched_user


def test_assign_legacy_sqlite_id_backfills_next_safe_id(monkeypatch):
    import models as models_module

    class DialectStub:
        name = "sqlite"

    class PragmaResult:
        def mappings(self):
            return self

        def all(self):
            return [{"name": "id", "pk": 0}]

    class ScalarResult:
        def scalar(self):
            return 184

    class ConnectionStub:
        def __init__(self):
            self.dialect = DialectStub()
            self.pragma_calls = []
            self.execute_calls = []

        def exec_driver_sql(self, sql):
            self.pragma_calls.append(sql)
            return PragmaResult()

        def execute(self, statement):
            self.execute_calls.append(str(statement))
            return ScalarResult()

    class TableStub:
        name = "pbo_reports"
        c = {"id": object()}

    class TargetStub:
        __table__ = TableStub()

        def __init__(self):
            self.id = None

    monkeypatch.setattr(models_module, "LEGACY_SQLITE_ID_PK_CACHE", {})

    connection = ConnectionStub()
    target = TargetStub()

    models_module.assign_legacy_sqlite_id(None, connection, target)

    assert target.id == 184
    assert connection.pragma_calls == ['PRAGMA table_info("pbo_reports")']
    assert any("COALESCE(MAX(COALESCE(id, rowid)), 0) + 1" in sql for sql in connection.execute_calls)


def test_ensure_persisted_primary_key_is_noop_for_postgres(app_code, monkeypatch):
    class TableStub:
        c = {"id": object()}

    class TargetStub:
        __table__ = TableStub()

        def __init__(self):
            self.id = 77

    monkeypatch.setattr(app_code, "current_database_dialect", lambda: "postgresql")

    assert app_code.ensure_persisted_primary_key(TargetStub(), label="report") == 77


def test_add_and_flush_new_instance_persists_sqlite_report_identity(app, models, app_code):
    db = models["db"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        report = PBOReport(
            pbo_name="SQLite Identity Guard Test",
            workflow_status="submitted",
            review_status="pending",
            data_source="form",
        )

        app_code.add_and_flush_new_instance(report, label="report")
        db.session.commit()

        persisted = db.session.execute(
            text("SELECT rowid AS sqlite_rowid, id FROM pbo_reports WHERE id = :report_id"),
            {"report_id": report.id},
        ).mappings().first()

        assert report.id is not None
        assert persisted is not None
        assert persisted["sqlite_rowid"] == report.id
        assert persisted["id"] == report.id


def test_superadmin_can_change_admin_role_via_json_endpoint(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(
            db,
            User,
            "boss@example.com",
            "BossUser@123",
            role="admin",
            is_superadmin=True,
            can_manage_all_records=True,
        )
        target = create_user(db, User, "member@example.com", "MemberUser@123", role="user")
        target_id = target.id

    login(client, "boss@example.com", "BossUser@123")

    promote = client.post(
        "/admin/users/change_admin",
        json={"user_id": target_id, "admin": True},
    )
    assert promote.status_code == 200
    assert promote.get_json()["success"] is True

    with app.app_context():
        promoted_user = db.session.get(User, target_id)
        assert promoted_user.role == "admin"
        assert promoted_user.is_authorized is True

    demote = client.post(
        "/admin/users/change_admin",
        json={"user_id": target_id, "admin": False},
    )
    assert demote.status_code == 200
    assert demote.get_json()["success"] is True

    with app.app_context():
        demoted_user = db.session.get(User, target_id)
        assert demoted_user.role == "user"


def test_admin_can_edit_user_account_from_management_flow(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "editor@example.com", "EditorUser@123", role="admin")
        target = create_user(
            db,
            User,
            "member@example.com",
            "MemberUser@123",
            department="ICT",
        )
        target_id = target.id

    login(client, "editor@example.com", "EditorUser@123")
    response = client.post(
        f"/admin/users/{target_id}/edit",
        data={
            "username": "updated.member",
            "email": "member@newdomain.org",
            "department": "finance and admin",
            "password": "UpdatedUser@123",
            "confirm_password": "UpdatedUser@123",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"updated successfully" in response.data

    with app.app_context():
        updated_user = db.session.get(User, target_id)
        assert updated_user.email == "updated.member@newdomain.org"
        assert updated_user.department == "FINANCE AND ADMIN"
        assert updated_user.check_password("UpdatedUser@123") is True


def test_admin_edit_user_account_blocks_duplicate_email(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "editor@example.com", "EditorUser@123", role="admin")
        target = create_user(db, User, "member@example.com", "MemberUser@123")
        create_user(db, User, "existing@example.com", "ExistingUser@123")
        target_id = target.id

    login(client, "editor@example.com", "EditorUser@123")
    response = client.post(
        f"/admin/users/{target_id}/edit",
        data={
            "username": "existing",
            "email": "existing@example.com",
            "department": "operations",
            "password": "",
            "confirm_password": "",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"already uses that email/username combination" in response.data

    with app.app_context():
        unchanged_user = db.session.get(User, target_id)
        assert unchanged_user.email == "member@example.com"


def test_regular_admin_cannot_edit_another_admin_account(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(db, User, "editor@example.com", "EditorUser@123", role="admin")
        other_admin = create_user(db, User, "otheradmin@example.com", "OtherAdmin@123", role="admin")
        other_admin_id = other_admin.id

    login(client, "editor@example.com", "EditorUser@123")
    response = client.post(
        f"/admin/users/{other_admin_id}/edit",
        data={
            "username": "blocked.admin",
            "email": "otheradmin@example.com",
            "department": "security",
            "password": "",
            "confirm_password": "",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Only a superadmin can edit another admin account." in response.data


def test_superadmin_access_cannot_be_revoked(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        create_user(
            db,
            User,
            "boss@example.com",
            "BossUser@123",
            role="admin",
            is_superadmin=True,
            can_manage_all_records=True,
        )
        protected = create_user(
            db,
            User,
            "protected@example.com",
            "ProtectedUser@123",
            role="admin",
            is_superadmin=True,
            can_manage_all_records=True,
        )
        protected_id = protected.id

    login(client, "boss@example.com", "BossUser@123")

    revoke_auth = client.post(
        f"/admin/users/{protected_id}/revoke",
        follow_redirects=True,
    )
    assert revoke_auth.status_code == 200
    assert b"Superadmin access cannot be revoked." in revoke_auth.data

    revoke_role = client.post(
        "/admin/users/change_admin",
        json={"user_id": protected_id, "admin": False},
    )
    assert revoke_role.status_code == 400
    assert revoke_role.get_json()["message"] == "Superadmin access cannot be revoked."

    with app.app_context():
        protected_user = db.session.get(User, protected_id)
        assert protected_user.is_superadmin is True
        assert protected_user.role == "admin"
        assert protected_user.is_authorized is True


def test_legacy_user_schema_is_bootstrapped_while_profile_and_password_routes_are_disabled(app, client, models, app_code):
    db = models["db"]

    with app.app_context():
        db.session.execute(text("DROP TABLE IF EXISTS users_for_form14"))
        db.session.execute(
            text(
                "CREATE TABLE users_for_form14 ("
                "id INTEGER PRIMARY KEY, "
                "email VARCHAR(255), "
                "password_hash VARCHAR(255), "
                "role VARCHAR(50), "
                "is_authorized BOOLEAN)"
            )
        )
        db.session.commit()

        app_code.ensure_legacy_user_schema()

        columns = {column["name"] for column in inspect(db.engine).get_columns("users_for_form14")}
        assert {"full_name", "phone", "department", "must_change_password", "password_changed_at"}.issubset(columns)

        db.session.execute(
            text(
                "INSERT INTO users_for_form14 "
                "(id, email, password_hash, role, is_authorized, must_change_password, failed_login_attempts) "
                "VALUES "
                "(:id, :email, :password_hash, :role, :is_authorized, :must_change_password, :failed_login_attempts)"
            ),
            {
                "id": 1,
                "email": "legacy@example.com",
                "password_hash": generate_password_hash("LegacyUser@123"),
                "role": "user",
                "is_authorized": True,
                "must_change_password": False,
                "failed_login_attempts": 0,
            },
        )
        db.session.commit()

    login_response = client.post(
        "/login",
        data={"email": "legacy@example.com", "password": "LegacyUser@123", "department": ""},
        follow_redirects=True,
    )

    assert login_response.status_code == 200
    assert b"Logged in successfully." in login_response.data

    profile_response = client.post(
        "/password/change",
        data={"action": "profile", "full_name": "Legacy User", "department": "Registry", "phone": "0712345678"},
        follow_redirects=True,
    )

    assert profile_response.status_code == 200
    assert b"Profile and password change are currently disabled." in profile_response.data

    password_response = client.post(
        "/password/change",
        data={
            "action": "password",
            "current_password": "LegacyUser@123",
            "password": "RenewedUser@123",
            "confirm_password": "RenewedUser@123",
        },
        follow_redirects=True,
    )

    assert password_response.status_code == 200
    assert b"Profile and password change are currently disabled." in password_response.data


def test_dedupe_sqlite_users_by_email_repoints_references(app, models, app_code):
    db = models["db"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        from models import UserActivityLog

        connection = db.session.connection()
        connection.exec_driver_sql("PRAGMA foreign_keys=OFF")
        connection.exec_driver_sql("DROP TABLE IF EXISTS users_for_form14")
        connection.exec_driver_sql(
            (
                "CREATE TABLE users_for_form14 ("
                "id INTEGER, "
                "email TEXT, "
                "full_name TEXT, "
                "phone TEXT, "
                "department TEXT, "
                "must_change_password INTEGER, "
                "password_hash TEXT, "
                "password_changed_at TIMESTAMP, "
                "role TEXT, "
                "is_superadmin INTEGER, "
                "can_manage_all_records INTEGER, "
                "is_authorized INTEGER, "
                "authorized_at TIMESTAMP, "
                "authorized_by_id INTEGER, "
                "last_login_at TIMESTAMP, "
                "last_login_ip TEXT, "
                "failed_login_attempts INTEGER, "
                "last_failed_login_at TIMESTAMP, "
                "report_id INTEGER)"
            )
        )
        db.session.commit()
        db.session.connection().exec_driver_sql("PRAGMA foreign_keys=OFF")

        app_code.ensure_legacy_user_schema()

        db.session.execute(
            text(
                "INSERT INTO users_for_form14 "
                "(id, email, full_name, department, must_change_password, password_hash, role, is_superadmin, can_manage_all_records, is_authorized, failed_login_attempts) "
                "VALUES "
                "(1, 'dupe@example.com', 'Legacy User', 'Field', 0, :user_hash, 'user', 0, 0, 0, 0), "
                "(2, 'dupe@example.com', 'Admin User', 'Registry', 0, :admin_hash, 'admin', 1, 1, 1, 0)"
            ),
            {
                "user_hash": generate_password_hash("User@123"),
                "admin_hash": generate_password_hash("Admin@123"),
            },
        )
        db.session.add(
            PBOReport(
                id=301,
                user_id=1,
                last_modified_by_id=1,
                workflow_status="submitted",
                review_status="pending",
                pbo_name="Duplicate User Report",
            )
        )
        db.session.add(
            UserActivityLog(
                id=401,
                user_id=1,
                report_id=301,
                action="report_updated",
            )
        )
        db.session.commit()

        summary = app_code.dedupe_sqlite_users_by_email()

        assert "dupe@example.com" in summary

        remaining_users = db.session.execute(
            text(
                "SELECT id, email, role, is_superadmin, can_manage_all_records, is_authorized, department "
                "FROM users_for_form14 WHERE lower(email) = 'dupe@example.com'"
            )
        ).mappings().all()
        assert len(remaining_users) == 1
        assert remaining_users[0]["id"] == 2
        assert remaining_users[0]["role"] == "admin"
        assert remaining_users[0]["is_superadmin"] == 1
        assert remaining_users[0]["can_manage_all_records"] == 1
        assert remaining_users[0]["is_authorized"] == 1
        assert remaining_users[0]["department"] == "Registry"

        report_row = db.session.execute(
            text("SELECT user_id, last_modified_by_id FROM pbo_reports WHERE id = 301")
        ).mappings().one()
        assert report_row["user_id"] == 2
        assert report_row["last_modified_by_id"] == 2

        activity_row = db.session.execute(
            text("SELECT user_id FROM user_activity_logs WHERE id = 401")
        ).mappings().one()
        assert activity_row["user_id"] == 2


def test_authenticated_user_with_missing_session_keys_is_rehydrated_not_logged_out(app, client, models):
    db = models["db"]
    User = models["User"]

    with app.app_context():
        user = create_user(db, User, "rehydrate@example.com", "RehydrateUser@123")

    login_response = login(client, "rehydrate@example.com", "RehydrateUser@123")
    assert login_response.status_code == 200

    with client.session_transaction() as session_state:
        session_state.pop("user_id", None)
        session_state.pop("user_email", None)

    response = client.get("/", follow_redirects=True)

    assert response.status_code == 200
    assert b"Logged in successfully." not in response.data
    assert b"Session expired. Please log in again." not in response.data

    with client.session_transaction() as session_state:
        assert session_state.get("user_id") is not None
        assert session_state.get("user_email") == "rehydrate@example.com"


def test_logged_out_user_is_redirected_from_protected_nav_targets(client):
    protected_paths = ["/reports", "/analysis", "/sectorreportdata"]

    for path in protected_paths:
        response = client.get(path, follow_redirects=False)
        assert response.status_code in (302, 401)
        if response.status_code == 302:
            assert "/login" in response.headers.get("Location", "")


def test_legacy_report_schema_is_bootstrapped_for_existing_post_deploy_tables(app, models, app_code):
    db = models["db"]

    with app.app_context():
        db.session.execute(text("DROP TABLE IF EXISTS pbo_reports"))
        db.session.execute(
            text(
                "CREATE TABLE pbo_reports ("
                "id INTEGER PRIMARY KEY, "
                "pbo_name VARCHAR(255), "
                "created_at DATETIME)"
            )
        )
        db.session.execute(
            text(
                "INSERT INTO pbo_reports (id, pbo_name, created_at) "
                "VALUES (1, 'LEGACY NGO', '2026-07-01 00:00:00')"
            )
        )
        db.session.commit()

        app_code.ensure_legacy_report_schema()

        columns = {column["name"] for column in inspect(db.engine).get_columns("pbo_reports")}
        assert {
            "workflow_status",
            "review_status",
            "duplicate_flag",
            "reporting_period_start_raw",
            "reporting_period_end_raw",
            "countries_of_operation",
            "income_b2_total",
            "receipts_total",
            "return_date",
        }.issubset(columns)
        legacy_return_date = db.session.execute(
            text("SELECT return_date FROM pbo_reports WHERE id = 1")
        ).scalar_one()
        assert str(legacy_return_date) == "9999-09-09"


def test_ensure_model_string_capacities_promotes_legacy_varchar_to_text_for_postgresql(app, models, app_code, monkeypatch):
    executed_statements = []

    class DummyInspector:
        def get_table_names(self):
            return ["pbo_reports"]

        def get_columns(self, table_name):
            assert table_name == "pbo_reports"
            return [{"name": "countries_of_operation", "type": SimpleNamespace(length=255)}]

    with app.app_context():
        monkeypatch.setattr(app_code, "inspect", lambda engine: DummyInspector())
        monkeypatch.setattr(app_code.db.engine.dialect, "name", "postgresql", raising=False)
        monkeypatch.setattr(
            app_code.db.session,
            "execute",
            lambda statement, *args, **kwargs: executed_statements.append(str(statement)),
        )
        monkeypatch.setattr(app_code.db.session, "commit", lambda: executed_statements.append("COMMIT"))

        updated_columns = app_code.ensure_model_string_capacities(models["PBOReport"])

    assert updated_columns == ["countries_of_operation=255->TEXT"]
    assert any(
        'ALTER TABLE "pbo_reports" ALTER COLUMN "countries_of_operation" TYPE TEXT' in statement
        for statement in executed_statements
    )
    assert "COMMIT" in executed_statements


def test_sync_db_schema_cli_runs_non_destructive_schema_sync(app, app_code, monkeypatch):
    initialize_calls = []

    monkeypatch.setattr(app_code, "initialize_database", lambda **kwargs: initialize_calls.append(kwargs))

    result = app.test_cli_runner().invoke(args=["sync-db-schema"])

    assert result.exit_code == 0
    assert initialize_calls == [
        {
            "reset": False,
            "seed_users": False,
            "apply_schema_changes": True,
            "run_postgresql_audit": False,
            "sync_default_admin": False,
        }
    ]


def test_fresh_schema_exposes_db_level_status_constraints(app, models):
    db = models["db"]

    with app.app_context():
        inspector = inspect(db.engine)
        pbo_checks = {constraint["name"] for constraint in inspector.get_check_constraints("pbo_reports")}
        import_checks = {constraint["name"] for constraint in inspector.get_check_constraints("import_batches")}
        uploaded_checks = {constraint["name"] for constraint in inspector.get_check_constraints("uploaded_files")}

        assert "ck_pbo_reports_workflow_status_allowed" in pbo_checks
        assert "ck_pbo_reports_review_status_allowed" in pbo_checks
        assert "ck_import_batches_status_allowed" in import_checks
        assert "ck_uploaded_files_status_allowed" in uploaded_checks


def test_db_rejects_invalid_report_status_values(app, models):
    db = models["db"]
    PBOReport = models["PBOReport"]

    with app.app_context():
        db.session.add(
            PBOReport(
                pbo_name="Invalid Status NGO",
                workflow_status="not_a_real_status",
                review_status="pending",
            )
        )
        with pytest.raises(IntegrityError):
            db.session.commit()
        db.session.rollback()


def test_postgresql_bootstrap_preserves_invalid_report_status_values(app, models, app_code):
    db = models["db"]

    if db.engine.dialect.name != "postgresql":
        pytest.skip("PostgreSQL-specific bootstrap hardening test.")

    with app.app_context():
        db.session.execute(text("DROP TABLE IF EXISTS pbo_reports CASCADE"))
        db.session.execute(
            text(
                "CREATE TABLE pbo_reports ("
                "id SERIAL PRIMARY KEY, "
                "workflow_status VARCHAR(500), "
                "review_status VARCHAR(500))"
            )
        )
        db.session.execute(
            text(
                "INSERT INTO pbo_reports (workflow_status, review_status) VALUES "
                "('Submitted', 'unknown'), "
                "(NULL, NULL)"
            )
        )
        db.session.commit()

        app_code.ensure_legacy_report_schema()
        result = app_code.ensure_postgresql_schema_integrity()
        status_audit = app_code.collect_postgresql_status_constraint_audit()

        rows = db.session.execute(
            text("SELECT workflow_status, review_status FROM pbo_reports ORDER BY id")
        ).all()
        columns = {column["name"]: column for column in inspect(db.engine).get_columns("pbo_reports")}
        check_names = {
            constraint["name"]
            for constraint in inspect(db.engine).get_check_constraints("pbo_reports")
        }

        assert rows == [("Submitted", "unknown"), (None, None)]
        assert columns["workflow_status"]["nullable"] is True
        assert columns["review_status"]["nullable"] is True
        assert "ck_pbo_reports_workflow_status_allowed" not in check_names
        assert "ck_pbo_reports_review_status_allowed" not in check_names
        assert not result["status_constraints"]
        assert result["skipped_status_constraints"]
        assert status_audit["blocking_rows"]


def test_postgresql_bootstrap_hardens_clean_report_status_columns(app, models, app_code):
    db = models["db"]

    if db.engine.dialect.name != "postgresql":
        pytest.skip("PostgreSQL-specific bootstrap hardening test.")

    with app.app_context():
        db.session.execute(text("DROP TABLE IF EXISTS pbo_reports CASCADE"))
        db.session.execute(
            text(
                "CREATE TABLE pbo_reports ("
                "id SERIAL PRIMARY KEY, "
                "workflow_status VARCHAR(500), "
                "review_status VARCHAR(500))"
            )
        )
        db.session.execute(
            text(
                "INSERT INTO pbo_reports (workflow_status, review_status) VALUES "
                "('submitted', 'pending')"
            )
        )
        db.session.commit()

        app_code.ensure_legacy_report_schema()
        result = app_code.ensure_postgresql_schema_integrity()

        rows = db.session.execute(
            text("SELECT workflow_status, review_status FROM pbo_reports ORDER BY id")
        ).all()
        columns = {column["name"]: column for column in inspect(db.engine).get_columns("pbo_reports")}
        check_names = {
            constraint["name"]
            for constraint in inspect(db.engine).get_check_constraints("pbo_reports")
        }

        assert rows == [("submitted", "pending")]
        assert columns["workflow_status"]["nullable"] is False
        assert columns["review_status"]["nullable"] is False
        assert "ck_pbo_reports_workflow_status_allowed" in check_names
        assert "ck_pbo_reports_review_status_allowed" in check_names
        assert result["status_constraints"]
        assert not result["skipped_status_constraints"]


def test_postgresql_bootstrap_preserves_invalid_required_report_foreign_key_rows(app, models, app_code):
    db = models["db"]

    if db.engine.dialect.name != "postgresql":
        pytest.skip("PostgreSQL-specific foreign-key hardening test.")

    with app.app_context():
        db.session.execute(text("DROP TABLE IF EXISTS pbo_payments"))
        db.session.execute(text("DROP TABLE IF EXISTS pbo_reports CASCADE"))
        db.session.execute(
            text(
                "CREATE TABLE pbo_reports ("
                "id INTEGER PRIMARY KEY, "
                "workflow_status VARCHAR(500), "
                "review_status VARCHAR(500))"
            )
        )
        db.session.execute(
            text(
                "CREATE TABLE pbo_payments ("
                "id SERIAL PRIMARY KEY, "
                "report_id INTEGER, "
                "description VARCHAR(255))"
            )
        )
        db.session.execute(
            text(
                "INSERT INTO pbo_reports (id, workflow_status, review_status) VALUES "
                "(1, 'draft', 'pending')"
            )
        )
        db.session.execute(
            text(
                "INSERT INTO pbo_payments (report_id, description) VALUES "
                "(NULL, 'NO REPORT'), "
                "(999, 'ORPHANED'), "
                "(1, 'VALID PAYMENT')"
            )
        )
        db.session.commit()

        result = app_code.ensure_postgresql_schema_integrity()

        remaining_rows = db.session.execute(
            text("SELECT report_id, description FROM pbo_payments ORDER BY id")
        ).all()
        columns = {column["name"]: column for column in inspect(db.engine).get_columns("pbo_payments")}
        matching_foreign_keys = [
            foreign_key
            for foreign_key in inspect(db.engine).get_foreign_keys("pbo_payments")
            if foreign_key.get("constrained_columns") == ["report_id"]
            and foreign_key.get("referred_table") == "pbo_reports"
        ]

        assert remaining_rows == [(None, "NO REPORT"), (999, "ORPHANED"), (1, "VALID PAYMENT")]
        assert columns["report_id"]["nullable"] is True
        assert not matching_foreign_keys
        assert not result["required_foreign_keys"]
        assert result["skipped_required_foreign_keys"]
        assert result["skipped_required_foreign_keys"][0]["table_name"] == "pbo_payments"
        assert result["skipped_required_foreign_keys"][0]["null_row_count"] == 1
        assert result["skipped_required_foreign_keys"][0]["orphan_row_count"] == 1


def test_postgresql_fk_audit_reports_blocking_required_foreign_key_tables(app, models, app_code):
    db = models["db"]

    if db.engine.dialect.name != "postgresql":
        pytest.skip("PostgreSQL-specific FK audit test.")

    with app.app_context():
        db.session.execute(text("DROP TABLE IF EXISTS pbo_payments"))
        db.session.execute(text("DROP TABLE IF EXISTS pbo_reports CASCADE"))
        db.session.execute(
            text(
                "CREATE TABLE pbo_reports ("
                "id INTEGER PRIMARY KEY, "
                "workflow_status VARCHAR(500), "
                "review_status VARCHAR(500))"
            )
        )
        db.session.execute(
            text(
                "CREATE TABLE pbo_payments ("
                "id SERIAL PRIMARY KEY, "
                "report_id INTEGER, "
                "description VARCHAR(255))"
            )
        )
        db.session.execute(
            text(
                "INSERT INTO pbo_reports (id, workflow_status, review_status) VALUES "
                "(1, 'draft', 'pending')"
            )
        )
        db.session.execute(
            text(
                "INSERT INTO pbo_payments (report_id, description) VALUES "
                "(NULL, 'NO REPORT'), "
                "(999, 'ORPHANED'), "
                "(1, 'VALID PAYMENT')"
            )
        )
        db.session.commit()

        audit_payload = app_code.collect_postgresql_required_foreign_key_audit()
        payment_row = next(
            row for row in audit_payload["rows"]
            if row["table_name"] == "pbo_payments" and row["column_name"] == "report_id"
        )

        assert audit_payload["available"] is True
        assert payment_row["status_key"] == "blocked"
        assert payment_row["null_row_count"] == 1
        assert payment_row["orphan_row_count"] == 1
        assert audit_payload["summary"]["blocked_relations"] >= 1
        assert audit_payload["summary"]["blocked_rows"] >= 2


def test_postgresql_bootstrap_hardens_required_report_foreign_keys_when_rows_are_clean(app, models, app_code):
    db = models["db"]

    if db.engine.dialect.name != "postgresql":
        pytest.skip("PostgreSQL-specific foreign-key hardening test.")

    with app.app_context():
        db.session.execute(text("DROP TABLE IF EXISTS pbo_payments"))
        db.session.execute(text("DROP TABLE IF EXISTS pbo_reports CASCADE"))
        db.session.execute(
            text(
                "CREATE TABLE pbo_reports ("
                "id INTEGER PRIMARY KEY, "
                "workflow_status VARCHAR(500), "
                "review_status VARCHAR(500))"
            )
        )
        db.session.execute(
            text(
                "CREATE TABLE pbo_payments ("
                "id SERIAL PRIMARY KEY, "
                "report_id INTEGER, "
                "description VARCHAR(255))"
            )
        )
        db.session.execute(
            text(
                "INSERT INTO pbo_reports (id, workflow_status, review_status) VALUES "
                "(1, 'draft', 'pending')"
            )
        )
        db.session.execute(
            text(
                "INSERT INTO pbo_payments (report_id, description) VALUES "
                "(1, 'VALID PAYMENT')"
            )
        )
        db.session.commit()

        result = app_code.ensure_postgresql_schema_integrity()

        remaining_rows = db.session.execute(
            text("SELECT report_id, description FROM pbo_payments ORDER BY id")
        ).all()
        columns = {column["name"]: column for column in inspect(db.engine).get_columns("pbo_payments")}
        matching_foreign_keys = [
            foreign_key
            for foreign_key in inspect(db.engine).get_foreign_keys("pbo_payments")
            if foreign_key.get("constrained_columns") == ["report_id"]
            and foreign_key.get("referred_table") == "pbo_reports"
        ]

        assert remaining_rows == [(1, "VALID PAYMENT")]
        assert columns["report_id"]["nullable"] is False
        assert matching_foreign_keys
        assert ((matching_foreign_keys[0].get("options") or {}).get("ondelete") or "").upper() == "CASCADE"
        assert result["required_foreign_keys"]
        assert not result["skipped_required_foreign_keys"]

        with pytest.raises(IntegrityError):
            db.session.execute(
                text(
                    "INSERT INTO pbo_payments (report_id, description) "
                    "VALUES (999, 'SHOULD FAIL')"
                )
            )
            db.session.commit()
        db.session.rollback()
